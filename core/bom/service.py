"""
Service Layer para BOM (Lista de Materiales).
Logica de negocio, workflow de aprobaciones, versionado y exportacion Excel.
"""

import asyncio
import logging
import json
import time
from collections import defaultdict
from contextlib import asynccontextmanager
from decimal import Decimal, InvalidOperation
from uuid import UUID, uuid4
from typing import Optional, List, Set

import asyncpg
import httpx
from jinja2 import TemplateError

from core.bom.compras_service import (
    BomComprasServiceMixin,
    ESTATUS_ITEM_CERRADO_COMPRA,
    ESTATUS_COMPRA_BLOQUEA_ADENDA,
    ESTATUS_ADENDA_APROBADA,
)
from core.bom.db_service import BomDBService
from core.bom.schemas import EstatusBOM, AccionHistorial, TipoAprobacion
from core.config import settings
from core.constants import ESTATUS_BOM_OCULTOS_PENDIENTES_PRECIO_COMPRAS
from core.materials.normalizer import normalizar_unidad
from core.materials.service import MaterialsService
from core.notifications.service import get_notifications_service
from core.timezone import now_mx, today_mx
from core.config_service import ConfigService

logger = logging.getLogger("BOM.Service")

# Campos que puede editar cada area
CAMPOS_INGENIERIA = {'id_categoria', 'descripcion', 'cantidad', 'unidad_medida', 'precio_unitario', 'origen_precio', 'tipo_partida', 'moneda'}
CAMPOS_CONSTRUCCION_BASE = {'fecha_requerida'}
CAMPOS_CONSTRUCCION_EJECUCION = {
    'entregado', 'comentarios', 'comentarios_operativos', 'cantidad_recibida',
    'fecha_estimada_entrega', 'fecha_llegada_real', 'estatus_ejecucion'
}
CAMPOS_CONSTRUCCION = CAMPOS_CONSTRUCCION_BASE | CAMPOS_CONSTRUCCION_EJECUCION
CAMPOS_BASE_BOM = CAMPOS_INGENIERIA | CAMPOS_CONSTRUCCION_BASE
# Campos que entran en `_calcular_estadisticas_costo`: solo estos justifican
# refrescar el grid de estadisticas OOB tras editar un item.
CAMPOS_AFECTAN_COSTO_ESTIMADO = {'precio_unitario', 'moneda', 'cantidad'}
CAMPOS_COMPRAS = {
    'id_proveedor', 'tipo_entrega', 'fecha_estimada_entrega',
    'fecha_llegada_real', 'comentarios', 'comentarios_operativos',
    'precio_unitario', 'precio_real', 'origen_precio', 'moneda',
    'moneda_real', 'estatus_ejecucion'
}

# Campos editables en lote: los de edicion individual menos los que varian por
# item. Derivado de los sets por area para que el bulk sea siempre un subconjunto
# del individual (evita drift al agregar/quitar campos).
_CAMPOS_BULK_EXCLUIDOS = {'descripcion', 'cantidad', 'cantidad_recibida'}
# Construccion no tiene autoridad de costos (precio_unitario/moneda son de
# Ingenieria o Compras) aunque controle el turno y tenga editar_base=True.
_CAMPOS_BULK_EXCLUIDOS_CONSTRUCCION = _CAMPOS_BULK_EXCLUIDOS | {'precio_unitario', 'moneda'}
CAMPOS_BULK = {
    'ingenieria': CAMPOS_BASE_BOM - _CAMPOS_BULK_EXCLUIDOS,
    'construccion': (CAMPOS_BASE_BOM | CAMPOS_CONSTRUCCION_EJECUCION) - _CAMPOS_BULK_EXCLUIDOS_CONSTRUCCION,
    'compras': CAMPOS_COMPRAS - _CAMPOS_BULK_EXCLUIDOS,
}
CAMPOS_BULK_BASE = CAMPOS_BASE_BOM - _CAMPOS_BULK_EXCLUIDOS

# Campos de bulk-edit que exigen moneda en la misma escritura (valor_secundario):
# capturar el monto sin la moneda deja el dato financiero ambiguo. No aplica cuando
# el campo se redirige a ejecucion (compras + precio_unitario -> precio_real).
CAMPOS_BULK_REQUIERE_MONEDA = {'precio_unitario'}

# Estados en los que NO se puede editar de ninguna forma
ESTATUS_BLOQUEADOS = {EstatusBOM.CANCELADO, EstatusBOM.APROBADO_FINAL}
ESTATUS_BLOQUEADOS_EJECUCION = {EstatusBOM.CANCELADO}
TIPO_ITEM_BASE = "BASE"
TIPO_ITEM_REEMPLAZO = "REEMPLAZO"
TIPO_ITEM_FUERA_SCOPE = "FUERA_SCOPE"
ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION = "PENDIENTE_CONSTRUCCION"
ESTATUS_ADENDA_PENDIENTE_INGENIERIA = "PENDIENTE_INGENIERIA"
ESTATUS_ADENDA_TERMINALES = {ESTATUS_ADENDA_APROBADA, "RECHAZADA", "CANCELADA"}
MONEDAS_VALIDAS = {"MXN", "USD"}
FLAG_ACTUALIZACION_PRECIOS_COMPRAS = "compras.actualizacion_precios_habilitada"

# Fechas de cabecera que representan el recorrido completo de aprobaciones.
FECHAS_FLUJO_BOM = (
    "fecha_envio_ing",
    "fecha_aprobacion_ing",
    "fecha_envio_obra",
    "fecha_aprobacion_obra",
    "fecha_envio_const",
    "fecha_aprobacion_const",
    "fecha_envio_final",
    "fecha_aprobacion_final",
)

# Estatus editables para agregar/eliminar items estructurales (ingenieria y construccion)
ESTATUS_EDITABLE_ING = set(EstatusBOM) - ESTATUS_BLOQUEADOS

# Campos especificos editables por construccion/compras en cualquier fase no cancelada
ESTATUS_EDITABLE_CONST_COMPRAS = set(EstatusBOM) - ESTATUS_BLOQUEADOS_EJECUCION

# Labels para historial
CAMPO_LABELS = {
    'id_categoria': 'Categoria',
    'descripcion': 'Descripcion',
    'cantidad': 'Cantidad',
    'unidad_medida': 'Unidad de medida',
    'fecha_requerida': 'Fecha requerida',
    'fecha_llegada_real': 'Fecha llegada real',
    'id_proveedor': 'Proveedor',
    'id_proveedor_real': 'Proveedor real',
    'tipo_entrega': 'Tipo entrega',
    'fecha_estimada_entrega': 'Fecha estimada entrega',
    'comentarios': 'Comentarios',
    'entregado': 'Entregado',
    'precio_unitario': 'Presupuesto unitario',
    'precio_real': 'Costo real',
    'moneda_real': 'Moneda real',
    'origen_precio': 'Origen precio',
    'cantidad_recibida': 'Cantidad recibida',
    'estatus_ejecucion': 'Estatus ejecucion',
    'comentarios_operativos': 'Comentarios operativos',
    'grupos_bom': 'Grupo',
    'grupos_operativos': 'Grupo operativo',
    'id_material_interno': 'Material vinculado',
    'id_material_ref': 'Material vinculado',
    'tipo_partida': 'Tipo de partida',
    'moneda': 'Moneda',
}

BOM_COSTOS_EVENTO = "BOM_ITEMS_SIN_COSTO"
BOM_COSTOS_REGLAS_MODULOS = {"BOM"}
BOM_COSTOS_ASUNTO_KEY = "bom.costos_notificacion_asunto"
BOM_COSTOS_TEMPLATE_KEY = "bom.costos_notificacion_template"
BOM_COSTOS_SSE_KEY = "bom.costos_notificacion_sse_activa"
BOM_COSTOS_DEFAULT_ASUNTO = (
    "BOM {proyecto_id} - {paquete_codigo} v{version} - Items sin costo estimado"
)
BOM_COSTOS_DEFAULT_TEMPLATE = (
    "El ingeniero ingreso {total_items} item(s) para {paquete_codigo} v{version} "
    "del proyecto {proyecto_id} sin costo estimado. Ingresa para actualizarlos."
)

# Estatus donde el BOM ya no admite el flujo de precios pendientes de Compras: a
# partir de APROBADO_CONST lo cubre la pestaña "Activos" de Compras (frontera
# compartida con modules.compras.db_service.get_proyectos_bom_pendientes_precio,
# que excluye el mismo set). Antes de eso (BORRADOR..EN_REVISION_CONST) un item
# sin costo puede haberse agregado en cualquier turno, no solo en BORRADOR.
ESTATUS_FUERA_DE_PRECIOS_PENDIENTES_COMPRAS = {
    *ESTATUS_BOM_OCULTOS_PENDIENTES_PRECIO_COMPRAS,
    EstatusBOM.APROBADO_FINAL.value,
}


class BomService(BomComprasServiceMixin):
    """Logica de negocio para BOM."""

    def __init__(self):
        self.db = BomDBService()
        self.materials = MaterialsService()

    # Cache en memoria (por worker) de get_catalogos: son catalogos que casi
    # no cambian (proveedores, tipos de entrega, usuarios por area, grupos
    # BOM) y se recargaban desde cero en cada apertura de modal de item (8
    # queries). No usa ConfigService/Redis porque los ids son UUID: el
    # roundtrip JSON de Redis los convertiria a str y rompería comparaciones
    # como `item.id_categoria == cat.id` en los templates.
    _cache_catalogos: Optional[tuple] = None
    _CATALOGOS_TTL_SECONDS = 60.0
    # Solo serializa el refresco (evita rehacer las 8 queries en cada request
    # concurrente cuando el TTL expira); no protege la asignacion en si, que
    # ya es atomica.
    _cache_catalogos_refresh_lock = asyncio.Lock()

    @staticmethod
    @asynccontextmanager
    async def _transaction(conn):
        """Transaccion real; el fallback solo facilita dobles unitarios sin I/O."""
        if hasattr(conn, "transaction"):
            async with conn.transaction():
                yield
        else:
            yield

    @staticmethod
    def _limpiar_fechas_flujo(*campos: str) -> dict:
        campos_limpieza = campos or FECHAS_FLUJO_BOM
        return {campo: None for campo in campos_limpieza}

    _ORDEN_AREA_POR_ETAPA = {
        "BORRADOR": ("ingenieria", "construccion", "compras"),
        "EN_REVISION_ING": ("ingenieria", "construccion", "compras"),
        "APROBADO_ING": ("ingenieria", "construccion", "compras"),
        "EN_REVISION_OBRA": ("construccion", "ingenieria", "compras"),
        "EN_REVISION_CONST": ("construccion", "ingenieria", "compras"),
        "APROBADO_CONST": ("construccion", "ingenieria", "compras"),
        "EN_REVISION_FINAL": ("construccion", "ingenieria", "compras"),
        "APROBADO_FINAL": ("construccion", "compras", "ingenieria"),
    }

    @classmethod
    def resolver_area_editor(cls, context: dict, bom: Optional[dict] = None) -> str:
        """Resuelve el area por la etapa activa, sin prioridad fija entre modulos."""
        role = context.get("role")
        module_roles = context.get("module_roles", {})
        areas = {
            area for area in ("ingenieria", "construccion", "compras")
            if role == "ADMIN" or module_roles.get(area) in ("editor", "admin")
        }
        estado = bom.get("estatus") if bom else None
        for area in cls._ORDEN_AREA_POR_ETAPA.get(
            estado, ("ingenieria", "construccion", "compras")
        ):
            if area in areas:
                return area
        return "viewer"

    @staticmethod
    def puede_editar_grupos(area_editor: str, estatus_bom: Optional[str]) -> bool:
        """Compras nunca clasifica items en Grupos BOM (es criterio tecnico de
        Ingenieria/Construccion): solo esas dos areas pueden tocar grupo_ids, y
        solo en la etapa que les corresponde. Fuente unica para el modal de
        edicion (checkboxes habilitados/deshabilitados) y el PATCH que aplica
        el cambio -- deben coincidir o el checkbox miente sobre lo que se puede
        guardar."""
        return (
            (area_editor == "ingenieria" and estatus_bom != "APROBADO_FINAL")
            or (
                area_editor == "construccion"
                and estatus_bom in {"EN_REVISION_OBRA", "EN_REVISION_CONST", "APROBADO_FINAL"}
            )
        )

    async def get_capacidades_bom(
        self, conn, bom: dict, user_id: Optional[UUID],
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        """Matriz unica de turno para una version exacta del BOM.

        Los permisos de modulo habilitan la clase de operacion; el turno y el
        ownership del paquete deciden si esa operacion puede ejecutarse.
        """
        capacidades = {
            "editar_base": False,
            "editar_ejecucion": False,
            "agregar_items": False,
            "eliminar_items": False,
            "restaurar_items": False,
            "editar_grupos": False,
            "refrescar_costos": False,
            "actor_turno": None,
            "es_cabeza_trabajo": bool(bom and bom.get("es_cabeza_trabajo")),
            "es_cabeza_oficial": bool(bom and bom.get("es_cabeza_oficial")),
        }
        if not bom or not user_id:
            return capacidades

        estado = EstatusBOM(bom["estatus"])
        activo = bom.get("estado_paquete", "ACTIVO") == "ACTIVO"
        es_cabeza_trabajo = bool(bom.get("es_cabeza_trabajo", True))
        es_cabeza_oficial = bool(bom.get("es_cabeza_oficial"))
        es_oficial_downstream = (
            estado == EstatusBOM.APROBADO_FINAL and es_cabeza_oficial
        )
        if (
            not activo
            or estado == EstatusBOM.CANCELADO
            or not (es_cabeza_trabajo or es_oficial_downstream)
        ):
            return capacidades

        roles_modulo = module_roles or {}
        representados = await self.get_titulares_que_representa(conn, user_id)

        ingeniero = bom.get("ingeniero_responsable_id") or bom.get("elaborado_por")
        ri = bom.get("responsable_ing")
        coordinador = bom.get("coordinador_obra")
        responsable_const = bom.get("jefe_construccion")
        direccion = await self.db.get_aprobador_final_id(conn)

        actores_por_estado = {
            EstatusBOM.BORRADOR: ({ingeniero, ri} - {None}, "Ingenieria"),
            EstatusBOM.EN_REVISION_ING: ({ri} - {None}, "Responsable de Ingenieria"),
            EstatusBOM.APROBADO_ING: ({ri} - {None}, "Responsable de Ingenieria"),
            EstatusBOM.EN_REVISION_OBRA: (
                {coordinador, responsable_const} - {None},
                "Coordinacion de Obra / Responsable de Construccion",
            ),
            EstatusBOM.EN_REVISION_CONST: (
                {responsable_const} - {None}, "Responsable de Construccion"
            ),
            EstatusBOM.APROBADO_CONST: (
                {responsable_const} - {None}, "Responsable de Construccion"
            ),
            EstatusBOM.EN_REVISION_FINAL: ({direccion} - {None}, "Direccion"),
            EstatusBOM.APROBADO_FINAL: (set(), "Operacion downstream"),
        }
        actores, turno = actores_por_estado.get(estado, (set(), None))
        capacidades["actor_turno"] = turno
        controla_turno = bool(representados & actores)

        tiene_ing = roles_modulo.get("ingenieria") in {"editor", "admin"}
        tiene_const = roles_modulo.get("construccion") in {"editor", "admin"}
        tiene_compras = roles_modulo.get("compras") in {"editor", "admin"}
        # Compatibilidad de llamadas internas existentes: el ownership sigue
        # siendo obligatorio aunque el caller aun no transporte module_roles.
        tiene_base = tiene_ing or tiene_const or not module_roles

        if estado not in {
            EstatusBOM.EN_REVISION_FINAL,
            EstatusBOM.APROBADO_FINAL,
            EstatusBOM.CANCELADO,
        } and controla_turno and tiene_base:
            capacidades.update({
                "editar_base": True,
                "agregar_items": True,
                "eliminar_items": True,
                "restaurar_items": True,
                "editar_grupos": True,
                "refrescar_costos": True,
            })

        capacidades["editar_ejecucion"] = (
            es_oficial_downstream and (tiene_const or tiene_compras)
        ) or (
            estado != EstatusBOM.APROBADO_FINAL
            and controla_turno
            and tiene_base
        )
        capacidades["aprobar_final"] = (
            estado == EstatusBOM.EN_REVISION_FINAL
            and direccion in representados
        )
        return capacidades

    @staticmethod
    def _resolver_revision(lock_version_esperado: Optional[int]) -> int:
        """Exige la revision esperada por el caller; sin ella no hay CAS real."""
        if lock_version_esperado is None:
            raise ValueError(
                "Falta la revision de la adenda; recarga el BOM e intenta de nuevo"
            )
        return lock_version_esperado

    async def _reservar_mutacion_base(
        self, conn, id_bom: UUID, user_id: UUID,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> tuple[dict, dict]:
        """Valida cabeza, turno y revision antes de una mutacion de base.

        Retorna (bom, capacidades): el caller puede reusar `capacidades` para la
        respuesta post-commit en vez de recalcularla — ningun campo del que
        depende get_capacidades_bom (estatus/actores/cabeza de paquete) cambia
        por las mutaciones que pasan por aqui.
        """
        bom = await self.db.get_bom_for_update(conn, id_bom)
        if not bom:
            raise ValueError("BOM no encontrado")
        if not bom.get("es_cabeza_trabajo", True):
            raise ValueError("Esta version es historica; abre la cabeza de trabajo del paquete")
        capacidades = await self.get_capacidades_bom(
            conn, bom, user_id, user_role, rol_org, module_roles
        )
        if not capacidades["editar_base"]:
            turno = capacidades.get("actor_turno") or "el actor asignado"
            raise ValueError(f"Solo {turno} puede modificar el BOM en esta etapa")
        if lock_version_esperado is None:
            raise ValueError(
                "Falta la revision del BOM; recarga el paquete e intenta de nuevo"
            )
        revision = lock_version_esperado
        reservado = await self.db.incrementar_lock_bom_cas(
            conn, id_bom, revision, bom["estatus"]
        )
        if not reservado:
            raise ValueError(
                "El BOM cambio desde que abriste el formulario; recarga el paquete e intenta de nuevo"
            )
        return {**bom, **reservado}, capacidades

    async def _validar_actor_asignado(
        self, conn, user_id: UUID, user_role: Optional[str],
        responsables: Set[Optional[UUID]], label: str,
    ) -> None:
        actores = responsables - {None}
        representados = await self.get_titulares_que_representa(conn, user_id)
        if not actores or not (actores & representados):
            raise ValueError(f"Solo {label} (o su suplente) puede ejecutar esta accion")

    async def _transicionar_bom(
        self, conn, id_bom: UUID, user_id: UUID,
        estado_esperado: EstatusBOM, nuevo_estado: EstatusBOM,
        tipo_aprobacion: TipoAprobacion,
        comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
        invalidar_ciclo: bool = False,
        destino_rechazo: Optional[str] = None,
        **campos,
    ) -> dict:
        """Aplica estado, aprobacion, auditoria y outbox en una transaccion."""
        async with self._transaction(conn):
            bom = await self.db.get_bom_for_update(conn, id_bom)
            if not bom:
                raise ValueError("BOM no encontrado")
            if not bom.get("es_cabeza_trabajo", True):
                raise ValueError("Solo la cabeza de trabajo puede cambiar de estado")
            if bom.get("estado_paquete", "ACTIVO") != "ACTIVO":
                raise ValueError("El paquete no esta activo")
            if EstatusBOM(bom["estatus"]) != estado_esperado:
                raise ValueError(
                    f"El BOM debe estar {estado_esperado.value} para ejecutar esta accion"
                )
            if lock_version_esperado is None:
                raise ValueError(
                    "Falta la revision del BOM; recarga el paquete e intenta de nuevo"
                )
            revision = lock_version_esperado
            actualizado = await self.db.update_bom_estatus_cas(
                conn, id_bom, estado_esperado.value, revision,
                nuevo_estado.value, **campos,
            )
            if not actualizado:
                raise ValueError(
                    "El BOM cambio desde que abriste la pagina; recarga el paquete e intenta de nuevo"
                )
            if invalidar_ciclo:
                await self.db.invalidar_aprobaciones_vigentes(conn, id_bom, user_id)
            await self.db.registrar_aprobacion(
                conn, id_bom, tipo_aprobacion,
                bom["version"], user_id, bom["id_paquete"],
                comentarios=comentarios,
                destino_rechazo=destino_rechazo,
            )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM:{id_bom}:{actualizado['lock_version']}:{tipo_aprobacion.value}",
                tipo_aprobacion.value,
                bom["id_proyecto"],
                user_id,
                {
                    "version": bom["version"],
                    "estado_anterior": estado_esperado.value,
                    "estado_nuevo": nuevo_estado.value,
                    "comentarios": comentarios,
                },
                id_paquete=bom.get("id_paquete"),
                id_bom=id_bom,
            )
        return await self.db.get_bom_by_id(conn, id_bom)

    async def puede_crear_o_retomar_bom(
        self, conn, id_proyecto: UUID, user_id: Optional[UUID],
        ingeniero_asignado: Optional[dict] = None,
    ) -> bool:
        """Permite BOM solo a jefe de Ingenieria o ingeniero asignado al proyecto.

        Si el caller ya resolvio la asignacion de ingeniero_asignado (p.ej. para
        mostrar un mensaje), puede pasarla para evitar una consulta adicional.
        """
        if not user_id:
            return False
        if await self.db.usuario_tiene_rol_org(conn, user_id, "jefe_ingenieria"):
            return True
        if ingeniero_asignado is not None:
            return str(ingeniero_asignado["id_usuario"]) == str(user_id)
        return await self.db.usuario_tiene_asignacion_proyecto(
            conn, id_proyecto, user_id, "ingeniero_asignado", "INGENIERIA"
        )

    async def puede_administrar_paquete(
        self, conn, id_proyecto: UUID, user_id: Optional[UUID],
        role: Optional[str] = None,
    ) -> bool:
        """ADMIN global u owner/suplente del jefe de Ingenieria del proyecto."""
        if role == "ADMIN":
            return True
        if not user_id:
            return False
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, id_proyecto, "jefe_ingenieria"
        )
        if not responsable:
            return False
        representados = await self.get_titulares_que_representa(conn, user_id)
        return responsable["id_usuario"] in representados

    async def get_permiso_configurar_paneles(
        self, conn, id_proyecto: UUID, user_id: Optional[UUID],
    ) -> tuple[bool, Optional[str]]:
        """Resuelve si el usuario puede configurar el panel FV y, si no, el label del jefe a contactar."""
        puede_configurar = await self.puede_crear_o_retomar_bom(conn, id_proyecto, user_id)
        jefe_label = await self.get_jefe_ingenieria_label(conn) if not puede_configurar else None
        return puede_configurar, jefe_label

    async def get_ingeniero_asignado(self, conn, id_proyecto: UUID) -> Optional[dict]:
        """Asignacion activa de ingeniero_asignado (INGENIERIA) del proyecto."""
        return await self.db.get_asignacion_proyecto(
            conn, id_proyecto, "ingeniero_asignado", "INGENIERIA"
        )

    async def _validar_retomar_bom_ingenieria(
        self, conn, id_proyecto: UUID, user_id: UUID
    ) -> None:
        if not await self.puede_crear_o_retomar_bom(conn, id_proyecto, user_id):
            raise ValueError(
                "Solo el jefe de Ingenieria o el ingeniero asignado pueden crear o retomar el BOM"
            )

    async def _validar_aprobador_bom(
        self, conn, user_id: UUID, user_role: str, rol_org: Optional[str],
        responsable_id: Optional[UUID], label: str, fallback_rol_org: str,
        representados: Optional[Set[UUID]] = None,
        permitir_admin_gestion: bool = False,
    ) -> None:
        if not fallback_rol_org:
            raise ValueError(
                "_validar_aprobador_bom requiere fallback_rol_org: sin el, un BOM con "
                "responsable_id=None y gestion_solo_responsable=True quedaria sin validar"
            )
        if permitir_admin_gestion and user_role == 'ADMIN':
            return
        solo_responsable = await ConfigService.get_global_config(
            conn, 'bom.gestion_solo_responsable', True, bool
        )
        if solo_responsable and responsable_id:
            # El titular del rol o su suplente activo pueden ejecutar la accion
            if representados is None:
                representados = await self.get_titulares_que_representa(conn, user_id)
            if responsable_id not in representados:
                raise ValueError(
                    f"Solo el {label} del proyecto (o su suplente) puede ejecutar esta accion"
                )
            return  # responsable_id=None cae al fallback de rol global
        if not await self.db.usuario_tiene_rol_org(conn, user_id, fallback_rol_org):
            raise ValueError(f"Solo el {label} puede ejecutar esta accion")

    async def puede_aprobar_bom(
        self, conn, user_id: UUID, user_role: str, rol_org: Optional[str],
        responsable_id: Optional[UUID], fallback_rol_org: str,
        representados: Optional[Set[UUID]] = None
    ) -> bool:
        """Version booleana de _validar_aprobador_bom para la UI (no lanza).

        Permite que el template oculte botones que el service rechazaria, usando
        exactamente la misma logica (propiedad, suplencia y fallback de rol
        global). Acepta `representados` precalculado para no repetir la consulta de
        suplencias dentro del mismo render.
        """
        try:
            await self._validar_aprobador_bom(
                conn, user_id, user_role, rol_org, responsable_id, "",
                fallback_rol_org, representados=representados
            )
            return True
        except ValueError:
            return False

    # ─── CREAR BOM ──────────────────────────────────────────

    async def crear_paquete(
        self, conn, id_proyecto: UUID, creado_por: UUID,
        tipo_alcance: str, nombre: str,
        descripcion_alcance: Optional[str] = None,
        notas: Optional[str] = None,
        user_role: Optional[str] = None,
        aceptar_responsabilidad: bool = False,
        clave_idempotencia: Optional[str] = None,
    ) -> dict:
        """Crea paquete, v1 y cabeza de trabajo de forma atomica."""
        proyecto = await self.db.get_proyecto_info(conn, id_proyecto)
        if not proyecto:
            raise ValueError("Proyecto no encontrado")
        if user_role == "ADMIN":
            if not aceptar_responsabilidad:
                raise ValueError(
                    "Confirma que asumirás la responsabilidad de Ingeniería del paquete"
                )
        else:
            await self._validar_retomar_bom_ingenieria(conn, id_proyecto, creado_por)

        tipo = (tipo_alcance or "").strip().upper()
        nombre_limpio = (nombre or "").strip()
        alcance_limpio = (descripcion_alcance or "").strip() or None
        if tipo not in {"COMPLETO", "PARCIAL"}:
            raise ValueError("El tipo de paquete debe ser COMPLETO o PARCIAL")
        if not nombre_limpio:
            raise ValueError("El nombre del paquete es obligatorio")
        if tipo == "PARCIAL" and not alcance_limpio:
            raise ValueError("Describe el alcance del paquete parcial")
        clave_limpia = (clave_idempotencia or "").strip()
        if not clave_limpia or len(clave_limpia) > 120:
            raise ValueError(
                "Falta la clave de reintento del formulario; recarga la pagina"
            )

        try:
            responsable = await self.db.get_responsable_proyecto_o_global(
                conn, id_proyecto, "jefe_ingenieria", estricto=True
            )
        except ValueError:
            raise ValueError(
                "Hay más de un jefe de Ingeniería activo y el proyecto aún no tiene "
                "Responsable asignado. Un Director debe asignarlo en 'Equipo del Proyecto' "
                "antes de crear el BOM."
            ) from None
        if not responsable:
            raise ValueError("No hay Responsable de Ingenieria activo configurado")
        try:
            jefe_const = await self.db.get_responsable_proyecto_o_global(
                conn, id_proyecto, "jefe_construccion", estricto=True
            )
        except ValueError:
            raise ValueError(
                "Hay más de un jefe de Construcción activo y el proyecto aún no tiene "
                "Responsable asignado. Un Director debe asignarlo en 'Equipo del Proyecto' "
                "antes de crear el BOM."
            ) from None
        coordinador = await self.db.get_asignacion_proyecto(
            conn, id_proyecto, "coordinador_obra", "CONSTRUCCION"
        )

        async with conn.transaction():
            estado = await self.db.get_estado_proyecto_for_update(conn, id_proyecto)
            existente = await self.db.get_paquete_por_clave_idempotencia(
                conn, id_proyecto, clave_limpia
            )
            if existente:
                if existente["creado_por"] != creado_por:
                    raise ValueError("La clave de reintento pertenece a otra solicitud")
                bom_existente = await self.db.get_bom_by_id(
                    conn, existente["cabeza_trabajo_id"]
                )
                if not bom_existente:
                    raise ValueError(
                        "El paquete existente no tiene una cabeza de trabajo valida"
                    )
                return bom_existente
            paquetes = await self.db.listar_paquetes_proyecto(conn, id_proyecto)
            multi_habilitado = await ConfigService.get_global_config(
                conn, "bom.multi_paquete_habilitado", False, bool
            )
            if paquetes and not multi_habilitado:
                raise ValueError(
                    "La creacion de multiples paquetes BOM aun no esta habilitada"
                )
            if estado["captura_cerrada"]:
                raise ValueError(
                    "La captura de paquetes BOM esta cerrada; reabre el conjunto antes de continuar"
                )
            paquetes_activos = [
                p for p in paquetes if p.get("estado_paquete") == "ACTIVO"
            ]
            tipos_activos = {p["tipo_alcance"] for p in paquetes_activos}
            if tipo == "COMPLETO" and paquetes_activos:
                raise ValueError(
                    "No se puede crear un BOM completo mientras existan otros paquetes activos"
                )
            if tipo == "PARCIAL" and "COMPLETO" in tipos_activos:
                raise ValueError(
                    "No se puede crear un paquete parcial mientras exista un BOM completo activo"
                )

            codigo = await self.db.get_siguiente_codigo_paquete(conn, id_proyecto)
            paquete = await self.db.crear_paquete(
                conn,
                id_proyecto,
                codigo,
                nombre_limpio,
                tipo,
                alcance_limpio,
                creado_por,
                creado_por,
                responsable["id_usuario"],
                coordinador["id_usuario"] if coordinador else None,
                jefe_const["id_usuario"] if jefe_const else None,
                clave_limpia,
            )
            bom = await self.db.crear_bom(
                conn,
                id_proyecto,
                creado_por,
                responsable_ing=responsable["id_usuario"],
                jefe_construccion=jefe_const["id_usuario"] if jefe_const else None,
                coordinador_obra=coordinador["id_usuario"] if coordinador else None,
                notas=notas,
                version=1,
                id_paquete=paquete["id_paquete"],
                ingeniero_responsable_id=creado_por,
            )
            cabeza = await self.db.actualizar_cabeza_trabajo(
                conn, paquete["id_paquete"], bom["id_bom"], paquete["lock_version"]
            )
            if not cabeza:
                raise ValueError(
                    "El paquete cambio mientras se creaba; actualiza e intenta de nuevo"
                )
            await self.db.registrar_historial(
                conn, bom["id_bom"], AccionHistorial.CREADO, 1, creado_por
            )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM:{bom['id_bom']}:0:PAQUETE_CREADO",
                "PAQUETE_CREADO",
                id_proyecto,
                creado_por,
                {
                    "codigo": codigo,
                    "nombre": nombre_limpio,
                    "tipo_alcance": tipo,
                    "version": 1,
                },
                id_paquete=paquete["id_paquete"],
                id_bom=bom["id_bom"],
            )

        logger.info(
            "Paquete BOM creado: proyecto=%s paquete=%s bom=%s actor=%s",
            id_proyecto, paquete["id_paquete"], bom["id_bom"], creado_por,
        )
        return await self.db.get_bom_by_id(conn, bom["id_bom"])

    async def crear_bom(
        self, conn, id_proyecto: UUID, elaborado_por: UUID,
        responsable_ing: Optional[UUID] = None,
        jefe_construccion: Optional[UUID] = None,
        coordinador_obra: Optional[UUID] = None,
        notas: Optional[str] = None,
    ) -> dict:
        """Compatibilidad: crea un paquete completo con su v1."""
        return await self.crear_paquete(
            conn,
            id_proyecto,
            elaborado_por,
            "COMPLETO",
            "BOM completo",
            notas=notas,
            clave_idempotencia=str(uuid4()),
        )

    # ─── OBTENER BOM ────────────────────────────────────────

    async def listar_paquetes(self, conn, id_proyecto: UUID) -> list[dict]:
        return await self.db.listar_paquetes_proyecto(conn, id_proyecto)

    async def get_estado_conjunto(self, conn, id_proyecto: UUID) -> dict:
        estado = await self.db.get_estado_proyecto(conn, id_proyecto)
        return estado or {
            "id_proyecto": id_proyecto,
            "captura_cerrada": False,
            "lock_version": 0,
            "motivo": None,
        }

    async def resolver_tipo_cambio(
        self, conn, id_proyecto: UUID, estado_proyecto: Optional[dict] = None,
    ) -> dict:
        """Punto unico de resolucion del TC "vigente" a nivel proyecto.

        Cadena: TC manual del proyecto (si esta fijado) > Banxico mas reciente >
        promedio 7 dias. NO considera el TC del XML de factura por item — esa prioridad
        superior la resuelve el caller (get_items/get_item) antes de caer aqui.

        Args:
            estado_proyecto: fila ya leida de tb_bom_proyecto_estado (evita una query
                duplicada en callers que ya la obtuvieron, ej. _get_consolidado_proyecto_snapshot).

        Returns:
            {"tasa": Decimal | None, "origen": "manual" | "banxico" | "promedio" | None,
             "fecha": date | None}
        """
        if estado_proyecto is None:
            estado_proyecto = await self.db.get_estado_proyecto(conn, id_proyecto)

        tc_manual = estado_proyecto.get("tipo_cambio_manual") if estado_proyecto else None
        if tc_manual:
            fijado_en = estado_proyecto.get("tipo_cambio_manual_fijado_en")
            return {
                "tasa": Decimal(str(tc_manual)), "origen": "manual",
                "fecha": fijado_en.date() if fijado_en else today_mx(),
            }

        from core.tipo_cambio.db_service import TipoCambioDBService
        tc_svc = TipoCambioDBService()
        tasa = await tc_svc.get_tasa_mas_reciente(conn)
        if tasa:
            return {
                "tasa": Decimal(str(tasa["tasa_mxn"])), "origen": "banxico",
                "fecha": tasa.get("fecha"),
            }

        promedio = await self.db.get_tasa_promedio(conn)
        if promedio:
            return {"tasa": promedio, "origen": "promedio", "fecha": today_mx()}

        return {"tasa": None, "origen": None, "fecha": None}

    async def fijar_tipo_cambio_manual(
        self, conn, id_proyecto: UUID, tasa: Decimal, actor_id: UUID,
        lock_version_esperado: int,
    ) -> dict:
        """Fija el TC manual del proyecto (CEO/ADMIN). RBAC se valida en el router."""
        if tasa <= 0:
            raise ValueError("El tipo de cambio manual debe ser mayor a 0")
        async with conn.transaction():
            estado = await self.db.get_estado_proyecto_for_update(conn, id_proyecto)
            if estado["lock_version"] != lock_version_esperado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
            actualizado = await self.db.set_tipo_cambio_manual_cas(
                conn, id_proyecto, lock_version_esperado, tasa, actor_id,
            )
            if not actualizado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
        logger.info(
            "TC manual fijado: proyecto=%s tasa=%s actor=%s", id_proyecto, tasa, actor_id,
        )
        return actualizado

    async def quitar_tipo_cambio_manual(
        self, conn, id_proyecto: UUID, actor_id: UUID, lock_version_esperado: int,
    ) -> dict:
        """Quita el TC manual del proyecto; vuelve a Banxico/promedio. RBAC en el router."""
        async with conn.transaction():
            estado = await self.db.get_estado_proyecto_for_update(conn, id_proyecto)
            if estado["lock_version"] != lock_version_esperado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
            actualizado = await self.db.limpiar_tipo_cambio_manual_cas(
                conn, id_proyecto, lock_version_esperado,
            )
            if not actualizado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
        logger.info("TC manual quitado: proyecto=%s actor=%s", id_proyecto, actor_id)
        return actualizado

    async def get_metricas_paneles(self, conn, id_proyecto: UUID) -> dict:
        return await self.db.get_metricas_paneles_proyecto(conn, id_proyecto)

    async def get_paquete(self, conn, id_paquete: UUID) -> dict:
        paquete = await self.db.get_paquete_by_id(conn, id_paquete)
        if not paquete:
            raise ValueError("Paquete BOM no encontrado")
        return paquete

    async def get_bom_cabeza_trabajo(self, conn, id_paquete: UUID) -> dict:
        bom = await self.db.get_bom_cabeza_trabajo(conn, id_paquete)
        if not bom:
            raise ValueError("El paquete no tiene una cabeza de trabajo valida")
        return bom

    async def get_bom_cabeza_oficial(self, conn, id_paquete: UUID) -> Optional[dict]:
        return await self.db.get_bom_cabeza_oficial(conn, id_paquete)

    async def get_versiones_paquete(self, conn, id_paquete: UUID) -> list[dict]:
        return await self.db.listar_versiones_paquete(conn, id_paquete)

    async def cambiar_captura_paquetes(
        self, conn, id_proyecto: UUID, actor_id: UUID,
        user_role: str, rol_org: Optional[str], captura_cerrada: bool,
        lock_version_esperado: int, motivo: str,
    ) -> dict:
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, id_proyecto, "jefe_ingenieria"
        )
        await self._validar_aprobador_bom(
            conn,
            actor_id,
            user_role,
            rol_org,
            responsable["id_usuario"] if responsable else None,
            "Responsable de Ingenieria",
            "jefe_ingenieria",
            permitir_admin_gestion=True,
        )
        motivo_limpio = (motivo or "").strip()
        if not motivo_limpio:
            raise ValueError("El motivo es obligatorio")
        async with conn.transaction():
            estado = await self.db.get_estado_proyecto_for_update(conn, id_proyecto)
            if estado["lock_version"] != lock_version_esperado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
            metricas_fv = (
                await self.db.get_metricas_paneles_proyecto(conn, id_proyecto)
                if captura_cerrada else {}
            )
            actualizado = await self.db.actualizar_captura_proyecto_cas(
                conn,
                id_proyecto,
                lock_version_esperado,
                captura_cerrada,
                actor_id,
                motivo_limpio,
                metricas_fv.get("modulos_fv"),
                metricas_fv.get("potencia_pico_kwp"),
            )
            if not actualizado:
                raise ValueError(
                    "El estado del conjunto BOM cambio; actualiza la pagina e intenta de nuevo"
                )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-PROYECTO:{id_proyecto}:{actualizado['lock_version']}:CAPTURA",
                "CAPTURA_CERRADA" if captura_cerrada else "CAPTURA_REABIERTA",
                id_proyecto,
                actor_id,
                {
                    "motivo": motivo_limpio,
                    "modulos_fv_snapshot": metricas_fv.get("modulos_fv"),
                    "potencia_pico_kwp_snapshot": str(
                        metricas_fv.get("potencia_pico_kwp") or ""
                    ),
                },
            )
        return actualizado

    async def cambiar_estado_paquete(
        self, conn, id_paquete: UUID, actor_id: UUID,
        user_role: str, rol_org: Optional[str], nuevo_estado: str,
        lock_version_esperado: int, motivo: str,
    ) -> dict:
        """Archiva o cancela sin eliminar la trazabilidad del paquete."""
        if nuevo_estado not in {"ACTIVO", "ARCHIVADO", "CANCELADO"}:
            raise ValueError("Estado de paquete invalido")
        motivo_limpio = (motivo or "").strip()
        if not motivo_limpio:
            raise ValueError("El motivo es obligatorio")
        paquete = await self.get_paquete(conn, id_paquete)
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, paquete["id_proyecto"], "jefe_ingenieria"
        )
        await self._validar_aprobador_bom(
            conn, actor_id, user_role, rol_org,
            responsable["id_usuario"] if responsable else None,
            "Responsable de Ingenieria", "jefe_ingenieria",
            permitir_admin_gestion=True,
        )
        async with conn.transaction():
            bloqueado = await self.db.get_paquete_for_update(conn, id_paquete)
            if not bloqueado:
                raise ValueError("Paquete BOM no encontrado")
            if nuevo_estado == "ACTIVO":
                paquetes = await self.db.listar_paquetes_proyecto(
                    conn, bloqueado["id_proyecto"]
                )
                otros_activos = [
                    p for p in paquetes
                    if p["id_paquete"] != id_paquete
                    and p.get("estado_paquete") == "ACTIVO"
                ]
                if bloqueado["tipo_alcance"] == "COMPLETO" and otros_activos:
                    raise ValueError(
                        "Un BOM completo no puede reactivarse junto con otros paquetes activos"
                    )
                if bloqueado["tipo_alcance"] != "COMPLETO" and any(
                    p.get("tipo_alcance") == "COMPLETO" for p in otros_activos
                ):
                    raise ValueError(
                        "No se puede reactivar el paquete mientras exista un BOM completo activo"
                    )
            if nuevo_estado == "CANCELADO":
                actividad = await self.db.get_actividad_downstream_paquete(conn, id_paquete)
                if bloqueado.get("cabeza_oficial_id") or actividad:
                    raise ValueError(
                        "El paquete tiene una version oficial o actividad downstream; debe archivarse"
                    )
            actualizado = await self.db.actualizar_estado_paquete_cas(
                conn, id_paquete, lock_version_esperado,
                bloqueado["estado_paquete"], nuevo_estado,
            )
            if not actualizado:
                raise ValueError(
                    "El paquete cambio desde que abriste la pagina; recarga e intenta de nuevo"
                )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-PAQUETE:{id_paquete}:{actualizado['lock_version']}:ESTADO",
                f"PAQUETE_{nuevo_estado}",
                bloqueado["id_proyecto"],
                actor_id,
                {"motivo": motivo_limpio, "estado_anterior": bloqueado["estado_paquete"]},
                id_paquete=id_paquete,
                id_bom=bloqueado.get("cabeza_trabajo_id"),
            )
        return actualizado

    async def reclasificar_paquete(
        self, conn, id_paquete: UUID, actor_id: UUID,
        user_role: str, rol_org: Optional[str], tipo_alcance: str,
        nombre: str, descripcion_alcance: Optional[str],
        lock_version_esperado: int, motivo: str,
    ) -> dict:
        """Cambia metadatos de alcance con validacion de coexistencia y auditoria."""
        tipo = (tipo_alcance or "").strip().upper()
        nombre_limpio = (nombre or "").strip()
        alcance = (descripcion_alcance or "").strip() or None
        motivo_limpio = (motivo or "").strip()
        if tipo not in {"COMPLETO", "PARCIAL", "LEGACY"}:
            raise ValueError("Tipo de alcance invalido")
        if not nombre_limpio:
            raise ValueError("El nombre del paquete es obligatorio")
        if tipo == "PARCIAL" and not alcance:
            raise ValueError("Describe el alcance del paquete parcial")
        if not motivo_limpio:
            raise ValueError("El motivo es obligatorio")

        paquete = await self.get_paquete(conn, id_paquete)
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, paquete["id_proyecto"], "jefe_ingenieria"
        )
        await self._validar_aprobador_bom(
            conn, actor_id, user_role, rol_org,
            responsable["id_usuario"] if responsable else None,
            "Responsable de Ingenieria", "jefe_ingenieria",
            permitir_admin_gestion=True,
        )

        async with conn.transaction():
            bloqueado = await self.db.get_paquete_for_update(conn, id_paquete)
            if not bloqueado or bloqueado["estado_paquete"] == "CANCELADO":
                raise ValueError("El paquete no esta disponible para reclasificacion")
            if bloqueado["lock_version"] != lock_version_esperado:
                raise ValueError(
                    "El paquete cambio desde que abriste la pagina; recarga e intenta de nuevo"
                )
            paquetes = await self.db.listar_paquetes_proyecto(
                conn, bloqueado["id_proyecto"]
            )
            otros_activos = [
                p for p in paquetes
                if p["id_paquete"] != id_paquete
                and p.get("estado_paquete") == "ACTIVO"
            ]
            if bloqueado["estado_paquete"] == "ACTIVO":
                if tipo == "COMPLETO" and otros_activos:
                    raise ValueError(
                        "Archiva los otros paquetes antes de clasificar este BOM como completo"
                    )
                if tipo != "COMPLETO" and any(
                    p.get("tipo_alcance") == "COMPLETO" for p in otros_activos
                ):
                    raise ValueError(
                        "No puede coexistir un paquete parcial con un BOM completo activo"
                    )
            actualizado = await self.db.reclasificar_paquete_cas(
                conn, id_paquete, lock_version_esperado,
                tipo, nombre_limpio, alcance,
            )
            if not actualizado:
                raise ValueError(
                    "El paquete cambio desde que abriste la pagina; recarga e intenta de nuevo"
                )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-PAQUETE:{id_paquete}:{actualizado['lock_version']}:RECLASIFICADO",
                "PAQUETE_RECLASIFICADO", bloqueado["id_proyecto"], actor_id,
                {
                    "motivo": motivo_limpio,
                    "tipo_anterior": bloqueado["tipo_alcance"],
                    "tipo_nuevo": tipo,
                },
                id_paquete=id_paquete,
                id_bom=bloqueado.get("cabeza_trabajo_id"),
            )
        return actualizado

    async def reasignar_paquete(
        self, conn, id_paquete: UUID, actor_id: UUID,
        user_role: str, rol_org: Optional[str], motivo: str,
        lock_version_paquete: int, lock_version_bom: int,
        ingeniero_responsable_id: UUID,
        responsable_ing_id: Optional[UUID],
        coordinador_obra_id: Optional[UUID],
        jefe_construccion_id: Optional[UUID],
    ) -> dict:
        motivo_limpio = (motivo or "").strip()
        if not motivo_limpio:
            raise ValueError("El motivo es obligatorio")
        paquete = await self.get_paquete(conn, id_paquete)
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, paquete["id_proyecto"], "jefe_ingenieria"
        )
        await self._validar_aprobador_bom(
            conn, actor_id, user_role, rol_org,
            responsable["id_usuario"] if responsable else None,
            "Responsable de Ingenieria", "jefe_ingenieria",
            permitir_admin_gestion=True,
        )
        async with conn.transaction():
            bom = await self.db.get_bom_for_update(conn, paquete["cabeza_trabajo_id"])
            if not bom or bom["estatus"] != EstatusBOM.BORRADOR:
                raise ValueError("Solo se puede reasignar un paquete en BORRADOR")
            actualizado = await self.db.reasignar_paquete_borrador_cas(
                conn, id_paquete, bom["id_bom"], lock_version_paquete,
                lock_version_bom, ingeniero_responsable_id, responsable_ing_id,
                coordinador_obra_id, jefe_construccion_id,
            )
            if not actualizado:
                raise ValueError(
                    "El paquete cambio desde que abriste la pagina; recarga e intenta de nuevo"
                )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-PAQUETE:{id_paquete}:{lock_version_paquete + 1}:REASIGNADO",
                "PAQUETE_REASIGNADO",
                paquete["id_proyecto"], actor_id,
                {"motivo": motivo_limpio},
                id_paquete=id_paquete, id_bom=bom["id_bom"],
            )
        return await self.get_paquete(conn, id_paquete)

    async def get_bom(self, conn, id_bom: UUID) -> dict:
        """Obtiene un BOM por ID. Lanza error si no existe."""
        bom = await self.db.get_bom_by_id(conn, id_bom)
        if not bom:
            raise ValueError("BOM no encontrado")
        return bom

    async def get_bom_subtitulo(self, conn, id_bom: UUID) -> dict:
        """Lookup liviano (version + codigo de paquete) para el subtitulo de los
        modales de log. Lanza error si no existe."""
        bom = await self.db.get_bom_subtitulo(conn, id_bom)
        if not bom:
            raise ValueError("BOM no encontrado")
        return bom

    @staticmethod
    def item_sin_costo(item: dict) -> bool:
        """True si el item activo no tiene costo util (oficial) para presupuesto:
        sin precio, o con un precio que Ingenieria capturo pero Compras aun no
        confirma (precio_pendiente_confirmacion) — no cuenta como resuelto.
        Aplica tanto a items BASE como a items de adenda (FUERA_SCOPE/REEMPLAZO):
        estos ultimos solo existen con el BOM en APROBADO_FINAL, pero igual
        necesitan quedar en el radar de Compras si nadie les capturo costo real."""
        if not item.get("activo", True):
            return False
        if item.get("precio_pendiente_confirmacion"):
            return True
        precio = item.get("precio_unitario")
        if precio is None:
            return True
        try:
            return Decimal(str(precio)) <= 0
        except (InvalidOperation, TypeError, ValueError):
            return True

    @staticmethod
    def mensaje_item_sin_costo() -> str:
        return (
            "No puede haber items sin costo. Revisa con Compras para "
            "actualizar el costo antes de avanzar el BOM."
        )

    @staticmethod
    def mensaje_item_costo_pendiente_confirmacion() -> str:
        return "Costo capturado; queda pendiente de que Compras lo confirme."

    @classmethod
    def mensaje_advertencia_costo_item(cls, item: dict) -> str:
        """Elige el mensaje segun el motivo real de item_sin_costo(): si ya se
        capturo un precio y solo falta la confirmacion de Compras, avisar eso
        en vez del mensaje generico de 'No puede haber items sin costo'
        (confuso cuando Ingenieria acaba de capturarlo)."""
        if item.get("precio_pendiente_confirmacion"):
            return cls.mensaje_item_costo_pendiente_confirmacion()
        return cls.mensaje_item_sin_costo()

    @staticmethod
    def mensaje_costo_ya_configurado() -> str:
        return (
            "Este item ya tiene un costo configurado. Si necesitas "
            "corregirlo, dejalo en cero para reabrirlo y que Compras "
            "lo capture de nuevo."
        )

    @staticmethod
    def mensaje_item_agregado(item: dict) -> str:
        return f"'{item['descripcion']}' se agrego al BOM"

    @staticmethod
    def _format_template(template: str, context: dict) -> str:
        return template.format_map(defaultdict(str, context))

    @staticmethod
    def _resumen_costos_pendientes(items: list[dict], limit: int = 5) -> str:
        previews = [
            str(item.get("descripcion") or "Item sin descripcion")
            for item in items[:limit]
        ]
        suffix = f" y {len(items) - limit} mas" if len(items) > limit else ""
        return "; ".join(previews) + suffix

    def _build_costos_pendientes_error(self, items: list[dict]) -> str:
        detalle = self._resumen_costos_pendientes(items)
        return (
            f"No se puede avanzar el BOM: hay {len(items)} item(s) sin costo estimado. "
            "Captura el costo estimado antes de continuar. "
            f"Pendientes: {detalle}."
        )

    async def get_items_sin_costo(self, conn, id_bom: UUID) -> list[dict]:
        """Lista items activos sin costo asignado."""
        return await self.db.get_items_sin_costo_bom(conn, id_bom)

    async def resolver_costos_pendientes_compras(
        self, conn, id_bom: UUID, user_id: UUID, items_payload: List[dict],
    ) -> dict:
        """Compras captura precio_unitario/moneda (presupuesto) sobre items sin costo
        de un BOM. Items BASE: solo en cualquier etapa activa previa a APROBADO_CONST
        (ver ESTATUS_FUERA_DE_PRECIOS_PENDIENTES_COMPRAS) — a partir de ahi los cubre
        la pestaña "Activos" de Compras. Items de adenda (FUERA_SCOPE/REEMPLAZO): SIEMPRE
        que el BOM no este CANCELADO, porque solo existen con el BOM en APROBADO_FINAL
        (dentro del mismo set) y no tienen ningun otro flujo donde resolverse. Este metodo
        no es una variante de `editar_item(area_editor='compras')`: ese metodo redirige precio_unitario/moneda
        a precio_real/moneda_real (costo real, no presupuesto), asi que reusarlo
        escribiria en la columna equivocada sin error visible. La misma trampa aplica
        a `id_material_interno`: `update_item` (BomDBService) lo excluye a proposito de
        su set `allowed` y lanza ValueError si se le pasa — el enlace/creacion de
        material para items de entrada manual (doc 42, BOM 6.1) va SIEMPRE via
        `actualizar_precios_items_compras_cas_batch`, gateado por el mismo CAS que el
        precio, nunca via el editor generico de items.

        CAS a nivel de item (tb_bom_items.lock_version) en vez del CAS de BOM completo
        de `_reservar_mutacion_base` — evita que un guardado de Compras sobre el item X
        invalide el lock_version que Ingenieria tiene cargado editando el item Y.

        `get_bom_for_update` SI se usa (bloquea la fila del BOM/paquete) para releer
        `estatus` fresco dentro de la misma transaccion y serializar contra una
        transicion de estado concurrente — sin la semantica de CAS por version que
        tiene `incrementar_lock_bom_cas` (aqui no se compara ni incrementa lock_version
        de tb_bom, solo se usa el lock de fila como exclusion mutua).

        Patron validar-todo-luego-aplicar-subconjunto (core.materials.service
        `_parse_y_validar`), no todo-o-nada: un item invalido no descarta el lote
        completo, se reporta en `rechazados` con motivo."""
        if not items_payload:
            raise ValueError("No hay items para actualizar")

        async with self._transaction(conn):
            bom = await self.db.get_bom_for_update(conn, id_bom)
            if not bom:
                raise ValueError("BOM no encontrado")
            if bom["estatus"] == EstatusBOM.CANCELADO.value:
                raise ValueError(
                    "El BOM esta cancelado; tus cambios no se aplicaron."
                )
            # A partir de aqui items BASE ya no se resuelven en este flujo (los cubre
            # la pestaña "Activos" de Compras); items de adenda si, por eso el gate se
            # aplica por item mas abajo en vez de cortar toda la funcion.
            bloqueado_para_base = bom["estatus"] in ESTATUS_FUERA_DE_PRECIOS_PENDIENTES_COMPRAS
            if bom.get("estado_paquete") != "ACTIVO":
                raise ValueError(
                    "Este paquete BOM ya no esta activo; tus cambios no se aplicaron."
                )

            ids_solicitados = [entrada["id_item"] for entrada in items_payload]
            items_actuales = {
                item["id_item"]: item
                for item in await self.db.get_items_por_ids_para_bom(
                    conn, id_bom, ids_solicitados
                )
            }

            rechazados: List[dict] = []
            validos: List[dict] = []
            for entrada in items_payload:
                id_item = entrada["id_item"]
                actual = items_actuales.get(id_item)
                # `actual is None` cubre tanto item inexistente como -por el WHERE
                # id_bom=$1 de get_items_por_ids_para_bom- un id_item de OTRO BOM
                # colado en el payload (proteccion IDOR: nunca se toca).
                if actual is None:
                    rechazados.append({"id_item": id_item, "motivo": "Item no encontrado en este BOM"})
                    continue
                if not self.item_sin_costo(actual):
                    rechazados.append({
                        "id_item": id_item,
                        "motivo": "Este item ya no aplica (fue resuelto, no es base, o esta inactivo)",
                    })
                    continue
                es_item_base = (
                    actual.get("tipo_origen_item") or TIPO_ITEM_BASE
                ) == TIPO_ITEM_BASE
                if bloqueado_para_base and es_item_base:
                    rechazados.append({
                        "id_item": id_item,
                        "motivo": (
                            "El BOM ya llego a Construccion aprobada; este item base "
                            "no se resuelve aqui, contacta a Ingenieria."
                        ),
                    })
                    continue

                try:
                    precio = self._decimal_o_error(
                        entrada.get("precio_unitario"), "El precio unitario no puede ser negativo"
                    )
                except ValueError as exc:
                    rechazados.append({"id_item": id_item, "motivo": str(exc)})
                    continue
                if precio <= 0:
                    rechazados.append({"id_item": id_item, "motivo": "El precio debe ser mayor a cero"})
                    continue
                moneda = entrada.get("moneda")
                if moneda not in MONEDAS_VALIDAS:
                    rechazados.append({"id_item": id_item, "motivo": "Selecciona la moneda"})
                    continue

                validos.append({
                    "id_item": id_item, "actual": actual, "precio": precio,
                    "moneda": moneda, "entrada": entrada,
                })

            # Crear/vincular material interno para items sin id_material_interno
            # (entrada manual) ANTES del batch de precios: si el anti-duplicado
            # rechaza la creacion, el item completo se rechaza aqui (no se aplica
            # el precio con un vinculo a medias) — doc 42, BOM 6.1.
            id_materiales_resueltos: dict[UUID, UUID] = {}
            # Savepoint por material recien creado en este intento: si el CAS de
            # precio del item falla mas abajo (lock_version desalineado), se hace
            # rollback del savepoint para no dejar el material huerfano en catalogo
            # -- de lo contrario el reintento lo detecta como "similar" y bloquea
            # la re-creacion (buscar_internos_similares, threshold=0.9).
            savepoints_material_nuevo: dict[UUID, object] = {}
            # Materiales creados mas temprano en ESTE MISMO batch: Postgres los
            # hace visibles a buscar_internos_similares aunque su savepoint no se
            # haya liberado, y si ese savepoint termina en rollback (CAS fallido)
            # no deben contar como "ya existe" para los items siguientes del loop.
            ids_creados_este_intento: List[UUID] = []
            alias_map: Optional[dict] = None
            validos_confirmados: List[dict] = []
            for v in validos:
                entrada, actual, id_item = v["entrada"], v["actual"], v["id_item"]
                if actual.get("id_material_interno"):
                    validos_confirmados.append(v)
                    continue
                id_vinculado = entrada.get("id_material_vinculado")
                if id_vinculado:
                    id_materiales_resueltos[id_item] = id_vinculado
                    validos_confirmados.append(v)
                    continue
                if not entrada.get("crear_catalogo"):
                    validos_confirmados.append(v)
                    continue

                descripcion = (actual.get("descripcion") or "").strip()
                if not descripcion:
                    rechazados.append({
                        "id_item": id_item,
                        "motivo": "El item no tiene descripcion para crear el material",
                    })
                    continue
                similares = await self.materials.buscar_internos_similares(
                    conn, descripcion, threshold=0.9, limit=1,
                    excluir_ids=ids_creados_este_intento,
                )
                if similares:
                    rechazados.append({
                        "id_item": id_item,
                        "motivo": (
                            f"Ya existe un material muy similar en el catalogo: "
                            f"\"{similares[0]['descripcion_canonica']}\". Usa la "
                            "sugerencia de homologacion para vincularlo en vez de "
                            "crear uno nuevo."
                        ),
                    })
                    continue

                if alias_map is None:
                    alias_map = await self.materials.db.get_unidad_alias_map(conn)
                unidad_txt = actual.get("unidad_medida")
                id_unidad = alias_map.get(normalizar_unidad(unidad_txt)) if unidad_txt else None

                savepoint = conn.transaction()
                await savepoint.start()
                nuevo_material = await self.materials.crear_interno(conn, {
                    "descripcion_canonica": descripcion,
                    "id_unidad_medida": id_unidad,
                    "id_categoria": actual.get("id_categoria"),
                    "clave_prod_serv": None,
                    "precio_referencia": v["precio"],
                    "notas": (
                        f"Creado desde BOM {bom.get('proyecto_id_estandar') or id_bom} "
                        f"por Compras, {today_mx().strftime('%d/%m/%Y')}"
                    ),
                    "material": None, "tipo": None, "acabado": None,
                    "marca": None, "adicional": None, "medida": None,
                    "moneda": v["moneda"],
                    "creado_por": user_id, "actualizado_por": user_id,
                }, puede_editar_costos=True)
                id_materiales_resueltos[id_item] = nuevo_material["id"]
                savepoints_material_nuevo[id_item] = savepoint
                ids_creados_este_intento.append(nuevo_material["id"])
                validos_confirmados.append(v)
            validos = validos_confirmados

            # Un solo UPDATE en lote (arrays desanidados) en vez de un CAS por item:
            # ver `actualizar_precios_items_compras_cas_batch`.
            actualizados_ids = set(await self.db.actualizar_precios_items_compras_cas_batch(
                conn,
                [
                    (
                        v["id_item"], v["precio"], v["moneda"],
                        v["entrada"].get("lock_version", -1),
                        id_materiales_resueltos.get(v["id_item"]),
                    )
                    for v in validos
                ],
            ))

            for id_item_material, savepoint in savepoints_material_nuevo.items():
                if id_item_material in actualizados_ids:
                    await savepoint.commit()
                else:
                    await savepoint.rollback()

            aplicados: List[UUID] = []
            historial_entradas: List[tuple] = []
            catalogo_entradas: List[tuple] = []
            for v in validos:
                id_item = v["id_item"]
                if id_item not in actualizados_ids:
                    rechazados.append({
                        "id_item": id_item,
                        "motivo": "Este item fue modificado por alguien mas; recarga e intenta de nuevo",
                    })
                    continue

                actual, precio, moneda = v["actual"], v["precio"], v["moneda"]
                for campo, valor_anterior, valor_nuevo in (
                    ("precio_unitario", actual.get("precio_unitario"), precio),
                    ("moneda", actual.get("moneda"), moneda),
                ):
                    historial_entradas.append((
                        id_bom, id_item, AccionHistorial.EDITADO,
                        CAMPO_LABELS.get(campo, campo),
                        str(valor_anterior) if valor_anterior is not None else None,
                        str(valor_nuevo), bom["version"], user_id,
                    ))
                if id_item in id_materiales_resueltos:
                    historial_entradas.append((
                        id_bom, id_item, AccionHistorial.EDITADO,
                        CAMPO_LABELS.get('id_material_interno', 'id_material_interno'),
                        None, str(id_materiales_resueltos[id_item]), bom["version"], user_id,
                    ))

                if v["entrada"].get("actualizar_catalogo") and actual.get("id_material_interno"):
                    catalogo_entradas.append(
                        (precio, moneda, user_id, actual["id_material_interno"])
                    )

                aplicados.append(id_item)

            # Historial y catalogo tambien en lote (executemany): ver
            # `registrar_historial_batch` / `actualizar_precios_catalogo_interno_batch`.
            await self.db.registrar_historial_batch(conn, historial_entradas)
            await self.db.actualizar_precios_catalogo_interno_batch(conn, catalogo_entradas)

        return {"aplicados": aplicados, "rechazados": rechazados}

    async def validar_sin_costos_pendientes(self, conn, id_bom: UUID) -> None:
        """Bloquea avances del workflow si quedan items activos sin costo."""
        items_sin_costo = await self.get_items_sin_costo(conn, id_bom)
        if items_sin_costo:
            raise ValueError(self._build_costos_pendientes_error(items_sin_costo))

    async def refrescar_costos_catalogo(
        self, conn, id_bom: UUID, user_id: UUID,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        """Sincroniza precio_unitario de items sin costo desde el catalogo interno
        (precio_referencia o factura XML vinculada mas reciente). Permitido para el
        actor de turno en cualquier estado previo al cierre (hasta APROBADO_CONST):
        el costo es un presupuesto aproximado, no bloquea el avance hacia Compras.
        Se congela en EN_REVISION_FINAL/APROBADO_FINAL porque ahi ya afecta a todos
        los actores downstream."""
        bom = await self.get_bom(conn, id_bom)
        async with conn.transaction():
            bom, capacidades = await self._reservar_mutacion_base(
                conn, id_bom, user_id, lock_version_esperado,
                user_role, rol_org, module_roles,
            )
            sincronizados = await self.db.sincronizar_costos_catalogo(conn, id_bom)
            if sincronizados:
                await self.db.registrar_historial(
                    conn, id_bom, AccionHistorial.EDITADO,
                    bom["version"], user_id,
                    campo_modificado="presupuesto_catalogo",
                    valor_nuevo=f"{len(sincronizados)} item(s) sincronizados",
                )
        return {"sincronizados": len(sincronizados), "bom": bom, "capacidades": capacidades}

    async def _get_aprobador_final_direccion_id(self, conn) -> UUID:
        aprobador_id = await self.db.get_aprobador_final_id(conn)
        if not aprobador_id:
            raise ValueError("Configura un aprobador final de Dirección antes de avanzar el BOM")
        if not await self.db.usuario_tiene_rol_org(conn, aprobador_id, "director"):
            raise ValueError("El aprobador final del BOM debe ser un usuario activo de Dirección")
        return aprobador_id

    async def configurar_aprobador_final(self, conn, user_id: Optional[UUID]) -> None:
        if user_id and not await self.db.usuario_tiene_rol_org(conn, user_id, "director"):
            raise ValueError("El aprobador final debe ser un usuario activo de Dirección")
        await self.db.set_aprobador_final_id(conn, user_id)

    async def validar_responsables_workflow_bom(self, conn, bom: dict) -> None:
        """Bloquea el primer envio si falta el responsable inmediato o el aprobador final.

        Coordinador de Obra y Jefe de Construccion no se validan aqui: se resuelven
        en vivo hasta el envio de Ingenieria a Obra (enviar_revision_obra), que es
        cuando realmente se necesitan.
        """
        problemas = []
        if not bom.get("responsable_ing"):
            problemas.append("falta responsable de Ingenieria")
        try:
            await self._get_aprobador_final_direccion_id(conn)
        except ValueError as exc:
            problemas.append(str(exc))

        if problemas:
            raise ValueError(
                "No se puede enviar el BOM a revision: " + "; ".join(problemas)
            )

    # ─── ITEMS CRUD ─────────────────────────────────────────


    async def agregar_item(
        self, conn, id_bom: UUID, user_id: UUID,
        descripcion: str, cantidad, id_categoria: Optional[int] = None,
        unidad_medida: Optional[str] = None,
        comentarios: Optional[str] = None,
        precio_unitario=None,
        origen_precio: Optional[str] = "MANUAL",
        id_material_ref: Optional[UUID] = None,
        id_material_interno: Optional[UUID] = None,
        tipo_partida: Optional[str] = "MATERIAL",
        moneda: Optional[str] = "MXN",
        area_editor: str = "ingenieria",
        grupo_ids: Optional[List[int]] = None,
        grupo_porcentajes: Optional[dict[int, Decimal]] = None,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        """Agrega linea, grupos e historial como una sola mutacion de la cabeza."""
        if self._decimal_o_error(cantidad, "La cantidad debe ser mayor a cero") <= 0:
            raise ValueError("La cantidad debe ser mayor a cero")
        if (
            precio_unitario is not None
            and self._decimal_o_error(
                precio_unitario, "El precio unitario no puede ser negativo"
            ) < 0
        ):
            raise ValueError("El precio unitario no puede ser negativo")
        if not grupo_ids:
            raise ValueError("Selecciona al menos un grupo BOM")
        distribucion_grupos = self._normalizar_distribucion_grupos(
            grupo_ids, grupo_porcentajes
        )

        # Autoridad de edicion de costos (Fase 4): si Ingenieria captura el
        # precio en el mismo alta (no solo al editar despues), el costo queda
        # igual pendiente de que Compras lo confirme -- mismo criterio que
        # editar_item (linea ~2665), replicado aqui para no saltarse el gate.
        precio_pendiente_confirmacion = (
            area_editor == "ingenieria"
            and precio_unitario is not None
            and self._decimal_o_error(precio_unitario, "") > 0
        )

        async with conn.transaction():
            bom, capacidades = await self._reservar_mutacion_base(
                conn, id_bom, user_id, lock_version_esperado,
                user_role, rol_org, module_roles,
            )
            orden = await self.db.get_next_orden(conn, id_bom)
            item = await self.db.agregar_item(
                conn, id_bom, descripcion, cantidad,
                id_categoria=id_categoria,
                unidad_medida=unidad_medida,
                comentarios=comentarios,
                orden=orden,
                precio_unitario=precio_unitario,
                origen_precio=origen_precio,
                id_material_ref=id_material_ref,
                id_material_interno=id_material_interno,
                tipo_partida=tipo_partida,
                moneda=moneda,
                creado_por=user_id,
                precio_pendiente_confirmacion=precio_pendiente_confirmacion,
            )
            await self.db.set_item_grupos(conn, item["id_item"], grupo_ids)
            await self._guardar_distribucion_grupos(
                conn, item["id_item"], distribucion_grupos
            )
            await self.db.registrar_historial(
                conn, id_bom, AccionHistorial.AGREGADO,
                bom["version"], user_id,
                id_item=item["id_item"],
                campo_modificado="item",
                valor_nuevo=descripcion,
            )
        return {"item": item, "capacidades": capacidades}

    @staticmethod
    def _validar_motivo_adenda(motivo: Optional[str]) -> str:
        motivo_limpio = (motivo or "").strip()
        if not motivo_limpio:
            raise ValueError("El motivo de la adenda es obligatorio")
        return motivo_limpio

    @staticmethod
    def _decimal_o_error(valor, mensaje: str) -> Decimal:
        try:
            return Decimal(str(valor))
        except (InvalidOperation, TypeError, ValueError):
            raise ValueError(mensaje) from None

    @staticmethod
    def _normalizar_distribucion_grupos(
        grupo_ids: List[int], porcentajes: Optional[dict[int, Decimal]] = None,
    ) -> dict[int, Decimal]:
        ids = list(dict.fromkeys(grupo_ids))
        if not ids:
            raise ValueError("Selecciona al menos un grupo BOM")
        if len(ids) == 1 and not porcentajes:
            return {ids[0]: Decimal("1")}
        if not porcentajes or set(porcentajes) != set(ids):
            raise ValueError(
                "Indica el porcentaje de cada grupo cuando un item pertenece a varios grupos"
            )
        normalizados = {
            grupo_id: Decimal(str(porcentajes[grupo_id])) for grupo_id in ids
        }
        if any(valor <= 0 or valor > 1 for valor in normalizados.values()):
            raise ValueError(
                "Cada porcentaje de grupo debe ser mayor a cero y menor o igual a 100"
            )
        if abs(sum(normalizados.values()) - Decimal("1")) > Decimal("0.000001"):
            raise ValueError("Los porcentajes de grupo deben sumar 100")
        return normalizados

    async def _guardar_distribucion_grupos(
        self, conn, id_item: UUID, porcentajes: dict[int, Decimal],
    ) -> None:
        guardar = getattr(self.db, "set_distribucion_grupos_item", None)
        if guardar:
            await guardar(conn, id_item, porcentajes)

    @staticmethod
    def _datos_item_adenda(**campos) -> dict:
        """Normaliza datos propuestos de adenda para guardarlos como JSON."""
        data = {}
        for key, value in campos.items():
            if value is None:
                data[key] = None
            elif isinstance(value, UUID):
                data[key] = str(value)
            else:
                data[key] = str(value) if key in {"cantidad", "precio_unitario"} else value
        return data

    @staticmethod
    def _datos_item_desde_adenda(datos: Optional[dict]) -> dict:
        """Convierte JSON de adenda al formato esperado por agregar_item."""
        if isinstance(datos, str):
            data = json.loads(datos or "{}")
        else:
            data = dict(datos or {})
        if data.get("cantidad") is not None:
            data["cantidad"] = BomService._decimal_o_error(
                data["cantidad"], "La cantidad debe ser mayor a cero"
            )
        if data.get("precio_unitario") not in (None, ""):
            data["precio_unitario"] = BomService._decimal_o_error(
                data["precio_unitario"], "El precio unitario no puede ser negativo"
            )
        else:
            data["precio_unitario"] = None
        for key in ("id_material_ref", "id_material_interno"):
            if data.get(key):
                data[key] = UUID(str(data[key]))
        return data

    async def _validar_bom_aprobado_final_para_adenda(
        self, conn, id_bom: UUID
    ) -> dict:
        bom = await self.get_bom(conn, id_bom)
        if EstatusBOM(bom["estatus"]) != EstatusBOM.APROBADO_FINAL:
            raise ValueError("Solo se pueden registrar adendas cuando el BOM esta aprobado final")
        return bom

    async def _validar_item_base_para_adenda(
        self, conn, id_item: UUID
    ) -> tuple[dict, dict]:
        item = await self.db.get_item_by_id(conn, id_item)
        if not item:
            raise ValueError("Item no encontrado")
        if not item.get("activo", True):
            raise ValueError("No se puede registrar una adenda sobre un item eliminado")
        if (item.get("tipo_origen_item") or TIPO_ITEM_BASE) != TIPO_ITEM_BASE:
            raise ValueError("Solo los items base pueden cerrarse o reemplazarse")
        if item.get("estatus_ejecucion") in ESTATUS_ITEM_CERRADO_COMPRA:
            raise ValueError("El item ya esta cerrado para compra")
        if item.get("estatus_compra") in ESTATUS_COMPRA_BLOQUEA_ADENDA:
            raise ValueError(
                "No se puede registrar una adenda sobre un item cotizado, autorizado, pagado o facturado"
            )
        bloqueo = await self.db.get_item_compra_bloqueante(conn, id_item)
        if bloqueo.get("tiene_cotizacion_seleccionada") or bloqueo.get("tiene_autorizacion_activa"):
            raise ValueError(
                "No se puede registrar una adenda mientras el item tenga cotizacion seleccionada o autorizacion activa"
            )
        bom = await self._validar_bom_aprobado_final_para_adenda(conn, item["id_bom"])
        return item, bom

    async def cerrar_item_sin_compra(
        self, conn, id_item: UUID, user_id: UUID, motivo: str
    ) -> dict:
        """Crea una adenda pendiente para cerrar un item base sin compra."""
        motivo = self._validar_motivo_adenda(motivo)
        item, bom = await self._validar_item_base_para_adenda(conn, id_item)

        async with conn.transaction():
            adenda = await self.db.crear_adenda(
                conn, item["id_bom"], "NO_ADQUIRIDO", motivo, user_id
            )
            await self.db.registrar_adenda_item(
                conn, adenda["id_adenda"], item["id_bom"], "NO_ADQUIRIDO", motivo,
                id_item_origen=id_item,
            )
            await self.db.registrar_historial(
                conn, item["id_bom"], AccionHistorial.EDITADO,
                bom["version"], user_id,
                id_item=id_item,
                campo_modificado="adenda",
                valor_nuevo="PENDIENTE_CONSTRUCCION",
            )

        return {**adenda, "id_bom": item["id_bom"]}

    async def crear_reemplazo_item(
        self, conn, id_item_origen: UUID, user_id: UUID,
        descripcion: str, cantidad, grupo_ids: List[int], motivo: str,
        grupo_porcentajes: Optional[dict[int, Decimal]] = None,
        id_categoria: Optional[int] = None,
        unidad_medida: Optional[str] = None,
        comentarios: Optional[str] = None,
        precio_unitario=None,
        origen_precio: Optional[str] = "MANUAL",
        id_material_ref: Optional[UUID] = None,
        id_material_interno: Optional[UUID] = None,
        tipo_partida: Optional[str] = "MATERIAL",
        moneda: Optional[str] = "MXN",
    ) -> dict:
        """Crea una adenda pendiente para sustituir un item base."""
        motivo = self._validar_motivo_adenda(motivo)
        descripcion = (descripcion or "").strip()
        if not descripcion:
            raise ValueError("La descripcion del reemplazo es obligatoria")
        if self._decimal_o_error(cantidad, "La cantidad debe ser mayor a cero") <= 0:
            raise ValueError("La cantidad debe ser mayor a cero")
        if (
            precio_unitario is not None
            and self._decimal_o_error(
                precio_unitario, "El precio unitario no puede ser negativo"
            ) < 0
        ):
            raise ValueError("El precio unitario no puede ser negativo")
        if not grupo_ids:
            raise ValueError("Selecciona al menos un grupo BOM")
        distribucion_grupos = self._normalizar_distribucion_grupos(
            grupo_ids, grupo_porcentajes
        )

        item_origen, bom = await self._validar_item_base_para_adenda(conn, id_item_origen)

        async with conn.transaction():
            adenda = await self.db.crear_adenda(
                conn, item_origen["id_bom"], "REEMPLAZO", motivo, user_id
            )
            await self.db.registrar_adenda_item(
                conn, adenda["id_adenda"], item_origen["id_bom"], "REEMPLAZO", motivo,
                id_item_origen=id_item_origen,
                datos_item=self._datos_item_adenda(
                    descripcion=descripcion,
                    cantidad=cantidad,
                    id_categoria=id_categoria,
                    unidad_medida=unidad_medida,
                    comentarios=comentarios,
                    precio_unitario=precio_unitario,
                    origen_precio=origen_precio,
                    id_material_ref=id_material_ref,
                    id_material_interno=id_material_interno,
                    tipo_partida=tipo_partida,
                    moneda=moneda,
                    _grupo_porcentajes={
                        str(grupo_id): str(porcentaje)
                        for grupo_id, porcentaje in distribucion_grupos.items()
                    },
                ),
                grupo_ids=grupo_ids,
            )
            await self.db.registrar_historial(
                conn, item_origen["id_bom"], AccionHistorial.AGREGADO,
                bom["version"], user_id,
                id_item=id_item_origen,
                campo_modificado="adenda_reemplazo",
                valor_nuevo=descripcion,
            )

        return {**adenda, "id_bom": item_origen["id_bom"]}

    async def agregar_fuera_scope(
        self, conn, id_bom: UUID, user_id: UUID,
        descripcion: str, cantidad, grupo_ids: List[int], motivo: str,
        grupo_porcentajes: Optional[dict[int, Decimal]] = None,
        id_categoria: Optional[int] = None,
        unidad_medida: Optional[str] = None,
        comentarios: Optional[str] = None,
        precio_unitario=None,
        origen_precio: Optional[str] = "MANUAL",
        id_material_ref: Optional[UUID] = None,
        id_material_interno: Optional[UUID] = None,
        tipo_partida: Optional[str] = "MATERIAL",
        moneda: Optional[str] = "MXN",
    ) -> dict:
        """Crea una adenda pendiente para una linea fuera del alcance aprobado."""
        motivo = self._validar_motivo_adenda(motivo)
        descripcion = (descripcion or "").strip()
        if not descripcion:
            raise ValueError("La descripcion del item fuera de alcance es obligatoria")
        if self._decimal_o_error(cantidad, "La cantidad debe ser mayor a cero") <= 0:
            raise ValueError("La cantidad debe ser mayor a cero")
        if (
            precio_unitario is not None
            and self._decimal_o_error(
                precio_unitario, "El precio unitario no puede ser negativo"
            ) < 0
        ):
            raise ValueError("El precio unitario no puede ser negativo")
        if not grupo_ids:
            raise ValueError("Selecciona al menos un grupo BOM")
        distribucion_grupos = self._normalizar_distribucion_grupos(
            grupo_ids, grupo_porcentajes
        )

        bom = await self._validar_bom_aprobado_final_para_adenda(conn, id_bom)

        async with conn.transaction():
            adenda = await self.db.crear_adenda(
                conn, id_bom, "FUERA_SCOPE", motivo, user_id
            )
            await self.db.registrar_adenda_item(
                conn, adenda["id_adenda"], id_bom, "FUERA_SCOPE", motivo,
                datos_item=self._datos_item_adenda(
                    descripcion=descripcion,
                    cantidad=cantidad,
                    id_categoria=id_categoria,
                    unidad_medida=unidad_medida,
                    comentarios=comentarios,
                    precio_unitario=precio_unitario,
                    origen_precio=origen_precio,
                    id_material_ref=id_material_ref,
                    id_material_interno=id_material_interno,
                    tipo_partida=tipo_partida,
                    moneda=moneda,
                    _grupo_porcentajes={
                        str(grupo_id): str(porcentaje)
                        for grupo_id, porcentaje in distribucion_grupos.items()
                    },
                ),
                grupo_ids=grupo_ids,
            )
            await self.db.registrar_historial(
                conn, id_bom, AccionHistorial.AGREGADO,
                bom["version"], user_id,
                campo_modificado="adenda_fuera_scope",
                valor_nuevo=descripcion,
            )

        return {**adenda, "id_bom": id_bom}

    async def get_adendas(self, conn, id_bom: UUID) -> list:
        """Lista adendas registradas para el BOM."""
        return await self.db.get_adendas_by_bom(conn, id_bom)

    async def get_item_grupos(self, conn, id_item: UUID) -> tuple[list, list]:
        grupos = await self.db.get_grupos_por_item(conn, id_item)
        grupos_operativos = await self.db.get_grupos_operativos_por_item(conn, id_item)
        return grupos, grupos_operativos

    async def get_item_grupos_base(self, conn, id_item: UUID) -> list:
        return await self.db.get_grupos_por_item(conn, id_item)

    async def get_adenda(self, conn, id_adenda: UUID) -> Optional[dict]:
        return await self.db.get_adenda_by_id(conn, id_adenda)

    async def get_adenda_comentarios(self, conn, id_adenda: UUID) -> list:
        """Lista comentarios de una adenda."""
        return await self.db.get_adenda_comentarios(conn, id_adenda)

    async def get_adenda_comentarios_by_bom(self, conn, id_bom: UUID) -> dict:
        """Lista comentarios de adendas agrupados por adenda."""
        return await self.db.get_adenda_comentarios_by_bom(conn, id_bom)

    async def get_jefe_ingenieria_label(self, conn) -> str:
        jefe = await self.db.get_usuario_activo_por_rol_org(conn, "jefe_ingenieria")
        return jefe["nombre"] if jefe else "el jefe de Ingeniería"

    async def get_jefe_construccion_label(self, conn) -> str:
        jefe = await self.db.get_usuario_activo_por_rol_org(conn, "jefe_construccion")
        return jefe["nombre"] if jefe else "el Jefe de Construccion"

    async def get_responsable_ingenieria_label(self, conn, id_proyecto: UUID) -> str:
        """Nombre del RI del proyecto (o el jefe global si aun no tiene RI asignado)."""
        responsable = await self.db.get_responsable_proyecto_o_global(
            conn, id_proyecto, "jefe_ingenieria"
        )
        return responsable["nombre"] if responsable else "el Responsable de Ingeniería"

    async def get_mensaje_hub_sin_bom(self, conn, id_proyecto: UUID, module_roles: dict) -> str:
        """Mensaje del hub cuando aun no hay paquetes BOM y el usuario no puede crearlos.

        El contacto sugerido depende del departamento de quien entra: Ingenieria
        siempre puede iniciar el BOM, asi que apunta al RI. Construccion/Compras
        solo ven un contacto si les falta su propio rol operativo asignado
        (Coordinador de Obra / Comprador Asignado); si ya lo tienen, no hay nada
        que ellos puedan resolver y se les muestra un mensaje neutro.
        """
        mensaje_espera_inicio = (
            "El BOM aún no ha sido iniciado. Este proyecto está a la espera de que Ingeniería lo inicie."
        )

        async def mensaje_contacta_ingenieria() -> str:
            label = await self.get_responsable_ingenieria_label(conn, id_proyecto)
            return f"Solicita a {label} que asigne este proyecto para poder iniciar el BOM."

        if module_roles.get("ingenieria"):
            return await mensaje_contacta_ingenieria()

        if module_roles.get("construccion"):
            coordinador = await self.db.get_asignacion_proyecto(
                conn, id_proyecto, "coordinador_obra", "CONSTRUCCION"
            )
            if not coordinador:
                responsable = await self.db.get_responsable_proyecto_o_global(
                    conn, id_proyecto, "jefe_construccion"
                )
                label = responsable["nombre"] if responsable else "el Responsable de Construcción"
                return f"Contacta a {label} para que asigne un Coordinador de Obra a este proyecto."
            return mensaje_espera_inicio

        if module_roles.get("compras"):
            comprador = await self.db.get_asignacion_proyecto(
                conn, id_proyecto, "comprador_asignado", "COMPRAS"
            )
            if not comprador:
                jefe = await self.db.get_usuario_activo_por_rol_org(conn, "jefe_compras")
                label = jefe["nombre"] if jefe else "el Jefe de Compras"
                return f"Contacta a {label} para que asigne un responsable de Compras a este proyecto."
            return mensaje_espera_inicio

        return await mensaje_contacta_ingenieria()

    async def comentar_adenda(
        self, conn, id_adenda: UUID, user_id: UUID, comentario: str
    ) -> dict:
        """Agrega un comentario de workflow a una adenda."""
        comentario_limpio = (comentario or "").strip()
        if not comentario_limpio:
            raise ValueError("El comentario es obligatorio")
        adenda = await self.db.get_adenda_by_id(conn, id_adenda)
        if not adenda:
            raise ValueError("Adenda no encontrada")
        return await self.db.registrar_adenda_comentario(
            conn, id_adenda, comentario_limpio, user_id
        )

    async def _calcular_snapshot_adenda(self, conn, adenda: dict) -> dict:
        """Congela el impacto monetario firmado de una adenda aprobada."""
        lineas = await self.db.get_adenda_items(conn, adenda["id_adenda"])
        if not lineas:
            raise ValueError("La adenda no tiene lineas para aplicar")
        ids_origen = sorted(
            {
                linea["id_item_origen"]
                for linea in lineas
                if linea.get("id_item_origen")
            },
            key=str,
        )
        if ids_origen:
            bloqueados = await self.db.lock_items_context_by_ids(conn, ids_origen)
            if len(bloqueados) != len(ids_origen) or any(
                str(item["id_bom"]) != str(adenda["id_bom_base"])
                for item in bloqueados
            ):
                raise ValueError("La adenda contiene un item invalido")

        impacto_mxn = Decimal("0")
        impacto_usd = Decimal("0")

        def acumular(moneda, importe):
            nonlocal impacto_mxn, impacto_usd
            moneda_normalizada = (moneda or "").strip().upper()
            if moneda_normalizada not in {"MXN", "USD"}:
                raise ValueError(
                    "La adenda contiene una moneda desconocida; corrige el item antes de aprobar"
                )
            if moneda_normalizada == "USD":
                impacto_usd += importe
            else:
                impacto_mxn += importe

        for linea in lineas:
            tipo = linea["tipo_linea"]
            if tipo in {"REEMPLAZO", "NO_ADQUIRIDO"}:
                if (
                    linea.get("origen_cantidad") is None
                    or linea.get("origen_precio_unitario") is None
                ):
                    raise ValueError(
                        "La adenda contiene un costo base desconocido; corrige el item antes de aprobar"
                    )
                cantidad_origen = Decimal(str(linea["origen_cantidad"]))
                precio_origen = Decimal(str(linea["origen_precio_unitario"]))
                acumular(linea.get("origen_moneda"), -(cantidad_origen * precio_origen))
            if tipo in {"REEMPLAZO", "FUERA_SCOPE"}:
                datos = self._datos_item_desde_adenda(linea.get("datos_item"))
                if datos.get("cantidad") is None or datos.get("precio_unitario") is None:
                    raise ValueError(
                        "La adenda contiene un costo nuevo desconocido; corrige el item antes de aprobar"
                    )
                cantidad = Decimal(str(datos["cantidad"]))
                precio = Decimal(str(datos["precio_unitario"]))
                acumular(datos.get("moneda"), cantidad * precio)

        id_proyecto_adenda = await self.db.get_id_proyecto_by_bom(conn, adenda["id_bom_base"])
        resuelto = await self.resolver_tipo_cambio(conn, id_proyecto_adenda)
        tasa = resuelto["tasa"]
        if impacto_usd and not tasa:
            raise ValueError(
                "No hay tipo de cambio vigente para congelar el impacto de la adenda"
            )
        impacto_total = impacto_mxn + (impacto_usd * tasa if tasa else 0)
        return {
            "tipo_cambio_aprobacion": tasa,
            "fecha_tipo_cambio_aprobacion": resuelto["fecha"] if tasa else None,
            "impacto_base_mxn_snapshot": impacto_mxn,
            "impacto_base_usd_snapshot": impacto_usd,
            "impacto_aprobado_mxn": impacto_total,
        }

    async def _aplicar_adenda(self, conn, adenda: dict, user_id: UUID) -> None:
        """Aplica lineas de adenda aprobadas dentro de la transaccion activa."""
        lineas = await self.db.get_adenda_items(conn, adenda["id_adenda"])
        if not lineas:
            raise ValueError("La adenda no tiene lineas para aplicar")

        for linea in lineas:
            tipo_linea = linea["tipo_linea"]
            motivo = linea.get("motivo") or adenda["motivo"]

            if tipo_linea == "NO_ADQUIRIDO":
                id_item_origen = linea["id_item_origen"]
                item_origen, _ = await self._validar_item_base_para_adenda(
                    conn, id_item_origen
                )
                ejecucion = await self.db.upsert_item_ejecucion(
                    conn, id_item_origen, updated_by=user_id,
                    lock_version_esperado=item_origen.get(
                        "ejecucion_lock_version", 0
                    ),
                    estatus_ejecucion="NO_ADQUIRIDO",
                    comentarios_operativos=motivo,
                )
                if not ejecucion:
                    raise ValueError(
                        "La ejecución del ítem cambió; recarga la adenda"
                    )
                await self.db.registrar_historial(
                    conn, adenda["id_bom_base"], AccionHistorial.EDITADO,
                    adenda["bom_version"], user_id,
                    id_item=id_item_origen,
                    campo_modificado="adenda",
                    valor_nuevo="NO_ADQUIRIDO",
                )
                continue

            datos = self._datos_item_desde_adenda(linea.get("datos_item"))
            grupo_ids = list(linea.get("grupo_ids") or [])
            porcentajes_json = datos.pop("_grupo_porcentajes", None)
            grupo_porcentajes = (
                {
                    int(grupo_id): Decimal(str(porcentaje))
                    for grupo_id, porcentaje in porcentajes_json.items()
                }
                if porcentajes_json
                else None
            )
            distribucion_grupos = self._normalizar_distribucion_grupos(
                grupo_ids, grupo_porcentajes
            )
            orden = await self.db.get_next_orden(conn, adenda["id_bom_base"])
            tipo_origen = (
                TIPO_ITEM_REEMPLAZO
                if tipo_linea == "REEMPLAZO"
                else TIPO_ITEM_FUERA_SCOPE
            )
            id_item_origen = linea.get("id_item_origen")
            item_origen = None
            if tipo_linea == "REEMPLAZO":
                item_origen, _ = await self._validar_item_base_para_adenda(
                    conn, id_item_origen
                )

            item = await self.db.agregar_item(
                conn,
                adenda["id_bom_base"],
                datos["descripcion"],
                datos["cantidad"],
                id_categoria=datos.get("id_categoria"),
                unidad_medida=datos.get("unidad_medida"),
                comentarios=datos.get("comentarios"),
                orden=orden,
                precio_unitario=datos.get("precio_unitario"),
                origen_precio=datos.get("origen_precio") or "MANUAL",
                id_material_ref=datos.get("id_material_ref"),
                id_material_interno=datos.get("id_material_interno"),
                tipo_partida=datos.get("tipo_partida") or "MATERIAL",
                moneda=datos.get("moneda") or "MXN",
                tipo_origen_item=tipo_origen,
                id_item_reemplazado=id_item_origen,
                motivo_adenda=motivo,
                creado_en_adenda=adenda["id_adenda"],
                creado_por=user_id,
            )
            await self.db.set_item_grupos_operativos(
                conn, item["id_item"], grupo_ids, user_id
            )
            await self._guardar_distribucion_grupos(
                conn, item["id_item"], distribucion_grupos
            )
            await self.db.vincular_adenda_item_bom(
                conn, linea["id_adenda_item"], item["id_item"]
            )

            if tipo_linea == "REEMPLAZO":
                ejecucion = await self.db.upsert_item_ejecucion(
                    conn, id_item_origen, updated_by=user_id,
                    lock_version_esperado=item_origen.get(
                        "ejecucion_lock_version", 0
                    ),
                    estatus_ejecucion="REEMPLAZADO",
                    comentarios_operativos=motivo,
                )
                if not ejecucion:
                    raise ValueError(
                        "La ejecución del ítem cambió; recarga la adenda"
                    )

            await self.db.registrar_historial(
                conn, adenda["id_bom_base"], AccionHistorial.AGREGADO,
                adenda["bom_version"], user_id,
                id_item=item["id_item"],
                campo_modificado=(
                    "adenda_reemplazo"
                    if tipo_linea == "REEMPLAZO"
                    else "adenda_fuera_scope"
                ),
                valor_nuevo=datos["descripcion"],
            )

    async def aprobar_adenda_construccion(
        self, conn, id_adenda: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, requiere_ingenieria: bool = False,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprueba una adenda por Construccion y la aplica si no requiere Ingenieria."""
        adenda = await self.db.get_adenda_by_id(conn, id_adenda)
        if not adenda:
            raise ValueError("Adenda no encontrada")
        if adenda["estatus"] != ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION:
            raise ValueError("La adenda no esta pendiente de Construccion")
        if EstatusBOM(adenda["bom_estatus"]) != EstatusBOM.APROBADO_FINAL:
            raise ValueError("Solo se pueden aprobar adendas en BOM aprobado final")
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            adenda.get("jefe_construccion"), "Jefe de Construccion", "jefe_construccion"
        )

        async with conn.transaction():
            bom_bloqueado = await self.db.get_bom_for_update(
                conn, adenda["id_bom_base"]
            )
            if not bom_bloqueado or bom_bloqueado["estatus"] != "APROBADO_FINAL":
                raise ValueError("El BOM ya no esta aprobado final")
            bloqueada = await self.db.get_adenda_for_update(conn, id_adenda)
            if not bloqueada or bloqueada["estatus"] != ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            revision = self._resolver_revision(lock_version_esperado)
            snapshot = (
                await self._calcular_snapshot_adenda(conn, adenda)
                if not requiere_ingenieria
                else {}
            )
            if not requiere_ingenieria:
                await self._aplicar_adenda(conn, adenda, user_id)
            updated = await self.db.marcar_adenda_construccion(
                conn, id_adenda, user_id, requiere_ingenieria, revision, **snapshot
            )
            if not updated:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-ADENDA:{id_adenda}:{updated['lock_version']}:CONSTRUCCION",
                "ADENDA_APROBADA" if not requiere_ingenieria else "ADENDA_PENDIENTE_INGENIERIA",
                adenda["id_proyecto"], user_id,
                {"tipo_adenda": adenda["tipo_adenda"], "version": adenda["bom_version"]},
                id_paquete=adenda.get("id_paquete"), id_bom=adenda["id_bom_base"],
                id_documento=id_adenda,
            )
        if not requiere_ingenieria:
            await self._avisar_costos_pendientes_tras_adenda(
                conn, adenda["id_bom_base"], user_id
            )
        return updated

    async def _avisar_costos_pendientes_tras_adenda(
        self, conn, id_bom: UUID, user_id: UUID
    ) -> None:
        """Dispara el mismo aviso (SSE/correo) del boton manual "Notificar a
        Compras", pero automatico al aplicar una adenda: FUERA_SCOPE/REEMPLAZO
        solo existen con el BOM en APROBADO_FINAL, donde ninguna transicion de
        etapa (ni su modal de aprobacion, unico lugar con ese boton hoy) vuelve
        a correr -- sin este disparo el item sin costo quedaba sin ningun canal
        activo de recordatorio, solo visible si Compras entraba por su cuenta a
        la pestaña de precios pendientes. Best effort, llamado DESPUES del
        commit de la adenda: nunca debe tumbar una aprobacion ya aplicada."""
        try:
            await self.notificar_items_sin_costo_compras(conn, id_bom, user_id)
        except ValueError:
            # Sin items sin costo (lo normal, la mayoria de adendas SI trae
            # precio) o sin canal de aviso configurado -- no es un error de la
            # adenda en si, solo no hay nada (o como) que avisar.
            pass
        except asyncpg.PostgresError:
            logger.exception(
                "Error de BD al avisar costos pendientes tras aplicar adenda (bom=%s)",
                id_bom,
            )

    async def aprobar_adenda_ingenieria(
        self, conn, id_adenda: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprueba tecnicamente una adenda y aplica sus cambios."""
        adenda = await self.db.get_adenda_by_id(conn, id_adenda)
        if not adenda:
            raise ValueError("Adenda no encontrada")
        if adenda["estatus"] != ESTATUS_ADENDA_PENDIENTE_INGENIERIA:
            raise ValueError("La adenda no esta pendiente de Ingenieria")
        if EstatusBOM(adenda["bom_estatus"]) != EstatusBOM.APROBADO_FINAL:
            raise ValueError("Solo se pueden aprobar adendas en BOM aprobado final")
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            adenda.get("responsable_ing"), "Responsable de Ingenieria", "jefe_ingenieria"
        )

        async with conn.transaction():
            bom_bloqueado = await self.db.get_bom_for_update(
                conn, adenda["id_bom_base"]
            )
            if not bom_bloqueado or bom_bloqueado["estatus"] != "APROBADO_FINAL":
                raise ValueError("El BOM ya no esta aprobado final")
            bloqueada = await self.db.get_adenda_for_update(conn, id_adenda)
            if not bloqueada or bloqueada["estatus"] != ESTATUS_ADENDA_PENDIENTE_INGENIERIA:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            revision = self._resolver_revision(lock_version_esperado)
            snapshot = await self._calcular_snapshot_adenda(conn, adenda)
            await self._aplicar_adenda(conn, adenda, user_id)
            updated = await self.db.aprobar_adenda_ingenieria(
                conn, id_adenda, user_id, revision, **snapshot
            )
            if not updated:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM-ADENDA:{id_adenda}:{updated['lock_version']}:INGENIERIA",
                "ADENDA_APROBADA", adenda["id_proyecto"], user_id,
                {"tipo_adenda": adenda["tipo_adenda"], "version": adenda["bom_version"]},
                id_paquete=adenda.get("id_paquete"), id_bom=adenda["id_bom_base"],
                id_documento=id_adenda,
            )
        await self._avisar_costos_pendientes_tras_adenda(
            conn, adenda["id_bom_base"], user_id
        )
        return updated

    async def rechazar_adenda(
        self, conn, id_adenda: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str], motivo_rechazo: str,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Rechaza una adenda pendiente sin mutar items."""
        motivo = (motivo_rechazo or "").strip()
        if not motivo:
            raise ValueError("El motivo de rechazo es obligatorio")
        adenda = await self.db.get_adenda_by_id(conn, id_adenda)
        if not adenda:
            raise ValueError("Adenda no encontrada")
        if adenda["estatus"] in ESTATUS_ADENDA_TERMINALES:
            raise ValueError("La adenda ya esta cerrada")
        if adenda["estatus"] == ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION:
            await self._validar_aprobador_bom(
                conn, user_id, user_role, rol_org,
                adenda.get("jefe_construccion"), "Jefe de Construccion", "jefe_construccion"
            )
        elif adenda["estatus"] == ESTATUS_ADENDA_PENDIENTE_INGENIERIA:
            await self._validar_aprobador_bom(
                conn, user_id, user_role, rol_org,
                adenda.get("responsable_ing"), "Responsable de Ingenieria", "jefe_ingenieria"
            )
        else:
            raise ValueError("La adenda no esta pendiente de aprobacion")
        async with conn.transaction():
            bloqueada = await self.db.get_adenda_for_update(conn, id_adenda)
            if not bloqueada or bloqueada["estatus"] != adenda["estatus"]:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            revision = self._resolver_revision(lock_version_esperado)
            updated = await self.db.rechazar_adenda(
                conn, id_adenda, user_id, motivo, adenda["estatus"], revision
            )
            if not updated:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            await self.db.registrar_evento_outbox(
                conn, f"BOM-ADENDA:{id_adenda}:{updated['lock_version']}:RECHAZADA",
                "ADENDA_RECHAZADA", adenda["id_proyecto"], user_id,
                {"motivo": motivo, "version": adenda["bom_version"]},
                id_paquete=adenda.get("id_paquete"), id_bom=adenda["id_bom_base"],
                id_documento=id_adenda,
            )
        return updated

    async def cancelar_adenda(
        self, conn, id_adenda: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Cancela una adenda pendiente de construccion sin mutar items."""
        adenda = await self.db.get_adenda_by_id(conn, id_adenda)
        if not adenda:
            raise ValueError("Adenda no encontrada")
        if adenda["estatus"] != ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION:
            raise ValueError("Solo se pueden cancelar adendas pendientes de Construccion")
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            adenda.get("jefe_construccion"), "Jefe de Construccion", "jefe_construccion"
        )
        async with conn.transaction():
            bloqueada = await self.db.get_adenda_for_update(conn, id_adenda)
            if not bloqueada or bloqueada["estatus"] != ESTATUS_ADENDA_PENDIENTE_CONSTRUCCION:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
            revision = self._resolver_revision(lock_version_esperado)
            updated = await self.db.cancelar_adenda(
                conn, id_adenda, user_id, revision
            )
            if not updated:
                raise ValueError("La adenda cambio; recarga el BOM e intenta de nuevo")
        return updated

    @staticmethod
    def _fecha_requerida_solo_ejecucion(area_editor: str, campos: set) -> bool:
        """fecha_requerida es autoridad de ejecucion para Construccion (no de
        turno/base como el resto de CAMPOS_BASE_BOM): se sigue escribiendo en
        tb_bom_items via update_item (unica tabla que la tiene, la de ejecucion
        no), pero no exige el lock completo del BOM cuando es lo unico que se
        esta tocando -- asi sigue disponible en APROBADO_FINAL (operacion
        downstream), igual que fecha estimada/llegada real.
        """
        return bool(
            area_editor == "construccion"
            and campos
            and campos <= CAMPOS_CONSTRUCCION_BASE
        )

    async def editar_item(
        self, conn, id_item: UUID, user_id: UUID,
        area_editor: str, lock_version_esperado: Optional[int] = None,
        ejecucion_lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
        grupo_ids: Optional[List[int]] = None,
        grupo_porcentajes: Optional[dict[int, Decimal]] = None,
        reservar_bom: bool = True,
        capacidades: Optional[dict] = None,
        bom: Optional[dict] = None,
        **campos,
    ) -> dict:
        """Edita base por turno y ejecucion por modulo, sin dividir secciones por actor.

        `capacidades`/`bom`, si se pasan (ej. desde editar_items_bulk), evitan
        recalcularlos/releerlos aqui adentro — el caller ya los obtuvo una vez
        para todo el lote y, con reservar_bom=False, ninguno cambia por item."""
        item = await self.db.get_item_by_id(conn, id_item)
        if not item:
            raise ValueError("Item no encontrado")
        if not item.get("activo", True):
            raise ValueError("No se puede editar un item eliminado")
        if item.get("bloqueado") and any(k in CAMPOS_BASE_BOM for k in campos):
            raise ValueError("Este item pertenece a una linea historica bloqueada")

        campos_base = {k: v for k, v in campos.items() if k in CAMPOS_BASE_BOM}
        if area_editor == "compras":
            for campo_real in {"precio_unitario", "moneda", "comentarios"}:
                campos_base.pop(campo_real, None)
        elif area_editor == "construccion":
            # Autoridad de costos (doc 38): precio_unitario/moneda son de
            # Ingenieria o Compras, nunca de Construccion, aunque controle el
            # turno (editar_base=True) en EN_REVISION_OBRA/EN_REVISION_CONST.
            # Excepcion acotada: SI puede reabrir un costo ya configurado
            # dejandolo exactamente en 0 (bandera "sin costo, esperando
            # Compras") -- cualquier otro valor se descarta igual que antes.
            precio_construccion = campos_base.get("precio_unitario")
            if precio_construccion is None or self._decimal_o_error(
                precio_construccion, "El precio unitario no puede ser negativo"
            ) != 0:
                campos_base.pop("precio_unitario", None)
                campos_base.pop("moneda", None)
        campos_ejecucion_permitidos = (
            CAMPOS_CONSTRUCCION_EJECUCION
            if area_editor == "construccion"
            else CAMPOS_COMPRAS if area_editor == "compras" else set()
        )
        campos_ejecucion_entrada = {
            k: v for k, v in campos.items()
            if k in campos_ejecucion_permitidos
        }
        entregado_manual = campos_ejecucion_entrada.get("entregado")
        if item.get("id_material_interno"):
            for protegido in {"descripcion", "id_material_ref"}:
                campos_base.pop(protegido, None)
        if "cantidad" in campos_base:
            cantidad = self._decimal_o_error(
                campos_base["cantidad"], "La cantidad debe ser mayor a cero"
            )
            if cantidad <= 0:
                raise ValueError("La cantidad debe ser mayor a cero")
        if campos_base.get("precio_unitario") is not None:
            precio = self._decimal_o_error(
                campos_base["precio_unitario"],
                "El precio unitario no puede ser negativo",
            )
            if precio < 0:
                raise ValueError("El precio unitario no puede ser negativo")
        if "precio_unitario" in campos_base and area_editor != "compras":
            if campos_base.get("moneda") not in MONEDAS_VALIDAS:
                raise ValueError(
                    "Selecciona la moneda del presupuesto antes de aplicar el cambio"
                )

        mapping_ejecucion = {
            "id_proveedor": "id_proveedor_real",
            "precio_unitario": "precio_real",
            "moneda": "moneda_real",
            "comentarios": "comentarios_operativos",
        }
        campos_ejecucion = {
            mapping_ejecucion.get(k, k): v
            for k, v in campos_ejecucion_entrada.items()
            if k not in {"origen_precio", "entregado"}
        }
        if "cantidad_recibida" in campos_ejecucion:
            recibida = self._decimal_o_error(
                campos_ejecucion["cantidad_recibida"],
                "La cantidad recibida no puede ser negativa",
            )
            total = self._decimal_o_error(item["cantidad"], "Cantidad de item invalida")
            if recibida < 0 or recibida > total:
                raise ValueError("La cantidad recibida debe estar entre cero y la cantidad total")
            campos_ejecucion.setdefault(
                "estatus_ejecucion",
                "RECIBIDO_TOTAL" if recibida == total else (
                    "RECIBIDO_PARCIAL" if recibida > 0 else "PENDIENTE"
                ),
            )
        elif entregado_manual is True:
            campos_ejecucion["cantidad_recibida"] = item["cantidad"]
            campos_ejecucion.setdefault("estatus_ejecucion", "RECIBIDO_TOTAL")
        elif entregado_manual is False:
            campos_ejecucion["cantidad_recibida"] = 0
        if campos_ejecucion.get("precio_real") is not None:
            precio_real = self._decimal_o_error(
                campos_ejecucion["precio_real"], "El costo real no puede ser negativo"
            )
            if precio_real < 0:
                raise ValueError("El costo real no puede ser negativo")
            campos_ejecucion.setdefault("estatus_ejecucion", "COTIZADO")
        if grupo_ids is not None and not grupo_ids:
            # El form de edicion de item manda los checkboxes de grupos aunque
            # el usuario solo haya tocado otro tab (ej. precio en Ingenieria):
            # si el item nunca tuvo grupos, no mandar ninguno es un no-op, no
            # un intento de vaciarlos. Solo bloquea si SI tenia grupos antes.
            grupos_actuales = await self.get_item_grupos_base(conn, id_item)
            if grupos_actuales:
                raise ValueError("Selecciona al menos un grupo BOM")
            # No-op confirmado: tratarlo igual que "el campo no se envio" para
            # que ni _normalizar_distribucion_grupos (que si truena en vacio)
            # ni la escritura de grupos de mas abajo (`if grupo_ids is not None`)
            # se disparen sobre una lista vacia legitima.
            grupo_ids = None
        distribucion_grupos = (
            self._normalizar_distribucion_grupos(grupo_ids, grupo_porcentajes)
            if grupo_ids is not None else None
        )
        if not campos_base and not campos_ejecucion and grupo_ids is None:
            raise ValueError("No hay campos validos para actualizar")

        async with conn.transaction():
            if bom is None:
                bom = await self.db.get_bom_by_id(conn, item["id_bom"])
            if not bom:
                raise ValueError("BOM no encontrado")
            grupos_base = (
                grupo_ids is not None
                and EstatusBOM(bom["estatus"]) != EstatusBOM.APROBADO_FINAL
            )
            muta_ejecucion = bool(
                campos_ejecucion or (grupo_ids is not None and not grupos_base)
            )
            if muta_ejecucion and ejecucion_lock_version_esperado is None:
                raise ValueError(
                    "Falta la revisión de ejecución; recarga el paquete e intenta de nuevo"
                )
            fecha_requerida_solo_ejecucion = self._fecha_requerida_solo_ejecucion(
                area_editor, campos_base.keys()
            )
            reservara_base = (
                (campos_base or grupos_base)
                and reservar_bom
                and not fecha_requerida_solo_ejecucion
            )
            if reservara_base:
                # capacidades se calcula aqui adentro, contra la fila recien
                # bloqueada (FOR UPDATE) — no hace falta calcularla antes.
                bom, capacidades = await self._reservar_mutacion_base(
                    conn, item["id_bom"], user_id, lock_version_esperado,
                    user_role, rol_org, module_roles,
                )
            else:
                if capacidades is None:
                    capacidades = await self.get_capacidades_bom(
                        conn, bom, user_id, user_role, rol_org, module_roles
                    )
                if fecha_requerida_solo_ejecucion:
                    if not capacidades["editar_ejecucion"]:
                        raise ValueError(
                            "Solo Construccion o Compras pueden actualizar la ejecucion"
                        )
                elif (campos_base or grupos_base) and not capacidades["editar_base"]:
                    turno = capacidades.get("actor_turno") or "el actor asignado"
                    raise ValueError(f"Solo {turno} puede modificar el BOM en esta etapa")
            if campos_ejecucion and not capacidades["editar_ejecucion"]:
                raise ValueError("Solo Construccion o Compras pueden actualizar la ejecucion")
            if grupo_ids is not None and not grupos_base and not capacidades["editar_ejecucion"]:
                raise ValueError("Solo Construccion puede actualizar grupos operativos")

            if (
                area_editor in ("ingenieria", "construccion")
                and campos_base.get("precio_unitario") is not None
            ):
                # Autoridad de edicion de costos (Fase 4): Ingenieria puede capturar
                # el costo de un item sin costo, pero una vez que el item YA tiene
                # un costo configurado (sin importar si es de entrada manual o de
                # catalogo) ya no puede modificarlo — solo dejarlo en cero para
                # reabrirlo. El valor que deje Ingenieria queda pendiente de que
                # Compras lo confirme o edite (precio_pendiente_confirmacion); ese
                # es el que se vuelve costo oficial. Se valida aqui (despues del
                # gate de turno/estatus de arriba) para que "no es tu turno"
                # siga siendo el mensaje correcto cuando ambas cosas aplican.
                # Construccion solo llega aqui con precio_unitario == 0 (el
                # popeo de arriba descarta cualquier otro valor), asi que para
                # esa area este bloque es siempre el camino de "reabrir".
                precio_nuevo = precio
                precio_actual = item.get("precio_unitario")
                try:
                    # precio_pendiente_confirmacion=True significa que el precio
                    # actual todavia no es oficial (Compras no lo ha confirmado) --
                    # Ingenieria puede seguir corrigiendolo directo, sin pasar por
                    # el "dejalo en cero" que solo aplica a un costo ya oficial.
                    tiene_costo_configurado = (
                        precio_actual is not None and Decimal(str(precio_actual)) > 0
                        and not item.get("precio_pendiente_confirmacion")
                    )
                except (InvalidOperation, TypeError, ValueError):
                    tiene_costo_configurado = False
                if tiene_costo_configurado and precio_nuevo != 0:
                    raise ValueError(self.mensaje_costo_ya_configurado())
                campos_base["precio_pendiente_confirmacion"] = precio_nuevo > 0

            cambios = []
            for campo, valor in campos_base.items():
                if campo == "precio_pendiente_confirmacion":
                    # Derivado internamente (Fase 4), no un campo que el usuario
                    # haya editado — no genera su propia linea de historial.
                    continue
                cambios.append((campo, item.get(campo), valor))
            publicos = {
                "id_proveedor_real": "id_proveedor",
                "precio_real": "precio_real",
                "moneda_real": "moneda_real",
            }
            for campo, valor in campos_ejecucion.items():
                cambios.append((campo, item.get(publicos.get(campo, campo)), valor))

            if campos_base:
                # CAS a nivel de item solo cuando se toca precio_unitario: protege
                # contra el caso donde Compras confirma un costo (bump de
                # lock_version via actualizar_precios_items_compras_cas_batch) al
                # mismo tiempo que Ingenieria envia su cambio (ej. reabrir a cero)
                # sobre el mismo item con un lock_version ya obsoleto. Otros campos
                # base (descripcion, fecha_requerida, etc.) no llevan este CAS —
                # su unica proteccion de concurrencia sigue siendo el lock del BOM
                # completo en _reservar_mutacion_base.
                lock_item_esperado = (
                    item["lock_version"] if "precio_unitario" in campos_base else None
                )
                item_base_actualizado = await self.db.update_item(
                    conn, id_item, lock_version_esperado=lock_item_esperado, **campos_base
                )
                if lock_item_esperado is not None and item_base_actualizado is None:
                    raise ValueError(
                        "El costo de este item fue modificado por alguien mas "
                        "(probablemente Compras); recarga el item e intenta de nuevo"
                    )
            if campos_ejecucion:
                ejecucion = await self.db.upsert_item_ejecucion(
                    conn, id_item, updated_by=user_id,
                    lock_version_esperado=ejecucion_lock_version_esperado,
                    **campos_ejecucion,
                )
                if not ejecucion:
                    raise ValueError(
                        "La ejecución del item cambió; recarga el paquete e intenta de nuevo"
                    )
            elif grupo_ids is not None and not grupos_base:
                ejecucion = await self.db.upsert_item_ejecucion(
                    conn, id_item, updated_by=user_id,
                    lock_version_esperado=ejecucion_lock_version_esperado,
                    estatus_ejecucion=(
                        item.get("estatus_ejecucion") or "PENDIENTE"
                    ),
                )
                if not ejecucion:
                    raise ValueError(
                        "La ejecución del item cambió; recarga el paquete e intenta de nuevo"
                    )
            if grupo_ids is not None:
                anteriores = (
                    await self.db.get_grupos_por_item(conn, id_item)
                    if grupos_base
                    else await self.db.get_grupos_operativos_por_item(conn, id_item)
                )
                if grupos_base:
                    await self.db.set_item_grupos(conn, id_item, grupo_ids)
                    campo_grupo = "grupos_bom"
                else:
                    await self.db.set_item_grupos_operativos(
                        conn, id_item, grupo_ids, user_id
                    )
                    campo_grupo = "grupos_operativos"
                await self._guardar_distribucion_grupos(
                    conn, id_item, distribucion_grupos
                )
                await self.db.registrar_historial(
                    conn, item["id_bom"], AccionHistorial.EDITADO,
                    item["bom_version"], user_id,
                    id_item=id_item,
                    campo_modificado=CAMPO_LABELS[campo_grupo],
                    valor_anterior=", ".join(anteriores) if anteriores else None,
                    valor_nuevo=", ".join(str(value) for value in grupo_ids),
                )
            for campo, anterior, nuevo in cambios:
                if str(anterior) != str(nuevo):
                    await self.db.registrar_historial(
                        conn, item["id_bom"], AccionHistorial.EDITADO,
                        item["bom_version"], user_id,
                        id_item=id_item,
                        campo_modificado=CAMPO_LABELS.get(campo, campo),
                        valor_anterior=str(anterior) if anterior is not None else None,
                        valor_nuevo=str(nuevo) if nuevo is not None else None,
                    )
        item_actualizado = await self.db.get_item_by_id(conn, id_item)
        return {"item": item_actualizado, "capacidades": capacidades}

    async def _validar_item_bulk(
        self, item: Optional[dict], id_bom: UUID, area_editor: str,
    ) -> None:
        """Valida que un item pueda editarse dentro de una operacion bulk."""
        if not item:
            raise ValueError("Item no encontrado")
        if str(item.get('id_bom')) != str(id_bom):
            raise ValueError("El item no pertenece a este BOM")
        if not item.get('activo', True):
            raise ValueError("No se puede editar un item eliminado")
        if item.get('bloqueado'):
            raise ValueError("Este item fue completado en una version anterior del BOM y no se puede modificar")

        if area_editor not in {"ingenieria", "construccion", "compras"}:
            raise ValueError("Sin permisos para editar items del BOM")

    async def editar_items_bulk(
        self, conn, id_bom: UUID, item_ids: List[UUID], user_id: UUID,
        area_editor: str, campo: str,
        valor=None, grupo_ids: Optional[List[int]] = None,
        valor_secundario: Optional[dict] = None,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        """Aplica un cambio atomico al lote con una sola reserva de la base."""
        if area_editor not in ('ingenieria', 'construccion', 'compras'):
            raise ValueError("Sin permisos para editar items del BOM")
        if not item_ids:
            raise ValueError("No hay items seleccionados")
        item_ids = list(dict.fromkeys(item_ids))

        es_grupos = campo == 'grupos'
        if es_grupos:
            if not grupo_ids:
                raise ValueError("Selecciona al menos un grupo BOM")
            if len(set(grupo_ids)) != 1:
                raise ValueError(
                    "La edición masiva de grupos permite un solo grupo; distribuye porcentajes por item"
                )
        elif (
            campo not in CAMPOS_BULK_BASE
            and campo not in CAMPOS_BULK.get(area_editor, set())
        ):
            raise ValueError("Campo no editable en lote para tu area")

        if campo in CAMPOS_BULK_REQUIERE_MONEDA and area_editor != 'compras':
            moneda_secundaria = (valor_secundario or {}).get('moneda')
            if moneda_secundaria not in MONEDAS_VALIDAS:
                raise ValueError(
                    "Selecciona la moneda del presupuesto antes de aplicar el cambio"
                )

        async with conn.transaction():
            bom = await self.db.get_bom_by_id(conn, id_bom)
            if not bom:
                raise ValueError("BOM no encontrado")
            grupos_base = es_grupos and EstatusBOM(bom["estatus"]) != EstatusBOM.APROBADO_FINAL
            redirigido_a_ejecucion = (
                area_editor == "compras"
                and campo in {"precio_unitario", "moneda", "comentarios"}
            ) or self._fecha_requerida_solo_ejecucion(area_editor, {campo})
            muta_base = (
                campo in CAMPOS_BASE_BOM and not redirigido_a_ejecucion
            ) or grupos_base
            if muta_base:
                if lock_version_esperado is None:
                    raise ValueError(
                        "Falta la revision del BOM; recarga el paquete e intenta de nuevo"
                    )
                bom, capacidades = await self._reservar_mutacion_base(
                    conn, id_bom, user_id, lock_version_esperado,
                    user_role, rol_org, module_roles,
                )
            else:
                capacidades = await self.get_capacidades_bom(
                    conn, bom, user_id, user_role, rol_org, module_roles
                )

            items_context = await self.db.lock_items_context_by_ids(conn, item_ids)
            items_por_id = {str(item["id_item"]): item for item in items_context}
            for id_item in item_ids:
                item = items_por_id.get(str(id_item))
                await self._validar_item_bulk(item, id_bom, area_editor)

            # bom/capacidades ya calculados arriba (una sola vez para todo el lote,
            # dentro de la misma transaccion): se pasan a cada editar_item para que
            # no los recalcule/relea por item (antes eran N repeticiones, una por
            # item) — con reservar_bom=False ninguno de los dos cambia entre items.
            for id_item in item_ids:
                item = items_por_id[str(id_item)]
                await self.editar_item(
                    conn, id_item, user_id, area_editor,
                    ejecucion_lock_version_esperado=item.get("ejecucion_lock_version"),
                    user_role=user_role,
                    rol_org=rol_org,
                    module_roles=module_roles,
                    grupo_ids=grupo_ids if es_grupos else None,
                    grupo_porcentajes=(
                        {grupo_ids[0]: Decimal("1")} if es_grupos else None
                    ),
                    reservar_bom=False,
                    capacidades=capacidades,
                    bom=bom,
                    **(
                        {} if es_grupos
                        else {campo: valor, **(valor_secundario or {})}
                    ),
                )
        return {"actualizados": len(item_ids), "omitidos": [], "capacidades": capacidades}

    async def eliminar_item(
        self, conn, id_item: UUID, user_id: UUID,
        area_editor: str = "ingenieria",
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        item = await self.db.get_item_by_id(conn, id_item)
        if not item:
            raise ValueError("Item no encontrado")
        if not item.get("activo", True):
            raise ValueError("El item ya esta eliminado")
        if item.get("bloqueado"):
            raise ValueError("Este item pertenece a una linea historica bloqueada")
        async with conn.transaction():
            bom, capacidades = await self._reservar_mutacion_base(
                conn, item["id_bom"], user_id, lock_version_esperado,
                user_role, rol_org, module_roles,
            )
            deleted = await self.db.soft_delete_item(conn, id_item)
            await self.db.registrar_historial(
                conn, item["id_bom"], AccionHistorial.ELIMINADO,
                bom["version"], user_id,
                id_item=id_item,
                campo_modificado="item",
                valor_anterior=item.get("descripcion"),
            )
        return {"deleted": deleted, "capacidades": capacidades}

    async def get_items(self, conn, id_bom: UUID) -> list:
        """Lista items activos del BOM, enriched with grupos and costo_mxn.

        Para items USD, el tipo de cambio se obtiene en este orden:
        1. TC del XML de la factura asociada (tb_materiales_historial.tipo_cambio_xml)
        2. TC manual del proyecto (si esta fijado)
        3. Ultima tasa Banxico registrada (tb_tipo_cambio)
        4. Promedio 7 dias Banxico (fallback final)
        """
        items = await self.db.get_items_by_bom(conn, id_bom)
        if not items:
            return items
        grupos_map = await self.db.get_grupos_por_bom(conn, id_bom)
        grupos_operativos_map = await self.db.get_grupos_operativos_por_bom(conn, id_bom)

        usd_ids = [
            item['id_item'] for item in items
            if (
                (item.get('moneda') == 'USD' and item.get('precio_unitario'))
                or (item.get('moneda_real') == 'USD' and item.get('precio_real'))
            )
        ]

        tc_from_xml = {}
        tc_resuelto = None

        if usd_ids:
            tc_from_xml = await self.db.get_tc_from_linked_materials(conn, usd_ids)

            still_need = [iid for iid in usd_ids if str(iid) not in tc_from_xml]
            if still_need:
                id_proyecto = await self.db.get_id_proyecto_by_bom(conn, id_bom)
                resuelto = await self.resolver_tipo_cambio(conn, id_proyecto)
                tc_resuelto = resuelto["tasa"]

        for item in items:
            item['grupos'] = grupos_map.get(str(item['id_item']), [])
            item['grupos_operativos'] = grupos_operativos_map.get(str(item['id_item']), [])
            moneda = item.get('moneda', 'MXN')
            moneda_real = item.get('moneda_real')
            if moneda == 'USD' and item.get('precio_unitario'):
                iid = str(item['id_item'])
                tc = tc_from_xml.get(iid) or tc_resuelto
                if tc:
                    item['costo_mxn'] = round(Decimal(str(item['precio_unitario'])) * tc, 2)
            if moneda_real == 'USD' and item.get('precio_real'):
                iid = str(item['id_item'])
                tc = tc_from_xml.get(iid) or tc_resuelto
                if tc:
                    item['costo_real_mxn'] = round(Decimal(str(item['precio_real'])) * tc, 2)

        # Enriquecer con gasto real desde materiales vinculados
        all_ids = [item['id_item'] for item in items]
        gasto_map = await self.db.get_gasto_real_por_item(conn, all_ids)
        for item in items:
            gasto = gasto_map.get(str(item['id_item']))
            if gasto is not None:
                item['gasto_real'] = round(gasto, 2)

        return items

    async def get_item(self, conn, id_item: UUID) -> dict:
        """Obtiene un item por ID, enriquecido con costo_mxn y gasto_real."""
        item = await self.db.get_item_by_id(conn, id_item)
        if not item:
            raise ValueError("Item no encontrado")
        # Enriquecer costo_mxn para items USD
        necesita_tc = (
            (item.get('moneda') == 'USD' and item.get('precio_unitario'))
            or (item.get('moneda_real') == 'USD' and item.get('precio_real'))
        )
        tc = None
        if necesita_tc:
            resuelto = await self.resolver_tipo_cambio(conn, item['id_proyecto'])
            tc = resuelto["tasa"]
        if item.get('moneda') == 'USD' and item.get('precio_unitario') and tc:
            item['costo_mxn'] = round(Decimal(str(item['precio_unitario'])) * tc, 2)
        if item.get('moneda_real') == 'USD' and item.get('precio_real') and tc:
            item['costo_real_mxn'] = round(Decimal(str(item['precio_real'])) * tc, 2)
        # Enriquecer gasto_real
        gasto_map = await self.db.get_gasto_real_por_item(conn, [id_item])
        gasto = gasto_map.get(str(id_item))
        if gasto is not None:
            item['gasto_real'] = round(gasto, 2)
        return item

    # ─── RESUMEN DE COMPRA ───────────────────────────────────

    async def get_resumen_compra(self, conn, id_bom: UUID) -> dict:
        """Arma el comparativo Presupuesto vs Facturado vs Pagado del BOM.

        Agrupa las categorias por grupo BOM (rollup en memoria), calcula totales,
        desviaciones (presupuesto - facturado / - pagado) y métricas normalizadas
        (MXN por modulo FV y por kWp). Solo agregacion; no escribe nada.
        """
        filas = await self.db.get_resumen_compra(conn, id_bom)
        divisores = await self.db.get_divisores_bom(conn, id_bom)

        def _decimal(valor):
            return None if valor is None else Decimal(str(valor))

        def _sumar(valores):
            valores = list(valores)
            if any(valor is None for valor in valores):
                return None
            return sum(valores, Decimal("0"))

        def _restar(minuendo, sustraendo):
            if minuendo is None or sustraendo is None:
                return None
            return minuendo - sustraendo

        por_grupo: dict[str, dict] = {}
        for f in filas:
            grupo_codigo = f["grupo_codigo"] or "SIN_CLASIFICAR"
            grupo = por_grupo.setdefault(grupo_codigo, {
                "codigo": grupo_codigo,
                "nombre": f["grupo_nombre"] or "Sin clasificar",
                "orden": int(f["grupo_orden"] or 999),
                "categorias": [],
            })
            facturado = _decimal(f["facturado_confirmado_mxn"])
            facturado_sugerido = _decimal(f["facturado_sugerido_mxn"])
            cat = {
                "categoria_id": f["categoria_id"],
                "categoria_nombre": f["categoria_nombre"],
                "presupuesto": _decimal(f["presupuesto_mxn"]),
                "real": _decimal(f.get("compra_real_mxn")),
                "real_base": _decimal(f.get("compra_real_base_mxn")),
                "reemplazos": _decimal(f.get("reemplazos_mxn")),
                "fuera_scope": _decimal(f.get("fuera_scope_mxn")),
                "no_adquirido": _decimal(f.get("no_adquirido_mxn")),
                "facturado": facturado,
                "facturado_sugerido": facturado_sugerido,
                "pagado": _decimal(f["pagado_mxn"]),
                "valores_pendientes": int(f.get("valores_pendientes") or 0),
                "grupos_pendientes": int(f.get("grupos_pendientes") or 0),
            }
            cat["dif_real"] = _restar(cat["presupuesto"], cat["real"])
            cat["dif_facturado"] = _restar(
                cat["presupuesto"], cat["facturado"]
            )
            cat["dif_pagado"] = _restar(cat["presupuesto"], cat["pagado"])
            grupo["categorias"].append(cat)

        secciones = []

        for grupo in sorted(por_grupo.values(), key=lambda g: (g["orden"], g["codigo"])):
            cats = grupo["categorias"]
            s_presup = _sumar(c["presupuesto"] for c in cats)
            s_real = _sumar(c["real"] for c in cats)
            s_real_base = _sumar(c["real_base"] for c in cats)
            s_reemplazos = _sumar(c["reemplazos"] for c in cats)
            s_fuera_scope = _sumar(c["fuera_scope"] for c in cats)
            s_no_adquirido = _sumar(c["no_adquirido"] for c in cats)
            s_fact = _sumar(c["facturado"] for c in cats)
            s_sug = _sumar(c["facturado_sugerido"] for c in cats)
            s_pag = _sumar(c["pagado"] for c in cats)
            secciones.append({
                "codigo": grupo["codigo"],
                "nombre": grupo["nombre"],
                "presupuesto": s_presup,
                "real": s_real,
                "real_base": s_real_base,
                "reemplazos": s_reemplazos,
                "fuera_scope": s_fuera_scope,
                "no_adquirido": s_no_adquirido,
                "facturado": s_fact,
                "facturado_sugerido": s_sug,
                "pagado": s_pag,
                "dif_real": _restar(s_presup, s_real),
                "dif_facturado": _restar(s_presup, s_fact),
                "dif_pagado": _restar(s_presup, s_pag),
                "valores_pendientes": sum(
                    c["valores_pendientes"] for c in cats
                ),
                "grupos_pendientes": sum(c["grupos_pendientes"] for c in cats),
                "categorias": cats,
            })

        tot_presup = _sumar(s["presupuesto"] for s in secciones)
        tot_real = _sumar(s["real"] for s in secciones)
        tot_real_base = _sumar(s["real_base"] for s in secciones)
        tot_reemplazos = _sumar(s["reemplazos"] for s in secciones)
        tot_fuera_scope = _sumar(s["fuera_scope"] for s in secciones)
        tot_no_adquirido = _sumar(s["no_adquirido"] for s in secciones)
        tot_fact = _sumar(s["facturado"] for s in secciones)
        tot_sugerido = _sumar(s["facturado_sugerido"] for s in secciones)
        tot_pag = _sumar(s["pagado"] for s in secciones)

        totales = {
            "presupuesto": tot_presup,
            "real": tot_real,
            "real_base": tot_real_base,
            "reemplazos": tot_reemplazos,
            "fuera_scope": tot_fuera_scope,
            "no_adquirido": tot_no_adquirido,
            "facturado": tot_fact,
            "facturado_sugerido": tot_sugerido,
            "pagado": tot_pag,
            "dif_real": _restar(tot_presup, tot_real),
            "dif_facturado": _restar(tot_presup, tot_fact),
            "dif_pagado": _restar(tot_presup, tot_pag),
            "valores_pendientes": sum(
                s["valores_pendientes"] for s in secciones
            ),
            "grupos_pendientes": sum(s["grupos_pendientes"] for s in secciones),
        }

        modulos = divisores["modulos_fv"]
        kwp = divisores["kwp"]

        def _por(divisor, valor):
            if not divisor or valor is None:
                return None
            return round(valor / Decimal(str(divisor)), 2)

        metricas = {
            "modulos_fv": modulos,
            "kwp": kwp,
            "presup_por_modulo": _por(modulos, tot_presup),
            "real_por_modulo": _por(modulos, tot_real),
            "facturado_por_modulo": _por(modulos, tot_fact),
            "sugerido_por_modulo": _por(modulos, tot_sugerido),
            "presup_por_kwp": _por(kwp, tot_presup),
            "real_por_kwp": _por(kwp, tot_real),
            "facturado_por_kwp": _por(kwp, tot_fact),
            "sugerido_por_kwp": _por(kwp, tot_sugerido),
        }

        return {
            "secciones": secciones,
            "totales": totales,
            "metricas": metricas,
            "sin_datos": not secciones,
            "reconciliacion_completa": not (
                totales["valores_pendientes"] or totales["grupos_pendientes"]
            ),
        }

    # ─── WORKFLOW DE APROBACION ──────────────────────────────

    async def enviar_revision_ing(
        self, conn, id_bom: UUID, user_id: UUID,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Envia BOM a revision de responsable de ingenieria."""
        bom = await self.get_bom(conn, id_bom)
        capacidades = await self.get_capacidades_bom(conn, bom, user_id)
        if not capacidades["editar_base"]:
            raise ValueError("Solo el Ingeniero responsable o el RI pueden enviar el borrador")
        await self.validar_responsables_workflow_bom(conn, bom)

        # Verificar que tenga items
        items = await self.db.get_items_by_bom(conn, id_bom)
        if not items:
            raise ValueError("El BOM debe tener al menos un item")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.BORRADOR, EstatusBOM.EN_REVISION_ING,
            TipoAprobacion.ENVIO_REVISION_ING,
            lock_version_esperado=lock_version_esperado,
            fecha_envio_ing=now_mx(),
        )

        logger.info("BOM %s enviado a revision ing por %s", id_bom, user_id)

        return bom_updated

    async def aprobar_ing(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprueba BOM por responsable de ingenieria."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            bom.get('responsable_ing'), "Responsable de Ingenieria", "jefe_ingenieria"
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_ING:
            raise ValueError("El BOM debe estar EN_REVISION_ING para aprobar")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_ING, EstatusBOM.APROBADO_ING,
            TipoAprobacion.APROBACION_ING,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            fecha_aprobacion_ing=now_mx(),
        )
        logger.info("BOM %s aprobado por ing %s", id_bom, user_id)
        return bom_updated

    async def rechazar_ing(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Rechaza BOM por responsable de ingenieria. Vuelve a BORRADOR."""
        if not comentarios or not comentarios.strip():
            raise ValueError("El motivo del rechazo es obligatorio")

        bom = await self.get_bom(conn, id_bom)
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            bom.get('responsable_ing'), "Responsable de Ingenieria", "jefe_ingenieria"
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_ING:
            raise ValueError("El BOM debe estar EN_REVISION_ING para rechazar")

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_ING, EstatusBOM.BORRADOR,
            TipoAprobacion.RECHAZO_ING,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            invalidar_ciclo=True,
            **self._limpiar_fechas_flujo(),
        )
        logger.info("BOM %s rechazado por ing %s: %s", id_bom, user_id, comentarios)
        return bom_updated

    async def aprobar_const(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprueba BOM por jefe de construccion. Estado final antes de compras."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            bom.get('jefe_construccion'), "Jefe de Construccion", "jefe_construccion"
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_CONST:
            raise ValueError("El BOM debe estar EN_REVISION_CONST para aprobar")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_CONST, EstatusBOM.APROBADO_CONST,
            TipoAprobacion.APROBACION_CONST,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            fecha_aprobacion_const=now_mx(),
        )
        logger.info("BOM %s aprobado por const %s", id_bom, user_id)
        return bom_updated

    async def rechazar_const(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        destino_rechazo: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Rechaza BOM por construccion. Vuelve a Obra o a Borrador."""
        if not comentarios or not comentarios.strip():
            raise ValueError("El motivo del rechazo es obligatorio")
        bom = await self.get_bom(conn, id_bom)
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            bom.get('jefe_construccion'), "Jefe de Construccion", "jefe_construccion"
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_CONST:
            raise ValueError("El BOM debe estar EN_REVISION_CONST para rechazar")

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_CONST, EstatusBOM.BORRADOR,
            TipoAprobacion.RECHAZO_CONST,
            comentarios=comentarios,
            destino_rechazo="ingenieria",
            lock_version_esperado=lock_version_esperado,
            invalidar_ciclo=True,
            **self._limpiar_fechas_flujo(),
        )
        logger.info(
            "BOM %s rechazado por const %s hacia %s: %s",
            id_bom, user_id, "ingenieria", comentarios
        )
        return bom_updated

    async def enviar_revision_obra(
        self, conn, id_bom: UUID, user_id: UUID,
        user_role: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Envia BOM aprobado por ing a revision del coordinador de obra."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role, {bom.get("responsable_ing")},
            "el Responsable de Ingenieria",
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.APROBADO_ING:
            raise ValueError("El BOM debe estar APROBADO_ING para enviar a obra")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        # Resueltos en vivo (no la foto de tb_bom tomada al crear el paquete): si el
        # proyecto asigno o cambio Coordinador de Obra/Jefe de Construccion despues,
        # este es el ultimo punto donde se puede recapturar antes de que se necesiten.
        coordinador_obra = await self.db.get_asignacion_proyecto(
            conn, bom["id_proyecto"], "coordinador_obra", "CONSTRUCCION"
        )
        jefe_construccion = await self.db.get_responsable_proyecto_o_global(
            conn, bom["id_proyecto"], "jefe_construccion"
        )
        problemas = []
        if not coordinador_obra:
            problemas.append("falta Coordinador de Obra")
        if not jefe_construccion:
            problemas.append("falta Jefe de Construccion")
        if problemas:
            jefe_label = (
                jefe_construccion["nombre"] if jefe_construccion
                else "el Jefe de Construccion"
            )
            raise ValueError(
                "No se puede enviar a Obra: " + "; ".join(problemas)
                + f". Solicita a {jefe_label} que lo asigne."
            )

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.APROBADO_ING, EstatusBOM.EN_REVISION_OBRA,
            TipoAprobacion.ENVIO_REVISION_OBRA,
            lock_version_esperado=lock_version_esperado,
            fecha_envio_obra=now_mx(),
            coordinador_obra=coordinador_obra["id_usuario"],
            jefe_construccion=jefe_construccion["id_usuario"],
        )
        logger.info("BOM %s enviado a revision obra por %s", id_bom, user_id)
        return bom_updated

    async def aprobar_revision_obra(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprueba BOM por coordinador de obra. Avanza automaticamente a EN_REVISION_CONST."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role,
            {bom.get("coordinador_obra"), bom.get("jefe_construccion")},
            "el Coordinador de Obra o el Responsable de Construccion",
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_OBRA:
            raise ValueError("El BOM debe estar EN_REVISION_OBRA para aprobar")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_OBRA, EstatusBOM.EN_REVISION_CONST,
            TipoAprobacion.APROBACION_OBRA,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            fecha_aprobacion_obra=now_mx(),
            fecha_envio_const=now_mx(),
        )
        logger.info("BOM %s aprobado por obra %s, avanza a EN_REVISION_CONST", id_bom, user_id)
        return bom_updated

    async def rechazar_obra(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None, comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Rechaza BOM por coordinador de obra. Vuelve a BORRADOR."""
        if not comentarios or not comentarios.strip():
            raise ValueError("El motivo del rechazo es obligatorio")

        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role,
            {bom.get("coordinador_obra"), bom.get("jefe_construccion")},
            "el Coordinador de Obra o el Responsable de Construccion",
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_OBRA:
            raise ValueError("El BOM debe estar EN_REVISION_OBRA para rechazar")

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_OBRA, EstatusBOM.BORRADOR,
            TipoAprobacion.RECHAZO_OBRA,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            invalidar_ciclo=True,
            **self._limpiar_fechas_flujo(),
        )
        logger.info("BOM %s rechazado por obra %s: %s", id_bom, user_id, comentarios)
        return bom_updated

    async def devolver_a_borrador(
        self, conn, id_bom: UUID, user_id: UUID,
        comentarios: Optional[str] = None,
        user_role: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Devuelve BOM de APROBADO_ING a BORRADOR para corregir tras rechazo de construccion."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role, {bom.get("responsable_ing")},
            "el Responsable de Ingenieria",
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.APROBADO_ING:
            raise ValueError("Solo se puede devolver a borrador desde APROBADO_ING")

        actualizado = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.APROBADO_ING, EstatusBOM.BORRADOR,
            TipoAprobacion.DEVOLUCION_BORRADOR,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            invalidar_ciclo=True,
            **self._limpiar_fechas_flujo(),
        )

        logger.info("BOM %s devuelto a borrador por %s", id_bom, user_id)
        return actualizado

    async def cancelar_bom(
        self, conn, id_bom: UUID, user_id: UUID,
        comentarios: Optional[str] = None,
        user_role: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Cancela un BOM en BORRADOR."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role,
            {bom.get("ingeniero_responsable_id"), bom.get("responsable_ing")},
            "el Ingeniero responsable o el Responsable de Ingenieria",
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.BORRADOR:
            raise ValueError("Solo se puede cancelar un BOM en BORRADOR")
        if lock_version_esperado is None:
            raise ValueError(
                "Falta la revision del BOM; recarga el paquete e intenta de nuevo"
            )

        async with conn.transaction():
            paquete = await self.db.get_paquete_for_update(conn, bom["id_paquete"])
            bloqueado = await self.db.get_bom_for_update(conn, id_bom)
            if not paquete or not bloqueado or paquete["cabeza_trabajo_id"] != id_bom:
                raise ValueError("Solo se puede cancelar la cabeza de trabajo vigente")
            cancelado = await self.db.update_bom_estatus_cas(
                conn, id_bom, EstatusBOM.BORRADOR.value,
                lock_version_esperado, EstatusBOM.CANCELADO.value,
            )
            if not cancelado:
                raise ValueError(
                    "El BOM cambio desde que abriste la pagina; recarga el paquete e intenta de nuevo"
                )
            await self.db.registrar_aprobacion(
                conn, id_bom, TipoAprobacion.CANCELACION,
                bom["version"], user_id, bom["id_paquete"], comentarios=comentarios,
            )
            if paquete.get("cabeza_oficial_id"):
                cabeza = await self.db.actualizar_cabeza_trabajo(
                    conn, bom["id_paquete"], paquete["cabeza_oficial_id"],
                    paquete["lock_version"],
                )
            else:
                cabeza = await self.db.actualizar_estado_paquete_cas(
                    conn, bom["id_paquete"], paquete["lock_version"],
                    paquete["estado_paquete"], "CANCELADO",
                )
            if not cabeza:
                raise ValueError("No se pudo restaurar la cabeza vigente del paquete")
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM:{id_bom}:{cancelado['lock_version']}:CANCELACION",
                "CANCELACION", bom["id_proyecto"], user_id,
                {"version": bom["version"], "comentarios": comentarios},
                id_paquete=bom["id_paquete"], id_bom=id_bom,
            )

        logger.info("BOM %s cancelado por %s", id_bom, user_id)
        return await self.db.get_bom_by_id(conn, id_bom)

    async def get_ultimo_rechazo(self, conn, id_bom: UUID) -> Optional[dict]:
        """Obtiene el ultimo rechazo/devolucion del BOM."""
        return await self.db.get_ultimo_rechazo(conn, id_bom)

    async def solicitar_modificacion(
        self, conn, id_bom: UUID, user_id: UUID,
        comentarios: Optional[str] = None,
        user_role: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
        paquete_lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Crea vN+1 unicamente desde la cabeza oficial vigente del paquete."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_actor_asignado(
            conn, user_id, user_role,
            {bom.get("ingeniero_responsable_id"), bom.get("responsable_ing")},
            "el Ingeniero responsable o el Responsable de Ingenieria",
        )
        if EstatusBOM(bom["estatus"]) != EstatusBOM.APROBADO_FINAL:
            raise ValueError("Solo la cabeza oficial APROBADO_FINAL puede originar otra version")
        if lock_version_esperado is None or paquete_lock_version_esperado is None:
            raise ValueError("El paquete o BOM cambió; recarga la página.")

        async with conn.transaction():
            paquete = await self.db.get_paquete_for_update(conn, bom["id_paquete"])
            origen = await self.db.get_bom_for_update(conn, id_bom)
            if (
                not paquete
                or not origen
                or paquete["lock_version"] != paquete_lock_version_esperado
                or origen["lock_version"] != lock_version_esperado
                or paquete.get("cabeza_oficial_id") != id_bom
                or paquete.get("cabeza_trabajo_id") != id_bom
            ):
                raise ValueError(
                    "El paquete ya tiene otra version en curso o esta version dejo de ser oficial"
                )
            nueva_version = origen["version"] + 1
            nuevo_bom = await self.db.crear_bom(
                conn, origen["id_proyecto"], user_id,
                responsable_ing=origen.get("responsable_ing"),
                jefe_construccion=origen.get("jefe_construccion"),
                coordinador_obra=origen.get("coordinador_obra"),
                notas=(
                    f"Modificacion solicitada sobre v{origen['version']}. "
                    f"{comentarios or ''}"
                ).strip(),
                version=nueva_version,
                id_paquete=origen["id_paquete"],
                ingeniero_responsable_id=paquete["ingeniero_responsable_id"],
            )
            items_copiados = await self.db.copiar_items_a_nueva_version(
                conn, id_bom, nuevo_bom["id_bom"]
            )
            cabeza = await self.db.actualizar_cabeza_trabajo(
                conn, origen["id_paquete"], nuevo_bom["id_bom"],
                paquete_lock_version_esperado,
            )
            if not cabeza:
                raise ValueError("Otra solicitud de modificacion gano la carrera")
            await self.db.registrar_aprobacion(
                conn, id_bom, TipoAprobacion.SOLICITUD_MODIFICACION,
                origen["version"], user_id, origen["id_paquete"], comentarios=comentarios,
            )
            await self.db.registrar_historial(
                conn, nuevo_bom["id_bom"], AccionHistorial.CREADO,
                nueva_version, user_id,
                campo_modificado="version",
                valor_anterior=str(origen["version"]),
                valor_nuevo=str(nueva_version),
            )
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM:{nuevo_bom['id_bom']}:0:NUEVA_VERSION",
                "NUEVA_VERSION", origen["id_proyecto"], user_id,
                {"version_origen": origen["version"], "version_nueva": nueva_version},
                id_paquete=origen["id_paquete"], id_bom=nuevo_bom["id_bom"],
            )

        logger.info(
            "Nueva version BOM creada: proyecto=%s paquete=%s v%d->v%d items=%d",
            origen["id_proyecto"], origen["id_paquete"], origen["version"],
            nueva_version, items_copiados,
        )
        return await self.db.get_bom_by_id(conn, nuevo_bom["id_bom"])

    # ─── HISTORIAL Y APROBACIONES ────────────────────────────

    async def get_historial(
        self, conn, id_bom: UUID,
        usuario_id: Optional[UUID] = None,
        q: Optional[str] = None,
    ) -> list:
        """Lista historial de cambios, con filtro opcional por usuario y texto."""
        return await self.db.get_historial_by_bom(conn, id_bom, usuario_id, q)

    async def get_historial_usuarios(self, conn, id_bom: UUID) -> list:
        """Usuarios distintos con cambios registrados, para poblar el filtro de historial."""
        return await self.db.get_historial_usuarios(conn, id_bom)

    async def get_aprobaciones(self, conn, id_bom: UUID) -> list:
        """Lista aprobaciones/rechazos."""
        return await self.db.get_aprobaciones_by_bom(conn, id_bom)

    async def get_estadisticas(self, conn, id_bom: UUID, items: Optional[List[dict]] = None) -> dict:
        """Estadisticas del BOM. Los campos de costo se calculan en Python sobre
        `items` (ya trae costo_mxn resuelto por get_items via cadena de TC de 3
        niveles: XML->Banxico reciente->promedio 7d) en vez de en SQL, que no
        tenia forma de convertir USD y descartaba el total completo con solo un
        item en esa moneda."""
        estadisticas = await self.db.get_estadisticas_bom(conn, id_bom)
        if items is None:
            items = await self.get_items(conn, id_bom)
        estadisticas.update(self._calcular_estadisticas_costo(items))
        return estadisticas

    @staticmethod
    def _calcular_estadisticas_costo(items: List[dict]) -> dict:
        """Costo estimado total en MXN + contadores, sobre items BASE activos.

        Todo-o-nada: si algun item visible no resuelve costo en MXN (sin precio
        capturado, o en USD sin ningun TC disponible), el total es None — nunca
        un numero parcial que aparente estar completo."""
        base_activos = [
            i for i in items
            if i.get("activo", True) and (i.get("tipo_origen_item") or "BASE") == "BASE"
        ]
        items_con_precio = 0
        items_sin_costo = 0
        items_sin_tc = 0
        total = Decimal("0")
        total_resuelto = True
        total_mxn_usd = Decimal("0")
        total_usd_nativo = Decimal("0")
        for item in base_activos:
            # Mismo criterio que item_sin_costo: un precio capturado por
            # Ingenieria pero aun no confirmado por Compras no cuenta como
            # costo oficial, aunque precio_unitario > 0.
            if item.get("precio_pendiente_confirmacion"):
                items_sin_costo += 1
                total_resuelto = False
                continue
            precio = item.get("precio_unitario")
            tiene_precio = precio is not None and Decimal(str(precio)) > 0
            if not tiene_precio:
                items_sin_costo += 1
                total_resuelto = False
                continue
            items_con_precio += 1
            moneda = item.get("moneda")
            if moneda == "MXN":
                total += Decimal(str(item.get("importe") or 0))
            elif moneda == "USD" and item.get("costo_mxn") is not None:
                # costo_mxn es precio unitario en MXN; se multiplica por cantidad
                # para obtener el importe total del item (mismo patron que costo_real_mxn).
                cantidad_raw = item.get("cantidad")
                cantidad = Decimal(str(cantidad_raw)) if cantidad_raw is not None else Decimal("1")
                monto_mxn = Decimal(str(item["costo_mxn"])) * cantidad
                total += monto_mxn
                total_mxn_usd += monto_mxn
                total_usd_nativo += Decimal(str(item.get("importe") or 0))
            else:
                items_sin_tc += 1
                total_resuelto = False
        return {
            "costo_total_estimado": total if total_resuelto else None,
            "items_con_precio": items_con_precio,
            "items_sin_costo": items_sin_costo,
            "items_sin_tc": items_sin_tc,
            "tc_promedio": (
                (total_mxn_usd / total_usd_nativo)
                if total_usd_nativo and items_sin_tc == 0 else None
            ),
        }

    # ─── PERMISOS BOM-ROLE ───────────────────────────────────

    async def get_titulares_que_representa(self, conn, user_id: UUID) -> Set[UUID]:
        """Retorna el user_id + los titulares cuya suplencia activa tiene este usuario."""
        getter = getattr(self.db, "get_titulares_que_representa", None)
        titulares = await getter(conn, user_id) if getter else []
        result = set(titulares)
        result.add(user_id)
        return result

    async def es_bom_role(
        self, conn, bom: dict, user_id: UUID,
        representados: Optional[Set[UUID]] = None
    ) -> bool:
        """True si el usuario es (o representa via suplencia) alguno de los 3 roles del BOM."""
        if not bom:
            return False
        if representados is None:
            representados = await self.get_titulares_que_representa(conn, user_id)
        bom_roles = {
            bom.get('elaborado_por'),
            bom.get('responsable_ing'),
            bom.get('jefe_construccion'),
            bom.get('coordinador_obra'),
        } - {None}
        return bool(representados & bom_roles)

    @staticmethod
    def puede_versionar_bom(
        bom: dict, representados: Set[UUID], role: Optional[str] = None,
    ) -> bool:
        """ADMIN o titular/suplente del ingeniero responsable/de diseno del BOM."""
        if role == "ADMIN":
            return True
        return bool(
            representados
            & {bom.get("ingeniero_responsable_id"), bom.get("responsable_ing")}
            - {None}
        )

    # ─── GRUPOS BOM ─────────────────────────────────────────

    # ─── SUPLENCIAS ─────────────────────────────────────────

    async def set_item_grupos(
        self, conn, id_item: UUID, user_id: UUID, grupo_ids: List[int],
        area_editor: str = "ingenieria",
        grupo_porcentajes: Optional[dict[int, Decimal]] = None,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        if not grupo_ids:
            raise ValueError("Selecciona al menos un grupo BOM")
        resultado = await self.editar_item(
            conn, id_item, user_id, area_editor,
            lock_version_esperado=lock_version_esperado,
            user_role=user_role,
            rol_org=rol_org,
            module_roles=module_roles,
            grupo_ids=grupo_ids,
            grupo_porcentajes=grupo_porcentajes,
        )
        return {"capacidades": resultado["capacidades"]}

    async def get_suplencia_activa(self, conn, user_id: UUID) -> Optional[dict]:
        """Suplencia activa vigente del usuario (como titular)."""
        return await self.db.get_suplencia_activa_del_titular(conn, user_id)

    async def configurar_suplente(
        self, conn, titular_id: UUID, suplente_id: UUID, fecha_fin,
        id_esperado: Optional[int] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Configura suplente para el usuario. Valida que la fecha sea futura."""
        from datetime import date as date_type
        if isinstance(fecha_fin, str):
            fecha_fin = date_type.fromisoformat(fecha_fin)
        if fecha_fin < today_mx():
            raise ValueError("La fecha fin de la suplencia debe ser futura")
        if suplente_id == titular_id:
            raise ValueError("No puedes designarte como tu propio suplente")
        suplente = await self.db.get_usuario_activo_basico(conn, suplente_id)
        if not suplente:
            raise ValueError("El usuario suplente no existe o no esta activo")
        async with conn.transaction():
            suplencia = await self.db.crear_suplencia(
                conn, titular_id, suplente_id, fecha_fin,
                id_esperado, lock_version_esperado,
            )
            if not suplencia:
                raise ValueError("La suplencia cambió; vuelve a abrir el formulario.")
        return suplencia

    async def eliminar_suplencia(
        self, conn, titular_id: UUID, id_esperado: int,
        lock_version_esperado: int,
    ) -> None:
        """Desactiva la suplencia activa del usuario."""
        async with conn.transaction():
            eliminada = await self.db.desactivar_suplencia(
                conn, titular_id, id_esperado, lock_version_esperado
            )
            if not eliminada:
                raise ValueError("La suplencia cambió; vuelve a abrir el formulario.")

    # ─── APROBADOR FINAL ────────────────────────────────────

    async def enviar_revision_final(
        self, conn, id_bom: UUID, user_id: UUID, user_role: str,
        rol_org: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Envia BOM APROBADO_CONST a revision del aprobador final."""
        bom = await self.get_bom(conn, id_bom)
        await self._validar_aprobador_bom(
            conn, user_id, user_role, rol_org,
            bom.get('jefe_construccion'), "Jefe de Construccion", "jefe_construccion"
        )

        if EstatusBOM(bom['estatus']) != EstatusBOM.APROBADO_CONST:
            raise ValueError("El BOM debe estar APROBADO_CONST para enviar al aprobador final")
        await self.validar_sin_costos_pendientes(conn, id_bom)

        aprobador_id = await self._get_aprobador_final_direccion_id(conn)
        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.APROBADO_CONST, EstatusBOM.EN_REVISION_FINAL,
            TipoAprobacion.ENVIO_REVISION_FINAL,
            lock_version_esperado=lock_version_esperado,
            fecha_envio_final=now_mx(),
        )
        logger.info("BOM %s enviado a revision final por %s", id_bom, user_id)
        return bom_updated

    async def aprobar_final(
        self, conn, id_bom: UUID, user_id: UUID,
        comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Aprobacion final del BOM por el aprobador final."""
        bom = await self.get_bom(conn, id_bom)
        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_FINAL:
            raise ValueError("El BOM debe estar EN_REVISION_FINAL para aprobar")
        await self.validar_sin_costos_pendientes(conn, id_bom)
        if lock_version_esperado is None:
            raise ValueError(
                "Falta la revision del BOM; recarga el paquete e intenta de nuevo"
            )

        aprobador_id = await self._get_aprobador_final_direccion_id(conn)
        representados = await self.get_titulares_que_representa(conn, user_id)
        if aprobador_id not in representados:
            raise ValueError(
                "Solo el aprobador final designado o su suplente puede ejecutar esta accion"
            )

        async with conn.transaction():
            paquete = await self.db.get_paquete_for_update(conn, bom["id_paquete"])
            bloqueado = await self.db.get_bom_for_update(conn, id_bom)
            if not paquete or not bloqueado or not bloqueado.get("es_cabeza_trabajo"):
                raise ValueError("Solo la cabeza de trabajo vigente puede aprobarse")
            if paquete.get("estado_paquete") != "ACTIVO":
                raise ValueError("El paquete no esta activo")
            if EstatusBOM(bloqueado["estatus"]) != EstatusBOM.EN_REVISION_FINAL:
                raise ValueError("El BOM ya no esta en revision final")
            await self.db.lock_configuracion_proyecto(conn, bom["id_proyecto"])
            estado_proyecto = await self.db.get_estado_proyecto_for_update(conn, bom["id_proyecto"])
            metricas_fv = await self.db.get_metricas_paneles_proyecto(
                conn, bom["id_proyecto"]
            )
            totales = await self.db.get_totales_base_por_moneda(conn, id_bom)
            if totales.get("costos_desconocidos"):
                raise ValueError(
                    "No se puede congelar el presupuesto: hay costos o monedas desconocidos"
                )
            resuelto = await self.resolver_tipo_cambio(conn, bom["id_proyecto"], estado_proyecto)
            total_usd = Decimal(str(totales["total_usd"]))
            tasa = resuelto["tasa"]
            if total_usd and not tasa:
                raise ValueError(
                    "No hay tipo de cambio vigente para congelar el total oficial en MXN"
                )
            total_aprobado = Decimal(str(totales["total_mxn"]))
            if tasa:
                total_aprobado += total_usd * tasa
            actualizado = await self.db.update_bom_estatus_cas(
                conn, id_bom, EstatusBOM.EN_REVISION_FINAL.value,
                lock_version_esperado, EstatusBOM.APROBADO_FINAL.value,
                fecha_aprobacion_final=now_mx(),
                modulos_fv_snapshot=metricas_fv.get("modulos_fv"),
                potencia_pico_kwp_snapshot=metricas_fv.get("potencia_pico_kwp"),
                tipo_cambio_aprobacion=tasa,
                fecha_tipo_cambio_aprobacion=resuelto["fecha"] if tasa else None,
                subtotal_base_mxn_snapshot=totales["total_mxn"],
                subtotal_base_usd_snapshot=totales["total_usd"],
                total_aprobado_mxn=total_aprobado,
            )
            if not actualizado:
                raise ValueError(
                    "El BOM cambio desde que abriste la pagina; recarga el paquete e intenta de nuevo"
                )
            await self.db.registrar_aprobacion(
                conn, id_bom, TipoAprobacion.APROBACION_FINAL,
                bom["version"], user_id, bom["id_paquete"], comentarios=comentarios,
            )
            cabeza = await self.db.actualizar_cabeza_oficial(
                conn, bom["id_paquete"], id_bom, paquete["lock_version"]
            )
            if not cabeza:
                raise ValueError("No se pudo actualizar la cabeza oficial del paquete")
            await self.db.registrar_evento_outbox(
                conn,
                f"BOM:{id_bom}:{actualizado['lock_version']}:APROBACION_FINAL",
                "APROBACION_FINAL", bom["id_proyecto"], user_id,
                {
                    "version": bom["version"],
                    "modulos_fv_snapshot": metricas_fv.get("modulos_fv"),
                    "potencia_pico_kwp_snapshot": str(
                        metricas_fv.get("potencia_pico_kwp") or ""
                    ),
                    "total_aprobado_mxn": str(total_aprobado),
                },
                id_paquete=bom["id_paquete"], id_bom=id_bom,
            )
        logger.info("BOM %s aprobado final por %s", id_bom, user_id)
        bom_updated = await self.db.get_bom_by_id(conn, id_bom)
        return bom_updated

    async def rechazar_final(
        self, conn, id_bom: UUID, user_id: UUID,
        comentarios: Optional[str] = None,
        lock_version_esperado: Optional[int] = None,
    ) -> dict:
        """Rechazo por aprobador final. Vuelve a BORRADOR."""
        if not comentarios or not comentarios.strip():
            raise ValueError("El motivo del rechazo es obligatorio")

        bom = await self.get_bom(conn, id_bom)
        if EstatusBOM(bom['estatus']) != EstatusBOM.EN_REVISION_FINAL:
            raise ValueError("El BOM debe estar EN_REVISION_FINAL para rechazar")

        aprobador_id = await self._get_aprobador_final_direccion_id(conn)
        representados = await self.get_titulares_que_representa(conn, user_id)
        if aprobador_id not in representados:
            raise ValueError(
                "Solo el aprobador final designado o su suplente puede ejecutar esta accion"
            )

        bom_updated = await self._transicionar_bom(
            conn, id_bom, user_id,
            EstatusBOM.EN_REVISION_FINAL, EstatusBOM.BORRADOR,
            TipoAprobacion.RECHAZO_FINAL,
            comentarios=comentarios,
            lock_version_esperado=lock_version_esperado,
            invalidar_ciclo=True,
            **self._limpiar_fechas_flujo(),
        )
        logger.info("BOM %s rechazado final por %s: %s", id_bom, user_id, comentarios)
        return bom_updated

    async def get_aprobador_final_id(self, conn) -> Optional[UUID]:
        return await self.db.get_aprobador_final_id(conn)

    # ─── NOTIFICACIONES ──────────────────────────────────────

    async def notificar_items_sin_costo_compras(
        self, conn, id_bom: UUID, user_id: UUID
    ) -> dict:
        """Envia a Compras la lista consolidada de items del BOM sin costo."""
        # items primero: si no hay nada que notificar (el caso comun cuando esto
        # se dispara automatico tras aprobar una adenda), evita el get_bom extra.
        items = await self.get_items_sin_costo(conn, id_bom)
        if not items:
            raise ValueError("No hay items sin costo estimado para notificar.")
        bom = await self.get_bom(conn, id_bom)

        try:
            from core.workflow.notification_service import NotificationService
            notif = NotificationService()

            to_emails = await notif._get_emails_for_event(
                conn, BOM_COSTOS_EVENTO, "TO", BOM_COSTOS_REGLAS_MODULOS
            )
            cc_emails = await notif._get_emails_for_event(
                conn, BOM_COSTOS_EVENTO, "CC", BOM_COSTOS_REGLAS_MODULOS
            )
            bcc_emails = await notif._get_emails_for_event(
                conn, BOM_COSTOS_EVENTO, "CCO", BOM_COSTOS_REGLAS_MODULOS
            )
            sse_activa = await ConfigService.get_global_config(
                conn, BOM_COSTOS_SSE_KEY, False, bool
            )
            hay_correo = bool(to_emails or cc_emails or bcc_emails)
            if not hay_correo and not sse_activa:
                raise ValueError(
                    "No hay ningun canal configurado para notificar a Compras. "
                    "Activa el aviso interno o captura al menos un correo en "
                    "Admin > Configuracion BOM > Costos pendientes."
                )

            sse_notificados = 0
            if sse_activa:
                sse_notificados = await self._broadcast_costos_pendientes(conn, bom, items)

            correo_enviado = False
            if hay_correo:
                por_nombre = await self.db.get_usuario_nombre(conn, user_id)
                proyecto_id = bom.get("proyecto_id_estandar") or str(bom.get("id_proyecto"))
                format_ctx = {
                    "bom": bom,
                    "items": items,
                    "total_items": len(items),
                    "proyecto_id": proyecto_id,
                    "proyecto_nombre": bom.get("proyecto_nombre") or "",
                    "paquete_codigo": bom.get("paquete_codigo") or "BOM",
                    "paquete_nombre": bom.get("paquete_nombre") or "",
                    "version": bom.get("version", ""),
                    "por_nombre": por_nombre or "Sistema",
                    "app_url": f"{settings.APP_BASE_URL}/bom/versiones/{bom.get('id_bom')}/ui",
                }
                sender = await notif._get_notification_sender(conn, "BOM")
                subject_template = await ConfigService.get_global_config(
                    conn, BOM_COSTOS_ASUNTO_KEY, BOM_COSTOS_DEFAULT_ASUNTO, str
                )
                body_template = await ConfigService.get_global_config(
                    conn, BOM_COSTOS_TEMPLATE_KEY, BOM_COSTOS_DEFAULT_TEMPLATE, str
                )
                subject = self._format_template(subject_template, format_ctx).strip()
                mensaje = self._format_template(body_template, format_ctx).strip()
                if not subject:
                    subject = self._format_template(BOM_COSTOS_DEFAULT_ASUNTO, format_ctx)
                if not mensaje:
                    mensaje = self._format_template(BOM_COSTOS_DEFAULT_TEMPLATE, format_ctx)

                html = notif._render_template("shared/emails/bom/items_sin_costo.html", {
                    **format_ctx,
                    "mensaje": mensaje,
                })
                correo_enviado = await notif._send_email(
                    to_emails,
                    cc_emails,
                    subject,
                    html,
                    sender["email"],
                    bcc_emails=bcc_emails,
                )
                if not correo_enviado:
                    logger.warning(
                        "BOM costos pendientes: correo no enviado (revisar destinatarios TO) bom=%s",
                        id_bom,
                    )

            if not correo_enviado and not sse_notificados:
                raise ValueError("No se pudo notificar a Compras por ningun canal.")

            logger.info(
                "BOM costos pendientes notificados: bom=%s items=%d to=%d cc=%d cco=%d sse=%d por=%s",
                id_bom, len(items), len(to_emails), len(cc_emails), len(bcc_emails),
                sse_notificados, user_id,
            )
            return {
                "items_sin_costo": len(items),
                "destinatarios": len(to_emails),
                "cc": len(cc_emails),
                "cco": len(bcc_emails),
                "sse": sse_notificados,
                "correo_enviado": correo_enviado,
            }
        except ValueError:
            raise
        except (TemplateError, httpx.HTTPError, RuntimeError, TypeError, KeyError) as exc:
            logger.warning("BOM costos pendientes: error notificando compras: %s", exc)
            raise ValueError("No se pudo enviar la notificacion a Compras.") from exc

    async def _broadcast_costos_pendientes(self, conn, bom: dict, items: list[dict]) -> int:
        """Crea aviso SSE para usuarios activos de Compras usando un tipo permitido."""
        usuarios_compras = await self.db.get_usuarios_por_area(conn, "compras", solo_jefes=False)
        if not usuarios_compras:
            return 0

        notif_svc = get_notifications_service()
        titulo = (
            f"BOM {bom.get('proyecto_id_estandar', '')} - "
            f"{bom.get('paquete_codigo', 'BOM')} v{bom.get('version', '')} - "
            "Items sin costo estimado"
        )
        mensaje = f"{len(items)} item(s) sin costo estimado pendiente(s) de actualizar."
        count = 0
        for usuario in usuarios_compras:
            usuario_id = usuario.get("id_usuario")
            if not usuario_id:
                continue
            notification_data = await notif_svc.create_notification(
                conn=conn,
                usuario_id=usuario_id,
                tipo="CAMBIO_ESTATUS",
                titulo=titulo,
                mensaje=mensaje,
                id_oportunidad=bom.get("id_oportunidad"),
                modulo_origen="bom",
            )
            await notif_svc.broadcast_to_user(conn, usuario_id, notification_data)
            count += 1
        return count

    async def _broadcast_bom(self, conn, to_user_id, tipo: str, titulo: str, proyecto_nombre: str) -> None:
        notif_svc = get_notifications_service()
        tipo_notificacion = (
            tipo
            if tipo in {"ASIGNACION", "CAMBIO_ESTATUS", "NUEVO_COMENTARIO"}
            else "CAMBIO_ESTATUS"
        )
        notification_data = await notif_svc.create_notification(
            conn=conn,
            usuario_id=to_user_id,
            tipo=tipo_notificacion,
            titulo=titulo,
            mensaje=f"Proyecto: {proyecto_nombre}",
            modulo_origen="bom",
        )
        await notif_svc.broadcast_to_user(conn, to_user_id, notification_data)

    async def _notify_bom(
        self, conn, bom: dict,
        to_user_id,
        evento: str,
        por_user_id=None,
        comentarios: Optional[str] = None
    ) -> None:
        """Envia email de notificacion de cambio de estado BOM. Fire-and-forget."""
        if not to_user_id:
            return
        try:
            from core.workflow.notification_service import NotificationService
            notif = NotificationService()

            to_email = await self.db.get_usuario_email(conn, to_user_id)
            if not to_email:
                return

            sender_email = await self.db.get_sender_email(conn, 'DEFAULT')
            if not sender_email:
                logger.warning("BOM notify: no DEFAULT sender email configurado")
                return

            por_nombre = None
            if por_user_id:
                por_nombre = await self.db.get_usuario_nombre(conn, por_user_id)

            html = notif._render_template('shared/emails/bom/bom_revision.html', {
                'bom': bom,
                'evento': evento,
                'por_nombre': por_nombre or 'Sistema',
                'comentarios': comentarios,
                'app_url': f"{settings.APP_BASE_URL}/bom/versiones/{bom.get('id_bom')}/ui",
            })

            identidad = (
                f"{bom.get('proyecto_id_estandar', '')} - "
                f"{bom.get('paquete_codigo', 'BOM')} v{bom.get('version', '')}"
            )

            subject_map = {
                'ENVIADO_REVISION_ING':   f"BOM {identidad} - Revision requerida (Ingenieria)",
                'APROBADO_ING':           f"BOM {identidad} - Aprobado por Ingenieria",
                'RECHAZADO_ING':          f"BOM {identidad} - Devuelto por Ingenieria",
                'ENVIADO_REVISION_OBRA':  f"BOM {identidad} - Revision requerida (Obra)",
                'RECHAZADO_OBRA':         f"BOM {identidad} - Devuelto por Obra",
                'ENVIADO_REVISION_CONST': f"BOM {identidad} - Revision requerida (Construccion)",
                'APROBADO_CONST':         f"BOM {identidad} - Aprobado por Construccion",
                'RECHAZADO_CONST':        f"BOM {identidad} - Devuelto por Construccion",
                'ENVIADO_REVISION_FINAL': f"BOM {identidad} - Aprobacion final requerida",
                'APROBADO_FINAL':         f"BOM {identidad} - Aprobado definitivamente",
                'RECHAZADO_FINAL':        f"BOM {identidad} - Devuelto por Aprobador Final",
                'FALTA_COORDINADOR_OBRA': f"BOM {identidad} - Asignar coordinador de obra",
            }
            subject = subject_map.get(evento, f"BOM {identidad} - Actualizacion")

            await notif._send_email({to_email}, set(), subject, html, sender_email)
            logger.info("BOM notify enviada: evento=%s to_user=%s", evento, to_user_id)
            await self._broadcast_bom(conn, to_user_id, f"BOM_{evento}", subject, bom.get('proyecto_nombre', ''))
        except (asyncpg.PostgresError, KeyError, RuntimeError, TemplateError, TypeError, ValueError) as exc:
            logger.warning("BOM notify: error enviando email, evento=%s: %s", evento, exc)

    # ─── CATALOGOS ──────────────────────────────────────────

    def _catalogos_cache_vigente(self) -> Optional[dict]:
        cached = BomService._cache_catalogos
        if cached is None:
            return None
        ts, data = cached
        return data if time.time() - ts < BomService._CATALOGOS_TTL_SECONDS else None

    async def get_catalogos(self, conn) -> dict:
        """Obtiene todos los catalogos necesarios para formularios. Cacheado en memoria (TTL corto)."""
        data = self._catalogos_cache_vigente()
        if data is not None:
            return data

        async with BomService._cache_catalogos_refresh_lock:
            # Reintenta tras esperar el lock: otro request concurrente pudo
            # haber refrescado el cache mientras esperabamos.
            data = self._catalogos_cache_vigente()
            if data is not None:
                return data

            tipos_entrega = await self.db.get_tipos_entrega(conn)
            categorias = await self.db.get_categorias_compra(conn)
            unidades_medida = await self.db.get_unidades_medida(conn)
            proveedores = await self.db.get_proveedores(conn)
            usuarios_ing_jefes = await self.db.get_usuarios_por_area(conn, 'ingenieria', solo_jefes=True)
            usuarios_ing = await self.db.get_usuarios_por_area(conn, 'ingenieria', solo_jefes=False)

            usuarios_const_jefes = await self.db.get_usuarios_por_area(conn, 'construccion', solo_jefes=True)
            usuarios_const = await self.db.get_usuarios_por_area(conn, 'construccion', solo_jefes=False)
            grupos_bom = await self.db.get_grupos_bom(conn)

            data = {
                'tipos_entrega': tipos_entrega,
                'categorias': categorias,
                'unidades_medida': unidades_medida,
                'proveedores': proveedores,
                'usuarios_ing': usuarios_ing,           # Lista completa (por si se requiere)
                'usuarios_ing_jefes': usuarios_ing_jefes, # Solo jefes (para Responsable de Ing)
                'usuarios_const': usuarios_const,       # Lista completa (para Coordinador de Obra)
                'usuarios_const_jefes': usuarios_const_jefes, # Solo jefes (para Jefe de Construccion)
                'grupos_bom': grupos_bom,
            }
            BomService._cache_catalogos = (time.time(), data)
            return data

    # ─── PANELES FV DEL PROYECTO ────────────────────────────

    async def paneles_configurados(self, conn, id_proyecto: UUID) -> bool:
        return await self.db.existen_paneles_proyecto(conn, id_proyecto)

    async def get_paneles_proyecto(self, conn, id_proyecto: UUID) -> list[dict]:
        return await self.db.get_paneles_proyecto(conn, id_proyecto)

    async def get_paneles_fv_activos(self, conn) -> list[dict]:
        return await self.db.get_paneles_fv_activos(conn)

    async def guardar_paneles_proyecto(
        self, conn, id_proyecto: UUID, paneles: list[dict], user_id: UUID
    ) -> None:
        """Reemplaza los paneles FV del proyecto. Solo Ingenieria (mismo guard que el BOM)."""
        await self._validar_retomar_bom_ingenieria(conn, id_proyecto, user_id)
        if not paneles:
            raise ValueError("Agrega al menos un panel")
        for p in paneles:
            if p["cantidad"] <= 0:
                raise ValueError("La cantidad debe ser mayor a 0")
        async with conn.transaction():
            await self.db.lock_configuracion_proyecto(conn, id_proyecto)
            await self.db.reemplazar_paneles_proyecto(conn, id_proyecto, paneles, user_id)

    # ─── EXPORT EXCEL ────────────────────────────────────────

    async def get_consolidado_proyecto(
        self, conn, id_proyecto: UUID, modo: str = "CURSO",
        proyecto: Optional[dict] = None,
    ) -> dict:
        """Lee cabezas, hechos y divisor dentro de un snapshot consistente.

        `proyecto` es opcional: si el caller ya lo consulto en la misma
        peticion (ej. bom_hub_ui), se pasa aqui para evitar volver a pedirlo.
        Es seguro reusarlo porque solo se usa como chequeo de existencia, no
        alimenta los totales financieros. `todos_paquetes`/`estado`, en
        cambio, SI deben leerse dentro de esta transaccion repeatable_read:
        pasarlos desde afuera rompe la consistencia frente a `paquetes`/
        `lineas`, que si se leen frescos aqui adentro.
        """
        if conn.__class__.__module__.startswith("asyncpg"):
            async with conn.transaction(isolation="repeatable_read", readonly=True):
                return await self._get_consolidado_proyecto_snapshot(
                    conn, id_proyecto, modo, proyecto
                )
        return await self._get_consolidado_proyecto_snapshot(
            conn, id_proyecto, modo, proyecto
        )

    async def _get_consolidado_proyecto_snapshot(
        self, conn, id_proyecto: UUID, modo: str = "CURSO",
        proyecto: Optional[dict] = None,
    ) -> dict:
        """Read model del conjunto sin crear un BOM consolidado persistente."""
        modo_normalizado = (modo or "CURSO").strip().upper()
        if modo_normalizado not in {"CURSO", "OFICIAL"}:
            raise ValueError("La vista consolidada debe ser CURSO u OFICIAL")
        if not (proyecto or await self.db.get_proyecto_info(conn, id_proyecto)):
            raise ValueError("Proyecto no encontrado")

        paquetes = await self.db.get_consolidado_paquetes(
            conn, id_proyecto, modo_normalizado
        )
        lineas = await self.db.get_consolidado_lineas(
            conn, id_proyecto, modo_normalizado
        )
        todos_paquetes = await self.db.listar_paquetes_proyecto(conn, id_proyecto)
        estado = await self.get_estado_conjunto(conn, id_proyecto)
        resuelto_curso = await self.resolver_tipo_cambio(conn, id_proyecto, estado)
        tasa_curso = resuelto_curso["tasa"]

        if modo_normalizado == "CURSO":
            for paquete in paquetes:
                total_raw = paquete.get("presupuesto_mxn")
                usd_raw = paquete.get("presupuesto_usd")
                if total_raw is None or usd_raw is None:
                    paquete["presupuesto_total_mxn"] = None
                    continue
                total = Decimal(str(total_raw))
                usd = Decimal(str(usd_raw))
                paquete["presupuesto_total_mxn"] = (
                    total + usd * tasa_curso
                    if tasa_curso is not None
                    else (None if usd != 0 else total)
                )

        for paquete in paquetes:
            cot_mxn = Decimal(str(paquete.get("cotizado_mxn") or 0))
            cot_usd = Decimal(str(paquete.get("cotizado_usd") or 0))
            paquete["cotizado_total_mxn"] = (
                cot_mxn + cot_usd * tasa_curso
                if tasa_curso is not None
                else (None if cot_usd != 0 else cot_mxn)
            )

        nombres_totales = (
            "presupuesto_mxn", "presupuesto_usd", "presupuesto_total_mxn",
            "cotizado_mxn", "cotizado_usd", "cotizado_total_mxn",
            "autorizado_mxn", "autorizado_usd",
            "autorizado_total_mxn", "facturado_mxn", "facturado_usd",
            "facturado_total_mxn", "pagado_mxn", "pagado_usd", "pagado_total_mxn",
        )
        totales = {}
        for nombre in nombres_totales:
            valores = [paquete.get(nombre) for paquete in paquetes]
            totales[nombre] = (
                None
                if any(valor is None for valor in valores)
                else sum((Decimal(str(valor)) for valor in valores), Decimal("0"))
            )
        conversion_pendiente = (
            modo_normalizado == "CURSO"
            and tasa_curso is None
            and any(
                p.get("presupuesto_usd") is not None
                and Decimal(str(p["presupuesto_usd"])) != 0
                for p in paquetes
            )
        )
        if conversion_pendiente:
            totales["presupuesto_total_mxn"] = None
        presupuesto = totales["presupuesto_total_mxn"]
        autorizado = totales["autorizado_total_mxn"]
        totales["diferencia_mxn"] = (
            presupuesto - autorizado
            if presupuesto is not None and autorizado is not None else None
        )
        totales["porcentaje_autorizado"] = (
            autorizado / presupuesto * 100
            if presupuesto and autorizado is not None else None
        )

        if modo_normalizado == "OFICIAL":
            divisor = await self.db.get_divisor_oficial_consolidado(conn, id_proyecto)
            modulos = divisor.get("modulos_fv_snapshot") if divisor else None
            potencia = divisor.get("potencia_pico_kwp_snapshot") if divisor else None
        else:
            divisor = None
            metricas = await self.db.get_metricas_paneles_proyecto(conn, id_proyecto)
            modulos = metricas.get("modulos_fv")
            potencia = metricas.get("potencia_pico_kwp")

        totales["mxn_por_modulo"] = (
            presupuesto / modulos if presupuesto is not None and modulos else None
        )
        totales["mxn_por_kwp"] = (
            presupuesto / Decimal(str(potencia))
            if presupuesto is not None and potencia else None
        )

        grupos = defaultdict(lambda: {
            "presupuesto_mxn": Decimal("0"),
            "presupuesto_usd": Decimal("0"),
            "facturado_mxn": Decimal("0"),
            "lineas": 0,
            "paquetes": set(),
            "presupuesto_pendiente": False,
        })
        for linea in lineas:
            codigos = list(linea.get("grupos") or []) or ["SIN_GRUPO"]
            distribucion = linea.get("distribucion_grupos") or {}
            if isinstance(distribucion, str):
                distribucion = json.loads(distribucion or "{}")
            if distribucion:
                asignaciones = [
                    (codigo, Decimal(str(porcentaje)))
                    for codigo, porcentaje in distribucion.items()
                ]
            elif len(codigos) == 1:
                asignaciones = [(codigos[0], Decimal("1"))]
            else:
                asignaciones = [("PENDIENTE_ASIGNACION", Decimal("1"))]
            importe_raw = linea.get("costo_estimado")
            for codigo, porcentaje in asignaciones:
                grupos[codigo]["lineas"] += 1
                grupos[codigo]["paquetes"].add(linea["paquete_codigo"])
                if importe_raw is None or linea.get("moneda") not in {"MXN", "USD"}:
                    grupos[codigo]["presupuesto_pendiente"] = True
                    continue
                importe = Decimal(str(importe_raw)) * porcentaje
                clave = (
                    "presupuesto_usd" if linea.get("moneda") == "USD"
                    else "presupuesto_mxn"
                )
                grupos[codigo][clave] += importe
            facturado_por_grupo = linea.get("facturado_por_grupo") or {}
            if isinstance(facturado_por_grupo, str):
                facturado_por_grupo = json.loads(facturado_por_grupo or "{}")
            for codigo, importe_facturado in facturado_por_grupo.items():
                if importe_facturado is None:
                    grupos[codigo]["presupuesto_pendiente"] = True
                else:
                    grupos[codigo]["facturado_mxn"] += Decimal(
                        str(importe_facturado)
                    )

        desglose_grupos = []
        for codigo, datos in sorted(grupos.items()):
            desglose_grupos.append({
                **datos,
                "codigo": codigo,
                "paquetes": sorted(datos["paquetes"]),
            })

        solapamientos = [
            linea for linea in lineas if linea.get("posible_solapamiento")
        ]
        pendientes_oficiales = sum(
            1 for paquete in todos_paquetes
            if paquete.get("estado_paquete") == "ACTIVO"
            and not paquete.get("version_oficial")
        )
        return {
            "modo": modo_normalizado,
            "paquetes": paquetes,
            "lineas": lineas,
            "totales": totales,
            "desglose_grupos": desglose_grupos,
            "solapamientos": solapamientos,
            "captura_cerrada": bool(estado.get("captura_cerrada")),
            "estado_lock_version": estado.get("lock_version", 0),
            "pendientes_oficiales": pendientes_oficiales,
            "divisor_fv": {
                "modulos_fv": modulos,
                "potencia_pico_kwp": potencia,
                "origen": divisor,
            },
            "tipo_cambio_curso": resuelto_curso if modo_normalizado == "CURSO" else None,
            "conversion_pendiente": conversion_pendiente,
            "tipo_cambio_manual_info": (
                await self.db.get_tipo_cambio_manual_info(conn, id_proyecto)
                if modo_normalizado == "CURSO" and resuelto_curso["origen"] == "manual"
                else None
            ),
        }

    async def export_consolidado_excel(
        self, conn, id_proyecto: UUID, modo: str = "CURSO",
    ) -> bytes:
        """Exporta el read model consolidado con procedencia por paquete y linea."""
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        consolidado = await self.get_consolidado_proyecto(conn, id_proyecto, modo)
        proyecto = await self.db.get_proyecto_info(conn, id_proyecto)
        wb = Workbook()
        resumen = wb.active
        resumen.title = "Resumen"
        encabezado = PatternFill("solid", fgColor="1F4E79")
        fuente = Font(bold=True, color="FFFFFF")

        resumen.append([
            "Paquete", "Nombre", "Versión", "Estatus", "Presupuesto MXN",
            "Presupuesto USD", "Presupuesto total MXN", "Cotizado total MXN",
            "Autorizado total MXN", "Facturado total MXN", "Pagado total MXN",
            "Fecha snapshot",
        ])
        for celda in resumen[1]:
            celda.fill = encabezado
            celda.font = fuente
        for paquete in consolidado["paquetes"]:
            resumen.append([
                paquete["codigo"], paquete["nombre"], paquete["version"],
                paquete["estatus"],
                (
                    float(paquete["presupuesto_mxn"])
                    if paquete.get("presupuesto_mxn") is not None else None
                ),
                (
                    float(paquete["presupuesto_usd"])
                    if paquete.get("presupuesto_usd") is not None else None
                ),
                (
                    float(paquete["presupuesto_total_mxn"])
                    if paquete.get("presupuesto_total_mxn") is not None else None
                ),
                (
                    float(paquete["cotizado_total_mxn"])
                    if paquete.get("cotizado_total_mxn") is not None else None
                ),
                (
                    float(paquete["autorizado_total_mxn"])
                    if paquete.get("autorizado_total_mxn") is not None else None
                ),
                (
                    float(paquete["facturado_total_mxn"])
                    if paquete.get("facturado_total_mxn") is not None else None
                ),
                (
                    float(paquete["pagado_total_mxn"])
                    if paquete.get("pagado_total_mxn") is not None else None
                ),
                paquete.get("fecha_aprobacion_final"),
            ])

        lineas_ws = wb.create_sheet("Líneas")
        lineas_ws.append([
            "Proyecto", "Paquete", "Nombre paquete", "Versión", "ID BOM",
            "ID línea estable", "Descripción", "Cantidad", "Unidad", "Grupos",
            "Moneda", "Precio unitario", "Costo estimado", "Costo facturado",
            "Estado", "Posible solapamiento", "Paquetes solapados",
        ])
        for celda in lineas_ws[1]:
            celda.fill = encabezado
            celda.font = fuente
        proyecto_id = proyecto.get("proyecto_id_estandar") if proyecto else str(id_proyecto)
        for linea in consolidado["lineas"]:
            lineas_ws.append([
                proyecto_id, linea["paquete_codigo"], linea["paquete_nombre"],
                linea["version"], str(linea["id_bom"]), str(linea["id_linea_bom"]),
                linea["descripcion"], (
                    float(linea["cantidad"]) if linea.get("cantidad") is not None else None
                ),
                linea.get("unidad_medida"), ", ".join(linea.get("grupos") or []),
                linea["moneda"], (
                    float(linea["precio_unitario"])
                    if linea.get("precio_unitario") is not None else None
                ),
                (
                    float(linea["costo_estimado"])
                    if linea.get("costo_estimado") is not None else None
                ),
                (
                    float(linea["costo_facturado"])
                    if linea.get("costo_facturado") is not None else None
                ),
                linea.get("estado_ejecucion"),
                "Sí" if linea.get("posible_solapamiento") else "No",
                ", ".join(linea.get("paquetes_solapados") or []),
            ])

        for hoja in (resumen, lineas_ws):
            hoja.freeze_panes = "A2"
            hoja.auto_filter.ref = hoja.dimensions
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def export_to_excel(self, conn, id_bom: UUID) -> bytes:
        """Genera archivo Excel con los items del BOM."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        bom = await self.get_bom(conn, id_bom)
        items = await self.get_items(conn, id_bom)

        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Materiales"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Info cabecera BOM (filas 1-5)
        info_font = Font(bold=True, size=10)
        ws.cell(row=1, column=1, value="Proyecto:").font = info_font
        ws.cell(row=1, column=2, value=f"{bom.get('proyecto_id_estandar', '')} - {bom.get('proyecto_nombre', '')}")
        ws.cell(row=2, column=1, value="Version:").font = info_font
        ws.cell(row=2, column=2, value=bom.get('version', 1))
        ws.cell(row=3, column=1, value="Estatus:").font = info_font
        ws.cell(row=3, column=2, value=bom.get('estatus', ''))
        ws.cell(row=4, column=1, value="Elaborado por:").font = info_font
        ws.cell(row=4, column=2, value=bom.get('elaborado_por_nombre', ''))
        ws.cell(row=5, column=1, value="Responsable Ing:").font = info_font
        ws.cell(row=5, column=2, value=bom.get('responsable_ing_nombre', ''))

        # Headers de tabla (fila 7)
        headers_row = 7
        headers = [
            "#", "Categoria", "Descripcion", "Cantidad", "Unidad",
            "Precio Unitario", "Importe", "Costo Real", "Estado",
            "Fecha Requerida", "Fecha Llegada Real", "Recepcion", "Proveedor",
            "Tipo Entrega", "Fecha Estimada Entrega", "Comentarios", "Entregado"
        ]

        ESTADO_LABELS = {
            'FACTURADO': 'Facturado', 'PAGADO': 'Pagado',
            'AUTORIZADO': 'Autorizado', 'COTIZADO': 'Cotizado',
        }

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=headers_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Datos
        total_importe = 0
        for row_num, item in enumerate(items, headers_row + 1):
            precio = float(item.get('precio_unitario') or 0)
            cantidad = float(item.get('cantidad') or 0)
            importe = cantidad * precio
            total_importe += importe

            precio_real = item.get('precio_real')
            if precio_real is not None:
                importe_real = item.get('importe_real')
                costo_real = float(importe_real) if importe_real is not None else float(precio_real) * cantidad
            else:
                gasto_real = item.get('gasto_real')
                costo_real = float(gasto_real) if gasto_real is not None else None

            cantidad_recibida = float(item.get('cantidad_recibida') or 0)
            pct_recepcion = (cantidad_recibida / cantidad * 100) if cantidad else 0

            row_data = [
                row_num - headers_row,
                item.get('categoria_nombre', ''),
                item.get('descripcion', ''),
                item.get('cantidad', 0),
                item.get('unidad_medida', ''),
                float(precio) if precio else None,
                importe if precio else None,
                costo_real,
                ESTADO_LABELS.get(item.get('estatus_compra'), 'Pendiente'),
                item['fecha_requerida'].strftime("%d/%m/%Y") if item.get('fecha_requerida') else '',
                item['fecha_llegada_real'].strftime("%d/%m/%Y") if item.get('fecha_llegada_real') else '',
                f"{min(pct_recepcion, 100):.0f}%" if pct_recepcion > 0 else '',
                item.get('proveedor_nombre', ''),
                item.get('tipo_entrega', ''),
                item['fecha_estimada_entrega'].strftime("%d/%m/%Y") if item.get('fecha_estimada_entrega') else '',
                item.get('comentarios', ''),
                'Si' if item.get('entregado') else 'No',
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num == 4:
                    cell.number_format = '#,##0.0000'
                    cell.alignment = Alignment(horizontal="right")
                elif col_num in (6, 7, 8):
                    cell.number_format = '$#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Fila de total
        if items:
            total_row = headers_row + len(items) + 1
            total_font = Font(bold=True, size=11)
            ws.cell(row=total_row, column=3, value="TOTAL").font = total_font
            ws.cell(row=total_row, column=3).border = thin_border
            cell_total = ws.cell(row=total_row, column=7, value=total_importe)
            cell_total.font = total_font
            cell_total.number_format = '$#,##0.00'
            cell_total.alignment = Alignment(horizontal="right")
            cell_total.border = thin_border

        # Anchos de columna
        column_widths = [5, 20, 40, 12, 10, 16, 16, 16, 14, 16, 16, 12, 25, 16, 18, 30, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = f"A{headers_row + 1}"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ─── HELPERS INTERNOS ────────────────────────────────────

    async def _validar_edicion_items(self, conn, id_bom: UUID, area_editor: str) -> dict:
        """Valida que el BOM estÃ© en estado editable para agregar/eliminar items segun el area."""
        bom = await self.get_bom(conn, id_bom)
        estatus = EstatusBOM(bom['estatus'])

        if area_editor == 'ingenieria':
            if estatus not in ESTATUS_EDITABLE_ING:
                raise ValueError(
                    f"El BOM esta en estado {estatus} y no permite edicion estructural por Ingenieria."
                )
        elif area_editor == 'construccion':
            if estatus not in ESTATUS_EDITABLE_CONST_COMPRAS:
                raise ValueError(
                    f"El BOM esta en estado {estatus} y no permite edicion estructural por Construccion."
                )
        else:
            if estatus not in ESTATUS_EDITABLE_ING:
                raise ValueError("Area de edicion no reconocida o estado invalido.")
        
        return bom

    # ========================================
    # DB PROXIES (router -> service -> db)
    # ========================================

    async def get_proyecto_info(self, conn, id_proyecto: UUID) -> Optional[dict]:
        return await self.db.get_proyecto_info(conn, id_proyecto)

    async def restaurar_item(
        self, conn, id_item: UUID, user_id: UUID,
        lock_version_esperado: Optional[int] = None,
        user_role: Optional[str] = None,
        rol_org: Optional[str] = None,
        module_roles: Optional[dict] = None,
    ) -> dict:
        item = await self.db.get_item_by_id(conn, id_item)
        if not item:
            raise ValueError("Item no encontrado")
        if item.get("activo", True):
            raise ValueError("El item ya esta activo")
        async with conn.transaction():
            bom, capacidades = await self._reservar_mutacion_base(
                conn, item["id_bom"], user_id, lock_version_esperado,
                user_role, rol_org, module_roles,
            )
            restaurado = await self.db.restaurar_item(conn, id_item)
            await self.db.registrar_historial(
                conn, item["id_bom"], AccionHistorial.RESTAURADO,
                bom["version"], user_id,
                id_item=id_item,
                campo_modificado="item",
                valor_nuevo=item.get("descripcion"),
            )
        return {"restaurado": restaurado, "capacidades": capacidades}

    async def buscar_materiales_para_bom(self, conn, query: str, **kwargs) -> dict:
        return await self.db.buscar_materiales_para_bom(conn, query, **kwargs)

    async def get_materiales_recientes(self, conn, limite: int = 10, offset: int = 0) -> dict:
        return await self.db.get_materiales_recientes(conn, limite=limite, offset=offset)

    async def get_usuarios_por_area(self, conn, module_slug: str, solo_jefes: bool = False) -> List[dict]:
        return await self.db.get_usuarios_por_area(conn, module_slug, solo_jefes=solo_jefes)

    async def get_bom_by_id(self, conn, id_bom: UUID) -> Optional[dict]:
        return await self.db.get_bom_by_id(conn, id_bom)

    async def get_cotizacion_by_id(self, conn, cotizacion_id: UUID) -> Optional[dict]:
        return await self.db.get_cotizacion_by_id(conn, cotizacion_id)

    async def get_items_by_ids(self, conn, item_ids: List[UUID]) -> List[dict]:
        return await self.db.get_items_by_ids(conn, item_ids)

    async def get_items_cotizacion(self, conn, cotizacion_id: UUID) -> List[dict]:
        return await self.db.get_items_cotizacion(conn, cotizacion_id)

    async def actualizar_pdf_cotizacion(
        self, conn, cotizacion_id: UUID, pdf_url: str,
        lock_version_esperado: Optional[int] = None,
    ) -> Optional[dict]:
        if lock_version_esperado is None:
            raise ValueError("La cotizacion cambio; recarga la pestaña")
        updated = await self.db.actualizar_pdf_cotizacion(
            conn, cotizacion_id, pdf_url, lock_version_esperado
        )
        if not updated:
            cotizacion = await self.db.get_cotizacion_by_id(conn, cotizacion_id)
            if not cotizacion:
                raise ValueError("Cotización no encontrada.")
            raise ValueError(f"La cotización está en estatus {cotizacion['estatus']} y no puede modificarse.")
        return updated


def get_bom_service():
    """Dependency injection para FastAPI."""
    return BomService()
