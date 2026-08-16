# Archivo: core/materials/service.py
"""
Service Layer para Materiales compartido.
Logica de negocio, conversion de tipos y exportacion Excel.
"""

from uuid import UUID
from typing import List, Optional, Tuple
from decimal import Decimal
import logging
import re
import unicodedata
from core.materials.normalizer import normalizar_descripcion, normalizar_unidad

from .db_service import get_materials_db_service

logger = logging.getLogger("MaterialsService")

# Columnas esperadas en la plantilla de carga masiva (orden de salida).
PLANTILLA_COLUMNAS = [
    "material", "tipo", "acabado", "marca", "adicional", "medida",
    "concepto", "unidad", "categoria", "precio_referencia", "moneda",
    "clave_sat", "notas",
]

# Columnas de la plantilla de actualizacion masiva de precios (orden de salida).
# id, descripcion y unidad van bloqueados; moneda y precio son editables.
PLANTILLA_PRECIOS_COLUMNAS = [
    "id", "descripcion", "unidad", "moneda", "precio_referencia",
]

# Alias de nombres de categoria del Excel de origen -> nombre real en BD.
_CATEGORIA_ALIASES = {
    "INVERSOR FV": "Inversores",
    "MODULO FV": "Panel",
}


def _norm_cat(texto: str) -> str:
    """Normaliza un nombre de categoria para comparacion (UPPER, sin acentos)."""
    if not texto:
        return ""
    t = unicodedata.normalize("NFKD", str(texto).strip().upper())
    return "".join(c for c in t if not unicodedata.combining(c))


class MaterialsService:
    """Logica de negocio del modulo Materiales."""

    def __init__(self):
        self.db = get_materials_db_service()

    async def get_materiales(
        self, conn, filtros: dict, page: int = 1, per_page: int = 50
    ) -> Tuple[List[dict], int]:
        """Obtiene materiales con filtros y paginacion."""
        total = await self.db.get_materiales_filtered(
            conn, filtros, page, per_page, count_only=True
        )
        rows = await self.db.get_materiales_filtered(
            conn, filtros, page, per_page, count_only=False
        )

        materiales = self._decimals_a_float(
            [dict(row) for row in rows], ('cantidad', 'precio_unitario', 'importe')
        )
        return materiales, total

    async def get_material_precios(
        self, conn, material_id: UUID
    ) -> Tuple[Optional[dict], List[dict], List[dict]]:
        """Obtiene material + analisis de precios por proveedor + productos con misma clave SAT."""
        material = await self.db.get_material_by_id(conn, material_id)
        if not material:
            return None, [], []

        precios = self._decimals_a_float(
            await self.db.get_material_precios(conn, material['descripcion_proveedor']),
            ('min_precio', 'max_precio', 'avg_precio'),
        )

        # Productos similares por clave SAT (excluye misma descripcion)
        precios_sat = []
        if material.get('clave_prod_serv'):
            precios_sat = self._decimals_a_float(
                await self.db.get_precios_por_clave_sat(
                    conn, material['clave_prod_serv'],
                    exclude_descripcion=material['descripcion_proveedor']
                ),
                ('min_precio', 'max_precio', 'avg_precio'),
            )

        return material, precios, precios_sat

    async def get_material_by_id(self, conn, material_id: UUID) -> Optional[dict]:
        """Obtiene un material del historial XML por id (con JOINs de proveedor/categoria)."""
        return await self.db.get_material_by_id(conn, material_id)

    async def update_material(
        self, conn, material_id: UUID, updates: dict
    ) -> Optional[dict]:
        """Actualiza clasificacion interna de un material."""
        success = await self.db.update_material(conn, material_id, updates)
        if not success:
            return None
        material = await self.db.get_material_by_id(conn, material_id)
        if material:
            self._decimals_a_float([material], ('cantidad', 'precio_unitario', 'importe'))
        return material

    async def get_estadisticas(self, conn, filtros: dict) -> dict:
        """Obtiene estadisticas de materiales."""
        return await self.db.get_estadisticas(conn, filtros)

    async def get_catalogos(self, conn) -> dict:
        """Obtiene catalogos para dropdowns."""
        return await self.db.get_catalogos(conn)

    async def buscar_materiales_similares(
        self, conn, query: str, threshold: float = 0.3, limit: int = 20
    ) -> List[dict]:
        """Busqueda fuzzy de materiales por descripcion."""
        rows = await self.db.buscar_similar_materiales(
            conn, query, threshold, limit
        )
        return self._decimals_a_float(rows, ('precio_unitario', 'importe', 'similitud'))

    async def buscar_internos_similares(
        self, conn, query: str, threshold: float = 0.3, limit: int = 10,
        excluir_ids: Optional[list] = None,
    ) -> List[dict]:
        """Homologacion: posibles coincidencias en el catalogo interno antes de dar
        de alta un material nuevo (anti-duplicados)."""
        query_norm = normalizar_descripcion(query)
        rows = await self.db.buscar_internos_similares(
            conn, query_norm, threshold, limit, excluir_ids
        )
        return self._decimals_a_float(rows, ('precio_referencia', 'similitud'))

    async def get_internos(
        self, conn, filtros: dict, page: int = 1, per_page: int = 50
    ) -> Tuple[List[dict], int]:
        total = await self.db.get_internos_filtered(conn, filtros, page, per_page, count_only=True)
        rows  = await self.db.get_internos_filtered(conn, filtros, page, per_page, count_only=False)
        internos = self._decimals_a_float([dict(row) for row in rows], ('precio_referencia',))
        return internos, total

    async def get_interno_by_id(self, conn, id: UUID) -> Optional[dict]:
        m = await self.db.get_interno_by_id(conn, id)
        return self._precios_referencia_a_float([m])[0] if m else None

    async def crear_interno(self, conn, data: dict) -> dict:
        return await self.db.crear_interno(conn, data)

    async def actualizar_interno(self, conn, id: UUID, data: dict) -> Optional[dict]:
        ok = await self.db.actualizar_interno(conn, id, data)
        if not ok:
            return None
        m = await self.db.get_interno_by_id(conn, id)
        if m:
            self._decimals_a_float([m], ('precio_referencia',))
        return m

    async def desactivar_interno(self, conn, id: UUID) -> bool:
        return await self.db.desactivar_interno(conn, id)

    async def get_estadisticas_internos(self, conn) -> dict:
        return await self.db.get_estadisticas_internos(conn)

    async def get_cat_unidades(self, conn) -> list:
        return await self.db.get_cat_unidades(conn)

    async def get_vinculos_xml(self, conn, id_interno: UUID) -> list:
        rows = await self.db.get_vinculos_xml(conn, id_interno)
        return self._precio_unitario_a_float(rows)

    async def buscar_xml_para_vincular(self, conn, id_interno: UUID, q: str) -> list:
        rows = await self.db.buscar_xml_para_vincular(conn, id_interno, q)
        return self._precio_unitario_a_float(rows)

    async def crear_vinculo_xml(self, conn, id_interno: UUID, id_xml: UUID) -> None:
        await self.db.crear_vinculo_xml(conn, id_interno, id_xml)

    # ====================================================================
    # MATCHER AUTOMATICO CATALOGO INTERNO <-> XML (doc 39, punto 6.2)
    # ====================================================================

    async def match_conceptos_a_internos(self, conn, conceptos: list, id_proveedor: UUID) -> dict:
        """Corre el matcher CLAVE_SAT->MEMORIA->TEXTO para los conceptos de una
        factura contra el catalogo interno activo. Ver core/materials/matcher.py
        para el algoritmo puro; aqui solo se resuelven las dependencias de BD
        (memoria del proveedor + snapshot del catalogo)."""
        from core.materials.matcher import match_conceptos_a_internos as _match

        claves = sorted({
            (c.get('clave_prod_serv') or '').strip()
            for c in conceptos if (c.get('clave_prod_serv') or '').strip()
        })
        memoria_map = (
            await self.db.get_memoria_match_interno(conn, id_proveedor, claves)
            if claves else {}
        )
        catalogo = await self.db.get_catalogo_interno_para_matching(conn)
        return _match(conceptos, catalogo, memoria_map)

    async def aplicar_matches_interno_alta(
        self, conn, uuid_factura: str, alta_map: dict
    ) -> None:
        """Aplica (auto-confirma) los matches ALTA (CLAVE_SAT/MEMORIA) directamente
        en tb_materiales_interno_xml con origen AUTO_* -- nunca sobreescribe un
        vinculo HUMANO (ver vincular_interno_a_xml). Incluye el backfill organico
        de clave_prod_serv del catalogo interno (doc 39, decision D+B).

        Debe llamarse DESPUES de guardar_conceptos_historial: necesita los id
        reales de tb_materiales_historial generados en ese INSERT (executemany
        no soporta RETURNING en asyncpg)."""
        if not alta_map:
            return
        historial_rows = await self.db.get_historial_ids_por_factura(conn, uuid_factura)
        id_por_linea = {r['numero_linea_cfdi']: r['id'] for r in historial_rows}
        for idx, match in alta_map.items():
            historial_id = id_por_linea.get(idx + 1)
            if not historial_id:
                continue
            vinculo_aplicado = await self.db.vincular_interno_a_xml(
                conn, historial_id, match['id_material_interno'],
                origen=f"AUTO_{match['origen']}", confianza=match['confianza'],
            )
            if vinculo_aplicado:
                await self.db.backfill_clave_sat_interno(
                    conn, match['id_material_interno'], match.get('clave_prod_serv')
                )

    async def get_conceptos_para_conciliacion_interno(self, conn, limite: int = 100) -> list:
        rows = await self.db.get_conceptos_para_conciliacion_interno(conn, limite)
        return self._decimals_a_float(rows, ('precio_unitario', 'importe'))

    async def confirmar_match_interno(
        self, conn, historial_id: UUID, id_material_interno: Optional[UUID],
        lock_version_esperado: int,
    ) -> dict:
        """Confirma (tal cual o editada) o rechaza una sugerencia del matcher
        automatico. Lanza ValueError si el concepto cambio de version (CAS) --
        mismo criterio que BomService.confirmar_match_concepto para el matcher
        factura<->BOM."""
        result = await self.db.confirmar_match_interno(
            conn, historial_id, id_material_interno, lock_version_esperado
        )
        if not result:
            raise ValueError(
                "El concepto cambió (alguien más lo actualizó); recarga la conciliación."
            )
        return result

    async def sugerir_internos_para_vincular(self, conn, id_xml: UUID, descripcion_material: str) -> list:
        """Sugerencias por similitud difusa cuando aun no hay texto de busqueda manual."""
        rows = await self.db.sugerir_internos_por_similitud(conn, id_xml, descripcion_material)
        return self._precios_referencia_a_float(rows)

    async def sugerir_xml_para_vincular(self, conn, id_interno: UUID, descripcion_interno: str) -> list:
        """Sugerencias por similitud difusa cuando aun no hay texto de busqueda manual."""
        rows = await self.db.sugerir_xml_por_similitud(conn, id_interno, descripcion_interno)
        return self._precio_unitario_a_float(rows)

    async def resolver_internos_para_vincular(
        self, conn, material_id: UUID, q: str, incluir_ancla: bool = True
    ) -> Tuple[Optional[dict], list]:
        """Busqueda textual (3+ caracteres) o sugerencias por similitud difusa (texto
        corto) para vincular un material XML a un item del catalogo interno. Devuelve
        (material, resultados). `incluir_ancla=False` evita re-consultar el material
        cuando el llamador no lo necesita mostrar (ej. respuesta parcial de cada tecleo
        con busqueda textual ya activa)."""
        if len(q) >= 3:
            resultados = await self.buscar_internos_para_vincular(conn, material_id, q)
            material = await self.get_material_by_id(conn, material_id) if incluir_ancla else None
            return material, resultados

        if incluir_ancla:
            material = await self.get_material_by_id(conn, material_id)
            descripcion = material["descripcion_proveedor"] if material else None
        else:
            material = None
            descripcion = await self.db.get_material_descripcion(conn, material_id)

        resultados = (
            await self.sugerir_internos_para_vincular(conn, material_id, descripcion)
            if descripcion else []
        )
        return material, resultados

    async def resolver_xml_para_vincular(
        self, conn, interno_id: UUID, q: str, incluir_ancla: bool = True
    ) -> Tuple[Optional[dict], list]:
        """Busqueda textual (3+ caracteres) o sugerencias por similitud difusa (texto
        corto) para vincular un item del catalogo interno a un registro XML. Devuelve
        (interno, resultados). `incluir_ancla=False` evita re-consultar el item cuando
        el llamador no lo necesita mostrar (ej. respuesta parcial de cada tecleo con
        busqueda textual ya activa)."""
        if len(q) >= 3:
            resultados = await self.buscar_xml_para_vincular(conn, interno_id, q)
            interno = await self.get_interno_by_id(conn, interno_id) if incluir_ancla else None
            return interno, resultados

        if incluir_ancla:
            interno = await self.get_interno_by_id(conn, interno_id)
            descripcion = interno["descripcion_canonica"] if interno else None
        else:
            interno = None
            descripcion = await self.db.get_interno_descripcion(conn, interno_id)

        resultados = (
            await self.sugerir_xml_para_vincular(conn, interno_id, descripcion)
            if descripcion else []
        )
        return interno, resultados

    async def eliminar_vinculo_xml(self, conn, id_interno: UUID, id_xml: UUID) -> None:
        await self.db.eliminar_vinculo_xml(conn, id_interno, id_xml)

    @staticmethod
    def _decimals_a_float(rows: list, keys: tuple) -> list:
        """Convierte a float, in-place, las columnas Decimal indicadas de cada fila.
        `isinstance` directo (no `r.get(key) and ...`) para no saltarse Decimal('0')."""
        for r in rows:
            for key in keys:
                if isinstance(r.get(key), Decimal):
                    r[key] = float(r[key])
        return rows

    @classmethod
    def _precios_referencia_a_float(cls, rows: list) -> list:
        """Convierte precio_referencia de Decimal a float en una lista de filas."""
        return cls._decimals_a_float(rows, ('precio_referencia',))

    @classmethod
    def _precio_unitario_a_float(cls, rows: list) -> list:
        """Convierte precio_unitario de Decimal a float en una lista de filas."""
        return cls._decimals_a_float(rows, ('precio_unitario',))

    async def get_vinculos_interno_por_xml(self, conn, id_xml: UUID) -> list:
        rows = await self.db.get_vinculos_interno_por_xml(conn, id_xml)
        return self._precios_referencia_a_float(rows)

    async def buscar_internos_para_vincular(self, conn, id_xml: UUID, q: str) -> list:
        rows = await self.db.buscar_internos_para_vincular(conn, id_xml, q)
        return self._precios_referencia_a_float(rows)

    async def vincular_interno_a_xml(self, conn, id_xml: UUID, id_interno: UUID) -> None:
        await self.db.vincular_interno_a_xml(conn, id_xml, id_interno)

    async def eliminar_vinculo_interno(self, conn, id_xml: UUID, id_interno: UUID) -> None:
        """Quita el vinculo entre un registro XML y un item del catalogo interno.
        Misma tabla que eliminar_vinculo_xml; se expone con este nombre para que la
        direccion de los argumentos coincida con el flujo 'vincular-interno'."""
        await self.db.eliminar_vinculo_xml(conn, id_interno, id_xml)

    # ====================================================================
    # CARGA MASIVA DE CATALOGO INTERNO (template + validar + cargar)
    # ====================================================================

    # Sinonimos de encabezado aceptados -> campo canonico de la plantilla.
    _HEADER_ALIASES = {
        'material': 'material', 'tipo': 'tipo', 'acabado': 'acabado',
        'marca': 'marca', 'adicional': 'adicional', 'medida': 'medida',
        'concepto': 'concepto', 'descripcion': 'concepto', 'descripcion_canonica': 'concepto',
        'unidad': 'unidad', 'unidad_medida': 'unidad',
        'categoria': 'categoria',
        'precio_referencia': 'precio_referencia', 'precio referencia': 'precio_referencia',
        'precio': 'precio_referencia', 'p. u. mxn': 'precio_referencia', 'p.u. mxn': 'precio_referencia',
        'moneda': 'moneda',
        'clave_sat': 'clave_sat', 'clave sat': 'clave_sat', 'clave_prod_serv': 'clave_sat',
        'notas': 'notas',
    }

    async def _build_resolucion(self, conn) -> Tuple[dict, dict]:
        """Construye los mapas de resolucion: unidades (alias->id) y categorias (norm->id)."""
        unidad_map = await self.db.get_unidad_alias_map(conn)
        categorias = (await self.db.get_catalogos(conn)).get('categorias', [])
        cat_map = {_norm_cat(c['nombre']): c['id'] for c in categorias}
        for alias, nombre_db in _CATEGORIA_ALIASES.items():
            cid = cat_map.get(_norm_cat(nombre_db))
            if cid:
                cat_map[alias] = cid
        return unidad_map, cat_map

    def _parse_y_validar(self, archivo_bytes: bytes, unidad_map: dict,
                         cat_map: dict, norms_existentes: set) -> dict:
        """Parsea el Excel y valida cada fila SIN escribir en BD.
        Devuelve filas listas para insertar + detalles por fila + resumen de conteos."""
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        from io import BytesIO
        from zipfile import BadZipFile

        try:
            wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as e:
            raise ValueError("Archivo Excel invalido o corrupto") from e
        ws = wb.active

        validas: List[dict] = []
        detalles: List[dict] = []
        norms_sesion: set = set()
        a_cargar = advertencias = errores = duplicados = 0

        headers: Optional[List[str]] = None
        fila_num = 0
        for row in ws.iter_rows(values_only=True):
            fila_num += 1
            # Localizar la fila de encabezados (la que contiene 'concepto' o 'material')
            if headers is None:
                celdas = [self._HEADER_ALIASES.get(str(c).strip().lower(), '') if c else '' for c in row]
                if 'concepto' in celdas or 'material' in celdas:
                    headers = celdas
                continue
            if not any(row):
                continue

            fila = {}
            for h, val in zip(headers, row):
                if h:
                    fila[h] = ('' if val is None else str(val).strip())

            concepto = fila.get('concepto', '').strip()
            partes = [fila.get(k, '').strip() for k in
                      ('material', 'tipo', 'acabado', 'marca', 'adicional', 'medida')]
            if not concepto:
                concepto = ' '.join(p for p in partes if p and p.upper() != 'NA').strip()
            concepto = re.sub(r'\s{2,}', ' ', concepto).strip()
            if not concepto:
                continue  # fila totalmente vacia: se ignora en silencio

            errs: List[str] = []
            warns: List[str] = []

            # Unidad
            unidad_txt = fila.get('unidad', '').strip()
            id_unidad = None
            if unidad_txt:
                id_unidad = unidad_map.get(normalizar_unidad(unidad_txt))
                if id_unidad is None:
                    errs.append(f"unidad '{unidad_txt}' no reconocida")

            # Categoria (opcional)
            cat_txt = fila.get('categoria', '').strip()
            id_categoria = None
            if cat_txt:
                id_categoria = cat_map.get(_norm_cat(cat_txt))
                if id_categoria is None:
                    warns.append(f"categoria '{cat_txt}' no existe (se carga sin categoria)")

            # Precio
            precio = None
            precio_raw = fila.get('precio_referencia', '')
            if precio_raw not in ('', None):
                try:
                    precio = float(str(precio_raw).replace(',', ''))
                    if precio < 0:
                        errs.append("precio negativo")
                except (ValueError, TypeError):
                    errs.append(f"precio '{precio_raw}' no es numerico")

            # Moneda
            moneda = (fila.get('moneda', '') or '').strip().upper() or 'MXN'
            if moneda not in ('MXN', 'USD'):
                warns.append(f"moneda '{moneda}' no valida (se usa MXN)")
                moneda = 'MXN'

            norm = normalizar_descripcion(concepto)

            if errs:
                errores += 1
                detalles.append({'fila': fila_num, 'concepto': concepto,
                                 'estado': 'error', 'mensaje': '; '.join(errs)})
                continue
            if norm in norms_existentes or norm in norms_sesion:
                duplicados += 1
                detalles.append({'fila': fila_num, 'concepto': concepto,
                                 'estado': 'duplicado', 'mensaje': 'ya existe en el catalogo'})
                continue

            norms_sesion.add(norm)
            registro = {
                'descripcion_canonica': concepto,
                'descripcion_norm': norm,
                'id_unidad_medida': id_unidad,
                'id_categoria': id_categoria,
                'clave_prod_serv': fila.get('clave_sat', '').strip() or None,
                'precio_referencia': precio,
                'notas': fila.get('notas', '').strip() or None,
                'material': partes[0] or None, 'tipo': partes[1] or None,
                'acabado': partes[2] or None, 'marca': partes[3] or None,
                'adicional': partes[4] or None, 'medida': partes[5] or None,
                'moneda': moneda,
            }
            validas.append(registro)
            a_cargar += 1
            if warns:
                advertencias += 1
                detalles.append({'fila': fila_num, 'concepto': concepto,
                                 'estado': 'advertencia', 'mensaje': '; '.join(warns)})

        if headers is None:
            raise ValueError("No se encontro la fila de encabezados (se requiere 'concepto' o 'material')")

        resumen = {
            'a_cargar': a_cargar, 'advertencias': advertencias,
            'errores': errores, 'duplicados': duplicados,
        }
        return {'validas': validas, 'detalles': detalles, 'resumen': resumen}

    async def validar_internos_excel(self, conn, archivo_bytes: bytes) -> dict:
        """Fase 1: parsea y valida sin escribir. Devuelve preview."""
        unidad_map, cat_map = await self._build_resolucion(conn)
        norms_existentes = await self.db.get_norms_existentes(conn)
        parsed = self._parse_y_validar(archivo_bytes, unidad_map, cat_map, norms_existentes)
        return {
            'fase': 'validacion',
            'resumen': parsed['resumen'],
            'detalles': parsed['detalles'],
        }

    async def cargar_internos_excel(self, conn, archivo_bytes: bytes, creado_por=None) -> dict:
        """Fase 2: re-valida e inserta solo las filas validas en una transaccion.
        Cada fila queda estampada con el usuario que ejecuta la carga."""
        unidad_map, cat_map = await self._build_resolucion(conn)
        norms_existentes = await self.db.get_norms_existentes(conn)
        parsed = self._parse_y_validar(archivo_bytes, unidad_map, cat_map, norms_existentes)
        for fila in parsed['validas']:
            fila['creado_por'] = creado_por
            fila['actualizado_por'] = creado_por
        async with conn.transaction():
            creados = await self.db.crear_internos_bulk(conn, parsed['validas'])
        return {
            'fase': 'carga',
            'creados': creados,
            'resumen': parsed['resumen'],
            'detalles': parsed['detalles'],
        }

    # ====================================================================
    # ACTUALIZACION MASIVA DE PRECIOS (solo precio_referencia + moneda)
    # ====================================================================

    def _parse_actualizacion_precios(self, archivo_bytes: bytes, actuales: dict) -> dict:
        """Parsea el Excel de actualizacion de precios y valida cada fila sin escribir."""
        from io import BytesIO
        from uuid import UUID
        from zipfile import BadZipFile

        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        try:
            wb = load_workbook(BytesIO(archivo_bytes), read_only=True, data_only=True)
        except (BadZipFile, InvalidFileException, OSError) as e:
            raise ValueError("Archivo Excel invalido o corrupto") from e
        ws = wb.active

        validas: List[dict] = []
        detalles: List[dict] = []
        ids_sesion: set = set()
        a_actualizar = errores = sin_cambios = 0

        headers: Optional[List[str]] = None
        fila_num = 0
        for row in ws.iter_rows(values_only=True):
            fila_num += 1
            if headers is None:
                celdas = [str(c).strip().lower() if c else '' for c in row]
                if 'id' in celdas and 'precio_referencia' in celdas:
                    headers = celdas
                continue
            if not any(row):
                continue

            fila = {}
            for h, val in zip(headers, row):
                if h:
                    fila[h] = '' if val is None else str(val).strip()

            id_txt = fila.get('id', '').strip()
            if not id_txt:
                continue

            try:
                mid = UUID(id_txt)
            except (ValueError, AttributeError):
                errores += 1
                detalles.append({
                    'fila': fila_num,
                    'id': id_txt,
                    'estado': 'error',
                    'mensaje': "id no es un UUID valido",
                })
                continue
            if mid not in actuales:
                errores += 1
                detalles.append({
                    'fila': fila_num,
                    'id': id_txt,
                    'estado': 'error',
                    'mensaje': "id no existe o esta desactivado",
                })
                continue
            if mid in ids_sesion:
                errores += 1
                detalles.append({
                    'fila': fila_num,
                    'id': id_txt,
                    'estado': 'error',
                    'mensaje': "id repetido en el archivo",
                })
                continue

            errs: List[str] = []

            precio = None
            precio_raw = fila.get('precio_referencia', '')
            if precio_raw not in ('', None):
                try:
                    precio = float(str(precio_raw).replace(',', ''))
                    if precio < 0:
                        errs.append("precio negativo")
                except (ValueError, TypeError):
                    errs.append(f"precio '{precio_raw}' no es numerico")

            moneda = (fila.get('moneda', '') or '').strip().upper() or 'MXN'
            if moneda not in ('MXN', 'USD'):
                errs.append(f"moneda '{moneda}' no valida (use MXN o USD)")

            if errs:
                errores += 1
                detalles.append({
                    'fila': fila_num,
                    'id': id_txt,
                    'estado': 'error',
                    'mensaje': '; '.join(errs),
                })
                continue

            ids_sesion.add(mid)
            actual = actuales[mid]
            precio_actual = actual['precio']
            moneda_actual = (actual['moneda'] or 'MXN').upper()
            mismo_precio = (
                (precio_actual is None and precio is None)
                or (
                    precio_actual is not None
                    and precio is not None
                    and abs(precio_actual - precio) < 0.005
                )
            )
            if mismo_precio and moneda_actual == moneda:
                sin_cambios += 1
                continue

            validas.append({'id': mid, 'precio_referencia': precio, 'moneda': moneda})
            a_actualizar += 1

        if headers is None:
            raise ValueError(
                "No se encontro la fila de encabezados (se requieren 'id' y 'precio_referencia')"
            )

        resumen = {'a_actualizar': a_actualizar, 'errores': errores, 'sin_cambios': sin_cambios}
        return {'validas': validas, 'detalles': detalles, 'resumen': resumen}

    async def validar_actualizacion_precios(self, conn, archivo_bytes: bytes) -> dict:
        """Fase 1: parsea y valida sin escribir. Devuelve preview."""
        actuales = await self.db.get_precios_actuales(conn)
        parsed = self._parse_actualizacion_precios(archivo_bytes, actuales)
        return {
            'fase': 'validacion',
            'modo': 'precios',
            'resumen': parsed['resumen'],
            'detalles': parsed['detalles'],
        }

    async def actualizar_precios_excel(self, conn, archivo_bytes: bytes, actualizado_por=None) -> dict:
        """Fase 2: re-valida y actualiza solo las filas con cambio real."""
        actuales = await self.db.get_precios_actuales(conn)
        parsed = self._parse_actualizacion_precios(archivo_bytes, actuales)
        for fila in parsed['validas']:
            fila['actualizado_por'] = actualizado_por
        async with conn.transaction():
            actualizados = await self.db.actualizar_precios_bulk(conn, parsed['validas'])
        return {
            'fase': 'carga',
            'modo': 'precios',
            'actualizados': actualizados,
            'resumen': parsed['resumen'],
            'detalles': parsed['detalles'],
        }

    async def generar_plantilla_internos(self, conn) -> bytes:
        """Genera la plantilla .xlsx de carga masiva con hoja de catalogos y validaciones."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        unidades = await self.db.get_cat_unidades(conn)
        categorias = (await self.db.get_catalogos(conn)).get('categorias', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Materiales"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, name in enumerate(PLANTILLA_COLUMNAS, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col)].width = 16

        # Fila de ejemplo
        ejemplo = ["ABRAZADERA", "CLIP", "PARED DELGADA", "", "", '(1/2")',
                   'ABRAZADERA CLIP PARED DELGADA (1/2")', "pza", "Accesorios electricos",
                   "5.24", "MXN", "", "Ejemplo: borrar esta fila"]
        for col, val in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col, value=val)

        # Hoja de catalogos (referencia + origen de las listas desplegables)
        cat_ws = wb.create_sheet("Catalogos")
        cat_ws.cell(row=1, column=1, value="unidad (codigo)").font = header_font
        cat_ws.cell(row=1, column=2, value="unidad (nombre)").font = header_font
        cat_ws.cell(row=1, column=4, value="categoria").font = header_font
        for i, u in enumerate(unidades, 2):
            cat_ws.cell(row=i, column=1, value=u['codigo'])
            cat_ws.cell(row=i, column=2, value=u['nombre'])
        for i, c in enumerate(categorias, 2):
            cat_ws.cell(row=i, column=4, value=c['nombre'])
        cat_ws.column_dimensions['A'].width = 14
        cat_ws.column_dimensions['B'].width = 24
        cat_ws.column_dimensions['D'].width = 24

        # Validaciones de datos (dropdowns) sobre las primeras 1000 filas
        n_uni = len(unidades) + 1
        n_cat = len(categorias) + 1
        dv_unidad = DataValidation(type="list", formula1=f"Catalogos!$A$2:$A${n_uni}", allow_blank=True)
        dv_cat = DataValidation(type="list", formula1=f"Catalogos!$D$2:$D${n_cat}", allow_blank=True)
        dv_moneda = DataValidation(type="list", formula1='"MXN,USD"', allow_blank=True)
        ws.add_data_validation(dv_unidad)
        ws.add_data_validation(dv_cat)
        ws.add_data_validation(dv_moneda)
        col_unidad = get_column_letter(PLANTILLA_COLUMNAS.index('unidad') + 1)
        col_cat = get_column_letter(PLANTILLA_COLUMNAS.index('categoria') + 1)
        col_mon = get_column_letter(PLANTILLA_COLUMNAS.index('moneda') + 1)
        dv_unidad.add(f"{col_unidad}2:{col_unidad}1000")
        dv_cat.add(f"{col_cat}2:{col_cat}1000")
        dv_moneda.add(f"{col_mon}2:{col_mon}1000")

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def generar_plantilla_precios(self, conn) -> bytes:
        """Genera la plantilla .xlsx para actualizar precio_referencia y moneda."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Protection
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from io import BytesIO

        internos, _ = await self.get_internos(conn, {}, page=1, per_page=100000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Precios"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        locked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        anchos = {
            "id": 38,
            "descripcion": 50,
            "unidad": 12,
            "moneda": 10,
            "precio_referencia": 16,
        }

        for col, name in enumerate(PLANTILLA_PRECIOS_COLUMNAS, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col)].width = anchos[name]

        dv_moneda = DataValidation(type="list", formula1='"MXN,USD"', allow_blank=True)
        ws.add_data_validation(dv_moneda)
        max_row = max(len(internos) + 1, 1000)
        dv_moneda.add(f"D2:D{max_row}")

        for i, material in enumerate(internos, 2):
            c_id = ws.cell(row=i, column=1, value=str(material['id']))
            c_desc = ws.cell(row=i, column=2, value=material.get('descripcion_canonica') or '')
            c_uni = ws.cell(row=i, column=3, value=material.get('unidad_codigo') or '')
            c_mon = ws.cell(row=i, column=4, value=material.get('moneda') or 'MXN')
            precio = material.get('precio_referencia')
            c_pre = ws.cell(row=i, column=5, value=float(precio) if precio is not None else None)
            c_pre.number_format = '#,##0.00'

            for cell in (c_id, c_desc, c_uni):
                cell.protection = Protection(locked=True)
                cell.fill = locked_fill
            c_mon.protection = Protection(locked=False)
            c_pre.protection = Protection(locked=False)

        ws.protection.sheet = True
        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    async def export_to_excel(self, conn, filtros: dict) -> bytes:
        """Genera archivo Excel con materiales filtrados."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        materiales, _ = await self.get_materiales(
            conn, filtros=filtros, per_page=100000
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Materiales"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "Proveedor",
            "RFC",
            "Descripcion Proveedor",
            "Descripcion Interna",
            "Categoria",
            "Cantidad",
            "P. Unitario",
            "Importe",
            "Unidad",
            "Clave SAT",
            "Fecha Factura",
            "Origen",
            "Proyecto",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, m in enumerate(materiales, 2):
            row_data = [
                m.get('proveedor_nombre', ''),
                m.get('proveedor_rfc', ''),
                m.get('descripcion_proveedor', ''),
                m.get('descripcion_interna', ''),
                m.get('categoria_nombre', ''),
                m.get('cantidad', 0),
                m.get('precio_unitario', 0),
                m.get('importe', 0),
                m.get('unidad', ''),
                m.get('clave_prod_serv', ''),
                m['fecha_factura'].strftime("%d/%m/%Y") if m.get('fecha_factura') else '',
                m.get('origen', ''),
                m.get('proyecto_nombre', ''),
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

                # Formato numerico para cantidad, precio e importe
                if col_num in (6, 7, 8):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Anchos de columna
        column_widths = [30, 15, 40, 35, 20, 12, 14, 14, 10, 12, 14, 10, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


def get_materials_service():
    """Dependency injection para FastAPI."""
    return MaterialsService()
