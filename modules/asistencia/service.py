from __future__ import annotations

import asyncio
import logging
import re
from datetime import date, datetime, time, timedelta
from typing import Any
from uuid import UUID

import asyncpg
import httpx
from fastapi import UploadFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.config import settings
from core.config_service import ConfigService
from core.database import DB_TIMEOUT_ERRORS, get_db_pool
from core.integrations.sharepoint import SharePointService
from core.microsoft import get_ms_auth
from core.permissions import user_has_module_access
from core.timezone import fmt_time_mx, now_mx, today_mx
from modules.asistencia import db_service as db
from modules.asistencia.biotime_client import BioTimeClient
from modules.asistencia.constants import (
    ASISTENCIA_ESTADOS_CON_MODALIDAD_METADATA,
    BIOTIME_CONFIG_KEYS,
)
from modules.asistencia.logic import (
    AttendanceCheck,
    MX_TZ,
    ScheduleConfig,
    build_labor_window,
    calcular_resumen_dia,
    ensure_mx,
    extender_salida_descanso_medianoche,
    is_break_state,
    is_in_state,
    is_out_state,
)
from modules.shared.utils import format_minutes
from modules.vacaciones import db_service as vacaciones_db

logger = logging.getLogger("asistencia.service")

# Fallas de conexion/timeout post-freeze (command_timeout del pool compartido);
# se agrupan aqui porque se repiten en varios puntos de entrada de BioTime.
# OSError se agrega aparte: aqui tambien cubre fallas de red del cliente httpx hacia BioTime,
# no solo la BD (a diferencia de DB_TIMEOUT_ERRORS, que es BD-only).
BIOTIME_CONNECTION_ERRORS = DB_TIMEOUT_ERRORS + (OSError,)
_BIOTIME_FINISH_ERRORS = (asyncpg.PostgresError,) + BIOTIME_CONNECTION_ERRORS
_BIOTIME_SYNC_ERRORS = (httpx.HTTPError, ValueError, TypeError, RuntimeError) + _BIOTIME_FINISH_ERRORS


class HEAutorizacionError(Exception):
    """El actor no tiene permiso para autorizar HE/compensatorio de un tercero (403)."""


class HEFallbackVacioError(Exception):
    """Un cambio administrativo dejaria a empleados con aprobador HE exclusivo sin fallback (409)."""


_SOLICITUD_MANUAL_ESTADOS = {
    "pendiente": "Pendiente",
    "aprobado": "Aprobado",
    "rechazado": "Rechazado",
}


def validate_aprobacion(minutos_aprobados: int, minutos_extra: int, comentario: str) -> None:
    if not comentario or not comentario.strip():
        raise ValueError("El comentario es obligatorio")
    if minutos_aprobados < 30:
        raise ValueError("El mínimo aprobable es 30 minutos")
    if minutos_aprobados > minutos_extra:
        raise ValueError(f"No puede aprobar más de {minutos_extra} minutos registrados")


_RRHH_EDITOR_EQUIPO_MAX_EMPLEADOS = 10000


async def _ids_todos_empleados_rh_editor(conn) -> list[UUID]:
    """Bypass RH editor/admin: todos los empleados con ficha (consulta y autorizacion HE).

    Usa la query liviana (solo IDs, sin los joins de detalle de
    get_all_empleados_con_datos) porque este bypass corre en cada carga de
    Mi Perfil / detalle de solicitud para RH editor/admin, no solo en listados."""
    return await vacaciones_db.get_all_empleado_ids_activos(
        conn, limit=_RRHH_EDITOR_EQUIPO_MAX_EMPLEADOS
    )


async def get_equipo_ids(conn, user_id: UUID, user_ctx: dict) -> list[UUID]:
    if user_has_module_access("rrhh", user_ctx, "editor"):
        return await _ids_todos_empleados_rh_editor(conn)

    ids_jefe = await vacaciones_db.get_empleados_donde_soy_jefe(conn, user_id)
    ids_aprobador = await vacaciones_db.get_empleados_donde_soy_aprobador(conn, user_id)
    return list({*ids_jefe, *ids_aprobador})


async def get_equipo_ids_para_autorizacion_he(conn, user_id: UUID, user_ctx: dict) -> list[UUID]:
    """Alcance de AUTORIZACION de HE/compensatorio (no de consulta general).

    RH editor conserva bypass total; cualquier otro usuario obtiene solo los empleados
    donde es aprobador exclusivo HE, jefe directo sin override, o aprobador de vacaciones
    sin override (ver db.get_empleados_para_autorizacion_he).
    """
    if user_has_module_access("rrhh", user_ctx, "editor"):
        return await _ids_todos_empleados_rh_editor(conn)
    return await db.get_empleados_para_autorizacion_he(conn, user_id)


async def get_equipo_visible_he(
    conn, user_id: UUID, user_ctx: dict, base_ids: list[UUID] | set[UUID] = ()
) -> tuple[list[UUID], set[UUID]]:
    """Union de `base_ids` (equipo de consulta general, ya resuelto por el llamador) con los
    IDs autorizables de HE/compensatorio. Devuelve tambien el set autorizable para marcar
    `puede_autorizar_he` por fila en listados de aprobaciones."""
    autorizable_he_set = set(await get_equipo_ids_para_autorizacion_he(conn, user_id, user_ctx))
    equipo_visible_he = list({*base_ids, *autorizable_he_set})
    return equipo_visible_he, autorizable_he_set


def marcar_puede_autorizar_he(rows: list[dict], autorizable_he_set: set[UUID]) -> None:
    """Anota `puede_autorizar_he` in-place en cada row de un listado HE/compensatorio,
    a partir del set devuelto por get_equipo_visible_he."""
    for row in rows:
        row["puede_autorizar_he"] = row["usuario_id"] in autorizable_he_set


async def anexar_modalidad_metadata_asistencia(conn, rows: list[dict]) -> list[dict]:
    """Anexa metadata visual de modalidades sin alterar la asistencia persistida."""
    if not rows:
        return rows

    usuario_ids = list({row["usuario_id"] for row in rows})
    fechas = [row["fecha_laboral"] for row in rows]
    modalidades = await db.get_modalidades_metadata_en_rango(
        conn,
        usuario_ids=usuario_ids,
        fecha_inicio=min(fechas),
        fecha_fin=max(fechas),
    )

    metadata_por_fecha: dict[tuple[UUID, date], dict] = {}
    for modalidad in modalidades:
        key = (modalidad["usuario_id"], modalidad["fecha_laboral"])
        if key in metadata_por_fecha:
            logger.warning(
                "[ASISTENCIA_ANOMALIA] Modalidades de asistencia duplicadas para usuario=%s fecha=%s; se conserva la primera",
                key[0],
                key[1],
            )
            continue
        metadata_por_fecha[key] = {
            "slug": modalidad["tipo_slug"],
            "nombre": modalidad["tipo_nombre"],
            "abreviatura": modalidad["tipo_abreviatura"],
            "solicitud_id": modalidad["solicitud_id"],
        }

    for row in rows:
        tiene_checada = bool(row.get("primera_entrada") or row.get("ultima_salida"))
        mostrar = (
            tiene_checada
            and (row.get("minutos_programados") or 0) > 0
            and row.get("estado") in ASISTENCIA_ESTADOS_CON_MODALIDAD_METADATA
        )
        key = (row["usuario_id"], row["fecha_laboral"])
        row["modalidad_metadata"] = metadata_por_fecha.get(key) if mostrar else None
    return rows


async def _lock_y_exigir_autorizacion_he(conn, *, empleado_id: UUID, aprobador_id: UUID) -> None:
    """Lock actor+empleado y revalida en BD que `aprobador_id` puede autorizar HE/compensatorio
    de terceros para `empleado_id`. Usar dentro de la transaccion, antes de mutar el registro.

    NOTA anti-enumeracion (aceptado 2026-07-14): todo llamador hace `if not row: raise
    ValueError("... no encontrado")` ANTES de invocar esta funcion, así que un actor sin
    autorizacion sobre un registro real recibe HEAutorizacionError/403 (distinto del 400 que
    recibe ante un id inexistente). Antes de esta feature, el chequeo equivalente
    (`_validar_permiso_compensatorio`, ya eliminado) devolvia el mismo "no encontrada" en ambos
    casos, ocultando deliberadamente si el id existia. Se decidio conservar el comportamiento
    actual (distinguible) porque esto es una herramienta interna autenticada, no una API publica,
    y lo que se filtra es solo "existe un registro con este UUID", no su contenido. Si se
    necesita restaurar la ambiguedad, hacer que esta funcion levante el mismo ValueError/mensaje
    "no encontrado" en vez de HEAutorizacionError."""
    await db.lock_he_actor(conn, aprobador_id)
    await db.lock_he_usuario(conn, empleado_id)
    if not await db.puede_autorizar_he(conn, empleado_id, aprobador_id):
        raise HEAutorizacionError("No tienes permiso para autorizar este registro")


async def aprobar_horas_extra_svc(
    conn,
    *,
    asistencia_id: UUID,
    aprobador_id: UUID,
    minutos_aprobados: int,
    comentario: str,
) -> dict:
    async with conn.transaction():
        row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
        if not row:
            raise ValueError("Registro no encontrado")
        await _lock_y_exigir_autorizacion_he(
            conn, empleado_id=row["usuario_id"], aprobador_id=aprobador_id
        )
        if row["horas_extra_estado"] not in ("pendiente", "solicitado"):
            raise ValueError("Este registro ya fue procesado")
        if row.get("minutos_he_compensatorio"):
            raise ValueError("Este dia ya tiene horas extra tomadas")
        validate_aprobacion(minutos_aprobados, row["minutos_extra"], comentario)

        acreditados = await db.aprobar_horas_extra(
            conn,
            asistencia_id=asistencia_id,
            aprobador_id=aprobador_id,
            minutos_aprobados=minutos_aprobados,
            comentario=comentario,
        )
        if acreditados != 1:
            raise ValueError("Este registro ya fue procesado o no aplica a bolsa")
    return {
        "empleado_nombre": row["empleado_nombre"],
        "empleado_email": row.get("empleado_email"),
        "fecha_laboral": row["fecha_laboral"],
        "minutos_aprobados": minutos_aprobados,
        "comentario": comentario,
    }


def parse_biotime_check_time(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return ensure_mx(value)
    raw = str(value).strip()
    if not raw:
        return None
    normalized = raw.replace("T", " ")
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        try:
            parsed = datetime.strptime(normalized[:19], "%Y-%m-%d %H:%M:%S")
        except ValueError:
            return None
    return ensure_mx(parsed)


async def sync_biotime_once(conn, *, force: bool = False) -> dict:
    activo = await ConfigService.get_global_config(
        conn, BIOTIME_CONFIG_KEYS["sync_activo"], "false", bool
    )
    if not activo and not force:
        return {"status": "disabled", "message": "Sincronizacion BioTime desactivada"}

    cfg = await _load_biotime_client_config(conn)
    base_url = cfg["base_url"]
    username = cfg["username"]
    password = cfg["password"]
    page_size = cfg["page_size"]
    timeout_seconds = cfg["timeout_seconds"]
    lookback_hours = await ConfigService.get_global_config(
        conn, BIOTIME_CONFIG_KEYS["lookback_hours"], 48, int
    )
    recalc_days = await ConfigService.get_global_config(
        conn, BIOTIME_CONFIG_KEYS["recalc_days"], 7, int
    )

    last_id = await db.get_last_transaction_id(conn)
    window_end = now_mx()
    window_start = window_end - timedelta(hours=max(1, min(lookback_hours, 31 * 24)))
    run_id = await db.create_sync_run(
        conn,
        from_transaction_id=last_id,
        window_start=window_start,
        window_end=window_end,
    )

    try:
        async with BioTimeClient(base_url, username, password, timeout_seconds=timeout_seconds) as client:
            items = await client.fetch_transactions(
                starttime=window_start,
                endtime=window_end,
                page_size=page_size,
            )
            employee_sync = await _sync_employee_mappings_from_biotime(conn, client)
        normalized = await _normalize_transactions(conn, items)
        inserted = await db.insert_checks_batch(conn, normalized)
        reassigned = await db.assign_unmapped_checks_from_mappings(conn)
        affected_targets = _targets_from_inserted(inserted + reassigned)

        if affected_targets:
            await recalcular_asistencia(conn, affected_targets)
        if recalc_days > 0:
            await recalcular_asistencia_reciente(conn, days=recalc_days)

        # Se calcula antes de cerrar el run en 'success': si esta consulta informativa
        # falla, no debe revertir a 'error' un run cuyo trabajo real ya se completo.
        unmapped = await db.get_unmapped_biotime_checks_summary(
            conn,
            fecha_inicio=ensure_mx(window_start).date(),
            fecha_fin=ensure_mx(window_end).date(),
        )
        to_id = _max_transaction_id(normalized) or last_id
        await db.finish_sync_run(
            conn,
            run_id=run_id,
            status="success",
            to_transaction_id=to_id,
            records_read=len(items),
            records_inserted=len(inserted),
            records_skipped=max(0, len(normalized) - len(inserted)),
        )
        insert_metrics = _check_insert_metrics(inserted)
        return {
            "status": "success",
            "records_read": len(items),
            "records_inserted": len(inserted),
            "records_skipped": max(0, len(normalized) - len(inserted)),
            **insert_metrics,
            "historical_checks_mapped": len(reassigned),
            "affected_targets": len(affected_targets),
            "unmapped_biotime_codes": [row["biotime_emp_code"] for row in unmapped],
            "unmapped_biotime_count": sum(row["total"] for row in unmapped),
            **employee_sync,
        }
    except _BIOTIME_SYNC_ERRORS as exc:
        try:
            await db.finish_sync_run(
                conn,
                run_id=run_id,
                status="error",
                records_read=0,
                error_message=str(exc)[:1000],
            )
        except _BIOTIME_FINISH_ERRORS as finish_exc:
            logger.error(
                "[BIOTIME_SYNC] No se pudo marcar run %s como error: %s",
                run_id, finish_exc,
            )
        raise


async def probar_conexion_biotime(
    *,
    base_url: str,
    username: str,
    password: str,
    timeout_seconds: int = 30,
) -> dict:
    async with BioTimeClient(base_url, username, password, timeout_seconds=timeout_seconds) as client:
        total = await client.ping()
    return {"status": "success", "records_read": total}


_MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _chunk_label(fecha_inicio: date, fecha_fin: date) -> str:
    if fecha_inicio.year == fecha_fin.year and fecha_inicio.month == fecha_fin.month:
        return f"{_MESES[fecha_inicio.month]} {fecha_inicio.year}"
    return f"{fecha_inicio.strftime('%d/%m/%Y')} – {fecha_fin.strftime('%d/%m/%Y')}"


async def _load_biotime_client_config(conn) -> dict:
    configs = await ConfigService.get_global_configs_bulk(conn, {
        BIOTIME_CONFIG_KEYS["base_url"]: ("", str),
        BIOTIME_CONFIG_KEYS["username"]: ("", str),
        BIOTIME_CONFIG_KEYS["password"]: ("", str),
        BIOTIME_CONFIG_KEYS["page_size"]: (200, int),
        BIOTIME_CONFIG_KEYS["timeout_seconds"]: (30, int),
    })
    return {
        "base_url": configs[BIOTIME_CONFIG_KEYS["base_url"]],
        "username": configs[BIOTIME_CONFIG_KEYS["username"]],
        "password": configs[BIOTIME_CONFIG_KEYS["password"]],
        "page_size": configs[BIOTIME_CONFIG_KEYS["page_size"]],
        "timeout_seconds": configs[BIOTIME_CONFIG_KEYS["timeout_seconds"]],
    }


async def backfill_biotime_chunk(
    conn,
    fecha_inicio: date,
    fecha_fin: date,
) -> dict:
    if (fecha_fin - fecha_inicio).days > 31:
        raise ValueError("El rango del chunk no puede superar 31 días")

    cfg = await _load_biotime_client_config(conn)
    base_url = cfg["base_url"]
    username = cfg["username"]
    password = cfg["password"]
    page_size = cfg["page_size"]
    timeout_seconds = cfg["timeout_seconds"]

    window_start = datetime.combine(fecha_inicio, time.min, tzinfo=MX_TZ)
    window_end = datetime.combine(fecha_fin + timedelta(days=1), time.min, tzinfo=MX_TZ)
    chunk_label = _chunk_label(fecha_inicio, fecha_fin)

    run_id = await db.create_sync_run(
        conn,
        from_transaction_id=None,
        window_start=window_start,
        window_end=window_end,
    )

    try:
        async with BioTimeClient(base_url, username, password, timeout_seconds=timeout_seconds) as client:
            items = await client.fetch_transactions(
                starttime=window_start,
                endtime=window_end,
                page_size=page_size,
            )
            employee_sync = await _sync_employee_mappings_from_biotime(conn, client)

        normalized = await _normalize_transactions(conn, items)
        inserted = await db.insert_checks_batch(conn, normalized)
        reassigned = await db.assign_unmapped_checks_from_mappings(conn)
        targets = _targets_from_inserted(inserted + reassigned)

        if targets:
            await recalcular_asistencia(conn, targets)

        # Se calcula antes de cerrar el run en 'success': si esta consulta informativa
        # falla, no debe revertir a 'error' un run cuyo trabajo real ya se completo.
        unmapped = await db.get_unmapped_biotime_checks_summary(
            conn,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
        )
        to_id = _max_transaction_id(normalized)
        await db.finish_sync_run(
            conn,
            run_id=run_id,
            status="success",
            to_transaction_id=to_id,
            records_read=len(items),
            records_inserted=len(inserted),
            records_skipped=max(0, len(normalized) - len(inserted)),
        )
        return {
            "chunk_label": chunk_label,
            "records_read": len(items),
            "records_inserted": len(inserted),
            "records_skipped": max(0, len(normalized) - len(inserted)),
            **_check_insert_metrics(inserted),
            "historical_checks_mapped": len(reassigned),
            "unmapped_biotime_codes": [row["biotime_emp_code"] for row in unmapped],
            "unmapped_biotime_count": sum(row["total"] for row in unmapped),
            "targets_recalculated": len(targets),
            **employee_sync,
        }
    except _BIOTIME_SYNC_ERRORS as exc:
        try:
            await db.finish_sync_run(
                conn,
                run_id=run_id,
                status="error",
                records_read=0,
                error_message=str(exc)[:1000],
            )
        except _BIOTIME_FINISH_ERRORS as finish_exc:
            logger.error(
                "[BIOTIME_SYNC] No se pudo marcar run %s como error: %s",
                run_id, finish_exc,
            )
        raise


async def _normalize_transactions(conn, items: list[dict[str, Any]]) -> list[dict]:
    emp_codes = sorted({
        str(_first_value(item, "emp_code", "pin", "badgenumber", "employee_code") or "").strip()
        for item in items
        if _first_value(item, "emp_code", "pin", "badgenumber", "employee_code")
    })
    employee_map = await db.get_employee_map(conn, emp_codes)
    normalized: list[dict] = []

    for item in items:
        emp_code = str(_first_value(item, "emp_code", "pin", "badgenumber", "employee_code") or "").strip()
        raw_time = _first_value(item, "transaction_punch_time", "checktime", "punch_time", "check_time")
        punch_date = item.get("transaction_punch_date")
        if raw_time and punch_date and ":" in str(raw_time) and "-" not in str(raw_time):
            raw_time = f"{punch_date} {raw_time}"
        check_time = parse_biotime_check_time(raw_time)
        if not emp_code or not check_time:
            continue

        mapping = employee_map.get(emp_code, {})
        normalized.append({
            "biotime_transaction_id": _safe_int(item.get("id")),
            "biotime_emp_code": emp_code,
            "usuario_id": str(mapping["usuario_id"]) if mapping.get("usuario_id") else None,
            "check_time": check_time.isoformat(),
            "punch_state": _string_or_none(_first_value(item, "stateno", "punch_state", "state")),
            "verify_type": _string_or_none(_first_value(item, "verify_type", "verify")),
            "terminal_sn": _string_or_none(_first_value(item, "sn", "terminal_sn")),
            "terminal_alias": _string_or_none(_first_value(item, "alias", "terminal_alias")),
            "deptnumber": _string_or_none(item.get("deptnumber")),
            "deptname": _string_or_none(_first_value(item, "deptname", "employee_department")),
            "raw_payload": item,
        })
    return normalized


async def _sync_employee_mappings_from_biotime(
    conn,
    client: BioTimeClient,
) -> dict:
    try:
        employees = await client.fetch_employees()
    except (httpx.HTTPError, ValueError, TypeError) as exc:
        logger.warning("[BIOTIME_SYNC] No se pudieron consultar empleados BioTime: %s", exc)
        return {"employees_read": 0, "employee_mappings": 0}

    normalized = [
        employee
        for employee in (_normalize_biotime_employee(item) for item in employees)
        if employee
    ]
    mapped = await db.upsert_biotime_employee_mappings(conn, normalized)
    return {"employees_read": len(employees), "employee_mappings": len(mapped)}


def _normalize_biotime_employee(item: dict[str, Any]) -> dict | None:
    emp_code = _string_or_none(_first_value(item, "emp_code", "pin", "userpin", "badgenumber", "employee_code"))
    if not emp_code:
        return None
    first = _string_or_none(_first_value(item, "first_name", "name", "username", "ename"))
    last = _string_or_none(_first_value(item, "last_name", "surname"))
    nombre = " ".join(part for part in [first, last] if part) or None
    dept_name = _string_or_none(_first_value(
        item, "employee_department", "department", "dept_name", "deptname", "department_name",
    ))
    dept_code = _string_or_none(_first_value(item, "dept_code", "deptnumber", "department_code", "department_id"))
    return {
        "biotime_emp_id": _safe_int(item.get("id")),
        "biotime_emp_code": emp_code,
        "biotime_pin": _string_or_none(_first_value(item, "pin", "userpin", "emp_code")),
        "email": _normalize_email(_first_value(item, "email", "mail")),
        "nombre": nombre,
        "biotime_deptnumber": dept_code,
        "biotime_deptname": dept_name,
    }


def _normalize_email(value: Any) -> str | None:
    text = _string_or_none(value)
    return text.lower() if text else None


def _first_value(item: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = item.get(key)
        if value is not None and value != "":
            return value
    return None


def _string_or_none(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _safe_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _max_transaction_id(rows: list[dict]) -> int | None:
    values = [row["biotime_transaction_id"] for row in rows if row.get("biotime_transaction_id")]
    return max(values) if values else None


def _check_insert_metrics(rows: list[dict]) -> dict:
    mapped = sum(1 for row in rows if row.get("usuario_id"))
    unmapped = len(rows) - mapped
    return {
        "records_inserted_mapped": mapped,
        "records_inserted_unmapped": unmapped,
    }


def _targets_from_inserted(inserted: list[dict]) -> list[tuple[UUID, date]]:
    targets: set[tuple[UUID, date]] = set()
    for row in inserted:
        usuario_id = row.get("usuario_id")
        check_time = row.get("check_time")
        if not usuario_id or not check_time:
            continue
        local_date = ensure_mx(check_time).date()
        targets.add((usuario_id, local_date))
        targets.add((usuario_id, local_date - timedelta(days=1)))
    return sorted(targets, key=lambda item: (str(item[0]), item[1]))


def _agrupar_fechas_contiguas(fechas: list[date]) -> list[list[date]]:
    """Agrupa fechas ordenadas en corridas de dias consecutivos. `recalcular_asistencia`
    fetchea BD por el rango [min(fechas), max(fechas)] de cada llamada -- sin agrupar,
    un conjunto disperso (ej. feriados del anio, ~8 fechas sueltas) forzaria un fetch de
    casi un anio completo de checadas para recalcular solo esas fechas puntuales."""
    grupos: list[list[date]] = []
    for fecha in fechas:
        if grupos and fecha - grupos[-1][-1] == timedelta(days=1):
            grupos[-1].append(fecha)
        else:
            grupos.append([fecha])
    return grupos


async def recalcular_asistencia_usuarios_activos(conn, fechas) -> list[tuple[UUID, date]]:
    """Recalcula asistencia de todos los usuarios activos para un conjunto de fechas.

    Compartido por el recalculo periodico reciente y por el CRUD de feriados
    (`rrhh/service.py`), que necesita recalcular fechas puntuales (no un rango
    contiguo) al crear, mover o eliminar un feriado."""
    fechas_validas = sorted({fecha for fecha in fechas if fecha})
    if not fechas_validas:
        return []
    usuario_ids = await db.get_active_attendance_users(conn)
    if not usuario_ids:
        return []
    targets: list[tuple[UUID, date]] = []
    for grupo in _agrupar_fechas_contiguas(fechas_validas):
        grupo_targets = [(usuario_id, fecha) for usuario_id in usuario_ids for fecha in grupo]
        await recalcular_asistencia(conn, grupo_targets)
        targets.extend(grupo_targets)
    return targets


async def recalcular_asistencia_reciente(conn, *, days: int) -> int:
    end = today_mx()
    start = end - timedelta(days=max(0, days - 1))
    fechas = {start + timedelta(days=offset) for offset in range((end - start).days + 1)}
    targets = await recalcular_asistencia_usuarios_activos(conn, fechas)
    return len(targets)


async def recalcular_asistencia_reciente_usuario(conn, usuario_id: UUID, *, days: int = 7) -> int:
    end = today_mx()
    start = end - timedelta(days=max(0, days - 1))
    targets = [
        (usuario_id, start + timedelta(days=offset))
        for offset in range((end - start).days + 1)
    ]
    await recalcular_asistencia(conn, targets)
    return len(targets)


async def recalcular_asistencia(conn, targets: list[tuple[UUID, date]]) -> list[dict]:
    targets = _dedupe_targets(targets)
    if not targets:
        return []

    usuario_ids = sorted({target[0] for target in targets}, key=str)
    fecha_inicio = min(target[1] for target in targets)
    fecha_fin = max(target[1] for target in targets)

    min_minutos_he = await ConfigService.get_global_config(
        conn, "ASISTENCIA_HE_MINIMO_MINUTOS", 30, int
    )

    context_rows = await db.get_attendance_contexts(conn, usuario_ids)
    schedule_by_user_day, sucursal_by_user = _build_context_maps(context_rows)
    ausencias = await db.get_ausencias_justificadas(
        conn,
        usuario_ids=usuario_ids,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )
    festivos = await db.get_festivos_range(conn, fecha_inicio, fecha_fin)
    compensatorios = await db.get_he_compensatorio_aprobado_por_fechas(
        conn,
        usuario_ids=usuario_ids,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
    )
    comp_by_user_date = {
        (row["usuario_id"], row["fecha_descanso"]): row
        for row in compensatorios
    }

    broad_start = datetime.combine(fecha_inicio - timedelta(days=1), time.min, tzinfo=MX_TZ)
    broad_end = datetime.combine(fecha_fin + timedelta(days=2), time.min, tzinfo=MX_TZ)
    checks = await db.get_checks_for_users_window(
        conn,
        usuario_ids=usuario_ids,
        start=broad_start,
        end=broad_end,
    )
    checks_by_user = _group_checks_by_user(checks)

    calculated_at = now_mx()
    rows_to_save: list[dict] = []
    # `targets` viene ordenado por (usuario_id, fecha) ascendente (`_dedupe_targets`), asi que
    # una checada tomada prestada del dia siguiente para cerrar un descanso (ver
    # `extender_salida_descanso_medianoche`) ya fue removida de `checks_by_user` antes de
    # procesar ese dia siguiente como su propio target -- evita que la misma checada se
    # cuente dos veces (una extendiendo el descanso, otra como salida huerfana del dia siguiente).
    for usuario_id, fecha_laboral in targets:
        schedule = schedule_by_user_day.get((usuario_id, fecha_laboral.weekday()))
        window = build_labor_window(fecha_laboral, schedule)
        checks_usuario = checks_by_user.get(usuario_id, [])
        checks_dia = [
            check for check in checks_usuario
            if window.start <= ensure_mx(check.check_time) < window.end
        ]
        es_dia_descanso = bool(schedule and not schedule.es_laboral)
        if es_dia_descanso:
            checks_dia, prestada = extender_salida_descanso_medianoche(
                checks_dia, checks_usuario, window, schedule.margen_salida_despues_min
            )
            if prestada is not None:
                checks_by_user[usuario_id] = [c for c in checks_usuario if c is not prestada]
        es_feriado = fecha_laboral in festivos
        ausencia = _find_ausencia_justificada(ausencias, usuario_id, fecha_laboral)
        ausencia_aplica_laboralmente = bool(
            ausencia
            and not es_feriado
            and not es_dia_descanso
        )
        tipo_slug = ausencia.get("tipo_slug") if ausencia_aplica_laboralmente else None
        solicitud_id = ausencia["id"] if ausencia_aplica_laboralmente else None
        tiene_vacaciones = tipo_slug == "vacaciones"
        tiene_ausencia_justificada = ausencia_aplica_laboralmente
        resumen = calcular_resumen_dia(
            checks=checks_dia,
            schedule=schedule,
            tiene_vacaciones=tiene_vacaciones,
            tiene_ausencia_justificada=tiene_ausencia_justificada,
            ausencia_tipo_nombre=ausencia.get("tipo_nombre") if ausencia_aplica_laboralmente else None,
            es_feriado=es_feriado,
            fecha_laboral=fecha_laboral,
            now=calculated_at,
            min_minutos_he=min_minutos_he,
        )
        compensatorio = comp_by_user_date.get((usuario_id, fecha_laboral))
        minutos_compensatorio_aprobado = int(compensatorio["minutos_solicitados"]) if compensatorio else 0
        minutos_compensatorio = 0
        if minutos_compensatorio_aprobado > 0 and not checks_dia:
            minutos_compensatorio = minutos_compensatorio_aprobado
            observaciones = [resumen.get("observaciones")] if resumen.get("observaciones") else []
            observaciones.append("Horas extra tomadas")
            resumen["estado"] = "he_compensatorio"
            resumen["minutos_extra"] = 0
            resumen["observaciones"] = "; ".join(observaciones)

        horas_extra_estado = "pendiente"
        if es_feriado and resumen.get("minutos_extra", 0) > 0:
            horas_extra_estado = "feriado"
        rows_to_save.append({
            "usuario_id": usuario_id,
            "sucursal_id": sucursal_by_user.get(usuario_id),
            "fecha_laboral": fecha_laboral,
            "tiene_vacaciones": tiene_vacaciones,
            "tiene_ausencia_justificada": tiene_ausencia_justificada,
            "solicitud_ausencia_id": solicitud_id,
            "calculated_at": calculated_at,
            "horas_extra_estado": horas_extra_estado,
            "minutos_he_compensatorio": minutos_compensatorio,
            "he_compensatorio_solicitud_id": compensatorio["id"] if minutos_compensatorio > 0 else None,
            **resumen,
        })

    await db.upsert_asistencia_diaria_batch(conn, rows_to_save)
    return rows_to_save


def _dedupe_targets(targets: list[tuple[UUID, date]]) -> list[tuple[UUID, date]]:
    """Contrato: el orden ascendente por (usuario_id, fecha) es load-bearing -- `recalcular_asistencia`
    depende de que, para un mismo usuario, las fechas se procesen de menor a mayor, para que una
    checada prestada de un descanso hacia el dia siguiente (`extender_salida_descanso_medianoche`)
    ya haya sido removida de `checks_by_user` antes de llegar a ese dia siguiente como su propio
    target. No cambiar el criterio de orden sin revisar ese uso."""
    return sorted(set(targets), key=lambda item: (str(item[0]), item[1]))


def _build_context_maps(rows: list[dict]) -> tuple[dict[tuple[UUID, int], ScheduleConfig], dict[UUID, UUID | None]]:
    schedules: dict[tuple[UUID, int], ScheduleConfig] = {}
    sucursales: dict[UUID, UUID | None] = {}
    for row in rows:
        usuario_id = row["usuario_id"]
        sucursales.setdefault(usuario_id, row.get("sucursal_id"))
        if row.get("dia_semana") is None or row.get("horario_id") is None:
            continue
        schedules[(usuario_id, row["dia_semana"])] = ScheduleConfig(
            hora_entrada=row.get("hora_entrada"),
            hora_salida=row.get("hora_salida"),
            minutos_programados=row.get("minutos_programados") or 0,
            es_laboral=row.get("es_laboral", True),
            cruza_medianoche=row.get("cruza_medianoche", False),
            margen_entrada_antes_min=row.get("margen_entrada_antes_min") or 0,
            margen_salida_despues_min=row.get("margen_salida_despues_min") or 0,
            tolerancia_extra_min=row.get("tolerancia_extra_min") or 0,
            descuento_comida_min=row.get("descuento_comida_min") or 0,
        )
    return schedules, sucursales


def _group_checks_by_user(rows: list[dict]) -> dict[UUID, list[AttendanceCheck]]:
    grouped: dict[UUID, list[AttendanceCheck]] = {}
    for row in rows:
        grouped.setdefault(row["usuario_id"], []).append(
            AttendanceCheck(check_time=row["check_time"], punch_state=row.get("punch_state"))
        )
    return grouped


def _find_ausencia_justificada(
    ausencias: list[dict],
    usuario_id: UUID,
    fecha_laboral: date,
) -> dict | None:
    """`ausencias` ya viene ordenada (updated_at DESC, id) por
    db.get_ausencias_justificadas -- ante un solapamiento anomalo de dos solicitudes
    aprobadas para la misma fecha (dato historico, bloqueado hoy en creacion/aprobacion
    para solicitudes nuevas), esta funcion registra la anomalia y usa de forma deterministica
    la mas recientemente actualizada, en vez de una eleccion arbitraria por orden de BD."""
    encontradas = [
        row for row in ausencias
        if row["usuario_id"] == usuario_id and row["fecha_inicio"] <= fecha_laboral <= row["fecha_fin"]
    ]
    if len(encontradas) > 1:
        logger.warning(
            "[ASISTENCIA_ANOMALIA] %d solicitudes aprobadas solapadas para usuario=%s fecha=%s; "
            "se usa la mas reciente (id=%s)",
            len(encontradas), usuario_id, fecha_laboral, encontradas[0]["id"],
        )
    return encontradas[0] if encontradas else None


async def omitir_horas_extra_propio_svc(
    conn,
    *,
    asistencia_id: UUID,
    usuario_id: UUID,
) -> dict:
    """Retiro propio: el dueno del registro descarta 'pendiente' o retira 'solicitado'.

    No es autorizacion de terceros (valida propiedad, no puede_autorizar_he), pero usa
    el mismo lock de empleado para linealizar contra aprobaciones/reasignaciones concurrentes.
    """
    async with conn.transaction():
        await db.lock_he_usuario(conn, usuario_id)
        row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
        if not row:
            raise ValueError("Registro no encontrado")
        if row["usuario_id"] != usuario_id:
            raise ValueError("No tienes permiso para modificar este registro")
        if row["horas_extra_estado"] not in ("pendiente", "solicitado"):
            raise ValueError("Solo puedes retirar registros pendientes o solicitados")
        estado_anterior = row["horas_extra_estado"]
        if not await db.omitir_horas_extra(conn, asistencia_id):
            raise ValueError("Este registro ya fue procesado")
    if estado_anterior == "solicitado":
        await _notificar_retiro_horas_extra(conn, usuario_id=usuario_id, row=row)
    return {"empleado_nombre": row["empleado_nombre"], "estado_anterior": estado_anterior}


async def _notificar_retiro_horas_extra(conn, *, usuario_id: UUID, row: dict) -> None:
    from core.workflow.notification_service import NotificationService

    resuelto = await resolver_destinatarios_he(conn, usuario_id)
    destinatarios = resuelto["to"] | resuelto["cc"] | resuelto["bcc"]
    if not destinatarios:
        return
    svc_notif = NotificationService()
    await svc_notif.notify_he_solicitud_retirada(
        conn,
        empleado_nombre=row["empleado_nombre"],
        fecha_laboral=row["fecha_laboral"],
        extra_fmt=format_minutes(row["minutos_extra"] or 0),
        destinatarios=destinatarios,
    )


async def omitir_horas_extra_svc(
    conn,
    *,
    asistencia_id: UUID,
    aprobador_id: UUID,
    comentario: str,
) -> dict:
    comentario_limpio = (comentario or "").strip()
    if not comentario_limpio:
        raise ValueError("El motivo del rechazo es obligatorio")
    async with conn.transaction():
        row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
        if not row:
            raise ValueError("Registro no encontrado")
        await _lock_y_exigir_autorizacion_he(
            conn, empleado_id=row["usuario_id"], aprobador_id=aprobador_id
        )
        if row["horas_extra_estado"] not in ("pendiente", "solicitado"):
            raise ValueError("Solo se pueden descartar registros pendientes o solicitados")
        if not await db.omitir_horas_extra(conn, asistencia_id, comentario_limpio):
            raise ValueError("Este registro ya fue procesado")
    return {"empleado_nombre": row["empleado_nombre"]}


async def recuperar_horas_extra_svc(
    conn,
    *,
    asistencia_id: UUID,
    aprobador_id: UUID,
) -> dict:
    async with conn.transaction():
        row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
        if not row:
            raise ValueError("Registro no encontrado")
        await _lock_y_exigir_autorizacion_he(
            conn, empleado_id=row["usuario_id"], aprobador_id=aprobador_id
        )
        if row["horas_extra_estado"] != "omitido":
            raise ValueError("El registro no está descartado")
        if not await db.recuperar_horas_extra(conn, asistencia_id):
            raise ValueError("Este registro ya fue procesado")
    return {"empleado_nombre": row["empleado_nombre"]}


async def revertir_dia_horas_extra_svc(
    conn,
    *,
    asistencia_id: UUID,
    revertido_por: UUID,
) -> dict:
    """Correccion manual RH: reabre un dia 'feriado' o 'aprobado' congelado por BioTime.

    Uso exclusivo RH (sin credito que revertir en 'feriado'; reversion de credito
    en 'aprobado' via revertir_horas_extra_aprobado). Sin exposicion en UI todavia
    — ver PENDIENTES_RH.md seccion 4. Fuera del override de aprobador exclusivo (3.6
    del plan): el gate de acceso es RH manager, no puede_autorizar_he.
    """
    row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
    if not row:
        raise ValueError("Registro no encontrado")

    async with conn.transaction():
        await db.lock_he_usuario(conn, row["usuario_id"])
        estado = row["horas_extra_estado"]
        if estado == "feriado":
            ok = await db.recuperar_dia_feriado(conn, asistencia_id)
        elif estado == "aprobado":
            ok = await db.revertir_horas_extra_aprobado(conn, asistencia_id, revertido_por)
        else:
            raise ValueError("Solo se puede corregir un registro en estado 'feriado' o 'aprobado'")

        if not ok:
            raise ValueError("El registro ya no esta en el estado esperado")
    return {
        "empleado_nombre": row["empleado_nombre"],
        "fecha_laboral": row["fecha_laboral"],
        "estado_anterior": estado,
    }


def _validar_registro_propio(row: dict | None, usuario_id: UUID, motivo: str) -> str:
    if not row:
        raise ValueError("Registro no encontrado")
    if row["usuario_id"] != usuario_id:
        raise ValueError("No tienes permiso para este registro")
    if not motivo or not motivo.strip():
        raise ValueError("El motivo es obligatorio")
    return motivo.strip()


async def _validar_solicitud_horas_extra(conn, row: dict, usuario_id: UUID, motivo: str) -> str:
    motivo_limpio = _validar_registro_propio(row, usuario_id, motivo)
    if row["horas_extra_estado"] == "feriado":
        raise ValueError("Las horas trabajadas en feriado son pago economico y no van a bolsa")
    festivos = await db.get_festivos_range(conn, row["fecha_laboral"], row["fecha_laboral"])
    if row["fecha_laboral"] in festivos:
        raise ValueError("Las horas trabajadas en feriado son pago economico y no van a bolsa")
    if row["horas_extra_estado"] != "pendiente":
        raise ValueError("Solo puedes solicitar aprobacion de registros pendientes")
    if row.get("minutos_he_compensatorio"):
        raise ValueError("Este dia ya tiene horas extra tomadas")
    return motivo_limpio


async def solicitar_aprobacion_svc(
    conn,
    *,
    asistencia_id: UUID,
    usuario_id: UUID,
    motivo: str,
    empleado_nombre: str | None = None,
) -> dict:
    row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
    motivo_limpio = await _validar_solicitud_horas_extra(conn, row, usuario_id, motivo)
    async with conn.transaction():
        await db.lock_he_usuario(conn, usuario_id)
        if not await db.solicitar_aprobacion_horas_extra(conn, asistencia_id, usuario_id, motivo_limpio):
            raise ValueError("Este registro ya fue procesado o no aplica")
    if empleado_nombre:
        await _notificar_solicitud_horas_extra(
            conn,
            usuario_id=usuario_id,
            empleado_nombre=empleado_nombre,
            fecha_laboral=row["fecha_laboral"],
            minutos_extra=row["minutos_extra"],
            motivo=motivo_limpio,
        )
    return {"fecha_laboral": row["fecha_laboral"], "minutos_extra": row["minutos_extra"]}


def resolver_destinatarios_he_puro(
    *,
    tiene_override: bool,
    override_email: str | None,
    jefe_emails,
    tiene_director: bool,
    aprobador_vac_email: str | None,
    fallback_emails: set[str],
    escalacion_cc: set[str] | None = None,
    escalacion_cco: set[str] | None = None,
) -> dict:
    """Logica pura del resolver unico de destinatarios TO/CC/CCO/motivo/URL para HE y compensatorio.

    Sin acceso a BD: recibe columnas ya resueltas para poder reusarse tanto en la solicitud
    inmediata/retiro propio (round trip via resolver_destinatarios_he) como en los recordatorios
    batched del worker (columnas ya traidas por la query, sin N+1).

    Prioridad: aprobador exclusivo activo (TO exclusivo, CC/CCO vacios) > aprobador exclusivo
    inactivo (fallback RH/ADMIN por rol, CC/CCO vacios) > regla normal: jefes activos union
    aprobador de vacaciones activo (CC/CCO = reglas configurables en Admin > Reglas de correo,
    evento HE_EVENTO_ESCALACION_DIRECTOR, solo si algun jefe jerarquico es director Y su correo
    quedo en TO) > sin destinatarios normales -> fallback.

    `fallback_emails` es el pool de enrutamiento real (RH editor/admin + ADMIN global por rol,
    ver get_fallback_rh_admin_emails) -- se usa solo cuando no hay ningun destinatario valido,
    porque alguien tiene que poder aprobar. `escalacion_cc`/`escalacion_cco` es una lista de
    correos manual editable desde Admin (no atada a rol) para la visibilidad informativa cuando
    la aprobacion escala a un director: si RH no configuro nada, no se agrega nadie.

    ATENCION: el input `tiene_override` codifica la misma regla de precedencia (exclusivo >
    jefe/aprobador de vacaciones) que modules.asistencia.db_service._HE_AUTORIZACION_PRECEDENCIA_QUERY
    (fuente unica compartida por get_empleados_para_autorizacion_he y puede_autorizar_he). No se
    unifico tambien aqui porque `tiene_override` viene de get_datos_resolucion_notificacion_he,
    que devuelve una fila de datos de notificacion (no un booleano/lista de UUIDs) y solo necesita
    el discriminador simple `id_aprobador_horas_extra IS NOT NULL`, sin la union con jefe/aprobador
    de vacaciones. Si cambias la regla de precedencia, replicala tambien ahi --
    tests/test_aprobador_he_exclusivo.py::test_consistencia_* la valida contra fixtures compartidos.
    """
    url_perfil = f"{settings.APP_BASE_URL}/perfil/ui?tab=aprobaciones"
    url_rh = f"{settings.APP_BASE_URL}/rrhh/ui?tab=aprobaciones"

    def _fallback_rh() -> dict:
        return {
            "to": set(fallback_emails),
            "cc": set(),
            "bcc": set(),
            "url": url_rh,
            "label_boton": "Revisar en RRHH",
        }

    if tiene_override:
        if override_email:
            return {
                "to": {override_email},
                "cc": set(),
                "bcc": set(),
                "url": url_perfil,
                "label_boton": "Revisar en Aprobaciones",
            }
        return _fallback_rh()

    jefe_emails_set = {email for email in (jefe_emails or []) if email}
    to_emails = jefe_emails_set | ({aprobador_vac_email} if aprobador_vac_email else set())
    if not to_emails:
        return _fallback_rh()
    hay_escalacion = tiene_director and jefe_emails_set
    return {
        "to": to_emails,
        "cc": set(escalacion_cc or set()) if hay_escalacion else set(),
        "bcc": set(escalacion_cco or set()) if hay_escalacion else set(),
        "url": url_perfil,
        "label_boton": "Revisar en Aprobaciones",
    }


async def verificar_fallback_aprobador_he_svc(
    conn, *, user_id: UUID, sera_fallback_despues: bool
) -> None:
    """Guard para deactivate_user/update_user_role/update_user_modules/save_user_all.

    Llamar dentro de la misma transaccion, bajo lock_he_fallback, antes de persistir el cambio.
    Lanza HEFallbackVacioError (409) si algun empleado con aprobador HE exclusivo quedaria sin
    ningun destinatario de respaldo (RH editor/admin o ADMIN global) con correo.

    NOTA arquitectura: este invariante ("todo aprobador HE exclusivo tiene fallback vivo") solo
    se garantiza por convencion -- los 4 mutadores de modules/admin/service.py que pueden afectar
    el pool de fallback (update_user_role, update_user_modules, save_user_all, deactivate_user)
    llaman este guard. No hay backstop a nivel de esquema (es un invariante cross-row, no
    expresable como CHECK de una sola fila); un futuro camino de escritura que cambie rol_sistema
    o tb_permisos_modulos sin pasar por admin/service.py lo saltaria en silencio.
    """
    resultado = await db.verificar_fallback_aprobador_he(
        conn, user_id=user_id, sera_fallback_despues=sera_fallback_despues
    )
    if resultado["afectados_count"] > 0 and not resultado["tiene_fallback"]:
        raise HEFallbackVacioError(
            f"No se puede completar el cambio: {resultado['afectados_count']} empleado(s) con "
            "aprobador de horas extra exclusivo quedarian sin ningun RH o Administrador de respaldo "
            "con correo. Asigna otro RH/Administrador activo antes de continuar."
        )


async def notify_aprobador_he_inactivo(
    conn, *, aprobador_id: UUID, afectados: list[dict] | None = None
) -> None:
    """Post-commit, best-effort: avisa a RH/ADMIN de los empleados cuyo aprobador HE exclusivo
    acaba de quedar inactivo. Un fallo aqui solo se registra (no revierte la baja ya confirmada) --
    el propio try/except vive aqui (no en el llamador) para que el contrato "best-effort" se
    cumpla sin importar quien invoque esta funcion.

    `afectados`, si se pasa, evita repetir la consulta que el llamador (ej. deactivate_user via
    _lock_y_guardar_fallback_he) ya hizo dentro de la misma transaccion recien commiteada."""
    try:
        if afectados is None:
            afectados = await db.get_empleados_con_aprobador_he_exclusivo(conn, aprobador_id)
        if not afectados:
            return
        fallback_emails = await db.get_fallback_rh_admin_emails(conn)
        if not fallback_emails:
            logger.warning(
                "[HE_APROBADOR_INACTIVO] Sin destinatarios de fallback para avisar de %d afectado(s)",
                len(afectados),
            )
            return
        from core.workflow.notification_service import NotificationService

        svc_notif = NotificationService()
        await svc_notif.notify_aprobador_he_inactivo(
            conn, empleados_afectados=afectados, destinatarios=fallback_emails
        )
    except (asyncpg.PostgresError, asyncpg.InterfaceError) as exc:
        logger.error(
            "[HE_APROBADOR_INACTIVO] Error avisando fallback para aprobador %s: %s",
            aprobador_id,
            exc,
        )


async def get_escalacion_director_cc_cco(conn) -> tuple[set[str], set[str]]:
    """CC/CCO configurables en Admin > Reglas de correo para el evento de escalacion a
    director (HE_EVENTO_ESCALACION_DIRECTOR). Usado por resolver_destinatarios_he (solicitud
    inmediata) y por el worker de recordatorios (core/tasks.py) una sola vez antes de su loop,
    para no repetir la consulta por fila. Un solo round trip via db.get_escalacion_director_emails
    (agrupa CC/CCO en una sola query, no dos)."""
    resuelto = await db.get_escalacion_director_emails(conn)
    return resuelto["cc"], resuelto["bcc"]


async def resolver_destinatarios_he(conn, usuario_id: UUID) -> dict:
    """Wrapper con round trip a BD de resolver_destinatarios_he_puro; usar en solicitud
    inmediata y retiro propio. Los recordatorios batched del worker llaman la version pura
    directamente con columnas ya traidas por su query (ver core/tasks.py). El CC/CCO de
    escalacion viene en el mismo round trip que el resto de datos (get_datos_resolucion_
    notificacion_he) -- no se llama get_escalacion_director_cc_cco aqui para no pagar una
    segunda query en un path sincrono de solicitud inmediata."""
    datos = await db.get_datos_resolucion_notificacion_he(conn, usuario_id)
    return resolver_destinatarios_he_puro(
        tiene_override=datos["tiene_override"],
        override_email=datos.get("override_email"),
        jefe_emails=datos.get("jefe_emails"),
        tiene_director=datos.get("tiene_director", False),
        aprobador_vac_email=datos.get("aprobador_vac_email"),
        fallback_emails=set(datos.get("fallback_emails") or []),
        escalacion_cc=set(datos.get("escalacion_cc") or []),
        escalacion_cco=set(datos.get("escalacion_cco") or []),
    )


async def _notificar_solicitud_horas_extra(
    conn,
    *,
    usuario_id: UUID,
    empleado_nombre: str,
    fecha_laboral: date,
    minutos_extra: int,
    motivo: str,
) -> None:
    from core.workflow.notification_service import NotificationService

    svc_notif = NotificationService()
    resuelto = await resolver_destinatarios_he(conn, usuario_id)
    await svc_notif.notify_horas_extra_solicitud(
        conn,
        empleado_nombre=empleado_nombre,
        fecha_laboral=fecha_laboral,
        extra_fmt=format_minutes(minutos_extra),
        motivo=motivo,
        destinatarios=resuelto["to"],
        cc_emails=resuelto["cc"],
        bcc_emails=resuelto["bcc"],
        url_aprobacion=resuelto["url"],
        label_boton=resuelto["label_boton"],
    )


async def attach_he_evidencias(conn, rows: list[dict]) -> list[dict]:
    evidencias = await db.get_he_evidencias_for_aprobador(conn, [row["id"] for row in rows])
    for row in rows:
        row["evidencias"] = evidencias.get(str(row["id"]), [])
    return rows


def build_horas_extra_grupos(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    grupos_map: dict[str, dict] = {}
    json_rows: list[dict] = []
    for row in rows:
        row["extra_fmt"] = format_minutes(row.get("minutos_extra") or 0)
        row["entrada_fmt"] = fmt_time_mx(row.pop("primera_entrada", None))
        row["salida_fmt"] = fmt_time_mx(row.pop("ultima_salida", None))
        json_rows.append({
            "id": str(row["id"]),
            "usuario_id": str(row["usuario_id"]),
            "empleado_nombre": row["empleado_nombre"],
            "fecha_fmt": row["fecha_laboral"].strftime("%d/%m/%Y"),
            "minutos_extra": int(row.get("minutos_extra") or 0),
            "horas_extra_estado": row.get("horas_extra_estado", "pendiente"),
            "motivo_solicitud": row.get("motivo_solicitud"),
            "entrada_fmt": row["entrada_fmt"],
            "salida_fmt": row["salida_fmt"],
        })
        uid = str(row["usuario_id"])
        if uid not in grupos_map:
            grupos_map[uid] = {
                "usuario_id": uid,
                "empleado_nombre": row["empleado_nombre"],
                "rows": [],
                "tiene_solicitado": False,
            }
        grupos_map[uid]["rows"].append(row)
        if row.get("horas_extra_estado") == "solicitado":
            grupos_map[uid]["tiene_solicitado"] = True
    return list(grupos_map.values()), json_rows


def _format_he_item(row: dict) -> dict:
    item = dict(row)
    item["minutos_fmt"] = format_minutes(item.get("minutos") or item.get("minutos_solicitados") or 0)
    return item


def _format_aprobador_he_label(usuario: dict) -> str:
    aprobador_he_nombre = usuario.get("aprobador_he_nombre")
    if aprobador_he_nombre:
        return f"Aprobador de horas extra (exclusivo): {aprobador_he_nombre}"
    return "Aprobador de horas extra: regla normal (jefe(s) jerarquico(s) de arriba)"


async def get_he_nivel_ctx(conn, usuario_id: UUID) -> dict | None:
    nivel = await db.get_he_nivel_usuario(conn, usuario_id, today_mx().year)
    return dict(nivel) if nivel else None


async def get_he_niveles_equipo_ctx(conn, usuario_ids: list[UUID]) -> list[dict]:
    return await db.get_he_niveles_equipo(conn, usuario_ids, today_mx().year)


async def get_he_niveles_escalera_ctx(conn, usuario_id: UUID, nivel_actual: dict | None = None) -> dict:
    catalogo = await db.get_he_niveles_catalogo(conn)
    if nivel_actual is None:
        nivel_actual = await get_he_nivel_ctx(conn, usuario_id)
    horas_actuales = nivel_actual["horas_actuales"] if nivel_actual else 0
    niveles = [
        {
            **fila,
            "alcanzado": fila["umbral_horas"] <= horas_actuales,
            "actual": bool(nivel_actual) and fila["nivel"] == nivel_actual["nivel"],
        }
        for fila in catalogo
    ]
    return {
        "niveles": niveles,
        "horas_actuales": horas_actuales,
        "horas_faltantes": nivel_actual["horas_faltantes"] if nivel_actual else None,
        "es_maximo": bool(nivel_actual) and nivel_actual["es_maximo"],
    }


async def get_he_bolsa_ctx(conn, usuario_id: UUID) -> dict:
    saldo = await db.get_he_saldo_usuario(conn, usuario_id)
    movimientos = await db.get_he_movimientos_usuario(conn, usuario_id)
    solicitudes = await db.get_he_solicitudes_compensatorio_usuario(conn, usuario_id)
    nivel_ctx = await get_he_nivel_ctx(conn, usuario_id)
    return {
        "saldo": saldo,
        "saldo_fmt": {
            "acumulado": format_minutes(saldo["minutos_acumulados"]),
            "tomado": format_minutes(saldo["minutos_tomados"]),
            "en_proceso": format_minutes(saldo["minutos_en_proceso"]),
            "disponible": format_minutes(saldo["minutos_disponibles"]),
        },
        "movimientos": [_format_he_item(row) for row in movimientos],
        "solicitudes": [_format_he_item(row) for row in solicitudes],
        "nivel_ctx": nivel_ctx,
        "niveles_ctx": await get_he_niveles_escalera_ctx(conn, usuario_id, nivel_actual=nivel_ctx),
    }


def validar_debito_compensatorio(saldo_disponible: int, minutos_solicitados: int) -> None:
    if minutos_solicitados <= 0:
        raise ValueError("Los minutos solicitados deben ser mayores a 0")
    if saldo_disponible < minutos_solicitados:
        raise ValueError("Saldo insuficiente en la bolsa de horas extra")


def _build_he_bolsa_workbook(usuarios: list[dict], saldos: dict, movimientos: list[dict], feriados: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bolsa HE"
    title_fill = PatternFill("solid", fgColor="00BABB")
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    mov_by_user: dict[UUID, list[dict]] = {}
    for row in movimientos:
        mov_by_user.setdefault(row["usuario_id"], []).append(row)
    fer_by_user: dict[UUID, list[dict]] = {}
    for row in feriados:
        fer_by_user.setdefault(row["usuario_id"], []).append(row)

    headers = ["Fecha", "Concepto", "Horas", "Saldo despues"]
    for usuario in usuarios:
        uid = usuario["id_usuario"]
        saldo = saldos.get(uid, {
            "minutos_acumulados": 0,
            "minutos_tomados": 0,
            "minutos_disponibles": 0,
        })
        ws.append([
            f"Empleado: {usuario.get('nombre') or ''}",
            f"Email: {usuario.get('email') or ''}",
            f"Jefe(s) jerarquico(s): {usuario.get('jefes_nombres') or '(sin asignar)'}",
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = title_fill
        ws.append([_format_aprobador_he_label(usuario)])
        ws.append([
            "Horas acumuladas",
            "Horas tomadas",
            "Horas disponibles",
        ])
        ws.append([
            round((saldo.get("minutos_acumulados") or 0) / 60, 2),
            round((saldo.get("minutos_tomados") or 0) / 60, 2),
            round((saldo.get("minutos_disponibles") or 0) / 60, 2),
        ])
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for mov in mov_by_user.get(uid, []):
            horas = round((mov["minutos"] if mov["tipo"] == "CREDITO" else -mov["minutos"]) / 60, 2)
            ws.append([
                mov["fecha_referencia"],
                mov["concepto"],
                horas,
                round((mov["saldo_despues"] or 0) / 60, 2),
            ])
        for feriado in fer_by_user.get(uid, []):
            ws.append([
                feriado["fecha_referencia"],
                feriado["concepto"],
                round((feriado["minutos_extra"] or 0) / 60, 2),
                "",
            ])
        ws.append([])
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        ws.column_dimensions[letter].width = min(42, max(12, max(len(str(c.value or "")) for c in column_cells) + 2))
    return wb


async def generar_reporte_bolsa_he_svc(conn, *, scope: str, usuario_id: UUID, context: dict) -> Workbook:
    if scope == "propio":
        usuario_ids = [usuario_id]
    elif scope == "equipo":
        equipo_consulta = await get_equipo_ids(conn, usuario_id, context)
        equipo_visible_he, _ = await get_equipo_visible_he(conn, usuario_id, context, equipo_consulta)
        usuario_ids = equipo_visible_he or [usuario_id]
    else:
        if not user_has_module_access("rrhh", context, "viewer"):
            raise PermissionError("No tienes acceso a este reporte")
        todos_ids = await vacaciones_db.get_all_empleado_ids_activos(conn, limit=10000)
        usuario_ids = await db.get_usuario_ids_con_he_aprobada(conn, todos_ids)

    usuarios = await db.get_he_reporte_usuarios(conn, usuario_ids)
    saldos = await db.get_he_saldo_reporte(conn, usuario_ids)
    movimientos = await db.get_he_movimientos_reporte(conn, usuario_ids)
    feriados = await db.get_he_feriados_reporte(conn, usuario_ids)
    return _build_he_bolsa_workbook(usuarios, saldos, movimientos, feriados)


async def _get_schedule_fecha(conn, usuario_id: UUID, fecha_descanso: date) -> ScheduleConfig | None:
    rows = await db.get_attendance_contexts(conn, [usuario_id])
    schedule_by_user_day, _ = _build_context_maps(rows)
    return schedule_by_user_day.get((usuario_id, fecha_descanso.weekday()))


async def _validar_fecha_compensatorio(conn, usuario_id: UUID, fecha_descanso: date) -> ScheduleConfig:
    hoy = today_mx()
    if fecha_descanso <= hoy:
        raise ValueError("La fecha debe ser a partir de manana")
    if fecha_descanso.weekday() >= 5:
        raise ValueError("El tiempo compensatorio debe solicitarse en un dia laboral")
    festivos = await db.get_festivos_range(conn, fecha_descanso, fecha_descanso)
    if fecha_descanso in festivos:
        raise ValueError("No se puede solicitar tiempo compensatorio en feriado")
    schedule = await _get_schedule_fecha(conn, usuario_id, fecha_descanso)
    if not schedule or not schedule.es_laboral or schedule.minutos_programados <= 0:
        raise ValueError("No hay jornada laboral programada para esa fecha")
    return schedule


async def solicitar_compensatorio_svc(
    conn,
    *,
    usuario_id: UUID,
    fecha_descanso: date,
    minutos_solicitados: int,
    motivo: str,
) -> dict:
    motivo_limpio = (motivo or "").strip()
    if not motivo_limpio:
        raise ValueError("El motivo es obligatorio")

    async with conn.transaction():
        await db.lock_he_usuario(conn, usuario_id)
        schedule = await _validar_fecha_compensatorio(conn, usuario_id, fecha_descanso)
        if minutos_solicitados < 30 or minutos_solicitados > min(1440, schedule.minutos_programados):
            raise ValueError("Los minutos solicitados deben estar dentro de la jornada programada")

        # Sin solo_justificadas: bloquea contra cualquier solicitud activa (incluye
        # home_office/permiso_llegar_tarde/permiso_salir_temprano), simetrico con
        # crear_solicitud, que ya bloquea la creacion de una ausencia si hay compensatorio activo.
        ausencias = await vacaciones_db.get_solicitudes_activas_en_rango(
            conn, usuario_id, fecha_descanso, fecha_descanso
        )
        if ausencias:
            raise ValueError("Ya existe una ausencia activa para esa fecha")

        saldo = await db.get_he_saldo_usuario(conn, usuario_id)
        validar_debito_compensatorio(saldo["minutos_disponibles"], minutos_solicitados)
        try:
            solicitud = await db.crear_he_solicitud_compensatorio(
                conn,
                usuario_id=usuario_id,
                fecha_descanso=fecha_descanso,
                minutos_solicitados=minutos_solicitados,
                motivo=motivo_limpio,
            )
        except asyncpg.UniqueViolationError as exc:
            raise ValueError("Ya existe una solicitud activa para esa fecha") from exc

    await _notificar_compensatorio_solicitud(conn, solicitud)
    return solicitud


async def _notificar_compensatorio_solicitud(conn, solicitud: dict | None) -> None:
    if not solicitud:
        return
    from core.workflow.notification_service import NotificationService

    svc_notif = NotificationService()
    resuelto = await resolver_destinatarios_he(conn, solicitud["usuario_id"])
    await svc_notif.notify_compensatorio_solicitud(
        conn,
        empleado_nombre=solicitud["empleado_nombre"],
        fecha_descanso=solicitud["fecha_descanso"],
        minutos_fmt=format_minutes(solicitud["minutos_solicitados"]),
        motivo=solicitud["motivo"],
        destinatarios=resuelto["to"],
        cc_emails=resuelto["cc"],
        bcc_emails=resuelto["bcc"],
        url_aprobacion=resuelto["url"],
        label_boton=resuelto["label_boton"],
    )


async def _lock_solicitud_compensatorio_pendiente(
    conn, solicitud_id: UUID, aprobador_id: UUID
) -> dict:
    """FOR UPDATE del recurso, luego lock actor + lock empleado, luego revalida autorizacion."""
    solicitud = await db.get_he_compensatorio_by_id(conn, solicitud_id, for_update=True)
    if not solicitud:
        raise ValueError("Solicitud no encontrada")
    await _lock_y_exigir_autorizacion_he(
        conn, empleado_id=solicitud["usuario_id"], aprobador_id=aprobador_id
    )
    if solicitud["estatus"] != "pendiente":
        raise ValueError("La solicitud ya fue procesada")
    return solicitud


async def aprobar_compensatorio_svc(
    conn,
    *,
    solicitud_id: UUID,
    aprobador_id: UUID,
    comentario: str | None = None,
) -> dict:
    async with conn.transaction():
        solicitud = await _lock_solicitud_compensatorio_pendiente(conn, solicitud_id, aprobador_id)
        if solicitud["fecha_descanso"] < today_mx():
            raise ValueError("La solicitud vencio; rechazala para liberar el saldo")

        estado_dia = await db.get_horas_extra_estado_en_fecha(
            conn, solicitud["usuario_id"], solicitud["fecha_descanso"]
        )
        if estado_dia in ("solicitado", "aprobado"):
            raise ValueError(
                "Ese dia ya tiene horas extra en proceso o aprobadas; no se puede aplicar como compensatorio"
            )

        saldo = await db.get_he_saldo_usuario(
            conn, solicitud["usuario_id"], excluir_solicitud_pendiente_id=solicitud_id
        )
        validar_debito_compensatorio(
            saldo["minutos_disponibles"], solicitud["minutos_solicitados"]
        )
        updated = await db.aprobar_he_compensatorio(
            conn,
            solicitud_id=solicitud_id,
            aprobador_id=aprobador_id,
            comentario=(comentario or "").strip(),
        )
        if not updated:
            raise ValueError("La solicitud ya fue procesada")
        await recalcular_asistencia(conn, [(solicitud["usuario_id"], solicitud["fecha_descanso"])])

    result = {**solicitud, **updated}
    await _notificar_compensatorio_resuelto(conn, result, aprobado=True)
    return result


async def rechazar_compensatorio_svc(
    conn,
    *,
    solicitud_id: UUID,
    aprobador_id: UUID,
    comentario: str,
) -> dict:
    comentario_limpio = (comentario or "").strip()
    if not comentario_limpio:
        raise ValueError("El comentario es obligatorio")
    async with conn.transaction():
        solicitud = await _lock_solicitud_compensatorio_pendiente(conn, solicitud_id, aprobador_id)
        updated = await db.rechazar_he_compensatorio(
            conn,
            solicitud_id=solicitud_id,
            aprobador_id=aprobador_id,
            comentario=comentario_limpio,
        )
        if not updated:
            raise ValueError("La solicitud ya fue procesada")
    result = {**solicitud, **updated}
    await _notificar_compensatorio_resuelto(conn, result, aprobado=False)
    return result


async def cancelar_compensatorio_svc(conn, *, solicitud_id: UUID, usuario_id: UUID) -> dict:
    """Toma el row-lock de la solicitud antes que el advisory lock_he_usuario, en el mismo
    orden que _lock_solicitud_compensatorio_pendiente (aprobar/rechazar) — el orden inverso
    (advisory antes que row-lock) puede producir un deadlock detectado por Postgres si esta
    funcion corre a la vez que una aprobacion/rechazo sobre la misma solicitud."""
    async with conn.transaction():
        solicitud = await db.get_he_compensatorio_by_id(conn, solicitud_id, for_update=True)
        if not solicitud or solicitud["usuario_id"] != usuario_id or solicitud["estatus"] != "pendiente":
            raise ValueError("Solo puedes cancelar solicitudes pendientes propias")
        await db.lock_he_usuario(conn, usuario_id)
        updated = await db.cancelar_he_compensatorio(
            conn,
            solicitud_id=solicitud_id,
            usuario_id=usuario_id,
        )
        if not updated:
            raise ValueError("Solo puedes cancelar solicitudes pendientes propias")
    return updated


async def _notificar_compensatorio_resuelto(conn, solicitud: dict | None, *, aprobado: bool) -> None:
    if not solicitud:
        return
    from core.workflow.notification_service import NotificationService

    svc_notif = NotificationService()
    await svc_notif.notify_compensatorio_resuelto(conn, solicitud, aprobado=aprobado)


async def get_he_bolsa_fecha_corte(conn) -> date:
    valor = await ConfigService.get_global_config(conn, "HE_BOLSA_FECHA_CORTE", "2026-07-07", str)
    return date.fromisoformat(valor)


async def confirmar_saldo_inicial_svc(
    conn,
    *,
    usuario_id: UUID,
    minutos: int,
    confirmado_por: UUID,
    context: dict,
) -> dict:
    if minutos < 0:
        raise ValueError("El saldo inicial no puede ser negativo")
    es_rrhh = user_has_module_access("rrhh", context, "editor")
    ids_jefe = await vacaciones_db.get_empleados_donde_soy_jefe(conn, confirmado_por)
    if not es_rrhh and usuario_id not in ids_jefe:
        raise ValueError("Solo el jefe directo o RRHH pueden confirmar este saldo")
    fecha_corte = await get_he_bolsa_fecha_corte(conn)
    async with conn.transaction():
        await db.lock_he_usuario(conn, usuario_id)
        try:
            return await db.confirmar_saldo_inicial(
                conn,
                usuario_id=usuario_id,
                minutos=minutos,
                confirmado_por=confirmado_por,
                fecha_corte=fecha_corte,
            )
        except asyncpg.UniqueViolationError as exc:
            raise ValueError("El saldo inicial ya fue confirmado") from exc


async def ajuste_manual_svc(
    conn,
    *,
    usuario_id: UUID,
    tipo: str,
    minutos: int,
    concepto: str,
    creado_por: UUID,
) -> UUID:
    tipo_limpio = (tipo or "").strip().upper()
    concepto_limpio = (concepto or "").strip()
    if tipo_limpio not in {"CREDITO", "DEBITO"}:
        raise ValueError("Tipo de ajuste invalido")
    if minutos <= 0:
        raise ValueError("Los minutos deben ser mayores a 0")
    if not concepto_limpio:
        raise ValueError("El concepto es obligatorio")
    async with conn.transaction():
        await db.lock_he_usuario(conn, usuario_id)
        if tipo_limpio == "DEBITO":
            saldo = await db.get_he_saldo_usuario(conn, usuario_id)
            validar_debito_compensatorio(saldo["minutos_disponibles"], minutos)
        return await db.crear_he_ajuste_manual(
            conn,
            usuario_id=usuario_id,
            tipo=tipo_limpio,
            minutos=minutos,
            concepto=concepto_limpio,
            fecha_referencia=today_mx(),
            creado_por=creado_por,
        )


_HE_EVIDENCIA_ALLOWED_TYPES = {"application/pdf", "image/jpeg", "image/png", "image/webp"}


def _safe_filename(filename: str) -> str:
    name = filename.rsplit("\\", 1)[-1].rsplit("/", 1)[-1]
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._") or "evidencia"


async def _validate_he_evidencias(
    conn, evidencias: list[UploadFile] | None
) -> list[tuple[UploadFile, int]]:
    files = [f for f in (evidencias or []) if f and f.filename]
    configs = await ConfigService.get_global_configs_bulk(
        conn,
        {
            "HE_EVIDENCIA_MAX_ARCHIVOS": (3, int),
            "HE_EVIDENCIA_MAX_MB": (4, int),
        },
    )
    max_archivos = configs["HE_EVIDENCIA_MAX_ARCHIVOS"]
    max_mb = configs["HE_EVIDENCIA_MAX_MB"]
    if len(files) > max_archivos:
        raise ValueError(f"Solo puedes adjuntar hasta {max_archivos} archivos")
    max_bytes = max_mb * 1024 * 1024
    result: list[tuple[UploadFile, int]] = []
    for file in files:
        if file.content_type not in _HE_EVIDENCIA_ALLOWED_TYPES:
            raise ValueError("Solo se aceptan PDF o imagenes como evidencia")
        await file.seek(0)
        content = await file.read()
        await file.seek(0)
        if len(content) > max_bytes:
            raise ValueError(f"Cada evidencia debe pesar maximo {max_mb} MB")
        result.append((file, len(content)))
    return result


async def subir_evidencias_he_y_solicitar_svc(
    conn,
    *,
    asistencia_id: UUID,
    usuario_id: UUID,
    motivo: str,
    empleado_nombre: str,
    evidencias: list[UploadFile] | None,
) -> dict:
    row = await db.get_asistencia_para_aprobar(conn, asistencia_id)
    motivo_limpio = await _validar_solicitud_horas_extra(conn, row, usuario_id, motivo)
    files_with_sizes = await _validate_he_evidencias(conn, evidencias)
    upload_results: list[tuple[UploadFile, dict, int]] = []
    config = None
    sp_service = None
    if files_with_sizes:
        base_folder = await ConfigService.get_global_config(conn, "SHAREPOINT_BASE_FOLDER", "", str)
        if not base_folder:
            raise ValueError("Falta configurar SHAREPOINT_BASE_FOLDER")
        token = await get_ms_auth().get_application_token()
        if not token:
            raise ValueError("No se pudo obtener token de aplicacion para SharePoint")
        sp_service = SharePointService(token)
        config = await sp_service._resolve_config(conn)
        if not config.get("site_id") and not config.get("drive_id"):
            raise ValueError("Falta configurar SharePoint para evidencias")
        folder = f"{base_folder.strip('/')}/he_evidencia/{usuario_id}/{asistencia_id}"
        timestamp = now_mx().strftime("%Y%m%d_%H%M%S")
        for file, _size in files_with_sizes:
            file.filename = f"{timestamp}_{_safe_filename(file.filename)}"

        results = [
            await sp_service.upload_file(conn, file, folder, _config=config)
            for file, _size in files_with_sizes
        ]
        upload_results = [
            (file, result, size)
            for (file, size), result in zip(files_with_sizes, results)
        ]

    try:
        async with conn.transaction():
            await db.lock_he_usuario(conn, usuario_id)
            if not await db.solicitar_aprobacion_horas_extra(conn, asistencia_id, usuario_id, motivo_limpio):
                raise ValueError("Este registro ya fue procesado o no aplica")
            for file, result, size in upload_results:
                await db.insertar_he_evidencia(
                    conn,
                    upload_result=result,
                    usuario_id=usuario_id,
                    asistencia_id=asistencia_id,
                    subido_por_id=usuario_id,
                    content_type=file.content_type or "application/octet-stream",
                    tamano_bytes=size,
                )
    except (asyncpg.PostgresError, ValueError):
        if sp_service and config:
            for _, result, _ in upload_results:
                item_id = result.get("id")
                if item_id:
                    try:
                        await sp_service.delete_file_by_item_id(conn, item_id, _config=config)
                    except (httpx.HTTPError, RuntimeError, ValueError) as exc:
                        logger.warning("No se pudo limpiar evidencia HE huerfana: %s", exc)
        raise

    await _notificar_solicitud_horas_extra(
        conn,
        usuario_id=usuario_id,
        empleado_nombre=empleado_nombre,
        fecha_laboral=row["fecha_laboral"],
        minutos_extra=row["minutos_extra"],
        motivo=motivo_limpio,
    )
    return {"fecha_laboral": row["fecha_laboral"], "minutos_extra": row["minutos_extra"]}


async def _get_config_manual_asistencia(conn) -> tuple[int, int]:
    configs = await ConfigService.get_global_configs_bulk(
        conn,
        {
            "ASISTENCIA_MANUAL_DIAS_RETROACTIVO": (7, int),
            "ASISTENCIA_MANUAL_MAX_HORAS": (16, int),
        },
    )
    return configs["ASISTENCIA_MANUAL_DIAS_RETROACTIVO"], configs["ASISTENCIA_MANUAL_MAX_HORAS"]


async def get_dias_retroactivo_manual(conn) -> int:
    dias_retroactivo, _ = await _get_config_manual_asistencia(conn)
    return dias_retroactivo


def _parse_manual_datetime(fecha: date | None, hora: str | None) -> datetime:
    if not fecha or not hora:
        raise ValueError("Fecha y hora son obligatorias")
    try:
        parsed_time = time.fromisoformat(hora)
    except ValueError as exc:
        raise ValueError("Hora invalida") from exc
    return datetime.combine(fecha, parsed_time, tzinfo=MX_TZ)


def _format_manual_datetime(value: datetime | None) -> str | None:
    if not value:
        return None
    return ensure_mx(value).strftime("%d/%m/%Y %H:%M")


def _format_manual_request(row: dict) -> dict:
    formatted = dict(row)
    formatted["estado_label"] = _SOLICITUD_MANUAL_ESTADOS.get(row.get("estado"), row.get("estado", ""))
    formatted["fecha_laboral_fmt"] = row["fecha_laboral"].strftime("%d/%m/%Y")
    formatted["entrada_fmt"] = _format_manual_datetime(row.get("entrada_tiempo"))
    formatted["salida_fmt"] = _format_manual_datetime(row.get("salida_tiempo"))
    formatted["created_at_fmt"] = _format_manual_datetime(row.get("created_at"))
    return formatted


def format_solicitudes_manuales(rows: list[dict]) -> list[dict]:
    return [_format_manual_request(row) for row in rows]


async def _get_labor_window_usuario_fecha(conn, usuario_id: UUID, fecha_laboral: date):
    context_rows = await db.get_attendance_contexts(conn, [usuario_id])
    schedule_by_user_day, _ = _build_context_maps(context_rows)
    schedule = schedule_by_user_day.get((usuario_id, fecha_laboral.weekday()))
    return build_labor_window(fecha_laboral, schedule), schedule


async def _detectar_huecos_manual(conn, usuario_id: UUID, fecha_laboral: date) -> dict:
    labor_window, schedule = await _get_labor_window_usuario_fecha(conn, usuario_id, fecha_laboral)
    checks = await db.get_biotime_checks_usuario_window(
        conn,
        usuario_id=usuario_id,
        start=labor_window.start,
        end=labor_window.end,
    )
    huecos = _clasificar_huecos_biotime(checks)
    return {
        "labor_window": labor_window,
        "schedule": schedule,
        "checks": checks,
        "huecos": huecos,
    }


def _clasificar_huecos_biotime(checks: list[dict]) -> dict:
    real_checks = sorted(
        [check for check in checks if not check.get("es_manual")],
        key=lambda check: ensure_mx(check["check_time"]),
    )
    if not real_checks:
        return {
            "falta_entrada": True,
            "falta_salida": True,
            "entrada_real": None,
            "salida_real": None,
            "bloqueado": False,
            "mensaje_bloqueo": None,
        }

    relevantes = [check for check in real_checks if not is_break_state(check.get("punch_state"))]
    entradas = [check for check in relevantes if is_in_state(check.get("punch_state"))]
    salidas = [check for check in relevantes if is_out_state(check.get("punch_state"))]
    desconocidas = [
        check
        for check in relevantes
        if not is_in_state(check.get("punch_state")) and not is_out_state(check.get("punch_state"))
    ]
    if desconocidas or (len(entradas) > 1 and len(salidas) > 1):
        return {
            "falta_entrada": False,
            "falta_salida": False,
            "entrada_real": ensure_mx(entradas[0]["check_time"]) if entradas else None,
            "salida_real": ensure_mx(salidas[-1]["check_time"]) if salidas else None,
            "bloqueado": True,
            "mensaje_bloqueo": "El dia tiene checadas BioTime ambiguas. Solicita revision de RRHH.",
        }

    entrada_real = ensure_mx(entradas[0]["check_time"]) if entradas else None
    salida_real = ensure_mx(salidas[-1]["check_time"]) if salidas else None
    return {
        "falta_entrada": entrada_real is None,
        "falta_salida": salida_real is None,
        "entrada_real": entrada_real,
        "salida_real": salida_real,
        "bloqueado": False,
        "mensaje_bloqueo": None,
    }


def _validar_fechas_manual(
    *,
    fecha_laboral: date,
    fecha_entrada: date | None,
    fecha_salida: date | None,
    solicita_entrada: bool,
    solicita_salida: bool,
    entrada_tiempo: datetime | None,
    salida_tiempo: datetime | None,
    dias_retroactivo: int,
    max_horas: int,
    labor_window,
    huecos: dict,
    schedule: ScheduleConfig | None,
) -> None:
    hoy = today_mx()
    if fecha_laboral > hoy:
        raise ValueError("La fecha laboral no puede ser futura")
    if fecha_laboral < hoy - timedelta(days=dias_retroactivo):
        raise ValueError(f"Solo puedes solicitar registros de los ultimos {dias_retroactivo} dias")
    if solicita_entrada and fecha_entrada != fecha_laboral:
        raise ValueError("La fecha de entrada debe coincidir con la fecha laboral")
    if solicita_salida and fecha_salida not in {fecha_laboral, fecha_laboral + timedelta(days=1)}:
        raise ValueError("La fecha de salida debe ser la fecha laboral o el dia siguiente")

    _validar_tiempos_manual(
        solicita_entrada=solicita_entrada,
        solicita_salida=solicita_salida,
        entrada_tiempo=entrada_tiempo,
        salida_tiempo=salida_tiempo,
        max_horas=max_horas,
        labor_window=labor_window,
        huecos=huecos,
        schedule=schedule,
    )


def _validar_tiempos_manual(
    *,
    solicita_entrada: bool,
    solicita_salida: bool,
    entrada_tiempo: datetime | None,
    salida_tiempo: datetime | None,
    max_horas: int,
    labor_window,
    huecos: dict,
    schedule: ScheduleConfig | None,
) -> None:
    ahora = now_mx()
    if solicita_entrada:
        if not entrada_tiempo:
            raise ValueError("La fecha y hora de entrada son obligatorias")
        if entrada_tiempo > ahora:
            raise ValueError("La fecha y hora capturada no puede ser futura.")
        if not labor_window.start <= entrada_tiempo <= labor_window.end:
            raise ValueError("La entrada esta fuera de la ventana laboral esperada")

    if solicita_salida:
        if not salida_tiempo:
            raise ValueError("La fecha y hora de salida son obligatorias")
        if salida_tiempo > ahora:
            raise ValueError("La fecha y hora capturada no puede ser futura.")
        limite_salida = labor_window.end
        if not schedule or not schedule.es_laboral or not schedule.hora_entrada or not schedule.hora_salida:
            # build_labor_window corta en medianoche para dias de descanso/sin horario (para no
            # arrastrar la entrada del dia siguiente al bucketing automatico) -- aqui si extendemos
            # el limite con el margen de salida, igual que ya hace el recalculo automatico via
            # extender_salida_descanso_medianoche, para permitir capturar una salida real pasada medianoche.
            margen = schedule.margen_salida_despues_min if schedule else 0
            limite_salida = limite_salida + timedelta(minutes=margen)
        if not labor_window.start <= salida_tiempo <= limite_salida:
            raise ValueError("La salida esta fuera de la ventana laboral esperada")

    entrada_ref = entrada_tiempo if solicita_entrada else huecos.get("entrada_real")
    salida_ref = salida_tiempo if solicita_salida else huecos.get("salida_real")
    if entrada_ref and salida_ref:
        if salida_ref <= entrada_ref:
            raise ValueError("La hora de salida es anterior a la entrada. Revisa la fecha y hora correcta.")
        duracion_horas = (salida_ref - entrada_ref).total_seconds() / 3600
        if duracion_horas > max_horas:
            raise ValueError(f"La jornada no puede exceder {max_horas} horas")


def _validar_solicitud_vs_huecos(
    *,
    solicita_entrada: bool,
    solicita_salida: bool,
    huecos: dict,
) -> dict:
    if huecos.get("bloqueado"):
        raise ValueError(huecos.get("mensaje_bloqueo") or "El dia tiene checadas BioTime ambiguas")

    falta_entrada = bool(huecos.get("falta_entrada"))
    falta_salida = bool(huecos.get("falta_salida"))
    if not falta_entrada and not falta_salida:
        raise ValueError("El dia ya tiene entrada y salida registradas")
    if falta_entrada and not solicita_entrada:
        raise ValueError("La solicitud ya no coincide con los huecos actuales de BioTime.")
    if falta_salida and not solicita_salida:
        raise ValueError("La solicitud ya no coincide con los huecos actuales de BioTime.")

    insertar_entrada = solicita_entrada and falta_entrada
    insertar_salida = solicita_salida and falta_salida
    if not insertar_entrada and not insertar_salida:
        raise ValueError("La solicitud ya no coincide con los huecos actuales de BioTime.")
    return {"insertar_entrada": insertar_entrada, "insertar_salida": insertar_salida}


def _prefill_manual_times(fecha_laboral: date, schedule: ScheduleConfig | None) -> dict:
    fecha_salida = fecha_laboral
    hora_entrada = "08:00"
    hora_salida = "17:00"
    if schedule:
        if schedule.hora_entrada:
            hora_entrada = schedule.hora_entrada.strftime("%H:%M")
        if schedule.hora_salida:
            hora_salida = schedule.hora_salida.strftime("%H:%M")
        if schedule.cruza_medianoche or (
            schedule.hora_entrada and schedule.hora_salida and schedule.hora_salida < schedule.hora_entrada
        ):
            fecha_salida = fecha_laboral + timedelta(days=1)
    return {
        "fecha_entrada": fecha_laboral,
        "fecha_salida": fecha_salida,
        "hora_entrada": hora_entrada,
        "hora_salida": hora_salida,
    }


async def preparar_solicitud_manual_svc(conn, usuario_id: UUID, fecha_laboral: date) -> dict:
    dias_retroactivo, _ = await _get_config_manual_asistencia(conn)
    hoy = today_mx()
    base = {
        "fecha_laboral": fecha_laboral,
        "fecha_laboral_iso": fecha_laboral.isoformat(),
        "fecha_laboral_fmt": fecha_laboral.strftime("%d/%m/%Y"),
    }
    if fecha_laboral > hoy:
        return {**base, "bloqueado": True, "mensaje_bloqueo": "La fecha laboral no puede ser futura"}
    if fecha_laboral < hoy - timedelta(days=dias_retroactivo):
        return {
            **base,
            "bloqueado": True,
            "mensaje_bloqueo": f"Solo puedes solicitar registros de los ultimos {dias_retroactivo} dias",
        }

    deteccion = await _detectar_huecos_manual(conn, usuario_id, fecha_laboral)
    huecos = deteccion["huecos"]
    if huecos.get("bloqueado"):
        return {**base, "bloqueado": True, "mensaje_bloqueo": huecos["mensaje_bloqueo"]}
    if not huecos["falta_entrada"] and not huecos["falta_salida"]:
        return {**base, "bloqueado": True, "mensaje_bloqueo": "El dia ya tiene entrada y salida registradas"}

    return {
        **base,
        **_prefill_manual_times(fecha_laboral, deteccion["schedule"]),
        "bloqueado": False,
        "mensaje_bloqueo": None,
        "solicita_entrada": huecos["falta_entrada"],
        "solicita_salida": huecos["falta_salida"],
    }


async def crear_solicitud_manual_svc(conn, usuario_id: UUID, payload) -> dict:
    motivo = (payload.motivo or "").strip()
    if not motivo:
        raise ValueError("El motivo es obligatorio")

    existente = await db.get_solicitud_manual_existente_activa(conn, usuario_id, payload.fecha_laboral)
    if existente:
        raise ValueError("Ya existe una solicitud activa para esa fecha")

    dias_retroactivo, max_horas = await _get_config_manual_asistencia(conn)
    deteccion = await _detectar_huecos_manual(conn, usuario_id, payload.fecha_laboral)
    huecos = deteccion["huecos"]
    solicita_entrada = huecos["falta_entrada"]
    solicita_salida = huecos["falta_salida"]
    _validar_solicitud_vs_huecos(
        solicita_entrada=solicita_entrada,
        solicita_salida=solicita_salida,
        huecos=huecos,
    )

    entrada_tiempo = _parse_manual_datetime(payload.fecha_entrada, payload.hora_entrada) if solicita_entrada else None
    salida_tiempo = _parse_manual_datetime(payload.fecha_salida, payload.hora_salida) if solicita_salida else None
    _validar_fechas_manual(
        fecha_laboral=payload.fecha_laboral,
        fecha_entrada=payload.fecha_entrada,
        fecha_salida=payload.fecha_salida,
        solicita_entrada=solicita_entrada,
        solicita_salida=solicita_salida,
        entrada_tiempo=entrada_tiempo,
        salida_tiempo=salida_tiempo,
        dias_retroactivo=dias_retroactivo,
        max_horas=max_horas,
        labor_window=deteccion["labor_window"],
        huecos=huecos,
        schedule=deteccion["schedule"],
    )

    row = await db.insert_solicitud_manual(
        conn,
        usuario_id=usuario_id,
        fecha_laboral=payload.fecha_laboral,
        solicita_entrada=solicita_entrada,
        solicita_salida=solicita_salida,
        entrada_tiempo=entrada_tiempo,
        salida_tiempo=salida_tiempo,
        motivo=motivo,
    )
    return _format_manual_request(row)


async def aprobar_solicitud_manual_svc(
    conn,
    *,
    solicitud_id: UUID,
    aprobador_id: UUID,
    equipo_ids: list[UUID],
) -> dict:
    async with conn.transaction():
        solicitud = await db.get_solicitud_manual_for_update(conn, solicitud_id)
        if not solicitud:
            raise ValueError("Solicitud no encontrada")
        if solicitud["usuario_id"] not in equipo_ids:
            raise ValueError("Solicitud no encontrada")
        if solicitud["estado"] != "pendiente":
            raise ValueError("Esta solicitud ya fue procesada")

        _, max_horas = await _get_config_manual_asistencia(conn)
        deteccion = await _detectar_huecos_manual(conn, solicitud["usuario_id"], solicitud["fecha_laboral"])
        acciones = _validar_solicitud_vs_huecos(
            solicita_entrada=solicitud["solicita_entrada"],
            solicita_salida=solicitud["solicita_salida"],
            huecos=deteccion["huecos"],
        )
        _validar_tiempos_manual(
            solicita_entrada=acciones["insertar_entrada"],
            solicita_salida=acciones["insertar_salida"],
            entrada_tiempo=solicitud["entrada_tiempo"] if acciones["insertar_entrada"] else None,
            salida_tiempo=solicitud["salida_tiempo"] if acciones["insertar_salida"] else None,
            max_horas=max_horas,
            labor_window=deteccion["labor_window"],
            huecos=deteccion["huecos"],
            schedule=deteccion["schedule"],
        )

        emp_code = await db.get_biotime_emp_code_para_manual(conn, solicitud["usuario_id"])
        if not emp_code:
            raise ValueError("No hay codigo BioTime configurado para este usuario")

        check_entrada_id = None
        check_salida_id = None
        if acciones["insertar_entrada"]:
            check_entrada_id = await db.insert_manual_check(
                conn,
                usuario_id=solicitud["usuario_id"],
                biotime_emp_code=emp_code,
                check_time=solicitud["entrada_tiempo"],
                punch_state="0",
                solicitud_manual_id=solicitud_id,
            )
        if acciones["insertar_salida"]:
            check_salida_id = await db.insert_manual_check(
                conn,
                usuario_id=solicitud["usuario_id"],
                biotime_emp_code=emp_code,
                check_time=solicitud["salida_tiempo"],
                punch_state="1",
                solicitud_manual_id=solicitud_id,
            )

        await db.aprobar_solicitud_manual(
            conn,
            solicitud_id=solicitud_id,
            revisado_por=aprobador_id,
            check_entrada_id=check_entrada_id,
            check_salida_id=check_salida_id,
        )
        await recalcular_asistencia(conn, [(solicitud["usuario_id"], solicitud["fecha_laboral"])])

    return _format_manual_request({**solicitud, "estado": "aprobado"})


async def rechazar_solicitud_manual_svc(
    conn,
    *,
    solicitud_id: UUID,
    aprobador_id: UUID,
    equipo_ids: list[UUID],
    comentario: str,
) -> dict:
    comentario = (comentario or "").strip()
    if not comentario:
        raise ValueError("El comentario es obligatorio")

    async with conn.transaction():
        solicitud = await db.get_solicitud_manual_for_update(conn, solicitud_id)
        if not solicitud:
            raise ValueError("Solicitud no encontrada")
        if solicitud["usuario_id"] not in equipo_ids:
            raise ValueError("Solicitud no encontrada")
        if solicitud["estado"] != "pendiente":
            raise ValueError("Esta solicitud ya fue procesada")
        await db.rechazar_solicitud_manual(
            conn,
            solicitud_id=solicitud_id,
            revisado_por=aprobador_id,
            comentario_revision=comentario,
        )

    return _format_manual_request({**solicitud, "estado": "rechazado", "comentario_revision": comentario})


async def sync_biotime_periodically() -> None:
    logger.info("[BIOTIME_SYNC] Tarea inicializada")
    interval_seconds = 900
    while True:
        try:
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                reaped = await db.reap_stale_sync_runs(conn)
                if reaped:
                    logger.warning("[BIOTIME_SYNC] %s run(s) atascado(s) en 'running' marcados como error", reaped)
                interval_seconds = await ConfigService.get_global_config(
                    conn, BIOTIME_CONFIG_KEYS["interval_seconds"], 900, int
                )
                result = await sync_biotime_once(conn)
                if result.get("status") == "disabled":
                    logger.debug("[BIOTIME_SYNC] Sincronizacion desactivada")
                else:
                    logger.info("[BIOTIME_SYNC] Resultado: %s", result)
        except asyncpg.PostgresError as exc:
            logger.error("[BIOTIME_SYNC] Error de BD: %s", exc)
        except BIOTIME_CONNECTION_ERRORS as exc:
            logger.error("[BIOTIME_SYNC] Error de conexion (post-freeze?): %s: %s", type(exc).__name__, exc)
        except httpx.HTTPError as exc:
            logger.error("[BIOTIME_SYNC] Error HTTP BioTime: %s: %s", type(exc).__name__, exc or "(sin mensaje)")
        except (ValueError, TypeError, RuntimeError) as exc:
            logger.error("[BIOTIME_SYNC] Error de sincronizacion: %s", exc)

        await asyncio.sleep(max(60, interval_seconds))
