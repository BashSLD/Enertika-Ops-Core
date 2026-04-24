# modules/calculadora_polizas/router.py
"""
Router del Módulo Calculadora Pólizas.

Endpoints:
- GET  /calculadora-polizas/ui                          — Dashboard + calculadora
- GET  /calculadora-polizas/api/plantas                 — JSON dropdown
- POST /calculadora-polizas/api/calcular                — HTMX → resultado.html
- GET  /oym/polizas/configuracion/ui                    — Editar precios/costos (manager+)
- POST /calculadora-polizas/admin/precios-zona
- PATCH /calculadora-polizas/admin/precios-zona/{zona}
- PATCH /calculadora-polizas/admin/wattabit/{id}
- PATCH /calculadora-polizas/admin/costos-fijos/{concepto}

Rutas de gestión OyM (prefijo /oym):
- POST /oym/cotizaciones/guardar
- GET  /oym/cotizaciones/ui
- PATCH /oym/cotizaciones/{id}/estatus
- GET  /oym/cotizaciones/{id}/asignar-modal
- PATCH /oym/cotizaciones/{id}/asignar
- GET  /oym/cotizaciones/{id}/editar-modal
- GET  /oym/cotizaciones/{id}/comparar-precios
- GET  /oym/cotizaciones/{id}/renovar-modal
- GET  /oym/cotizaciones/{id}/aceptar-modal
- GET  /oym/cotizaciones/{id}/info-modal
- GET  /oym/cotizaciones/{id}/cambiar-estatus-resumen-modal
- PATCH /oym/cotizaciones/{id}/estatus-resumen
- PUT  /oym/cotizaciones/{id}
- GET  /oym/partials/polizas-resumen
- GET  /oym/plantas/nueva-modal
- GET  /oym/plantas/ui
- POST /oym/plantas/import-excel
- POST /oym/plantas
- GET  /oym/plantas/{id}/editar-modal
- PUT  /oym/plantas/{id}
- POST /oym/plantas/{id}/toggle
"""

from fastapi import APIRouter, Depends, Request, Form, File, UploadFile, Query, HTTPException
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, Response
from typing import Optional
from uuid import UUID
import json
import logging
import re
from datetime import datetime, date
import pytz
from core.timezone import today_mx

from core.database import get_db_connection
from core.security import get_current_user_context
from core.permissions import require_module_access, require_manager_access, user_has_module_access
from core.config import settings
from core.config_service import ConfigService

from .service import CalculadoraService, get_service, tiene_garantia_produccion
from .schemas import CalcularRequest, EstatusCotizacion
from core.pdf_service.service import get_pdf_service, PDFService

logger = logging.getLogger("CalculadoraPolizas.Router")
templates = Jinja2Templates(directory="templates")
templates.env.globals["DEBUG_MODE"] = settings.DEBUG_MODE

from core.jinja_filters import register_timezone_filters
register_timezone_filters(templates.env)

router = APIRouter(prefix="/calculadora-polizas", tags=["Modulo Calculadora Polizas"])
oym_router = APIRouter(prefix="/oym", tags=["Modulo O&M - Polizas y Plantas"])

SLUG = "oym"          # sub-herramienta de O&M — hereda permisos del módulo oym
TPL = "calculadora_polizas"

_MESES_ES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre",
}

_ZONAS_INCIDENCIA_VALIDAS = {"Zona 1", "Zona 2"}
_PATTERN_PLANTA_MX = re.compile(r"^(MX-50\d{3})(?:-(\d{2}))?$")
_PATTERN_INCIDENCIA = re.compile(r"^MX-800\d{2,}$")


def _normalize_zona_incidencia(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    valor = value.strip()
    if not valor:
        return None
    if valor not in _ZONAS_INCIDENCIA_VALIDAS:
        raise ValueError("Zona Operativa / Incidencias no válida. Usa: Zona 1 o Zona 2")
    return valor


def _normalize_zona_catalogo(value: str) -> str:
    valor = re.sub(r"\s+", " ", (value or "").strip())
    if not valor:
        raise ValueError("El nombre de zona es obligatorio")
    if len(valor) > 100:
        raise ValueError("El nombre de zona no puede exceder 100 caracteres")
    return valor


def _to_bool_flag(value: Optional[str]) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "si", "on"}


def _normalize_id_incidencia(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    normalized = value.strip().upper()
    return normalized or None


def _build_warning_message_from_ids(rows: list) -> str:
    if not rows:
        return ""
    labels = [f"{row['id']} ({row['nombre']})" for row in rows[:5]]
    suffix = "" if len(rows) <= 5 else f" y {len(rows) - 5} más"
    return ", ".join(labels) + suffix


async def _build_planta_warnings(
    conn,
    service: CalculadoraService,
    planta_id: str,
    id_incidencia: Optional[str],
    *,
    exclude_id: Optional[str] = None,
    check_existing_exact_id: bool = False,
) -> list:
    warnings = []
    planta_id_clean = (planta_id or "").strip().upper()
    match = _PATTERN_PLANTA_MX.fullmatch(planta_id_clean)

    if not match:
        warnings.append(
            "ID de planta con formato no estándar. Se recomienda MX-50### o MX-50###-NN (subproyecto)."
        )
        mx_raw = re.match(r"^MX-50(\d+)", planta_id_clean)
        if mx_raw and len(mx_raw.group(1)) != 3:
            warnings.append(
                f"Posible inconsistencia en longitud del ID ({planta_id_clean}). Después de MX-50 deben venir 3 dígitos."
            )
    else:
        base_id = match.group(1)
        related = await service.db.get_plantas_by_mx_base(conn, base_id, exclude_id)
        if related:
            warnings.append(
                "Ya existen plantas relacionadas al mismo ID base "
                f"{base_id}: {_build_warning_message_from_ids(related)}"
            )

    if check_existing_exact_id:
        existing = await service.db.get_planta_by_id(conn, planta_id_clean)
        if existing:
            warnings.append(
                f"El ID {planta_id_clean} ya existe. Si confirmas, se actualizará la planta existente ({existing['nombre']})."
            )

    if id_incidencia:
        if not _PATTERN_INCIDENCIA.fullmatch(id_incidencia):
            warnings.append(
                "ID de incidencia con formato no estándar. Se recomienda MX-800## (o más dígitos)."
            )
        duplicated_incidencia = await service.db.get_plantas_by_id_incidencia(conn, id_incidencia, exclude_id)
        if duplicated_incidencia:
            warnings.append(
                "El ID de incidencia ya está asignado a otras plantas: "
                f"{_build_warning_message_from_ids(duplicated_incidencia)}"
            )

    return warnings


def _warning_confirmation_context(
    endpoint: str,
    method: str,
    form_values: dict,
    warnings: list,
    *,
    hx_target: str,
    hx_swap: str,
) -> dict:
    payload = {k: "" if v is None else str(v) for k, v in form_values.items()}
    payload["confirm_warnings"] = "true"
    return {
        "endpoint": endpoint,
        "method": method,
        "values": payload,
        "warnings": warnings,
        "hx_target": hx_target,
        "hx_swap": hx_swap,
    }


async def _build_admin_ctx(context: dict, conn, service: CalculadoraService, admin_notice: Optional[str] = None) -> dict:
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return {
        **_base_ctx(context, mod_role),
        "precios_zona": await service.db.get_precios_zona_list(conn),
        "wattabit": await service.db.get_wattabit_list(conn),
        "costos_fijos": await service.db.get_costos_fijos_list(conn),
        "vigencia_dias": await ConfigService.get_global_config(conn, "calc_poliza_vigencia_dias", 30, int),
        "admin_notice": admin_notice,
    }


def _base_ctx(context: dict, mod_role: str) -> dict:
    is_admin = context.get("role") == "ADMIN"
    return {
        "user_name": context.get("user_name"),
        "role": context.get("role"),
        "module_roles": context.get("module_roles", {}),
        "current_module_role": mod_role,
        "puede_editar": mod_role in ("editor", "admin") or is_admin,
        "puede_admin": mod_role == "admin" or is_admin or (
            context.get("role") == "MANAGER" and mod_role in ("editor", "admin")
        ),
    }


def _parse_cotizacion_id(cotizacion_id: str) -> UUID:
    try:
        return UUID(cotizacion_id)
    except ValueError:
        raise HTTPException(404, "Cotizacion no encontrada")


def _plantas_for_template(plantas_db: list) -> list:
    return [
        {
            "id": p["id"], "nombre": p["nombre"], "zona": p["zona"],
            "potencia_kw": float(p["potencia_kw"]) if p["potencia_kw"] else None,
            "num_paneles": p["num_paneles"],
        }
        for p in plantas_db
    ]


# ============================================================
# UI PRINCIPAL
# ============================================================

@router.api_route("/ui", methods=["GET", "HEAD"], include_in_schema=False)
async def get_calculadora_ui(
    request: Request,
    planta_id: Optional[str] = Query(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    plantas_db = await service.db.get_plantas_dropdown(conn)
    plantas = _plantas_for_template(plantas_db)
    costos = await service.db.get_costos_fijos(conn)

    ctx = {
        **_base_ctx(context, mod_role),
        "plantas": plantas,
        "utilidad_default": costos.get("utilidad_default", 0.30),
        "resultado": None,
        "planta_id_default": planta_id or "",
    }

    if request.headers.get("hx-request") and not request.headers.get("hx-history-restore-request"):
        return templates.TemplateResponse(request, f"{TPL}/partials/content.html", ctx)
    return templates.TemplateResponse(request, f"{TPL}/dashboard.html", ctx)


# ============================================================
# API: PLANTAS DROPDOWN (JSON)
# ============================================================

@router.get("/api/plantas", include_in_schema=False)
async def api_plantas(
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    plantas = await service.db.get_plantas_dropdown(conn)
    return JSONResponse(_plantas_for_template(plantas))


# ============================================================
# CALCULAR (HTMX)
# ============================================================

@router.post("/api/calcular", include_in_schema=False)
async def calcular(
    request: Request,
    planta_id: str = Form(...),
    tipo_poliza: str = Form(...),
    utilidad: float = Form(0.30),
    descuento_pct_1: Optional[float] = Form(None),
    descuento_pct_3: Optional[float] = Form(None),
    descuento_pct_5: Optional[float] = Form(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    try:
        req = CalcularRequest(
            planta_id=planta_id, tipo_poliza=tipo_poliza, utilidad=utilidad,
            descuento_pct_1=descuento_pct_1,
            descuento_pct_3=descuento_pct_3,
            descuento_pct_5=descuento_pct_5,
        )
        resultado = await service.calcular(conn, req)
    except ValueError as exc:
        return templates.TemplateResponse(
            request, f"{TPL}/partials/resultado.html",
            {**_base_ctx(context, mod_role), "resultado": None, "error": str(exc)},
            status_code=422,
        )

    return templates.TemplateResponse(
        request, f"{TPL}/partials/resultado.html",
        {**_base_ctx(context, mod_role), "resultado": resultado, "error": None},
    )


# ============================================================
# GUARDAR COTIZACIÓN
# ============================================================

@router.get("/partials/guardar-modal", include_in_schema=False)
async def guardar_modal(
    request: Request,
    planta_id: str = Query(...),
    tipo_poliza: str = Query(...),
    utilidad: float = Query(0.30),
    descuento_pct_1: Optional[str] = Query(None),
    descuento_pct_3: Optional[str] = Query(None),
    descuento_pct_5: Optional[str] = Query(None),
    fecha_fin_poliza_anterior: str = Query(""),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    def _pct(v: Optional[str]) -> Optional[float]:
        if not v:
            return None
        try:
            f = float(v)
            return f if f > 0 else None
        except (ValueError, TypeError):
            return None

    d1, d3, d5 = _pct(descuento_pct_1), _pct(descuento_pct_3), _pct(descuento_pct_5)
    usuarios_comercial = await service.db.get_usuarios_comercial(conn)
    vigencia_dias_default = await ConfigService.get_global_config(
        conn, "calc_poliza_vigencia_dias", 30, int
    )
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/guardar_cotizacion_modal.html",
        {
            **_base_ctx(context, mod_role),
            "planta_id": planta_id,
            "tipo_poliza": tipo_poliza,
            "utilidad": utilidad,
            "descuento_pct_1": d1,
            "descuento_pct_3": d3,
            "descuento_pct_5": d5,
            "usuarios_comercial": usuarios_comercial,
            "fecha_fin_poliza_anterior": fecha_fin_poliza_anterior,
            "vigencia_dias_default": vigencia_dias_default,
        },
    )


@oym_router.post("/cotizaciones/guardar", include_in_schema=False)
async def guardar_cotizacion(
    request: Request,
    planta_id: str = Form(...),
    tipo_poliza: str = Form(...),
    utilidad: float = Form(0.30),
    descuento_pct_1: Optional[float] = Form(None),
    descuento_pct_3: Optional[float] = Form(None),
    descuento_pct_5: Optional[float] = Form(None),
    solicitante_id: Optional[str] = Form(None),
    fecha_inicio_poliza: Optional[str] = Form(None),
    fecha_fin_poliza: Optional[str] = Form(None),
    poliza_anterior_id: Optional[str] = Form(None),
    fecha_fin_poliza_anterior: Optional[str] = Form(None),
    vigencia_cotizacion_dias: int = Form(30),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    user_id = context.get("user_db_id")
    sol_id = None
    if solicitante_id and solicitante_id.strip():
        try:
            sol_id = UUID(solicitante_id)
        except ValueError:
            pass

    anterior_id = None
    if poliza_anterior_id and poliza_anterior_id.strip():
        try:
            anterior_id = UUID(poliza_anterior_id)
        except ValueError:
            pass

    def _parse_date(val: Optional[str]) -> Optional[date]:
        if not val or not val.strip():
            return None
        try:
            return date.fromisoformat(val.strip())
        except ValueError:
            return None

    if vigencia_cotizacion_dias < 1:
        vigencia_cotizacion_dias = 30

    try:
        req = CalcularRequest(
            planta_id=planta_id, tipo_poliza=tipo_poliza, utilidad=utilidad,
            descuento_pct_1=descuento_pct_1,
            descuento_pct_3=descuento_pct_3,
            descuento_pct_5=descuento_pct_5,
        )
        resultado = await service.calcular(conn, req)
        await service.guardar_cotizacion(
            conn, resultado, user_id, solicitante_id=sol_id,
            fecha_inicio_poliza=_parse_date(fecha_inicio_poliza),
            fecha_fin_poliza=_parse_date(fecha_fin_poliza),
            poliza_anterior_id=anterior_id,
            fecha_fin_poliza_anterior=_parse_date(fecha_fin_poliza_anterior),
            vigencia_cotizacion_dias=vigencia_cotizacion_dias,
        )
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Error", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )

    return templates.TemplateResponse(
        request, "shared/toast.html",
        {"type": "success", "title": "Guardado", "message": "Cotizacion guardada correctamente"},
        headers={"HX-Reswap": "none"},
    )


# ============================================================
# POLIZAS GENERADAS (cotizaciones)
# ============================================================

async def _build_cotizaciones_ctx(
    context, conn, service,
    limit: int,
    estatus_filter: Optional[str],
    planta_filter: Optional[str] = None,
    tipo_filter: Optional[str] = None,
    solicitante_id_filter: Optional[str] = None,
) -> dict:
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    cotizaciones = await service.db.get_cotizaciones(
        conn, limit=limit,
        estatus_filter=estatus_filter or None,
        planta_filter=planta_filter or None,
        tipo_filter=tipo_filter or None,
        solicitante_id_filter=solicitante_id_filter or None,
    )
    resumen = await service.db.get_resumen_estatus(conn)
    alertas = await service.db.get_alertas_vencimiento(conn)
    filter_options = await service.db.get_polizas_filter_options(conn)
    return {
        **_base_ctx(context, mod_role),
        "cotizaciones": cotizaciones,
        "limit": limit,
        "estatus_filter": estatus_filter or "",
        "planta_filter": planta_filter or "",
        "tipo_filter": tipo_filter or "",
        "solicitante_id_filter": solicitante_id_filter or "",
        "filter_options": filter_options,
        "resumen": resumen,
        "alertas": alertas,
        "fecha_hoy": today_mx(),
    }


@oym_router.get("/cotizaciones/ui", include_in_schema=False)
async def cotizaciones_ui(
    request: Request,
    limit: int = Query(15, ge=0),
    estatus_filter: Optional[str] = Query(None),
    planta_filter: Optional[str] = Query(None),
    tipo_filter: Optional[str] = Query(None),
    solicitante_id_filter: Optional[str] = Query(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    ctx = await _build_cotizaciones_ctx(
        context, conn, service, limit,
        estatus_filter, planta_filter, tipo_filter, solicitante_id_filter,
    )
    if request.headers.get("hx-request") and not request.headers.get("hx-history-restore-request"):
        return templates.TemplateResponse(request, f"{TPL}/partials/cotizaciones_tabla.html", ctx)
    return templates.TemplateResponse(request, f"{TPL}/cotizaciones.html", ctx)


@oym_router.patch("/cotizaciones/{cotizacion_id}/estatus", include_in_schema=False)
async def update_cotizacion_estatus(
    request: Request,
    cotizacion_id: str,
    estatus: EstatusCotizacion = Form(...),
    estatus_filter: str = Form(""),
    limit: int = Form(15),
    fecha_inicio_poliza: Optional[str] = Form(None),
    fecha_fin_poliza: Optional[str] = Form(None),
    anios_contratados: Optional[int] = Form(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    def _parse_date(val: Optional[str]) -> Optional[date]:
        if not val:
            return None
        try:
            return date.fromisoformat(val.strip())
        except ValueError:
            return None

    uid = _parse_cotizacion_id(cotizacion_id)
    user_id = context.get("user_db_id")
    rol_sistema = context.get("role", "USER")
    mod_role = context.get("module_roles", {}).get(SLUG, "viewer")

    try:
        await service.cambiar_estatus_cotizacion(
            conn, uid, estatus.value, user_id, rol_sistema, mod_role,
            fecha_inicio=_parse_date(fecha_inicio_poliza),
            fecha_fin=_parse_date(fecha_fin_poliza),
            anios_contratados=anios_contratados,
        )
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "No permitido", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )

    ctx = await _build_cotizaciones_ctx(context, conn, service, limit, estatus_filter or None)
    return templates.TemplateResponse(request, f"{TPL}/partials/cotizaciones_tabla.html", ctx)


# ============================================================
# RESUMEN EMBEBIDO (para tab en OyM)
# ============================================================

@oym_router.get("/partials/polizas-resumen", include_in_schema=False)
async def polizas_resumen(
    request: Request,
    limit: int = Query(15, ge=0),
    estatus_filter: Optional[str] = Query(None),
    planta_filter: Optional[str] = Query(None),
    tipo_filter: Optional[str] = Query(None),
    solicitante_id_filter: Optional[str] = Query(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    cotizaciones = await service.db.get_cotizaciones(
        conn, limit=limit,
        estatus_filter=estatus_filter or None,
        planta_filter=planta_filter or None,
        tipo_filter=tipo_filter or None,
        solicitante_id_filter=solicitante_id_filter or None,
    )
    resumen = await service.db.get_resumen_estatus(conn)
    alertas = await service.db.get_alertas_vencimiento(conn)
    filter_options = await service.db.get_polizas_filter_options(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/polizas_resumen.html",
        {
            **_base_ctx(context, mod_role),
            "cotizaciones": cotizaciones,
            "resumen": resumen,
            "alertas": alertas,
            "limit": limit,
            "estatus_filter": estatus_filter or "",
            "planta_filter": planta_filter or "",
            "tipo_filter": tipo_filter or "",
            "solicitante_id_filter": solicitante_id_filter or "",
            "filter_options": filter_options,
            "fecha_hoy": today_mx(),
        },
    )


# ============================================================
# CAMBIAR ESTATUS DESDE RESUMEN OYM (todos los estatus)
# ============================================================

@oym_router.get("/cotizaciones/{cotizacion_id}/cambiar-estatus-resumen-modal", include_in_schema=False)
async def cambiar_estatus_resumen_modal(
    request: Request,
    cotizacion_id: str,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/cambiar_estatus_resumen_modal.html",
        {
            **_base_ctx(context, mod_role),
            "cotizacion": cotizacion,
        },
    )


@oym_router.patch("/cotizaciones/{cotizacion_id}/estatus-resumen", include_in_schema=False)
async def update_estatus_resumen(
    request: Request,
    cotizacion_id: str,
    estatus: EstatusCotizacion = Form(...),
    motivo_cancelacion: Optional[str] = Form(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    user_id = context.get("user_db_id")
    rol_sistema = context.get("role", "USER")
    mod_role = context.get("module_roles", {}).get(SLUG, "viewer")

    try:
        await service.cambiar_estatus_cotizacion(
            conn, uid, estatus.value, user_id, rol_sistema, mod_role,
            motivo=motivo_cancelacion,
        )
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "No permitido", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )

    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    cotizaciones = await service.db.get_cotizaciones(conn, limit=15)
    resumen = await service.db.get_resumen_estatus(conn)
    alertas = await service.db.get_alertas_vencimiento(conn)
    filter_options = await service.db.get_polizas_filter_options(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/polizas_resumen.html",
        {
            **_base_ctx(context, mod_role),
            "cotizaciones": cotizaciones,
            "resumen": resumen,
            "alertas": alertas,
            "limit": 15,
            "estatus_filter": "",
            "planta_filter": "",
            "tipo_filter": "",
            "solicitante_id_filter": "",
            "filter_options": filter_options,
            "fecha_hoy": today_mx(),
        },
    )


# ============================================================
# PÓLIZAS PARA MÓDULO COMERCIAL
# ============================================================

@oym_router.get("/partials/polizas-comercial", include_in_schema=False)
async def polizas_comercial(
    request: Request,
    page: int = Query(1, ge=1),
    estatus_filter: Optional[str] = Query(None),
    context=Depends(get_current_user_context),
    _=require_module_access("comercial"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    role = context.get("role", "USER")
    user_id = context.get("user_db_id")
    es_admin_o_manager = role in ("ADMIN", "MANAGER")
    es_admin_modulo = user_has_module_access("comercial", context, "admin")
    ver_todas = es_admin_o_manager or es_admin_modulo

    per_page = 50
    offset = (page - 1) * per_page
    ef = estatus_filter or None

    cotizaciones = await service.db.get_cotizaciones_comercial(
        conn, limit=per_page, offset=offset,
        ver_todas=ver_todas, user_id=user_id, estatus_filter=ef,
    )
    total = await service.db.count_cotizaciones_comercial(
        conn, ver_todas=ver_todas, user_id=user_id, estatus_filter=ef,
    )

    comercial_role = context.get("module_roles", {}).get("comercial", "viewer")
    return templates.TemplateResponse(
        request, "comercial/partials/polizas_tab.html",
        {
            **_base_ctx(context, comercial_role),
            "cotizaciones": cotizaciones,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max(1, -(-total // per_page)),
            "estatus_filter": estatus_filter or "",
            "ver_todas": ver_todas,
        },
    )


# ============================================================
# EDITAR COTIZACIÓN (editor+)
# ============================================================

@oym_router.get("/cotizaciones/{cotizacion_id}/editar-modal", include_in_schema=False)
async def editar_cotizacion_modal(
    request: Request,
    cotizacion_id: str,
    estatus_filter: str = Query(""),
    limit: int = Query(15, ge=0),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    plantas_db = await service.db.get_plantas_dropdown(conn)
    plantas = _plantas_for_template(plantas_db)
    usuarios_comercial = await service.db.get_usuarios_comercial(conn)

    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/editar_cotizacion_modal.html",
        {
            **_base_ctx(context, mod_role),
            "cotizacion": cotizacion,
            "plantas": plantas,
            "usuarios_comercial": usuarios_comercial,
            "estatus_filter": estatus_filter,
            "limit": limit,
        },
    )


@oym_router.get("/cotizaciones/{cotizacion_id}/comparar-precios", include_in_schema=False)
async def comparar_precios_cotizacion(
    cotizacion_id: str,
    planta_id: str = Query(...),
    tipo_poliza: str = Query(...),
    utilidad: float = Query(0.30),
    descuento_pct_1: Optional[float] = Query(None),
    descuento_pct_3: Optional[float] = Query(None),
    descuento_pct_5: Optional[float] = Query(None),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    try:
        req = CalcularRequest(
            planta_id=planta_id, tipo_poliza=tipo_poliza, utilidad=utilidad,
            descuento_pct_1=descuento_pct_1,
            descuento_pct_3=descuento_pct_3,
            descuento_pct_5=descuento_pct_5,
        )
        nuevo = await service.calcular(conn, req)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)

    snapshot = cotizacion.get("resultado_json") or {}
    if isinstance(snapshot, str):
        snapshot = json.loads(snapshot)

    snap_tipo = snapshot.get("tipo_poliza", tipo_poliza)
    snap_paneles = int(snapshot.get("num_paneles") or 0)
    snap_mtto_principal = float(snapshot.get("mtto_principal") or 0)
    snap_mtto_fijo = float(snapshot.get("mtto_fijo") or 0)
    snap_wattabit = float(snapshot.get("wattabit") or 0)
    snap_internet = float(snapshot.get("internet") or 0)
    snap_gestion = float(snapshot.get("gestion") or 0)

    nuevo_paneles = nuevo.num_paneles

    # Precios unitarios derivados del snapshot
    if snap_tipo == "premium" and snap_paneles > 0:
        snap_precio_panel = snap_mtto_principal / (snap_paneles * 2)
    else:
        snap_precio_panel = snap_mtto_principal  # costo fijo estandar
    snap_gestion_panel = snap_gestion / snap_paneles if snap_paneles > 0 else 0

    # Precios unitarios del nuevo calculo
    if nuevo.tipo_poliza == "premium" and nuevo_paneles > 0:
        nuevo_precio_panel = nuevo.mtto_principal / (nuevo_paneles * 2)
    else:
        nuevo_precio_panel = nuevo.mtto_principal
    nuevo_gestion_panel = nuevo.gestion / nuevo_paneles if nuevo_paneles > 0 else 0

    def _cambio(original: float, nuevo_val: float, etiqueta: str, unitario: bool = False):
        if abs(original - nuevo_val) > 0.01:
            return {"etiqueta": etiqueta, "original": original, "nuevo": nuevo_val, "unitario": unitario}
        return None

    precios_cambios = []

    # Comparar precio unitario solo si el tipo de poliza no cambio entre snapshot y nuevo calculo
    if snap_tipo == nuevo.tipo_poliza:
        if snap_tipo == "premium":
            d = _cambio(snap_precio_panel, nuevo_precio_panel,
                        f"Precio por panel — zona {nuevo.zona}", unitario=True)
            if d:
                precios_cambios.append(d)
            d = _cambio(snap_mtto_fijo, nuevo.mtto_fijo, "Mantenimiento correctivo fijo")
            if d:
                precios_cambios.append(d)
        else:
            d = _cambio(snap_precio_panel, nuevo_precio_panel, "Mantenimiento diagnostico estandar")
            if d:
                precios_cambios.append(d)

    d = _cambio(snap_wattabit, nuevo.wattabit, f"Wattabit — {nuevo.nombre_wattabit}")
    if d:
        precios_cambios.append(d)

    d = _cambio(snap_internet, nuevo.internet, "Internet anual")
    if d:
        precios_cambios.append(d)

    if snap_paneles > 0 and nuevo_paneles > 0:
        d = _cambio(snap_gestion_panel, nuevo_gestion_panel,
                    "Gestion energetica por panel", unitario=True)
        if d:
            precios_cambios.append(d)

    # Montos resultantes
    montos_cambios = []
    for campo, etiqueta in [
        ("sub_total", "Sub total"),
        ("sub_total_utilidad", "Sub total + utilidad"),
        ("total_final", "Total final (con IVA)"),
    ]:
        d = _cambio(float(snapshot.get(campo) or 0), float(getattr(nuevo, campo)), etiqueta)
        if d:
            montos_cambios.append(d)

    tiene_cambios = bool(precios_cambios or montos_cambios)
    return JSONResponse({
        "tiene_cambios": tiene_cambios,
        "precios_cambios": precios_cambios,
        "montos_cambios": montos_cambios,
    })


@oym_router.put("/cotizaciones/{cotizacion_id}", include_in_schema=False)
async def update_cotizacion(
    request: Request,
    cotizacion_id: str,
    planta_id: str = Form(...),
    tipo_poliza: str = Form(...),
    utilidad: float = Form(0.30),
    descuento_pct_1: Optional[float] = Form(None),
    descuento_pct_3: Optional[float] = Form(None),
    descuento_pct_5: Optional[float] = Form(None),
    solicitante_id: Optional[str] = Form(None),
    fecha_inicio_poliza: Optional[str] = Form(None),
    fecha_fin_poliza: Optional[str] = Form(None),
    poliza_anterior_id: Optional[str] = Form(None),
    fecha_fin_poliza_anterior: Optional[str] = Form(None),
    usar_snapshot: str = Form(""),
    estatus_filter: str = Form(""),
    limit: int = Form(15),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)

    sol_id = None
    if solicitante_id and solicitante_id.strip():
        try:
            sol_id = UUID(solicitante_id)
        except ValueError:
            pass

    anterior_id = None
    if poliza_anterior_id and poliza_anterior_id.strip():
        try:
            anterior_id = UUID(poliza_anterior_id)
        except ValueError:
            pass

    def _parse_date(val: Optional[str]) -> Optional[date]:
        if not val or not val.strip():
            return None
        try:
            return date.fromisoformat(val.strip())
        except ValueError:
            return None

    if usar_snapshot == "mantener":
        cotizacion_actual = await service.db.get_cotizacion_by_id(conn, uid)
        if not cotizacion_actual:
            raise HTTPException(404, "Cotizacion no encontrada")

        snap_json = cotizacion_actual.get("resultado_json") or {}
        if isinstance(snap_json, str):
            snap_json = json.loads(snap_json)

        nombre_planta = cotizacion_actual["nombre_planta"]
        if planta_id != str(cotizacion_actual.get("planta_id") or ""):
            planta_info = await service.db.get_planta_by_id(conn, planta_id)
            if planta_info:
                nombre_planta = planta_info["nombre"]

        # Precios del catálogo congelados en el snapshot; solo se recalculan
        # los montos dependientes de utilidad/descuentos que el usuario puede cambiar.
        costos = await service.db.get_costos_fijos(conn)
        iva = costos.get("iva", 0.16)
        factor = costos.get("incremento_anual", 0.03)

        snap_sub_total = float(snap_json.get("sub_total", cotizacion_actual["sub_total"]))
        snap_sub_total_utilidad = round(snap_sub_total / (1.0 - utilidad), 2)
        snap_total_final = round(snap_sub_total_utilidad * (1.0 + iva), 2)

        # Proyección a 5 años con la nueva utilidad (mismo factor de crecimiento que el original)
        _a1 = round(snap_sub_total_utilidad, 2)
        _a2 = round(_a1 * (1 + factor), 2)
        _a3 = round(_a1 * (1 + factor) ** 2, 2)
        _a4 = round(_a1 * (1 + factor) ** 3, 2)
        _a5 = round(_a1 * (1 + factor) ** 4, 2)
        _acum_3 = round(_a1 + _a2 + _a3, 2)
        _acum_5 = round(_a1 + _a2 + _a3 + _a4 + _a5, 2)

        def _apply_dto(base: float, pct: Optional[float]) -> float:
            return round(base * (1.0 - pct), 2) if pct else base

        snap_json["utilidad"] = utilidad
        snap_json["sub_total_utilidad"] = snap_sub_total_utilidad
        snap_json["total_final"] = snap_total_final
        snap_json["anio_1"] = _a1
        snap_json["anio_3"] = _a3
        snap_json["anio_5"] = _a5
        snap_json["acumulado_1_3"] = _acum_3
        snap_json["acumulado_1_5"] = _acum_5
        snap_json["descuento_pct_1"] = descuento_pct_1
        snap_json["descuento_pct_3"] = descuento_pct_3
        snap_json["descuento_pct_5"] = descuento_pct_5
        snap_json["anio_1_desc"] = _apply_dto(_a1, descuento_pct_1)
        snap_json["anio_3_desc"] = _apply_dto(_a3, descuento_pct_3)
        snap_json["anio_5_desc"] = _apply_dto(_a5, descuento_pct_5)
        snap_json["acumulado_1_3_desc"] = _apply_dto(_acum_3, descuento_pct_3)
        snap_json["acumulado_1_5_desc"] = _apply_dto(_acum_5, descuento_pct_5)

        ok = await service.db.update_cotizacion_full(conn, uid, {
            "planta_id": planta_id,
            "nombre_planta": nombre_planta,
            "tipo_poliza": tipo_poliza,
            "utilidad": utilidad,
            "sub_total": snap_sub_total,
            "sub_total_utilidad": snap_sub_total_utilidad,
            "total_final": snap_total_final,
            "resultado_json": snap_json,
            "solicitante_id": sol_id,
            "descuento_pct": None,
            "descuento_anios": None,
            "descuento_pct_1": descuento_pct_1,
            "descuento_pct_3": descuento_pct_3,
            "descuento_pct_5": descuento_pct_5,
            "fecha_inicio_poliza": _parse_date(fecha_inicio_poliza),
            "fecha_fin_poliza": _parse_date(fecha_fin_poliza),
            "poliza_anterior_id": anterior_id,
            "fecha_fin_poliza_anterior": _parse_date(fecha_fin_poliza_anterior),
        })
    else:
        try:
            req = CalcularRequest(
                planta_id=planta_id, tipo_poliza=tipo_poliza, utilidad=utilidad,
                descuento_pct_1=descuento_pct_1,
                descuento_pct_3=descuento_pct_3,
                descuento_pct_5=descuento_pct_5,
            )
            resultado = await service.calcular(conn, req)
        except ValueError as exc:
            return templates.TemplateResponse(
                request, "shared/toast.html",
                {"type": "error", "title": "Error al recalcular", "message": str(exc)},
                headers={"HX-Reswap": "none"},
            )

        ok = await service.db.update_cotizacion_full(conn, uid, {
            "planta_id": resultado.planta_id,
            "nombre_planta": resultado.nombre_planta,
            "tipo_poliza": resultado.tipo_poliza,
            "utilidad": resultado.utilidad,
            "sub_total": resultado.sub_total,
            "sub_total_utilidad": resultado.sub_total_utilidad,
            "total_final": resultado.total_final,
            "resultado_json": resultado.model_dump(),
            "solicitante_id": sol_id,
            "descuento_pct": None,
            "descuento_anios": None,
            "descuento_pct_1": resultado.descuento_pct_1,
            "descuento_pct_3": resultado.descuento_pct_3,
            "descuento_pct_5": resultado.descuento_pct_5,
            "fecha_inicio_poliza": _parse_date(fecha_inicio_poliza),
            "fecha_fin_poliza": _parse_date(fecha_fin_poliza),
            "poliza_anterior_id": anterior_id,
            "fecha_fin_poliza_anterior": _parse_date(fecha_fin_poliza_anterior),
        })

    if not ok:
        raise HTTPException(404, "Cotizacion no encontrada")

    ctx = await _build_cotizaciones_ctx(context, conn, service, limit, estatus_filter or None)
    return templates.TemplateResponse(request, f"{TPL}/partials/cotizaciones_tabla.html", ctx)


# ============================================================
# DECISION ACEPTADA/RECHAZADA (desde Comercial, editor+)
# ============================================================

@oym_router.get("/cotizaciones/{cotizacion_id}/asignar-modal", include_in_schema=False)
async def asignar_modal(
    request: Request,
    cotizacion_id: str,
    page: int = Query(1, ge=1),
    estatus_filter: str = Query(""),
    context=Depends(get_current_user_context),
    _=require_module_access("comercial", "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    comercial_role = context.get("module_roles", {}).get("comercial", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/asignar_cotizacion_modal.html",
        {
            **_base_ctx(context, comercial_role),
            "cotizacion": cotizacion,
            "page": page,
            "estatus_filter": estatus_filter,
        },
    )


@oym_router.patch("/cotizaciones/{cotizacion_id}/asignar", include_in_schema=False)
async def asignar_cotizacion(
    request: Request,
    cotizacion_id: str,
    estatus: str = Form(...),
    page: int = Form(1),
    estatus_filter: str = Form(""),
    context=Depends(get_current_user_context),
    _=require_module_access("comercial", "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    from core.workflow.notification_service import NotificationService

    if estatus not in {"ACEPTADA", "RECHAZADA"}:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Error", "message": "Solo se permite Aceptada o Rechazada"},
            headers={"HX-Reswap": "none"},
        )

    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    user_id = context.get("user_db_id")
    rol_sistema = context.get("role", "USER")
    mod_role = context.get("module_roles", {}).get("comercial", "viewer")

    try:
        await service.cambiar_estatus_cotizacion(
            conn, uid, estatus, user_id, rol_sistema, mod_role,
        )
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "No permitido", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )

    # Notificar al creador por email (fire-and-forget)
    cotizacion_actualizada = {**cotizacion, "estatus": estatus}
    try:
        notif = NotificationService()
        await notif.notify_poliza_estatus_change(
            conn=conn,
            cotizacion_id=uid,
            cotizacion=cotizacion_actualizada,
            nuevo_estatus=estatus,
            changed_by_ctx=context,
        )
    except Exception as e:
        logger.error(f"[ASIGNAR] Error al notificar poliza {uid}: {e}", exc_info=True)

    # Reconstruir el listado del tab de Comercial con los mismos filtros activos
    role = context.get("role", "USER")
    es_admin_o_manager = role in ("ADMIN", "MANAGER")
    es_admin_modulo = user_has_module_access("comercial", context, "admin")
    ver_todas = es_admin_o_manager or es_admin_modulo

    per_page = 50
    offset = (page - 1) * per_page
    ef = estatus_filter or None

    cotizaciones = await service.db.get_cotizaciones_comercial(
        conn, limit=per_page, offset=offset,
        ver_todas=ver_todas, user_id=user_id, estatus_filter=ef,
    )
    total = await service.db.count_cotizaciones_comercial(
        conn, ver_todas=ver_todas, user_id=user_id, estatus_filter=ef,
    )

    comercial_role = context.get("module_roles", {}).get("comercial", "viewer")
    return templates.TemplateResponse(
        request, "comercial/partials/polizas_tab.html",
        {
            "user_name": context.get("user_name"),
            "role": role,
            "module_roles": context.get("module_roles", {}),
            "current_module_role": comercial_role,
            "cotizaciones": cotizaciones,
            "total": total,
            "page": page,
            "per_page": per_page,
            "pages": max(1, -(-total // per_page)),
            "estatus_filter": estatus_filter,
            "ver_todas": ver_todas,
        },
    )


# ============================================================
# PLANTAS — CRUD (editor+)
# ============================================================

@oym_router.get("/plantas/nueva-modal", include_in_schema=False)
async def nueva_planta_modal(
    request: Request,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    precios_zona = await service.db.get_precios_zona(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/nueva_planta_modal.html",
        {**_base_ctx(context, mod_role), "zonas": sorted(precios_zona.keys())},
    )


@oym_router.get("/plantas/ui", include_in_schema=False)
async def plantas_ui(
    request: Request,
    q: Optional[str] = Query(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    plantas = await service.db.get_plantas_list(conn, q)
    precios_zona = await service.db.get_precios_zona(conn)

    ctx = {
        **_base_ctx(context, mod_role),
        "plantas": plantas,
        "zonas": sorted(precios_zona.keys()),
        "q": q or "",
        "today": today_mx(),
    }

    if request.headers.get("hx-request") and not request.headers.get("hx-history-restore-request"):
        return templates.TemplateResponse(request, f"{TPL}/partials/plantas_tabla.html", ctx)
    return templates.TemplateResponse(request, f"{TPL}/plantas.html", ctx)


@oym_router.get("/plantas/{planta_id}/editar-modal", include_in_schema=False)
async def editar_planta_modal(
    request: Request,
    planta_id: str,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    planta = await service.db.get_planta_by_id(conn, planta_id)
    if not planta:
        raise HTTPException(404, "Planta no encontrada")
    precios_zona = await service.db.get_precios_zona(conn)
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/editar_planta_modal.html",
        {
            **_base_ctx(context, mod_role),
            "planta": planta,
            "zonas": sorted(precios_zona.keys()),
        },
    )


@oym_router.put("/plantas/{planta_id}", include_in_schema=False)
async def actualizar_planta(
    request: Request,
    planta_id: str,
    nombre: str = Form(...),
    zona: str = Form(...),
    potencia_kw: Optional[float] = Form(None),
    num_paneles: Optional[int] = Form(None),
    cliente: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    es_externa: Optional[str] = Form(None),
    zona_incidencia: Optional[str] = Form(None),
    id_incidencia: Optional[str] = Form(None),
    confirm_warnings: Optional[str] = Form(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    planta = await service.db.get_planta_by_id(conn, planta_id)
    if not planta:
        raise HTTPException(404, "Planta no encontrada")
    try:
        nombre_clean = nombre.strip()
        zona_clean = zona.strip()
        cliente_clean = cliente.strip() if cliente else None
        direccion_clean = direccion.strip() if direccion else None
        zona_incidencia_clean = _normalize_zona_incidencia(zona_incidencia)
        id_incidencia_clean = _normalize_id_incidencia(id_incidencia)

        warnings = await _build_planta_warnings(
            conn,
            service,
            planta_id,
            id_incidencia_clean,
            exclude_id=planta_id,
            check_existing_exact_id=False,
        )

        if warnings and not _to_bool_flag(confirm_warnings):
            return templates.TemplateResponse(
                request,
                f"{TPL}/partials/plantas_warning_confirm_oob.html",
                _warning_confirmation_context(
                    endpoint=f"/oym/plantas/{planta_id}",
                    method="put",
                    form_values={
                        "nombre": nombre_clean,
                        "zona": zona_clean,
                        "potencia_kw": potencia_kw,
                        "num_paneles": num_paneles,
                        "cliente": cliente_clean,
                        "direccion": direccion_clean,
                        "es_externa": "true" if es_externa == "true" else "",
                        "zona_incidencia": zona_incidencia_clean,
                        "id_incidencia": id_incidencia_clean,
                    },
                    warnings=warnings,
                    hx_target="#main-content",
                    hx_swap="innerHTML",
                ),
                headers={"HX-Reswap": "none"},
            )

        await service.db.upsert_planta(conn, {
            "id": planta_id,
            "nombre": nombre_clean,
            "zona": zona_clean,
            "potencia_kw": potencia_kw,
            "num_paneles": num_paneles,
            "cliente": cliente_clean,
            "direccion": direccion_clean,
            "es_externa": es_externa == "true",
            "activa": planta["activa"],
            "id_proyecto": planta.get("id_proyecto"),
            "zona_incidencia": zona_incidencia_clean,
            "id_incidencia": id_incidencia_clean,
        })
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Validación", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )
    except Exception as exc:
        logger.error("Error actualizando planta %s: %s", planta_id, exc)
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Error", "message": "Error al actualizar la planta"},
            headers={"HX-Reswap": "none"},
        )

    plantas = await service.db.get_plantas_list(conn)
    precios_zona = await service.db.get_precios_zona(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/plantas_tabla.html",
        {**_base_ctx(context, mod_role), "plantas": plantas,
         "zonas": sorted(precios_zona.keys()), "q": "", "today": today_mx()},
        headers={"HX-Trigger": "plantaSaved"},
    )


@oym_router.post("/plantas", include_in_schema=False)
async def crear_planta(
    request: Request,
    id: str = Form(...),
    nombre: str = Form(...),
    zona: str = Form(...),
    potencia_kw: Optional[float] = Form(None),
    num_paneles: Optional[int] = Form(None),
    cliente: Optional[str] = Form(None),
    direccion: Optional[str] = Form(None),
    es_externa: Optional[str] = Form(None),
    zona_incidencia: Optional[str] = Form(None),
    next: Optional[str] = Form(None),
    confirm_warnings: Optional[str] = Form(None),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    try:
        id_clean = id.strip().upper()
        nombre_clean = nombre.strip()
        zona_clean = zona.strip()
        cliente_clean = cliente.strip() if cliente else None
        direccion_clean = direccion.strip() if direccion else None
        zona_incidencia_clean = _normalize_zona_incidencia(zona_incidencia)

        warnings = await _build_planta_warnings(
            conn,
            service,
            id_clean,
            None,
            exclude_id=None,
            check_existing_exact_id=True,
        )

        if warnings and not _to_bool_flag(confirm_warnings):
            hx_target = "#modal-nueva-planta-oym" if next == "oym" else "#main-content"
            hx_swap = "outerHTML" if next == "oym" else "innerHTML"
            return templates.TemplateResponse(
                request,
                f"{TPL}/partials/plantas_warning_confirm_oob.html",
                _warning_confirmation_context(
                    endpoint="/oym/plantas",
                    method="post",
                    form_values={
                        "id": id_clean,
                        "nombre": nombre_clean,
                        "zona": zona_clean,
                        "potencia_kw": potencia_kw,
                        "num_paneles": num_paneles,
                        "cliente": cliente_clean,
                        "direccion": direccion_clean,
                        "es_externa": "true" if es_externa == "true" else "",
                        "zona_incidencia": zona_incidencia_clean,
                        "next": next,
                    },
                    warnings=warnings,
                    hx_target=hx_target,
                    hx_swap=hx_swap,
                ),
                headers={"HX-Reswap": "none"},
            )

        await service.db.upsert_planta(conn, {
            "id": id_clean,
            "nombre": nombre_clean,
            "zona": zona_clean,
            "potencia_kw": potencia_kw,
            "num_paneles": num_paneles,
            "cliente": cliente_clean,
            "direccion": direccion_clean,
            "es_externa": es_externa == "true",
            "activa": True,
            "id_proyecto": None,
            "zona_incidencia": zona_incidencia_clean,
            "id_incidencia": None,
        })
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Validación", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )
    except Exception as exc:
        logger.error("Error creando planta: %s", exc)
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Error", "message": "Error al crear la planta"},
            headers={"HX-Reswap": "none"},
        )

    if next == "oym":
        return templates.TemplateResponse(
            request, f"{TPL}/partials/nueva_planta_oym_ok.html", {},
        )

    plantas = await service.db.get_plantas_list(conn)
    precios_zona = await service.db.get_precios_zona(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/plantas_tabla.html",
        {**_base_ctx(context, mod_role), "plantas": plantas,
         "zonas": sorted(precios_zona.keys()), "q": "", "today": today_mx()},
    )


@oym_router.post("/plantas/{planta_id}/toggle", include_in_schema=False)
async def toggle_planta(
    request: Request,
    planta_id: str,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    nuevo_estado = await service.db.toggle_planta_activa(conn, planta_id)
    if nuevo_estado is None:
        raise HTTPException(404, "Planta no encontrada")
    plantas = await service.db.get_plantas_list(conn)
    precios_zona = await service.db.get_precios_zona(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/plantas_tabla.html",
        {**_base_ctx(context, mod_role), "plantas": plantas,
         "zonas": sorted(precios_zona.keys()), "q": "", "today": today_mx()},
    )


@oym_router.get("/plantas/plantilla-excel", include_in_schema=False)
async def plantilla_excel(
    request: Request,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    """Genera y descarga un .xlsx con la plantilla de importación de plantas."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl no está instalado")

    from fastapi.responses import StreamingResponse
    import io as _io

    precios_zona = await service.db.get_precios_zona(conn)
    zonas_validas = sorted(precios_zona.keys())

    wb = openpyxl.Workbook()

    # ─── Hoja principal ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Plantas"

    headers = [
        "id", "nombre", "zona", "potencia_kw", "num_paneles",
        "cliente", "direccion", "es_externa",
        "zona_incidencia",
        "tipo_poliza", "fecha_inicio_poliza", "fecha_fin_poliza",
    ]
    notas = [
        "Ej: MX-01 (único, mayúsculas)", "Nombre completo de la planta", "Zona de precios (ver hoja Zonas)",
        "Potencia en kWp (decimal)", "Cantidad de paneles (entero)", "Nombre del cliente/empresa",
        "Dirección del sitio", "Escribe: SI si es externa, déjalo vacío si no",
        "Opcional: Zona 1 o Zona 2",
        "Opcional: premium o estandar (solo si ya tiene póliza activa)",
        "Opcional: fecha inicio póliza (AAAA-MM-DD)",
        "Opcional: fecha fin póliza (AAAA-MM-DD)",
    ]

    # Encabezado con estilo — columnas de póliza en color distinto
    header_fill      = PatternFill("solid", fgColor="1E3A5F")
    header_fill_pol  = PatternFill("solid", fgColor="065F46")   # verde oscuro para póliza
    header_font      = Font(color="FFFFFF", bold=True)
    for col, (h, _) in enumerate(zip(headers, notas), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill_pol if col > 8 else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fila de ejemplo (en gris claro)
    ejemplo = [
        "MX-01", "Planta Ejemplo SA", zonas_validas[0] if zonas_validas else "ZONA1",
        150.5, 300, "Cliente Ejemplo", "Calle Principal 123, Ciudad", "",
        "Zona 1",
        "premium", "2024-01-01", "2025-12-31",
    ]
    ex_fill = PatternFill("solid", fgColor="F2F2F2")
    for col, val in enumerate(ejemplo, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = ex_fill
        cell.font = Font(italic=True, color="888888")

    # Ajustar anchos de columnas
    anchos = [12, 35, 15, 14, 14, 25, 35, 14, 16, 14, 20, 20]
    for col, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

    # Nota informativa sobre columnas opcionales de póliza
    ws.cell(row=4, column=10, value="* Las columnas zona_incidencia, tipo_poliza, fecha_inicio_poliza y fecha_fin_poliza son opcionales.")
    ws.cell(row=4, column=10).font = Font(italic=True, color="065F46", size=9)
    ws.merge_cells(start_row=4, start_column=10, end_row=4, end_column=12)

    # ─── Hoja Zonas Válidas ────────────────────────────────────────
    ws2 = wb.create_sheet("Zonas Válidas")
    ws2.cell(row=1, column=1, value="Zona").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Precio por panel (MXP)").font = Font(bold=True)
    for i, zona in enumerate(zonas_validas, start=2):
        ws2.cell(row=i, column=1, value=zona)
        ws2.cell(row=i, column=2, value=precios_zona[zona])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 25

    # Serializar a bytes
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_plantas.xlsx"},
    )


@oym_router.post("/plantas/preview-excel", include_in_schema=False)
async def preview_excel(
    request: Request,
    archivo: UploadFile = File(...),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    """Valida el Excel fila por fila sin guardar. Retorna HTML con preview de la importación."""
    mod_role = context.get("module_roles", {}).get("oym", "viewer")

    if not archivo.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse(
            request, f"{TPL}/partials/import_preview.html",
            {
                **_base_ctx(context, mod_role),
                "filas": [],
                "errores_globales": ["Solo se aceptan archivos .xlsx o .xls"],
                "zonas_validas": [],
                "resumen": None,
            },
        )

    contenido = await archivo.read()
    resultado = await service.preview_plantas_excel(conn, contenido)

    filas = resultado["filas"]
    resumen = {
        "nuevas":       sum(1 for f in filas if f["estado"] == "nueva"),
        "actualizadas": sum(1 for f in filas if f["estado"] == "actualiza"),
        "errores":      sum(1 for f in filas if f["estado"] == "error"),
        "total":        len(filas),
        "con_poliza":   sum(1 for f in filas if f.get("tiene_poliza")),
    }

    return templates.TemplateResponse(
        request, f"{TPL}/partials/import_preview.html",
        {
            **_base_ctx(context, mod_role),
            "filas": filas,
            "errores_globales": resultado["errores_globales"],
            "zonas_validas": resultado["zonas_validas"],
            "resumen": resumen,
        },
    )


@oym_router.post("/plantas/import-excel", include_in_schema=False)
async def import_excel(
    request: Request,
    archivo: UploadFile = File(...),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    if not archivo.filename.endswith((".xlsx", ".xls")):
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Archivo invalido", "message": "Solo se aceptan archivos .xlsx"},
            headers={"HX-Reswap": "none"},
        )

    contenido = await archivo.read()
    user_id = context.get("user_id")
    try:
        resultado = await service.importar_plantas_excel(conn, contenido, user_id=user_id)
    except ValueError as exc:
        return templates.TemplateResponse(
            request, "shared/toast.html",
            {"type": "error", "title": "Error", "message": str(exc)},
            headers={"HX-Reswap": "none"},
        )

    msg = f"{resultado.insertadas} plantas nuevas, {resultado.actualizadas} actualizadas"
    if resultado.polizas_legacy:
        msg += f", {resultado.polizas_legacy} pólizas legacy creadas"
    if resultado.errores:
        msg += f". {len(resultado.errores)} con errores."
    toast_type = "success" if not resultado.errores else "warning"

    plantas = await service.db.get_plantas_list(conn)
    precios_zona = await service.db.get_precios_zona(conn)
    return templates.TemplateResponse(
        request, f"{TPL}/partials/plantas_tabla.html",
        {**_base_ctx(context, mod_role), "plantas": plantas,
         "zonas": sorted(precios_zona.keys()), "q": "",
         "today": today_mx(), "import_msg": msg, "import_type": toast_type},
    )


# ============================================================
# ADMIN — EDICIÓN DE PRECIOS/COSTOS (manager+)
# ============================================================

@oym_router.get("/polizas/configuracion/ui", include_in_schema=False)
async def admin_ui(
    request: Request,
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    ctx = await _build_admin_ctx(context, conn, service)
    if request.headers.get("hx-request") and not request.headers.get("hx-history-restore-request"):
        return templates.TemplateResponse(request, f"{TPL}/partials/admin_content.html", ctx)
    return templates.TemplateResponse(request, f"{TPL}/admin.html", ctx)


@router.patch("/admin/vigencia-cotizacion", include_in_schema=False)
async def update_vigencia_cotizacion(
    request: Request,
    dias: int = Form(...),
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    if dias < 1:
        raise ValueError("La vigencia debe ser al menos 1 dia")
    await conn.execute(
        "UPDATE tb_configuracion_global SET valor = $1 WHERE clave = 'calc_poliza_vigencia_dias'",
        str(dias),
    )
    ConfigService.invalidar_cache()
    ctx = await _build_admin_ctx(context, conn, service)
    ctx["vigencia_dias"] = dias
    return templates.TemplateResponse(request, f"{TPL}/partials/admin_content.html", ctx)


@router.post("/admin/precios-zona", include_in_schema=False)
async def create_precio_zona(
    request: Request,
    zona: str = Form(...),
    precio: float = Form(...),
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    if precio <= 0:
        raise HTTPException(400, "El precio debe ser mayor a 0")

    zona_normalizada = _normalize_zona_catalogo(zona)
    created = await service.db.create_precio_zona(conn, zona_normalizada, precio)
    if not created:
        return templates.TemplateResponse(
            request,
            "shared/toast.html",
            {
                "type": "warning",
                "title": "Zona duplicada",
                "message": f"La zona '{zona_normalizada}' ya existe en el catálogo",
            },
            headers={"HX-Reswap": "none"},
        )

    ctx = await _build_admin_ctx(
        context,
        conn,
        service,
        admin_notice=f"Zona '{zona_normalizada}' agregada correctamente",
    )
    return templates.TemplateResponse(request, f"{TPL}/partials/admin_content.html", ctx)


@router.patch("/admin/precios-zona/{zona}", include_in_schema=False)
async def update_precio_zona(
    request: Request,
    zona: str,
    precio: float = Form(...),
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    if precio <= 0:
        raise HTTPException(400, "El precio debe ser mayor a 0")
    ok = await service.db.update_precio_zona(conn, zona, precio)
    if not ok:
        raise HTTPException(404, f"Zona '{zona}' no encontrada")
    return templates.TemplateResponse(
        request, "shared/toast.html",
        {"type": "success", "title": "Actualizado", "message": f"Precio de {zona} actualizado"},
        headers={"HX-Reswap": "none"},
    )


@router.patch("/admin/wattabit/{wattabit_id}", include_in_schema=False)
async def update_wattabit(
    request: Request,
    wattabit_id: int,
    precio: float = Form(...),
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    if precio <= 0:
        raise HTTPException(400, "El precio debe ser mayor a 0")
    ok = await service.db.update_wattabit(conn, wattabit_id, precio)
    if not ok:
        raise HTTPException(404, "Registro Wattabit no encontrado")
    return templates.TemplateResponse(
        request, "shared/toast.html",
        {"type": "success", "title": "Actualizado", "message": "Precio Wattabit actualizado"},
        headers={"HX-Reswap": "none"},
    )


@router.patch("/admin/costos-fijos/{concepto}", include_in_schema=False)
async def update_costo_fijo(
    request: Request,
    concepto: str,
    valor: float = Form(...),
    notas: Optional[str] = Form(None),
    context=Depends(get_current_user_context),
    _=require_manager_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    if valor < 0:
        raise HTTPException(400, "El valor no puede ser negativo")
    ok = await service.db.update_costo_fijo(conn, concepto, valor, notas)
    if not ok:
        raise HTTPException(404, f"Concepto '{concepto}' no encontrado")
    return templates.TemplateResponse(
        request, "shared/toast.html",
        {"type": "success", "title": "Actualizado", "message": f"{concepto} actualizado"},
        headers={"HX-Reswap": "none"},
    )


# ============================================================
# RENOVAR PÓLIZA (editor+)
# ============================================================

@oym_router.get("/cotizaciones/{cotizacion_id}/renovar-modal", include_in_schema=False)
async def renovar_cotizacion_modal(
    request: Request,
    cotizacion_id: str,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    plantas_db = await service.db.get_plantas_dropdown(conn)
    plantas = _plantas_for_template(plantas_db)
    usuarios_comercial = await service.db.get_usuarios_comercial(conn)
    mod_role = context.get("module_roles", {}).get("oym", "viewer")

    return templates.TemplateResponse(
        request, f"{TPL}/partials/renovar_cotizacion_modal.html",
        {
            **_base_ctx(context, mod_role),
            "cotizacion": cotizacion,
            "plantas": plantas,
            "usuarios_comercial": usuarios_comercial,
            "fecha_hoy": today_mx().isoformat(),
        },
    )


@oym_router.get("/cotizaciones/{cotizacion_id}/aceptar-modal", include_in_schema=False)
async def aceptar_cotizacion_modal(
    request: Request,
    cotizacion_id: str,
    estatus_filter: str = Query(""),
    page: int = Query(1),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG, "editor"),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/aceptar_cotizacion_modal.html",
        {
            **_base_ctx(context, mod_role),
            "cotizacion": cotizacion,
            "cotizacion_id": str(uid),
            "estatus_filter": estatus_filter,
            "page": page,
            "fecha_hoy": today_mx().isoformat(),
        },
    )


@oym_router.get("/cotizaciones/{cotizacion_id}/info-modal", include_in_schema=False)
async def cotizacion_info_modal(
    request: Request,
    cotizacion_id: str,
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")
    mod_role = context.get("module_roles", {}).get("oym", "viewer")
    return templates.TemplateResponse(
        request, f"{TPL}/partials/cotizacion_info_modal.html",
        {
            **_base_ctx(context, mod_role),
            "cotizacion": cotizacion,
            "fecha_hoy": today_mx(),
        },
    )


# ============================================================
# PDF — PROPUESTA DE PÓLIZA
# ============================================================

@oym_router.get("/cotizaciones/{cotizacion_id}/pdf", include_in_schema=False)
async def descargar_pdf_poliza(
    cotizacion_id: str,
    show_projection: bool = Query(True),
    context=Depends(get_current_user_context),
    _=require_module_access(SLUG),
    conn=Depends(get_db_connection),
    service: CalculadoraService = Depends(get_service),
    pdf_service: PDFService = Depends(get_pdf_service),
):
    uid = _parse_cotizacion_id(cotizacion_id)
    cotizacion = await service.db.get_cotizacion_by_id(conn, uid)
    if not cotizacion:
        raise HTTPException(404, "Cotizacion no encontrada")

    resultado = cotizacion["resultado_json"]
    if isinstance(resultado, str):
        resultado = json.loads(resultado)

    planta = None
    if cotizacion.get("planta_id"):
        planta = await service.db.get_planta_by_id(conn, cotizacion["planta_id"])

    # Garantía de producción: resolver fecha_fin anterior
    es_externa = bool(planta.get("es_externa")) if planta else False
    fecha_fin_ant = (
        cotizacion.get("anterior_fecha_fin")   # del JOIN con póliza anterior en sistema
        or cotizacion.get("fecha_fin_poliza_anterior")  # entrada manual
    )
    garantia_produccion = tiene_garantia_produccion(
        tipo_poliza=cotizacion.get("tipo_poliza", ""),
        es_externa=es_externa,
        fecha_inicio=cotizacion.get("fecha_inicio_poliza"),
        fecha_fin_anterior=fecha_fin_ant,
    )

    tz = pytz.timezone("America/Mexico_City")
    _dt = cotizacion["created_at"].astimezone(tz) if cotizacion.get("created_at") else datetime.now(tz)
    fecha_emision = f"{_dt.day} de {_MESES_ES[_dt.month]} de {_dt.year}"

    factor = 1.03
    anio_1 = resultado.get("anio_1", resultado.get("sub_total_utilidad", 0))

    # resultado_json usa descuento_pct_1/3/5 independientes (schema actual, mig 035+)
    _dto_1 = resultado.get("descuento_pct_1") or 0.0
    _dto_3 = resultado.get("descuento_pct_3") or 0.0
    _dto_5 = resultado.get("descuento_pct_5") or 0.0
    # descuento_anios: años que tienen descuento aplicado
    descuento_anios = [y for y, pct in [(1, _dto_1), (3, _dto_3), (5, _dto_5)] if pct > 0]
    # descuento_pct: porcentaje para el badge del price-box (año 1)
    descuento_pct = _dto_1

    proyeccion = [
        {
            "anio": 1,
            "valor": round(anio_1, 2),
            "acumulado": round(anio_1, 2),
            "acumulado_desc": resultado.get("anio_1_desc") if 1 in descuento_anios else None,
        },
        {
            "anio": 3,
            "valor": round(anio_1 * (factor ** 2), 2),
            "acumulado": resultado.get("acumulado_1_3", 0),
            "acumulado_desc": resultado.get("acumulado_1_3_desc") if 3 in descuento_anios else None,
        },
        {
            "anio": 5,
            "valor": round(anio_1 * (factor ** 4), 2),
            "acumulado": resultado.get("acumulado_1_5", 0),
            "acumulado_desc": resultado.get("acumulado_1_5_desc") if 5 in descuento_anios else None,
        },
    ]

    sub_total_utilidad = resultado.get("sub_total_utilidad", 0)
    total_final = resultado.get("total_final", 0)

    ctx = {
        "folio": str(cotizacion["id"])[:8].upper(),
        "fecha_emision": fecha_emision,
        "ejecutivo": cotizacion.get("creado_por_nombre") or "Enertika Mexico",
        "nombre_planta": cotizacion.get("nombre_planta") or resultado.get("nombre_planta", ""),
        "cliente": planta.get("cliente") if planta else None,
        "direccion": planta.get("direccion") if planta else None,
        "zona": resultado.get("zona", ""),
        "potencia_kw": resultado.get("potencia_kw", 0),
        "num_paneles": resultado.get("num_paneles", 0),
        "tipo_poliza": resultado.get("tipo_poliza", "premium"),
        "nombre_wattabit": resultado.get("nombre_wattabit", ""),
        "sub_total_utilidad": sub_total_utilidad,
        "total_final": total_final,
        "proyeccion": proyeccion,
        "mostrar_proyeccion": show_projection,
        "descuento_pct": descuento_pct,
        "descuento_anios": descuento_anios,
        "descuento_monto": resultado.get("descuento_monto", 0.0),
        "garantia_produccion": garantia_produccion,
        "vigencia_cotizacion_dias": cotizacion.get("vigencia_cotizacion_dias") or 30,
    }

    pdf_bytes = await pdf_service.generate("poliza_oym.html", ctx)
    nombre_planta_clean = (cotizacion.get("nombre_planta") or "poliza").replace(" ", "_")
    filename = pdf_service.generate_filename("Propuesta_Poliza", nombre_planta_clean)

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
