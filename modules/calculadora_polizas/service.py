# modules/calculadora/service.py
from typing import Optional
from fastapi import Depends, HTTPException
from datetime import date
import logging
import io

from .db_service import CalculadoraDBService
from .schemas import CalcularRequest, CalcularResponse, ImportExcelResult

logger = logging.getLogger("CalculadoraPolizas.Service")


class CalculadoraService:

    def __init__(self):
        self.db = CalculadoraDBService()

    # ----------------------------------------
    # CÁLCULO PRINCIPAL
    # ----------------------------------------

    async def calcular(self, conn, req: CalcularRequest) -> CalcularResponse:
        planta = await self.db.get_planta_by_id(conn, req.planta_id)
        if not planta:
            raise ValueError(f"Planta '{req.planta_id}' no encontrada")

        potencia_kw = float(planta["potencia_kw"] or 0)
        num_paneles = int(planta["num_paneles"] or 0)

        if potencia_kw <= 0:
            raise ValueError("La planta no tiene potencia kW registrada")
        if num_paneles <= 0:
            raise ValueError("La planta no tiene cantidad de paneles registrada")

        precios_zona = await self.db.get_precios_zona(conn)
        costos = await self.db.get_costos_fijos(conn)
        wattabit_tier = await self.db.get_wattabit_para_kwp(conn, potencia_kw)

        zona = planta["zona"]
        precio_panel = precios_zona.get(zona)
        if precio_panel is None:
            raise ValueError(f"No existe precio configurado para la zona '{zona}'")

        if wattabit_tier is None:
            raise ValueError(f"La potencia {potencia_kw} kWp está fuera del rango Wattabit (máx 1,500 kWp)")

        tipo = req.tipo_poliza if isinstance(req.tipo_poliza, str) else req.tipo_poliza.value
        utilidad = req.utilidad

        # Componentes de costo
        if tipo == "premium":
            mtto_principal = precio_panel * num_paneles * 2   # 2 visitas/año
            mtto_fijo = costos["mtto_correctivo"]
        else:
            mtto_principal = costos["mtto_diagnostico_estandar"]
            mtto_fijo = 0.0

        wattabit_precio = float(wattabit_tier["precio_anual_mxp"])
        internet = costos["internet_anual"]
        gestion = costos["gestion_energetica_por_panel"] * num_paneles

        sub_total = mtto_principal + mtto_fijo + wattabit_precio + internet + gestion
        sub_total_utilidad = sub_total / (1 - utilidad)
        total_final = sub_total_utilidad * (1 + costos["iva"])

        peso_kwp = sub_total_utilidad / potencia_kw
        peso_panel = sub_total_utilidad / num_paneles

        # Proyección 5 años (3% anual)
        factor = costos.get("incremento_anual", 0.03)
        anos = []
        valor = sub_total_utilidad
        for i in range(5):
            if i > 0:
                valor *= (1 + factor)
            anos.append(round(valor, 2))

        # Descuento por duración de contrato (porcentaje independiente por opción)
        pct_1 = float(req.descuento_pct_1 or 0.0)
        pct_3 = float(req.descuento_pct_3 or 0.0)
        pct_5 = float(req.descuento_pct_5 or 0.0)

        def _con_dto(valor_base: float, pct: float) -> float:
            return round(valor_base * (1.0 - pct), 2) if pct > 0 else valor_base

        anio_1_desc = _con_dto(anos[0], pct_1)
        anio_3_desc = _con_dto(anos[2], pct_3)
        anio_5_desc = _con_dto(anos[4], pct_5)
        acumulado_1_3_desc = _con_dto(round(sum(anos[:3]), 2), pct_3)
        acumulado_1_5_desc = _con_dto(round(sum(anos), 2), pct_5)

        return CalcularResponse(
            planta_id=planta["id"],
            nombre_planta=planta["nombre"],
            zona=zona,
            potencia_kw=potencia_kw,
            num_paneles=num_paneles,
            tipo_poliza=tipo,
            utilidad=utilidad,
            mtto_principal=round(mtto_principal, 2),
            mtto_fijo=round(mtto_fijo, 2),
            wattabit=round(wattabit_precio, 2),
            internet=round(internet, 2),
            gestion=round(gestion, 2),
            sub_total=round(sub_total, 2),
            sub_total_utilidad=round(sub_total_utilidad, 2),
            total_final=round(total_final, 2),
            peso_kwp=round(peso_kwp, 4),
            peso_panel=round(peso_panel, 4),
            anio_1=anos[0],
            anio_3=anos[2],
            anio_5=anos[4],
            acumulado_1_3=round(sum(anos[:3]), 2),
            acumulado_1_5=round(sum(anos), 2),
            nombre_wattabit=wattabit_tier["nombre"],
            descuento_pct_1=req.descuento_pct_1,
            descuento_pct_3=req.descuento_pct_3,
            descuento_pct_5=req.descuento_pct_5,
            anio_1_desc=anio_1_desc,
            anio_3_desc=anio_3_desc,
            anio_5_desc=anio_5_desc,
            acumulado_1_3_desc=acumulado_1_3_desc,
            acumulado_1_5_desc=acumulado_1_5_desc,
        )

    # ----------------------------------------
    # PREVIEW EXCEL (sin guardar)
    # ----------------------------------------

    async def preview_plantas_excel(self, conn, contenido: bytes) -> dict:
        """
        Valida el Excel fila por fila sin guardar nada en BD.
        Retorna un dict con:
          - filas: list de dicts con estado 'nueva'|'actualiza'|'error'
          - errores_globales: list de str (columnas faltantes, etc.)
          - zonas_validas: list de str
        """
        try:
            import openpyxl
        except ImportError:
            return {"filas": [], "errores_globales": ["openpyxl no está instalado en el servidor"], "zonas_validas": []}

        precios_zona = await self.db.get_precios_zona(conn)
        zonas_validas = set(precios_zona.keys())

        try:
            wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        except Exception as exc:
            return {"filas": [], "errores_globales": [f"No se pudo leer el archivo: {exc}"], "zonas_validas": sorted(zonas_validas)}

        ws = wb.active

        # Leer encabezados
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            wb.close()
            return {"filas": [], "errores_globales": ["El archivo está vacío"], "zonas_validas": sorted(zonas_validas)}

        headers = [str(c.value).strip().lower() if c.value else "" for c in header_row]

        col_map = {}
        for campo, aliases in {
            "id":          ["id", "codigo", "código"],
            "nombre":      ["nombre", "planta", "name"],
            "zona":        ["zona", "zone"],
            "potencia_kw": ["potencia_kw", "potencia", "kw", "kwp"],
            "num_paneles": ["num_paneles", "paneles", "panels", "cantidad_paneles"],
            "cliente":     ["cliente", "client", "razon_social", "razón_social"],
            "direccion":   ["direccion", "dirección", "address", "ubicacion", "ubicación"],
            "es_externa":  ["es_externa", "externa", "external"],
        }.items():
            for alias in aliases:
                if alias in headers:
                    col_map[campo] = headers.index(alias)
                    break

        required = ["id", "nombre", "zona"]
        missing = [f for f in required if f not in col_map]
        if missing:
            wb.close()
            return {
                "filas": [],
                "errores_globales": [f"Columnas requeridas no encontradas en el Excel: {', '.join(missing)}. Descarga la plantilla para ver el formato correcto."],
                "zonas_validas": sorted(zonas_validas),
            }

        filas = []
        ids_en_excel = []  # para detectar duplicados dentro del mismo archivo

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            planta_id_raw = row[col_map["id"]] if row[col_map["id"]] is not None else ""
            nombre_raw    = row[col_map["nombre"]] if row[col_map["nombre"]] is not None else ""
            zona_raw      = row[col_map["zona"]] if row[col_map["zona"]] is not None else ""

            planta_id = str(planta_id_raw).strip().upper()
            nombre    = str(nombre_raw).strip()
            zona      = str(zona_raw).strip()

            # Fila completamente vacía → saltar
            if not planta_id and not nombre and not zona:
                continue

            errores_fila = []

            if not planta_id:
                errores_fila.append("ID vacío")
            if not nombre:
                errores_fila.append("Nombre vacío")
            if not zona:
                errores_fila.append("Zona vacía")
            elif zona not in zonas_validas:
                errores_fila.append(f"Zona '{zona}' no válida (válidas: {', '.join(sorted(zonas_validas))})")

            potencia_kw = None
            if "potencia_kw" in col_map and row[col_map["potencia_kw"]] is not None:
                try:
                    potencia_kw = float(row[col_map["potencia_kw"]])
                    if potencia_kw < 0:
                        errores_fila.append("Potencia no puede ser negativa")
                except (ValueError, TypeError):
                    errores_fila.append("Potencia no es un número válido")

            num_paneles = None
            if "num_paneles" in col_map and row[col_map["num_paneles"]] is not None:
                try:
                    num_paneles = int(row[col_map["num_paneles"]])
                    if num_paneles < 0:
                        errores_fila.append("Número de paneles no puede ser negativo")
                except (ValueError, TypeError):
                    errores_fila.append("Número de paneles no es un entero válido")

            cliente = None
            if "cliente" in col_map and row[col_map["cliente"]] is not None:
                cliente = str(row[col_map["cliente"]]).strip() or None

            direccion = None
            if "direccion" in col_map and row[col_map["direccion"]] is not None:
                direccion = str(row[col_map["direccion"]]).strip() or None

            es_externa = False
            if "es_externa" in col_map and row[col_map["es_externa"]] is not None:
                val = str(row[col_map["es_externa"]]).strip().lower()
                es_externa = val in ("1", "true", "sí", "si", "yes", "x")

            # Determinar estado si no hay errores de campos requeridos
            estado = "error" if errores_fila else "nueva"
            es_duplicado_bd = False

            if estado != "error" and planta_id:
                existente = await self.db.get_planta_by_id(conn, planta_id)
                if existente:
                    estado = "actualiza"
                    es_duplicado_bd = True

            # Duplicado dentro del mismo archivo
            es_duplicado_excel = planta_id in ids_en_excel
            if es_duplicado_excel and estado != "error":
                errores_fila.append(f"ID '{planta_id}' aparece más de una vez en el archivo")
                estado = "error"

            if planta_id:
                ids_en_excel.append(planta_id)

            filas.append({
                "fila": row_num,
                "id": planta_id,
                "nombre": nombre,
                "zona": zona,
                "potencia_kw": potencia_kw,
                "num_paneles": num_paneles,
                "cliente": cliente,
                "direccion": direccion,
                "es_externa": es_externa,
                "estado": estado,
                "errores": errores_fila,
                "es_duplicado_bd": es_duplicado_bd,
            })

        wb.close()
        return {
            "filas": filas,
            "errores_globales": [],
            "zonas_validas": sorted(zonas_validas),
        }

    # ----------------------------------------
    # IMPORTACIÓN EXCEL
    # ----------------------------------------

    async def importar_plantas_excel(self, conn, contenido: bytes) -> ImportExcelResult:
        try:
            import openpyxl
        except ImportError:
            raise ValueError("openpyxl no está instalado en el servidor")

        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active

        insertadas = 0
        actualizadas = 0
        errores = []

        # Leer encabezados de la primera fila
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

        col_map = {}
        for campo, aliases in {
            "id":          ["id", "codigo", "código"],
            "nombre":      ["nombre", "planta", "name"],
            "zona":        ["zona", "zone"],
            "potencia_kw": ["potencia_kw", "potencia", "kw", "kwp"],
            "num_paneles": ["num_paneles", "paneles", "panels", "cantidad_paneles"],
            "cliente":     ["cliente", "client", "razon_social", "razón_social"],
            "direccion":   ["direccion", "dirección", "address", "ubicacion", "ubicación"],
            "es_externa":  ["es_externa", "externa", "external"],
        }.items():
            for alias in aliases:
                if alias in headers:
                    col_map[campo] = headers.index(alias)
                    break

        required = ["id", "nombre", "zona"]
        missing = [f for f in required if f not in col_map]
        if missing:
            raise ValueError(f"Columnas requeridas no encontradas: {', '.join(missing)}")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                planta_id = str(row[col_map["id"]]).strip() if row[col_map["id"]] is not None else ""
                nombre = str(row[col_map["nombre"]]).strip() if row[col_map["nombre"]] is not None else ""
                zona = str(row[col_map["zona"]]).strip() if row[col_map["zona"]] is not None else ""

                if not planta_id or not nombre or not zona:
                    continue  # fila vacía, saltar sin error

                potencia_kw = None
                if "potencia_kw" in col_map and row[col_map["potencia_kw"]] is not None:
                    potencia_kw = float(row[col_map["potencia_kw"]])

                num_paneles = None
                if "num_paneles" in col_map and row[col_map["num_paneles"]] is not None:
                    num_paneles = int(row[col_map["num_paneles"]])

                cliente = None
                if "cliente" in col_map and row[col_map["cliente"]] is not None:
                    cliente = str(row[col_map["cliente"]]).strip() or None

                direccion = None
                if "direccion" in col_map and row[col_map["direccion"]] is not None:
                    direccion = str(row[col_map["direccion"]]).strip() or None

                es_externa = False
                if "es_externa" in col_map and row[col_map["es_externa"]] is not None:
                    val = str(row[col_map["es_externa"]]).strip().lower()
                    es_externa = val in ("1", "true", "sí", "si", "yes", "x")

                was_insert = await self.db.upsert_planta(conn, {
                    "id": planta_id,
                    "nombre": nombre,
                    "zona": zona,
                    "potencia_kw": potencia_kw,
                    "num_paneles": num_paneles,
                    "cliente": cliente,
                    "direccion": direccion,
                    "es_externa": es_externa,
                    "activa": True,
                })

                if was_insert:
                    insertadas += 1
                else:
                    actualizadas += 1

            except Exception as exc:
                errores.append(f"Fila {row_num}: {exc}")

        wb.close()
        return ImportExcelResult(insertadas=insertadas, actualizadas=actualizadas, errores=errores)

    # ----------------------------------------
    # CAMBIO DE ESTATUS (con validación de transiciones)
    # ----------------------------------------

    _ALLOWED_TRANSITIONS: dict = {
        "CREADA":          {"ENVIADA", "EN_NEGOCIACION", "RECHAZADA"},
        "ENVIADA":         {"CREADA", "EN_NEGOCIACION", "ACEPTADA", "RECHAZADA"},
        "EN_NEGOCIACION":  {"ENVIADA", "ACEPTADA", "RECHAZADA"},
        "ACEPTADA":        {"CANCELADA"},
        "RECHAZADA":       {"CREADA"},
        "VENCIDA":         set(),
        "TERMINADA":       set(),
        "CANCELADA":       set(),
    }

    async def cambiar_estatus_cotizacion(
        self, conn, cotizacion_id, nuevo_estatus: str,
        user_id, rol_sistema: str, mod_role: str,
        motivo: Optional[str] = None,
        fecha_inicio=None, fecha_fin=None,
        anios_contratados=None,
    ) -> None:
        """
        Cambia el estatus de una cotización aplicando las reglas de negocio:
        - Transiciones permitidas por estatus actual
        - CANCELADA: solo admin/manager, motivo obligatorio
        - ACEPTADA: verifica solapamiento de fechas y termina la póliza anterior si es renovación
        Lanza ValueError con mensaje legible si alguna validación falla.
        """
        cotizacion = await self.db.get_cotizacion_by_id(conn, cotizacion_id)
        if not cotizacion:
            raise ValueError("Cotizacion no encontrada")

        estatus_actual = cotizacion["estatus"]
        permitidos = self._ALLOWED_TRANSITIONS.get(estatus_actual, set())
        if nuevo_estatus not in permitidos:
            label = {
                "TERMINADA": "Terminada", "CANCELADA": "Cancelada",
                "VENCIDA": "Vencida",
            }.get(estatus_actual, estatus_actual)
            raise ValueError(f"Una poliza en estatus '{label}' no puede cambiar a '{nuevo_estatus}'")

        if nuevo_estatus == "CANCELADA":
            puede = rol_sistema in ("ADMIN", "MANAGER") or mod_role == "admin"
            if not puede:
                raise ValueError("Solo administradores pueden cancelar una poliza activa")
            if not motivo or not motivo.strip():
                raise ValueError("Se requiere un motivo para cancelar la poliza")

        if nuevo_estatus == "ACEPTADA":
            f_inicio = fecha_inicio or cotizacion.get("fecha_inicio_poliza")
            f_fin = fecha_fin or cotizacion.get("fecha_fin_poliza")
            planta_id = cotizacion.get("planta_id")

            if planta_id and f_inicio and f_fin:
                conflicto = await self.db.check_solapamiento_poliza(
                    conn, planta_id, f_inicio, f_fin, exclude_id=cotizacion_id
                )
                if conflicto:
                    raise ValueError(
                        "Ya existe una poliza activa para esta planta en ese rango de fechas. "
                        "Cancela o termina la poliza vigente antes de aceptar esta."
                    )

            poliza_anterior_id = cotizacion.get("poliza_anterior_id")
            if poliza_anterior_id:
                await self.db.terminar_poliza_anterior(conn, poliza_anterior_id, user_id)

        await self.db.update_cotizacion_estatus(
            conn, cotizacion_id, nuevo_estatus, user_id,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            anios_contratados=anios_contratados,
            motivo_cancelacion=motivo if nuevo_estatus == "CANCELADA" else None,
        )

    # ----------------------------------------
    # GUARDAR COTIZACIÓN
    # ----------------------------------------

    async def guardar_cotizacion(
        self, conn, resultado: CalcularResponse, user_id,
        solicitante_id=None,
        fecha_inicio_poliza=None,
        fecha_fin_poliza=None,
        poliza_anterior_id=None,
        fecha_fin_poliza_anterior=None,
        vigencia_cotizacion_dias: int = 30,
    ) -> str:
        cotizacion_id = await self.db.save_cotizacion(conn, {
            "planta_id": resultado.planta_id,
            "nombre_planta": resultado.nombre_planta,
            "tipo_poliza": resultado.tipo_poliza,
            "utilidad": resultado.utilidad,
            "sub_total": resultado.sub_total,
            "sub_total_utilidad": resultado.sub_total_utilidad,
            "total_final": resultado.total_final,
            "resultado_json": resultado.model_dump(),
            "creado_por": user_id,
            "solicitante_id": solicitante_id,
            "descuento_pct": None,
            "descuento_anios": None,
            "descuento_pct_1": resultado.descuento_pct_1,
            "descuento_pct_3": resultado.descuento_pct_3,
            "descuento_pct_5": resultado.descuento_pct_5,
            "fecha_inicio_poliza": fecha_inicio_poliza,
            "fecha_fin_poliza": fecha_fin_poliza,
            "poliza_anterior_id": poliza_anterior_id,
            "fecha_fin_poliza_anterior": fecha_fin_poliza_anterior,
            "vigencia_cotizacion_dias": vigencia_cotizacion_dias,
        })
        return str(cotizacion_id)


def get_service() -> CalculadoraService:
    return CalculadoraService()


def tiene_garantia_produccion(
    tipo_poliza: str,
    es_externa: bool,
    fecha_inicio: Optional[date],
    fecha_fin_anterior: Optional[date],
) -> bool:
    """True si el hito 'Garantia de produccion' aplica en el PDF.

    Condiciones (todas deben cumplirse):
    - Poliza premium
    - Planta instalada por Enertika (es_externa = False)
    - La renovation ocurre dentro de los 6 meses posteriores al vencimiento anterior
    """
    if tipo_poliza != "premium":
        return False
    if es_externa:
        return False
    if fecha_fin_anterior is None:
        return False
    inicio = fecha_inicio or date.today()
    # 6 meses ≈ 183 días (contamos desde vencimiento anterior hasta inicio de nueva)
    return (inicio - fecha_fin_anterior).days <= 183
