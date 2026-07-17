"""Builder openpyxl del reporte de clientes/empresas de Comercial (modo general y por cliente)."""

from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO

from openpyxl import Workbook

from core.timezone import ensure_mx, now_mx
from modules.rrhh.excel_utils import autofit_columns, style_sheet
from modules.shared.utils import safe_sheet_title

_CARACTERES_PELIGROSOS = ("=", "+", "-", "@")

_RESUMEN_HEADERS = ["Cliente", "Total de solicitudes", "Desglose por estatus"]
_DETALLE_GENERAL_HEADERS = ["Cliente", "Folio", "Fecha de solicitud", "Estatus", "Fase de proyecto"]
_DETALLE_CLIENTE_HEADERS = [
    "Folio",
    "Fecha de solicitud",
    "Estatus",
    "Sitio",
    "Direccion",
    "Proyecto",
    "Fase de proyecto",
]

# (clave del dict, es_fecha) en el mismo orden que los headers de cada hoja de detalle.
_DETALLE_GENERAL_COLUMNAS = [
    ("cliente_nombre", False),
    ("folio", False),
    ("fecha_solicitud", True),
    ("estatus_nombre", False),
    ("fase_proyecto", False),
]
_DETALLE_CLIENTE_COLUMNAS = [
    ("folio", False),
    ("fecha_solicitud", True),
    ("estatus_nombre", False),
    ("sitio_nombre", False),
    ("sitio_direccion", False),
    ("proyecto_id_estandar", False),
    ("fase_proyecto", False),
]


def _texto_seguro(value) -> str:
    """Neutraliza inyeccion de formulas: antepone comilla si el texto empieza con =, +, - o @."""
    text = "" if value is None else str(value)
    if text and text[0] in _CARACTERES_PELIGROSOS:
        return "'" + text
    return text


def _set_fecha_solicitud(worksheet, row: int, col: int, value) -> None:
    """Escribe la fecha como valor nativo de Excel (no strftime), convertida a MX."""
    if isinstance(value, datetime):
        dt_mx = ensure_mx(value).replace(tzinfo=None)
        cell = worksheet.cell(row=row, column=col, value=dt_mx)
        cell.number_format = "dd/mm/yyyy hh:mm"
    elif isinstance(value, date):
        cell = worksheet.cell(row=row, column=col, value=value)
        cell.number_format = "dd/mm/yyyy"
    else:
        worksheet.cell(row=row, column=col, value="")


def _escribir_fila_detalle(worksheet, row_num: int, row: dict, columnas: list[tuple[str, bool]]) -> None:
    for col_index, (clave, es_fecha) in enumerate(columnas, start=1):
        if es_fecha:
            _set_fecha_solicitud(worksheet, row_num, col_index, row.get(clave))
        else:
            worksheet.cell(row=row_num, column=col_index, value=_texto_seguro(row.get(clave)))


def _agregar_hoja_detalle(
    workbook: Workbook,
    titulo: str,
    headers: list[str],
    columnas: list[tuple[str, bool]],
    detalle: list[dict],
) -> None:
    ws = workbook.create_sheet(titulo)
    style_sheet(ws, headers)
    for offset, row in enumerate(detalle, start=2):
        _escribir_fila_detalle(ws, offset, row, columnas)
    autofit_columns(ws, visible_columns=len(headers))


def build_reporte_clientes_general_workbook(resumen: list[dict], detalle: list[dict]) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    ws_resumen = workbook.create_sheet("Resumen por cliente")
    style_sheet(ws_resumen, _RESUMEN_HEADERS)
    for row in resumen:
        ws_resumen.append([
            _texto_seguro(row.get("cliente_nombre")),
            int(row.get("total_solicitudes") or 0),
            _texto_seguro(row.get("desglose_estatus")),
        ])
    autofit_columns(ws_resumen, visible_columns=len(_RESUMEN_HEADERS))

    _agregar_hoja_detalle(
        workbook, "Detalle de solicitudes", _DETALLE_GENERAL_HEADERS, _DETALLE_GENERAL_COLUMNAS, detalle
    )

    return workbook


def build_reporte_clientes_por_cliente_workbook(detalle: list[dict], cliente_nombre: str | None = None) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    titulo_hoja = safe_sheet_title(
        f"Detalle - {cliente_nombre}" if cliente_nombre else "Detalle por cliente",
        set(),
        fallback="Detalle por cliente",
    )
    _agregar_hoja_detalle(workbook, titulo_hoja, _DETALLE_CLIENTE_HEADERS, _DETALLE_CLIENTE_COLUMNAS, detalle)

    return workbook


def _workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def construir_bytes_general(resumen: list[dict], detalle: list[dict]) -> bytes:
    """Sincrono: pensado para ejecutarse en un executor (CPU-bound, no bloquear el loop)."""
    workbook = build_reporte_clientes_general_workbook(resumen, detalle)
    return _workbook_to_bytes(workbook)


def construir_bytes_por_cliente(detalle: list[dict], cliente_nombre: str | None = None) -> bytes:
    """Sincrono: pensado para ejecutarse en un executor (CPU-bound, no bloquear el loop)."""
    workbook = build_reporte_clientes_por_cliente_workbook(detalle, cliente_nombre)
    return _workbook_to_bytes(workbook)


def generar_nombre_archivo(cliente_nombre: str | None = None) -> str:
    ts = f"{now_mx():%Y%m%d_%H%M}"
    if cliente_nombre:
        slug = re.sub(r"[^\w\-]", "_", cliente_nombre).strip("_")[:40]
        if slug:
            return f"reporte_clientes_{slug}_{ts}.xlsx"
    return f"reporte_clientes_{ts}.xlsx"
