

from fastapi import APIRouter, Request, Depends, HTTPException, Form, UploadFile, File, status
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from uuid import UUID, uuid4
from typing import Optional, List
import pandas as pd  # Solo para generación de Excel (escritura)
import io
import json
import logging
import asyncpg  # Para manejo específico de excepciones de PostgreSQL
import re  # Para sanitización de nombres en IDs
from openpyxl import load_workbook  # Para lectura ligera de Excel
from datetime import datetime, timedelta, time as dt_time # Keep timedelta and dt_time as they are used later


# --- Imports de Core ---
from core.database import get_db_connection
from core.microsoft import get_ms_auth, MicrosoftAuth # Keep MicrosoftAuth as it's used in get_ms_auth
from core.config import settings
from core.security import get_current_user_context, get_valid_graph_token  # NUEVO: Sistema de renovación de tokens
from core.permissions import require_module_access  # NUEVO: Sistema de permisos
from .schemas import OportunidadCreate, SitioImportacion, OportunidadListOut, OportunidadCreateCompleta, DetalleBessCreate
from .service import ComercialService, get_comercial_service  # NUEVO: Service Layer
from .email_handler import EmailHandler, get_email_handler  # NUEVO: Email Handler

# Configuración básica de logging
logger = logging.getLogger("ComercialModule")

templates = Jinja2Templates(directory="templates")

# Registrar filtros de timezone (México)
from core.jinja_filters import register_timezone_filters
register_timezone_filters(templates.env)

router = APIRouter(
    prefix="/comercial",
    tags=["Módulo Comercial"],
)


# ----------------------------------------
# ENDPOINTS
# ----------------------------------------

@router.head("/ui", include_in_schema=False)
async def check_comercial_ui(
    request: Request,
    context = Depends(get_current_user_context),
    _ = require_module_access("comercial")  # VALIDACIÓN DE ACCESO
):
    """Heartbeat endpoint to check session status without rendering."""
    return HTMLResponse("", status_code=200)

@router.get("/ui", include_in_schema=False)
async def get_comercial_ui(
    request: Request,
    context = Depends(get_current_user_context),  # Usar dependencia completa
    _ = require_module_access("comercial")  # VALIDACIÓN DE ACCESO
):
    """Main Entry: Shows the Tabbed Dashboard (Graphs + Records)."""
    user_name = context.get("user_name", "Usuario")
    role = context.get("role", "USER")
    
    # Detección inteligente de contexto:
    # 1. Si es HTMX (Navegación parcial desde Sidebar), devolvemos solo contenido interno (tabs.html)
    # 2. Si es Carga Completa (F5 o URL directa), devolvemos el wrapper completo (dashboard.html)
    if request.headers.get("hx-request"):
        template = "comercial/tabs.html"
    else:
        template = "comercial/dashboard.html"
        
    return templates.TemplateResponse(template, {
        "request": request,
        "user_name": user_name,
        "role": role,  # Pasar rol de sistema para el sidebar
        "module_roles": context.get("module_roles", {}),  # IMPORTANTE para el sidebar
        "current_module_role": context.get("module_roles", {}).get("comercial", "viewer")  # Rol específico en este módulo
    }, headers={"HX-Title": "Enertika Ops Core | Comercial"})

@router.get("/form", include_in_schema=False)
async def get_comercial_form(
    request: Request,
    user_context = Depends(get_current_user_context),
    conn = Depends(get_db_connection),
    service: ComercialService = Depends(get_comercial_service),
    _ = require_module_access("comercial", "editor")  # REQUIERE ROL EDITOR O SUPERIOR
):
    """Shows the creation form (Partial or Full Page)."""
    
    # Validación Estricta: Si no hay email, cortamos aquí.
    if not user_context.get("email"):
        # Retornamos 401 SIN redirección automática. HTMX lo atrapará.
        return HTMLResponse(status_code=401)
    
    # PREVENCIÓN CRÍTICA: Validar token ANTES de mostrar formulario
    # Esto evita que el usuario pierda su trabajo si el token expira mientras lo llena.
    # Si el token está cerca de expirar, get_valid_graph_token lo renovará automáticamente.
    token = await get_valid_graph_token(request)
    if not token:
        # Token expirado y no se pudo renovar - redirigir al login AHORA
        # Mejor que el usuario lo sepa de inmediato en lugar de perder 10 minutos de trabajo
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})
    
    # Generar canal default desde el servicio
    canal_default = ComercialService.get_canal_from_user_name(
        user_context.get("user_name")
    )
    
    # CORRECCI\u00d3N: Cargar catálogo completo si es modo homologación (legacy_term presente)
    # Esto permite que ACTUALIZACIÓN esté disponible en el template
    if request.query_params.get('legacy_term'):
        catalogos = await service.get_catalogos_ui(conn)  # TODOS los tipos
        
        # Buscar ACTUALIZACIÓN directamente en BD por codigo_interno (más confiable que regex)
        tipo_act = await conn.fetchrow(
            "SELECT id, nombre FROM tb_cat_tipos_solicitud WHERE codigo_interno = 'ACTUALIZACION' AND activo = true"
        )
        catalogos['tipo_actualizacion_id'] = tipo_act['id'] if tipo_act else None
    else:
        catalogos = await service.get_catalogos_creacion(conn)  # Filtrado (PRE_OFERTA, LICITACION)

    return templates.TemplateResponse("comercial/form.html", {
        "request": request, 
        "canal_default": canal_default,
        "catalogos": catalogos,  # Catálogos filtrados
        "user_name": user_context.get("user_name"),
        "role": user_context.get("role"),
        "module_roles": user_context.get("module_roles", {})
    }, headers={"HX-Title": "Enertika Ops Core | Nuevo Comercial"})

@router.get("/partials/graphs", include_in_schema=False)
async def get_graphs_partial(
    request: Request,
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection),
    user_context: dict = Depends(get_current_user_context),
    _ = require_module_access("comercial")
):
    """Partial: Graphs Tab Content."""
    stats = await service.get_dashboard_stats(conn, user_context)
    return templates.TemplateResponse("comercial/partials/graphs.html", {"request": request, "stats": stats})

@router.get("/partials/cards", include_in_schema=False)
async def get_cards_partial(
    request: Request,
    tab: str = "activos",
    q: Optional[str] = None,
    limit: int = 15,
    subtab: Optional[str] = None,
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection),
    user_context: dict = Depends(get_current_user_context),
    _ = require_module_access("comercial")
):
    """Partial: List of Opportunities (Cards/Grid)."""
    
    # OPTIMIZACIÓN: Token se obtiene solo cuando el usuario hace click en "Enviar Correo"
    # Esto evita una llamada HTTP a Microsoft Graph en cada cambio de pestaña
    # El token se validará en tiempo real cuando se necesite (lazy loading)
    user_token = request.session.get("access_token")  # Solo verificar si existe en sesión
    
    items = await service.get_oportunidades_list(conn, user_context=user_context, tab=tab, q=q, limit=limit, subtab=subtab)
    
    return templates.TemplateResponse(
        "comercial/partials/cards.html", 
        {
            "request": request, 
            "oportunidades": items,
            "user_token": user_token,
            "current_tab": tab,
            "subtab": subtab,
            "q": q,
            "limit": limit
        }
    )

@router.get("/partials/sitios/{id_oportunidad}", include_in_schema=False)
async def get_sitios_partial(
    request: Request,
    id_oportunidad: UUID,
    conn = Depends(get_db_connection)
):
    """Retorna la sub-tabla de sitios para una oportunidad."""
    rows = await conn.fetch(
        "SELECT * FROM tb_sitios_oportunidad WHERE id_oportunidad = $1 ORDER BY id_sitio",
        id_oportunidad
    )
    return templates.TemplateResponse(
        "comercial/partials/sitios_list.html",
        {"request": request, "sitios": rows}
    )

@router.get("/partials/comentarios/{id_oportunidad}", include_in_schema=False)
async def get_comentarios_partial(
    request: Request,
    id_oportunidad: UUID,
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("comercial")
):
    """Retorna los comentarios de simulación para una oportunidad."""
    comentarios = await service.get_comentarios_simulacion(conn, id_oportunidad)
    return templates.TemplateResponse(
        "comercial/partials/detalles/comentarios_list.html",
        {"request": request, "comentarios": comentarios}
    )

@router.get("/partials/bess/{id_oportunidad}", include_in_schema=False)
async def get_bess_partial(
    request: Request,
    id_oportunidad: UUID,
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("comercial")
):
    """Retorna los detalles BESS para una oportunidad."""
    bess = await service.get_detalles_bess(conn, id_oportunidad)
    return templates.TemplateResponse(
        "comercial/partials/detalles/bess_info.html",
        {"request": request, "bess": bess}
    )

@router.post("/notificar/{id_oportunidad}")
async def notificar_oportunidad(
    request: Request,
    id_oportunidad: UUID,
    recipients_str: str = Form(""), # Chips de TO
    fixed_to: List[str] = Form([]), # Hidden fixed TOs
    fixed_cc: List[str] = Form([]), # Hidden fixed CCs
    extra_cc: str = Form(""),       # Input manual CC
    subject: str = Form(...),
    body: str = Form(""),           # Mensaje adicional del usuario
    auto_message: str = Form(...),  # Mensaje automático
    prioridad: str = Form("normal"),  # Prioridad del email
    archivos_extra: List[UploadFile] = File(default=[]),
    service: ComercialService = Depends(get_comercial_service),
    ms_auth = Depends(get_ms_auth),
    conn = Depends(get_db_connection),
    email_handler = Depends(get_email_handler)  # NUEVO: Inyectar EmailHandler
):
    """Envía el correo de notificación usando el token de la sesión."""
    
    # Preparar datos del formulario
    form_data = {
        "recipients_str": recipients_str,
        "fixed_to": fixed_to,
        "fixed_cc": fixed_cc,
        "extra_cc": extra_cc,
        "subject": subject,
        "body": body,
        "auto_message": auto_message,
        "prioridad": prioridad,
        "archivos_extra": archivos_extra
    }
    
    # Delegar toda la lógica al EmailHandler
    success, result = await email_handler.procesar_y_enviar_notificacion(
        request=request,
        conn=conn,
        service=service,
        ms_auth=ms_auth,
        id_oportunidad=id_oportunidad,
        form_data=form_data
    )
    
    return result

@router.get("/plantilla", response_class=StreamingResponse)
async def descargar_plantilla_sitios():
    """Genera y descarga la plantilla Excel oficial."""
    # Columnas actualizadas según requerimiento
    cols = ["#", "NOMBRE", "# DE SERVICIO", "TARIFA", "LINK GOOGLE", "DIRECCION", "COMENTARIOS"]
    df = pd.DataFrame(columns=cols)
    
    # Fila de ejemplo actualizada
    df.loc[0] = [1, "SUCURSAL NORTE", "123456789012", "GDMTO", "https://maps.google.com/?q=19.4326,-99.1332", "Av. Reforma 123, Col. Centro", "Ejemplo de comentario"]
    
    # Guardar en buffer de memoria
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Sitios')
    buffer.seek(0)
    
    headers = {"Content-Disposition": 'attachment; filename="plantilla_sitios_enertika.xlsx"'}
    return StreamingResponse(buffer, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@router.post("/validate-thread-check")
async def validate_thread_check(
    request: Request,
    search_term: str = Form(...),
    ms_auth = Depends(get_ms_auth)
):
    """Valida si existe un hilo de correo antes de permitir avanzar al usuario (Modo Homologación)."""
    token = await get_valid_graph_token(request)
    if not token:
        return JSONResponse({"found": False, "error": "Sesión expirada"}, status_code=401)

    thread_id = ms_auth.find_thread_id(token, search_term)
    
    if thread_id:
        # Retorna éxito y el término para que el frontend lo pase al formulario
        return JSONResponse({"found": True, "clean_term": search_term})
    else:
        return JSONResponse({"found": False, "message": "No se encontró ningún hilo con ese texto."}, status_code=404)

@router.post("/form")
async def handle_oportunidad_creation(
    request: Request,
    # --- Campos Estándar ---
    nombre_proyecto: str = Form(...),
    nombre_cliente: str = Form(...),
    canal_venta: str = Form(...),
    id_tecnologia: int = Form(...),
    id_tipo_solicitud: int = Form(...),
    cantidad_sitios: int = Form(...),
    prioridad: str = Form(...),
    direccion_obra: str = Form(...),
    coordenadas_gps: Optional[str] = Form(None),
    google_maps_link: Optional[str] = Form(None),
    sharepoint_folder_url: Optional[str] = Form(None),
    
    # --- Campo Fecha Manual (Gerentes) ---
    fecha_manual: Optional[str] = Form(None),
    
    # --- Campos BESS (HTMX Conditional) ---
    bess_cargas_criticas: Optional[float] = Form(None),
    bess_tiene_motores: bool = Form(False),
    bess_potencia_motor: Optional[float] = Form(None),
    bess_autonomia: Optional[str] = Form(None),
    bess_voltaje: Optional[str] = Form(None),
    bess_planta_emergencia: bool = Form(False),
    bess_cargas_separadas: Optional[bool] = Form(False), # NUEVO CAMPO
    bess_objetivos: List[str] = Form([]),  # Recibe lista de checkboxes

    # --- Campo Legacy (Homologación) ---
    legacy_search_term: Optional[str] = Form(None),  # Asunto del correo antiguo (solo viaja, no se guarda)

    # --- Dependencias ---
    conn = Depends(get_db_connection),
    service: ComercialService = Depends(get_comercial_service),
    context = Depends(get_current_user_context)
):
    try:
        # 1. Seguridad Token
        token = await get_valid_graph_token(request)
        if not token: 
            from fastapi import Response
            return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})

        # 2. Construir Objeto BESS (Pydantic v2)
        # Solo si se enviaron datos relevantes o la tecnología implica BESS
        datos_bess = None
        if bess_cargas_criticas or bess_tiene_motores or bess_objetivos:
            # Importar schema BESS
            from .schemas import DetalleBessCreate
            datos_bess = DetalleBessCreate(
                cargas_criticas_kw=bess_cargas_criticas,
                tiene_motores=bess_tiene_motores,
                potencia_motor_hp=bess_potencia_motor,
                tiempo_autonomia=bess_autonomia,
                voltaje_operacion=bess_voltaje,
                cargas_separadas=bess_cargas_separadas,
                objetivos_json=bess_objetivos,
                tiene_planta_emergencia=bess_planta_emergencia
            )

        # 3. Validar Permiso Fecha Manual
        role = context.get("role")
        fecha_final_str = None
        if role in ['ADMIN', 'MANAGER']:
            fecha_final_str = fecha_manual
        
        # 3.5 MODO HOMOLOGACIÓN: Forzar tipo ACTUALIZACIÓN
        # Si viene legacy_search_term, sobrescribir id_tipo_solicitud
        if legacy_search_term:
            id_tipo_actualizacion = await conn.fetchval(
                "SELECT id FROM tb_cat_tipos_solicitud WHERE UPPER(codigo_interno) = 'ACTUALIZACION'"
            )
            if id_tipo_actualizacion:
                id_tipo_solicitud = id_tipo_actualizacion
                logger.info(f"MODO HOMOLOGACIÓN: Tipo de solicitud forzado a ACTUALIZACIÓN")
            else:
                logger.warning("No se encontró tipo ACTUALIZACIÓN en catálogo, usando el seleccionado por usuario")

        # 4. Construir Objeto Principal (Pydantic v2)
        from .schemas import OportunidadCreateCompleta
        oportunidad_data = OportunidadCreateCompleta(
            nombre_proyecto=nombre_proyecto,
            cliente_nombre=nombre_cliente,
            canal_venta=canal_venta,
            id_tecnologia=id_tecnologia,
            id_tipo_solicitud=id_tipo_solicitud,
            cantidad_sitios=cantidad_sitios,
            prioridad=prioridad,
            direccion_obra=direccion_obra,
            coordenadas_gps=coordenadas_gps,
            google_maps_link=google_maps_link,
            sharepoint_folder_url=sharepoint_folder_url,
            fecha_manual_str=fecha_final_str,
            detalles_bess=datos_bess
        )

        # 5. Ejecutar Transacción en Servicio
        new_id, op_id_estandar, es_fuera_horario = await service.crear_oportunidad_transaccional(
            conn, oportunidad_data, context
        )

        # 6. Redirección (Lógica Multisitos vs Unisitio)
        if cantidad_sitios == 1:
            # Auto-creación de sitio único (Legacy logic mantenida por consistencia)
            try:
                await conn.execute("""
                    INSERT INTO tb_sitios_oportunidad (id_sitio, id_oportunidad, nombre_sitio, direccion, google_maps_link)
                    VALUES ($1, $2, $3, $4, $5)
                """, uuid4(), new_id, nombre_proyecto, direccion_obra, google_maps_link)
            except Exception as e:
                logger.error(f"Error auto-creando sitio único: {e}")
            target_url = f"/comercial/paso3/{new_id}"
        else:
            target_url = f"/comercial/paso2/{new_id}"
        
        # Construir parámetros de URL
        params = f"?new_op={op_id_estandar}&fh={str(es_fuera_horario).lower()}"
        
        # PUENTE: Si existe legacy_search_term, agregarlo a la URL para que viaje hasta el Paso 3
        if legacy_search_term:
            import urllib.parse
            safe_legacy = urllib.parse.quote(legacy_search_term)
            params += f"&legacy_term={safe_legacy}"
        
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": f"{target_url}{params}"})

    except Exception as e:
        logger.error(f"Error en creación de oportunidad: {e}")
        # En producción, mejorar el manejo de error para no exponer detalles técnicos
        return templates.TemplateResponse(
            "comercial/error_message.html", 
            {"request": request, "detail": str(e)},
            status_code=500
        )

# ===== FORMULARIO EXTRAORDINARIO (ADMIN/MANAGER ONLY) =====
@router.get("/form-extraordinario", include_in_schema=False)
async def get_comercial_form_extraordinario(
    request: Request,
    user_context = Depends(get_current_user_context),
    conn = Depends(get_db_connection),
    service: ComercialService = Depends(get_comercial_service)
):
    """Shows the extraordinary creation form (ADMIN/MANAGER ONLY)."""
    
    # Validación de Rol: SOLO ADMIN o MANAGER
    role = user_context.get("role")
    if role not in ['ADMIN', 'MANAGER']:
        from fastapi import Response
        return Response(status_code=403, content="Acceso denegado. Solo ADMIN y MANAGER pueden acceder.")
    
    # Validación de sesión
    if not user_context.get("email"):
        return HTMLResponse(status_code=401)
    
    # Validar token
    token = await get_valid_graph_token(request)
    if not token:
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})
    
    # Generar canal default
    canal_default = ComercialService.get_canal_from_user_name(user_context.get("user_name"))
    
    # Obtener catálogos
    catalogos = await service.get_catalogos_creacion(conn)

    return templates.TemplateResponse("comercial/form_extraordinario.html", {
        "request": request, 
        "canal_default": canal_default,
        "catalogos": catalogos,
        "user_name": user_context.get("user_name"),
        "role": user_context.get("role"),
        "module_roles": user_context.get("module_roles", {})
    }, headers={"HX-Title": "Enertika Ops Core | Solicitud Extraordinaria"})

@router.post("/form-extraordinario")
async def handle_oportunidad_extraordinaria(
    request: Request,
    # --- Campos Estándar ---
    nombre_proyecto: str = Form(...),
    nombre_cliente: str = Form(...),
    canal_venta: str = Form(...),
    id_tecnologia: int = Form(...),
    id_tipo_solicitud: int = Form(...),
    cantidad_sitios: int = Form(...),
    prioridad: str = Form(...),
    direccion_obra: str = Form(...),
    coordenadas_gps: Optional[str] = Form(None),
    google_maps_link: Optional[str] = Form(None),
    sharepoint_folder_url: Optional[str] = Form(None),
    
    # --- Campo Fecha Manual (OBLIGATORIO en extraordinarias) ---
    fecha_manual: str = Form(...),  # REQUIRED en extraordinarias
    
    # --- Campos BESS (Opcionales) ---
    bess_cargas_criticas: Optional[float] = Form(None),
    bess_tiene_motores: bool = Form(False),
    bess_potencia_motor: Optional[float] = Form(None),
    bess_autonomia: Optional[str] = Form(None),
    bess_voltaje: Optional[str] = Form(None),
    bess_planta_emergencia: bool = Form(False),
    bess_cargas_separadas: Optional[bool] = Form(False), # NUEVO CAMPO
    bess_objetivos: List[str] = Form([]),

    # --- Dependencias ---
    conn = Depends(get_db_connection),
    service: ComercialService = Depends(get_comercial_service),
    context = Depends(get_current_user_context)
):
    """Procesa solicitud extraordinaria: SIN envío de correo, email_enviado=TRUE automático."""
    try:
        # 1. Seguridad: Validar rol ADMIN/MANAGER
        role = context.get("role")
        if role not in ['ADMIN', 'MANAGER']:
            from fastapi import HTTPException
            raise HTTPException(status_code=403, detail="Solo ADMIN y MANAGER pueden crear solicitudes extraordinarias")
        
        # Validar token
        token = await get_valid_graph_token(request)
        if not token: 
            from fastapi import Response
            return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})

        # 2. Construir Objeto BESS (si aplica)
        datos_bess = None
        if bess_cargas_criticas or bess_tiene_motores or bess_objetivos:
            from .schemas import DetalleBessCreate
            datos_bess = DetalleBessCreate(
                cargas_criticas_kw=bess_cargas_criticas,
                tiene_motores=bess_tiene_motores,
                potencia_motor_hp=bess_potencia_motor,
                tiempo_autonomia=bess_autonomia,
                voltaje_operacion=bess_voltaje,
                cargas_separadas=bess_cargas_separadas,
                objetivos_json=bess_objetivos,
                tiene_planta_emergencia=bess_planta_emergencia
            )

        # 3. Construir Objeto Principal (fecha_manual siempre presente)
        from .schemas import OportunidadCreateCompleta
        oportunidad_data = OportunidadCreateCompleta(
            nombre_proyecto=nombre_proyecto,
            cliente_nombre=nombre_cliente,
            canal_venta=canal_venta,
            id_tecnologia=id_tecnologia,
            id_tipo_solicitud=id_tipo_solicitud,
            cantidad_sitios=cantidad_sitios,
            prioridad=prioridad,
            direccion_obra=direccion_obra,
            coordenadas_gps=coordenadas_gps,
            google_maps_link=google_maps_link,
            sharepoint_folder_url=sharepoint_folder_url,
            fecha_manual_str=fecha_manual,  # SIEMPRE presente en extraordinarias
            detalles_bess=datos_bess
        )

        # 4. Ejecutar Transacción en Servicio
        new_id, op_id_estandar, es_fuera_horario = await service.crear_oportunidad_transaccional(
            conn, oportunidad_data, context
        )
        
        # 5. MARCAR COMO EXTRAORDINARIA: email_enviado = TRUE (sin envío real)
        await conn.execute("""
            UPDATE tb_oportunidades 
            SET email_enviado = TRUE
            WHERE id_oportunidad = $1
        """, new_id)
        
        logger.info(f"Solicitud extraordinaria {op_id_estandar} marcada con email_enviado=TRUE (sin envío real)")

        # 6. Redirección (Lógica Multisitos vs Unisitio)
        if cantidad_sitios == 1:
            # Auto-creación de sitio único para extraordinarias
            try:
                await conn.execute("""
                    INSERT INTO tb_sitios_oportunidad (id_sitio, id_oportunidad, nombre_sitio, direccion, google_maps_link)
                    VALUES ($1, $2, $3, $4, $5)
                """, uuid4(), new_id, nombre_proyecto, direccion_obra, google_maps_link)
            except Exception as e:
                logger.error(f"Error auto-creando sitio único: {e}")
            
            # EXTRAORDINARIAS: Unisitio va directamente al HOME (sin paso 3 de correo)
            target_url = "/comercial/ui"
            params = f"?new_op={op_id_estandar}&fh={str(es_fuera_horario).lower()}&extraordinaria=1"
        else:
            # Multisitio: va a paso 2 (carga Excel), luego volverá al home
            target_url = f"/comercial/paso2/{new_id}"
            params = f"?new_op={op_id_estandar}&fh={str(es_fuera_horario).lower()}&extraordinaria=1"
        
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": f"{target_url}{params}"})

    except Exception as e:
        logger.error(f"Error en creación de solicitud extraordinaria: {e}")
        return templates.TemplateResponse(
            "comercial/error_message.html", 
            {"request": request, "detail": str(e)},
            status_code=500
        )

# (El endpoint de Excel se mantiene igual, asegúrate de que esté incluido en tu archivo)
@router.delete("/{id_oportunidad}", response_class=HTMLResponse)
async def cancelar_oportunidad(
    request: Request,
    id_oportunidad: UUID,
    conn = Depends(get_db_connection)
):
    """Elimina borrador y fuerza una recarga completa al Dashboard."""
    
    # 1. Protección de Sesión con Token Inteligente
    access_token = await get_valid_graph_token(request)
    if not access_token:
        # Token expirado y no se pudo renovar
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})
        
    # 2. Borrar datos en BD
    await conn.execute("DELETE FROM tb_sitios_oportunidad WHERE id_oportunidad = $1", id_oportunidad)
    await conn.execute("DELETE FROM tb_oportunidades WHERE id_oportunidad = $1", id_oportunidad)
    
    # 3. CAMBIO CLAVE: Usamos HX-Redirect en lugar de HX-Location
    # HX-Redirect obliga al navegador a cambiar de URL (como un F5 + ir a nueva página)
    # Esto limpia la memoria y carga los estilos/scripts correctamente.
    from fastapi import Response
    return Response(status_code=200, headers={"HX-Redirect": "/comercial/ui"}) 


# ----------------------------------------
# NEW ENDPOINTS FOR STEP 3 & DEBUG
# ----------------------------------------

@router.get("/paso3/{id_oportunidad}", include_in_schema=False)
async def get_paso3_email_form(
    request: Request,
    id_oportunidad: UUID,
    legacy_term: Optional[str] = None,  # Captura del término legacy desde la URL
    conn = Depends(get_db_connection)
):
    """Muestra el formulario final de envío de correo (Paso 3)."""
    
    # PREVENCIÓN CRÍTICA: Validar token ANTES de mostrar formulario de correo
    # El usuario puede tardar varios minutos redactando el mensaje
    # Si el token expira mientras escribe, perderá todo su trabajo
    token = await get_valid_graph_token(request)
    if not token:
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})
    
    # Recuperar datos para pre-llenar con JOINs para traer nombres de catálogos
    # CORRECCIÓN: Agregar JOINs para tipo_tecnologia y status_global (reglas de email)
    row = await conn.fetchrow(
        """SELECT o.*, 
                  tec.nombre as tipo_tecnologia,
                  tipo_sol.nombre as tipo_solicitud,
                  tipo_sol.es_seguimiento,
                  eg.nombre as status_global,
                  -- DETALLES BESS
                  db.cargas_criticas_kw,
                  db.tiene_motores,
                  db.potencia_motor_hp,
                  db.tiempo_autonomia,
                  db.voltaje_operacion,
                  db.cargas_separadas,
                  db.objetivos_json,
                  db.tiene_planta_emergencia
           FROM tb_oportunidades o
           LEFT JOIN tb_cat_tecnologias tec ON o.id_tecnologia = tec.id
           LEFT JOIN tb_cat_tipos_solicitud tipo_sol ON o.id_tipo_solicitud = tipo_sol.id
           LEFT JOIN tb_cat_estatus_global eg ON o.id_estatus_global = eg.id
           LEFT JOIN tb_detalles_bess db ON o.id_oportunidad = db.id_oportunidad
           WHERE o.id_oportunidad = $1""", 
        id_oportunidad
    )
    if not row:
        return HTMLResponse("Oportunidad no encontrada", 404)
        
    # Verificar si es multisitio para mostrar badge
    has_multisitio = (row['cantidad_sitios'] or 0) > 1
    
    # Determinar si es seguimiento desde BD (SIN HARDCODING)
    es_seguimiento = row.get('es_seguimiento', False)
    
    # Editable solo si es seguimiento Y tiene multisitios
    editable = es_seguimiento and has_multisitio
    
    # [MODIFICACIÓN] Traer los sitios para poder editarlos en el frontend
    sitios_rows = await conn.fetch("SELECT * FROM tb_sitios_oportunidad WHERE id_oportunidad = $1 ORDER BY nombre_sitio", id_oportunidad)
    
    # --- LOGICA DE CORREOS DINÁMICA (Desde tb_config_emails + Defaults) ---
    
    # 0. Defaults
    defaults_row = await conn.fetchrow("SELECT * FROM tb_email_defaults WHERE id = 1")
    # Parseamos a listas para el frontend (evitar nulos)
    def_to = (defaults_row['default_to'] or "").replace(";", ",").split(",") if defaults_row else []
    def_cc = (defaults_row['default_cc'] or "").replace(";", ",").split(",") if defaults_row else []
    # No pasamos CCO al frontend por privacidad/regla de negocio
    
    fixed_to = [d.strip() for d in def_to if d.strip()] 
    fixed_cc = [d.strip() for d in def_cc if d.strip()]

    # 1. Traer todas las reglas del módulo COMERCIAL
    rules = await conn.fetch("SELECT * FROM tb_config_emails WHERE modulo = 'COMERCIAL'")
    
    # 2. Evaluar reglas
    # MAPEO: Nombre Visual (Admin) -> Columna DB (Row)
    FIELD_MAPPING = {
        "Tecnología": "id_tecnologia",
        "Tipo Solicitud": "id_tipo_solicitud",
        "Estatus": "id_estatus_global",
        "Cliente": "cliente_nombre"
    }

    for rule in rules:
        field_admin = rule['trigger_field']    # e.g., 'Tecnología'
        val_trigger = str(rule['trigger_value']).strip().upper() # e.g., '1' (ID)
        
        # Mapear al campo real en la BD
        db_key = FIELD_MAPPING.get(field_admin, field_admin)
        
        # Obtener valor real de la oportunidad
        val_actual = row.get(db_key)
        
        # Lógica de comparación
        match = False
        
        if field_admin == "Cliente":
            # Búsqueda Parcial de Texto (Contains)
            if val_trigger in str(val_actual or "").upper():
                match = True
        else:
            # Comparación Exacta de ID (String vs String)
            # val_actual suele ser int (ID), val_trigger es str
            if str(val_actual or "") == val_trigger:
                match = True
        
        if match:
            email = rule['email_to_add']
            tipo = rule['type'] # TO o CC
            
            if tipo == 'TO':
                if email not in fixed_to: fixed_to.append(email)
            else:
                if email not in fixed_cc: fixed_cc.append(email)
    
    
    # 3. Determinar Template (HTMX vs Full Load)
    if request.headers.get("hx-request"):
        template = "comercial/email_form.html"
    else:
        template = "comercial/email_full.html" # Wrapper que extiende base.html

    # --- LOGICA DE FORMATEO DE OBJETIVOS BESS ---
    bess_objetivos_str = ""
    raw_objs = row.get('objetivos_json')
    if raw_objs:
        try:
            # Si asyncpg ya lo devolvió como lista
            if isinstance(raw_objs, list):
                bess_objetivos_str = ", ".join(raw_objs)
            # Si es string JSON
            elif isinstance(raw_objs, str):
                import json
                loaded = json.loads(raw_objs)
                if isinstance(loaded, list):
                    bess_objetivos_str = ", ".join(loaded)
        except Exception as e:
            logger.warning(f"Error parseando objetivos_json: {e}")
            pass

    return templates.TemplateResponse(
        template,
        {
            "request": request,
            "op": row,
            "bess_objetivos_str": bess_objetivos_str,  # <--- VARIABLE NUEVA
            "has_multisitio_file": has_multisitio,
            "sitios": sitios_rows,
            "editable": editable,              # Basado en BD, no hardcoded
            "is_followup": es_seguimiento,     # NUEVO: indica si es seguimiento
            "legacy_term": legacy_term,        # PUENTE: Término legacy para homologación
            "fixed_to": fixed_to,
            "fixed_cc": fixed_cc,
            # Contexto necesario para base.html en Full Load
            "user_name": request.session.get("user_name", "Usuario"),
            "role": request.session.get("role", "USER")
        }
    )

# ----------------------------------------
# NUEVOS ENDPOINTS PARA EXCEL PREVIEW
# ----------------------------------------

@router.post("/upload-preview", response_class=HTMLResponse)
async def upload_preview_endpoint(
    request: Request,
    id_oportunidad: str = Form(...), # Llega como string
    file: UploadFile = File(...),
    extraordinaria: int = Form(0),  # <-- Capturar del form
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection)
):
    try:
        # 1. Resetear puntero del archivo (CRÍTICO para reintentos)
        await file.seek(0)

        # 2. Validación de Seguridad: Tamaño Máximo (10MB)
        MAX_FILE_SIZE = 10 * 1024 * 1024
        file.file.seek(0, 2)  # Ir al final
        file_size = file.file.tell()  # Ver tamaño
        await file.seek(0)  # Volver al inicio - CRÍTICO
        
        if file_size > MAX_FILE_SIZE:
            logger.warning(f"Excel rechazado (excede 10MB): {file.filename} ({file_size} bytes)")
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Archivo muy grande",
                "message": "El archivo excede el tamaño máximo permitido de 10MB.",
                "action_btn": "removeFile(event)",
                "action_text": "Intentar de nuevo"
            })

        # 3. Validar extensión
        if not file.filename.endswith((".xlsx", ".xls")):
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Formato inválido",
                "message": "Solo se permiten archivos Excel (.xlsx o .xls).",
                "action_btn": "removeFile(event)",
                "action_text": "Intentar de nuevo"
            })

        # 4. Leer contenido en memoria (ahora es seguro)
        contents = await file.read()
        
        # Nota: io y openpyxl ya están importados al inicio del archivo
        
        # 4. Leer Excel con openpyxl (más ligero que pandas)
        try:
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
            ws = wb.active
            
            # Obtener cabeceras (primera fila)
            headers = [str(cell.value).strip().upper() for cell in ws[1] if cell.value]
            
            # Obtener datos (resto de filas)
            preview_rows = []
            full_data_list = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Crear diccionario mapeando header -> valor
                row_data = dict(zip(headers, row))
                
                # Filtrar filas completamente vacías
                if not any(row_data.values()):
                    continue
                
                # Limpieza básica (None -> "")
                clean_data = {k: (v if v is not None else "") for k, v in row_data.items()}
                
                preview_rows.append(list(clean_data.values()))  # Para la vista simple
                full_data_list.append(clean_data)  # Para el JSON
            
            columns = headers
            total_rows = len(full_data_list)
            
        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}")
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Error de archivo",
                "message": f"El archivo no es un Excel válido o está corrupto. ({e})"
            })

        # 5. VALIDACIÓN ESTRUCTURA (Columnas)
        cols_req = ["NOMBRE", "DIRECCION"]
        if not all(col in columns for col in cols_req):
            missing_cols = ', '.join([c for c in cols_req if c not in columns])
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Formato Incorrecto",
                "message": f"Faltan columnas requeridas: {missing_cols}. Usa la plantilla oficial.",
                "action_btn": "removeFile(event)",
                "action_text": "Intentar de nuevo"
            })

        # 6. VALIDACIÓN CANTIDAD
        # Convertimos el string id_oportunidad a UUID para la DB
        try:
            uuid_op = UUID(id_oportunidad)
        except ValueError:
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Error interno",
                "message": "ID de oportunidad inválido. Por favor recarga la página e intenta nuevamente."
            })

        expected_qty = await conn.fetchval(
            "SELECT cantidad_sitios FROM tb_oportunidades WHERE id_oportunidad = $1", 
            uuid_op
        )
        
        # Si por alguna razón no existe la oportunidad
        if expected_qty is None:
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Oportunidad no encontrada",
                "message": "La oportunidad no existe en la base de datos. Por favor verifica e intenta nuevamente."
            })

        real_qty = total_rows
        
        if real_qty != expected_qty:
            return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
                "request": request,
                "title": "Error de Cantidad",
                "message": f"Declaraste <strong>{expected_qty}</strong> sitios, pero el archivo tiene <strong>{real_qty}</strong> filas. Corrige el Excel y vuelve a seleccionarlo.",
                "action_btn": "removeFile(event)",
                "action_text": "Intentar de nuevo"
            })

        # 7. Generar Preview y Retornar Respuesta
        json_payload = json.dumps(full_data_list, default=str)
        
        return templates.TemplateResponse(
            "comercial/partials/upload_preview.html",
            {
                "request": request,
                "columns": columns,
                "preview_rows": preview_rows,
                "total_rows": total_rows,
                "json_data": json_payload,
                "op_id": id_oportunidad,
                "extraordinaria": extraordinaria  # <-- Pasar al template
            }
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"Error técnico en upload_preview_endpoint: {e}")
        return templates.TemplateResponse("comercial/partials/toasts/toast_error.html", {
            "request": request,
            "title": "Error técnico",
            "message": f"Ocurrió un error inesperado: {str(e)}. Por favor intenta nuevamente.",
            "action_btn": "removeFile(event)",
            "action_text": "Intentar de nuevo"
        })


@router.post("/upload-confirm", response_class=HTMLResponse)
async def upload_confirm_endpoint(
    request: Request,
    sitios_json: str = Form(...), # <--- Recibimos el JSON
    op_id: str = Form(...),
    extraordinaria: int = Form(0),  # <--- Nuevo parámetro
    conn = Depends(get_db_connection)
):
    try:
        uuid_op = UUID(op_id)
        
        # 1. Deserializar
        try:
            raw_data = json.loads(sitios_json)
        except json.JSONDecodeError:
             return HTMLResponse("<div class='text-red-500'>Error: Data corrupta (JSON inválido).</div>", 400)
             
        # 2. Borrar sitios anteriores (Idempotencia)
        await conn.execute("DELETE FROM tb_sitios_oportunidad WHERE id_oportunidad = $1", uuid_op)
        
        # 3. Preparar Registros
        records = []
        
        for item in raw_data:
            # Validación con Pydantic (Opcional pero recomendada para sanear)
            # Mapeamos los alias del Excel a los campos internos
            # Como Pydantic usa alias 'NOMBRE' -> 'nombre_sitio', le pasamos el dict tal cual
            try:
                # Validamos contra el schema. 
                # NOTA: SitioImportacion espera alias (NOMBRE, DIRECCION).
                # El json tiene las claves en mayúsculas (NOMBRE, DIRECCION...)
                sitio_obj = SitioImportacion(**item)
                
                records.append((
                    uuid4(), # id_sitio
                    uuid_op,
                    sitio_obj.nombre_sitio,
                    sitio_obj.direccion,
                    sitio_obj.tipo_tarifa,
                    sitio_obj.google_maps_link,
                    sitio_obj.numero_servicio,
                    sitio_obj.comentarios
                ))
            except Exception as e:
                logger.error(f"Error parseando fila: {item} -> {e}")
                # Podríamos fallar o saltar. Aquí saltamos filas malas.
                continue

        # 4. Insertar
        logger.info(f"Preparados {len(records)} registros para insertar en tb_sitios_oportunidad")
        q = """
            INSERT INTO tb_sitios_oportunidad (
                id_sitio, id_oportunidad, nombre_sitio, direccion, tipo_tarifa, google_maps_link, numero_servicio, comentarios
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
        """
        if records:
            await conn.executemany(q, records)
            logger.info(f"Insertados {len(records)} sitios exitosamente")
        
        # 5. Redirección Condicional
        if extraordinaria == 1:
            # EXTRAORDINARIAS: Redirección completa al HOME (no HTMX partial)
            logger.info("Solicitud extraordinaria multisitio completada, redirigiendo a HOME")
            return HTMLResponse(content=f"""
            <div class="bg-green-100 border-l-4 border-green-500 text-green-700 p-4 mt-4 animate-fade-in-down" role="alert">
                <p class="font-bold">Carga Exitosa</p>
                <p>Se confirmaron e insertaron {len(records)} sitios.</p>
                <p class="mt-2 text-sm">Redirigiendo al inicio...</p>
            </div>
            
            <!-- Redirección completa para extraordinarias -->
            <script>
                setTimeout(function() {{
                    window.location.href = '/comercial/ui';
                }}, 1500);
            </script>
            """, status_code=200)
        else:
            # NORMALES: Ir a paso 3 (envío de correo) usando HTMX
            next_url = f"/comercial/paso3/{op_id}"
            return HTMLResponse(content=f"""
            <div class="bg-green-100 border-l-4 border-green-500 text-green-700 p-4 mt-4 animate-fade-in-down" role="alert">
                <p class="font-bold">Carga Exitosa</p>
                <p>Se confirmaron e insertaron {len(records)} sitios.</p>
            </div>
            
            <!-- Transición automática a paso 3 -->
            <div hx-trigger="load delay:1s" hx-get="{next_url}" hx-target="#main-content"></div> 
            """, status_code=200)
        
    except Exception as e:
        logger.error(f"Error Confirm: {e}")
        return HTMLResponse(f"<div class='text-red-500'>Error confirmando carga: {e}</div>", 500)

@router.get("/paso2/{id_oportunidad}", include_in_schema=False)
async def get_paso_2_form(request: Request, id_oportunidad: UUID, extraordinaria: int = 0, conn = Depends(get_db_connection)):
    """Re-renderiza el formulario de carga multisitio (Paso 2)."""
    row = await conn.fetchrow(
        "SELECT id_interno_simulacion, titulo_proyecto, cliente_nombre, cantidad_sitios FROM tb_oportunidades WHERE id_oportunidad = $1", 
        id_oportunidad
    )
    if not row:
         return HTMLResponse("Oportunidad no encontrada", 404)
         
    return templates.TemplateResponse(
        "comercial/paso2.html",
        {
            "request": request,
            "oportunidad_id": id_oportunidad, 
            "nombre_cliente": row['cliente_nombre'],
            "id_interno": row['id_interno_simulacion'],
            "titulo_proyecto": row['titulo_proyecto'],
            "cantidad_declarada": row['cantidad_sitios'],
            "extraordinaria": extraordinaria
        }
    )

@router.post("/crear-seguimiento/{parent_id}")
async def crear_seguimiento(
    request: Request,
    parent_id: UUID,
    tipo_solicitud: str = Form(...), # "COTIZACION", "ACTUALIZACION"
    prioridad: str = Form(...),
    service: ComercialService = Depends(get_comercial_service),
    conn = Depends(get_db_connection),
    user_context = Depends(get_current_user_context)
):
    """Acción del Historial: Crea seguimiento y salta directo al correo."""
    if not user_context.get("email"): 
        return HTMLResponse(status_code=401)

    new_id = await service.create_followup_oportunidad(
        parent_id, tipo_solicitud, prioridad, conn, 
        user_context['user_db_id'], user_context['user_name']
    )
    
    # Salto directo al Paso 3 (El usuario ya no carga Excel)
    return HTMLResponse(headers={"HX-Location": f"/comercial/paso3/{new_id}"})

@router.delete("/sitios/{id_sitio}", response_class=HTMLResponse)
async def delete_sitio_endpoint(request: Request, id_sitio: UUID, conn = Depends(get_db_connection)):
    """Elimina un sitio específico (Usado en el filtrado de seguimiento)."""
    # Validar sesión con token inteligente
    access_token = await get_valid_graph_token(request)
    if not access_token:
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": "/auth/login?expired=1"})
    
    await conn.execute("DELETE FROM tb_sitios_oportunidad WHERE id_sitio = $1", id_sitio)
    # Retorna vacío para que HTMX elimine la fila de la tabla
    return HTMLResponse("", status_code=200)
    