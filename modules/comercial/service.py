from datetime import datetime, timedelta, date, time as dt_time
from uuid import UUID, uuid4
from typing import List, Optional, Tuple
import json
import logging
import asyncpg
from fastapi import HTTPException
from zoneinfo import ZoneInfo
import io

from openpyxl import load_workbook
from .schemas import SitioImportacion, DetalleBessCreate
from .constants import STATUS_PENDIENTE, DEFAULT_STATUS_ID_PENDIENTE
import asyncio
from core.config_service import ConfigService
from .db_service import (
    QUERY_GET_OPORTUNIDADES_LIST,
    QUERY_INSERT_OPORTUNIDAD,
    QUERY_INSERT_FOLLOWUP,
    QUERY_CLONE_SITIOS,
    QUERY_GET_TECNOLOGIAS,
    QUERY_GET_TIPOS_SOLICITUD,
    QUERY_GET_ESTATUS_GLOBAL,
    QUERY_GET_OPORTUNIDAD_OWNER,
    QUERY_GET_OPORTUNIDAD_FROM_SITIO,
    QUERY_UPDATE_OPORTUNIDAD_OWNER,
    QUERY_GET_USUARIOS_COMERCIAL,
    QUERY_GET_ALL_USUARIOS,
    QUERY_GET_TIPO_ACTUALIZACION_ID,
    QUERY_CHECK_USER_TOKEN,
    QUERY_GET_TECNOLOGIA_NAME,
    QUERY_GET_TIPO_SOLICITUD_ID_BY_CODE,
    QUERY_GET_TIPO_SOLICITUD_CODE,
    QUERY_GET_TIPO_SOLICITUD_NAME,
    QUERY_GET_USER_NAME,
    QUERY_GET_DETALLES_BESS,
    QUERY_GET_COMENTARIOS_WORKFLOW,
    QUERY_GET_CANTIDAD_SITIOS,
    QUERY_GET_TIPO_SOLICITUD_FROM_OP,
    QUERY_DELETE_SITIOS_OP,
    QUERY_INSERT_SITIO_BULK,
    QUERY_DELETE_SITIO,
    QUERY_GET_SITIOS_SIMPLE,
    QUERY_INSERT_SITIO_UNICO,
    QUERY_UPDATE_EMAIL_ENVIADO,
    QUERY_UPDATE_PRIORIDAD,
    QUERY_DELETE_OPORTUNIDAD,
    QUERY_DELETE_COMENTARIOS_WF,
    QUERY_DELETE_NOTIFICACIONES,
    QUERY_DELETE_DOCS,
    QUERY_DELETE_LEVANTAMIENTOS,
    QUERY_DELETE_BESS,
    QUERY_SEARCH_CLIENTES,
    # Fase 5: queries extraídas de inline
    QUERY_GET_CLIENTE_BY_ID,
    QUERY_GET_OLDEST_OP_BY_CLIENTE,
    QUERY_UPDATE_CLIENTE_ID_INTERNO,
    QUERY_GET_OPORTUNIDAD_FULL,
    QUERY_GET_PASO2_DATA,
    QUERY_GET_SITIO_IDS_BY_OP,
    QUERY_DELETE_SITIOS_BY_IDS,
    QUERY_RELINK_LEVANTAMIENTOS,
    QUERY_UPDATE_CANTIDAD_SITIOS,
    QUERY_COUNT_SITIOS_BY_OP,
    QUERY_GET_OP_ESTATUS,
    QUERY_UPDATE_OP_ESTATUS,
    QUERY_UPDATE_SITIOS_ESTATUS_BY_IDS,
    QUERY_UPDATE_SITIOS_ESTATUS_OTHERS,
    QUERY_UPDATE_SITIOS_ESTATUS_ALL,
)

# Shared Services
from modules.shared.services import IdGeneratorService, ClientService, BessService

# Sub-Services
from .services import DashboardService, NotificationService
from .sla_calculator import SLACalculator
from core.config_service import ConfigService
import logging

logger = logging.getLogger("ComercialModule")
from core.permissions import user_has_module_access

logger = logging.getLogger("ComercialModule")


# Constante para evitar magic strings
EVENTO_EXTRAORDINARIA = "EXTRAORDINARIA"

class ComercialService:
    """Encapsula la lógica de negocio del módulo Comercial."""

    def __init__(
        self,
        dashboard_service: Optional[DashboardService] = None,
        notification_service: Optional[NotificationService] = None
    ):
        self.dashboard_service = dashboard_service or DashboardService()
        self.notification_service = notification_service or NotificationService()
        
        # Shared Helpers
        # (Static methods don't need instantiation but we use them directly)

    async def should_show_popup(self, conn, user_email: str) -> bool:
        """
        Verifica si el usuario debe ver el popup comercial.
        Lee la lista de correos desde la configuración global.
        """
        if not user_email:
            return False
            
        # Obtener lista de emails (string separado por comas)
        targets_str = await ConfigService.get_global_config(conn, "COMERCIAL_POPUP_TARGETS", "")
        
        if not targets_str:
            return False
            
        # Normalizar y comparar
        targets = [e.strip().lower() for e in targets_str.split(",") if e.strip()]
        return user_email.lower() in targets

    
    def _handle_legacy_mode(self, search_term: Optional[str]):
        """
        Maneja la lógica de 'Modo Homologación' (continuidad de hilos de correo).
        Centraliza cualquier efecto secundario o logging relacionado.
        """
        if search_term:
            logger.info(f"MODO HOMOLOGACIÓN ACTIVADO: Búsqueda de hilo con término '{search_term}'")
            # Futura expansión: Guardar en tabla temporal, emitir evento, etc.

    async def get_zona_horaria_default(self, conn) -> ZoneInfo:
        """
        Lee ZONA_HORARIA_DEFAULT de la base de datos (ConfigService).
        Fallback: America/Mexico_City
        """
        tz_str = await ConfigService.get_global_config(conn, "ZONA_HORARIA_DEFAULT", "America/Mexico_City")
        try:
            return ZoneInfo(tz_str)
        except Exception:
             logger.error(f"Zona horaria inválida: {tz_str}, usando fallback.")
             return ZoneInfo("America/Mexico_City")
    
    async def get_current_datetime_mx(self, conn) -> datetime:
        """
        Obtiene la hora actual EXACTA respetando la configuración de zona horaria en BD.
        Fuente de verdad para timestamps.
        """
        zona_horaria = await self.get_zona_horaria_default(conn)
        return datetime.now(zona_horaria)

    
    @staticmethod
    def get_canal_from_user_name(user_name: str) -> str:
        """
        Genera el canal de venta basado en el nombre del usuario.
        
        Regla de negocio:
        - Toma las 2 PRIMERAS palabras significativas (>2 caracteres, sin puntos)
        - Formato: PRIMER_NOMBRE_PRIMER_APELLIDO
        - Si solo hay 1 palabra significativa: esa palabra
        - Si vacío o None: cadena vacía
        
        Ejemplos:
        - "Sharon V. Morales Perez" → "SHARON_MORALES" 
        - "Moises Jimenez" → "MOISES_JIMENEZ" 
        - "Admin" → "ADMIN" 
        """
        parts = (user_name or "").strip().split()
        
        # Filtrar palabras significativas (más de 2 caracteres, sin puntos)
        meaningful_parts = [
            p.replace('.', '') for p in parts 
            if len(p.replace('.', '')) > 2
        ]
        
        if len(meaningful_parts) >= 2:
            # Tomar las 2 PRIMERAS palabras significativas (no primera y última)
            return f"{meaningful_parts[0]}_{meaningful_parts[1]}".upper()
        elif len(meaningful_parts) == 1:
            return meaningful_parts[0].upper()
        elif len(parts) >= 2:
            # Fallback si no hay palabras significativas: usar primera_segunda original
            return f"{parts[0]}_{parts[1]}".upper()
        elif len(parts) == 1:
            return parts[0].upper()
        else:
            return ""

    @staticmethod
    def is_originally_multisite(row: dict) -> bool:
        """
        Determina si una oportunidad fue concebida como multisitio,
        incluso si actualmente tiene 1 o 0 sitios activos (por borrado).
        
        Heurística:
        1. Cantidad Sitios > 1 (Obvio)
        2. Naming Convention del ID Interno ({BASE}_{PROYECTO})
           El ID de multisitio SIEMPRE sufija el nombre del proyecto.
           El ID de unisitio NO sufija el nombre del proyecto (usa el BASE generado).
        """
        # 1. Chequeo directo de cantidad (Active count)
        if (row.get('cantidad_sitios') or 0) > 1:
            return True
            
        # 2. Heurística de Nombre (Fallback para cuando borran sitios hasta quedar 1)
        try:
            proj_name = (row.get('nombre_proyecto') or "").strip().upper()
            id_interno = (row.get('id_interno_simulacion') or "").strip().upper()
            
            if not proj_name or not id_interno:
                return False
                
            # Si el ID interno termina con _NOMBRE_PROYECTO, es estructura multisitio
            # (A menos que el nombre del proyecto sea vacío, validado arriba)
            suffix = f"_{proj_name}"
            return id_interno.endswith(suffix)
        except Exception:
            return False

    @staticmethod
    def build_bess_detail(
        uso_sistema: List[str],
        cargas_criticas: Optional[float],
        tiene_motores: bool,
        potencia_motor: Optional[float],
        tiempo_autonomia: Optional[str],
        voltaje_operacion: Optional[str],
        cargas_separadas: bool,
        tiene_planta_emergencia: bool
    ) -> Optional[DetalleBessCreate]:
        """
        Factory method para construir el objeto DetalleBessCreate.
        Centraliza la lógica de conversión de form -> objeto.
        """
        if not uso_sistema:
            return None
            
        return DetalleBessCreate(
            uso_sistema_json=uso_sistema,
            cargas_criticas_kw=cargas_criticas,
            tiene_motores=tiene_motores,
            potencia_motor_hp=potencia_motor,
            tiempo_autonomia=tiempo_autonomia,
            voltaje_operacion=voltaje_operacion,
            cargas_separadas=cargas_separadas,
            tiene_planta_emergencia=tiene_planta_emergencia
        )

    async def get_configuracion_global(self, conn):
        """Obtiene la configuración de horarios desde la BD (usando cache)."""
        # Obtenemos valores individuales cacheados
        hora_corte = await ConfigService.get_global_config(conn, "HORA_CORTE_L_V", "18:00")
        return {"HORA_CORTE_L_V": hora_corte}


    async def get_catalog_ids(self, conn) -> dict:
        """
        Carga IDs de catálogos para filtros rápidos.
        Usa caché centralizado en ConfigService (30s).
        NOTA: asyncpg no soporta operaciones concurrentes en la misma conexión,
        por lo que se deben esperar secuencialmente.
        """
        # Ejecución secuencial para evitar "InterfaceError: cannot perform operation: another operation is in progress"
        estatus_map = await ConfigService.get_catalog_map(conn, "tb_cat_estatus_global", "nombre", "id")
        tipos_map = await ConfigService.get_catalog_map(conn, "tb_cat_tipos_solicitud", "codigo_interno", "id")
        
        return {
            "estatus": estatus_map,
            "tipos": tipos_map
        }


    async def calcular_fuera_de_horario(self, conn, fecha_creacion: datetime) -> bool:
        """
        Valida si la fecha dada cae fuera del horario laboral configurado.
        Delegación: SLACalculator.
        """
        config = await self.get_configuracion_global(conn)
        hora_corte, dias_fin_semana, _ = SLACalculator.parse_config(config)
        return SLACalculator.is_out_of_hours(fecha_creacion, hora_corte, dias_fin_semana)

    async def calcular_deadline_inicial(self, conn, fecha_creacion: datetime) -> datetime:
        """
        Calcula el deadline inicial (Meta).
        Delegación: SLACalculator.
        """
        config = await self.get_configuracion_global(conn)
        hora_corte, _, dias_sla = SLACalculator.parse_config(config)
        
        return SLACalculator.calculate_deadline(fecha_creacion, hora_corte, dias_sla)

    async def get_catalogos_ui(self, conn) -> dict:
        """Recupera los catálogos para llenar los <select> del formulario y filtros."""
        # Cache Strategy (5 min TTL)
        cache_key = "COMERCIAL_UI_CATALOGS"
        cached = await ConfigService.get_cached_value(cache_key, ttl=300.0)
        if cached:
            return cached

        tecnologias = await conn.fetch(QUERY_GET_TECNOLOGIAS)
        tipos = await conn.fetch(QUERY_GET_TIPOS_SOLICITUD)
        estatus = await conn.fetch(QUERY_GET_ESTATUS_GLOBAL)
        usuarios = await conn.fetch(QUERY_GET_USUARIOS_COMERCIAL)
        
        data = {
            "tecnologias": [dict(t) for t in tecnologias],
            "tipos_solicitud": [dict(t) for t in tipos],
            "estatus_global": [dict(t) for t in estatus],
            "usuarios": [dict(t) for t in usuarios]
        }
        
        await ConfigService.set_cached_value(cache_key, data)
        return data

    async def get_catalogos_creacion(self, conn, include_simulacion: bool = False) -> dict:
        """
        Carga catálogos filtrados específicamente para el Formulario de Creación (Paso 1).
        
        Args:
            include_simulacion: Si True, incluye 'SIMULACION' en la lista (para extraordinarias).
                              Si False, solo 'PRE_OFERTA' y 'LICITACION' (para normal).
        """
        # Tecnologías (Todas)
        tecnologias = await conn.fetch(QUERY_GET_TECNOLOGIAS)
        
        # Tipos de Solicitud (Filtrado Dinámico)
        # Base: PRE_OFERTA, LEVANTAMIENTO ('LICITACION' removido, ahora es flag transversal)
        codigos = ['PRE_OFERTA', 'LEVANTAMIENTO']
        
        if include_simulacion:
            codigos.append('SIMULACION')
            
        # Generar placeholders para la query IN ($1, $2, ...)
        placeholders = ",".join([f"${i+1}" for i in range(len(codigos))])
        
        tipos = await conn.fetch(f"""
            SELECT id, nombre, codigo_interno 
            FROM tb_cat_tipos_solicitud 
            WHERE activo = true 
            AND codigo_interno IN ({placeholders})
            ORDER BY nombre
        """, *codigos)
        
        # Usuarios (Para delegación)
        usuarios = await conn.fetch(QUERY_GET_ALL_USUARIOS)

        return {
            "tecnologias": [dict(t) for t in tecnologias],
            "tipos_solicitud": [dict(t) for t in tipos],
            "usuarios": [dict(u) for u in usuarios]
        }

    async def get_catalogos_extraordinario(self, conn) -> dict:
        """
        Carga catálogos filtrados específicamente para el Formulario Extraordinario.
        Solo incluye PRE_OFERTA y SIMULACION (NO incluye LEVANTAMIENTO).
        
        Returns:
            dict con tecnologias, tipos_solicitud (solo PRE_OFERTA y SIMULACION), y usuarios
        """
        # Tecnologías (Todas)
        tecnologias = await conn.fetch(QUERY_GET_TECNOLOGIAS)
        
        # Tipos de Solicitud (PRE_OFERTA, SIMULACION, CAPTURA_RECIBOS)
        codigos = ['PRE_OFERTA', 'SIMULACION', 'CAPTURA_RECIBOS']
        placeholders = ",".join([f"${i+1}" for i in range(len(codigos))])
        
        tipos = await conn.fetch(f"""
            SELECT id, nombre, codigo_interno 
            FROM tb_cat_tipos_solicitud 
            WHERE activo = true 
            AND codigo_interno IN ({placeholders})
            ORDER BY nombre
        """, *codigos)
        
        # Usuarios (Para delegación)
        usuarios = await conn.fetch(QUERY_GET_ALL_USUARIOS)

        return {
            "tecnologias": [dict(t) for t in tecnologias],
            "tipos_solicitud": [dict(t) for t in tipos],
            "usuarios": [dict(u) for u in usuarios]
        }

    async def get_id_tipo_actualizacion(self, conn) -> Optional[int]:
        """
        Obtiene el ID del tipo de solicitud 'ACTUALIZACION' desde catálogo.
        Usado en modo homologación para forzar tipo de solicitud.
        
        Returns:
            ID del tipo ACTUALIZACION o None si no existe
        """
        tipo_act = await conn.fetchrow(QUERY_GET_TIPO_ACTUALIZACION_ID)
        return tipo_act['id'] if tipo_act else None

    async def procesar_fecha_manual(self, conn, fecha_input_str: Optional[str]) -> datetime:
        """
        Regla de Negocio: Solicitudes Extraordinarias (Gerentes).
        Input: "2025-10-20T10:00" (String ISO del navegador, usualmente naive).
        Output: Datetime con timezone America/Mexico_City.
        """
        zona_mx = await self.get_zona_horaria_default(conn)
        
        if not fecha_input_str:
            return datetime.now(zona_mx)
            
        try:
            # 1. Parsear string ISO
            dt = datetime.fromisoformat(fecha_input_str)
            
            # 2. Asignar Zona Horaria
            # El input datetime-local NO envía zona horaria, asumimos que el gerente
            # está capturando la hora de CDMX.
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=zona_mx)
            else:
                # Si viniera con zona, convertimos a CDMX
                dt = dt.astimezone(zona_mx)
                
            return dt
        except ValueError:
            logger.error(f"Fecha manual inválida: {fecha_input_str}, usando NOW()")
            return datetime.now(zona_mx)

    async def verify_ownership(self, conn, id_oportunidad: UUID, user_context: dict) -> None:
        """
        Verifica que el usuario sea dueño de la oportunidad o tenga rol de MANAGER/ADMIN.
        Raises HTTPException 403 si no tiene permiso.
        """
        # Refactor: Use centralized module permission check
        # "admin" module role (or Global ADMIN/MANAGER+Admin) allows bypassing ownership
        if user_has_module_access("comercial", user_context, "admin"):
            return

        owner_id = await conn.fetchval(QUERY_GET_OPORTUNIDAD_OWNER, id_oportunidad)

        if owner_id is None:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada.")

        user_id = user_context.get("user_db_id")
        if str(owner_id) != str(user_id):
            logger.warning(f"ACCESS DENIED: User {user_id} tried to access op {id_oportunidad} owned by {owner_id}")
            raise HTTPException(status_code=403, detail="No tiene permiso para acceder a esta oportunidad.")



    async def crear_oportunidad_transaccional(self, conn, datos, user_context: dict, legacy_search_term: Optional[str] = None) -> tuple:
        """
        Orquestador Transaccional (Fase 2).
        Maneja: Fechas, Identificadores, Oportunidad Base y BESS.
        
        Args:
            conn: Conexión asyncpg
            datos: OportunidadCreateCompleta (Pydantic v2)
            user_context: dict con user_db_id, user_name, role
            legacy_search_term: Término de búsqueda legacy para modo homologación (opcional)
            
        Returns:
            tuple: (new_id, op_id_estandar, es_fuera_horario)
        """
        logger.info(f"Iniciando creación de oportunidad para cliente {datos.cliente_nombre}")
        
        fecha_solicitud = await self.procesar_fecha_manual(conn, datos.fecha_manual_str)
        es_fuera_horario = await self.calcular_fuera_de_horario(conn, fecha_solicitud)
        deadline = await self.calcular_deadline_inicial(conn, fecha_solicitud)

        # BUSINESS LOGIC: Modo Homologación
        self._handle_legacy_mode(legacy_search_term)

        # Generar Identificadores Temporales (Para calcular)
        new_id = uuid4()
        now_mx = await self.get_current_datetime_mx(conn)
        op_id_estandar = IdGeneratorService.generate_standard_op_id(now_mx)
        
        # GESTIÓN INTELIGENTE DE CLIENTES (Shared Service)
        final_cliente_id = datos.cliente_id
        final_cliente_nombre = datos.cliente_nombre.strip().upper()

        if not final_cliente_id:
            # Caso 1: Cliente Nuevo (o no seleccionado del dropdown)
            # Intentamos buscarlo o crearlo.
            # Pasamos 'op_id_estandar' + Name como posible ID Maestro si resulta ser nuevo.
            potential_master_id = IdGeneratorService.generate_internal_id(
                op_id_estandar, final_cliente_nombre, "", 1
            )
            
            final_cliente_id, final_cliente_nombre, frozen_id = await ClientService.get_or_create_client_by_name(
                conn, final_cliente_nombre, initial_id_interno=potential_master_id
            )
        else:
            # Caso 2: Cliente Seleccionado (Explicit ID)
            # Solo recuperamos sus datos para ver si ya tiene ID congelado
            row = await conn.fetchrow(QUERY_GET_CLIENTE_BY_ID, final_cliente_id)
            if row:
                final_cliente_nombre = row['nombre_fiscal']
                frozen_id = row['id_interno_simulacion']

                # LÓGICA DE RECUPERACIÓN (BACKFILL)
                # Si el cliente existe pero NO tiene ID Interno (Legacy Bug), lo arreglamos ahora.
                if not frozen_id:
                    logger.info(f"Cliente {final_cliente_id} sin ID Interno. Iniciando recuperación histórica...")
                    
                    # 1. Buscar la oportunidad MÁS ANTIGUA de este cliente para heredar su fecha
                    oldest_op = await conn.fetchrow(QUERY_GET_OLDEST_OP_BY_CLIENTE, final_cliente_id)
                    
                    if oldest_op and oldest_op['op_id_estandar']:
                        # Reconstruir ID Ancestral
                        frozen_id = IdGeneratorService.generate_internal_id(
                            oldest_op['op_id_estandar'], final_cliente_nombre, "", 1
                        )
                        logger.info(f"RECUPERADO HISTÓRICO: Usando '{frozen_id}' basado en op antigua {oldest_op['op_id_estandar']}")
                    else:
                        # Cliente sin historial previo (o error de datos), generamos uno nuevo HOY.
                        frozen_id = IdGeneratorService.generate_internal_id(
                             op_id_estandar, final_cliente_nombre, "", 1
                        )
                        logger.info(f"SIN HISTORIAL: Generando nuevo ID Interno '{frozen_id}'")

                    # 2. GUARDAR EN EL CLIENTE (Persistencia futura)
                    await conn.execute(QUERY_UPDATE_CLIENTE_ID_INTERNO, frozen_id, final_cliente_id)
                    logger.info("BACKFILL EXITOSO: Cliente actualizado.")

            else:
                frozen_id = None # Should not happen if client exists

        # LÓGICA DE ID PERSEVERANTE
        # Si el cliente ya tiene un ID maestro (frozen_id), LO USAMOS como base reemplazando al op_id_estandar generado por tiempo.
        # Si no tiene (caso raro o legacy sin migrar), usamos el op_id_estandar actual.
        
        base_id_para_proyecto = frozen_id if frozen_id else IdGeneratorService.generate_internal_id(op_id_estandar, final_cliente_nombre, "", 1)

        # Generar ID Final de la Oportunidad
        # Si es multisitio, le agregamos el nombre del proyecto. Si es unisitio, se queda con el base.
        if datos.cantidad_sitios > 1:
             # Multisitio: BASE + PROYECTO
             clean_proyecto = (datos.nombre_proyecto or "").strip().upper()
             id_interno = f"{base_id_para_proyecto}_{clean_proyecto}"[:150]
        else:
             # Unisitio: BASE (Es igual al del cliente)
             id_interno = base_id_para_proyecto

        # Obtener nombres de catálogos
        nombre_tec = await conn.fetchval(QUERY_GET_TECNOLOGIA_NAME, datos.id_tecnologia)
        nombre_tipo = await conn.fetchval(QUERY_GET_TIPO_SOLICITUD_NAME, datos.id_tipo_solicitud)
        
        # Generar Títulos Legacy (Standardized)
        titulo = IdGeneratorService.generate_project_title(
            nombre_tipo, final_cliente_nombre, datos.nombre_proyecto, nombre_tec, datos.canal_venta
        )

        # Insertar Oportunidad
        
        # Obtener ID de estatus inicial
        cats = await self.get_catalog_ids(conn)
        # Use constant key for lookup, fallback to constant ID
        id_status_inicial = cats['estatus'].get(STATUS_PENDIENTE) or DEFAULT_STATUS_ID_PENDIENTE

        
        # Lógica de Solicitado Por
        solicitado_por_nombre = user_context.get('user_name', 'Usuario')
        if datos.solicitado_por_id:
             solicitado_por_nombre = await conn.fetchval(QUERY_GET_USER_NAME, datos.solicitado_por_id) or solicitado_por_nombre


        query_op = QUERY_INSERT_OPORTUNIDAD
        es_manual = bool(datos.fecha_manual_str)


        await conn.execute(query_op, 
            new_id,                 # $1: id_oportunidad
            op_id_estandar,         # $2: op_id_estandar
            id_interno,             # $3: id_interno_simulacion
            titulo,                 # $4: titulo_proyecto
            datos.nombre_proyecto,  # $5: nombre_proyecto
            final_cliente_nombre,   # $6: cliente_nombre
            datos.canal_venta,      # $7: canal_venta
            datos.id_tecnologia,    # $8: id_tecnologia
            datos.id_tipo_solicitud,# $9: id_tipo_solicitud
            datos.cantidad_sitios,  # $10: cantidad_sitios
            datos.prioridad,        # $11: prioridad
            datos.direccion_obra,   # $12: direccion_obra
            datos.coordenadas_gps,  # $13: coordenadas_gps
            datos.google_maps_link, # $14: google_maps_link
            datos.sharepoint_folder_url, # $15: sharepoint_folder_url
            user_context['user_db_id'],  # $16: creado_por_id
            fecha_solicitud,        # $17: fecha_solicitud
            es_fuera_horario,       # $18: es_fuera_horario
            deadline,               # $19: deadline_calculado
            solicitado_por_nombre,  # $20: solicitado_por
            es_manual,              # $21: es_carga_manual
            id_status_inicial,      # $22: id_estatus_global
            datos.clasificacion_solicitud, # $23: clasificacion_solicitud
            datos.solicitado_por_id,       # $24: solicitado_por_id
            datos.es_licitacion,           # $25: es_licitacion
            final_cliente_id,        # $26: cliente_id
            datos.fecha_ideal_usuario # $27: fecha_ideal_usuario
        )

        # Insertar BESS (Shared Service)
        if datos.detalles_bess:
            await BessService.create_bess_details(conn, new_id, datos.detalles_bess)

        # 6. Insertar Historial Estatus Inicial (Reemplazo de Trigger)
        # Calculamos la fecha SLA usando la misma lógica que el deadline inicial pero con offset 0
        # O simplemente usamos el deadline inicial si asumimos que el primer paso consume tiempo?
        # NO, fecha_cambio_sla es la FECHA DE INICIO del estatus ajustada a horario laboral.
        # Es decir, si se crea sábado, cuenta desde lunes.
        
        config = await self.get_configuracion_global(conn)
        hora_corte, dias_fin_semana, _ = SLACalculator.parse_config(config)
        
        # Ajustar fecha de inicio si es fuera de horario (Lógica SLACalculator.adjust_start_date interna o similar)
        # Dado que calculate_deadline hace "start + dias", si usamos dias=0 obtenemos el start ajustado.
        fecha_inicio_sla = SLACalculator.calculate_deadline(fecha_solicitud, hora_corte, 0)
        
        from .db_service import QUERY_INSERT_HISTORIAL_ESTATUS
        await conn.execute(QUERY_INSERT_HISTORIAL_ESTATUS,
            new_id,             # $1 id_oportunidad
            None,               # $2 id_estatus_anterior (NULL al inicio)
            id_status_inicial,  # $3 id_estatus_nuevo
            fecha_solicitud,    # $4 fecha_cambio_real
            fecha_inicio_sla,   # $5 fecha_cambio_sla
            user_context['user_db_id'] # $6 id_responsable
        )
        
        # ========================================
        # HOOK: Crear levantamiento automáticamente si es tipo LEVANTAMIENTO
        # ========================================
        tipo_datos = await conn.fetchrow(
            QUERY_GET_TIPO_SOLICITUD_CODE,
            datos.id_tipo_solicitud
        )
        if tipo_datos and tipo_datos['codigo_interno'] == 'LEVANTAMIENTO':
            try:
                from modules.levantamientos.service import LevantamientoService
                lev_service = LevantamientoService()
                lev_id = await lev_service.crear_desde_oportunidad(conn, new_id, user_context)
                logger.info(f"Levantamiento {lev_id} creado automáticamente para oportunidad {op_id_estandar}")
            except Exception as e:
                logger.error(f"Error creando levantamiento automático: {e}")
                # No fallar la creación de oportunidad por esto
        
        # Site Unico (Si aplica, pasar id_tipo_solicitud)
        if datos.cantidad_sitios == 1:
            #Evitar doble creación para LEVANTAMIENTO (Ya lo creó el hook arriba)
            es_levantamiento = (tipo_datos and tipo_datos['codigo_interno'] == 'LEVANTAMIENTO')
            
            if not es_levantamiento:
                await self.auto_crear_sitio_unico(
                    conn, new_id, 
                    datos.nombre_proyecto, 
                    datos.direccion_obra, 
                    datos.google_maps_link,
                    datos.id_tipo_solicitud,
                    id_status_inicial
                )
        
        logger.info(f"Oportunidad {op_id_estandar} creada exitosamente por usuario {user_context.get('user_db_id')}")
        return new_id, op_id_estandar, es_fuera_horario




    async def get_oportunidad_for_email(self, conn, id_oportunidad: UUID, user_context: dict) -> Optional[dict]:
        await self.verify_ownership(conn, id_oportunidad, user_context)
        return await self.notification_service.get_oportunidad_for_email(conn, id_oportunidad)

    async def get_parent_titulo(self, conn, parent_id: UUID) -> Optional[str]:
        return await self.notification_service.get_parent_titulo(conn, parent_id)
    
    async def get_email_threading_context(self, conn, row: dict, legacy_search_term: Optional[str] = None) -> dict:
        return await self.notification_service.get_email_threading_context(conn, row, legacy_search_term)

    async def get_oportunidades_list(
        self, 
        conn, 
        user_context: dict, 
        tab: str = "activos", 
        q: str = None, 
        limit: int = 15, 
        subtab: str = None,
        # Nuevos filtros globales
        filtro_usuario_id: Optional[UUID] = None,
        filtro_tipo_id: Optional[int] = None,
        filtro_estatus_id: Optional[int] = None,
        filtro_tecnologia_id: Optional[int] = None,
        filtro_fecha_inicio: Optional[str] = None, # YYYY-MM-DD
        filtro_fecha_fin: Optional[str] = None     # YYYY-MM-DD
    ) -> List[dict]:
        """Recupera lista filtrada de oportunidades con permisos y paginación."""
        user_id = user_context.get("user_db_id")  # CORREGIDO: era "user_id"
        role = user_context.get("role", "USER")
        
        logger.debug(f"Consultando oportunidades - Tab: {tab}, Filtro: {q}, Usuario: {user_id}")
        
        # Parse Dates if strings
        if filtro_fecha_inicio and isinstance(filtro_fecha_inicio, str):
            try:
                filtro_fecha_inicio = datetime.strptime(filtro_fecha_inicio, '%Y-%m-%d').date()
            except ValueError:
                logger.warning(f"Fecha inicio inválida ignorada: {filtro_fecha_inicio}")
                filtro_fecha_inicio = None

        if filtro_fecha_fin and isinstance(filtro_fecha_fin, str):
            try:
                filtro_fecha_fin = datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                logger.warning(f"Fecha fin inválida ignorada: {filtro_fecha_fin}")
                filtro_fecha_fin = None

        # Cargar IDs de catálogos
        cats = await self.get_catalog_ids(conn)



        query = QUERY_GET_OPORTUNIDADES_LIST
        
        params = []
        param_idx = 1

        # --- Lógica de Filtros Globales ---
        if filtro_usuario_id:
            query += f" AND o.creado_por_id = ${param_idx}"
            params.append(filtro_usuario_id)
            param_idx += 1
            
        if filtro_tipo_id:
            query += f" AND o.id_tipo_solicitud = ${param_idx}"
            params.append(filtro_tipo_id)
            param_idx += 1
            
        if filtro_estatus_id:
            query += f" AND o.id_estatus_global = ${param_idx}"
            params.append(filtro_estatus_id)
            param_idx += 1
            
        if filtro_tecnologia_id:
            query += f" AND o.id_tecnologia = ${param_idx}"
            params.append(filtro_tecnologia_id)
            param_idx += 1
            
        if filtro_fecha_inicio:
            query += f" AND (o.fecha_solicitud AT TIME ZONE 'America/Mexico_City')::date >= ${param_idx}::date"
            params.append(filtro_fecha_inicio)
            param_idx += 1
            
        if filtro_fecha_fin:
            # +1 day para incluir todo el día final si es timestamp, o cast a date
            query += f" AND (o.fecha_solicitud AT TIME ZONE 'America/Mexico_City') < (${param_idx}::date + INTERVAL '1 day')"
            params.append(filtro_fecha_fin)
            param_idx += 1

        # Filtro por tab (Usa IDs)
        if tab == "historial": # Renombrado en UI a "Solicitudes (Entregadas)"
            # Subtabs para historial: entregado vs cancelado-perdido
            if not subtab or subtab == 'entregado':
                # Por defecto: solo Entregado
                id_entregado = cats['estatus'].get('entregado')
                if id_entregado:
                    query += f" AND o.id_estatus_global = ${param_idx}"
                    params.append(id_entregado)
                    param_idx += 1
            elif subtab == 'cancelado_perdido':
                # Alternativa: Cancelado + Perdido
                ids_fallidos = [
                    cats['estatus'].get('cancelado'),
                    cats['estatus'].get('perdido')
                ]
                ids_fallidos = [i for i in ids_fallidos if i is not None]
                if ids_fallidos:
                    placeholders = ','.join([f'${i}' for i in range(param_idx, param_idx + len(ids_fallidos))])
                    query += f" AND o.id_estatus_global IN ({placeholders})"
                    params.extend(ids_fallidos)
                    param_idx += len(ids_fallidos)
                
        elif tab == "levantamientos":
            # Buscar ID del tipo "levantamiento"
            id_levantamiento = cats['tipos'].get('levantamiento')
            if id_levantamiento:
                query += f" AND o.id_tipo_solicitud = ${param_idx}"
                params.append(id_levantamiento)
                param_idx += 1
                
            # Sub-filtro por subtab
            if subtab == 'realizados':
                # Realizados: Completado (11), Entregado (12)
                # Se filtra por el estatus del LEVANTAMIENTO (lev.id_estatus_global)
                ids_realizados = [11, 12]
                placeholders = ','.join([f'${param_idx + i}' for i in range(len(ids_realizados))])
                query += f" AND lev.id_estatus_global IN ({placeholders})"
                params.extend(ids_realizados)
                param_idx += len(ids_realizados)

            else:
                # Solicitados (Default): Pendiente (8), Agendado (9), En Proceso (10), Pospuesto (13)
                ids_solicitados = [8, 9, 10, 13]
                placeholders = ','.join([f'${param_idx + i}' for i in range(len(ids_solicitados))])
                query += f" AND lev.id_estatus_global IN ({placeholders})"
                params.extend(ids_solicitados)
                param_idx += len(ids_solicitados)
                    
        elif tab == "ganadas":
            id_ganada = cats['estatus'].get('ganada')
            if id_ganada:
                query += f" AND o.id_estatus_global = ${param_idx}"
                params.append(id_ganada)
                param_idx += 1
                
        else:  # activos
            # Estados ACTIVOS - Inclusión explícita
            ids_activos = [
                cats['estatus'].get('pendiente'),
                cats['estatus'].get('en revisión'),  # Con tilde según BD
                cats['estatus'].get('en proceso')
            ]
            ids_activos = [i for i in ids_activos if i is not None]
            
            # Si el usuario seleccionó un estatus específico en el filtro global, 
            # ese ya se aplicó arriba. PERO para la pestaña 'activos', 
            # debemos asegurarnos que SOLO se muestren los activos.
            # Si el filtro global es 'Entregado', esta pestaña mostrará vacío, lo cual es correcto.
            # Pero para evitar conflictos lógicos:
            # 1. Si NO hay filtro de estatus global, aplicamos el filtro por defecto de la pestaña.
            # 2. Si HAY filtro de estatus global, verificamos si es compatible con la pestaña (opcional, o dejamos que SQL filtre).
            #    Dejaremos que SQL filtre: (Global=X) AND (TabLimit IN (X,Y,Z)). Si X no está en TabLimit, retorna 0. Correcto.
            
            if ids_activos:
                placeholders = ','.join([f'${i}' for i in range(param_idx, param_idx + len(ids_activos))])
                query += f" AND o.id_estatus_global IN ({placeholders})"
                params.extend(ids_activos)
                param_idx += len(ids_activos)
                
            # Excluir levantamientos de activos
            id_levantamiento = cats['tipos'].get('levantamiento')
            if id_levantamiento:
                query += f" AND o.id_tipo_solicitud != ${param_idx}"
                params.append(id_levantamiento)
                param_idx += 1

        # Búsqueda
        if q:
            query += f" AND (o.titulo_proyecto ILIKE ${param_idx} OR o.nombre_proyecto ILIKE ${param_idx} OR o.cliente_nombre ILIKE ${param_idx})"
            params.append(f"%{q}%")
            param_idx += 1

        # Filtro de seguridad (solo ADMIN, MANAGER ven todo)
        # Filtro de seguridad:
        # Si NO es admin del módulo (ni global), solo ve sus propias oportunidades
        if not user_has_module_access("comercial", user_context, "admin"):
            query += f" AND o.creado_por_id = ${param_idx}"
            params.append(user_id)
            param_idx += 1

        query += " ORDER BY o.fecha_solicitud DESC"
        
        if limit > 0:
            query += f" LIMIT {limit}"
        
        rows = await conn.fetch(query, *params)
        
        logger.debug(f"Retornando {len(rows)} oportunidades")
        return [dict(row) for row in rows]


    async def update_email_status(self, conn, id_oportunidad: UUID, user_context: dict):
        await self.verify_ownership(conn, id_oportunidad, user_context)
        await self.notification_service.update_email_status(conn, id_oportunidad)
    
    async def update_oportunidad_prioridad(
        self, 
        conn, 
        id_oportunidad: UUID, 
        prioridad: str,
        user_context: dict
    ) -> None:
        """
        Updates the priority of an opportunity.
        
        Args:
            conn: Database connection
            id_oportunidad: UUID of the opportunity
            prioridad: Priority value ('normal', 'alta', 'baja')
            user_context: Context for ownership check
        """
        await self.verify_ownership(conn, id_oportunidad, user_context)
        await conn.execute(
            QUERY_UPDATE_PRIORIDAD,
            prioridad,
            id_oportunidad
        )
        logger.info(f"Prioridad actualizada a '{prioridad}' para oportunidad {id_oportunidad}")

    async def reasignar_oportunidad(self, conn, id_oportunidad: UUID, new_owner_id: UUID, user_context: dict) -> None:
        """
        Transfiere la propiedad de una oportunidad a otro usuario.
        Solo permitido para MANAGER o ADMIN.
        """
        # Validación: ADMIN global, admin del módulo, o MANAGER + editor del módulo
        role = user_context.get("role")
        com_role = user_context.get("module_roles", {}).get("comercial", "")
        can_reassign = (
            role == "ADMIN" or
            com_role == "admin" or
            (role == "MANAGER" and com_role in ["editor", "admin"])
        )
        if not can_reassign:
             raise HTTPException(status_code=403, detail="No tienes permisos para reasignar oportunidades.")

        await conn.execute(QUERY_UPDATE_OPORTUNIDAD_OWNER, new_owner_id, id_oportunidad)
        logger.info(f"Oportunidad {id_oportunidad} reasignada a {new_owner_id} por {user_context.get('user_db_id')}")
    
    async def check_user_has_access_token(
        self, 
        conn, 
        user_db_id: UUID
    ) -> bool:
        """
        Checks if a user has a valid access token stored.
        
        Args:
            conn: Database connection
            user_db_id: User's database ID
            
        Returns:
            True if user has an access token, False otherwise
        """
        has_token = await conn.fetchval(
            QUERY_CHECK_USER_TOKEN,
            user_db_id
        )
        return has_token or False

    async def create_followup_oportunidad(self, parent_id: UUID, nuevo_tipo_solicitud: str, prioridad: str, conn, user_id: UUID, user_name: str) -> UUID:
        """Crea seguimiento clonando padre + sitios."""
        
        # Fuente de verdad temporal (Corrección Zona Horaria)
        # Obtenemos la hora con timezone de México. Asyncpg la convertirá a UTC al guardar.
        now_mx = await self.get_current_datetime_mx(conn)

        parent = await conn.fetchrow(QUERY_GET_OPORTUNIDAD_FULL, parent_id)
        if not parent: 
            raise HTTPException(status_code=404, detail="Oportunidad original no encontrada")

        # Convertir string de tipo_solicitud a ID
        # El parámetro nuevo_tipo_solicitud viene como "OFERTA_FINAL", "ACTUALIZACION", etc.
        id_tipo_solicitud = await conn.fetchval(
            QUERY_GET_TIPO_SOLICITUD_ID_BY_CODE,
            nuevo_tipo_solicitud
        )
        
        if not id_tipo_solicitud:
            raise HTTPException(status_code=400, detail=f"Tipo de solicitud '{nuevo_tipo_solicitud}' no encontrado en catálogo")

        new_uuid = uuid4()
        timestamp_id = now_mx.strftime('%y%m%d%H%M')
        op_id_estandar_new = f"OP - {timestamp_id}"
        
        # Calcular es_fuera_horario para seguimientos (antes faltaba)
        es_fuera_horario = await self.calcular_fuera_de_horario(conn, now_mx)
        deadline = await self.calcular_deadline_inicial(conn, now_mx)
        
        # Obtener datos completos para construir título igual que en creación inicial
        nombre_tipo = await conn.fetchval(QUERY_GET_TIPO_SOLICITUD_NAME, id_tipo_solicitud)
        nombre_tec = await conn.fetchval(QUERY_GET_TECNOLOGIA_NAME, parent['id_tecnologia'])
        
        # Título completo con el MISMO formato que la creación inicial
        titulo_new = f"{nombre_tipo}_{parent['cliente_nombre']}_{parent['nombre_proyecto']}_{nombre_tec}_{parent['canal_venta']}".upper()

        # Obtener ID dinámico
        cats = await self.get_catalog_ids(conn)
        id_status_inicial = cats['estatus'].get('pendiente') or 1

        # Agregar id_tecnologia que faltaba en seguimientos
        # Usar placeholders $22 y $23

        query_insert = QUERY_INSERT_FOLLOWUP
        await conn.fetchval(query_insert,
            new_uuid, user_id, parent_id,
            titulo_new, parent['nombre_proyecto'], parent['cliente_nombre'], parent['cliente_id'],
            parent['canal_venta'], user_name,
            parent['id_tecnologia'], id_tipo_solicitud, parent['cantidad_sitios'], prioridad,
            parent['direccion_obra'], parent['coordenadas_gps'], parent['google_maps_link'], parent['sharepoint_folder_url'],
            parent['id_interno_simulacion'], op_id_estandar_new, deadline, es_fuera_horario,
            id_status_inicial,  # Parámetro $22
            now_mx,             # Parámetro $23
            parent['es_licitacion'], # Parámetro $24 (Herencia)
            (now_mx.date() + timedelta(days=7)) # Parámetro $25 (Default +7 dias para seguimiento)
        )

        # Clonar sitios (Heredan id_tipo_solicitud del NUEVO tipo)
        # Fix: Usar el mismo status inicial (pendiente) para los sitios clonados
        query_clone = QUERY_CLONE_SITIOS
        await conn.execute(query_clone, new_uuid, parent_id, id_tipo_solicitud, id_status_inicial)
        
        # ========================================
        # HOOK: Crear levantamiento automáticamente si es tipo LEVANTAMIENTO
        # ========================================
        if nuevo_tipo_solicitud == 'LEVANTAMIENTO':
            try:
                from modules.levantamientos.service import LevantamientoService
                lev_service = LevantamientoService()
                user_context = {'user_db_id': user_id, 'user_name': user_name}
                lev_id = await lev_service.crear_desde_oportunidad(conn, new_uuid, user_context)
                logger.info(f"Levantamiento {lev_id} creado automáticamente para seguimiento {new_uuid}")
            except Exception as e:
                logger.error(f"Error creando levantamiento automático en seguimiento: {e}")
                # No fallar la creación del seguimiento por esto
        
        return new_uuid


    async def generate_multisite_excel(self, conn, id_oportunidad: UUID, id_interno: str, user_context:dict) -> Optional[dict]:
        await self.verify_ownership(conn, id_oportunidad, user_context)
        return await self.dashboard_service.generate_multisite_excel(conn, id_oportunidad, id_interno)

    async def get_dashboard_stats(
        self, 
        conn, 
        user_context: dict,
        filtro_usuario_id: Optional[UUID] = None,
        filtro_tipo_id: Optional[int] = None,
        filtro_estatus_id: Optional[int] = None,
        filtro_tecnologia_id: Optional[int] = None,
        filtro_fecha_inicio: Optional[str] = None,
        filtro_fecha_fin: Optional[str] = None
    ) -> dict:
        """
        Calcula KPIs y datos para gráficos del Dashboard Comercial.
        Soporta filtrado dinámico.
        """
        cats = await self.get_catalog_ids(conn)
        return await self.dashboard_service.get_dashboard_stats(
            conn, user_context, cats,
            filtro_usuario_id, filtro_tipo_id, filtro_estatus_id,
            filtro_tecnologia_id, filtro_fecha_inicio, filtro_fecha_fin
        )

    async def get_comentarios_workflow(self, conn, id_oportunidad: UUID) -> List[dict]:
        """
        Obtiene el historial unificado de comentarios.
        Usa tb_comentarios_workflow (la única fuente de verdad).
        
        Args:
            conn: Conexión a la base de datos
            id_oportunidad: UUID de la oportunidad
            
        Returns:
            Lista de diccionarios con comentarios
        """
        rows = await conn.fetch(QUERY_GET_COMENTARIOS_WORKFLOW, id_oportunidad)
        return [dict(r) for r in rows]

    async def get_detalles_bess(self, conn, id_oportunidad: UUID, user_context: dict) -> Optional[dict]:
        """
        Obtiene detalles BESS si existen para la oportunidad.
        
        Args:
            conn: Conexión a la base de datos
            id_oportunidad: UUID de la oportunidad
            
        Returns:
            Diccionario con detalles BESS o None si no existen
            Diccionario con detalles BESS o None si no existen
        """
        await self.verify_ownership(conn, id_oportunidad, user_context)
        row = await conn.fetchrow(QUERY_GET_DETALLES_BESS, id_oportunidad)
        
        if not row:
            return None
            
        # Convertir a dict y parsear JSON
        bess_data = dict(row)
        
        # Parsear uso_sistema_json de string a lista
        if bess_data.get('uso_sistema_json'):
            try:
                if isinstance(bess_data['uso_sistema_json'], str):
                    bess_data['uso_sistema_json'] = json.loads(bess_data['uso_sistema_json'])
            except (json.JSONDecodeError, TypeError):
                bess_data['uso_sistema_json'] = []
        
        return bess_data
    
    async def get_data_for_email_form(self, conn, id_oportunidad: UUID, user_context: dict) -> dict:
        await self.verify_ownership(conn, id_oportunidad, user_context)
        return await self.notification_service.get_data_for_email_form(conn, id_oportunidad)

    async def get_email_recipients_context(self, conn, recipients_str: str, fixed_to: List[str], fixed_cc: List[str], extra_cc: str) -> dict:
        return await self.notification_service.get_email_recipients_context(conn, recipients_str, fixed_to, fixed_cc, extra_cc)

    async def preview_site_upload(self, conn, file_contents: bytes, id_oportunidad: UUID, user_context: dict) -> dict:
        """
        Procesa el Excel en memoria y valida estructura/cantidad.
        Retorna dict con datos para previsualización o raises HTTPException.
        
        EJECUCIÓN: CPU-Bound. Se ejecuta en ThreadPoolExecutor para no bloquear el Event Loop.
        """
        await self.verify_ownership(conn, id_oportunidad, user_context)

        # Validar Cantidad Esperada en BD
        expected_qty = await conn.fetchval(
            QUERY_GET_CANTIDAD_SITIOS, 
            id_oportunidad
        )
        if expected_qty is None:
            raise HTTPException(status_code=404, detail="Oportunidad no encontrada")

        # Función auxiliar para ejecutar en thread pool
        def _process_excel_sync(contents: bytes, expected_qty: int):
            try:
                # load_workbook es bloqueante
                wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
                ws = wb.active
                headers = [str(cell.value).strip().upper() for cell in ws[1] if cell.value]
                
                full_data_list = []
                preview_rows = []
                
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_data = dict(zip(headers, row))
                    if not any(row_data.values()): continue
                    
                    clean_data = {k: (v if v is not None else "") for k, v in row_data.items()}
                    preview_rows.append(list(clean_data.values()))
                    full_data_list.append(clean_data)
                
                # Validaciones de Negocio (Sync)
                cols_req = ["NOMBRE", "DIRECCION"]
                missing = [c for c in cols_req if c not in headers]
                if missing:
                    raise ValueError(f"Faltan columnas: {', '.join(missing)}")

                if len(full_data_list) != expected_qty:
                    raise ValueError(f"Cantidad incorrecta. Esperados: {expected_qty}, Encontrados: {len(full_data_list)}")

                return {
                    "columns": headers,
                    "preview_rows": preview_rows,
                    "total_rows": len(full_data_list),
                    "json_data": json.dumps(full_data_list, default=str)
                }
            except Exception as e:
                # Re-raise para que el thread capture
                raise e

        # Ejecutar en Thread Pool
        try:
            loop = asyncio.get_running_loop()
            result = await loop.run_in_executor(
                None, 
                _process_excel_sync, 
                file_contents, 
                expected_qty
            )
            return result
        except ValueError as ve:
             # Errores de negocio conocidos
             raise HTTPException(status_code=400, detail=str(ve))
        except HTTPException:
            raise
        except Exception as e:
            # Errores inesperados de parsing
            raise HTTPException(status_code=400, detail=f"Error leyendo Excel: {str(e)}")

    async def confirm_site_upload(self, conn, id_oportunidad: UUID, json_data: str, user_context: dict) -> int:
        """Deserializa JSON, valida y realiza INSERT masivo ATÓMICO."""
        await self.verify_ownership(conn, id_oportunidad, user_context)
        try:
            raw_data = json.loads(json_data)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="JSON corrupto")

        records = []
        for item in raw_data:
            try:
                sitio = SitioImportacion(**item)
                records.append((
                    uuid4(), id_oportunidad, sitio.nombre_sitio, sitio.direccion,
                    sitio.tipo_tarifa, sitio.google_maps_link, sitio.numero_servicio, sitio.comentarios
                ))
            except Exception as e:
                logger.warning(f"Saltando fila inválida: {e}")
                continue

        # Bloque Atómico: reads + writes en la misma transacción (fix race condition B12)
        async with conn.transaction():
            # Obtener id_tipo_solicitud y estatus inicial dentro de la transacción
            id_tipo_solicitud = await conn.fetchval(QUERY_GET_TIPO_SOLICITUD_FROM_OP, id_oportunidad)

            cats = await self.get_catalog_ids(conn)
            id_status_inicial = cats['estatus'].get(STATUS_PENDIENTE) or DEFAULT_STATUS_ID_PENDIENTE

            # Recuperar IDs de sitios existentes (para borrar selectivamente después)
            old_rows = await conn.fetch(QUERY_GET_SITIO_IDS_BY_OP, id_oportunidad)
            old_ids_list = [r['id_sitio'] for r in old_rows]

            # 1. Insertar NUEVOS sitios
            if records:
                q = QUERY_INSERT_SITIO_BULK
                records_with_status_and_type = [r + (id_status_inicial, id_tipo_solicitud) for r in records]
                await conn.executemany(q, records_with_status_and_type)

            # 2. Desvincular/Re-vincular Levantamientos (FIX FK Constraint)
            if records:
                new_main_id = records[0][0]
                await conn.execute(QUERY_RELINK_LEVANTAMIENTOS, new_main_id, id_oportunidad)

            # 3. Borrar los sitios VIEJOS explícitamente
            if old_ids_list:
                await conn.execute(QUERY_DELETE_SITIOS_BY_IDS, old_ids_list)

            # 4. Sincronizar Contador
            new_count = len(records)
            await conn.execute(QUERY_UPDATE_CANTIDAD_SITIOS, new_count, id_oportunidad)

        return len(records)

    async def delete_sitio(self, conn, id_sitio: UUID, user_context: dict):
        """Elimina un sitio específico con validación de seguridad (atómico)."""
        id_oportunidad = await conn.fetchval(QUERY_GET_OPORTUNIDAD_FROM_SITIO, id_sitio)
        if id_oportunidad:
            await self.verify_ownership(conn, id_oportunidad, user_context)

        async with conn.transaction():
            await conn.execute(QUERY_DELETE_SITIO, id_sitio)

            if id_oportunidad:
                real_count = await conn.fetchval(QUERY_COUNT_SITIOS_BY_OP, id_oportunidad)
                await conn.execute(QUERY_UPDATE_CANTIDAD_SITIOS, real_count, id_oportunidad)

    # --- MÉTODOS DE LIMPIEZA ---

    async def get_sitios_simple(self, conn, id_oportunidad: UUID, user_context: Optional[dict] = None) -> List[dict]:
        """Obtiene lista simple de sitios para la UI (Partial)."""
        if user_context:
            await self.verify_ownership(conn, id_oportunidad, user_context)
            
        rows = await conn.fetch(
            QUERY_GET_SITIOS_SIMPLE,
            id_oportunidad
        )
        return [dict(r) for r in rows]

    async def auto_crear_sitio_unico(self, conn, id_oportunidad: UUID, nombre: str, direccion: str, link: Optional[str], id_tipo_solicitud: int, id_estatus: int):
        """Crea automáticamente el registro de sitio para flujos de un solo sitio."""
        await conn.execute(QUERY_INSERT_SITIO_UNICO, uuid4(), id_oportunidad, nombre, direccion, link, id_tipo_solicitud, id_estatus)

    async def marcar_extraordinaria_enviada(self, conn, id_oportunidad: UUID):
        """Marca una solicitud extraordinaria como 'enviada' sin mandar correo real."""
        await conn.execute(QUERY_UPDATE_EMAIL_ENVIADO, id_oportunidad)

    async def cancelar_oportunidad(self, conn, id_oportunidad: UUID, user_context: dict):
        """
        Elimina de forma transaccional una oportunidad y TODAS sus dependencias.
        Orden crítico: eliminar hijos antes que padre para evitar FK violations.
        """
        await self.verify_ownership(conn, id_oportunidad, user_context)
        try:
            async with conn.transaction():
                # Eliminar TODAS las tablas hijas (en orden de dependencias)
                await conn.execute(QUERY_DELETE_COMENTARIOS_WF, id_oportunidad)
                await conn.execute(QUERY_DELETE_NOTIFICACIONES, id_oportunidad)
                await conn.execute(QUERY_DELETE_DOCS, id_oportunidad)
                await conn.execute(QUERY_DELETE_LEVANTAMIENTOS, id_oportunidad)
                await conn.execute(QUERY_DELETE_BESS, id_oportunidad)
                await conn.execute(QUERY_DELETE_SITIOS_OP, id_oportunidad)
                
                # Finalmente eliminar la oportunidad padre
                await conn.execute(QUERY_DELETE_OPORTUNIDAD, id_oportunidad)
                
        except asyncpg.ForeignKeyViolationError:
            # Solo debería fallar si hay proyectos/compras (dependencias externas críticas)
            logger.warning(f"Intento de eliminar oportunidad {id_oportunidad} con dependencias críticas.")
            raise HTTPException(
                status_code=409,
                detail="No se puede eliminar: La oportunidad ya tiene Proyectos o Registros de Compra asociados."
            )

    # Helper para inyección de dependencias
    # --- MÉTODO DE NOTIFICACIÓN ---
    async def enviar_notificacion_extraordinaria(self, conn, ms_auth, token: str, id_oportunidad: UUID, base_url: str, user_email: str):
        await self.notification_service.enviar_notificacion_extraordinaria(conn, ms_auth, token, id_oportunidad, base_url, user_email)

    async def buscar_clientes(self, conn, query: str) -> List[dict]:
        """
        Búsqueda inteligente de clientes por nombre fiscal.
        Args:
            conn: Conexión BD
            query: Texto a buscar (case insensitive)
        Returns:
            List[dict]: [{id, nombre_fiscal}, ...] lim 10
        """
        if not query or len(query.strip()) < 2:
            return []
            
        # Normalizar query para búsqueda ILIKE segura
        search_term = f"%{query.strip()}%"
        
        rows = await conn.fetch(QUERY_SEARCH_CLIENTES, search_term)
        
        return [{"id": str(r["id"]), "nombre_fiscal": r["nombre_fiscal"]} for r in rows]

    def get_redirection_params(
        self,
        new_id: UUID, 
        op_std_id: str, 
        cant_sitios: int, 
        es_fuera_horario: bool,
        legacy_term: Optional[str] = None,
        is_extraordinario: bool = False
    ) -> dict:
        """
        Calcula los parámetros y datos necesarios para la redirección.
        NO construye headers HTTP, solo retorna la información lógica.
        
        Returns:
            dict: {
                "redirect_url": "/path/to/step",
                "query_params": {key: value}
            }
        """
        # 1. Determinar URL Base
        if is_extraordinario:
            # Extraordinarias Unisitio -> Dashboard (UI)
            # Extraordinarias Multisitio -> Paso 2 (Carga Masiva)
            if cant_sitios == 1:
                target_url = "/comercial/ui"
            else:
                target_url = f"/comercial/paso2/{new_id}"
        else:
            # Flujo Normal: Paso 3 (Email) o Paso 2 (Excel)
            step = self.get_next_ui_step(cant_sitios)
            target_url = f"/comercial/{step}/{new_id}"

        # 2. Construir Query Params Diccionario
        params = {
            "new_op": op_std_id,
            "fh": str(es_fuera_horario).lower()
        }
        
        if legacy_term:
            params["legacy_term"] = legacy_term
            
        if is_extraordinario:
            params["extraordinaria"] = "1"

        return {
            "redirect_url": target_url,
            "query_params": params
        }

    async def marcar_como_ganada(
        self, 
        conn, 
        id_oportunidad: UUID, 
        sitios_ganados: List[UUID], 
        user_context: dict
    ) -> dict:
        """
        Marca una oportunidad como Ganada (cierre de venta exitoso).
        
        Reglas de negocio:
        1. Solo se puede marcar si status actual = Entregado
        2. Los KPIs ya fueron calculados al marcar como Entregado (se heredan)
        3. Para multisitio: solo cambia status de sitios seleccionados a Ganada
        4. Sitios no seleccionados pasan a Perdido (si hay selección parcial)
        5. La oportunidad padre siempre cambia a Ganada
        
        Args:
            conn: Conexión asyncpg
            id_oportunidad: UUID de la oportunidad
            sitios_ganados: Lista de UUIDs de sitios ganados (vacía = todos ganados)
            user_context: Contexto del usuario actual
            
        Returns:
            dict con información del cierre
            
        Raises:
            HTTPException: Si status actual no es Entregado
        """
        # Obtener mapa de estatus
        cats = await self.get_catalog_ids(conn)
        estatus_map = cats.get("estatus", {})
        
        id_entregado = estatus_map.get("entregado")
        id_ganada = estatus_map.get("ganada")
        id_perdido = estatus_map.get("perdido")
        
        if not all([id_entregado, id_ganada, id_perdido]):
            raise HTTPException(500, "Error de configuración: faltan estatus en catálogo")

        await self.verify_ownership(conn, id_oportunidad, user_context)

        # Validar status actual de la oportunidad
        current_status = await conn.fetchval(QUERY_GET_OP_ESTATUS, id_oportunidad)
        
        if current_status != id_entregado:
            raise HTTPException(
                400, 
                "Solo se puede marcar como Ganada desde status 'Entregado'"
            )
        
        # Obtener cantidad de sitios para determinar lógica
        sitios_count = await conn.fetchval(QUERY_COUNT_SITIOS_BY_OP, id_oportunidad)
        
        async with conn.transaction():
            if sitios_ganados and len(sitios_ganados) > 0:
                # Multisitio: seleccionados → Ganada
                await conn.execute(QUERY_UPDATE_SITIOS_ESTATUS_BY_IDS, id_ganada, sitios_ganados, id_oportunidad)
                # No seleccionados → Perdido (solo los que están en Entregado)
                await conn.execute(QUERY_UPDATE_SITIOS_ESTATUS_OTHERS, id_perdido, id_oportunidad, sitios_ganados, id_entregado)

                sitios_ganados_count = len(sitios_ganados)
                sitios_perdidos_count = sitios_count - sitios_ganados_count
            else:
                # Unisitio o sin selección → todos Ganada
                await conn.execute(QUERY_UPDATE_SITIOS_ESTATUS_ALL, id_ganada, id_oportunidad)

                sitios_ganados_count = sitios_count
                sitios_perdidos_count = 0

            # Actualizar oportunidad padre a Ganada
            await conn.execute(QUERY_UPDATE_OP_ESTATUS, id_ganada, id_oportunidad)
            
            # Insertar Historial (Ganada)
            # Reutilizamos lógica de SLACalculator
            config = await self.get_configuracion_global(conn)
            hora_corte, dias_fin_semana, _ = SLACalculator.parse_config(config)
            now_mx = await self.get_current_datetime_mx(conn)
            fecha_inicio_sla = SLACalculator.calculate_deadline(now_mx, hora_corte, 0)
            
            from .db_service import QUERY_INSERT_HISTORIAL_ESTATUS
            await conn.execute(QUERY_INSERT_HISTORIAL_ESTATUS,
                id_oportunidad,
                current_status, # Anterior (Entregado)
                id_ganada,      # Nuevo
                now_mx,
                fecha_inicio_sla,
                user_context['user_db_id']
            )
        
        logger.info(
            f"Cierre de venta: Oportunidad {id_oportunidad} marcada como Ganada "
            f"por {user_context.get('user_name')}. "
            f"Sitios ganados: {sitios_ganados_count}, perdidos: {sitios_perdidos_count}"
        )
        
        return {
            "success": True,
            "id_oportunidad": str(id_oportunidad),
            "sitios_ganados": sitios_ganados_count,
            "sitios_perdidos": sitios_perdidos_count
        }

    async def get_paso2_data(self, conn, id_oportunidad: UUID) -> Optional[dict]:
        """Recupera datos mínimos para renderizar el formulario de Paso 2."""
        row = await conn.fetchrow(QUERY_GET_PASO2_DATA, id_oportunidad)
        return dict(row) if row else None

    async def predict_followup_title(self, conn, parent_id: UUID, nuevo_tipo_solicitud: str) -> str:
        """
        Retorna el título ACTUAL de la oportunidad padre.
        Para encontrar el hilo de correo previo, debemos buscar por el título que YA existe (el del padre),
        no por el título futuro (que tendría el tipo de solicitud nuevo).
        """
        titulo_parent = await conn.fetchval(
            "SELECT titulo_proyecto FROM tb_oportunidades WHERE id_oportunidad = $1", 
            parent_id
        )
        return titulo_parent or ""

    def get_next_ui_step(self, cantidad_sitios: int) -> str:
        """Determina el siguiente paso del flujo UI basado en reglas de negocio."""
        # Regla: Unisitio (1) -> Paso 3 (Email)
        # Regla: Multisitio (>1) -> Paso 2 (Carga Masiva)
        return "paso3" if cantidad_sitios == 1 else "paso2"

def get_comercial_service():
    return ComercialService()