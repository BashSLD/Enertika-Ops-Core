# Archivo: modules/compras/service.py
"""
Service Layer del Módulo Compras.
Maneja la lógica de negocio para comprobantes de pago y facturas XML.
"""

from uuid import UUID
from datetime import datetime, date
from core.timezone import today_mx, now_mx
from typing import List, Dict, Optional, Tuple, Any
from fastapi import HTTPException
from decimal import Decimal, InvalidOperation
import logging
import time

import asyncpg
import base64
import httpx
from pdfminer.pdfexceptions import PSException
from core.materials.service import get_materials_service
from .pdf_extractor import process_pdf_bytes
from .xml_extractor import parse_cfdi_xml, validate_xml_content
from .schemas import (
    CfdiData, TipoFactura, XmlMatchResult, XmlUploadResult, XmlUploadError,
)

logger = logging.getLogger("ComprasService")

# Tolerancia de matching por monto (pesos/dolares)
MATCH_TOLERANCIA = Decimal("0.50")

# Constraints de unicidad que corresponden a duplicados de negocio (mismo PDF cargado dos veces).
# Cualquier otra UniqueViolationError en insert_comprobante es un error de infraestructura y debe propagarse.
# IMPORTANTE: si una migración renombra alguno de estos constraints, actualizar este set en paralelo.
_BUSINESS_DUPLICATE_CONSTRAINTS = frozenset({
    "uq_comprobante_pago_key",
    "uq_comprobante_duplicado_no_bom",
})


def parse_exceso_monto_error(msg: str) -> tuple[str | None, str | None, str]:
    parts = msg.split("|", 3)
    if len(parts) == 4:
        _, exceso_monto, monto_aplicado, user_msg = parts
        return exceso_monto, monto_aplicado, user_msg
    if len(parts) == 3:
        _, exceso_monto, user_msg = parts
        return exceso_monto, None, user_msg
    return None, None, msg


def _to_decimal(value) -> Decimal:
    return Decimal(str(value or 0))


def _saldo_factura_desde_resumen(total_cfdi, factura_resumen: dict) -> Decimal:
    monto_factura_base = max(
        _to_decimal(total_cfdi),
        _to_decimal(factura_resumen.get('monto_factura')),
    )
    return monto_factura_base - _to_decimal(factura_resumen.get('monto_aplicado'))


def _es_concepto_producto(clave: str | None) -> bool:
    """Filtra conceptos SAT que no son productos físicos.

    UNSPSC: segmentos 10–49 = productos, 50–99 = servicios.
    Excluye también el catch-all 01010101 y claves inválidas.
    """
    if not clave or len(clave) < 2:
        return False
    try:
        segmento = int(clave[:2])
    except ValueError:
        return False
    return 10 <= segmento <= 49


def _separar_matches_bom(match_result: dict) -> tuple[dict, dict, dict]:
    """Separa ligas reales de sugerencias para no facturar matches de baja confianza."""
    bom_item_map = {}
    match_meta_map = {}
    suggestion_map = {}

    for idx, match in match_result.items():
        if not match:
            continue
        meta = {'confianza': match['confianza'], 'origen': match['origen']}
        if match['confianza'] == 'ALTA':
            bom_item_map[idx] = match['id_item']
            match_meta_map[idx] = meta
        else:
            suggestion_map[idx] = {'id_item': match['id_item'], **meta}

    return bom_item_map, match_meta_map, suggestion_map


def _separar_matches_interno(match_result: dict) -> tuple[dict, dict]:
    """Separa matches ALTA (CLAVE_SAT/MEMORIA, auto-aplicables) de sugerencias
    BAJA (TEXTO, requieren revision humana) del matcher catalogo interno<->XML
    (doc 39). Mismo criterio que _separar_matches_bom, pero ALTA no se escribe
    en la misma fila del INSERT -- necesita el id real de tb_materiales_historial,
    que se aplica en un paso aparte (ver ComprasService.confirmar_match_xml)."""
    alta_map = {}
    suggestion_map = {}
    for idx, match in match_result.items():
        if not match:
            continue
        if match['confianza'] == 'ALTA':
            alta_map[idx] = match
        else:
            suggestion_map[idx] = match
    return alta_map, suggestion_map


class ComprasService:
    """Lógica de negocio del módulo Compras - Comprobantes de Pago."""
    
    # ========================================
    # CARGA DE COMPROBANTES (PDFs)
    # ========================================
    
    async def process_and_save_pdfs(
        self, 
        conn, 
        files: list, 
        user_id: UUID
    ) -> Dict[str, Any]:
        """
        Procesa múltiples PDFs y guarda los comprobantes válidos.
        
        Args:
            conn: Conexión a base de datos
            files: Lista de UploadFile de FastAPI
            user_id: UUID del usuario que realiza la carga
            
        Returns:
            {
                "insertados": int,
                "duplicados": List[dict],
                "errores": List[dict]
            }
        """
        insertados = 0
        duplicados = []
        errores = []
        
        for file in files:
            filename = file.filename
            
            # 1. Leer contenido del archivo
            try:
                content = await file.read()
                await file.seek(0)
            except Exception as e:
                logger.error(f"Error leyendo archivo {filename}: {e}")
                errores.append({
                    "archivo": filename,
                    "error": f"Error al leer archivo: {str(e)}"
                })
                continue
            
            # 2. Extraer datos del PDF
            try:
                data = process_pdf_bytes(content, filename)
            except (PSException, ValueError, TypeError, KeyError, IndexError, AttributeError, OSError) as e:
                logger.error(f"Error procesando PDF {filename}: {e}", exc_info=True)
                errores.append({
                    "archivo": filename,
                    "error": f"Error al procesar PDF: {str(e)}"
                })
                continue

            if data.error or not data.is_valid():
                errores.append({
                    "archivo": filename,
                    "error": data.error or "Datos incompletos"
                })
                continue
            
            # 3. Verificar duplicado using db_service
            fecha_pago_date = data.fecha_pago.date() if isinstance(data.fecha_pago, datetime) else data.fecha_pago
            
            from .db_service import get_db_service
            db_svc = get_db_service()
            
            exists = await db_svc.check_duplicate_comprobante(
                conn, fecha_pago_date, data.beneficiario, Decimal(str(data.monto))
            )
            
            entrada_duplicado = {
                "archivo": filename,
                "fecha": data.fecha_pago.strftime("%d/%m/%Y"),
                "beneficiario": data.beneficiario,
                "monto": data.monto,
                "moneda": data.moneda,
            }

            if exists:
                duplicados.append(entrada_duplicado)
                logger.info(f"Duplicado detectado: {filename}")
                continue

            # 4. Insertar en base de datos using db_service
            try:
                # Intentar auto-asignar el proveedor si ya existe la relación
                proveedor_rel = await db_svc.get_proveedor_by_beneficiario(conn, data.beneficiario)
                id_proveedor = proveedor_rel['id_proveedor'] if proveedor_rel else None

                comprobante_data = {
                    'fecha_pago': fecha_pago_date,
                    'beneficiario': data.beneficiario,
                    'monto': Decimal(str(data.monto)),
                    'moneda': data.moneda,
                    'user_id': user_id,
                    'id_proveedor': id_proveedor
                }
                try:
                    new_id = await db_svc.insert_comprobante(conn, comprobante_data)
                except asyncpg.exceptions.UniqueViolationError as e:
                    if e.constraint_name is None or e.constraint_name not in _BUSINESS_DUPLICATE_CONSTRAINTS:
                        raise
                    duplicados.append(entrada_duplicado)
                    logger.info("Duplicado concurrente detectado: %s", filename)
                    continue

                insertados += 1
                logger.info(f"Comprobante insertado: {filename} - {data.beneficiario} - ${data.monto}")

                # 5. Subir PDF a SharePoint
                try:
                    await file.seek(0)
                    now = now_mx()
                    subcarpeta = f"compras/comprobantes_pdf/{now.strftime('%Y-%m')}"
                    sp_result = await self.upload_archivo_sharepoint(
                        conn, file, subcarpeta,
                        new_id, "comprobante_pago", user_id,
                        metadata_extra={
                            "beneficiario": data.beneficiario,
                            "monto": str(data.monto),
                            "moneda": data.moneda,
                        }
                    )
                    if sp_result:
                        logger.info("PDF subido a SharePoint: %s", sp_result.get("url_sharepoint"))
                except Exception as e:
                    logger.error("Error subiendo PDF %s a SharePoint: %s (comprobante ya guardado en BD)", filename, e)

            except Exception as e:
                logger.error(f"Error insertando comprobante {filename}: {e}")
                errores.append({
                    "archivo": filename,
                    "error": f"Error de base de datos: {str(e)}"
                })
        
        logger.info(f"Proceso completado: {insertados} insertados, {len(duplicados)} duplicados, {len(errores)} errores")
        
        return {
            "insertados": insertados,
            "duplicados": duplicados,
            "errores": errores
        }
    
    # ========================================
    # CONSULTAS DE COMPROBANTES
    # ========================================
   
    async def get_comprobantes(
        self,
        conn,
        filtros: dict,
        page: int = 1,
        per_page: int = 50
    ) -> Tuple[List[dict], int]:
        """
        Obtiene comprobantes con filtros y paginación.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()

        # Obtener total
        total = await db_svc.get_comprobantes_filtered(
            conn, filtros, page, per_page, count_only=True
        )

        # Obtener datos
        rows = await db_svc.get_comprobantes_filtered(
            conn, filtros, page, per_page, count_only=False
        )
        
        # Convertir a diccionarios
        comprobantes = []
        for row in rows:
            comp = dict(row)
            # Convertir Decimal a float para serialización
            if comp.get('monto'):
                comp['monto'] = float(comp['monto'])
            comprobantes.append(comp)
        
        return comprobantes, total
    
    async def get_comprobantes_default_view(self, conn, user_id=None) -> Tuple[List[dict], int]:
        """Vista default: comprobantes abiertos. Si user_id se provee, filtra por ese usuario."""
        filtros = {"estatus": "SIN_COMPLETAR"}
        if user_id:
            filtros["id_usuario"] = user_id
        return await self.get_comprobantes(conn, filtros=filtros)
    
    async def get_comprobante_by_id(self, conn, id_comprobante: UUID) -> Optional[dict]:
        """
        Obtiene un comprobante específico por ID.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()
        
        comp = await db_svc.get_comprobante_by_id(conn, id_comprobante)
        if comp:
            if comp.get('monto'):
                comp['monto'] = float(comp['monto'])
            return comp
        return None
    
    # ========================================
    # EDICIÓN DE COMPROBANTES
    # ========================================
    
    async def update_comprobante(
        self,
        conn,
        id_comprobante: UUID,
        updates: dict,
        user_context: dict
    ) -> dict:
        """
        Actualiza campos editables de un comprobante.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()
        
        # PERMISOS: Admin/Manager, Editor Module, o DUEÑO DEL REGISTRO
        user_id = user_context.get("user_db_id")
        user_role = user_context.get("role")
        mod_role = user_context.get("module_roles", {}).get("compras")
        
        is_admin_or_editor = (user_role in ["ADMIN", "MANAGER"] or mod_role in ["admin", "editor"])
        
        if not is_admin_or_editor:
            # Verificar ownership
            current = await db_svc.get_comprobante_by_id(conn, id_comprobante)
            if not current:
                raise HTTPException(status_code=404, detail="Comprobante no encontrado")
                
            if current['capturado_por_id'] != user_id:
                raise HTTPException(
                    status_code=403, 
                    detail="Solo puedes editar los comprobantes que tú capturaste."
                )
        
        success = await db_svc.update_comprobante(conn, id_comprobante, updates)
        if not success:
             raise HTTPException(status_code=404, detail="Comprobante no encontrado o sin cambios")
        
        return await self.get_comprobante_by_id(conn, id_comprobante)
    
    async def bulk_update_comprobantes(
        self,
        conn,
        ids: List[UUID],
        updates: dict,
        user_context: dict
    ) -> int:
        """
        Actualización masiva de múltiples comprobantes.
        """
        if not ids:
            return 0
            
        from .db_service import get_db_service
        db_svc = get_db_service()
        
        # PERMISOS
        user_id = user_context.get("user_db_id")
        user_role = user_context.get("role")
        mod_role = user_context.get("module_roles", {}).get("compras")
        
        is_admin_or_editor = (user_role in ["ADMIN", "MANAGER"] or mod_role in ["admin", "editor"])
        
        if not is_admin_or_editor:
            # Verificar que TODOS los comprobantes sean del usuario using db_service
            not_owned_count = await db_svc.check_ownership_bulk(conn, ids, user_id)
            
            if not_owned_count > 0:
                raise HTTPException(
                    status_code=403,
                    detail=f"Seleccionaste {not_owned_count} comprobante(s) que no te pertenecen. Solo puedes editar tus propios registros."
                )
        
        count = await db_svc.bulk_update(conn, ids, updates)
        logger.info(f"Bulk update: {count} comprobantes actualizados")
        return count
    
    # ========================================
    # EXPORTACIÓN A EXCEL
    # ========================================
    
    async def export_to_excel(
        self,
        conn,
        filtros: dict
    ) -> bytes:
        """
        Genera archivo Excel con dos hojas:
        - Hoja 1: Comprobantes de Pago (incluye montos parciales)
        - Hoja 2: Facturas Vinculadas (una fila por factura)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        TIPOS_ES = {
            "NORMAL":          "Factura",
            "ANTICIPO":        "Anticipo",
            "CIERRE_ANTICIPO": "Cierre Anticipo",
            "NOTA_CREDITO":    "Nota de Crédito",
            "PAGO":            "Complemento Pago",
        }

        comprobantes, _ = await self.get_comprobantes(
            conn,
            filtros=filtros,
            per_page=100000
        )

        from .db_service import get_db_service
        db_svc = get_db_service()
        comp_ids = [c['id_comprobante'] for c in comprobantes if c.get('id_comprobante')]
        facturas_map = await db_svc.get_facturas_for_comprobantes(conn, comp_ids)

        wb = Workbook()

        # ── Estilos compartidos ──────────────────────────────────────────────
        def make_header_style():
            return (
                Font(bold=True, color="FFFFFF"),
                PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
                Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

        def make_subheader_fill(color="2E75B6"):
            return PatternFill(start_color=color, end_color=color, fill_type="solid")

        thin = Side(style='thin')
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        num_fmt = '#,##0.00'
        right_align = Alignment(horizontal="right")

        def write_header_row(ws, headers):
            h_font, h_fill, h_align = make_header_style()
            for col, text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=text)
                cell.font = h_font
                cell.fill = h_fill
                cell.alignment = h_align
                cell.border = thin_border

        # ── HOJA 1: Comprobantes ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Comprobantes de Pago"
        ws1.row_dimensions[1].height = 30

        headers1 = [
            "Comprador", "Proveedor", "Proyecto", "Zona",
            "Fecha de Pago", "Estatus", "Tipo CFDI",
            "Monto", "Monto Facturado", "Monto Pendiente", "Moneda",
            "Categoría", "Num. Facturas",
        ]
        write_header_row(ws1, headers1)

        # Colores de estatus para Hoja 1
        estatus_fills = {
            "FACTURADO":              PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "PARCIALMENTE_FACTURADO": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "CERRADO":                PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
            "PENDIENTE":              PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
            "ANTICIPO":               PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        }

        for row_num, comp in enumerate(comprobantes, 2):
            proveedor = comp.get('proveedor_nombre') or comp.get('beneficiario_orig', '')
            comp_id = comp.get('id_comprobante')
            comp_facturas = facturas_map.get(comp_id, [])
            monto = float(comp.get('monto') or 0)
            monto_facturado = float(comp.get('monto_facturado') or 0)
            monto_pendiente = monto - monto_facturado
            estatus = comp.get('estatus', '')

            row_data = [
                comp.get('comprador_nombre', ''),
                proveedor,
                comp.get('proyecto_nombre', ''),
                comp.get('zona_nombre', ''),
                comp['fecha_pago'].strftime("%d/%m/%Y") if comp.get('fecha_pago') else '',
                estatus,
                TIPOS_ES.get(comp.get('tipo_factura') or 'NORMAL', comp.get('tipo_factura', '')),
                float(monto),
                float(monto_facturado),
                float(monto_pendiente),
                comp.get('moneda', 'MXN'),
                comp.get('categoria_nombre', ''),
                len(comp_facturas),
            ]

            estatus_fill = estatus_fills.get(estatus)
            for col_num, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num in (8, 9, 10):
                    cell.number_format = num_fmt
                    cell.alignment = right_align
                if estatus_fill and col_num == 6:
                    cell.fill = estatus_fill

        col_widths1 = [20, 35, 30, 15, 15, 22, 16, 15, 16, 16, 10, 20, 13]
        for i, w in enumerate(col_widths1, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w
        ws1.freeze_panes = "A2"

        # ── HOJA 2: Facturas Vinculadas ──────────────────────────────────────
        ws2 = wb.create_sheet("Facturas Vinculadas")
        ws2.row_dimensions[1].height = 30

        headers2 = [
            "Fecha de Pago", "Comprador", "Proveedor", "Proyecto",
            "Estatus Comprobante", "Monto Comprobante", "Moneda",
            "UUID Factura", "Tipo", "Monto Factura", "Monto Aplicado",
            "Saldo Factura", "Estatus Factura", "Fecha Factura", "RFC Emisor", "Nombre Emisor",
        ]
        write_header_row(ws2, headers2)

        row_num2 = 2
        for comp in comprobantes:
            comp_id = comp.get('id_comprobante')
            comp_facturas = facturas_map.get(comp_id, [])
            if not comp_facturas:
                continue

            proveedor = comp.get('proveedor_nombre') or comp.get('beneficiario_orig', '')
            fecha_pago = comp['fecha_pago'].strftime("%d/%m/%Y") if comp.get('fecha_pago') else ''

            for f in comp_facturas:
                fecha_factura_str = f.get('fecha').strftime("%d/%m/%Y") if f.get('fecha') else ''
                monto_factura = float(f.get('monto') or 0)
                monto_aplicado_val = f.get('monto_aplicado')
                if monto_aplicado_val is None:
                    monto_aplicado_val = f.get('monto')
                monto_aplicado = float(monto_aplicado_val or 0)
                saldo_factura = float(f.get('saldo_factura') or 0)
                estatus_factura = "Cubierta" if saldo_factura <= 0.005 else "Parcial"
                row_data2 = [
                    fecha_pago,
                    comp.get('comprador_nombre', ''),
                    proveedor,
                    comp.get('proyecto_nombre', ''),
                    comp.get('estatus', ''),
                    float(comp.get('monto') or 0),
                    comp.get('moneda', 'MXN'),
                    str(f.get('uuid_factura', '')),
                    TIPOS_ES.get(f.get('tipo') or 'NORMAL', f.get('tipo', '')),
                    monto_factura,
                    monto_aplicado,
                    saldo_factura,
                    estatus_factura,
                    fecha_factura_str,
                    f.get('rfc_emisor', ''),
                    f.get('nombre_emisor', ''),
                ]
                for col_num, value in enumerate(row_data2, 1):
                    cell = ws2.cell(row=row_num2, column=col_num, value=value)
                    cell.border = thin_border
                    if col_num in (6, 10, 11, 12):
                        cell.number_format = num_fmt
                        cell.alignment = right_align
                row_num2 += 1

        col_widths2 = [15, 20, 35, 30, 22, 16, 10, 38, 14, 15, 15, 15, 15, 14, 18, 40]
        for i, w in enumerate(col_widths2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        ws2.freeze_panes = "A2"

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    # ========================================
    # CATÁLOGOS
    # ========================================
    
    async def get_catalogos(self, conn) -> dict:
        """
        Obtiene todos los catálogos necesarios para dropdowns.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.get_catalogos_data(conn)
    
    async def get_proveedores_search(
        self, 
        conn, 
        search_term: str, 
        limit: int = 10
    ) -> List[dict]:
        """
        Búsqueda de proveedores por nombre o RFC.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.search_proveedores(conn, search_term, limit)
    
    # ========================================
    # ESTADÍSTICAS (para dashboard futuro)
    # ========================================
    
    async def get_estadisticas_generales(
        self, 
        conn,
        filtros: Optional[dict] = None,
        # Legacy params support for ease of refactor, convert them to dict
        fecha_inicio: Optional[date] = None,
        fecha_fin: Optional[date] = None,
        estatus: Optional[str] = None,
        id_zona: Optional[int] = None,
        id_proyecto: Optional[UUID] = None,
        id_categoria: Optional[int] = None
    ) -> dict:
        """
        Obtiene estadísticas generales con filtros dinámicos.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()
        
        # Build filter dict if not provided
        if filtros is None:
            filtros = {
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "estatus": estatus,
                "id_zona": id_zona,
                "id_proyecto": id_proyecto,
                "id_categoria": id_categoria
            }
            
        stats = await db_svc.get_estadisticas(conn, filtros)
        xml_pendientes = await db_svc.get_xml_pendientes_count(conn)

        return {
            "total": stats['total'],
            "pendientes": stats['pendientes'],
            "facturados": stats['facturados'],
            "anticipos": stats.get('anticipos', 0),
            "parciales": stats.get('parciales', 0),
            "cerrados": stats.get('cerrados', 0),
            "total_mxn": float(stats['total_mxn']),
            "total_usd": float(stats['total_usd']),
            "xml_pendientes": xml_pendientes,
        }


    # ========================================
    # CARGA Y PROCESAMIENTO DE XMLs
    # ========================================

    async def procesar_xmls(
        self,
        conn,
        files: list,
        user_id: UUID
    ) -> XmlUploadResult:
        """
        Procesa multiples XMLs CFDI: parsea, busca match, prepara resultados.
        NO confirma match automaticamente — retorna candidatos para UI.

        Args:
            conn: Conexion a base de datos
            files: Lista de UploadFile
            user_id: UUID del usuario

        Returns:
            XmlUploadResult con procesados, duplicados y errores
        """
        from .db_service import get_db_service
        db_svc = get_db_service()

        result = XmlUploadResult()

        for file in files:
            filename = file.filename or "sin_nombre.xml"

            # 1. Leer contenido
            try:
                content = await file.read()
                await file.seek(0)
            except Exception as e:
                logger.error("Error leyendo XML %s: %s", filename, e)
                result.errores.append(XmlUploadError(
                    archivo=filename, error=f"Error al leer archivo: {e}"
                ))
                continue

            # 2. Validacion rapida
            error_msg = validate_xml_content(content, filename)
            if error_msg:
                result.errores.append(XmlUploadError(
                    archivo=filename, error=error_msg
                ))
                continue

            # 3. Parsear XML
            try:
                cfdi = parse_cfdi_xml(content, filename)
            except ValueError as e:
                result.errores.append(XmlUploadError(
                    archivo=filename, error=str(e)
                ))
                continue

            # 4. Verificar UUID duplicado. Una factura ya vinculada puede volver
            # a usarse si conserva saldo pendiente por aplicar.
            existe_legacy = await db_svc.uuid_factura_exists(conn, cfdi.uuid)
            existe_junction = await db_svc.uuid_factura_exists_in_junction(conn, cfdi.uuid)
            if existe_legacy or existe_junction:
                if not existe_junction:
                    result.duplicados.append(XmlUploadError(
                        archivo=filename,
                        error=f"UUID {cfdi.uuid[:8]}... ya existe en el sistema"
                    ))
                    continue

                factura_resumen = await db_svc.get_factura_aplicacion_resumen(conn, cfdi.uuid)
                saldo_factura = _saldo_factura_desde_resumen(cfdi.total, factura_resumen)
                if saldo_factura <= MATCH_TOLERANCIA:
                    result.duplicados.append(XmlUploadError(
                        archivo=filename,
                        error=f"UUID {cfdi.uuid[:8]}... ya esta cubierta por pagos relacionados"
                    ))
                    continue

            # 5. Buscar/crear proveedor
            proveedor = await db_svc.get_proveedor_by_rfc(conn, cfdi.emisor_rfc)
            if not proveedor:
                proveedor = await db_svc.create_proveedor(
                    conn, cfdi.emisor_rfc, cfdi.emisor_nombre
                )
                logger.info(
                    "Proveedor creado: RFC=%s, Nombre=%s",
                    cfdi.emisor_rfc, cfdi.emisor_nombre
                )

            # 6. Buscar matching con comprobantes
            match_result = await self._buscar_match(
                conn, db_svc, cfdi, proveedor
            )

            # 7. Almacenar contenido XML en base64 para upload posterior a SharePoint
            match_result.xml_content_b64 = base64.b64encode(content).decode('ascii')

            result.procesados.append(match_result)
            try:
                await db_svc.upsert_xml_staging(
                    conn, cfdi.uuid, cfdi.emisor_rfc, cfdi.emisor_nombre,
                    cfdi.total, cfdi.moneda, cfdi.tipo_factura.value,
                    match_result.match_type, user_id,
                    xml_content_b64=match_result.xml_content_b64
                )
            except asyncpg.PostgresError as e:
                logger.warning("No se pudo registrar XML en staging: %s", e)

        logger.info(
            "XMLs procesados: %d OK, %d duplicados, %d errores",
            len(result.procesados), len(result.duplicados), len(result.errores)
        )
        return result

    async def _buscar_match(
        self, conn, db_svc, cfdi: CfdiData, proveedor: dict
    ) -> XmlMatchResult:
        """
        Busca match para un CFDI parseado en 3 niveles:
        1. Relacion conocida (beneficiario↔proveedor)
        2. Solo por monto + moneda
        3. Sin match
        """
        id_proveedor = proveedor['id_proveedor']
        monto = cfdi.total
        moneda = cfdi.moneda or "MXN"

        if cfdi.tipo_factura == TipoFactura.CIERRE_ANTICIPO:
            uuid_anticipo_relacionado = None
            for rel in cfdi.relacionados or []:
                if rel.tipo_relacion == "07" and rel.uuid:
                    uuid_anticipo_relacionado = rel.uuid
                    break

            if uuid_anticipo_relacionado:
                try:
                    anticipo = await db_svc.get_comprobante_anticipo_by_uuid(
                        conn, uuid_anticipo_relacionado
                    )
                except ValueError:
                    anticipo = None

                if anticipo:
                    comprobante = await db_svc.get_comprobante_by_id(
                        conn, anticipo["id_comprobante"]
                    )
                    candidatos = [comprobante] if comprobante else []
                    return XmlMatchResult(
                        cfdi=cfdi,
                        match_type="AUTO_MATCH",
                        candidatos=self._format_candidatos(candidatos),
                        comprobante_id=anticipo["id_comprobante"],
                    )

        # Nivel 1: buscar por relacion conocida (batch — una sola query con todos los beneficiarios)
        relaciones = await db_svc.get_relaciones_beneficiario(conn, id_proveedor)
        if relaciones:
            nombres_rel = [rel['beneficiario_nombre'] for rel in relaciones]
            candidatos = await db_svc.buscar_comprobantes_por_nombres_proveedor(
                conn, nombres_rel, monto, moneda, MATCH_TOLERANCIA
            )
            if len(candidatos) == 1:
                return XmlMatchResult(
                    cfdi=cfdi,
                    match_type="AUTO_MATCH",
                    candidatos=self._format_candidatos(candidatos),
                    comprobante_id=candidatos[0]['id_comprobante'],
                )
            if candidatos:
                return XmlMatchResult(
                    cfdi=cfdi,
                    match_type="MULTIPLE_MATCH",
                    candidatos=self._format_candidatos(candidatos),
                )

        # Nivel 1.5: buscar por razon_social/nombre_comercial del proveedor
        nombres_proveedor = [proveedor.get('razon_social', '')]
        nombre_com = proveedor.get('nombre_comercial')
        if nombre_com and nombre_com != nombres_proveedor[0]:
            nombres_proveedor.append(nombre_com)

        candidatos = await db_svc.buscar_comprobantes_por_nombres_proveedor(
            conn, nombres_proveedor, monto, moneda, MATCH_TOLERANCIA
        )
        if len(candidatos) == 1:
            return XmlMatchResult(
                cfdi=cfdi,
                match_type="AUTO_MATCH",
                candidatos=self._format_candidatos(candidatos),
                comprobante_id=candidatos[0]['id_comprobante'],
            )
        if candidatos:
            return XmlMatchResult(
                cfdi=cfdi,
                match_type="MULTIPLE_MATCH",
                candidatos=self._format_candidatos(candidatos),
            )

        # Nivel 2: buscar solo por monto
        candidatos = await db_svc.buscar_comprobantes_por_monto(
            conn, monto, moneda, MATCH_TOLERANCIA
        )
        if len(candidatos) == 1:
            return XmlMatchResult(
                cfdi=cfdi,
                match_type="MONTO_MATCH",
                candidatos=self._format_candidatos(candidatos),
                comprobante_id=candidatos[0]['id_comprobante'],
            )
        if candidatos:
            return XmlMatchResult(
                cfdi=cfdi,
                match_type="MULTIPLE_MATCH",
                candidatos=self._format_candidatos(candidatos),
            )

        # Nivel 3: buscar por proveedor + saldo pendiente (facturas parciales)
        id_proveedor = proveedor.get('id_proveedor')
        if id_proveedor:
            candidatos = await db_svc.buscar_comprobantes_parciales_por_proveedor(
                conn, id_proveedor, moneda, monto, MATCH_TOLERANCIA
            )
            if candidatos:
                return XmlMatchResult(
                    cfdi=cfdi,
                    match_type="PARCIAL_MATCH",
                    candidatos=self._format_candidatos(candidatos),
                    comprobante_id=candidatos[0]['id_comprobante'] if len(candidatos) == 1 else None,
                )

        # Nivel 4: sin match
        return XmlMatchResult(
            cfdi=cfdi,
            match_type="NO_MATCH",
            candidatos=[],
        )

    def _format_candidatos(self, rows: List[dict]) -> List[dict]:
        """Formatea candidatos para la respuesta, convirtiendo Decimal a float."""
        formatted = []
        for r in rows:
            item = dict(r)
            if 'monto' in item and isinstance(item['monto'], Decimal):
                item['monto'] = float(item['monto'])
            if 'monto_facturado' in item and isinstance(item['monto_facturado'], Decimal):
                item['monto_facturado'] = float(item['monto_facturado'])
            if 'saldo_pendiente' in item and isinstance(item['saldo_pendiente'], Decimal):
                item['saldo_pendiente'] = float(item['saldo_pendiente'])
            if 'fecha_pago' in item and hasattr(item['fecha_pago'], 'strftime'):
                item['fecha_pago_str'] = item['fecha_pago'].strftime("%d/%m/%Y")
            formatted.append(item)
        return formatted

    async def confirmar_match_xml(
        self,
        conn,
        cfdi_data: dict,
        id_comprobante: UUID,
        user_id: UUID,
        guardar_relacion: bool = True,
        forzar_match: bool = False,
    ) -> dict:
        """
        Confirma el match entre un XML y un comprobante de pago.
        Actualiza el comprobante, guarda relacion, conceptos y CFDI relacionados.

        Args:
            conn: Conexion a BD
            cfdi_data: Datos del CFDI (dict del CfdiData)
            id_comprobante: UUID del comprobante seleccionado
            user_id: UUID del usuario
            guardar_relacion: Si guardar la relacion beneficiario↔proveedor

        Returns:
            dict con resultado
        """
        from .db_service import get_db_service
        db_svc = get_db_service()

        uuid_factura = cfdi_data['uuid']
        emisor_rfc = cfdi_data['emisor_rfc']
        tipo_factura = cfdi_data.get('tipo_factura', 'NORMAL')

        # Obtener/crear proveedor
        proveedor = await db_svc.get_proveedor_by_rfc(conn, emisor_rfc)
        if not proveedor:
            proveedor = await db_svc.create_proveedor(
                conn, emisor_rfc, cfdi_data['emisor_nombre']
            )
        id_proveedor = proveedor['id_proveedor']

        # Bloquear el comprobante primero; luego re-verificar el UUID dentro del lock
        # para evitar que una carrera concurrente produzca un 500 en lugar de ValueError.
        comprobante = await db_svc.get_comprobante_by_id(conn, id_comprobante, for_update=True)
        if not comprobante:
            raise ValueError("Comprobante no encontrado")

        if await db_svc.uuid_factura_exists_for_comprobante(conn, id_comprobante, uuid_factura):
            raise ValueError(f"UUID {uuid_factura[:8]}... ya esta vinculado a este comprobante")

        current_estatus = comprobante['estatus']
        if current_estatus not in ('PENDIENTE', 'ANTICIPO', 'PARCIALMENTE_FACTURADO'):
            raise ValueError("El comprobante ya no esta disponible para match")

        cfdi_moneda = (cfdi_data.get('moneda') or "MXN").upper()
        comprobante_moneda = (comprobante.get('moneda') or "MXN").upper()
        if cfdi_moneda != comprobante_moneda:
            raise ValueError(
                f"La moneda del CFDI ({cfdi_moneda}) no coincide con la del comprobante ({comprobante_moneda})"
            )

        monto_factura = Decimal(str(cfdi_data.get('total', 0)))
        monto_pago = Decimal(str(comprobante['monto']))
        monto_ya_facturado = Decimal(str(comprobante.get('monto_facturado') or 0))
        tolerancia_monto = Decimal("0.50")
        saldo_comprobante = monto_pago - monto_ya_facturado

        relacionados = cfdi_data.get('relacionados', [])
        id_comprobante_anticipo = None
        if tipo_factura == "CIERRE_ANTICIPO":
            uuid_anticipo_relacionado = None
            for rel in relacionados:
                tipo_rel = rel.get('tipo_relacion', '') if isinstance(rel, dict) else rel.tipo_relacion
                uuid_rel = rel.get('uuid', '') if isinstance(rel, dict) else rel.uuid
                if tipo_rel == "07" and uuid_rel:
                    uuid_anticipo_relacionado = uuid_rel
                    break

            if uuid_anticipo_relacionado:
                anticipo = await db_svc.get_comprobante_anticipo_by_uuid(
                    conn, uuid_anticipo_relacionado
                )
                if not anticipo:
                    raise ValueError("No se encontro el comprobante de anticipo original relacionado")

                id_comprobante_anticipo = anticipo["id_comprobante"]
                if str(id_comprobante_anticipo) != str(id_comprobante):
                    raise ValueError(
                        "El comprobante seleccionado no corresponde al anticipo "
                        f"relacionado en el CFDI (UUID esperado: {uuid_anticipo_relacionado})"
                    )

                anticipo_proveedor = anticipo.get("id_proveedor")
                if anticipo_proveedor and str(anticipo_proveedor) != str(id_proveedor):
                    raise ValueError("El proveedor del cierre no coincide con el proveedor del anticipo relacionado")
            else:
                if current_estatus != "ANTICIPO":
                    raise ValueError(
                        "El CFDI de cierre no declara la relacion 07 con el anticipo "
                        "original y el comprobante seleccionado no esta en estatus ANTICIPO"
                    )

                comprobante_proveedor = comprobante.get("id_proveedor")
                if comprobante_proveedor and str(comprobante_proveedor) != str(id_proveedor):
                    raise ValueError(
                        "El proveedor del cierre no coincide con el proveedor del "
                        "comprobante de anticipo seleccionado"
                    )

                if monto_factura > monto_pago + tolerancia_monto:
                    exceso = monto_factura - monto_pago
                    raise ValueError(
                        f"El cierre de anticipo excede el monto del comprobante por ${exceso:,.2f} "
                        f"(cierre: ${monto_factura:,.2f}, anticipo: ${monto_pago:,.2f})"
                    )

                logger.warning(
                    "CIERRE_ANTICIPO sin relacion 07; usando comprobante seleccionado como anticipo: %s",
                    id_comprobante,
                )
                id_comprobante_anticipo = id_comprobante

        factura_resumen = await db_svc.get_factura_aplicacion_resumen(conn, uuid_factura)
        saldo_factura = _saldo_factura_desde_resumen(monto_factura, factura_resumen)

        monto_aplicado_raw = cfdi_data.get("monto_aplicado")
        monto_aplicado_editado = monto_aplicado_raw not in (None, "")

        if tipo_factura in ("NOTA_CREDITO", "PAGO"):
            monto_aplicado = Decimal("0")
        else:
            if saldo_comprobante <= Decimal("0"):
                raise ValueError("El comprobante ya no tiene saldo disponible para aplicar facturas")
            if saldo_factura <= Decimal("0"):
                raise ValueError(f"UUID {uuid_factura[:8]}... ya esta cubierto por otros pagos")

            monto_aplicado_sugerido = min(saldo_comprobante, saldo_factura)
            if monto_aplicado_editado:
                try:
                    monto_aplicado = Decimal(str(monto_aplicado_raw))
                except (InvalidOperation, ValueError, TypeError):
                    raise ValueError("El monto a aplicar no es valido")
                if monto_aplicado <= Decimal("0"):
                    raise ValueError("El monto a aplicar debe ser mayor a cero")
                if monto_aplicado > saldo_comprobante + tolerancia_monto:
                    raise ValueError(
                        f"El monto a aplicar excede el saldo del comprobante "
                        f"(${saldo_comprobante:,.2f})"
                    )
                if monto_aplicado > saldo_factura + tolerancia_monto:
                    raise ValueError(
                        f"El monto a aplicar excede el saldo de la factura "
                        f"(${saldo_factura:,.2f})"
                    )
                monto_aplicado = min(monto_aplicado, saldo_comprobante, saldo_factura)
            else:
                monto_aplicado = monto_aplicado_sugerido

        # Validar anti-sobrefacturacion
        if tipo_factura not in ('NOTA_CREDITO', 'ANTICIPO', 'PAGO', 'CIERRE_ANTICIPO'):
            proyectado = monto_ya_facturado + monto_factura
            if proyectado > monto_pago + tolerancia_monto:
                exceso = proyectado - monto_pago
                if not forzar_match and not monto_aplicado_editado:
                    raise ValueError(
                        f"EXCESO_MONTO|{exceso:.2f}|{monto_aplicado:.2f}|"
                        f"La factura excede el monto del pago por ${exceso:,.2f} "
                        f"(ya facturado: ${monto_ya_facturado:,.2f}, "
                        f"nueva factura: ${monto_factura:,.2f}, "
                        f"pago total: ${monto_pago:,.2f}). "
                        f"Se aplicara ${monto_aplicado:,.2f} si confirmas."
                    )
                logger.warning(
                    "Match con monto aplicado menor al CFDI: comprobante=%s exceso=$%s aplicado=$%s user=%s",
                    id_comprobante, f"{exceso:.2f}", f"{monto_aplicado:.2f}", user_id,
                )

        # 1. Insertar en junction table PRIMERO (confirmar_match lee desde aqui)
        try:
            fecha_str = cfdi_data.get('fecha', '')
            fecha_factura = datetime.fromisoformat(fecha_str).date()
        except (ValueError, TypeError):
            fecha_factura = None

        await db_svc.insertar_comprobante_factura(
            conn, id_comprobante, uuid_factura, tipo_factura,
            monto=monto_factura,
            monto_aplicado=monto_aplicado,
            moneda=cfdi_moneda,
            fecha=fecha_factura,
            id_proveedor=id_proveedor,
            rfc_emisor=cfdi_data.get('emisor_rfc'),
            nombre_emisor=cfdi_data.get('emisor_nombre'),
        )

        # 2. Actualizar comprobante (calcula nuevo estatus usando monto_factura)
        await db_svc.confirmar_match(
            conn, id_comprobante, uuid_factura, id_proveedor,
            tipo_factura, current_estatus, monto_factura,
            id_comprobante_anticipo=id_comprobante_anticipo,
            monto_comprobante=monto_pago,
            monto_acumulado=monto_ya_facturado,
            monto_aplicado=monto_aplicado,
        )

        # 2. Guardar relaciones beneficiario↔proveedor (bidireccional)
        if guardar_relacion:
            beneficiario = comprobante['beneficiario_orig']
            # Relacion principal: nombre del beneficiario del PDF
            await db_svc.guardar_relacion_beneficiario(
                conn, beneficiario, id_proveedor, user_id
            )
            # Relacion inversa: razon_social del proveedor (XML)
            razon_social = proveedor.get('razon_social', '')
            if razon_social and razon_social != beneficiario:
                await db_svc.guardar_relacion_beneficiario(
                    conn, razon_social, id_proveedor, user_id
                )
            # Relacion adicional: nombre_comercial (si existe y es diferente)
            nombre_com = proveedor.get('nombre_comercial')
            if nombre_com and nombre_com != beneficiario and nombre_com != razon_social:
                await db_svc.guardar_relacion_beneficiario(
                    conn, nombre_com, id_proveedor, user_id
                )

        bom_item_map = {}
        match_meta_map = {}
        suggestion_map = {}
        bom_svc = None
        conceptos = cfdi_data.get('conceptos', [])
        # Lista filtrada a conceptos-producto. El match y el historial comparten ESTA
        # misma lista (mismos indices): correr el match sobre la lista sin filtrar
        # desalineaba id_bom_item/confianza/origen cuando se descartaba algun concepto.
        conceptos_dicts = [
            {
                'descripcion': c.get('descripcion', c) if isinstance(c, dict) else c.descripcion,
                'cantidad': c.get('cantidad', 0) if isinstance(c, dict) else c.cantidad,
                'valor_unitario': c.get('valor_unitario', 0) if isinstance(c, dict) else c.valor_unitario,
                'importe': c.get('importe', 0) if isinstance(c, dict) else c.importe,
                'unidad': c.get('unidad') if isinstance(c, dict) else c.unidad,
                'clave_prod_serv': c.get('clave_prod_serv') if isinstance(c, dict) else c.clave_prod_serv,
                'clave_unidad': c.get('clave_unidad') if isinstance(c, dict) else c.clave_unidad,
            }
            for c in conceptos
            if _es_concepto_producto(
                c.get('clave_prod_serv') if isinstance(c, dict) else c.clave_prod_serv
            )
        ]
        descartados = len(conceptos) - len(conceptos_dicts)
        if descartados:
            logger.debug(
                "Historial materiales: %d concepto(s) descartado(s) por clave SAT no-producto (UUID=%s)",
                descartados, uuid_factura[:8]
            )

        if comprobante.get('origen') == 'BOM' and comprobante.get('id_bom_pago'):
            try:
                from core.bom.service import get_bom_service
                bom_svc = get_bom_service()

                autorizacion = await bom_svc.get_autorizacion_por_bom_pago(
                    conn, comprobante['id_bom_pago']
                )
                if autorizacion:
                    bom_items = await bom_svc.get_items_por_autorizacion(
                        conn, autorizacion['id']
                    )
                    if bom_items and conceptos_dicts:
                        # Memoria proveedor-producto (clave SAT -> material) del historial confirmado.
                        claves = sorted({
                            (c.get('clave_prod_serv') or '').strip()
                            for c in conceptos_dicts
                            if (c.get('clave_prod_serv') or '').strip()
                        })
                        memoria_map = await bom_svc.get_memoria_match_proveedor(
                            conn, id_proveedor, claves
                        ) if claves else {}
                        # {idx: {id_item, confianza, origen}|None} sobre la lista FILTRADA.
                        match_result = bom_svc.match_conceptos_a_items(
                            conceptos_dicts, bom_items, memoria_map=memoria_map
                        )
                        bom_item_map, match_meta_map, suggestion_map = _separar_matches_bom(
                            match_result
                        )
                        matches_alta = sum(
                            1 for v in match_result.values() if v and v['confianza'] == 'ALTA'
                        )
                        logger.info(
                            "BOM link: autorizacion=%s conceptos=%d items_bom=%d matches=%d (alta=%d sugeridos=%d)",
                            autorizacion['id'], len(conceptos_dicts), len(bom_items),
                            len(bom_item_map) + len(suggestion_map), matches_alta,
                            len(suggestion_map)
                        )
            except Exception:
                logger.exception("BOM auto-link: error no critico, continuando sin vincular")

        # Matcher automatico catalogo interno <-> XML (doc 39, punto 6.2). Independiente
        # del origen BOM del comprobante -- aplica a cualquier factura con conceptos.
        interno_alta_map = {}
        interno_suggestion_map = {}
        if conceptos_dicts:
            try:
                materials_svc = get_materials_service()
                interno_result = await materials_svc.match_conceptos_a_internos(
                    conn, conceptos_dicts, id_proveedor
                )
                interno_alta_map, interno_suggestion_map = _separar_matches_interno(
                    interno_result
                )
                logger.info(
                    "Materiales interno auto-link: conceptos=%d matches_alta=%d sugeridos=%d",
                    len(conceptos_dicts), len(interno_alta_map), len(interno_suggestion_map)
                )
            except Exception:  # devtools: allow-broad-except
                # Feature auxiliar de UX (sugerencia de vinculo): una falla aqui NUNCA
                # debe abortar la confirmacion real de la factura/pago -- mismo criterio
                # que el bloque BOM auto-link de arriba.
                logger.exception(
                    "Materiales interno auto-link: error no critico (match), "
                    "continuando sin vincular"
                )

        # 3. Guardar conceptos en historial de materiales
        # Anticipos y cierres no contienen productos reales — omitir por completo
        if tipo_factura not in ('ANTICIPO', 'CIERRE_ANTICIPO') and conceptos_dicts:
            fecha_str = cfdi_data.get('fecha', '')
            try:
                fecha_factura = datetime.fromisoformat(fecha_str).date()
            except (ValueError, TypeError):
                fecha_factura = today_mx()

            tc_xml = cfdi_data.get('tipo_cambio_xml')
            await db_svc.guardar_conceptos_historial(
                conn, uuid_factura, id_comprobante, id_proveedor,
                conceptos_dicts, fecha_factura, user_id,
                tipo_cambio_xml=tc_xml,
                bom_item_map=bom_item_map,
                match_meta_map=match_meta_map,
                suggestion_map=suggestion_map,
                interno_suggestion_map=interno_suggestion_map,
                tipo_factura=tipo_factura,
                moneda=cfdi_moneda,
            )

            if bom_item_map and bom_svc:
                try:
                    item_ids_matched = [
                        str(v) for v in bom_item_map.values() if v is not None
                    ]
                    if item_ids_matched:
                        await bom_svc.actualizar_estatus_compra(
                            conn, item_ids_matched, 'FACTURADO'
                        )
                        logger.info(
                            "BOM estatus_compra actualizado: %d items → FACTURADO",
                            len(item_ids_matched)
                        )
                except Exception:
                    logger.exception("BOM actualizar estatus_compra: error no critico")

            if interno_alta_map:
                try:
                    materials_svc = get_materials_service()
                    await materials_svc.aplicar_matches_interno_alta(
                        conn, uuid_factura, interno_alta_map
                    )
                except Exception:  # devtools: allow-broad-except
                    # Mismo criterio que arriba: aplicar el vinculo automatico es
                    # best-effort, nunca debe abortar la confirmacion de la factura.
                    logger.exception(
                        "Materiales interno auto-link: error no critico (aplicar), "
                        "continuando sin vincular"
                    )

        # 4. Guardar CFDI relacionados
        if relacionados:
            rel_dicts = [
                {
                    'uuid': r.get('uuid', r) if isinstance(r, dict) else r.uuid,
                    'tipo_relacion': r.get('tipo_relacion', '') if isinstance(r, dict) else r.tipo_relacion,
                    'tipo_relacion_desc': r.get('tipo_relacion_desc') if isinstance(r, dict) else r.tipo_relacion_desc,
                }
                for r in relacionados
            ]
            await db_svc.guardar_cfdi_relacionados(conn, uuid_factura, rel_dicts)

        # 6. Validar integridad: suma de conceptos vs subtotal
        validacion_ok = True
        subtotal_str = cfdi_data.get('subtotal')
        if subtotal_str and conceptos:
            try:
                subtotal_expected = Decimal(str(subtotal_str))
                conceptos_sum = sum(
                    Decimal(str(c.get('importe', 0) if isinstance(c, dict) else c.importe))
                    for c in conceptos
                )
                diff = abs(conceptos_sum - subtotal_expected)
                if diff > Decimal('0.50'):
                    validacion_ok = False
                    logger.warning(
                        "Validacion conceptos: suma=%s != subtotal=%s (diff=%s) UUID=%s",
                        conceptos_sum, subtotal_expected, diff, uuid_factura[:8]
                    )
                else:
                    logger.info(
                        "Validacion conceptos OK: suma=%s ~= subtotal=%s (diff=%s) UUID=%s",
                        conceptos_sum, subtotal_expected, diff, uuid_factura[:8]
                    )
            except (ValueError, TypeError) as e:
                logger.warning("Error validando conceptos: %s", e)
                validacion_ok = False

        # Leer estatus y saldo resultante para informar al usuario
        comprobante_actualizado = await db_svc.get_comprobante_by_id(conn, id_comprobante)
        nuevo_estatus = comprobante_actualizado['estatus'] if comprobante_actualizado else "FACTURADO"
        monto_facturado_nuevo = float(comprobante_actualizado.get('monto_facturado') or 0) if comprobante_actualizado else 0.0
        monto_pago_total = float(comprobante_actualizado.get('monto') or 0) if comprobante_actualizado else 0.0
        saldo_pendiente = monto_pago_total - monto_facturado_nuevo
        saldo_factura_final = max(saldo_factura - monto_aplicado, Decimal("0"))

        logger.info(
            "Match confirmado: UUID=%s, Comprobante=%s, Proveedor=%s, Tipo=%s, Estatus=%s",
            uuid_factura[:8], id_comprobante, emisor_rfc, tipo_factura, nuevo_estatus
        )

        try:
            await db_svc.confirm_xml_staging(conn, uuid_factura)
        except Exception as e:
            logger.warning("No se pudo actualizar staging de XML: %s", e)

        return {
            "uuid_factura": uuid_factura,
            "id_comprobante": str(id_comprobante),
            "id_proveedor": str(id_proveedor),
            "tipo_factura": tipo_factura,
            "conceptos_guardados": len(conceptos),
            "relacionados_guardados": len(relacionados),
            "validacion_ok": validacion_ok,
            "nuevo_estatus": nuevo_estatus,
            "es_parcial": nuevo_estatus == "PARCIALMENTE_FACTURADO",
            "monto_facturado": monto_facturado_nuevo,
            "monto_aplicado": float(monto_aplicado),
            "monto_total": monto_pago_total,
            "saldo_pendiente": saldo_pendiente,
            "saldo_factura": float(saldo_factura_final),
        }

    async def buscar_comprobantes_pendientes(
        self, conn, q: Optional[str] = None, limit: int = 20
    ) -> List[dict]:
        """Busqueda libre de comprobantes pendientes para match manual."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        rows = await db_svc.buscar_comprobantes_pendientes(conn, q, limit)
        return self._format_candidatos(rows)

    # ========================================
    # RELACIONES BENEFICIARIO-PROVEEDOR
    # ========================================

    async def get_relaciones(
        self, conn, q: Optional[str] = None, limit: int = 100
    ) -> List[dict]:
        """Lista relaciones beneficiario-proveedor."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.get_relaciones_all(conn, q=q, limit=limit)

    async def delete_relacion(self, conn, relacion_id: int) -> bool:
        """Elimina una relacion beneficiario-proveedor."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.delete_relacion(conn, relacion_id)

    # ========================================
    # FACTURAS PARCIALES Y REMANENTES
    # ========================================

    async def get_facturas_vinculadas(self, conn, id_comprobante: UUID) -> List[dict]:
        """Lista todas las facturas vinculadas a un comprobante con sus montos."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        rows = await db_svc.get_facturas_comprobante(conn, id_comprobante)
        for r in rows:
            for field in ('monto', 'monto_aplicado', 'saldo_factura'):
                if field in r and isinstance(r[field], Decimal):
                    r[field] = float(r[field])
            if 'fecha' in r and r['fecha'] and hasattr(r['fecha'], 'strftime'):
                r['fecha_str'] = r['fecha'].strftime('%d/%m/%Y')
        return rows

    async def desvincular_factura(
        self, conn, id_comprobante: UUID, uuid_factura: str
    ) -> dict:
        """Desvincula una factura de un comprobante y recalcula su estado."""
        from .db_service import get_db_service
        db_svc = get_db_service()

        # Verificar que la factura existe en este comprobante
        exists = await db_svc.uuid_factura_exists_for_comprobante(conn, id_comprobante, uuid_factura)
        if not exists:
            raise ValueError("La factura no esta vinculada a este comprobante")

        return await db_svc.desvincular_factura(conn, id_comprobante, uuid_factura)

    async def cerrar_remanente(
        self, conn, id_comprobante: UUID, motivo: str, user_id: UUID
    ) -> bool:
        """Cierra un comprobante indicando que no habra mas facturas."""
        from .db_service import get_db_service
        db_svc = get_db_service()

        if not motivo or not motivo.strip():
            raise ValueError("El motivo de cierre es requerido")

        ok = await db_svc.cerrar_remanente(conn, id_comprobante, motivo.strip(), user_id)
        if not ok:
            raise ValueError(
                "El comprobante no puede cerrarse (debe estar en PENDIENTE o PARCIALMENTE_FACTURADO)"
            )
        return True

    async def reabrir_comprobante(self, conn, id_comprobante: UUID) -> bool:
        """Reabre un comprobante CERRADO."""
        from .db_service import get_db_service
        db_svc = get_db_service()

        ok = await db_svc.reabrir_comprobante(conn, id_comprobante)
        if not ok:
            raise ValueError("El comprobante no esta CERRADO o no existe")
        return True

    # ========================================
    # SHAREPOINT - ARCHIVOS
    # ========================================

    async def upload_archivo_sharepoint(
        self, conn, file, subcarpeta: str,
        id_comprobante: Optional[UUID],
        origen_slug: str, user_id: UUID,
        metadata_extra: Optional[dict] = None
    ) -> Optional[dict]:
        """
        Sube un archivo a SharePoint y registra en tb_documentos_attachments.
        Reutiliza patron de levantamientos.

        Args:
            conn: Conexion a BD
            file: UploadFile de FastAPI
            subcarpeta: Ruta relativa (ej: 'compras/facturas_xml/2026-02')
            id_comprobante: UUID del comprobante asociado (puede ser None)
            origen_slug: 'comprobante_pago' o 'factura_xml'
            user_id: UUID del usuario
            metadata_extra: Datos adicionales para JSONB

        Returns:
            dict con url_sharepoint y datos del upload, o None si falla
        """
        from .db_service import get_db_service
        db_svc = get_db_service()

        try:
            from core.microsoft import MicrosoftAuth
            from core.integrations.sharepoint import SharePointService

            ms_auth = MicrosoftAuth()
            app_token = await ms_auth.get_application_token()
            if not app_token:
                logger.error("No se pudo obtener token de SharePoint")
                return None

            sharepoint = SharePointService(access_token=app_token)

            # Construir ruta - Leer base_folder de configuración using db_service
            base_folder = await db_svc.get_config_valor(conn, 'SHAREPOINT_BASE_FOLDER')
            folder_path = f"{base_folder}/{subcarpeta}" if base_folder else subcarpeta

            # Nombre unico
            original_name = file.filename or "archivo"
            timestamp = int(time.time())
            file.filename = f"{timestamp}_{original_name}"

            # Validar tamano using db_service
            max_size_str = await db_svc.get_config_valor(conn, 'MAX_UPLOAD_SIZE_MB')
            max_size_mb = float(max_size_str) if max_size_str else 50.0

            file.file.seek(0, 2)
            f_size = file.file.tell()
            file.file.seek(0)

            if f_size / (1024 * 1024) > max_size_mb:
                logger.warning("Archivo %s excede limite: %d bytes", original_name, f_size)
                return None

            # Upload
            upload_result = await sharepoint.upload_file(conn, file, folder_path)

            # Metadata
            meta = {
                "nombre_original": original_name,
                "content_type": getattr(file, 'content_type', None) or 'application/octet-stream',
            }
            if id_comprobante:
                meta["id_comprobante"] = str(id_comprobante)
            if metadata_extra:
                meta.update(metadata_extra)

            # Registrar en BD
            doc_id = await db_svc.registrar_archivo_sharepoint(
                conn, id_comprobante, origen_slug,
                upload_result, user_id, meta
            )

            logger.info(
                "Archivo subido a SharePoint: %s -> %s",
                original_name, upload_result.get('webUrl', '')
            )

            parent_ref = upload_result.get('parentReference') or {}

            return {
                "id_documento_attachment": str(doc_id),
                "url_sharepoint": upload_result.get('webUrl', ''),
                "nombre": upload_result.get('name', ''),
                "drive_item_id": upload_result.get('id', ''),
                "parent_drive_id": parent_ref.get('driveId'),
                "folder_path": folder_path,
                "tipo_contenido": meta["content_type"],
                "tamano_bytes": upload_result.get('size', f_size),
            }

        except (ValueError, RuntimeError, OSError, asyncpg.PostgresError, httpx.HTTPError) as e:
            logger.error("Error subiendo archivo a SharePoint: %s", e, exc_info=True)
            return None

    async def get_archivos_comprobante(
        self, conn, id_comprobante: UUID
    ) -> List[dict]:
        """Obtiene archivos asociados a un comprobante."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.get_archivos_comprobante(conn, id_comprobante)

    async def get_xml_staging_pendientes(self, conn) -> list[dict]:
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.get_xml_staging_pendientes(conn)

    async def eliminar_xml_staging(self, conn, uuid_factura: str) -> bool:
        from .db_service import get_db_service
        db_svc = get_db_service()
        return await db_svc.delete_xml_staging(conn, uuid_factura)


    async def buscar_comprobantes_para_grupo(
        self, conn, q: Optional[str], moneda: str = 'MXN', limit: int = 30
    ) -> List[dict]:
        """Comprobantes pendientes para el panel de vinculacion en grupo."""
        from .db_service import get_db_service
        db_svc = get_db_service()
        rows = await db_svc.buscar_comprobantes_pendientes_para_grupo(conn, q, moneda, limit)
        return self._format_candidatos(rows)

    async def confirmar_match_grupo(
        self,
        conn,
        facturas_data: List[dict],
        comprobante_ids: List[UUID],
        user_id: UUID,
        forzar_excepcion: bool = False,
        motivo_excepcion: Optional[str] = None,
    ) -> dict:
        """Vincula N facturas XML a M comprobantes mediante distribucion greedy.

        Requiere que todas las facturas sean del mismo RFC.
        Diferencias mayores a la tolerancia normal requieren motivo de excepcion.
        """
        from .db_service import get_db_service
        db_svc = get_db_service()

        if not facturas_data:
            raise ValueError("No hay facturas para vincular")
        if not comprobante_ids:
            raise ValueError("No hay comprobantes seleccionados")

        rfcs = {f.get('emisor_rfc', '') for f in facturas_data}
        if len(rfcs) > 1:
            raise ValueError(f"Las facturas deben ser del mismo proveedor (RFCs: {', '.join(rfcs)})")

        comprobantes = await db_svc.get_comprobantes_by_ids(conn, comprobante_ids, for_update=True)
        if len(comprobantes) != len(comprobante_ids):
            raise ValueError("Uno o mas comprobantes no fueron encontrados")

        validos = ('PENDIENTE', 'PARCIALMENTE_FACTURADO', 'ANTICIPO')
        invalidos = [c for c in comprobantes if c['estatus'] not in validos]
        if invalidos:
            nombres = ', '.join(c.get('beneficiario_orig', '?') for c in invalidos)
            raise ValueError(f"Comprobantes no disponibles para match: {nombres}")

        def saldo_disponible(comprobante: dict) -> Decimal:
            monto = Decimal(str(comprobante['monto']))
            monto_facturado = Decimal(str(comprobante.get('monto_facturado') or 0))
            return monto - monto_facturado

        tolerancia_grupo = MATCH_TOLERANCIA
        sum_facturas = sum(Decimal(str(f.get('total', 0))) for f in facturas_data)
        sum_pagos = sum(saldo_disponible(c) for c in comprobantes)
        diferencia_grupo = abs(sum_facturas - sum_pagos)
        requiere_excepcion = diferencia_grupo > tolerancia_grupo

        motivo_limpio = (motivo_excepcion or "").strip()
        if requiere_excepcion and not forzar_excepcion:
            raise ValueError(
                f"La diferencia de montos supera ${tolerancia_grupo:,.2f}: "
                f"facturas ${sum_facturas:,.2f} vs pagos ${sum_pagos:,.2f} "
                f"(diferencia: ${diferencia_grupo:,.2f}). Captura un motivo de excepcion."
            )
        if requiere_excepcion and not motivo_limpio:
            raise ValueError("Captura el motivo de excepcion para vincular fuera de tolerancia")

        facturas_sorted = sorted(facturas_data, key=lambda f: Decimal(str(f.get('total', 0))), reverse=True)
        pagos_sorted = sorted(comprobantes, key=saldo_disponible, reverse=True)

        pago_balances = {
            str(c['id_comprobante']): saldo_disponible(c)
            for c in pagos_sorted
        }

        assignments = []
        cero = Decimal("0.005")

        for factura in facturas_sorted:
            total_factura = Decimal(str(factura.get('total', 0)))
            remaining = total_factura
            for comp_id in pago_balances:
                if remaining <= cero:
                    break
                balance = pago_balances[comp_id]
                if balance <= cero:
                    continue
                aplicar = min(remaining, balance)
                is_full = aplicar >= total_factura - cero
                assignments.append({
                    'factura': factura,
                    'id_comprobante': UUID(comp_id),
                    'uuid_factura': factura.get('uuid', ''),
                    'monto_factura': total_factura,
                    'monto_aplicado': aplicar,
                    'is_full': is_full,
                })
                pago_balances[comp_id] -= aplicar
                remaining -= aplicar

        comprobantes_con_asignacion = list(dict.fromkeys(
            assign['id_comprobante'] for assign in assignments
        ))
        resultados_match = []
        for assign in assignments:
            factura_data = dict(assign['factura'])
            if not assign['is_full']:
                factura_data['monto_aplicado'] = str(assign['monto_aplicado'])
            else:
                factura_data.pop('monto_aplicado', None)

            resultado_match = await self.confirmar_match_xml(
                conn, factura_data, assign['id_comprobante'], user_id,
                guardar_relacion=True, forzar_match=True,
            )
            resultados_match.append(resultado_match)

        comprobantes_cerrados = []
        for id_comprobante in comprobantes_con_asignacion:
            comprobante_actual = await db_svc.get_comprobante_by_id(conn, id_comprobante)
            if not comprobante_actual:
                continue

            monto_pago = Decimal(str(comprobante_actual.get('monto') or 0))
            monto_facturado = Decimal(str(comprobante_actual.get('monto_facturado') or 0))
            diferencia_comprobante = monto_pago - monto_facturado
            debe_cerrar_por_tolerancia = abs(diferencia_comprobante) <= tolerancia_grupo
            debe_cerrar_por_excepcion = requiere_excepcion and forzar_excepcion

            if comprobante_actual.get('estatus') not in ('PARCIALMENTE_FACTURADO', 'FACTURADO'):
                continue
            if abs(diferencia_comprobante) <= cero:
                continue
            if not debe_cerrar_por_tolerancia and not debe_cerrar_por_excepcion:
                continue

            if debe_cerrar_por_excepcion:
                motivo = (
                    f"Excepcion match grupal XML: {motivo_limpio}. "
                    f"Diferencia: ${diferencia_comprobante:,.2f}."
                )
            else:
                motivo = (
                    "Cerrado automaticamente por match grupal XML. "
                    f"Diferencia dentro de tolerancia: ${diferencia_comprobante:,.2f}."
                )

            cierre = await db_svc.cerrar_remanente_automatico(
                conn, id_comprobante, motivo, user_id
            )
            if cierre and cierre.get('cerrado'):
                comprobantes_cerrados.append({
                    'id_comprobante': str(id_comprobante),
                    'monto_remanente': float(cierre.get('monto_remanente') or 0),
                    'motivo': motivo,
                })

        asignaciones_result = [
            {
                'uuid_factura': assign['uuid_factura'],
                'id_comprobante': str(assign['id_comprobante']),
                'monto_factura': float(assign['monto_factura']),
                'monto_aplicado': float(assign['monto_aplicado']),
            }
            for assign in assignments
        ]

        return {
            'total_facturas': len(facturas_data),
            'total_comprobantes': len(comprobante_ids),
            'sum_facturas': float(sum_facturas),
            'sum_pagos': float(sum_pagos),
            'diferencia': float(diferencia_grupo),
            'tolerancia': float(tolerancia_grupo),
            'requiere_excepcion': requiere_excepcion,
            'motivo_excepcion': motivo_limpio if forzar_excepcion else '',
            'asignaciones': asignaciones_result,
            'matches': resultados_match,
            'comprobantes_cerrados': comprobantes_cerrados,
        }


def get_compras_service():
    """Dependency injection para FastAPI."""
    return ComprasService()
