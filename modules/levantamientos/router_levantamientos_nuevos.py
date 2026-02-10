# ==============================================================
# modules/levantamientos/router_levantamientos_nuevos.py
#
# Endpoints nuevos para posponer, reagendar y viaticos.
# Se agregan al router existente (modules/levantamientos/router.py).
#
# CÓMO INTEGRAR:
#   En router.py, al final del archivo, agregar:
#
#       from .router_levantamientos_nuevos import register_nuevos_endpoints
#       register_nuevos_endpoints(router)
#
#   Eso conecta todos los endpoints de este archivo al mismo
#   APIRouter que ya existe, sin duplicar prefijo ni tags.
# ==============================================================

from fastapi import APIRouter, Depends, Request, Form, HTTPException, File, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from typing import Optional, List
from uuid import UUID, uuid4
from datetime import datetime, date
from zoneinfo import ZoneInfo
from io import BytesIO
import json
import logging
import asyncio
import time

from core.security import get_current_user_context
from core.permissions import require_module_access
from core.database import get_db_connection
from core.microsoft import MicrosoftAuth

from .service import get_service, LevantamientoService
from .db_service import get_db_service, LevantamientosDBService

logger = logging.getLogger("Levantamientos.Router.Nuevos")

templates = Jinja2Templates(directory="templates")


def register_nuevos_endpoints(router: APIRouter):
    """
    Registra todos los endpoints nuevos en el router existente.
    Llamar una sola vez desde router.py.
    """

    # ==============================================================
    # GET — MODALES (renderizar templates con datos)
    # ==============================================================

    @router.get("/modal/posponer/{id_levantamiento}", include_in_schema=False)
    async def get_modal_posponer(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """Renderiza el modal de posponer con datos del levantamiento."""
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        return templates.TemplateResponse("levantamientos/modals/posponer_modal.html", {
            "request": request,
            "lev_data": lev,
            "has_active_viaticos": await db_svc.check_viaticos_sent(conn, id_levantamiento)
        })

    # ----------------------------------------------------------
    
    @router.get("/modals/detalle/{id_levantamiento}", include_in_schema=False)
    async def get_detalle_levantamiento_modal(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        # Accesible para anyone with viewer access (Shared Modal)
        _=require_module_access("levantamientos", "viewer"),
    ):
        """
        Renderiza el modal de DETALLE COMPLETO.
        Accesible desde Comercial y Simulación.
        """
        lev = await db_svc.get_detalle_completo(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        return templates.TemplateResponse("shared/modals/detalle_levantamiento_modal.html", {
            "request": request,
            "lev": lev,
        })

    # ----------------------------------------------------------

    @router.get("/modal/historial/{id_levantamiento}", include_in_schema=False)
    async def get_modal_historial(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "viewer"),
    ):
        """Renderiza el modal de historial con timeline de cambios.
        
        Usa el sistema de modales global (#modal-content).
        """
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        historial = await service.get_historial_estados(conn, id_levantamiento)

        return templates.TemplateResponse("shared/modals/historial_levantamiento_modal.html", {
            "request": request,
            "lev_data": lev,
            "historial": historial,
        })

    # ----------------------------------------------------------

    @router.get("/modal/reagendar/{id_levantamiento}", include_in_schema=False)
    async def get_modal_reagendar(
        request: Request,
        id_levantamiento: UUID,
        desde: str = "pendiente",   # pendiente | pospuesto
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Renderiza el modal de reagendar.
        desde=pendiente  → agendar desde estado 8
        desde=pospuesto  → reagendar desde estado 13
        """
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        today_str = datetime.now(ZoneInfo("America/Mexico_City")).strftime("%Y-%m-%dT%H:%M")

        return templates.TemplateResponse("levantamientos/modals/reagendar_modal.html", {
            "request": request,
            "lev_data": lev,
            "desde": desde,
            "today_str": today_str,
        })

    # ----------------------------------------------------------

    @router.get("/modal/viaticos/{id_levantamiento}", include_in_schema=False)
    async def get_modal_viaticos(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Renderiza el modal de viaticos con:
          - datos del levantamiento
          - lista de viaticos actuales
          - usuarios disponibles (select)
          - CC configurados desde tb_config_emails
          - historial de envíos previos
        """
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        viaticos        = await db_svc.get_viaticos(conn, id_levantamiento)
        usuarios        = await db_svc.get_usuarios_viaticos(conn)
        cc_configurados = await db_svc.get_cc_configurados_viaticos(conn)
        historial       = await db_svc.get_historial_envios(conn, id_levantamiento)

        return templates.TemplateResponse("levantamientos/modals/viaticos_modal.html", {
            "request": request,
            "lev_data": lev,
            "viaticos": viaticos,
            "usuarios": usuarios,
            "cc_configurados": cc_configurados,
            "historial_envios": historial,
        })

    # ==============================================================
    # POST — POSPONER
    # ==============================================================

    @router.post("/posponer/{id_levantamiento}")
    async def posponer_endpoint(
        request: Request,
        id_levantamiento: UUID,
        motivo_pospone: str = Form(...),
        devolver_viaticos: Optional[bool] = Form(False),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Valida motivo (min 10 chars).
        2. Obtiene estado actual para el historial.
        3. Ejecuta UPDATE via db_service.
        4. Registra en historial via service.
        5. Retorna Kanban actualizado (outerHTML).
        """
        if not motivo_pospone or len(motivo_pospone.strip()) < 10:
            raise HTTPException(status_code=400, detail="El motivo debe tener al menos 10 caracteres.")

        # Estado actual antes del cambio
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        estado_anterior = lev["id_estatus_global"]

        # UPDATE
        await db_svc.update_posponer(conn, id_levantamiento, motivo_pospone.strip(), context["user_db_id"])

        # Historial
        await service._registrar_en_historial(
            conn=conn,
            id_levantamiento=id_levantamiento,
            estatus_anterior=estado_anterior,
            estatus_nuevo=13,
            user_context=context,
            observaciones=motivo_pospone.strip(),
            metadata={"tipo_cambio": "posponer"}
        )

        # Si se marcó devolver viáticos
        if devolver_viaticos:
            await service.registrar_devolucion(conn, id_levantamiento, context)

        # Notificar (Background)
        asyncio.create_task(
            service._execute_notification_background(
                service._notificar_pospuesto_impl,
                id_oportunidad=lev["id_oportunidad"],
                motivo=motivo_pospone.strip(),
                user_context=context
            )
        )

        # Retornar kanban completo
        return await _render_kanban(request, conn, service, context)

    # ==============================================================
    # POST — REAGENDAR
    # ==============================================================

    @router.post("/reagendar/{id_levantamiento}")
    async def reagendar_endpoint(
        request: Request,
        id_levantamiento: UUID,
        nueva_fecha_visita: str = Form(...),
        observaciones: Optional[str] = Form(None),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Valida formato de fecha y que no sea pasada.
        2. Obtiene estado actual.
        3. Ejecuta UPDATE via db_service.
        4. Registra en historial.
        5. Retorna Kanban actualizado.
        """
        # Validar fecha y hora
        if not nueva_fecha_visita:
            raise HTTPException(status_code=400, detail="Se requiere fecha de visita.")

        try:
            # datetime-local envía formato "YYYY-MM-DDTHH:MM"
            fecha_obj = datetime.fromisoformat(nueva_fecha_visita).replace(tzinfo=ZoneInfo("America/Mexico_City"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha/hora inválido.")

        now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
        if fecha_obj.date() < now_mx.date():
            raise HTTPException(status_code=400, detail="La fecha no puede ser anterior a hoy.")

        # Estado actual
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        estado_anterior = lev["id_estatus_global"]

        # UPDATE
        # Si estaba en Pendiente (8), es cita inicial => NO es rescheduling
        is_rescheduling = (estado_anterior != 8)
        await db_svc.update_reagendar(conn, id_levantamiento, fecha_obj, context["user_db_id"], is_rescheduling=is_rescheduling)

        # Historial
        fecha_display = fecha_obj.strftime("%d/%m/%Y %H:%M")
        obs_text = observaciones or f"Visita reagendada para {fecha_display}"
        await service._registrar_en_historial(
            conn=conn,
            id_levantamiento=id_levantamiento,
            estatus_anterior=estado_anterior,
            estatus_nuevo=9,
            user_context=context,
            observaciones=obs_text,
            metadata={"tipo_cambio": "reagendar", "nueva_fecha": fecha_obj.isoformat()}
        )

        # Notificar (Background)
        asyncio.create_task(
            service._execute_notification_background(
                service._notificar_agendado_impl,
                id_oportunidad=lev["id_oportunidad"],
                fecha_visita=fecha_display,
                user_context=context
            )
        )

        # Verificar si hay ingenieros asignados para notificar
        # Debugging logging
        has_techs_new = await conn.fetchval("""
            SELECT EXISTS(SELECT 1 FROM tb_levantamiento_asignaciones WHERE id_levantamiento = $1)
        """, id_levantamiento)
        
        has_techs_legacy = await conn.fetchval("""
            SELECT (tecnico_asignado_id IS NOT NULL) FROM tb_levantamientos WHERE id_levantamiento = $1
        """, id_levantamiento)

        has_techs = has_techs_new or has_techs_legacy

        logger.info(f"Reagendar Validation - ID: {id_levantamiento}, HasTechsNew: {has_techs_new}, HasTechsLegacy: {has_techs_legacy}, Final: {has_techs}")
        
        notification = None
        if not has_techs:
            logger.info("Triggering Warning Toast: No technicians assigned.")
            notification = {
                "title": "Asignación Pendiente",
                "message": "El levantamiento ha sido agendado. Recuerda asignar un ingeniero.",
                "type": "warning"
            }

        return await _render_kanban(request, conn, service, context, notification)

    # ==============================================================
    # POST / DELETE — VIATICOS CRUD
    # ==============================================================

    @router.post("/viaticos/{id_levantamiento}")
    async def crear_viatico_endpoint(
        request: Request,
        id_levantamiento: UUID,
        usuario_id: UUID = Form(...),
        concepto: str = Form(...),
        monto: float = Form(...),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Crea un viatico y retorna el innerHTML de #tabla-viaticos-container
        con la tabla actualizada + total.
        """
        if not concepto or not concepto.strip():
            raise HTTPException(status_code=400, detail="El concepto es obligatorio.")
        if monto <= 0:
            raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0.")

        await db_svc.create_viatico(
            conn, id_levantamiento, usuario_id, concepto.strip(), monto, context["user_db_id"]
        )

        # Recargar lista completa para retornar partial
        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)

        return templates.TemplateResponse("levantamientos/partials/tabla_viaticos.html", {
            "request": request,
            "viaticos": viaticos,
            "id_levantamiento": id_levantamiento,
        })

    # ----------------------------------------------------------

    @router.delete("/viaticos/{id_levantamiento}/{viatico_id}")
    async def eliminar_viatico_endpoint(
        request: Request,
        id_levantamiento: UUID,
        viatico_id: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Elimina un viatico y retorna el innerHTML actualizado
        de #tabla-viaticos-container.
        """
        eliminado = await db_svc.delete_viatico(conn, id_levantamiento, viatico_id)
        if not eliminado:
            raise HTTPException(status_code=404, detail="Viatico no encontrado.")

        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)

        return templates.TemplateResponse("levantamientos/partials/tabla_viaticos.html", {
            "request": request,
            "viaticos": viaticos,
            "id_levantamiento": id_levantamiento,
        })

    # ==============================================================
    # POST — ENVIAR SOLICITUD DE VIATICOS
    # ==============================================================

    @router.post("/viaticos/solicitud/{id_levantamiento}")
    async def enviar_solicitud_viaticos_endpoint(
        request: Request,
        id_levantamiento: UUID,
        cc_adicionales: str = Form(""),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Obtiene viaticos actuales (debe haber al menos 1).
        2. Construye TO y CC (configurados + manuales).
        3. Renderiza solicitud_viaticos.html como body del correo.
        4. Envía via MicrosoftAuth.
        5. Registra en tb_levantamiento_viaticos_historico con snapshot.
        6. Retorna innerHTML de #historial-envios-container actualizado.
        """
        # 1. Viaticos
        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)
        if not viaticos:
            raise HTTPException(status_code=400, detail="No hay viaticos registrados.")

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        total_monto = sum(v["monto"] for v in viaticos)

        # 2. Destinatarios
        to_list = await db_svc.get_to_configurados_viaticos(conn)
        cc_configurados = await db_svc.get_cc_configurados_viaticos(conn)

        import re
        # CC manuales: separados por ";" o ","
        cc_manuales = [e.strip() for e in re.split(r'[;,]', cc_adicionales) if e.strip() and "@" in e.strip()]

        # Unir CC sin duplicados, quitar TO de CC
        cc_all = list(set(cc_configurados + cc_manuales) - set(to_list))

        if not to_list:
            raise HTTPException(status_code=500, detail="No hay destinatarios TO configurados para SOLICITUD_VIATICOS.")

        # 3. Renderizar template de email
        now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
        fecha_envio_str = now_mx.strftime("%d/%m/%Y %H:%M")

        email_template = templates.get_template("levantamientos/emails/solicitud_viaticos.html")
        html_body = email_template.render(
            proyecto_nombre=lev["nombre_proyecto"] or lev["titulo_proyecto"] or "Sin nombre",
            op_id=lev["op_id_estandar"],
            cliente_nombre=lev["cliente_nombre"],
            sitio_direccion=lev.get("sitio_direccion"),
            enviado_por=context.get("user_name", "Sistema"),
            fecha_envio=fecha_envio_str,
            viaticos=viaticos,
            total_monto=total_monto,
        )

        # 4. Enviar correo
        subject = f"Solicitud de Viaticos — {lev['op_id_estandar']} | {lev['cliente_nombre']}"

        # Usar buzón de notificaciones configurado (NO email personal del usuario)
        sender_config = await conn.fetchrow("""
            SELECT email_remitente FROM tb_correos_notificaciones
            WHERE departamento = 'LEVANTAMIENTOS' AND activo = true
            LIMIT 1
        """)
        if not sender_config:
            sender_config = await conn.fetchrow("""
                SELECT email_remitente FROM tb_correos_notificaciones
                WHERE departamento = 'DEFAULT' AND activo = true
                LIMIT 1
            """)
        sender_email = sender_config['email_remitente'] if sender_config else 'app-notifications@enertika.mx'

        estatus_envio = "enviado"
        error_detalle = None

        try:
            ms_auth = MicrosoftAuth()
            app_token = await ms_auth.get_application_token()

            if not app_token:
                raise Exception("No se pudo obtener token de aplicación de Microsoft Graph.")

            success, msg = await ms_auth.send_email_with_attachments(
                access_token=app_token,
                from_email=sender_email,
                subject=subject,
                body=html_body,
                recipients=to_list,
                cc_recipients=cc_all if cc_all else None,
                importance="normal",
            )

            if not success:
                estatus_envio = "error"
                error_detalle = msg
                logger.error(f"[VIATICOS] Error envío correo lev {id_levantamiento}: {msg}")
            else:
                logger.info(f"[VIATICOS] Correo enviado exitosamente lev {id_levantamiento}")

        except Exception as exc:
            estatus_envio = "error"
            error_detalle = str(exc)
            logger.error(f"[VIATICOS] Excepción envío correo lev {id_levantamiento}: {exc}")

        # 5. Registrar historial (siempre, incluso si falla el correo)
        snapshot = [
            {"usuario_nombre": v["usuario_nombre"], "concepto": v["concepto"], "monto": float(v["monto"])}
            for v in viaticos
        ]

        await db_svc.insert_historial_envio(
            conn=conn,
            id_levantamiento=id_levantamiento,
            enviado_por_id=context["user_db_id"],
            enviado_por_nombre=context.get("user_name", "Sistema"),
            to_destinatarios=to_list,
            cc_destinatarios=cc_all,
            viaticos_snapshot=snapshot,
            total_monto=total_monto,
            estatus=estatus_envio,
            error_detalle=error_detalle,
        )

        # 6. Retornar historial actualizado
        historial = await db_svc.get_historial_envios(conn, id_levantamiento)

        return templates.TemplateResponse("levantamientos/partials/historial_envios.html", {
            "request": request,
            "historial_envios": historial,
        })

    # ==============================================================
    # GET — MODAL DE ENTREGA
    # ==============================================================

    @router.get("/modal/entrega/{id_levantamiento}", include_in_schema=False)
    async def get_modal_entrega(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """Renderiza el modal de entrega con datos del levantamiento."""
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        return templates.TemplateResponse("levantamientos/modals/entrega_modal.html", {
            "request": request,
            "lev_data": lev,
        })

    # ==============================================================
    # POST — ENTREGAR LEVANTAMIENTO (con archivos opcionales)
    # ==============================================================

    @router.post("/entregar/{id_levantamiento}", include_in_schema=False)
    async def entregar_levantamiento(
        request: Request,
        id_levantamiento: UUID,
        observaciones: Optional[str] = Form(None),
        file_uploads: List[UploadFile] = File(None),
        conn=Depends(get_db_connection),
        service: LevantamientoService = Depends(get_service),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Marca un levantamiento como entregado (estado 12).
        Opcionalmente sube archivos a SharePoint y los registra en BD.
        """
        notification = None
        try:
            # 1. Cambiar estado a 12 (Entregado)
            await service.cambiar_estado(
                conn, id_levantamiento, 12, context,
                observaciones=observaciones.strip() if observaciones else None
            )

            # 2. Si hay archivos, subirlos a SharePoint
            files_uploaded = 0
            if file_uploads and any(f.filename for f in file_uploads):
                try:
                    lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
                    op_id_estandar = lev["op_id_estandar"] if lev else None

                    if not op_id_estandar:
                        logger.warning(f"[ENTREGA] No se pudo subir adjuntos: op_id_estandar es NULL para {id_levantamiento}")
                    else:
                        # Leer configuracion
                        config_rows = await conn.fetch("""
                            SELECT clave, valor FROM tb_configuracion_global
                            WHERE clave IN ('MAX_UPLOAD_SIZE_MB', 'SHAREPOINT_BASE_FOLDER')
                        """)
                        config_map = {row['clave']: row['valor'] for row in config_rows}
                        max_size_mb = int(config_map.get('MAX_UPLOAD_SIZE_MB', '10'))
                        base_folder = config_map.get('SHAREPOINT_BASE_FOLDER', '').strip().strip("/")

                        # Carpeta destino en SharePoint
                        relative_path = f"levantamiento/{op_id_estandar}/entrega"
                        folder_path = f"{base_folder}/{relative_path}" if base_folder else relative_path

                        # Obtener token de aplicacion
                        ms_auth = MicrosoftAuth()
                        app_token = await ms_auth.get_application_token()
                        if not app_token:
                            logger.error(f"[ENTREGA] No se pudo obtener token de aplicacion para SharePoint")
                        else:
                            from core.integrations.sharepoint import SharePointService
                            sharepoint = SharePointService(access_token=app_token)

                            for f_obj in file_uploads:
                                if not f_obj.filename:
                                    continue
                                try:
                                    # Validar tamano
                                    f_obj.file.seek(0, 2)
                                    f_size = f_obj.file.tell()
                                    f_obj.file.seek(0)

                                    if f_size / (1024 * 1024) > max_size_mb:
                                        logger.warning(f"[ENTREGA] Archivo {f_obj.filename} excede limite: {f_size} bytes")
                                        continue

                                    # Nombre unico
                                    timestamp = int(time.time())
                                    original_name = f_obj.filename
                                    f_obj.filename = f"{timestamp}_{original_name}"

                                    logger.info(f"[ENTREGA] Subiendo archivo: {f_obj.filename} a {folder_path}")

                                    # Upload a SharePoint
                                    upload_result = await sharepoint.upload_file(conn, f_obj, folder_path)

                                    # Registrar en BD
                                    doc_id = uuid4()
                                    parent_ref = upload_result.get('parentReference', {})

                                    await conn.execute("""
                                        INSERT INTO tb_documentos_attachments (
                                            id_documento, nombre_archivo, url_sharepoint, drive_item_id, parent_drive_id,
                                            tipo_contenido, tamano_bytes, id_oportunidad, subido_por_id,
                                            origen_slug, activo, metadata
                                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'comentario', TRUE, $10::jsonb)
                                    """,
                                        doc_id,
                                        upload_result['name'],
                                        upload_result['webUrl'],
                                        upload_result['id'],
                                        parent_ref.get('driveId'),
                                        f_obj.content_type,
                                        upload_result['size'],
                                        lev["id_oportunidad"],
                                        context["user_db_id"],
                                        json.dumps({
                                            "id_levantamiento": str(id_levantamiento),
                                            "tipo": "entrega",
                                            "nombre_original": original_name
                                        })
                                    )

                                    files_uploaded += 1
                                    logger.info(f"[ENTREGA] Adjunto registrado: {upload_result['name']}")

                                except Exception as e_file:
                                    logger.error(f"[ENTREGA] Error subiendo archivo {f_obj.filename}: {e_file}")

                except Exception as e_sp:
                    logger.error(f"[ENTREGA] Fallo general en adjuntos: {e_sp}")

            # 3. Toast de exito
            msg = "Levantamiento entregado exitosamente."
            if files_uploaded > 0:
                msg += f" {files_uploaded} archivo(s) adjuntado(s)."
            notification = {"title": "Entregado", "message": msg, "type": "success"}

        except HTTPException:
            raise
        except ValueError as ve:
            notification = {"title": "Error", "message": str(ve), "type": "error"}
        except Exception as exc:
            logger.error(f"[ENTREGA] Error inesperado: {exc}", exc_info=True)
            notification = {"title": "Error", "message": "Ocurrio un error al procesar la entrega.", "type": "error"}

        return await _render_kanban(request, conn, service, context, notification)

    # ==============================================================
    # GET — REPORTE EXCEL DE GASTOS (VIATICOS)
    # ==============================================================

    @router.get("/reporte-gastos/{id_levantamiento}", include_in_schema=False)
    async def reporte_gastos_excel(
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos"),
    ):
        """
        Genera y descarga un reporte Excel con los gastos (viaticos)
        y el historial de envios de un levantamiento.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)
        historial = await db_svc.get_historial_envios(conn, id_levantamiento)

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=11, color="333333")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_format = '#,##0.00'

        wb = Workbook()

        # ===================== HOJA 1: RESUMEN =====================
        ws1 = wb.active
        ws1.title = "Resumen"

        # Titulo
        ws1.merge_cells('A1:D1')
        cell_title = ws1.cell(row=1, column=1, value="Reporte de Gastos - Levantamiento")
        cell_title.font = title_font
        cell_title.alignment = Alignment(horizontal="center")

        # Info del proyecto
        info_data = [
            ("OP-ID:", lev.get("op_id_estandar", "")),
            ("Cliente:", lev.get("cliente_nombre", "")),
            ("Proyecto:", lev.get("nombre_proyecto") or lev.get("titulo_proyecto") or "Sin nombre"),
            ("Direccion:", lev.get("direccion") or lev.get("sitio_direccion") or "Sin direccion"),
        ]
        for i, (label, value) in enumerate(info_data, start=3):
            ws1.cell(row=i, column=1, value=label).font = subtitle_font
            ws1.cell(row=i, column=2, value=value)

        # Tabla de viaticos
        viaticos_start_row = len(info_data) + 5
        ws1.cell(row=viaticos_start_row - 1, column=1, value="Detalle de Viaticos").font = subtitle_font

        viat_headers = ["Usuario", "Concepto", "Monto"]
        for col_num, header in enumerate(viat_headers, 1):
            cell = ws1.cell(row=viaticos_start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        total_monto = 0
        for row_num, v in enumerate(viaticos, start=viaticos_start_row + 1):
            ws1.cell(row=row_num, column=1, value=v.get("usuario_nombre", "")).border = thin_border
            ws1.cell(row=row_num, column=2, value=v.get("concepto", "")).border = thin_border
            monto_cell = ws1.cell(row=row_num, column=3, value=float(v.get("monto", 0)))
            monto_cell.number_format = money_format
            monto_cell.border = thin_border
            total_monto += float(v.get("monto", 0))

        # Fila total
        total_row = viaticos_start_row + len(viaticos) + 1
        ws1.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True, size=11)
        total_cell = ws1.cell(row=total_row, column=3, value=total_monto)
        total_cell.font = Font(bold=True, size=11)
        total_cell.number_format = money_format

        # Ajustar anchos
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 40

        # ===================== HOJA 2: HISTORIAL ENVIOS =====================
        ws2 = wb.create_sheet("Historial Envios")

        ws2.merge_cells('A1:F1')
        cell_title2 = ws2.cell(row=1, column=1, value="Historial de Solicitudes de Viaticos")
        cell_title2.font = title_font
        cell_title2.alignment = Alignment(horizontal="center")

        hist_headers = ["Fecha Envio", "Enviado Por", "Destinatarios TO", "Destinatarios CC", "Total", "Estatus"]
        for col_num, header in enumerate(hist_headers, 1):
            cell = ws2.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_num, h in enumerate(historial, start=4):
            fecha = h.get("fecha_envio")
            fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if fecha else ""

            to_list = h.get("to_destinatarios") or []
            cc_list = h.get("cc_destinatarios") or []
            to_str = ", ".join(to_list) if isinstance(to_list, list) else str(to_list)
            cc_str = ", ".join(cc_list) if isinstance(cc_list, list) else str(cc_list)

            ws2.cell(row=row_num, column=1, value=fecha_str).border = thin_border
            ws2.cell(row=row_num, column=2, value=h.get("enviado_por_nombre", "")).border = thin_border
            ws2.cell(row=row_num, column=3, value=to_str).border = thin_border
            ws2.cell(row=row_num, column=4, value=cc_str).border = thin_border
            monto_h = ws2.cell(row=row_num, column=5, value=float(h.get("total_monto", 0)))
            monto_h.number_format = money_format
            monto_h.border = thin_border
            ws2.cell(row=row_num, column=6, value=h.get("estatus", "")).border = thin_border

        # Ajustar anchos
        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 35
        ws2.column_dimensions['D'].width = 35
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 15

        # Guardar a BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        op_id = lev.get("op_id_estandar", "SIN_ID").replace("/", "-")
        filename = f"Reporte_Gastos_{op_id}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ==============================================================
    # HELPER interno: renderiza el Kanban completo (outerHTML)
    # ==============================================================

    async def _render_kanban(request, conn, service, context, notification: Optional[dict] = None):
        """
        Recarga datos del kanban y retorna el template completo.
        Usado por posponer y reagendar para refrescar el tablero.
        """
        data = await service.get_kanban_data(conn)

        can_edit = (
            context.get("role") == "ADMIN"
            or context.get("module_roles", {}).get("levantamientos") in ["editor", "admin"]
        )

        return templates.TemplateResponse("levantamientos/partials/kanban.html", {
            "request": request,
            "pendientes": data["pendientes"],
            "agendados": data["agendados"],
            "en_proceso": data["en_proceso"],
            "completados": data["completados"],
            "entregados": data["entregados"],
            "pospuestos": data["pospuestos"],
            "can_edit": can_edit,
            "user_context": context,
            "notification": notification # OOB Toast support
        })
