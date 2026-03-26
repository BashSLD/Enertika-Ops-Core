# ==============================================================
# modules/levantamientos/router_operaciones.py
# Endpoints POST/DELETE de operaciones para el módulo Levantamientos.
# Incluye: posponer, reagendar, viaticos CRUD, solicitud viáticos,
# entregar y reporte Excel.
# Registrado en router_levantamientos_nuevos.py.
# ==============================================================

import asyncio
import json
import logging
import re
import time
from io import BytesIO
from typing import List, Optional
from uuid import UUID, uuid4
from datetime import datetime, date as date_type, time as time_type
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates

from core.security import get_current_user_context
from core.permissions import require_module_access, require_manager_access
from core.database import get_db_connection
from core.microsoft import MicrosoftAuth

from .service import get_service, LevantamientoService
from .db_service import get_db_service, LevantamientosDBService
from .db_service_visitas import (
    get_visitas_db_service,
    VisitasCampoDBService,
    calcular_prorrateo,
)

logger = logging.getLogger("Levantamientos.Router.Operaciones")

templates = Jinja2Templates(directory="templates")


_MSG_PRORRATEO_CAMBIO = "El prorrateo ha cambiado. Considera reenviar la solicitud."

# ==============================================================
# HELPER: renderiza el Kanban completo (outerHTML)
# Usado por posponer, reagendar y entregar.
# ==============================================================

async def _render_kanban(request, conn, service, context, notification: Optional[dict] = None):
    """
    Recarga datos del kanban y retorna el template completo.
    """
    data = await service.get_kanban_data(conn)

    can_edit = (
        context.get("role") == "ADMIN"
        or context.get("module_roles", {}).get("levantamientos") in ["editor", "admin"]
    )

    return templates.TemplateResponse(request, "levantamientos/partials/kanban.html", {"pendientes": data["pendientes"],
        "agendados": data["agendados"],
        "en_proceso": data["en_proceso"],
        "completados": data["completados"],
        "entregados": data["entregados"],
        "pospuestos": data["pospuestos"],
        "can_edit": can_edit,
        "user_context": context,
        "notification": notification
    })


def register_operaciones_endpoints(router: APIRouter):
    """
    Registra los endpoints de operaciones en el router existente.
    """

    # ==============================================================
    # POST — POSPONER
    # ==============================================================

    @router.post("/posponer/{id_levantamiento}")
    async def posponer_endpoint(
        request: Request,
        id_levantamiento: UUID,
        motivo_pospone: str = Form(...),
        devolver_viaticos: Optional[bool] = Form(False),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Valida motivo (min 10 chars).
        2. Obtiene estado actual para el historial.
        3. Ejecuta UPDATE via db_service.
        4. Registra en historial via service.
        5. Retorna Kanban actualizado (outerHTML).
        """
        if not motivo_pospone or len(motivo_pospone.strip()) < 10:
            raise HTTPException(status_code=400, detail="El motivo debe tener al menos 10 caracteres.")

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        estatus_map = await db_svc.get_estatus_map(conn)
        id_pospuesto = estatus_map['pospuesto']
        estado_anterior = lev["id_estatus_global"]

        await db_svc.update_posponer(conn, id_levantamiento, motivo_pospone.strip(), context["user_db_id"], id_pospuesto)

        await service._registrar_en_historial(
            conn=conn,
            id_levantamiento=id_levantamiento,
            estatus_anterior=estado_anterior,
            estatus_nuevo=id_pospuesto,
            user_context=context,
            observaciones=motivo_pospone.strip(),
            metadata={"tipo_cambio": "posponer"}
        )

        if devolver_viaticos:
            await service.registrar_devolucion(conn, id_levantamiento, context)

        asyncio.create_task(
            service._execute_notification_background(
                service._notificar_pospuesto_impl,
                id_oportunidad=lev["id_oportunidad"],
                motivo=motivo_pospone.strip(),
                user_context=context
            )
        )

        return await _render_kanban(request, conn, service, context)

    # ==============================================================
    # POST — REAGENDAR
    # ==============================================================

    @router.post("/reagendar/{id_levantamiento}")
    async def reagendar_endpoint(
        request: Request,
        id_levantamiento: UUID,
        nueva_fecha_visita: str = Form(...),
        observaciones: Optional[str] = Form(None),
        asumir_responsable: bool = Form(False),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Valida formato de fecha y que no sea pasada.
        2. Obtiene estado actual.
        3. Ejecuta UPDATE via db_service.
        4. Registra en historial.
        5. Retorna Kanban actualizado.
        """
        if not nueva_fecha_visita:
            raise HTTPException(status_code=400, detail="Se requiere fecha de visita.")

        try:
            fecha_obj = datetime.fromisoformat(nueva_fecha_visita).replace(tzinfo=ZoneInfo("America/Mexico_City"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha/hora inválido.")

        now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
        if fecha_obj.date() < now_mx.date():
            raise HTTPException(status_code=400, detail="La fecha no puede ser anterior a hoy.")

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        estatus_map_r = await db_svc.get_estatus_map(conn)
        id_agendado = estatus_map_r['agendado']
        id_pendiente = estatus_map_r['pendiente']
        estado_anterior = lev["id_estatus_global"]

        is_rescheduling = (estado_anterior != id_pendiente)
        await db_svc.update_reagendar(conn, id_levantamiento, fecha_obj, context["user_db_id"], id_agendado, is_rescheduling=is_rescheduling)

        fecha_display = fecha_obj.strftime("%d/%m/%Y %H:%M")
        obs_text = observaciones or f"Visita reagendada para {fecha_display}"
        await service._registrar_en_historial(
            conn=conn,
            id_levantamiento=id_levantamiento,
            estatus_anterior=estado_anterior,
            estatus_nuevo=id_agendado,
            user_context=context,
            observaciones=obs_text,
            metadata={"tipo_cambio": "reagendar", "nueva_fecha": fecha_obj.isoformat()}
        )

        asyncio.create_task(
            service._execute_notification_background(
                service._notificar_agendado_impl,
                id_oportunidad=lev["id_oportunidad"],
                fecha_visita=fecha_display,
                user_context=context
            )
        )

        # --- Auto-asignación como responsable (solo si el usuario no es jefe) ---
        user_db_id = context.get("user_db_id")
        jefe_area_id = lev.get("jefe_area_id")
        is_jefe = (jefe_area_id is not None and str(jefe_area_id) == str(user_db_id))
        auto_asignado = False
        notif_asignacion = None

        if not is_jefe:
            try:
                responsable_actual = await db_svc.get_responsable_asignado(conn, id_levantamiento)
                if asumir_responsable:
                    # Solo si el usuario confirmó explícitamente asumir la responsabilidad
                    await db_svc.update_responsable(conn, id_levantamiento, user_db_id, user_db_id)
                    auto_asignado = True
                    notif_asignacion = "Ahora eres el ingeniero responsable de este levantamiento."
                    logger.info(f"[REAGENDAR] Cambio de responsable confirmado: {user_db_id} en lev {id_levantamiento}")
            except Exception as e_resp:
                logger.error(f"[REAGENDAR] Error en auto-asignación responsable para lev {id_levantamiento}: {e_resp}")

        has_techs_new = await conn.fetchval("""
            SELECT EXISTS(SELECT 1 FROM tb_levantamiento_asignaciones WHERE id_levantamiento = $1)
        """, id_levantamiento)

        has_techs_legacy = await conn.fetchval("""
            SELECT (tecnico_asignado_id IS NOT NULL) FROM tb_levantamientos WHERE id_levantamiento = $1
        """, id_levantamiento)

        has_techs = has_techs_new or has_techs_legacy

        logger.info(f"Reagendar Validation - ID: {id_levantamiento}, HasTechsNew: {has_techs_new}, HasTechsLegacy: {has_techs_legacy}, Final: {has_techs}")

        action_label = "Reagendado" if is_rescheduling else "Agendado"
        if auto_asignado:
            notification = {
                "title": action_label,
                "message": f"Levantamiento {action_label.lower()} para {fecha_display}. {notif_asignacion}",
                "type": "success"
            }
        elif not has_techs:
            logger.info("Triggering Warning Toast: No technicians assigned.")
            notification = {
                "title": "Asignación Pendiente",
                "message": "El levantamiento ha sido agendado. Recuerda asignar un ingeniero.",
                "type": "warning"
            }
        else:
            notification = None

        return await _render_kanban(request, conn, service, context, notification)

    # ==============================================================
    # POST / DELETE — VIATICOS CRUD
    # ==============================================================

    @router.post("/viaticos/{id_levantamiento}")
    async def crear_viatico_endpoint(
        request: Request,
        id_levantamiento: UUID,
        usuario_id: UUID = Form(...),
        concepto: str = Form(...),
        monto: float = Form(...),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Crea un viatico y retorna el innerHTML de #tabla-viaticos-container
        con la tabla actualizada + total.
        """
        if not concepto or not concepto.strip():
            raise HTTPException(status_code=400, detail="El concepto es obligatorio.")
        if monto <= 0:
            raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0.")

        await db_svc.create_viatico(
            conn, id_levantamiento, usuario_id, concepto.strip(), monto, context["user_db_id"]
        )

        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)

        return templates.TemplateResponse(request, "levantamientos/partials/tabla_viaticos.html", {"viaticos": viaticos,
            "id_levantamiento": id_levantamiento,
        })

    # ----------------------------------------------------------

    @router.delete("/viaticos/{id_levantamiento}/{viatico_id}")
    async def eliminar_viatico_endpoint(
        request: Request,
        id_levantamiento: UUID,
        viatico_id: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Elimina un viatico y retorna el innerHTML actualizado
        de #tabla-viaticos-container.
        """
        eliminado = await db_svc.delete_viatico(conn, id_levantamiento, viatico_id)
        if not eliminado:
            raise HTTPException(status_code=404, detail="Viatico no encontrado.")

        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)

        return templates.TemplateResponse(request, "levantamientos/partials/tabla_viaticos.html", {"viaticos": viaticos,
            "id_levantamiento": id_levantamiento,
        })

    # ==============================================================
    # POST — ENVIAR SOLICITUD DE VIATICOS
    # ==============================================================

    @router.post("/viaticos/solicitud/{id_levantamiento}")
    async def enviar_solicitud_viaticos_endpoint(
        request: Request,
        id_levantamiento: UUID,
        to_destinatarios: str = Form(""),
        cc_adicionales: str = Form(""),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        1. Obtiene viaticos actuales (debe haber al menos 1).
        2. Construye TO (del form, con fallback al configurado) y CC.
        3. Renderiza solicitud_viaticos.html como body del correo.
        4. Envía via MicrosoftAuth.
        5. Registra en tb_levantamiento_viaticos_historico con snapshot.
        6. Retorna innerHTML de #historial-envios-container actualizado.
        """
        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)
        if not viaticos:
            raise HTTPException(status_code=400, detail="No hay viaticos registrados.")

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        total_monto = sum(v["monto"] for v in viaticos)

        # TO: usar lo que envió el usuario; si llegó vacío, caer a la configuración de admin
        to_configurado = await db_svc.get_to_configurados_viaticos(conn)
        to_del_form = [e.strip() for e in re.split(r'[;,]', to_destinatarios) if e.strip() and "@" in e.strip()]
        to_list = to_del_form if to_del_form else to_configurado

        cc_configurados = await db_svc.get_cc_configurados_viaticos(conn)
        cc_manuales = [e.strip() for e in re.split(r'[;,]', cc_adicionales) if e.strip() and "@" in e.strip()]
        cc_all = list(set(cc_configurados + cc_manuales) - set(to_list))

        if not to_list:
            raise HTTPException(status_code=500, detail="No hay destinatarios TO configurados para SOLICITUD_VIATICOS.")

        now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
        fecha_envio_str = now_mx.strftime("%d/%m/%Y %H:%M")

        email_template = templates.get_template("levantamientos/emails/solicitud_viaticos.html")
        html_body = email_template.render(
            proyecto_nombre=lev["nombre_proyecto"] or lev["titulo_proyecto"] or "Sin nombre",
            op_id=lev["op_id_estandar"],
            cliente_nombre=lev["cliente_nombre"],
            sitio_direccion=lev.get("sitio_direccion"),
            enviado_por=context.get("user_name", "Sistema"),
            fecha_envio=fecha_envio_str,
            viaticos=viaticos,
            total_monto=total_monto,
        )

        subject = f"Solicitud de Viaticos — {lev['op_id_estandar']} | {lev['cliente_nombre']}"

        sender_config = await conn.fetchrow("""
            SELECT email_remitente FROM tb_correos_notificaciones
            WHERE departamento = 'LEVANTAMIENTOS' AND activo = true
            LIMIT 1
        """)
        if not sender_config:
            sender_config = await conn.fetchrow("""
                SELECT email_remitente FROM tb_correos_notificaciones
                WHERE departamento = 'DEFAULT' AND activo = true
                LIMIT 1
            """)
        sender_email = sender_config['email_remitente'] if sender_config else 'app-notifications@enertika.mx'

        estatus_envio = "enviado"
        error_detalle = None

        try:
            ms_auth = MicrosoftAuth()
            app_token = await ms_auth.get_application_token()

            if not app_token:
                raise Exception("No se pudo obtener token de aplicación de Microsoft Graph.")

            success, msg = await ms_auth.send_email_with_attachments(
                access_token=app_token,
                from_email=sender_email,
                subject=subject,
                body=html_body,
                recipients=to_list,
                cc_recipients=cc_all if cc_all else None,
                importance="normal",
            )

            if not success:
                estatus_envio = "error"
                error_detalle = msg
                logger.error(f"[VIATICOS] Error envío correo lev {id_levantamiento}: {msg}")
            else:
                logger.info(f"[VIATICOS] Correo enviado exitosamente lev {id_levantamiento}")

        except Exception as exc:
            estatus_envio = "error"
            error_detalle = str(exc)
            logger.error(f"[VIATICOS] Excepción envío correo lev {id_levantamiento}: {exc}")

        snapshot = [
            {"usuario_nombre": v["usuario_nombre"], "concepto": v["concepto"], "monto": float(v["monto"])}
            for v in viaticos
        ]

        await db_svc.insert_historial_envio(
            conn=conn,
            id_levantamiento=id_levantamiento,
            enviado_por_id=context["user_db_id"],
            enviado_por_nombre=context.get("user_name", "Sistema"),
            to_destinatarios=to_list,
            cc_destinatarios=cc_all,
            viaticos_snapshot=snapshot,
            total_monto=total_monto,
            estatus=estatus_envio,
            error_detalle=error_detalle,
        )

        historial = await db_svc.get_historial_envios(conn, id_levantamiento)

        return templates.TemplateResponse(request, "levantamientos/partials/historial_envios.html", {"historial_envios": historial,
        })

    # ==============================================================
    # POST — ENTREGAR LEVANTAMIENTO (con archivos opcionales)
    # ==============================================================

    @router.post("/entregar/{id_levantamiento}", include_in_schema=False)
    async def entregar_levantamiento(
        request: Request,
        id_levantamiento: UUID,
        observaciones: Optional[str] = Form(None),
        file_uploads: List[UploadFile] = File(None),
        conn=Depends(get_db_connection),
        service: LevantamientoService = Depends(get_service),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Marca un levantamiento como entregado (estado 12).
        Opcionalmente sube archivos a SharePoint y los registra en BD.
        """
        notification = None
        try:
            _estatus_map_e = await db_svc.get_estatus_map(conn)
            id_entregado = _estatus_map_e['entregado']
            await service.cambiar_estado(
                conn, id_levantamiento, id_entregado, context,
                observaciones=observaciones.strip() if observaciones else None
            )

            files_uploaded = 0
            if file_uploads and any(f.filename for f in file_uploads):
                try:
                    lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
                    op_id_estandar = lev["op_id_estandar"] if lev else None

                    if not op_id_estandar:
                        logger.warning(f"[ENTREGA] No se pudo subir adjuntos: op_id_estandar es NULL para {id_levantamiento}")
                    else:
                        config_rows = await conn.fetch("""
                            SELECT clave, valor FROM tb_configuracion_global
                            WHERE clave IN ('MAX_UPLOAD_SIZE_MB', 'SHAREPOINT_BASE_FOLDER')
                        """)
                        config_map = {row['clave']: row['valor'] for row in config_rows}
                        max_size_mb = int(config_map.get('MAX_UPLOAD_SIZE_MB', '10'))
                        base_folder = config_map.get('SHAREPOINT_BASE_FOLDER', '').strip().strip("/")

                        relative_path = f"levantamiento/{op_id_estandar}/entrega"
                        folder_path = f"{base_folder}/{relative_path}" if base_folder else relative_path

                        ms_auth = MicrosoftAuth()
                        app_token = await ms_auth.get_application_token()
                        if not app_token:
                            logger.error(f"[ENTREGA] No se pudo obtener token de aplicacion para SharePoint")
                        else:
                            from core.integrations.sharepoint import SharePointService
                            sharepoint = SharePointService(access_token=app_token)

                            for f_obj in file_uploads:
                                if not f_obj.filename:
                                    continue
                                try:
                                    f_obj.file.seek(0, 2)
                                    f_size = f_obj.file.tell()
                                    f_obj.file.seek(0)

                                    if f_size / (1024 * 1024) > max_size_mb:
                                        logger.warning(f"[ENTREGA] Archivo {f_obj.filename} excede limite: {f_size} bytes")
                                        continue

                                    timestamp = int(time.time())
                                    original_name = f_obj.filename
                                    f_obj.filename = f"{timestamp}_{original_name}"

                                    logger.info(f"[ENTREGA] Subiendo archivo: {f_obj.filename} a {folder_path}")

                                    upload_result = await sharepoint.upload_file(conn, f_obj, folder_path)

                                    doc_id = uuid4()
                                    parent_ref = upload_result.get('parentReference', {})

                                    await conn.execute("""
                                        INSERT INTO tb_documentos_attachments (
                                            id_documento, nombre_archivo, url_sharepoint, drive_item_id, parent_drive_id,
                                            tipo_contenido, tamano_bytes, id_oportunidad, subido_por_id,
                                            origen_slug, activo, metadata
                                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, 'lev_entregados', TRUE, $10::jsonb)
                                    """,
                                        doc_id,
                                        upload_result['name'],
                                        upload_result['webUrl'],
                                        upload_result['id'],
                                        parent_ref.get('driveId'),
                                        f_obj.content_type,
                                        upload_result['size'],
                                        lev["id_oportunidad"],
                                        context["user_db_id"],
                                        json.dumps({
                                            "id_levantamiento": str(id_levantamiento),
                                            "tipo": "entrega",
                                            "nombre_original": original_name
                                        })
                                    )

                                    files_uploaded += 1
                                    logger.info(f"[ENTREGA] Adjunto registrado: {upload_result['name']}")

                                except Exception as e_file:
                                    logger.error(f"[ENTREGA] Error subiendo archivo {f_obj.filename}: {e_file}")

                except Exception as e_sp:
                    logger.error(f"[ENTREGA] Fallo general en adjuntos: {e_sp}")

            msg = "Levantamiento entregado exitosamente."
            if files_uploaded > 0:
                msg += f" {files_uploaded} archivo(s) adjuntado(s)."
            notification = {"title": "Entregado", "message": msg, "type": "success"}

        except HTTPException:
            raise
        except ValueError as ve:
            notification = {"title": "Error", "message": str(ve), "type": "error"}
        except Exception as exc:
            logger.error(f"[ENTREGA] Error inesperado: {exc}", exc_info=True)
            notification = {"title": "Error", "message": "Ocurrio un error al procesar la entrega.", "type": "error"}

        return await _render_kanban(request, conn, service, context, notification)

    # ==============================================================
    # POST — CANCELAR LEVANTAMIENTO
    # ==============================================================

    @router.post("/cancelar/{id_levantamiento}")
    async def cancelar_levantamiento_endpoint(
        request: Request,
        id_levantamiento: UUID,
        motivo: str = Form(...),
        conn=Depends(get_db_connection),
        service: LevantamientoService = Depends(get_service),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_manager_access("levantamientos", "editor"),
    ):
        """
        Cancela un levantamiento. NO afecta la oportunidad comercial.
        Limpia viáticos, registra historial y notifica a jefe + solicitante.
        Solo ADMIN, Admin del módulo o MANAGER+editor pueden cancelar.
        """
        user_role = context.get("role")
        mod_role = context.get("module_roles", {}).get("levantamientos")
        can_cancel = (
            user_role == "ADMIN" or
            mod_role == "admin" or
            (user_role == "MANAGER" and mod_role in ["editor", "admin"])
        )
        if not can_cancel:
            raise HTTPException(status_code=403, detail="No tienes permisos para cancelar levantamientos")

        if not motivo or len(motivo.strip()) < 10:
            raise HTTPException(status_code=400, detail="El motivo debe tener al menos 10 caracteres")

        await service.cancelar_levantamiento(
            conn=conn,
            id_levantamiento=id_levantamiento,
            motivo=motivo.strip(),
            user_context=context,
        )

        estatus_map = await db_svc.get_estatus_map(conn)
        id_cancelado = estatus_map.get('cancelado')
        ids_activos = [v for k, v in estatus_map.items() if k not in ('completado', 'entregado', 'cancelado')]
        levantamientos = await db_svc.get_lista_activos(conn, ids_activos=ids_activos)
        estatus_list = await db_svc.get_estatus_list(conn)
        tecnicos = await db_svc.get_usuarios_tecnicos(conn)
        can_edit = (
            user_role == "ADMIN"
            or mod_role in ["editor", "admin"]
        )
        return templates.TemplateResponse(request, "levantamientos/partials/lista.html", {"tab": "activos",
            "levantamientos": levantamientos,
            "tecnicos": tecnicos,
            "estatus_filtro": [e for e in estatus_list if e['grupo_kanban'] == 'activo'],
            "can_edit": can_edit,
            "can_manage": can_cancel,
            "filtros": {"q": "", "estado": None, "tecnico_id": "", "fecha_inicio": "", "fecha_fin": ""},
            "notification": {
                "title": "Levantamiento Cancelado",
                "message": "El levantamiento ha sido cancelado.",
                "type": "success",
            },
        })

    # ==============================================================
    # POST — REACTIVAR LEVANTAMIENTO
    # ==============================================================

    @router.post("/reactivar/{id_levantamiento}")
    async def reactivar_levantamiento_endpoint(
        request: Request,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        service: LevantamientoService = Depends(get_service),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Reactiva un levantamiento cancelado, volviéndolo a estado Pendiente.
        La oportunidad permanece cancelada; el equipo comercial decide qué hacer con ella.
        """
        user_role = context.get("role")
        mod_role = context.get("module_roles", {}).get("levantamientos")
        can_manage = (
            user_role == "ADMIN" or
            mod_role == "admin" or
            (user_role == "MANAGER" and mod_role in ["editor", "admin"])
        )
        if not can_manage:
            raise HTTPException(status_code=403, detail="No tienes permisos para reactivar levantamientos")

        await service.reactivar_levantamiento(
            conn=conn,
            id_levantamiento=id_levantamiento,
            user_context=context,
        )

        estatus_map = await db_svc.get_estatus_map(conn)
        id_cancelado = estatus_map.get('cancelado')
        cancelados = await db_svc.get_lista_cancelados(conn, id_cancelado=id_cancelado)
        estatus_list = await db_svc.get_estatus_list(conn)
        tecnicos = await db_svc.get_usuarios_tecnicos(conn)
        can_edit = (user_role == "ADMIN" or mod_role in ["editor", "admin"])
        return templates.TemplateResponse(request, "levantamientos/partials/lista.html", {"tab": "cancelados",
            "levantamientos": cancelados,
            "tecnicos": tecnicos,
            "estatus_filtro": [],
            "can_edit": can_edit,
            "can_manage": can_manage,
            "filtros": {"q": "", "estado": None, "tecnico_id": "", "fecha_inicio": "", "fecha_fin": ""},
            "notification": {
                "title": "Levantamiento Reactivado",
                "message": "El levantamiento volvio a estado Pendiente. Recuerda actualizar la oportunidad desde Comercial.",
                "type": "success",
            },
        })

    # ==============================================================
    # POST — SOLICITAR REASIGNACION
    # ==============================================================

    @router.post("/solicitar-reasignacion/{id_levantamiento}")
    async def solicitar_reasignacion_endpoint(
        request: Request,
        id_levantamiento: UUID,
        motivo: str = Form(...),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos"),
    ):
        """
        Solicita la reasignación de un levantamiento.
        Solo puede ser ejecutado por el ingeniero responsable actual.
        Envía notificación a quien asignó + quien solicitó el levantamiento.
        No modifica el estado del levantamiento.
        """
        if not motivo or len(motivo.strip()) < 10:
            raise HTTPException(status_code=400, detail="El motivo debe tener al menos 10 caracteres")

        responsable = await db_svc.get_responsable_asignado(conn, id_levantamiento)
        user_db_id = context.get("user_db_id")
        if not responsable or str(responsable['id_usuario']) != str(user_db_id):
            raise HTTPException(status_code=403, detail="Solo el ingeniero responsable puede solicitar reasignacion")

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        asyncio.create_task(
            service._execute_notification_background(
                service._notificar_solicitud_reasignacion_impl,
                id_levantamiento=id_levantamiento,
                id_oportunidad=lev['id_oportunidad'],
                motivo=motivo.strip(),
                user_context=context,
            )
        )

        return templates.TemplateResponse(request, "shared/toast.html", {"title": "Solicitud Enviada",
            "message": "El equipo responsable sera notificado para gestionar la reasignacion.",
            "type": "success",
        })

    # ==============================================================
    # GET — REPORTE EXCEL DE GASTOS (VIATICOS)
    # ==============================================================

    @router.get("/reporte-gastos/{id_levantamiento}", include_in_schema=False)
    async def reporte_gastos_excel(
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos"),
    ):
        """
        Genera y descarga un reporte Excel con los gastos (viaticos)
        y el historial de envios de un levantamiento.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        viaticos = await db_svc.get_viaticos(conn, id_levantamiento)
        # Si no hay viáticos individuales, buscar en visita de campo (prorrateo)
        if not viaticos:
            visitas_vc = await db_svc.get_visitas_campo_for_lev(conn, id_levantamiento)
            for vc in visitas_vc:
                if vc.get("monto_prorrateo", 0) > 0:
                    viaticos = [{
                        "usuario_nombre": "Prorrateo — " + (vc.get("nombre") or "Visita de Campo"),
                        "concepto": f"Prorrateo visita de campo ({vc['num_levantamientos']} sitios)",
                        "monto": vc["monto_prorrateo"],
                    }]
                    break
        historial = await db_svc.get_historial_envios(conn, id_levantamiento)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        title_font = Font(bold=True, size=14, color="1F4E79")
        subtitle_font = Font(bold=True, size=11, color="333333")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        money_format = '#,##0.00'

        wb = Workbook()

        # ===================== HOJA 1: RESUMEN =====================
        ws1 = wb.active
        ws1.title = "Resumen"

        ws1.merge_cells('A1:D1')
        cell_title = ws1.cell(row=1, column=1, value="Reporte de Gastos - Levantamiento")
        cell_title.font = title_font
        cell_title.alignment = Alignment(horizontal="center")

        info_data = [
            ("OP-ID:", lev.get("op_id_estandar", "")),
            ("Cliente:", lev.get("cliente_nombre", "")),
            ("Proyecto:", lev.get("nombre_proyecto") or lev.get("titulo_proyecto") or "Sin nombre"),
            ("Direccion:", lev.get("direccion") or lev.get("sitio_direccion") or "Sin direccion"),
        ]
        for i, (label, value) in enumerate(info_data, start=3):
            ws1.cell(row=i, column=1, value=label).font = subtitle_font
            ws1.cell(row=i, column=2, value=value)

        viaticos_start_row = len(info_data) + 5
        ws1.cell(row=viaticos_start_row - 1, column=1, value="Detalle de Viaticos").font = subtitle_font

        viat_headers = ["Usuario", "Concepto", "Monto"]
        for col_num, header in enumerate(viat_headers, 1):
            cell = ws1.cell(row=viaticos_start_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        total_monto = 0
        for row_num, v in enumerate(viaticos, start=viaticos_start_row + 1):
            ws1.cell(row=row_num, column=1, value=v.get("usuario_nombre", "")).border = thin_border
            ws1.cell(row=row_num, column=2, value=v.get("concepto", "")).border = thin_border
            monto_cell = ws1.cell(row=row_num, column=3, value=float(v.get("monto", 0)))
            monto_cell.number_format = money_format
            monto_cell.border = thin_border
            total_monto += float(v.get("monto", 0))

        total_row = viaticos_start_row + len(viaticos) + 1
        ws1.cell(row=total_row, column=2, value="TOTAL:").font = Font(bold=True, size=11)
        total_cell = ws1.cell(row=total_row, column=3, value=total_monto)
        total_cell.font = Font(bold=True, size=11)
        total_cell.number_format = money_format

        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 40
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 40

        # ===================== HOJA 2: HISTORIAL ENVIOS =====================
        ws2 = wb.create_sheet("Historial Envios")

        ws2.merge_cells('A1:F1')
        cell_title2 = ws2.cell(row=1, column=1, value="Historial de Solicitudes de Viaticos")
        cell_title2.font = title_font
        cell_title2.alignment = Alignment(horizontal="center")

        hist_headers = ["Fecha Envio", "Enviado Por", "Destinatarios TO", "Destinatarios CC", "Total", "Estatus"]
        for col_num, header in enumerate(hist_headers, 1):
            cell = ws2.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_num, h in enumerate(historial, start=4):
            fecha = h.get("fecha_envio")
            fecha_str = fecha.strftime("%d/%m/%Y %H:%M") if fecha else ""

            to_list = h.get("to_destinatarios") or []
            cc_list = h.get("cc_destinatarios") or []
            to_str = ", ".join(to_list) if isinstance(to_list, list) else str(to_list)
            cc_str = ", ".join(cc_list) if isinstance(cc_list, list) else str(cc_list)

            ws2.cell(row=row_num, column=1, value=fecha_str).border = thin_border
            ws2.cell(row=row_num, column=2, value=h.get("enviado_por_nombre", "")).border = thin_border
            ws2.cell(row=row_num, column=3, value=to_str).border = thin_border
            ws2.cell(row=row_num, column=4, value=cc_str).border = thin_border
            monto_h = ws2.cell(row=row_num, column=5, value=float(h.get("total_monto", 0)))
            monto_h.number_format = money_format
            monto_h.border = thin_border
            ws2.cell(row=row_num, column=6, value=h.get("estatus", "")).border = thin_border

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 25
        ws2.column_dimensions['C'].width = 35
        ws2.column_dimensions['D'].width = 35
        ws2.column_dimensions['E'].width = 18
        ws2.column_dimensions['F'].width = 15

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        op_id = lev.get("op_id_estandar", "SIN_ID").replace("/", "-")
        filename = f"Reporte_Gastos_{op_id}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    # ==============================================================
    # VISITAS DE CAMPO — Endpoints POST/DELETE
    # ==============================================================

    @router.post("/visitas-campo", include_in_schema=False)
    async def crear_visita_campo(
        request: Request,
        nombre: Optional[str] = Form(None),
        fecha_inicio: str = Form(...),
        fecha_fin: str = Form(...),
        levantamiento_ids: List[UUID] = Form(default=[]),
        ingeniero_id: Optional[UUID] = Form(None),
        acompaniante_id: Optional[UUID] = Form(None),
        notas: Optional[str] = Form(None),
        force_replace: bool = Form(False),
        keep_existing: bool = Form(False),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Crea una nueva Visita de Campo con sus levantamientos vinculados.
        Soporta fechas individuales por levantamiento y viáticos iniciales.
        Retorna HX-Location a la página de detalle tras crear la visita.
        """
        def _err(msg: str) -> HTMLResponse:
            return HTMLResponse(
                content=(
                    '<div x-data="{show:true}" x-show="show" x-transition'
                    ' x-init="setTimeout(()=>show=false,6000)"'
                    ' class="pointer-events-auto max-w-sm w-full bg-slate-800 border border-red-500/60'
                    ' rounded-xl shadow-lg p-4 flex items-start gap-3">'
                    '<div class="flex-shrink-0 w-8 h-8 rounded-full bg-red-500/20 flex items-center justify-center">'
                    '<i class="fas fa-exclamation-circle text-red-400 text-sm"></i></div>'
                    '<div class="flex-1 min-w-0">'
                    '<p class="text-sm font-semibold text-red-400">Error</p>'
                    f'<p class="text-xs text-slate-300 mt-0.5">{msg}</p>'
                    '</div>'
                    '<button @click="show=false" class="text-slate-500 hover:text-white transition-colors flex-shrink-0">'
                    '<i class="fas fa-times text-xs"></i></button>'
                    '</div>'
                ),
                status_code=200,
            )

        if not levantamiento_ids:
            return _err("Debes seleccionar al menos un levantamiento.")

        try:
            fecha_inicio_dt = datetime.fromisoformat(fecha_inicio).replace(tzinfo=ZoneInfo("America/Mexico_City"))
            fecha_fin_dt = datetime.fromisoformat(fecha_fin).replace(tzinfo=ZoneInfo("America/Mexico_City"))
        except ValueError:
            return _err("Formato de fecha/hora inválido.")

        if fecha_fin_dt <= fecha_inicio_dt:
            return _err("La fecha de fin debe ser posterior a la de inicio.")

        # Extraer fechas individuales por levantamiento del form
        form_data = await request.form()
        fechas_individuales = {}
        for lev_id in levantamiento_ids:
            fecha_key = f"fecha_lev_{lev_id}"
            fecha_val = form_data.get(fecha_key, "")
            if fecha_val:
                try:
                    dt = datetime.fromisoformat(str(fecha_val)).replace(tzinfo=ZoneInfo("America/Mexico_City"))
                    fechas_individuales[str(lev_id)] = dt
                except ValueError:
                    pass  # Skip invalid dates

        # Validar que las fechas individuales estén dentro del período de la visita
        fuera = [lev_id_str for lev_id_str, fecha_ind in fechas_individuales.items()
                 if not (fecha_inicio_dt <= fecha_ind <= fecha_fin_dt)]
        if fuera:
            op_ids_map = await db_svc.get_op_ids_by_ids(conn, levantamiento_ids)
            nombres = ", ".join(op_ids_map.get(lid, lid) for lid in fuera)
            return _err(f"La fecha de {nombres} está fuera del período de la visita.")

        con_viaticos = await db_svc.get_levantamientos_con_viaticos_activos(conn, levantamiento_ids)
        if con_viaticos:
            nombres = ", ".join(
                f"{r['op_id_estandar']} - {r['nombre_referencia']}" for r in con_viaticos
            )
            return _err(
                f"Los siguientes levantamientos ya tienen viáticos solicitados: {nombres}. "
                "Devuelve los viáticos antes de agendar la visita de campo."
            )

        # Verificar conflicto de ingeniero responsable previo
        if ingeniero_id and not force_replace and not keep_existing:
            n_conflictos = await visitas_db_svc.check_levantamientos_con_responsable(
                conn, levantamiento_ids
            )
            if n_conflictos > 0:
                sitios_txt = f"{n_conflictos} sitio{'s' if n_conflictos > 1 else ''}"
                return HTMLResponse(
                    content=(
                        '<div x-data="{show:true}" x-show="show" x-transition'
                        ' class="pointer-events-auto max-w-sm w-full bg-slate-800 border border-amber-500/60'
                        ' rounded-xl shadow-lg p-4">'
                        '<p class="text-sm font-semibold text-amber-400 mb-1">'
                        f'<i class="fas fa-exclamation-triangle mr-1"></i>{sitios_txt} ya tienen ingeniero asignado</p>'
                        '<p class="text-xs text-slate-300 mb-3">¿Deseas reemplazar el ingeniero en todos los sitios, o conservar los que ya están asignados?</p>'
                        '<div class="flex gap-2">'
                        '<button @click="show=false;'
                        'let f=document.getElementById(\'form-crear-visita\');'
                        'if(!f.querySelector(\'[name=keep_existing]\')){let i=document.createElement(\'input\');i.type=\'hidden\';i.name=\'keep_existing\';f.appendChild(i);}'
                        'f.querySelector(\'[name=keep_existing]\').value=\'true\';'
                        'f.dispatchEvent(new Event(\'submit\',{bubbles:true}))" '
                        'class="flex-1 text-xs py-1.5 rounded-lg bg-slate-700 text-slate-300 hover:bg-slate-600">No, conservar</button>'
                        '<button @click="show=false;'
                        'document.getElementById(\'form-crear-visita\').querySelector(\'[name=force_replace]\').value=\'true\';'
                        'document.getElementById(\'form-crear-visita\').dispatchEvent(new Event(\'submit\',{bubbles:true}))" '
                        'class="flex-1 text-xs py-1.5 rounded-lg bg-amber-600 hover:bg-amber-500 text-white">Sí, reemplazar</button>'
                        '</div>'
                        '</div>'
                    ),
                    status_code=200,
                )

        # Extraer viáticos iniciales del form (arrays paralelos) antes de la transacción
        viatico_conceptos = form_data.getlist("viatico_concepto")
        viatico_montos = form_data.getlist("viatico_monto")
        viatico_usuario_ids = form_data.getlist("viatico_usuario_id")

        # Parsear viáticos válidos antes de abrir la transacción
        viaticos_a_crear = []
        for i, concepto in enumerate(viatico_conceptos):
            if not concepto or not concepto.strip():
                continue
            try:
                monto = float(viatico_montos[i]) if i < len(viatico_montos) else 0.0
            except (ValueError, IndexError):
                continue
            if monto <= 0:
                continue
            usuario_id = None
            if i < len(viatico_usuario_ids) and viatico_usuario_ids[i]:
                try:
                    usuario_id = UUID(viatico_usuario_ids[i])
                except ValueError:
                    pass
            viaticos_a_crear.append((concepto.strip(), monto, usuario_id))

        async with conn.transaction():
            visita = await visitas_db_svc.create_visita(
                conn,
                nombre=nombre.strip() if nombre else None,
                fecha_inicio=fecha_inicio_dt,
                fecha_fin=fecha_fin_dt,
                levantamiento_ids=levantamiento_ids,
                creado_por_id=context["user_db_id"],
                fechas_individuales=fechas_individuales if fechas_individuales else None,
                notas=notas.strip() if notas and notas.strip() else None,
            )

            id_visita = visita["id_visita"]

            # Sincronizar levantamientos como "Agendado" si tienen fecha individual
            if fechas_individuales:
                await visitas_db_svc.sync_levantamientos_agendado(
                    conn, levantamiento_ids, fechas_individuales, context["user_db_id"]
                )

            # Propagar ingeniero responsable (seleccionado o creador por defecto)
            ingeniero_efectivo = ingeniero_id or context["user_db_id"]
            await visitas_db_svc.propagar_ingeniero_visita(
                conn, id_visita, ingeniero_efectivo, context["user_db_id"], keep_existing=keep_existing
            )

            # Propagar acompañante si fue seleccionado (y es distinto del ingeniero)
            if acompaniante_id and acompaniante_id != ingeniero_efectivo:
                await visitas_db_svc.propagar_acompaniante_visita(
                    conn, id_visita, acompaniante_id, context["user_db_id"]
                )

            for concepto, monto, usuario_id in viaticos_a_crear:
                await visitas_db_svc.create_viatico_visita(
                    conn, id_visita, usuario_id, concepto, monto, context["user_db_id"]
                )

        levantamientos_visita = await visitas_db_svc.get_levantamientos_en_visita(conn, id_visita)
        visita_completa = await visitas_db_svc.get_visita(conn, id_visita)
        usuarios = await visitas_db_svc.get_usuarios_para_visita(conn, id_visita)
        viaticos = await visitas_db_svc.get_viaticos_visita(conn, id_visita)
        to_configurados = await db_svc.get_to_configurados_viaticos(conn)
        cc_configurados = await db_svc.get_cc_configurados_viaticos(conn)

        # Redirigir a la página de detalle via HX-Location
        hx_location = json.dumps({
            "path": f"/levantamientos/visitas-campo/{id_visita}/ui",
            "target": "#main-content",
            "swap": "innerHTML",
        })
        return Response(content="", status_code=204, headers={"HX-Location": hx_location})

    # ----------------------------------------------------------

    @router.post("/visitas-campo/{id_visita}/viaticos", include_in_schema=False)
    async def crear_viatico_visita(
        request: Request,
        id_visita: UUID,
        usuario_id: Optional[UUID] = Form(None),
        concepto: str = Form(...),
        monto: float = Form(...),
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Agrega un viático a la visita de campo.
        Retorna la tabla de viáticos actualizada + prorrateo OOB.
        """
        if not concepto or not concepto.strip():
            raise HTTPException(status_code=400, detail="El concepto es obligatorio.")
        if monto <= 0:
            raise HTTPException(status_code=400, detail="El monto debe ser mayor a 0.")

        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        await visitas_db_svc.create_viatico_visita(
            conn, id_visita, usuario_id, concepto.strip(), monto, context["user_db_id"]
        )

        viaticos = await visitas_db_svc.get_viaticos_visita(conn, id_visita)
        levantamientos_visita = await visitas_db_svc.get_levantamientos_en_visita(conn, id_visita)
        total_viaticos = float(sum(v["monto"] for v in viaticos))
        prorrateo = calcular_prorrateo(total_viaticos, levantamientos_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/tabla_visita_campo_viaticos.html", {"viaticos": viaticos,
            "id_visita": id_visita,
            "levantamientos_visita": levantamientos_visita,
            "prorrateo": prorrateo,
            "total_viaticos": total_viaticos,
        })

    # ----------------------------------------------------------

    @router.delete("/visitas-campo/{id_visita}/viaticos/{id_viatico}", include_in_schema=False)
    async def eliminar_viatico_visita(
        request: Request,
        id_visita: UUID,
        id_viatico: UUID,
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Elimina un viático de la visita.
        Retorna la tabla y prorrateo actualizados.
        """
        eliminado = await visitas_db_svc.delete_viatico_visita(conn, id_visita, id_viatico)
        if not eliminado:
            raise HTTPException(status_code=404, detail="Viatico no encontrado.")

        viaticos = await visitas_db_svc.get_viaticos_visita(conn, id_visita)
        levantamientos_visita = await visitas_db_svc.get_levantamientos_en_visita(conn, id_visita)
        total_viaticos = float(sum(v["monto"] for v in viaticos))
        prorrateo = calcular_prorrateo(total_viaticos, levantamientos_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/tabla_visita_campo_viaticos.html", {"viaticos": viaticos,
            "id_visita": id_visita,
            "levantamientos_visita": levantamientos_visita,
            "prorrateo": prorrateo,
            "total_viaticos": total_viaticos,
        })

    # ----------------------------------------------------------

    @router.delete("/visitas-campo/{id_visita}", include_in_schema=False)
    async def eliminar_visita_campo(
        request: Request,
        id_visita: UUID,
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Elimina una Visita de Campo completa (CASCADE en viáticos, pivot y envíos).
        Retorna toast OOB. El frontend cierra el modal via hx-on::after-request.
        """
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        eliminada = await visitas_db_svc.delete_visita(conn, id_visita)
        if not eliminada:
            raise HTTPException(status_code=500, detail="No se pudo eliminar la visita.")

        logger.info(f"[VISITA_CAMPO] Visita {id_visita} eliminada por {context.get('user_name')}")

        hx_location = json.dumps({
            "path": "/levantamientos/ui",
            "target": "#main-content",
            "swap": "innerHTML",
        })
        return Response(content="", status_code=204, headers={"HX-Location": hx_location})

    # ----------------------------------------------------------

    @router.delete("/visitas-campo/{id_visita}/levantamientos/{id_levantamiento}", include_in_schema=False)
    async def desacoplar_levantamiento_visita(
        request: Request,
        id_visita: UUID,
        id_levantamiento: UUID,
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Desacopla un levantamiento de la visita y revierte su estatus a 'Pendiente'.
        Si era el último levantamiento, elimina la visita completa y cierra el modal.
        Si la visita ya tenía un envío, incluye advertencia OOB en la respuesta.
        """
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        ya_enviada = await visitas_db_svc.has_envios(conn, id_visita)
        restantes = await visitas_db_svc.remove_levantamiento_from_visita(conn, id_visita, id_levantamiento)
        await db_svc.revertir_estatus_pendiente(conn, id_levantamiento)

        logger.info(
            "Levantamiento %s desacoplado de visita %s por %s",
            id_levantamiento, id_visita, context.get('user_name')
        )

        if restantes == 0:
            await visitas_db_svc.delete_visita(conn, id_visita)
            logger.info(
                "Visita %s eliminada (sin levantamientos) por %s",
                id_visita, context.get('user_name')
            )
            hx_location = json.dumps({
                "path": "/levantamientos/ui",
                "target": "#main-content",
                "swap": "innerHTML",
            })
            return Response(content="", status_code=204, headers={"HX-Location": hx_location})

        levantamientos_visita, viaticos = await asyncio.gather(
            visitas_db_svc.get_levantamientos_en_visita(conn, id_visita),
            visitas_db_svc.get_viaticos_visita(conn, id_visita),
        )
        total_viaticos = float(sum(v["monto"] for v in viaticos))
        prorrateo = calcular_prorrateo(total_viaticos, levantamientos_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/visita_campo_levantamientos.html", {"levantamientos_visita": levantamientos_visita,
            "id_visita": id_visita,
            "oob_outerhtml": True,
            "oob_prorrateo": True,
            "prorrateo": prorrateo,
            "total_viaticos": total_viaticos,
            "oob_warning": _MSG_PRORRATEO_CAMBIO if ya_enviada else None,
        })

    # ----------------------------------------------------------

    @router.post("/visitas-campo/{id_visita}/periodo", include_in_schema=False)
    async def actualizar_periodo_visita(
        request: Request,
        id_visita: UUID,
        fecha_inicio: str = Form(...),
        fecha_fin: str = Form(...),
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Actualiza las fechas de inicio y fin de una visita existente.
        Retorna el bloque de info de período actualizado + toast OOB.
        """
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        try:
            fecha_inicio_dt = datetime.fromisoformat(fecha_inicio).replace(tzinfo=ZoneInfo("America/Mexico_City"))
            fecha_fin_dt = datetime.fromisoformat(fecha_fin).replace(tzinfo=ZoneInfo("America/Mexico_City"))
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha/hora invalido.")

        if fecha_fin_dt <= fecha_inicio_dt:
            raise HTTPException(status_code=400, detail="La fecha de fin debe ser posterior a la de inicio.")

        actualizada = await visitas_db_svc.update_periodo_visita(
            conn, id_visita, fecha_inicio_dt, fecha_fin_dt
        )
        if not actualizada:
            raise HTTPException(status_code=500, detail="No se pudo actualizar el periodo.")

        visita_actualizada = await visitas_db_svc.get_visita(conn, id_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/visita_campo_periodo.html", {"visita": visita_actualizada,
            "can_edit": True,
        })

    # ----------------------------------------------------------

    @router.post("/visitas-campo/{id_visita}/levantamientos", include_in_schema=False)
    async def agregar_levantamientos_visita(
        request: Request,
        id_visita: UUID,
        levantamiento_ids: List[UUID] = Form(...),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Agrega nuevos levantamientos a una visita existente.
        Retorna la lista actualizada de levantamientos + prorrateo OOB.
        """
        if not levantamiento_ids:
            raise HTTPException(status_code=400, detail="Debes seleccionar al menos un levantamiento.")

        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        con_viaticos = await db_svc.get_levantamientos_con_viaticos_activos(conn, levantamiento_ids)
        if con_viaticos:
            nombres = ", ".join(
                f"{r['op_id_estandar']} - {r['nombre_referencia']}" for r in con_viaticos
            )
            raise HTTPException(
                status_code=400,
                detail=f"Los siguientes levantamientos ya tienen viáticos solicitados: {nombres}. "
                       "Devuelve los viáticos antes de agregarlos a la visita de campo.",
            )

        ya_enviada = await visitas_db_svc.has_envios(conn, id_visita)
        await visitas_db_svc.add_levantamientos_to_visita(conn, id_visita, levantamiento_ids)
        await visitas_db_svc.propagar_ingeniero_visita(
            conn, id_visita, context["user_db_id"], context["user_db_id"]
        )

        levantamientos_visita, viaticos = await asyncio.gather(
            visitas_db_svc.get_levantamientos_en_visita(conn, id_visita),
            visitas_db_svc.get_viaticos_visita(conn, id_visita),
        )
        total_viaticos = float(sum(v["monto"] for v in viaticos))
        prorrateo = calcular_prorrateo(total_viaticos, levantamientos_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/visita_campo_levantamientos.html", {"levantamientos_visita": levantamientos_visita,
            "id_visita": id_visita,
            "oob_prorrateo": True,
            "prorrateo": prorrateo,
            "total_viaticos": total_viaticos,
            "oob_warning": _MSG_PRORRATEO_CAMBIO if ya_enviada else None,
        })

    # ----------------------------------------------------------

    @router.patch("/visitas-campo/{id_visita}/viaticos-opcionales", include_in_schema=False)
    async def toggle_viaticos_opcionales(
        id_visita: UUID,
        viaticos_opcionales: str = Form(...),
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Activa o desactiva el flag viaticos_opcionales de la visita.
        Recibe 'viaticos_opcionales' como string ('true'/'false'/'1'/'0').
        """
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")
        opcionales_bool = viaticos_opcionales.lower() in ("true", "1", "on")
        await visitas_db_svc.update_visita_opcionalidad(conn, id_visita, opcionales_bool)
        return Response(status_code=200)

    # ----------------------------------------------------------

    @router.patch("/visitas-campo/{id_visita}/notas", include_in_schema=False)
    async def update_notas_visita(
        request: Request,
        id_visita: UUID,
        notas: Optional[str] = Form(None),
        conn=Depends(get_db_connection),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        _=require_module_access("levantamientos", "editor"),
    ):
        """Guarda las notas de la visita de campo."""
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")
        await visitas_db_svc.update_notas_visita(conn, id_visita, notas.strip() if notas and notas.strip() else None)
        return HTMLResponse(
            content=(
                '<div x-data="{show:true}"'
                ' x-init="window.setTimeout(function(){show=false},3500)"'
                ' x-show="show" x-transition'
                ' class="pointer-events-auto max-w-sm w-full bg-slate-800 border border-emerald-500/60'
                ' rounded-xl shadow-lg p-4 flex items-start gap-3">'
                '<div class="flex-shrink-0 w-8 h-8 rounded-full bg-emerald-500/20 flex items-center justify-center">'
                '<i class="fas fa-check text-emerald-400 text-sm"></i></div>'
                '<div class="flex-1 min-w-0">'
                '<p class="text-sm font-semibold text-emerald-400">Notas guardadas</p>'
                '<p class="text-xs text-slate-400 mt-0.5">Los cambios quedaron registrados.</p>'
                '</div>'
                '<button @click="show=false" class="text-slate-500 hover:text-white transition-colors flex-shrink-0 ml-2">'
                '<i class="fas fa-times text-xs"></i></button>'
                '</div>'
            ),
            status_code=200,
        )

    # ----------------------------------------------------------

    @router.post("/visitas-campo/{id_visita}/enviar", include_in_schema=False)
    async def enviar_solicitud_visita_campo(
        request: Request,
        id_visita: UUID,
        to_destinatarios: str = Form(""),
        cc_adicionales: str = Form(""),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        visitas_db_svc: VisitasCampoDBService = Depends(get_visitas_db_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "editor"),
    ):
        """
        Envía correo con prorrateo de viáticos de la visita de campo.
        1. Obtiene viáticos y levantamientos.
        2. Calcula prorrateo.
        3. Renderiza email template.
        4. Envía via MicrosoftAuth.
        5. Registra en historial.
        6. Retorna partial historial actualizado.
        """
        visita = await visitas_db_svc.get_visita(conn, id_visita)
        if not visita:
            raise HTTPException(status_code=404, detail="Visita de campo no encontrada.")

        viaticos_opcionales = visita.get("viaticos_opcionales", False)

        viaticos = await visitas_db_svc.get_viaticos_visita(conn, id_visita)
        if not viaticos and not viaticos_opcionales:
            raise HTTPException(status_code=400, detail="No hay viaticos registrados en la visita.")

        levantamientos_visita = await visitas_db_svc.get_levantamientos_en_visita(conn, id_visita)
        total_monto = float(sum(v["monto"] for v in viaticos)) if not viaticos_opcionales else 0.0
        prorrateo = calcular_prorrateo(total_monto, levantamientos_visita) if not viaticos_opcionales else {}
        metodo_prorrateo = "Division igual"

        # Construir TO y CC
        to_configurado = await db_svc.get_to_configurados_viaticos(conn)
        to_del_form = [e.strip() for e in re.split(r'[;,]', to_destinatarios) if e.strip() and "@" in e.strip()]
        to_list = to_del_form if to_del_form else to_configurado

        cc_configurados = await db_svc.get_cc_configurados_viaticos(conn)
        cc_manuales = [e.strip() for e in re.split(r'[;,]', cc_adicionales) if e.strip() and "@" in e.strip()]
        cc_all = list(set(cc_configurados + cc_manuales) - set(to_list))

        if not to_list:
            raise HTTPException(status_code=500, detail="No hay destinatarios TO configurados.")

        now_mx = datetime.now(ZoneInfo("America/Mexico_City"))
        fecha_envio_str = now_mx.strftime("%d/%m/%Y %H:%M")

        email_template = templates.get_template("levantamientos/emails/solicitud_viaticos_visita.html")
        html_body = email_template.render(
            visita=visita,
            levantamientos=levantamientos_visita,
            viaticos=viaticos,
            prorrateo=prorrateo,
            total_monto=total_monto,
            metodo_prorrateo=metodo_prorrateo,
            viaticos_opcionales=viaticos_opcionales,
            enviado_por=context.get("user_name", "Sistema"),
            fecha_envio=fecha_envio_str,
        )

        nombre_visita = visita.get("nombre") or f"Visita {fecha_envio_str}"
        subject = f"Solicitud de Viaticos — Visita de Campo: {nombre_visita}"

        sender_config = await conn.fetchrow("""
            SELECT email_remitente FROM tb_correos_notificaciones
            WHERE departamento = 'LEVANTAMIENTOS' AND activo = true
            LIMIT 1
        """)
        if not sender_config:
            sender_config = await conn.fetchrow("""
                SELECT email_remitente FROM tb_correos_notificaciones
                WHERE departamento = 'DEFAULT' AND activo = true
                LIMIT 1
            """)
        sender_email = sender_config['email_remitente'] if sender_config else 'app-notifications@enertika.mx'

        estatus_envio = "enviado"
        error_detalle = None

        try:
            ms_auth = MicrosoftAuth()
            app_token = await ms_auth.get_application_token()
            if not app_token:
                raise Exception("No se pudo obtener token de aplicacion de Microsoft Graph.")

            success, msg = await ms_auth.send_email_with_attachments(
                access_token=app_token,
                from_email=sender_email,
                subject=subject,
                body=html_body,
                recipients=to_list,
                cc_recipients=cc_all if cc_all else None,
                importance="normal",
            )

            if not success:
                estatus_envio = "error"
                error_detalle = msg
                logger.error(f"[VISITA_CAMPO] Error envio correo visita {id_visita}: {msg}")
            else:
                logger.info(f"[VISITA_CAMPO] Correo enviado visita {id_visita}")

        except Exception as exc:
            estatus_envio = "error"
            error_detalle = str(exc)
            logger.error(f"[VISITA_CAMPO] Excepcion envio correo visita {id_visita}: {exc}")

        snapshot = {
            "viaticos_opcionales": viaticos_opcionales,
            "levantamientos": [
                {
                    "id_levantamiento": str(lev["id_levantamiento"]),
                    "op_id_estandar": lev["op_id_estandar"],
                    "nombre_proyecto": lev["nombre_proyecto"] or lev["titulo_proyecto"],
                    "cliente_nombre": lev["cliente_nombre"],
                }
                for lev in levantamientos_visita
            ],
            "viaticos": [
                {"usuario_nombre": v["usuario_nombre"], "concepto": v["concepto"], "monto": float(v["monto"])}
                for v in viaticos
            ],
            "prorrateo": prorrateo,
        }

        await visitas_db_svc.insert_envio_visita(
            conn=conn,
            id_visita=id_visita,
            enviado_por_id=context["user_db_id"],
            enviado_por_nombre=context.get("user_name", "Sistema"),
            to_destinatarios=to_list,
            cc_destinatarios=cc_all,
            snapshot=snapshot,
            total_monto=total_monto,
            estatus=estatus_envio,
            error_detalle=error_detalle,
        )

        # Al envío exitoso: registrar en historial individual de cada levantamiento
        # para que check_viaticos_sent() los reconozca y permitan cambiar a "En Proceso".
        if estatus_envio == "enviado":
            snapshot_viaticos = [
                {"usuario_nombre": v["usuario_nombre"], "concepto": v["concepto"], "monto": float(v["monto"])}
                for v in viaticos
            ]
            for lev in levantamientos_visita:
                monto_lev = float(prorrateo.get(str(lev["id_levantamiento"]), 0)) if prorrateo else 0.0
                await db_svc.insert_historial_envio(
                    conn=conn,
                    id_levantamiento=lev["id_levantamiento"],
                    enviado_por_id=context["user_db_id"],
                    enviado_por_nombre=context.get("user_name", "Sistema"),
                    to_destinatarios=to_list,
                    cc_destinatarios=cc_all,
                    viaticos_snapshot=snapshot_viaticos,
                    total_monto=monto_lev,
                    estatus="enviado",
                )

        historial_envios = await visitas_db_svc.get_envios_visita(conn, id_visita)

        return templates.TemplateResponse(request, "levantamientos/partials/historial_envios_visita.html", {"historial_envios": historial_envios,
        })

    # ==============================================================
    # POST — FECHA IDEAL DEL SOLICITANTE
    # ==============================================================

    @router.post("/operaciones/fecha-ideal/{id_levantamiento}", include_in_schema=False)
    async def actualizar_fecha_ideal_endpoint(
        request: Request,
        id_levantamiento: UUID,
        fecha_ideal: date_type = Form(...),
        hora_ideal: Optional[time_type] = Form(None),
        conn=Depends(get_db_connection),
        db_svc: LevantamientosDBService = Depends(get_db_service),
        service: LevantamientoService = Depends(get_service),
        context=Depends(get_current_user_context),
        _=require_module_access("levantamientos", "viewer"),
    ):
        """
        Actualiza la fecha ideal del solicitante en el levantamiento.
        Permitido para: el propio solicitante o usuarios con rol editor+ en el módulo.
        """
        lev = await db_svc.get_levantamiento_base(conn, id_levantamiento)
        if not lev:
            raise HTTPException(status_code=404, detail="Levantamiento no encontrado")

        # Validar permisos: solicitante propio O editor+
        user_db_id = context.get("user_db_id")
        user_role = context.get("role")
        module_role = context.get("module_roles", {}).get("levantamientos")
        is_solicitante = str(lev.get("solicitado_por_id", "")) == str(user_db_id)
        can_edit = (
            user_role == "ADMIN"
            or module_role in ["editor", "admin"]
            or is_solicitante
        )
        if not can_edit:
            raise HTTPException(status_code=403, detail="No tienes permiso para modificar la fecha ideal.")

        # Combinar fecha + hora
        hora = hora_ideal if hora_ideal else time_type(0, 0)
        fecha_ideal_dt = datetime.combine(fecha_ideal, hora)

        await db_svc.update_fecha_ideal_solicitante(conn, id_levantamiento, fecha_ideal_dt)

        return await _render_kanban(request, conn, service, context)
