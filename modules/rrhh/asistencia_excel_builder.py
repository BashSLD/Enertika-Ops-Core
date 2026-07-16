from __future__ import annotations

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from modules.asistencia.constants import formatear_estado_asistencia_label
from modules.rrhh.excel_utils import autofit_columns, format_date, format_datetime, style_sheet
from modules.shared.utils import format_minutes, safe_sheet_title

FORMATOS_REPORTE_ASISTENCIA = frozenset({
    "consolidado",
    "detalle",
    "departamentos",
    "completo",
})

_ASISTENCIA_HEADERS = [
    "Fecha",
    "Empleado",
    "Email",
    "Sucursal",
    "Departamento",
    "Primera entrada",
    "Ultima salida",
    "Horas trabajadas",
    "Horas a cubrir",
    "Horas extra",
    "Estado",
    "Ausencia aprobada",
    "Tipo de ausencia",
    "Observaciones",
]
_CONSOLIDADO_HEADERS = [
    "Empleado",
    "Email",
    "Departamento",
    "Horas trabajadas horario laboral",
    "Horas extra autorizadas",
    "Total autorizado",
]
_RESERVED_SHEET_TITLES = {"checadas sin mapear"}


def build_asistencia_workbook(
    rows: list[dict],
    unmapped: list[dict],
    formato: str,
) -> Workbook:
    if formato not in FORMATOS_REPORTE_ASISTENCIA:
        raise ValueError("Formato de asistencia no valido")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    if formato in {"consolidado", "completo"}:
        _append_consolidado_sheet(workbook, rows)
    if formato in {"detalle", "completo"}:
        _append_detail_sheet(workbook, "Asistencia", rows)
    if formato == "completo":
        _append_detail_sheet(
            workbook,
            "Por empleado",
            sorted(rows, key=_employee_date_sort_key),
        )
    if formato in {"departamentos", "completo"}:
        _append_departamento_sheets(workbook, rows)
    _append_unmapped_biotime_sheet(workbook, unmapped)
    return workbook


def _append_consolidado_sheet(workbook: Workbook, rows: list[dict]) -> None:
    worksheet = workbook.create_sheet("Consolidado")
    style_sheet(worksheet, _CONSOLIDADO_HEADERS)

    consolidated: dict[object, dict] = {}
    for row in rows:
        employee_id = row.get("usuario_id")
        if employee_id is None:
            employee_id = (row.get("empleado_email"), row.get("empleado_nombre"))
        values = consolidated.setdefault(
            employee_id,
            {
                "empleado_nombre": row.get("empleado_nombre") or "",
                "empleado_email": row.get("empleado_email") or "",
                "departamento": _department_name(row),
                "minutos_laboral": 0,
                "minutos_aprobados": 0,
            },
        )
        minutos_trabajados = int(row.get("minutos_trabajados") or 0)
        minutos_extra = int(row.get("minutos_extra") or 0)
        values["minutos_laboral"] += max(0, minutos_trabajados - minutos_extra)
        if row.get("horas_extra_estado") != "feriado":
            values["minutos_aprobados"] += int(row.get("minutos_aprobados") or 0)

    ordered_rows = sorted(
        consolidated.values(),
        key=lambda row: (row["empleado_nombre"].casefold(), row["empleado_email"].casefold()),
    )
    minutos_laboral_list: list[int] = []
    minutos_aprobados_list: list[int] = []
    for row in ordered_rows:
        minutos_laboral = row["minutos_laboral"]
        minutos_aprobados = row["minutos_aprobados"]
        worksheet.append([
            row["empleado_nombre"],
            row["empleado_email"],
            row["departamento"],
            format_minutes(minutos_laboral),
            format_minutes(minutos_aprobados),
            None,
        ])
        minutos_laboral_list.append(minutos_laboral)
        minutos_aprobados_list.append(minutos_aprobados)

    column_laboral = _header_column_letter(_CONSOLIDADO_HEADERS, "Horas trabajadas horario laboral")
    column_aprobadas = _header_column_letter(_CONSOLIDADO_HEADERS, "Horas extra autorizadas")
    column_total_index = _header_column_index(_CONSOLIDADO_HEADERS, "Total autorizado")
    column_total = get_column_letter(column_total_index)

    helper_columns = _append_minutes_helper_columns(
        worksheet,
        {
            column_laboral: ("minutos_laboral", minutos_laboral_list),
            column_aprobadas: ("minutos_aprobados", minutos_aprobados_list),
        },
    )
    helper_laboral = helper_columns[column_laboral]
    helper_aprobadas = helper_columns[column_aprobadas]
    for row_number in range(2, len(ordered_rows) + 2):
        worksheet.cell(
            row=row_number,
            column=column_total_index,
            value=_duration_formula(f"{helper_laboral}{row_number},{helper_aprobadas}{row_number}"),
        )

    _append_duration_total_row(
        worksheet,
        {
            column_laboral: [helper_laboral],
            column_aprobadas: [helper_aprobadas],
            column_total: [helper_laboral, helper_aprobadas],
        },
    )
    autofit_columns(worksheet, visible_columns=len(_CONSOLIDADO_HEADERS))


def _append_detail_sheet(workbook: Workbook, title: str, rows: list[dict]) -> None:
    worksheet = workbook.create_sheet(title)
    style_sheet(worksheet, _ASISTENCIA_HEADERS)
    minutos_trabajados_list: list[int] = []
    minutos_programados_list: list[int] = []
    minutos_extra_list: list[int] = []
    for row in rows:
        worksheet.append(_asistencia_row_values(row))
        minutos_trabajados_list.append(int(row.get("minutos_trabajados") or 0))
        minutos_programados_list.append(int(row.get("minutos_programados") or 0))
        minutos_extra_list.append(int(row.get("minutos_extra") or 0))

    column_trabajadas = _header_column_letter(_ASISTENCIA_HEADERS, "Horas trabajadas")
    column_programadas = _header_column_letter(_ASISTENCIA_HEADERS, "Horas a cubrir")
    column_extra = _header_column_letter(_ASISTENCIA_HEADERS, "Horas extra")

    helper_columns = _append_minutes_helper_columns(
        worksheet,
        {
            column_trabajadas: ("minutos_trabajados", minutos_trabajados_list),
            column_programadas: ("minutos_programados", minutos_programados_list),
            column_extra: ("minutos_extra", minutos_extra_list),
        },
    )
    _append_duration_total_row(
        worksheet, {column: [helper] for column, helper in helper_columns.items()}
    )
    autofit_columns(worksheet, visible_columns=len(_ASISTENCIA_HEADERS))


def _append_departamento_sheets(workbook: Workbook, rows: list[dict]) -> None:
    grouped_rows: dict[str, list[dict]] = {}
    for row in rows:
        department = _department_name(row)
        grouped_rows.setdefault(department, []).append(row)

    used_titles = _RESERVED_SHEET_TITLES | {worksheet.title.casefold() for worksheet in workbook.worksheets}
    for department in sorted(grouped_rows, key=str.casefold):
        title = safe_sheet_title(department, used_titles, fallback="Sin departamento")
        _append_detail_sheet(workbook, title, sorted(grouped_rows[department], key=_employee_date_sort_key))


def _append_unmapped_biotime_sheet(workbook: Workbook, rows: list[dict]) -> None:
    if not rows:
        return
    worksheet = workbook.create_sheet("Checadas sin mapear")
    headers = [
        "Codigo BioTime",
        "Departamento",
        "Checadas",
        "Primera checada",
        "Ultima checada",
    ]
    style_sheet(worksheet, headers)
    for row in rows:
        worksheet.append([
            row.get("biotime_emp_code") or "",
            row.get("deptname") or "",
            row.get("total") or 0,
            format_datetime(row.get("primera_checada")),
            format_datetime(row.get("ultima_checada")),
        ])
    autofit_columns(worksheet, visible_columns=len(headers))


def _append_minutes_helper_columns(
    worksheet, minutes_by_column: dict[str, tuple[str, list[int]]]
) -> dict[str, str]:
    """Returns a mapping of visible column letter to its hidden helper column letter."""
    helper_columns: dict[str, str] = {}
    start_column = worksheet.max_column + 1
    for index, (visible_column, (field, minutes)) in enumerate(minutes_by_column.items(), start=start_column):
        column_letter = get_column_letter(index)
        helper_columns[visible_column] = column_letter
        worksheet.cell(row=1, column=index, value=f"_{field}")
        for row_index, value in enumerate(minutes, start=2):
            worksheet.cell(row=row_index, column=index, value=value)
        worksheet.column_dimensions[column_letter].hidden = True
    return helper_columns


def _append_duration_total_row(
    worksheet, helper_columns: dict[str, list[str]]
) -> None:
    last_data_row = worksheet.max_row
    has_data_rows = last_data_row > 1
    total_row = last_data_row + 1
    worksheet.cell(row=total_row, column=1, value="Total")
    for visible_column, helper_column_list in helper_columns.items():
        if has_data_rows:
            ranges = ",".join(
                f"{column}2:{column}{last_data_row}" for column in helper_column_list
            )
            worksheet[f"{visible_column}{total_row}"] = _duration_formula(ranges)
        else:
            worksheet[f"{visible_column}{total_row}"] = "0m"
    for cell in worksheet[total_row]:
        cell.font = Font(bold=True)


def _duration_formula(range_ref: str) -> str:
    hours = f"INT(SUM({range_ref})/60)"
    minutes = f"MOD(SUM({range_ref}),60)"
    return (
        f'=IF(AND({hours}>0,{minutes}>0),{hours}&"h "&TEXT({minutes},"00")&"m",'
        f'IF({hours}>0,{hours}&"h",{minutes}&"m"))'
    )


def _asistencia_row_values(row: dict) -> list:
    return [
        format_date(row.get("fecha_laboral")),
        row.get("empleado_nombre") or "",
        row.get("empleado_email") or "",
        row.get("sucursal_nombre") or "",
        row.get("departamento") or "",
        format_datetime(row.get("primera_entrada")),
        format_datetime(row.get("ultima_salida")),
        format_minutes(row.get("minutos_trabajados")),
        format_minutes(row.get("minutos_programados")),
        format_minutes(row.get("minutos_extra")),
        formatear_estado_asistencia_label(row.get("estado"), row.get("tipo_ausencia_nombre")),
        "Si" if row.get("tiene_ausencia_justificada") else "No",
        row.get("tipo_ausencia_nombre") or "",
        row.get("observaciones") or "",
    ]


def _header_column_index(headers: list[str], header_name: str) -> int:
    return headers.index(header_name) + 1


def _header_column_letter(headers: list[str], header_name: str) -> str:
    return get_column_letter(_header_column_index(headers, header_name))


def _department_name(row: dict) -> str:
    return str(row.get("departamento") or "").strip() or "Sin departamento"


def _employee_date_sort_key(row: dict) -> tuple[str, date]:
    return (
        str(row.get("empleado_nombre") or "").casefold(),
        row.get("fecha_laboral") or date.min,
    )
