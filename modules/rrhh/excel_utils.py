from __future__ import annotations

from datetime import date, datetime

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.asistencia.logic import ensure_mx


def style_sheet(worksheet, headers: list[str], header_row: int = 1) -> None:
    """Escribe `headers` en `header_row` (agregalos despues de cualquier fila previa,
    ej. una nota) y congela los paneles justo debajo."""
    worksheet.append(headers)
    worksheet.freeze_panes = f"A{header_row + 1}"
    for cell in worksheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123456")


def autofit_columns(worksheet, visible_columns: int | None = None, min_row: int = 1) -> None:
    """`min_row` permite saltar filas previas al header (ej. una nota) que no deben
    influir en el ancho de columna."""
    max_column = visible_columns or worksheet.max_column
    max_lengths = [0] * max_column
    for row in worksheet.iter_rows(min_row=min_row, min_col=1, max_col=max_column, values_only=True):
        for index, value in enumerate(row):
            length = len(str(value)) if value else 0
            if length > max_lengths[index]:
                max_lengths[index] = length
    for index, length in enumerate(max_lengths, start=1):
        column_letter = get_column_letter(index)
        worksheet.column_dimensions[column_letter].width = min(length + 2, 36)


def format_date(value: date | None) -> str:
    return value.strftime("%d/%m/%Y") if value else ""


def format_datetime(value: datetime | None) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return ensure_mx(value).strftime("%d/%m/%Y %H:%M")
    return str(value)
