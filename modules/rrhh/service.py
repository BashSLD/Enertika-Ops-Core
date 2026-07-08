from __future__ import annotations

import asyncpg
import base64
import binascii
import json
import logging
import re
import time as time_module
import unicodedata
import zlib
from datetime import date, time, timedelta
from io import BytesIO
from uuid import UUID

from itsdangerous import BadData, SignatureExpired, TimestampSigner

from core.config import settings
from core.config_service import ConfigService
from core.timezone import today_mx
from modules.asistencia import db_service as asistencia_db
from modules.asistencia.service import (
    attach_he_evidencias,
    build_horas_extra_grupos,
    format_solicitudes_manuales,
    get_he_bolsa_fecha_corte,
    recalcular_asistencia,
    recalcular_asistencia_reciente_usuario,
)
from modules.asistencia.constants import ASISTENCIA_ESTADO_LABELS, ASISTENCIA_ESTADOS
from modules.rrhh import db_service as rrhh_db
from modules.vacaciones import db_service as vac_db
from modules.vacaciones.constants import ESTADOS_SOLICITUD
from modules.vacaciones.holidays import generar_feriados_mexico
from modules.vacaciones.logic import calcular_balance, calcular_periodos

logger = logging.getLogger("rrhh.service")

DIAS_SEMANA = [
    {"value": 0, "nombre": "Lunes"},
    {"value": 1, "nombre": "Martes"},
    {"value": 2, "nombre": "Miércoles"},
    {"value": 3, "nombre": "Jueves"},
    {"value": 4, "nombre": "Viernes"},
    {"value": 5, "nombre": "Sábado"},
    {"value": 6, "nombre": "Domingo"},
]

MINUTOS_DIA = 24 * 60
MIGRACION_PREVIEW_TTL_SECONDS = 20 * 60
MIGRACION_MAX_FILE_BYTES = 5 * 1024 * 1024
FESTIVOS_ANIO_MIN = 2026
FESTIVOS_ANIO_MAX = 2100


async def get_solicitudes_manuales_pendientes_todas_svc(conn) -> list[dict]:
    solicitudes = await asistencia_db.get_solicitudes_manuales_pendientes_todas(conn)
    return format_solicitudes_manuales(solicitudes)


async def get_aprobaciones_ctx_rrhh_svc(conn, context: dict, rrhh_perms: dict, **extra) -> dict:
    hoy = today_mx()
    pendientes = await vac_db.get_todas_solicitudes_pendientes(conn)
    horas_extra = await asistencia_db.get_horas_extra_todas(conn, hoy - timedelta(days=30), hoy)
    solicitudes_manuales = await get_solicitudes_manuales_pendientes_todas_svc(conn)
    await attach_he_evidencias(conn, horas_extra)
    horas_extra_grupos, horas_extra_json = build_horas_extra_grupos(horas_extra)
    comp_pendientes = (
        await asistencia_db.get_he_compensatorio_pendientes(conn)
        if rrhh_perms["can_edit"]
        else []
    )
    saldo_inicial_pendientes = (
        await asistencia_db.get_saldo_inicial_pendientes(conn, fecha_corte=await get_he_bolsa_fecha_corte(conn))
        if rrhh_perms["can_edit"]
        else []
    )
    return {
        "pendientes": pendientes,
        "horas_extra_grupos": horas_extra_grupos,
        "horas_extra_json": horas_extra_json,
        "solicitudes_manuales_pendientes": solicitudes_manuales,
        "comp_pendientes": comp_pendientes,
        "saldo_inicial_pendientes": saldo_inicial_pendientes,
        "can_manage_compensatorio": rrhh_perms["can_edit"],
        "context": context,
        "rrhh_perms": rrhh_perms,
        **extra,
    }


async def get_dashboard_data(conn) -> dict:
    hoy = today_mx()
    ausencias_hoy = await vac_db.get_ausencias_activas(conn, hoy, hoy)
    pendientes = await vac_db.get_todas_solicitudes_pendientes(conn)
    total_empleados = await vac_db.count_empleados(conn)
    migracion_vacaciones = await vac_db.count_empleados_migrados(conn)
    return {
        "ausencias_hoy": ausencias_hoy,
        "pendientes": pendientes,
        "total_empleados": total_empleados,
        "migracion_vacaciones": migracion_vacaciones,
        "hoy": hoy,
    }


async def get_empleado_edit_ctx(conn, usuario_id: UUID) -> dict:
    empleado = await vac_db.get_empleado_datos(conn, usuario_id)
    usuario = await rrhh_db.get_usuario_simple_by_id(conn, usuario_id)
    jefes = await vac_db.get_jefes_con_nombre(conn, usuario_id)
    usuarios = await vac_db.get_usuarios_activos_simples(conn)
    prorrogas = await vac_db.get_prorrogas_usuario(conn, usuario_id)
    jefes_ids = {j["id_usuario"] for j in jefes}
    return {
        "empleado": empleado,
        "usuario": usuario or {},
        "jefes": jefes,
        "jefes_ids": jefes_ids,
        "usuarios": usuarios,
        "sucursales": await asistencia_db.get_sucursales(conn),
        "prorrogas": prorrogas,
    }


def _encode_b64(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def _decode_b64(value: str) -> bytes:
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode((value + padding).encode("ascii"))


def _firmar_preview(rows: list[dict], ttl_seconds: int = MIGRACION_PREVIEW_TTL_SECONDS) -> str:
    raw = zlib.compress(json.dumps(rows, separators=(",", ":"), default=str).encode("utf-8"))
    signer = TimestampSigner(settings.SECRET_KEY)
    return signer.sign(_encode_b64(raw)).decode("ascii")


def _leer_preview_firmado(token: str) -> list[dict]:
    signer = TimestampSigner(settings.SECRET_KEY)
    try:
        encoded = signer.unsign(token, max_age=MIGRACION_PREVIEW_TTL_SECONDS).decode("ascii")
        rows = json.loads(zlib.decompress(_decode_b64(encoded)).decode("utf-8"))
    except SignatureExpired:
        raise ValueError("La vista previa expiro. Vuelve a importar el archivo.")
    except (BadData, binascii.Error) as exc:
        raise ValueError("La vista previa no es valida. Vuelve a importar el archivo.") from exc
    except (zlib.error, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("La vista previa no es valida. Vuelve a importar el archivo.") from exc
    if not isinstance(rows, list):
        raise ValueError("La vista previa no contiene filas validas.")
    return rows


def _normalizar_header_excel(value: object) -> str:
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def _parse_dias_excel(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return 0
        value = value.replace(",", ".")
    if isinstance(value, bool):
        raise ValueError("debe ser un numero entero")
    if isinstance(value, int):
        dias = value
    elif isinstance(value, float) and value.is_integer():
        dias = int(value)
    elif isinstance(value, str) and re.fullmatch(r"\d+(\.0+)?", value):
        dias = int(float(value))
    else:
        raise ValueError("debe ser un numero entero")
    if dias < 0:
        raise ValueError("no puede ser negativo")
    return dias


def _sumar_consumos_por_periodo(consumos: list[dict]) -> dict[int, int]:
    result: dict[int, int] = {}
    for consumo in consumos:
        num_periodo = int(consumo["num_periodo"])
        result[num_periodo] = result.get(num_periodo, 0) + int(consumo["dias_consumidos"])
    return result


def _token_rows_migracion(rows: list[dict]) -> list[dict]:
    token_rows = []
    for row in rows:
        if row.get("errores") or not row.get("limpiar"):
            continue
        token_rows.append({
            "usuario_id": str(row["usuario_id"]),
            "limpiar": True,
            "periodos": [
                {"num_periodo": p["num_periodo"], "dias": p["dias"]}
                for p in row.get("periodos", [])
            ],
        })
    return token_rows


def _periodo_preview(periodo: dict, dias: int, consumo_no_migrado: int) -> dict:
    return {
        "num_periodo": periodo["num_periodo"],
        "dias": dias,
        "dias_otorgados": periodo["dias_otorgados"],
        "dias_maximos": max(0, periodo["dias_otorgados"] - consumo_no_migrado),
        "fecha_aniversario": periodo["fecha_aniversario"],
        "fecha_expiracion": periodo["fecha_expiracion"],
        "expirado": periodo.get("expirado", False),
    }


def _periodos_migrables(
    empleado: dict,
    hoy: date,
    catalogo: list[dict],
    meses_exp: int,
    consumos_no_migracion: list[dict],
) -> tuple[list[dict], dict[int, int]]:
    periodos = calcular_periodos(
        empleado["fecha_contratacion"],
        hoy,
        catalogo,
        ajuste_dias=empleado.get("dias_vacaciones_ajuste") or 0,
        meses_expiracion=meses_exp,
    )
    balance = calcular_balance(periodos, consumos_no_migracion)
    consumo_por_periodo = _sumar_consumos_por_periodo(consumos_no_migracion)
    return balance, consumo_por_periodo


async def _validar_rows_migracion(conn, raw_rows: list[dict]) -> list[dict]:
    hoy = today_mx()
    empleados = await vac_db.get_empleados_para_migracion(conn)
    empleados_by_id = {str(emp["id_usuario"]): emp for emp in empleados}
    catalogo = await vac_db.get_catalogo_dias(conn)
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)

    normalizadas: list[dict] = []
    usuarios_validos: list[UUID] = []
    vistos: set[str] = set()

    for raw in raw_rows:
        errores = list(raw.get("errores", []))
        usuario_id_text = str(raw.get("usuario_id") or "").strip()
        usuario_uuid: UUID | None = None
        empleado = None
        if not usuario_id_text:
            errores.append("Falta usuario_id")
        else:
            try:
                usuario_uuid = UUID(usuario_id_text)
            except ValueError:
                errores.append("usuario_id no es valido")

        if usuario_uuid:
            empleado = empleados_by_id.get(str(usuario_uuid))
            if not empleado:
                errores.append("El empleado no existe, está inactivo o no tiene fecha de contratación")
            elif str(usuario_uuid) in vistos:
                errores.append("Empleado duplicado en la importacion")
            else:
                vistos.add(str(usuario_uuid))
                usuarios_validos.append(usuario_uuid)

        periodos_raw = raw.get("periodos") or []
        touched = bool(raw.get("limpiar") or raw.get("tocado") or periodos_raw)
        if not touched and not errores:
            continue

        dias_por_periodo: dict[int, int] = {}
        for periodo_raw in periodos_raw:
            try:
                num_periodo = int(periodo_raw.get("num_periodo"))
                dias = _parse_dias_excel(periodo_raw.get("dias"))
            except (TypeError, ValueError) as exc:
                errores.append(f"Periodo invalido: {exc}")
                continue
            if num_periodo <= 0:
                errores.append("El periodo debe ser mayor a cero")
                continue
            if num_periodo in dias_por_periodo:
                errores.append(f"Periodo {num_periodo} esta duplicado")
                continue
            dias_por_periodo[num_periodo] = dias

        normalizadas.append({
            "excel_row": raw.get("excel_row"),
            "usuario_id": usuario_uuid,
            "usuario_id_text": usuario_id_text,
            "empleado": empleado,
            "limpiar": touched,
            "dias_por_periodo": dias_por_periodo,
            "errores": errores,
        })

    consumos_no_migracion = await vac_db.get_consumos_no_migracion_bulk(conn, usuarios_validos)
    resultado = []
    for row in normalizadas:
        empleado = row["empleado"]
        usuario_id = row["usuario_id"]
        errores = row["errores"]
        periodos_preview = []
        total_dias = 0

        if empleado and usuario_id:
            disponibles, consumo_no_migrado = _periodos_migrables(
                empleado,
                hoy,
                catalogo,
                meses_exp,
                consumos_no_migracion.get(usuario_id, []),
            )
            disponibles_by_num = {p["num_periodo"]: p for p in disponibles}
            total_otorgado = 0
            total_consumo_no_migrado = 0
            for periodo in disponibles:
                num = periodo["num_periodo"]
                total_otorgado += int(periodo["dias_otorgados"])
                total_consumo_no_migrado += consumo_no_migrado.get(num, 0)

            for num_periodo, dias in sorted(row["dias_por_periodo"].items()):
                periodo = disponibles_by_num.get(num_periodo)
                if not periodo:
                    errores.append(f"Periodo {num_periodo} no existe o aun no esta disponible")
                    continue
                consumo_actual = consumo_no_migrado.get(num_periodo, 0)
                if not periodo.get("es_proximo") and dias + consumo_actual > periodo["dias_otorgados"]:
                    errores.append(
                        f"Periodo {num_periodo}: {dias} dias excede el maximo disponible "
                        f"({max(0, periodo['dias_otorgados'] - consumo_actual)})"
                    )
                    continue
                total_dias += dias
                if dias > 0:
                    periodos_preview.append(_periodo_preview(periodo, dias, consumo_actual))

            has_proximo = any(p.get("es_proximo") for p in disponibles)
            if not has_proximo and total_dias + total_consumo_no_migrado > total_otorgado:
                errores.append("La suma total excede los dias otorgados")

        resultado.append({
            "excel_row": row["excel_row"],
            "usuario_id": usuario_id,
            "usuario_id_text": row["usuario_id_text"],
            "nombre": empleado["nombre"] if empleado else "",
            "email": empleado["email"] if empleado else "",
            "limpiar": row["limpiar"],
            "periodos": periodos_preview,
            "total_dias": total_dias,
            "errores": errores,
        })
    return resultado


async def get_migracion_ctx(conn) -> dict:
    conteo = await vac_db.count_empleados_migrados(conn)
    empleados = await vac_db.get_empleados_para_migracion(conn)
    return {
        "conteo": conteo,
        "empleados": empleados,
    }


async def generar_plantilla_migracion(conn):
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill, Protection

    hoy = today_mx()
    empleados = await vac_db.get_empleados_para_migracion(conn)
    catalogo = await vac_db.get_catalogo_dias(conn)
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
    usuario_ids = [emp["id_usuario"] for emp in empleados]
    consumos_no_migracion = await vac_db.get_consumos_no_migracion_bulk(conn, usuario_ids)
    migraciones = await vac_db.get_consumos_migracion_bulk(conn, usuario_ids)

    periodos_por_usuario: dict[UUID, list[dict]] = {}
    max_periodo = 0
    max_por_periodo: dict[int, int] = {}
    for emp in empleados:
        uid = emp["id_usuario"]
        disponibles, consumo_no_migrado = _periodos_migrables(
            emp,
            hoy,
            catalogo,
            meses_exp,
            consumos_no_migracion.get(uid, []),
        )
        periodos_por_usuario[uid] = disponibles
        for periodo in disponibles:
            num_periodo = periodo["num_periodo"]
            max_migrable = max(0, periodo["dias_otorgados"] - consumo_no_migrado.get(num_periodo, 0))
            max_por_periodo[num_periodo] = max(max_por_periodo.get(num_periodo, 0), max_migrable)
        if disponibles:
            max_periodo = max(max_periodo, max(p["num_periodo"] for p in disponibles))

    workbook = Workbook()

    # --- Hoja de instrucciones ---
    instrucciones = workbook.active
    instrucciones.title = "Instrucciones"
    instrucciones.sheet_view.showGridLines = False
    instrucciones.column_dimensions["A"].width = 3
    instrucciones.column_dimensions["B"].width = 28
    instrucciones.column_dimensions["C"].width = 60

    titulo_font = Font(bold=True, size=14, color="123456")
    seccion_font = Font(bold=True, size=11, color="FFFFFF")
    seccion_fill = PatternFill("solid", fgColor="123456")
    normal_font = Font(size=10)
    ejemplo_fill = PatternFill("solid", fgColor="F0FFFE")

    def _ins(row, col, value, font=None, fill=None, wrap=False):
        cell = instrucciones.cell(row=row, column=col, value=value)
        cell.font = font or normal_font
        if fill:
            cell.fill = fill
        if wrap:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    instrucciones.row_dimensions[1].height = 8
    _ins(2, 2, "Plantilla de migración histórica de vacaciones", font=titulo_font)
    instrucciones.row_dimensions[2].height = 22

    instrucciones.row_dimensions[3].height = 6
    _ins(4, 2, "QUE ES ESTA PLANTILLA", font=seccion_font, fill=seccion_fill)
    instrucciones.merge_cells("B4:C4")
    instrucciones.row_dimensions[4].height = 18
    instrucciones.row_dimensions[5].height = 48
    _ins(5, 2, "Para que sirve:", font=Font(bold=True, size=10))
    _ins(5, 3,
         "Registrar los dias de vacaciones que cada empleado YA USO antes de que el sistema existiera. "
         "Esto permite que el saldo que muestra el sistema sea correcto desde el primer dia.",
         wrap=True)
    instrucciones.row_dimensions[6].height = 40
    _ins(6, 2, "Que NO capturar:", font=Font(bold=True, size=10))
    _ins(6, 3,
         "No captures dias pendientes ni saldos disponibles. Solo los dias que el empleado "
         "efectivamente tomo ANTES de que el sistema entrara en operacion.",
         wrap=True)

    instrucciones.row_dimensions[7].height = 8
    _ins(8, 2, "COLUMNAS DE LA HOJA DE DATOS", font=seccion_font, fill=seccion_fill)
    instrucciones.merge_cells("B8:C8")
    instrucciones.row_dimensions[8].height = 18

    columnas = [
        ("usuario_id", "Identificador interno. No modificar."),
        ("Nombre / Email", "Datos del empleado. No modificar."),
        ("Fecha contratación", "Fecha de ingreso. No modificar."),
        ("Periodos calculados", "Resumen de dias disponibles por periodo. Solo referencia."),
        ("Ya migrado", "Si = ya tiene historial cargado. No = pendiente."),
        ("Periodo N (max X dias)",
         "UNICO CAMPO A LLENAR. Escribe cuantos dias tomo el empleado en ese periodo "
         "antes del sistema. Si no tomo ninguno, deja la celda en blanco o pon 0. "
         "No puedes capturar mas del maximo indicado."),
    ]
    for i, (col_name, desc) in enumerate(columnas, start=9):
        instrucciones.row_dimensions[i].height = 36 if i == 14 else 20
        _ins(i, 2, col_name, font=Font(bold=True, size=10))
        _ins(i, 3, desc, wrap=True)

    instrucciones.row_dimensions[15].height = 8
    _ins(16, 2, "COLORES EN LA HOJA DE DATOS", font=seccion_font, fill=seccion_fill)
    instrucciones.merge_cells("B16:C16")
    instrucciones.row_dimensions[16].height = 18

    colores = [
        (PatternFill("solid", fgColor="E6FFFB"), "Verde claro", "Celda editable. Captura aqui los dias tomados."),
        (PatternFill("solid", fgColor="E5E7EB"), "Gris", "No aplica para este empleado en este periodo."),
        (PatternFill("solid", fgColor="F3F4F6"), "Gris oscuro", "Periodo aun no disponible o bloqueado."),
        (PatternFill("solid", fgColor="E5E7EB"), "Gris (vencido)", "El periodo ya expiro pero puedes registrar dias tomados."),
        (PatternFill("solid", fgColor="FEF3C7"), "Amarillo", "Empleado ya tiene historial cargado anteriormente."),
    ]
    for i, (fill, nombre, desc) in enumerate(colores, start=17):
        instrucciones.row_dimensions[i].height = 18
        color_cell = instrucciones.cell(row=i, column=2, value=nombre)
        color_cell.fill = fill
        color_cell.font = Font(size=10)
        _ins(i, 3, desc)

    instrucciones.row_dimensions[22].height = 8
    _ins(23, 2, "EJEMPLO", font=seccion_font, fill=seccion_fill)
    instrucciones.merge_cells("B23:C23")
    instrucciones.row_dimensions[23].height = 18

    ejemplo_rows = [
        ("Empleado con 2 periodos completos:",
         "Le corresponden 15 dias en P1 y 14 dias en P2 segun el catalogo Enertika."),
        ("Tomo 8 dias en P1 antes del sistema:",
         "Captura 8 en la columna Periodo 1. El sistema mostrara 7 dias disponibles en P1."),
        ("No tomo nada en P2:",
         "Deja Periodo 2 en blanco. El sistema mostrara 14 dias disponibles en P2."),
        ("Empleado nuevo (menos de 1 anio):",
         "El Periodo 1 aparece editable. Captura los dias que tomo como anticipo antes de que el sistema existiera."),
    ]
    for i, (concepto, explicacion) in enumerate(ejemplo_rows, start=24):
        instrucciones.row_dimensions[i].height = 20
        c = instrucciones.cell(row=i, column=2, value=concepto)
        c.font = Font(bold=True, size=10)
        c.fill = ejemplo_fill
        e = instrucciones.cell(row=i, column=3, value=explicacion)
        e.font = normal_font
        e.fill = ejemplo_fill
        e.alignment = Alignment(wrap_text=True, vertical="top")

    instrucciones.row_dimensions[28].height = 8
    _ins(29, 2, "PROCESO DE CARGA", font=seccion_font, fill=seccion_fill)
    instrucciones.merge_cells("B29:C29")
    instrucciones.row_dimensions[29].height = 18

    pasos = [
        ("1. Llenar", "Captura los dias en las celdas verdes de la hoja 'Migracion vacaciones'."),
        ("2. Guardar", "Guarda el archivo en formato .xlsx sin cambiar el nombre de las hojas."),
        ("3. Subir", "En RRHH > Migración histórica, selecciona el archivo y haz clic en 'Validar archivo'."),
        ("4. Revisar", "El sistema muestra una vista previa con errores si los hay. Corrige y vuelve a subir si es necesario."),
        ("5. Confirmar", "Si todo esta correcto, haz clic en 'Confirmar importacion'. Los saldos se actualizan al instante."),
    ]
    for i, (paso, desc) in enumerate(pasos, start=30):
        instrucciones.row_dimensions[i].height = 20
        _ins(i, 2, paso, font=Font(bold=True, size=10))
        _ins(i, 3, desc, wrap=True)

    instrucciones.protection.sheet = True
    instrucciones.protection.enable()

    # --- Hoja de datos ---
    worksheet = workbook.create_sheet("Migración vacaciones")

    headers = [
        "usuario_id",
        "Nombre",
        "Email",
        "Fecha contratación",
        "Periodos calculados",
        "Ya migrado",
    ] + [
        f"Periodo {num} (max {max_por_periodo.get(num, 0)} dias)"
        for num in range(1, max_periodo + 1)
    ]
    worksheet.append(headers)
    worksheet.freeze_panes = "G2"

    header_fill = PatternFill("solid", fgColor="123456")
    editable_fill = PatternFill("solid", fgColor="E6FFFB")
    locked_fill = PatternFill("solid", fgColor="F3F4F6")
    expired_fill = PatternFill("solid", fgColor="E5E7EB")
    migrated_fill = PatternFill("solid", fgColor="FEF3C7")

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.protection = Protection(locked=True)

    for row_index, emp in enumerate(empleados, start=2):
        uid = emp["id_usuario"]
        periodos = periodos_por_usuario.get(uid, [])
        periodos_by_num = {p["num_periodo"]: p for p in periodos}
        consumo_no_migrado = _sumar_consumos_por_periodo(consumos_no_migracion.get(uid, []))
        migrado_por_periodo = _sumar_consumos_por_periodo(migraciones.get(uid, []))
        resumen = "; ".join(
            f"P{p['num_periodo']}: {max(0, p['dias_otorgados'] - consumo_no_migrado.get(p['num_periodo'], 0))} dias"
            for p in periodos
        )
        worksheet.append([
            str(uid),
            emp["nombre"],
            emp["email"],
            emp["fecha_contratacion"],
            resumen,
            "Si" if emp.get("ya_migrado") else "No",
        ] + [None] * max_periodo)
        if emp.get("ya_migrado"):
            for cell in worksheet[row_index]:
                cell.fill = migrated_fill

        for fixed_cell in worksheet[row_index][:6]:
            fixed_cell.protection = Protection(locked=True)

        for num_periodo in range(1, max_periodo + 1):
            cell = worksheet.cell(row=row_index, column=6 + num_periodo)
            periodo = periodos_by_num.get(num_periodo)
            if not periodo:
                cell.fill = locked_fill
                cell.protection = Protection(locked=True)
                continue

            consumo_actual = consumo_no_migrado.get(num_periodo, 0)
            maximo = max(0, periodo["dias_otorgados"] - consumo_actual)
            es_proximo = periodo.get("es_proximo", False)
            cell.value = migrado_por_periodo.get(num_periodo) or None
            cell.number_format = "0"
            cell.protection = Protection(locked=False)
            cell.fill = expired_fill if periodo.get("expirado") else editable_fill
            vencido = periodo.get("expirado")
            cell.comment = Comment(
                f"{'PERIODO VENCIDO. ' if vencido else ''}"
                f"{'ANTICIPO: El empleado aun no cumple aniversario. Captura los dias tomados anticipadamente.\n' if es_proximo else ''}"
                f"Captura aqui cuantos dias tomo este empleado en este periodo ANTES de que el sistema existiera.\n"
                f"Si no tomo ninguno, deja en blanco o pon 0.\n\n"
                f"Dias otorgados en este periodo: {periodo['dias_otorgados']}\n"
                f"{'Sin limite maximo (son dias anticipados)' if es_proximo else f'Maximo que puedes capturar: {maximo}'}\n"
                f"Aniversario del periodo: {periodo['fecha_aniversario']:%d/%m/%Y}\n"
                f"Expira: {periodo['fecha_expiracion']:%d/%m/%Y}",
                "Enertika",
            )

    worksheet.column_dimensions["A"].hidden = True
    for column in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = min(max_len + 2, 48)
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.protection.sheet = True
    worksheet.protection.enable()
    return workbook


async def validar_importacion_migracion(conn, file_bytes: bytes) -> dict:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from zipfile import BadZipFile

    if not file_bytes:
        raise ValueError("El archivo esta vacio")
    if len(file_bytes) > MIGRACION_MAX_FILE_BYTES:
        raise ValueError("El archivo excede el tamano maximo permitido")

    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ValueError("No se pudo leer el archivo Excel") from exc

    worksheet = workbook.active
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_values:
        workbook.close()
        raise ValueError("La plantilla no contiene encabezados")
    headers = [_normalizar_header_excel(value) for value in header_values]
    try:
        usuario_col = headers.index("usuario_id") + 1
    except ValueError as exc:
        workbook.close()
        raise ValueError("La plantilla no contiene la columna usuario_id") from exc

    periodo_cols: list[tuple[int, int]] = []
    for index, header in enumerate(headers, start=1):
        match = re.search(r"\bperiodo\s+(\d+)\b", header)
        if match:
            periodo_cols.append((index, int(match.group(1))))
    if not periodo_cols:
        workbook.close()
        raise ValueError("La plantilla no contiene columnas de periodos")

    raw_rows = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        usuario_id = row[usuario_col - 1] if len(row) >= usuario_col else None
        periodos = []
        errores = []
        tocado = False
        for col_index, num_periodo in periodo_cols:
            value = row[col_index - 1] if len(row) >= col_index else None
            if value is not None and str(value).strip() != "":
                tocado = True
            try:
                dias = _parse_dias_excel(value)
            except ValueError as exc:
                errores.append(f"Periodo {num_periodo}: {exc}")
                continue
            if dias > 0:
                periodos.append({"num_periodo": num_periodo, "dias": dias})
        raw_rows.append({
            "excel_row": row_index,
            "usuario_id": usuario_id,
            "tocado": tocado,
            "periodos": periodos,
            "errores": errores,
        })
    workbook.close()

    filas = await _validar_rows_migracion(conn, raw_rows)
    token_rows = _token_rows_migracion(filas)
    tiene_errores = any(row["errores"] for row in filas)
    total_dias = sum(row["total_dias"] for row in filas if not row["errores"])
    return {
        "filas": filas,
        "tiene_errores": tiene_errores,
        "token": _firmar_preview(token_rows) if token_rows and not tiene_errores else None,
        "total_empleados": len(token_rows),
        "total_dias": total_dias,
    }


async def _get_tipo_vacaciones_id(conn) -> UUID:
    tipos = await vac_db.get_tipos_ausencia(conn)
    for tipo in tipos:
        if tipo["slug"] == "vacaciones":
            return tipo["id"]
    raise ValueError("No existe el tipo de ausencia vacaciones")


async def ejecutar_migracion(conn, token: str, ejecutado_por: UUID) -> dict:
    raw_rows = _leer_preview_firmado(token)
    filas = await _validar_rows_migracion(conn, raw_rows)
    errores = [row for row in filas if row["errores"]]
    if errores:
        raise ValueError("La informacion cambio o contiene errores. Vuelve a importar el archivo.")

    tipo_vacaciones_id = await _get_tipo_vacaciones_id(conn)
    actualizadas = 0
    total_periodos = 0
    total_dias = 0
    async with conn.transaction():
        for row in filas:
            if not row.get("limpiar"):
                continue
            await vac_db.limpiar_migracion_usuario(conn, row["usuario_id"])
            actualizadas += 1
            for periodo in row["periodos"]:
                solicitud_id = await vac_db.insertar_solicitud_migracion(
                    conn,
                    usuario_id=row["usuario_id"],
                    tipo_ausencia_id=tipo_vacaciones_id,
                    fecha_aniversario=periodo["fecha_aniversario"],
                    dias_solicitados=periodo["dias"],
                    num_periodo=periodo["num_periodo"],
                    ejecutado_por=ejecutado_por,
                )
                await vac_db.insert_consumos(
                    conn,
                    solicitud_id,
                    [{
                        "num_periodo": periodo["num_periodo"],
                        "dias_consumir": periodo["dias"],
                        "fecha_aniversario_periodo": periodo["fecha_aniversario"],
                    }],
                )
                total_periodos += 1
                total_dias += periodo["dias"]

    return {
        "empleados_actualizados": actualizadas,
        "periodos_insertados": total_periodos,
        "dias_insertados": total_dias,
    }


async def get_migracion_empleado_ctx(conn, usuario_id: UUID) -> dict:
    usuario = await rrhh_db.get_usuario_simple_by_id(conn, usuario_id)
    empleado = await vac_db.get_empleado_datos(conn, usuario_id)
    if not empleado or not empleado.get("fecha_contratacion"):
        return {
            "usuario": usuario or {},
            "empleado": empleado,
            "periodos": [],
            "aviso": "El empleado no tiene fecha de contratación.",
        }

    hoy = today_mx()
    catalogo = await vac_db.get_catalogo_dias(conn)
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
    consumos_no_migracion = await vac_db.get_consumos_no_migracion_bulk(conn, [usuario_id])
    migracion = await vac_db.get_migracion_usuario(conn, usuario_id)
    migrado_por_periodo = _sumar_consumos_por_periodo(migracion)
    visibles_base, consumo_no_migrado = _periodos_migrables(
        empleado,
        hoy,
        catalogo,
        meses_exp,
        consumos_no_migracion.get(usuario_id, []),
    )
    visibles = []
    for periodo in visibles_base:
        num_periodo = periodo["num_periodo"]
        consumo_actual = consumo_no_migrado.get(num_periodo, 0)
        visibles.append({
            **periodo,
            "dias_migrados": migrado_por_periodo.get(num_periodo, 0),
            "dias_maximos_migracion": max(0, periodo["dias_otorgados"] - consumo_actual),
            "consumo_no_migrado": consumo_actual,
        })

    return {
        "usuario": usuario or {},
        "empleado": empleado,
        "periodos": visibles,
        "total_migrado": sum(migrado_por_periodo.values()),
        "migracion": migracion,
    }


async def guardar_migracion_individual(
    conn,
    usuario_id: UUID,
    periodos_dias: list[dict],
    ejecutado_por: UUID,
) -> dict:
    raw_rows = [{
        "usuario_id": str(usuario_id),
        "limpiar": True,
        "periodos": [
            {"num_periodo": item["num_periodo"], "dias": item["dias"]}
            for item in periodos_dias
        ],
    }]
    filas = await _validar_rows_migracion(conn, raw_rows)
    errores = [error for row in filas for error in row["errores"]]
    if errores:
        raise ValueError("; ".join(errores))
    token = _firmar_preview(_token_rows_migracion(filas))
    return await ejecutar_migracion(conn, token, ejecutado_por)


async def limpiar_migracion_empleado(conn, usuario_id: UUID) -> int:
    async with conn.transaction():
        return await vac_db.limpiar_migracion_usuario(conn, usuario_id)


async def get_admin_ctx(conn, anio: int | None = None) -> dict:
    anio = anio or today_mx().year
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
    he_minimo = await ConfigService.get_global_config(conn, "ASISTENCIA_HE_MINIMO_MINUTOS", 30, int)
    horarios_rows = await rrhh_db.get_horarios_sucursal_admin(conn)
    return {
        "anio": anio,
        "tipos": await vac_db.get_tipos_ausencia_admin(conn),
        "dias_vacaciones": await vac_db.get_catalogo_dias_admin(conn),
        "vacaciones_meses_expiracion": meses_exp,
        "he_minimo_minutos": he_minimo,
        "sucursales": await rrhh_db.get_sucursales_admin(conn),
        "dias_semana": DIAS_SEMANA,
        "horarios_sucursal": _build_horarios_admin(horarios_rows),
    }


def _build_horarios_admin(rows: list[dict]) -> list[dict]:
    horarios: dict[UUID, dict] = {}
    for row in rows:
        horario_id = row["id"]
        horario = horarios.setdefault(
            horario_id,
            {
                "id": horario_id,
                "sucursal_id": row["sucursal_id"],
                "sucursal_nombre": row["sucursal_nombre"],
                "nombre": row["nombre"],
                "activo": row["activo"],
                "margen_entrada_antes_min": row["margen_entrada_antes_min"],
                "margen_salida_despues_min": row["margen_salida_despues_min"],
                "tolerancia_extra_min": row["tolerancia_extra_min"],
                "descuento_comida_min": row["descuento_comida_min"],
                "updated_at": row["updated_at"],
                "dias_by_index": {},
            },
        )
        if row.get("dia_semana") is not None:
            horario["dias_by_index"][row["dia_semana"]] = _format_horario_dia(row)

    result = []
    for horario in horarios.values():
        dias = []
        for dia in DIAS_SEMANA:
            dias.append(
                horario["dias_by_index"].get(
                    dia["value"],
                    {
                        "dia_semana": dia["value"],
                        "nombre": dia["nombre"],
                        "es_laboral": dia["value"] < 5,
                        "hora_entrada": None,
                        "hora_salida": None,
                        "hora_entrada_fmt": "08:00" if dia["value"] < 5 else "",
                        "hora_salida_fmt": "17:00" if dia["value"] < 5 else "",
                        "minutos_programados": 0,
                        "cruza_medianoche": False,
                        "descuento_comida_min": horario["descuento_comida_min"] if dia["value"] < 5 else 0,
                    },
                )
            )
        horario["dias"] = dias
        horario["resumen_dias"] = _resumen_dias_horario(dias)
        del horario["dias_by_index"]
        result.append(horario)
    return result


def _format_horario_dia(row: dict) -> dict:
    dia_semana = row["dia_semana"]
    return {
        "dia_semana": dia_semana,
        "nombre": DIAS_SEMANA[dia_semana]["nombre"],
        "es_laboral": row["es_laboral"],
        "hora_entrada": row["hora_entrada"],
        "hora_salida": row["hora_salida"],
        "hora_entrada_fmt": _format_time(row["hora_entrada"]),
        "hora_salida_fmt": _format_time(row["hora_salida"]),
        "minutos_programados": row["minutos_programados"],
        "cruza_medianoche": row["cruza_medianoche"],
        "descuento_comida_min": row["dia_descuento_comida_min"],
    }


def _format_time(value: time | None) -> str:
    return value.strftime("%H:%M") if value else ""


def _resumen_dias_horario(dias: list[dict]) -> str:
    laborables = [dia for dia in dias if dia["es_laboral"]]
    if not laborables:
        return "Sin dias laborales"
    compact = []
    for dia in laborables:
        if dia["hora_entrada_fmt"] and dia["hora_salida_fmt"]:
            compact.append(f"{dia['nombre'][:3]} {dia['hora_entrada_fmt']}-{dia['hora_salida_fmt']}")
        else:
            compact.append(dia["nombre"][:3])
    return ", ".join(compact)


async def guardar_horario_sucursal(
    conn,
    *,
    sucursal_id: UUID,
    nombre: str,
    activo: bool,
    margen_entrada_antes_min: int,
    margen_salida_despues_min: int,
    tolerancia_extra_min: int,
    descuento_comida_min: int,
    dias: list[dict],
    user_id: UUID,
    horario_id: UUID | None = None,
) -> int:
    nombre = (nombre or "").strip()
    if not nombre:
        raise ValueError("El nombre del horario es obligatorio")
    if len(nombre) > 100:
        raise ValueError("El nombre del horario no puede exceder 100 caracteres")
    sucursales = await rrhh_db.get_sucursales_admin(conn)
    if sucursal_id not in {sucursal["id"] for sucursal in sucursales}:
        raise ValueError("Sucursal no encontrada")

    _validar_minutos_config(
        margen_entrada_antes_min=margen_entrada_antes_min,
        margen_salida_despues_min=margen_salida_despues_min,
        tolerancia_extra_min=tolerancia_extra_min,
        descuento_comida_min=descuento_comida_min,
    )
    dias_normalizados = _normalizar_dias_horario(
        dias,
        descuento_comida_min,
        margen_entrada_antes_min=margen_entrada_antes_min,
        margen_salida_despues_min=margen_salida_despues_min,
    )
    if activo and not any(dia["es_laboral"] for dia in dias_normalizados):
        raise ValueError("Un horario activo debe tener al menos un dia laboral")
    sucursales_recalculo = {sucursal_id}

    async with conn.transaction():
        if horario_id:
            existente = await rrhh_db.get_horario_sucursal(conn, horario_id)
            if not existente:
                raise ValueError("Horario no encontrado")
            sucursales_recalculo.add(existente["sucursal_id"])

        if activo:
            await rrhh_db.deactivate_horarios_sucursal(
                conn, sucursal_id=sucursal_id, exclude_id=horario_id
            )

        if horario_id:
            updated = await rrhh_db.update_horario_sucursal(
                conn,
                horario_id=horario_id,
                sucursal_id=sucursal_id,
                nombre=nombre,
                activo=activo,
                margen_entrada_antes_min=margen_entrada_antes_min,
                margen_salida_despues_min=margen_salida_despues_min,
                tolerancia_extra_min=tolerancia_extra_min,
                descuento_comida_min=descuento_comida_min,
                updated_by=user_id,
            )
            if not updated:
                raise ValueError("Horario no encontrado")
            saved_id = updated["id"]
        else:
            saved_id = await rrhh_db.create_horario_sucursal(
                conn,
                sucursal_id=sucursal_id,
                nombre=nombre,
                activo=activo,
                margen_entrada_antes_min=margen_entrada_antes_min,
                margen_salida_despues_min=margen_salida_despues_min,
                tolerancia_extra_min=tolerancia_extra_min,
                descuento_comida_min=descuento_comida_min,
                updated_by=user_id,
            )

        await rrhh_db.replace_horario_sucursal_dias(conn, saved_id, dias_normalizados)
        return await _recalcular_sucursales_completo(conn, sucursales_recalculo)


async def desactivar_horario_sucursal(conn, horario_id: UUID, user_id: UUID) -> int:
    async with conn.transaction():
        updated = await rrhh_db.deactivate_horario_sucursal(conn, horario_id, user_id)
        if not updated:
            raise ValueError("Horario no encontrado")
        return await _recalcular_sucursales_completo(conn, {updated["sucursal_id"]})


def _validar_minutos_config(
    *,
    margen_entrada_antes_min: int,
    margen_salida_despues_min: int,
    tolerancia_extra_min: int,
    descuento_comida_min: int,
) -> None:
    valores = {
        "ventana antes de entrada": margen_entrada_antes_min,
        "ventana despues de salida": margen_salida_despues_min,
        "tolerancia para horas extra": tolerancia_extra_min,
        "descuento de comida": descuento_comida_min,
    }
    for nombre, valor in valores.items():
        if valor < 0:
            raise ValueError(f"El valor de {nombre} no puede ser negativo")
        if valor > 1440:
            raise ValueError(f"El valor de {nombre} no puede exceder 1440 minutos")


def _normalizar_dias_horario(
    dias: list[dict],
    descuento_comida_min: int,
    *,
    margen_entrada_antes_min: int,
    margen_salida_despues_min: int,
) -> list[dict]:
    if len(dias) != 7:
        raise ValueError("Debes configurar los 7 dias de la semana")
    dias_by_index = {dia["dia_semana"]: dia for dia in dias}
    if set(dias_by_index) != set(range(7)):
        raise ValueError("La configuracion semanal debe incluir lunes a domingo")

    normalizados = []
    for dia_semana in range(7):
        dia = dias_by_index[dia_semana]
        es_laboral = bool(dia.get("es_laboral"))
        if not es_laboral:
            normalizados.append({
                "dia_semana": dia_semana,
                "hora_entrada": None,
                "hora_salida": None,
                "minutos_programados": 0,
                "cruza_medianoche": False,
                "es_laboral": False,
                "descuento_comida_min": 0,
            })
            continue

        entrada = _parse_hora(dia.get("hora_entrada"), f"entrada de {DIAS_SEMANA[dia_semana]['nombre']}")
        salida = _parse_hora(dia.get("hora_salida"), f"salida de {DIAS_SEMANA[dia_semana]['nombre']}")
        cruza_medianoche = _calcular_cruza_medianoche(
            entrada,
            salida,
            DIAS_SEMANA[dia_semana]["nombre"],
        )
        comida_dia = _parse_minutos_dia(
            dia.get("descuento_comida_min"),
            f"comida de {DIAS_SEMANA[dia_semana]['nombre']}",
            descuento_comida_min,
        )
        minutos_programados = _calcular_minutos_programados(entrada, salida, comida_dia)
        normalizados.append({
            "dia_semana": dia_semana,
            "hora_entrada": entrada,
            "hora_salida": salida,
            "minutos_programados": minutos_programados,
            "cruza_medianoche": cruza_medianoche,
            "es_laboral": True,
            "descuento_comida_min": comida_dia,
        })
    _validar_ventanas_no_traslapadas(
        normalizados,
        margen_entrada_antes_min=margen_entrada_antes_min,
        margen_salida_despues_min=margen_salida_despues_min,
    )
    return normalizados


def _parse_hora(value: str | None, campo: str) -> time:
    value = (value or "").strip()
    if not value:
        raise ValueError(f"La hora de {campo} es obligatoria")
    try:
        parts = value.split(":")
        if len(parts) < 2:
            raise ValueError
        hour = int(parts[0])
        minute = int(parts[1])
        return time(hour=hour, minute=minute)
    except ValueError as exc:
        raise ValueError(f"La hora de {campo} no es valida") from exc


def _parse_minutos_dia(value: object, campo: str, default: int) -> int:
    if value is None or value == "":
        valor = default
    else:
        try:
            valor = int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"El valor de {campo} no es valido") from exc
    if valor < 0:
        raise ValueError(f"El valor de {campo} no puede ser negativo")
    if valor > 1440:
        raise ValueError(f"El valor de {campo} no puede exceder 1440 minutos")
    return valor


def _minutos_hora(value: time) -> int:
    return value.hour * 60 + value.minute


def _calcular_cruza_medianoche(entrada: time, salida: time, dia_nombre: str) -> bool:
    entrada_min = _minutos_hora(entrada)
    salida_min = _minutos_hora(salida)
    if salida_min == entrada_min:
        raise ValueError(f"La entrada y salida de {dia_nombre} no pueden ser iguales")
    return salida_min < entrada_min


def _calcular_minutos_programados(
    entrada: time,
    salida: time,
    descuento_comida_min: int,
) -> int:
    entrada_min = _minutos_hora(entrada)
    salida_min = _minutos_hora(salida)
    if salida_min < entrada_min:
        salida_min += MINUTOS_DIA
    duracion = salida_min - entrada_min
    if descuento_comida_min >= duracion:
        raise ValueError("El descuento de comida no puede ser mayor o igual a la jornada")
    return duracion - descuento_comida_min


def _inicio_ventana_semana(dia: dict, dia_offset: int, margen_entrada_antes_min: int) -> int:
    return dia_offset * MINUTOS_DIA + _minutos_hora(dia["hora_entrada"]) - margen_entrada_antes_min


def _fin_ventana_semana(dia: dict, dia_offset: int, margen_salida_despues_min: int) -> int:
    salida_min = _minutos_hora(dia["hora_salida"])
    ajuste_cruce = MINUTOS_DIA if dia["cruza_medianoche"] else 0
    return dia_offset * MINUTOS_DIA + salida_min + ajuste_cruce + margen_salida_despues_min


def _validar_ventanas_no_traslapadas(
    dias: list[dict],
    *,
    margen_entrada_antes_min: int,
    margen_salida_despues_min: int,
) -> None:
    dias_by_index = {dia["dia_semana"]: dia for dia in dias if dia["es_laboral"]}
    for dia_semana in range(7):
        actual = dias_by_index.get(dia_semana)
        siguiente_idx = (dia_semana + 1) % 7
        siguiente = dias_by_index.get(siguiente_idx)
        if not actual or not siguiente:
            continue

        fin_actual = _fin_ventana_semana(actual, dia_semana, margen_salida_despues_min)
        inicio_siguiente = _inicio_ventana_semana(
            siguiente,
            dia_semana + 1,
            margen_entrada_antes_min,
        )

        if fin_actual >= inicio_siguiente:
            dia_actual = DIAS_SEMANA[dia_semana]["nombre"]
            dia_siguiente = DIAS_SEMANA[siguiente_idx]["nombre"]
            logger.warning(
                "Ventanas de horario traslapadas",
                extra={
                    "dia_actual": dia_actual,
                    "dia_siguiente": dia_siguiente,
                    "fin_actual": _format_minuto_semana(fin_actual),
                    "inicio_siguiente": _format_minuto_semana(inicio_siguiente),
                    "hora_entrada_actual": actual["hora_entrada"].isoformat(timespec="minutes"),
                    "hora_salida_actual": actual["hora_salida"].isoformat(timespec="minutes"),
                    "hora_entrada_siguiente": siguiente["hora_entrada"].isoformat(timespec="minutes"),
                    "hora_salida_siguiente": siguiente["hora_salida"].isoformat(timespec="minutes"),
                    "margen_entrada_antes_min": margen_entrada_antes_min,
                    "margen_salida_despues_min": margen_salida_despues_min,
                },
            )
            raise ValueError(
                f"La ventana del {dia_actual} se cruza con la ventana del {dia_siguiente}. "
                "Reduce el margen de salida o el margen de entrada para evitar que una checada "
                "pueda pertenecer a dos dias."
            )


def _format_minuto_semana(minutos_abs: int) -> str:
    dia_offset = minutos_abs // MINUTOS_DIA
    minuto_dia = minutos_abs % MINUTOS_DIA
    dia = DIAS_SEMANA[dia_offset % 7]["nombre"]
    suffix = " siguiente" if dia_offset >= 7 else ""
    return f"{dia}{suffix} {minuto_dia // 60:02d}:{minuto_dia % 60:02d}"


async def _recalcular_sucursales_completo(conn, sucursal_ids: set[UUID]) -> int:
    usuarios = await rrhh_db.get_usuarios_asistencia_por_sucursales(conn, list(sucursal_ids))
    if not usuarios:
        return 0
    bounds = await asistencia_db.get_recalculo_bounds(conn, usuarios)
    if not bounds:
        return 0
    fecha_inicio = bounds["fecha_inicio"]
    fecha_fin = max(bounds["fecha_fin"], today_mx())
    targets = [
        (usuario_id, fecha_inicio + timedelta(days=offset))
        for usuario_id in usuarios
        for offset in range((fecha_fin - fecha_inicio).days + 1)
    ]
    await recalcular_asistencia(conn, targets)
    return len(targets)


async def guardar_config_vacaciones(conn, *, meses_expiracion: int) -> None:
    if meses_expiracion < 1 or meses_expiracion > 120:
        raise ValueError("Los meses de expiracion deben estar entre 1 y 120")
    await rrhh_db.upsert_vacaciones_meses_expiracion(conn, meses_expiracion)


async def guardar_config_asistencia(conn, *, he_minimo_minutos: int) -> None:
    if he_minimo_minutos < 1 or he_minimo_minutos > 480:
        raise ValueError("El umbral de horas extra debe estar entre 1 y 480 minutos")
    await rrhh_db.upsert_he_minimo_minutos(conn, he_minimo_minutos)


async def get_reportes_ctx(conn) -> dict:
    hoy = today_mx()
    return {
        "fecha_inicio": hoy - timedelta(days=30),
        "fecha_fin": hoy,
        "usuarios": await vac_db.get_usuarios_activos_simples(conn),
        "sucursales": await asistencia_db.get_sucursales(conn),
        "estados_asistencia": ASISTENCIA_ESTADOS,
        "estados_asistencia_labels": ASISTENCIA_ESTADO_LABELS,
        "estados_vacaciones": list(ESTADOS_SOLICITUD),
    }


def validar_rango_reportes(fecha_inicio: date, fecha_fin: date, *, max_dias: int = 92) -> None:
    if fecha_fin < fecha_inicio:
        raise ValueError("La fecha final no puede ser menor que la inicial")
    if (fecha_fin - fecha_inicio).days > max_dias:
        raise ValueError(f"El rango maximo permitido es de {max_dias} dias")


async def get_reporte_vacaciones(
    conn,
    *,
    fecha_inicio: date,
    fecha_fin: date,
    usuario_ids: list[UUID] | None = None,
    estado: str | None = None,
    incluir_dados_de_baja: bool = False,
) -> list[dict]:
    return await rrhh_db.get_reporte_vacaciones(
        conn,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        usuario_ids=usuario_ids,
        estado=estado,
        incluir_dados_de_baja=incluir_dados_de_baja,
    )


async def get_vacaciones_aprobadas(
    conn,
    *,
    fecha_desde: date,
    usuario_ids: list[UUID] | None = None,
    incluir_dados_de_baja: bool = False,
) -> list[dict]:
    return await rrhh_db.get_vacaciones_aprobadas(
        conn,
        fecha_desde=fecha_desde,
        usuario_ids=usuario_ids,
        incluir_dados_de_baja=incluir_dados_de_baja,
    )


async def build_empleados_vacaciones_export(
    conn,
    sucursal_ids: list[UUID] | None = None,
    usuario_ids: list[UUID] | None = None,
    incluir_dados_de_baja: bool = False,
) -> tuple[list[str], list[list], str]:
    hoy = today_mx()
    empleados = await vac_db.get_all_empleados_con_datos(
        conn,
        limit=10000,
        offset=0,
        sucursal_ids=sucursal_ids,
        usuario_ids=usuario_ids,
        incluir_dados_de_baja=incluir_dados_de_baja,
    )
    catalogo = await vac_db.get_catalogo_dias(conn)
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
    ids_para_bulk = [emp["id_usuario"] for emp in empleados]
    consumos_bulk = await vac_db.get_consumos_bulk(conn, ids_para_bulk)
    prorrogas_bulk = await vac_db.get_prorrogas_activas_bulk(conn, ids_para_bulk)

    headers = [
        "Empleado", "Email", "No. empleado", "Departamento", "Fecha contratacion",
        "Periodo", "Dias otorgados", "Dias tomados", "Dias restantes",
        "Fecha expiracion", "Dias para expirar", "Aprobador",
    ]
    rows = []
    for emp in empleados:
        uid = emp["id_usuario"]
        fecha_contratacion = emp.get("fecha_contratacion")
        base_row = [
            emp["nombre"],
            emp["email"],
            emp.get("numero_empleado"),
            emp.get("departamento") or emp.get("department"),
            fecha_contratacion,
        ]
        if not fecha_contratacion:
            rows.append(base_row + ["", "", "", "", "", "", emp.get("aprobador_nombre")])
            continue

        periodos = calcular_periodos(
            fecha_contratacion,
            hoy,
            catalogo,
            ajuste_dias=emp.get("dias_vacaciones_ajuste") or 0,
            meses_expiracion=meses_exp,
        )
        balance = calcular_balance(periodos, consumos_bulk.get(uid, []), prorrogas=prorrogas_bulk.get(uid, []))
        activos = [periodo for periodo in balance if not periodo.get("es_proximo")]
        if not activos:
            rows.append(base_row + ["", "", "", "", "", "", emp.get("aprobador_nombre")])
            continue

        for periodo in activos:
            rows.append(base_row + [
                periodo.get("num_periodo"),
                periodo.get("dias_otorgados"),
                periodo.get("dias_usados"),
                periodo.get("dias_restantes"),
                periodo.get("fecha_expiracion"),
                periodo.get("dias_para_expiracion"),
                emp.get("aprobador_nombre"),
            ])

    filename = f"empleados_vacaciones_{hoy.strftime('%Y%m%d')}.xlsx"
    return headers, rows, filename


async def generar_festivos_anio(conn, anio: int, user_id: UUID | None = None) -> int:
    _validar_anio_festivos(anio)
    async with conn.transaction():
        insertados = await vac_db.insert_festivos_generados(
            conn,
            generar_feriados_mexico(anio),
            created_by=user_id,
        )
        await vac_db.mark_festivos_validacion_pendiente(conn, anio, updated_by=user_id)
        return insertados


async def get_festivos_ctx(conn, anio: int | None = None) -> dict:
    anio_final = anio or today_mx().year
    _validar_anio_festivos(anio_final)
    validacion = await vac_db.get_festivos_validacion(conn, anio_final)
    if not validacion:
        validacion = {
            "anio": anio_final,
            "estado": "pendiente",
            "notas": None,
            "validado_at": None,
            "validado_by": None,
            "validado_por_nombre": None,
        }
    return {
        "anio": anio_final,
        "festivos": await vac_db.get_festivos_by_year(conn, anio_final),
        "validacion": validacion,
        "anio_min": FESTIVOS_ANIO_MIN,
        "anio_max": FESTIVOS_ANIO_MAX,
    }


async def guardar_festivo(
    conn,
    *,
    fecha,
    descripcion: str,
    es_oficial: bool,
    user_id: UUID,
    festivo_id: UUID | None = None,
) -> None:
    descripcion = (descripcion or "").strip()
    if not descripcion:
        raise ValueError("La descripción es obligatoria")
    _validar_anio_festivos(fecha.year)
    async with conn.transaction():
        anios_pendientes = {fecha.year}
        if festivo_id:
            actual = await vac_db.get_festivo_by_id(conn, festivo_id)
            if not actual:
                raise ValueError("Festivo no encontrado")
            anios_pendientes.add(actual["fecha"].year)
            updated = await vac_db.update_festivo(
                conn, festivo_id, fecha, descripcion, es_oficial, user_id
            )
            if not updated:
                raise ValueError("Festivo no encontrado")
        else:
            await vac_db.create_festivo(conn, fecha, descripcion, es_oficial, user_id)

        for anio in anios_pendientes:
            await vac_db.mark_festivos_validacion_pendiente(conn, anio, updated_by=user_id)


async def eliminar_festivo(conn, festivo_id: UUID, anio: int, user_id: UUID) -> None:
    _validar_anio_festivos(anio)
    async with conn.transaction():
        deleted = await vac_db.delete_festivo(conn, festivo_id)
        if not deleted:
            raise ValueError("Festivo no encontrado")
        await vac_db.mark_festivos_validacion_pendiente(conn, anio, updated_by=user_id)


async def validar_festivos_anio(conn, anio: int, notas: str | None, user_id: UUID) -> None:
    _validar_anio_festivos(anio)
    notas_clean = (notas or "").strip() or None
    await vac_db.validar_festivos_anio(conn, anio, notas_clean, user_id)


async def ensure_festivos_anio_worker(conn, anio: int) -> int:
    _validar_anio_festivos(anio)
    validacion = await vac_db.get_festivos_validacion(conn, anio)
    if validacion and validacion.get("estado") == "validado":
        return 0
    existentes = await vac_db.get_festivos_by_year(conn, anio)
    if existentes:
        if not validacion:
            await vac_db.mark_festivos_validacion_pendiente(conn, anio, updated_by=None)
        return 0
    return await generar_festivos_anio(conn, anio, user_id=None)


def _validar_anio_festivos(anio: int) -> None:
    if anio < FESTIVOS_ANIO_MIN or anio > FESTIVOS_ANIO_MAX:
        raise ValueError(f"El año debe estar entre {FESTIVOS_ANIO_MIN} y {FESTIVOS_ANIO_MAX}")


def _normalizar_slug(slug: str) -> str:
    value = (slug or "").strip().lower()
    value = re.sub(r"[^a-z0-9_]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        raise ValueError("El slug es obligatorio")
    if len(value) > 30:
        raise ValueError("El slug no puede exceder 30 caracteres")
    return value


async def crear_tipo_ausencia(
    conn,
    *,
    nombre: str,
    slug: str,
    abreviatura: str,
    afecta_saldo: bool,
    requiere_aprobacion: bool,
    is_active: bool,
    orden: int,
    user_id: UUID,
) -> None:
    nombre = (nombre or "").strip()
    abreviatura = (abreviatura or "").strip().upper()
    if not nombre:
        raise ValueError("El nombre es obligatorio")
    if not abreviatura or len(abreviatura) > 5:
        raise ValueError("La abreviatura es obligatoria y debe tener maximo 5 caracteres")
    await vac_db.create_tipo_ausencia(
        conn,
        nombre=nombre,
        slug=_normalizar_slug(slug),
        abreviatura=abreviatura,
        afecta_saldo=afecta_saldo,
        requiere_aprobacion=requiere_aprobacion,
        is_active=is_active,
        orden=orden,
        updated_by=user_id,
    )


async def actualizar_tipo_ausencia(
    conn,
    *,
    tipo_id: UUID,
    nombre: str,
    abreviatura: str,
    afecta_saldo: bool,
    requiere_aprobacion: bool,
    is_active: bool,
    orden: int,
    user_id: UUID,
) -> None:
    tipo = await vac_db.get_tipo_ausencia_admin_by_id(conn, tipo_id)
    if not tipo:
        raise ValueError("Tipo de permiso no encontrado")
    if tipo["es_sistema"] and not is_active:
        raise ValueError("Los tipos base del sistema no se pueden desactivar")
    nombre = (nombre or "").strip()
    abreviatura = (abreviatura or "").strip().upper()
    if not nombre:
        raise ValueError("El nombre es obligatorio")
    if not abreviatura or len(abreviatura) > 5:
        raise ValueError("La abreviatura es obligatoria y debe tener maximo 5 caracteres")
    await vac_db.update_tipo_ausencia(
        conn,
        tipo_id=tipo_id,
        nombre=nombre,
        abreviatura=abreviatura,
        afecta_saldo=afecta_saldo,
        requiere_aprobacion=requiere_aprobacion,
        is_active=is_active,
        orden=orden,
        updated_by=user_id,
    )


async def guardar_dias_vacaciones(
    conn,
    *,
    antiguedad_anios: int,
    antiguedad_anios_fin: int | None,
    dias_lft: int,
    dias_enertika: int,
    is_active: bool,
    user_id: UUID,
    row_id: UUID | None = None,
) -> None:
    _validar_rango_dias(antiguedad_anios, antiguedad_anios_fin, dias_lft, dias_enertika)
    await _validar_no_solapa_rango(
        conn, antiguedad_anios, antiguedad_anios_fin, is_active, excluir_id=row_id
    )
    if row_id:
        updated = await vac_db.update_dias_vacaciones(
            conn,
            row_id=row_id,
            antiguedad_anios=antiguedad_anios,
            antiguedad_anios_fin=antiguedad_anios_fin,
            dias_lft=dias_lft,
            dias_enertika=dias_enertika,
            is_active=is_active,
            updated_by=user_id,
        )
        if not updated:
            raise ValueError("Rango de antiguedad no encontrado")
    else:
        await vac_db.create_dias_vacaciones(
            conn,
            antiguedad_anios=antiguedad_anios,
            antiguedad_anios_fin=antiguedad_anios_fin,
            dias_lft=dias_lft,
            dias_enertika=dias_enertika,
            is_active=is_active,
            updated_by=user_id,
        )


def _validar_rango_dias(
    antiguedad_anios: int,
    antiguedad_anios_fin: int | None,
    dias_lft: int,
    dias_enertika: int,
) -> None:
    if antiguedad_anios <= 0:
        raise ValueError("La antiguedad inicial debe ser mayor a cero")
    if antiguedad_anios_fin is not None and antiguedad_anios_fin < antiguedad_anios:
        raise ValueError("La antiguedad final no puede ser menor a la inicial")
    if dias_lft <= 0 or dias_enertika <= 0:
        raise ValueError("Los dias deben ser mayores a cero")
    if dias_enertika < dias_lft:
        raise ValueError("Los dias Enertika no pueden ser menores que LFT")


async def _validar_no_solapa_rango(
    conn,
    inicio: int,
    fin: int | None,
    is_active: bool,
    excluir_id: UUID | None = None,
) -> None:
    if not is_active:
        return
    rows = await vac_db.get_catalogo_dias_admin(conn)
    nuevo_fin = fin if fin is not None else 999
    for row in rows:
        if not row["is_active"] or row["id"] == excluir_id:
            continue
        row_inicio = row["antiguedad_anios"]
        row_fin = row["antiguedad_anios_fin"] if row["antiguedad_anios_fin"] is not None else 999
        if inicio <= row_fin and nuevo_fin >= row_inicio:
            raise ValueError("El rango de antiguedad se empalma con otro rango activo")


async def guardar_empleado(
    conn,
    usuario_id: UUID,
    numero_empleado: str | None,
    fecha_contratacion,
    puesto: str | None,
    departamento: str | None,
    id_aprobador_vacaciones: UUID | None,
    dias_vacaciones_ajuste: int | None,
    sucursal_id: UUID | None,
    jefes_ids: list[UUID],
    updated_by: UUID,
) -> None:
    existing = await vac_db.get_empleado_datos(conn, usuario_id)
    old_sucursal_id = existing["sucursal_id"] if existing else None

    await vac_db.upsert_empleado_datos(
        conn,
        usuario_id=usuario_id,
        numero_empleado=numero_empleado,
        fecha_contratacion=fecha_contratacion,
        puesto=puesto,
        departamento=departamento,
        id_aprobador_vacaciones=id_aprobador_vacaciones,
        dias_vacaciones_ajuste=dias_vacaciones_ajuste,
        sucursal_id=sucursal_id,
        updated_by=updated_by,
    )
    await vac_db.set_jefes(conn, usuario_id, jefes_ids)

    if sucursal_id != old_sucursal_id:
        await recalcular_asistencia_reciente_usuario(conn, usuario_id)


# ─────────────────────────────────────────────
# Prórrogas de vacaciones
# ─────────────────────────────────────────────

async def get_prorrogas_empleado_ctx(conn, usuario_id: UUID) -> dict:
    usuario = await rrhh_db.get_usuario_simple_by_id(conn, usuario_id)
    empleado = await vac_db.get_empleado_datos(conn, usuario_id)
    prorrogas = await vac_db.get_prorrogas_usuario(conn, usuario_id)

    periodos_vencidos: list[dict] = []
    if empleado and empleado.get("fecha_contratacion"):
        hoy = today_mx()
        catalogo = await vac_db.get_catalogo_dias(conn)
        meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
        periodos = calcular_periodos(
            empleado["fecha_contratacion"],
            hoy,
            catalogo,
            ajuste_dias=empleado.get("dias_vacaciones_ajuste") or 0,
            meses_expiracion=meses_exp,
        )
        consumos = await vac_db.get_consumos_usuario(conn, usuario_id)
        balance = calcular_balance(periodos, consumos)
        prorrogas_activas = await vac_db.get_prorrogas_activas_usuario(conn, usuario_id)
        con_prorroga_activa = {
            (p["num_periodo"], p["fecha_aniversario_periodo"])
            for p in prorrogas_activas
        }
        periodos_vencidos = [
            {
                **p,
                "ya_tiene_prorroga_activa": (p["num_periodo"], p["fecha_aniversario"]) in con_prorroga_activa,
            }
            for p in balance
            if p.get("expirado") and p["dias_restantes"] > 0
        ]

    return {
        "usuario": usuario or {},
        "empleado": empleado,
        "periodos_vencidos": periodos_vencidos,
        "prorrogas": prorrogas,
    }


async def crear_prorroga_vacaciones(
    conn,
    usuario_id: UUID,
    num_periodo: int,
    fecha_aniversario_periodo: date,
    dias_prorrogados: int,
    fecha_expiracion_prorroga: date,
    motivo: str,
    created_by: UUID,
) -> dict:
    empleado = await vac_db.get_empleado_datos(conn, usuario_id)
    if not empleado or not empleado.get("fecha_contratacion"):
        raise ValueError("El empleado no existe o no tiene fecha de contratación")

    hoy = today_mx()
    catalogo = await vac_db.get_catalogo_dias(conn)
    meses_exp = await ConfigService.get_global_config(conn, "VACACIONES_MESES_EXPIRACION", 18, int)
    periodos = calcular_periodos(
        empleado["fecha_contratacion"],
        hoy,
        catalogo,
        ajuste_dias=empleado.get("dias_vacaciones_ajuste") or 0,
        meses_expiracion=meses_exp,
    )
    consumos = await vac_db.get_consumos_usuario(conn, usuario_id)
    balance = calcular_balance(periodos, consumos)

    periodo = next(
        (p for p in balance
         if p["num_periodo"] == num_periodo
         and p["fecha_aniversario"] == fecha_aniversario_periodo),
        None,
    )
    if not periodo:
        raise ValueError("El período no existe")
    if not periodo.get("expirado"):
        raise ValueError("Solo se pueden prorrogar períodos vencidos")
    if periodo["dias_restantes"] <= 0:
        raise ValueError("El período no tiene saldo disponible para prorrogar")
    if dias_prorrogados <= 0:
        raise ValueError("Los días prorrogados deben ser mayores a cero")
    if dias_prorrogados > periodo["dias_restantes"]:
        raise ValueError(
            f"Los días prorrogados ({dias_prorrogados}) no pueden exceder "
            f"el saldo vencido ({periodo['dias_restantes']})"
        )
    if fecha_expiracion_prorroga <= hoy:
        raise ValueError("La fecha límite de la prórroga debe ser futura")
    if fecha_expiracion_prorroga <= periodo["fecha_expiracion"]:
        raise ValueError("La fecha límite debe ser posterior a la expiración original del período")

    motivo = (motivo or "").strip()
    if not motivo:
        raise ValueError("El motivo es obligatorio")

    try:
        return await vac_db.create_prorroga(
            conn,
            usuario_id=usuario_id,
            num_periodo=num_periodo,
            fecha_aniversario_periodo=fecha_aniversario_periodo,
            fecha_expiracion_original=periodo["fecha_expiracion"],
            fecha_expiracion_prorroga=fecha_expiracion_prorroga,
            dias_prorrogados=dias_prorrogados,
            motivo=motivo,
            created_by=created_by,
        )
    except asyncpg.UniqueViolationError as exc:
        raise ValueError("Ya existe una prórroga activa para este período") from exc


async def cancelar_prorroga_vacaciones(
    conn,
    prorroga_id: UUID,
    motivo_cancelacion: str,
    cancelled_by: UUID,
) -> dict:
    motivo_cancelacion = (motivo_cancelacion or "").strip()
    if not motivo_cancelacion:
        raise ValueError("El motivo de cancelación es obligatorio")
    result = await vac_db.cancel_prorroga(conn, prorroga_id, cancelled_by, motivo_cancelacion)
    if not result:
        raise ValueError("Prórroga no encontrada o ya cancelada")
    return result
