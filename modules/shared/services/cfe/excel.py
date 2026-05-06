import calendar
import io
import logging
import math
from collections.abc import Sequence
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .extractor import MAX_XML_SIZE_BYTES, extraer_datos_xml
from .profiles import obtener_perfil_cfe
from .schemas import CfeExcelModo, CfeExcelProfile, CfeReceipt, CfeXmlInput


logger = logging.getLogger("SharedCfeExcelService")

MESES_NUMERO = {
    "ENE": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
}

CAMPOS_ENTEROS = {
    "consumo",
    "consumo_base",
    "consumo_intermedio",
    "consumo_punta",
    "potencia_base",
    "potencia_intermedia",
    "potencia_punta",
    "dias",
    "kw_cap",
    "kw_dist",
    "kwmax",
    "reactiva",
}

CAMPOS_PESOS = {
    "coste_energia_base",
    "coste_energia_intermedia",
    "coste_energia_punta",
    "transmision",
    "coste_distribucion",
    "coste_capacidad",
    "scnmem",
    "suministro",
    "cenace",
    "dos_por_ciento",
    "penalizacion_fp",
    "total",
}

CAMPOS_CALCULADOS_FORMULA = {"consumo", "kw_cap", "kw_dist"}


@dataclass(frozen=True)
class ResultadoValidacionXml:
    archivo: str
    estatus: str
    mensaje: str


async def generar_excel_cfe_desde_uploads(
    files: Sequence[Any],
    perfil_slug: str,
    modo_calculo: str = CfeExcelModo.CALCULADO.value,
) -> io.BytesIO:
    xml_inputs = await _leer_archivos_xml(files)
    return generar_excel_cfe(xml_inputs, perfil_slug, modo_calculo)


def generar_excel_cfe(
    xml_inputs: Sequence[CfeXmlInput],
    perfil_slug: str,
    modo_calculo: str = CfeExcelModo.CALCULADO.value,
) -> io.BytesIO:
    modo = _validar_modo_calculo(modo_calculo)
    perfil = obtener_perfil_cfe(perfil_slug)

    datos = []
    validaciones: list[ResultadoValidacionXml] = []
    for xml_input in xml_inputs:
        try:
            recibo = extraer_datos_xml(xml_input.content, xml_input.filename)
        except ValueError as exc:
            validaciones.append(
                ResultadoValidacionXml(
                    archivo=xml_input.filename,
                    estatus="Omitido",
                    mensaje=_mensaje_validacion_xml(exc),
                )
            )
            logger.warning(
                "cfe_xml_omitido archivo=%s motivo=%s",
                xml_input.filename,
                exc,
            )
            continue

        datos.append(recibo)
        validaciones.append(
            ResultadoValidacionXml(
                archivo=xml_input.filename,
                estatus="Procesado",
                mensaje="XML de CFE válido",
            )
        )

    if not datos:
        omitidos = "; ".join(
            f"{validacion.archivo}: {validacion.mensaje}"
            for validacion in validaciones
        )
        detalle = f" Detalle: {omitidos}" if omitidos else ""
        raise ValueError(f"No se encontraron XML de CFE válidos.{detalle}")

    datos.sort(key=lambda item: item["cfdi"]["fecha_emision"])
    grupos = _agrupar_por_servicio(datos)
    logger.info(
        "cfe_excel_generacion perfil=%s modo=%s recibos=%s servicios=%s",
        perfil.slug,
        modo.value,
        len(datos),
        len(grupos),
    )

    wb = Workbook()
    hoja_inicial = wb.active
    nombres_usados: set[str] = set()

    for index, servicio in enumerate(sorted(grupos)):
        recibos = grupos[servicio]
        ws = hoja_inicial if index == 0 else wb.create_sheet()
        ws.title = _nombre_hoja(servicio, nombres_usados)
        ws.append([col.header for col in perfil.columns])

        for recibo in recibos:
            row_idx = ws.max_row + 1
            fila = _construir_fila(recibo, perfil, modo, row_idx)
            ws.append(fila)

        _aplicar_estilo(ws, perfil)
        logger.info(
            "cfe_excel_hoja_creada perfil=%s servicio=%s hoja=%s recibos=%s",
            perfil.slug,
            servicio,
            ws.title,
            len(recibos),
        )

    if any(validacion.estatus == "Omitido" for validacion in validaciones):
        _agregar_hoja_validacion(wb, validaciones)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _mensaje_validacion_xml(exc: ValueError) -> str:
    mensaje = str(exc)
    if "clsRegArchFact" in mensaje:
        return "No es un XML de recibo CFE"
    if "no parece ser un XML CFDI" in mensaje:
        return "No es un XML CFDI válido"
    return mensaje


async def _leer_archivos_xml(files: Sequence[Any]) -> list[CfeXmlInput]:
    xml_files = [
        file for file in files
        if getattr(file, "filename", None)
        and str(file.filename).lower().endswith(".xml")
    ]

    if not xml_files:
        raise ValueError("No se encontraron archivos XML válidos")

    xml_inputs: list[CfeXmlInput] = []
    for file in xml_files:
        filename = str(file.filename)
        try:
            content = await file.read()
        except OSError as exc:
            raise ValueError(f"No se pudo leer el archivo {filename}: {exc}") from exc

        if len(content) > MAX_XML_SIZE_BYTES:
            raise ValueError(
                f"Archivo {filename} excede el límite de {MAX_XML_SIZE_BYTES // (1024 * 1024)} MB"
            )
        xml_inputs.append(CfeXmlInput(filename=filename, content=content))

    return xml_inputs


def _validar_modo_calculo(modo_calculo: str) -> CfeExcelModo:
    try:
        return CfeExcelModo(modo_calculo)
    except ValueError as exc:
        modos = ", ".join(modo.value for modo in CfeExcelModo)
        raise ValueError(f"Modo de Excel CFE no soportado: {modo_calculo}. Usa: {modos}") from exc


def _valor(
    importes: dict[str, float | None],
    concepto: str,
    observaciones: list[str],
) -> float | None:
    importe = importes.get(concepto)
    if importe is None:
        observaciones.append(f"Falta {concepto}")
        return None
    return importe


def _dias_del_periodo(periodo: dict[str, str], observaciones: list[str]) -> int | None:
    partes = periodo["hasta"].split()
    if len(partes) < 3:
        observaciones.append("Falta FECHASTA para calcular Dias")
        return None

    mes = MESES_NUMERO.get(partes[1].upper())
    if mes is None:
        observaciones.append(f"Mes no reconocido en FECHASTA: {periodo['hasta']}")
        return None

    try:
        anio = int(partes[2])
    except ValueError:
        observaciones.append(f"Año no reconocido en FECHASTA: {periodo['hasta']}")
        return None

    if anio < 100:
        anio += 2000
    return calendar.monthrange(anio, mes)[1]


def _mes_con_anio(periodo: dict[str, str]) -> str:
    mes = periodo["mes_nombre"]
    partes = periodo["hasta"].split()
    if len(partes) < 3:
        return mes

    anio = partes[2]
    if len(anio) > 2:
        anio = anio[-2:]
    return f"{mes}-{anio}"


def _calcular_kw_cap(
    consumo: float | None,
    potencia_punta: float | None,
    dias: int | None,
) -> int | None:
    if consumo is None or potencia_punta is None or not dias:
        return None
    demanda_promedio = consumo / (24 * dias * 0.57)
    return math.ceil(min(potencia_punta, demanda_promedio))


def _calcular_kw_dist(
    consumo: float | None,
    potencias: list[float | None],
    dias: int | None,
) -> int | None:
    if consumo is None or any(potencia is None for potencia in potencias) or not dias:
        return None
    demanda_promedio = consumo / (24 * dias * 0.57)
    return math.ceil(min(max(potencias), demanda_promedio))


def _calcular_fp(
    consumo: float | None,
    reactiva: float | None,
    fp_directo: float | None,
) -> float | None:
    if fp_directo is not None:
        return fp_directo
    if consumo is None or reactiva is None or (consumo == 0 and reactiva == 0):
        return None
    return consumo / math.sqrt(consumo ** 2 + reactiva ** 2) * 100


def _suma_si_completo(valores: list[float | None]) -> float | None:
    if any(valor is None for valor in valores):
        return None
    return sum(valores)


def _kwmax(
    importes: dict[str, float | None],
    potencias: list[float | None],
    observaciones: list[str],
) -> float | None:
    directo = importes.get("KWMax")
    if directo is not None:
        return directo
    if any(potencia is None for potencia in potencias):
        observaciones.append("Falta KWMax")
        return None
    observaciones.append("Falta KWMax, calculado con potencias")
    return max(potencias)


def _observacion_tarifa(servicio: dict[str, Any]) -> str:
    tarifa = servicio["tarifa"] or "N/D"
    tipo_tarifa = servicio["tipo_tarifa"] or "N/D"
    return f"Tarifa: {tarifa}; Tipo tarifa: {tipo_tarifa}"


def _construir_datos_fila(recibo: CfeReceipt) -> dict[str, Any]:
    importes = {
        linea["concepto"]: linea["importe"]
        for linea in recibo["lineas_excel"]
    }
    observaciones: list[str] = []

    consumo_base = _valor(importes, "kWh base", observaciones)
    consumo_intermedio = _valor(importes, "kWh intermedia", observaciones)
    consumo_punta = _valor(importes, "kWh punta", observaciones)
    consumo = _suma_si_completo([consumo_base, consumo_intermedio, consumo_punta])

    potencia_base = _valor(importes, "kW base", observaciones)
    potencia_intermedia = _valor(importes, "kW intermedia", observaciones)
    potencia_punta = _valor(importes, "kW punta", observaciones)
    potencias = [potencia_base, potencia_intermedia, potencia_punta]

    dias = _dias_del_periodo(recibo["periodo"], observaciones)
    reactiva = _valor(importes, "kVArh", observaciones)
    fp_directo = importes.get("Factor de potencia %")
    fp = _calcular_fp(consumo, reactiva, fp_directo)
    if fp_directo is None and fp is not None:
        observaciones.append("Falta Factor de potencia %, calculado con consumo y reactiva")
    elif fp_directo is None:
        observaciones.append("Falta Factor de potencia %")

    datos = {
        "mes": _mes_con_anio(recibo["periodo"]),
        "consumo": consumo,
        "consumo_base": consumo_base,
        "consumo_intermedio": consumo_intermedio,
        "consumo_punta": consumo_punta,
        "potencia_base": potencia_base,
        "potencia_intermedia": potencia_intermedia,
        "potencia_punta": potencia_punta,
        "dias": dias,
        "kw_cap": _calcular_kw_cap(consumo, potencia_punta, dias),
        "kw_dist": _calcular_kw_dist(consumo, potencias, dias),
        "kwmax": _kwmax(importes, potencias, observaciones),
        "fp": fp,
        "reactiva": reactiva,
        "coste_energia_base": _valor(importes, "Generación B", observaciones),
        "coste_energia_intermedia": _valor(importes, "Generación I", observaciones),
        "coste_energia_punta": _valor(importes, "Generación P", observaciones),
        "transmision": _valor(importes, "Transmisión", observaciones),
        "coste_distribucion": _valor(importes, "Distribución", observaciones),
        "coste_capacidad": _valor(importes, "Capacidad", observaciones),
        "scnmem": _valor(importes, "SCnMEM(1)", observaciones),
        "suministro": _valor(importes, "Suministro", observaciones),
        "cenace": _valor(importes, "CENACE", observaciones),
        "dos_por_ciento": _valor(importes, "2% Baja Tension((3))", observaciones),
        "penalizacion_fp": _valor(
            importes, "Bonificacion Factor de Potencia((3))", observaciones
        ),
        "total": _valor(importes, "Subtotal", observaciones),
    }
    datos["observaciones"] = "; ".join([
        _observacion_tarifa(recibo["servicio"]),
        *observaciones,
    ])

    if observaciones:
        logger.warning(
            "cfe_datos_faltantes archivo=%s servicio=%s mes=%s observaciones=%s",
            recibo["archivo"],
            recibo["servicio"]["rpu"] or "Sin servicio",
            datos["mes"],
            "; ".join(observaciones),
        )

    return datos


def _construir_fila(
    recibo: CfeReceipt,
    perfil: CfeExcelProfile,
    modo: CfeExcelModo,
    row_idx: int,
) -> list[Any]:
    datos = _construir_datos_fila(recibo)
    column_index = {column.key: idx for idx, column in enumerate(perfil.columns, start=1)}
    fila = []
    for column in perfil.columns:
        if modo == CfeExcelModo.FORMULAS and column.key in CAMPOS_CALCULADOS_FORMULA:
            fila.append(_formula_para_campo(column.key, row_idx, column_index) or datos.get(column.key))
        else:
            fila.append(datos.get(column.key))
    return fila


def _formula_para_campo(
    key: str,
    row_idx: int,
    column_index: dict[str, int],
) -> str | None:
    refs = {
        campo: f"{get_column_letter(col_idx)}{row_idx}"
        for campo, col_idx in column_index.items()
    }

    if key == "consumo" and all(
        campo in refs for campo in ("consumo_base", "consumo_intermedio", "consumo_punta")
    ):
        return (
            f"=SUM({refs['consumo_base']},{refs['consumo_intermedio']},"
            f"{refs['consumo_punta']})"
        )
    if key == "kw_cap" and all(
        campo in refs for campo in ("potencia_punta", "consumo", "dias")
    ):
        return (
            f"=ROUNDUP(MIN({refs['potencia_punta']},"
            f"{refs['consumo']}/(24*{refs['dias']}*0.57)),0)"
        )
    if key == "kw_dist" and all(
        campo in refs
        for campo in (
            "potencia_base",
            "potencia_intermedia",
            "potencia_punta",
            "consumo",
            "dias",
        )
    ):
        return (
            f"=ROUNDUP(MIN(MAX({refs['potencia_base']},{refs['potencia_intermedia']},"
            f"{refs['potencia_punta']}),{refs['consumo']}/(24*{refs['dias']}*0.57)),0)"
        )
    return None


def _agrupar_por_servicio(datos: Sequence[CfeReceipt]) -> dict[str, list[CfeReceipt]]:
    grupos: dict[str, list[CfeReceipt]] = {}
    for recibo in datos:
        servicio = recibo["servicio"]["rpu"] or "Sin servicio"
        grupos.setdefault(servicio, []).append(recibo)
    return grupos


def _nombre_hoja(servicio: str, usados: set[str]) -> str:
    invalidos = "[]:*?/\\"
    nombre = "".join("_" if caracter in invalidos else caracter for caracter in servicio).strip()
    nombre = nombre or "Sin servicio"
    nombre = nombre[:31]

    candidato = nombre
    contador = 2
    while candidato in usados:
        sufijo = f"_{contador}"
        candidato = f"{nombre[:31 - len(sufijo)]}{sufijo}"
        contador += 1

    usados.add(candidato)
    return candidato


def _agregar_hoja_validacion(
    wb: Workbook,
    validaciones: Sequence[ResultadoValidacionXml],
) -> None:
    ws = wb.create_sheet("Validacion")
    ws.append(["Archivo", "Estatus", "Mensaje"])

    for validacion in validaciones:
        ws.append([validacion.archivo, validacion.estatus, validacion.mensaje])

    header_fill = PatternFill("solid", fgColor="7F1D1D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin = Side(style="thin", color="FECACA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 48
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _aplicar_estilo(ws, perfil: CfeExcelProfile) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    key_by_column = {
        idx: column.key
        for idx, column in enumerate(perfil.columns, start=1)
    }
    for col_idx in range(2, ws.max_column + 1):
        key = key_by_column.get(col_idx)
        if key == "observaciones":
            for cells in ws.iter_cols(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in cells:
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=False,
                    )
            continue

        if key in CAMPOS_ENTEROS:
            number_format = "#,##0"
        elif key == "fp":
            number_format = '0.00"%"'
        elif key in CAMPOS_PESOS:
            number_format = "$#,##0.00;-$#,##0.00"
        else:
            number_format = "#,##0.00"
        for cells in ws.iter_cols(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in cells:
                cell.number_format = number_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value or ""
        key = key_by_column.get(col_idx)
        if key == "observaciones":
            ws.column_dimensions[col_letter].width = 45
        else:
            ws.column_dimensions[col_letter].width = max(12, min(len(str(header)) + 2, 28))

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
