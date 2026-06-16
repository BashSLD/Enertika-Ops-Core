"""
Router del Módulo Simulación
"""

from fastapi import APIRouter, Request, Depends, Form, HTTPException, Query
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, StreamingResponse, RedirectResponse
import io
from datetime import date
from uuid import UUID
from typing import Optional, List
from decimal import Decimal
import logging
import asyncpg
from dataclasses import asdict

# IMPORTS OBLIGATORIOS para permisos
from core.security import get_current_user_context
from core.permissions import require_module_access, require_manager_access, require_role
from core.config import settings
from core.config_service import ConfigService

from core.database import get_db_connection

# Import del Service Layer
from .service import SimulacionService, get_simulacion_service, resolve_update_permissions
from .db_service import SimulacionDBService, get_db_service
from ..comercial.service import ComercialService # Reusing logic from Comercial
from modules.shared.services import SiteService

from datetime import datetime
from core.timezone import today_mx

# Helper para conversión segura
def _safe_int(val: Optional[str]) -> Optional[int]:
    if not val:
        return None
    try:
        return int(val)
    except ValueError:
        return None
        
def _safe_uuid(val: Optional[str]) -> Optional[UUID]:
    if not val:
        return None
    try:
        return UUID(val)
    except ValueError:
        return None

from .schemas import OportunidadCreateCompleta, DetalleBessCreate, SimulacionUpdate, SitiosBatchUpdate, SimulacionAdicionalItem

# Import Workflow Service (Centralizado)
from core.workflow.service import get_workflow_service

logger = logging.getLogger("SimulacionModule")

templates = Jinja2Templates(directory="templates")
templates.env.globals["DEBUG_MODE"] = settings.DEBUG_MODE

# Registrar filtros de timezone (México)
from core.jinja_filters import register_timezone_filters
register_timezone_filters(templates.env)

router = APIRouter(
    prefix="/simulacion",
    tags=["Módulo Simulación"],
)


# ========================================
# ENDPOINT PRINCIPAL (UI)
# ========================================
@router.api_route("/ui", methods=["GET", "HEAD"], include_in_schema=False)
async def get_simulacion_ui(
    request: Request,
    context = Depends(get_current_user_context),
    service: SimulacionService = Depends(get_simulacion_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("simulacion")
):
    """
    Dashboard principal del módulo simulación con sistema de tabs.
    
    HTMX Detection:
    - Si viene desde sidebar (HTMX): retorna solo tabs.html (contenido)
    - Si es carga directa (F5/URL): retorna dashboard.html (wrapper completo)
    """
    # HTMX Detection
    # HX-History-Restore-Request: HTMX lo envía al restaurar historial (Back/Forward) — retornar full page
    is_htmx = request.headers.get("hx-request")
    is_history_restore = request.headers.get("hx-history-restore-request")
    if is_htmx and not is_history_restore:
        template = "simulacion/tabs.html"  # Solo contenido
    else:
        template = "simulacion/dashboard.html"  # Wrapper completo
    

    # Logic to determine effective role for UI
    effective_role = "viewer"
    if context.get("role") == "ADMIN":
         effective_role = "admin"
    else:
        effective_role = context.get("module_roles", {}).get("simulacion", "viewer")

    return templates.TemplateResponse(request, template, {"user_name": context.get("user_name"),
        "role": context.get("role"),
        "module_roles": context.get("module_roles", {}),
        "current_module_role": effective_role,
        "catalogos": await service.get_catalogos_ui(conn)
    })

# ========================================
# FORMULARIO EXTRAORDINARIO (Solo ADMIN/MANAGER)
# ========================================
@router.api_route("/form-extraordinario", methods=["GET", "HEAD"], include_in_schema=False)
async def get_form_extraordinario(
    request: Request,
    context = Depends(get_current_user_context),
    service: SimulacionService = Depends(get_simulacion_service),
    conn = Depends(get_db_connection),
    _ = require_manager_access("simulacion")
):
    """
    Formulario para registro extraordinario de oportunidades.
    Solo accesible para ADMIN y MANAGER.
    """
    # Validación de permiso: Delegada a require_manager_access
    role = context.get("role")

    
    # Cargar catálogos para el formulario
    catalogos = await service.get_catalogos_ui(conn)
    canal_default = service.get_canal_from_user_name(context.get("user_name", ""))
    
    return templates.TemplateResponse(request, "shared/forms/oportunidad_form.html", {"catalogos": catalogos,
        "canal_default": canal_default,
        "user_name": context.get("user_name"),
        "role": role,
        "module_roles": context.get("module_roles", {}),
        "post_url": "/simulacion/form-extraordinario",
        "cancel_url": "/simulacion/ui",
        "is_extraordinario": True
    })

@router.post("/form-extraordinario", include_in_schema=False)
async def create_oportunidad_extraordinaria(
    request: Request,
    fecha_manual: str = Form(...),
    cliente_nombre: str = Form(..., min_length=3),
    cliente_id: Optional[UUID] = Form(None),
    nombre_proyecto: str = Form(...),
    id_tecnologia: int = Form(...),
    id_tipo_solicitud: int = Form(...),
    canal_venta: str = Form(...),
    prioridad: str = Form("normal"),
    cantidad_sitios: int = Form(1),
    direccion_obra: str = Form(...),
    google_maps_link: str = Form(...),
    coordenadas_gps: Optional[str] = Form(None),
    sharepoint_folder_url: Optional[str] = Form(None),
    solicitado_por_id: Optional[UUID] = Form(None),
    es_licitacion: bool = Form(False),
    
    # Campos BESS
    bess_uso_sistema: List[str] = Form([]),
    bess_cargas_criticas: Optional[float] = Form(None),
    bess_voltaje: Optional[str] = Form(None),
    bess_autonomia: Optional[str] = Form(None),
    bess_tiene_motores: bool = Form(False),
    bess_potencia_motor: Optional[float] = Form(None),
    bess_cargas_separadas: bool = Form(False),
    bess_planta_emergencia: bool = Form(False),
    
    # Dependencies
    context = Depends(get_current_user_context),
    service: SimulacionService = Depends(get_simulacion_service),
    conn = Depends(get_db_connection),
    _ = require_manager_access("simulacion")
):
    """
    Procesa el formulario extraordinario y crea la oportunidad.
    Solo accesible para ADMIN y MANAGER.
    """
    # Validación de permiso: Delegada a require_manager_access
    
    try:
        # Construir objeto BESS (Reusando lógica centralizada de Comercial)
        detalles_bess = ComercialService.build_bess_detail(
            uso_sistema=bess_uso_sistema,
            cargas_criticas=bess_cargas_criticas,
            tiene_motores=bess_tiene_motores,
            potencia_motor=bess_potencia_motor,
            tiempo_autonomia=bess_autonomia,
            voltaje_operacion=bess_voltaje,
            cargas_separadas=bess_cargas_separadas,
            tiene_planta_emergencia=bess_planta_emergencia
        )
        
        # Crear objeto de datos completo
        datos = OportunidadCreateCompleta(
            fecha_manual_str=fecha_manual,
            cliente_nombre=cliente_nombre,
            cliente_id=cliente_id,
            nombre_proyecto=nombre_proyecto,
            id_tecnologia=id_tecnologia,
            id_tipo_solicitud=id_tipo_solicitud,
            id_estatus_global=1,
            canal_venta=canal_venta,
            prioridad=prioridad,
            cantidad_sitios=cantidad_sitios,
            direccion_obra=direccion_obra,
            google_maps_link=google_maps_link,
            coordenadas_gps=coordenadas_gps or "",
            sharepoint_folder_url=sharepoint_folder_url or "",
            detalles_bess=detalles_bess,
            clasificacion_solicitud="EXTRAORDINARIO",
            solicitado_por_id=solicitado_por_id,
            es_licitacion=es_licitacion
        )
        
        # Crear oportunidad
        new_id, op_id_estandar, es_fuera_horario = await service.crear_oportunidad_transaccional(
            conn, datos, context
        )
        
        # Auto-crear sitio si es unisitio (Para evitar proyectos huérfanos de sitio)
        if cantidad_sitios == 1:
            await SiteService.create_single_site(
                conn, new_id, nombre_proyecto, direccion_obra, google_maps_link, id_tipo_solicitud
            )
        
        target_url = "/simulacion/ui"
        # Params para mostrar alerta en dashboard
        params = f"?new_op={op_id_estandar}&fh={str(es_fuera_horario).lower()}&extraordinaria=1"
        
        from fastapi import Response
        return Response(status_code=200, headers={"HX-Redirect": f"{target_url}{params}"})
        
    except asyncpg.PostgresError as db_err:
        logger.error(f"DB Error creating simulacion op: {db_err}")
        return templates.TemplateResponse(request, "simulacion/partials/messages/error.html", {"title": "Error de Base de Datos",
            "message": "No se pudo crear la oportunidad. Verifique los datos o contacte a soporte."
        })
    except ValueError as val_err:
        return templates.TemplateResponse(request, "simulacion/partials/messages/error.html", {"title": "Datos Inválidos",
            "message": str(val_err)
        })
    except Exception as e:
        logger.error(f"Unexpected error creating simulacion op: {e}")
        return templates.TemplateResponse(request, "simulacion/partials/messages/error.html", {"title": "Error del Sistema",
            "message": "Ocurrió un error inesperado."
        })

# ========================================
# ENDPOINTS PARCIALES (HTMX)
# ========================================
@router.get("/partials/graphs", include_in_schema=False)
async def get_graphs_partial(
    request: Request,
    filtro_fecha_inicio: Optional[str] = None,
    filtro_fecha_fin: Optional[str] = None,
    filtro_tecnologia_id: Optional[str] = None,
    filtro_responsable_id: Optional[str] = None,
    context = Depends(get_current_user_context),
    conn = Depends(get_db_connection),
    _ = require_module_access("simulacion")
):
    """Partial: Tab de gráficas y reportes interactivos."""
    from .report_service import ReportesSimulacionService, FiltrosReporte, get_reportes_service

    # Instanciar servicio de reportes
    report_service = ReportesSimulacionService()

    today = today_mx()
    
    # Default: Start of current Year
    start_date = today.replace(month=1, day=1)
    end_date = today
    
    # Parse params if present
    if filtro_fecha_inicio:
        try:
            start_date = datetime.strptime(filtro_fecha_inicio, '%Y-%m-%d').date()
        except ValueError:
            pass # Keep default
            
    if filtro_fecha_fin:
        try:
            end_date = datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            pass # Keep default

    id_tecnologia = _safe_int(filtro_tecnologia_id)
    responsable_id = _safe_uuid(filtro_responsable_id)
    
    filtros = FiltrosReporte(
        fecha_inicio=start_date,
        fecha_fin=end_date,
        id_tecnologia=id_tecnologia,
        responsable_id=responsable_id
    )
    
    # Obtener datos para el dashboard
    catalogos = await report_service.get_catalogos_filtros(conn)
    metricas = await report_service.get_metricas_generales(conn, filtros)
    graficas = await report_service.get_datos_graficas(conn, filtros, metricas=metricas)
    
    return templates.TemplateResponse(request, "simulacion/reportes/tabs.html", {"user_name": context.get("user_name"),
        "role": context.get("role"),
        "module_roles": context.get("module_roles", {}),
        "current_module_role": context.get("module_roles", {}).get("simulacion", "viewer"),
        "catalogos": catalogos,
        "metricas": metricas,
        "graficas": {k: asdict(v) for k, v in graficas.items()},
        "filtros_aplicados": {
            "fecha_inicio": filtros.fecha_inicio.isoformat(),
            "fecha_fin": filtros.fecha_fin.isoformat(),
            "id_tecnologia": filtros.id_tecnologia if filtros.id_tecnologia else "",
            "responsable_id": str(filtros.responsable_id) if filtros.responsable_id else ""
        }
    })

@router.get("/exportar-excel", include_in_schema=False)
async def exportar_simulaciones_excel(
    request: Request,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    responsable_id: Optional[str] = None,
    id_tecnologia: Optional[str] = None,
    context = Depends(get_current_user_context),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("simulacion", "editor"),
):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = today_mx()
    fi = today.replace(month=1, day=1)
    ff = today
    try:
        if fecha_inicio:
            fi = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        if fecha_fin:
            ff = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    except ValueError:
        pass

    resp_uuid = _safe_uuid(responsable_id)
    tec_int   = _safe_int(id_tecnologia)

    rows = await db_service.get_simulaciones_para_excel(conn, fi, ff, resp_uuid, tec_int)

    # ── Estilos ──────────────────────────────────────────────────────────────
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)
    AZUL  = "1F4E79"; AZUL_C = "BDD7EE"; VERDE = "C6EFCE"; ROJO = "FFC7CE"
    NARAN = "FFEB9C"; GRIS  = "F2F2F2"; BLANC = "FFFFFF"; AMAR  = "FFF2CC"

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def put(ws, row, col, val, bg=BLANC, bold=False, center=False, txt_color="000000"):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = Font(bold=bold, size=10, color=txt_color)
        c.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True
        )
        c.border = brd
        return c

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulaciones"
    ws.freeze_panes = "A3"

    titulo = f"SIMULACIONES | {fi.strftime('%d/%m/%Y')} – {ff.strftime('%d/%m/%Y')} | Hora México"
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = titulo
    t.fill = fill(AZUL)
    t.font = Font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30

    headers = [
        "OP ID", "Responsable", "Cliente", "Título del Proyecto",
        "Fecha Solicitud", "Deadline Calculado", "Deadline Negociado",
        "Fecha Entrega", "Estatus", "KPI Interno", "KPI Compromiso",
    ]
    widths = [20, 26, 28, 52, 17, 20, 20, 17, 14, 18, 18]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = fill(AZUL)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
        ws.column_dimensions[get_column_letter(col)].width = w

    def fmt_dt(val):
        if val is None:
            return "Pendiente"
        return val.strftime("%Y-%m-%d %H:%M") if hasattr(val, "strftime") else str(val)

    for i, r in enumerate(rows, 3):
        tiene_dn = r["deadline_negociado"] is not None
        bg = AMAR if tiene_dn else (GRIS if i % 2 == 0 else BLANC)
        ki, kc = r["kpi_status_sla_interno"] or "", r["kpi_status_compromiso"] or ""

        put(ws, i, 1,  r["op_id_estandar"],          bg, center=True)
        put(ws, i, 2,  r["responsable"] or "Sin asignar", bg)
        put(ws, i, 3,  r["cliente_nombre"],            bg)
        put(ws, i, 4,  r["titulo_proyecto"],           bg)
        put(ws, i, 5,  fmt_dt(r["fecha_solicitud"]),   bg, center=True)
        put(ws, i, 6,  fmt_dt(r["deadline_calculado"]), bg, center=True)
        put(ws, i, 7,  fmt_dt(r["deadline_negociado"]) if tiene_dn else "—",
            bg, bold=tiene_dn, center=True, txt_color="7B3F00" if tiene_dn else "000000")
        put(ws, i, 8,  fmt_dt(r["fecha_entrega"]),     bg, center=True)
        est_bg = {"Entregado": VERDE, "En Proceso": NARAN, "En Revisión": AZUL_C,
                  "Pendiente": GRIS}.get(r["estatus"], bg)
        put(ws, i, 9,  r["estatus"] or "—",           est_bg, center=True)
        ki_bg = VERDE if "a tiempo" in ki else (ROJO if "tarde" in ki else bg)
        put(ws, i, 10, ki or "—",                     ki_bg, center=True)
        kc_bg = VERDE if "a tiempo" in kc else (ROJO if "tarde" in kc else bg)
        put(ws, i, 11, kc or "—",                     kc_bg, center=True)

    ws.auto_filter.ref = f"A2:K{max(2, 2 + len(rows))}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"Simulaciones_{fi.isoformat()}_{ff.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
    )


@router.get("/partials/cards", include_in_schema=False)
async def get_cards_partial(
    request: Request,
    tab: str = "activos",
    q: Optional[str] = None,
    limit: int = 20,
    page: int = 1,
    subtab: Optional[str] = None,
    filtro_tecnologia_id: Optional[str] = None,
    context = Depends(get_current_user_context),
    service: SimulacionService = Depends(get_simulacion_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("simulacion")
):
    """Partial: Tabla de oportunidades con filtros."""
    f_tecnologia = _safe_int(filtro_tecnologia_id)
    
    result = await service.get_oportunidades_list(
        conn, context, tab=tab, q=q, limit=limit, page=page, subtab=subtab, filtro_tecnologia_id=f_tecnologia
    )

    ops_processed = [{**op, 'es_multisitio': ComercialService.is_originally_multisite(op)} for op in result["items"]]

    return templates.TemplateResponse(request, "simulacion/partials/cards.html", {
        "oportunidades": ops_processed,
        "current_tab": tab,
        "subtab": subtab,
        "limit": result["limit"],
        "page": result["page"],
        "total": result["total"],
        "total_pages": result["total_pages"],
        "context": context,
        "catalogos": await service.get_tecnologias_only(conn),
        "filtro_tecnologia_id": f_tecnologia,
        "base_url": "/simulacion/partials/cards",
        "hx_target": "#sim-tab-content",
        "hx_include": "#active_tab_input, #search-input, .filter-select, #limit-selector",
    })

@router.get("/partials/comentarios/{id_oportunidad}", include_in_schema=False)
async def get_comentarios_partial(
    id_oportunidad: UUID,
    request: Request,
    mode: Optional[str] = None,
    workflow_service = Depends(get_workflow_service),
    conn = Depends(get_db_connection),
    _ = require_module_access("simulacion")
):
    """Partial: Lista de comentarios de simulación."""
    comentarios = await workflow_service.get_historial(conn, id_oportunidad)
    
    total_comentarios = len(comentarios)
    has_more = False
    
    # Modo compacto: muestra últimos 3 con opción de expandir
    if mode == 'compact':
        if total_comentarios > 3:
            comentarios = comentarios[:3]  # Primeros 3 (más recientes)
            has_more = True
        # Si hay 3 o menos, mostrar todos sin botón expandir
    
    # Modo latest: solo el más reciente
    elif mode == 'latest' and comentarios:
        comentarios = [comentarios[0]]
        if total_comentarios > 1:
            has_more = True
    
    # Modo para mostrar solo el último comentario (Historial)
    elif mode == 'last_one' and comentarios:
        comentarios = [comentarios[0]]
        has_more = False  # No mostrar botón "ver más" en historial
            
    return templates.TemplateResponse(request, "shared/partials/comentarios_list.html", {"comentarios": comentarios,
        "mode": mode,
        "has_more": has_more,
        "total_extra": total_comentarios - len(comentarios) if mode == 'compact' else total_comentarios - 1,
        "id_oportunidad": id_oportunidad
    })

@router.post("/comentarios/{id_oportunidad}")
async def create_comentario(
    id_oportunidad: UUID,
    request: Request,
    nuevo_comentario: str = Form(...),
    workflow_service = Depends(get_workflow_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion", "editor") 
):
    """Crea un nuevo comentario y devuelve la lista actualizada."""
    logger.info(f"[ROUTER] Recibido POST comentario para {id_oportunidad}: '{nuevo_comentario[:50]}...'")
    if nuevo_comentario.strip():
        await workflow_service.add_comentario(
            conn, context, id_oportunidad, nuevo_comentario,
            departamento_slug="SIMULACION",
            modulo_origen="simulacion"
        )
    
    # Retornar la lista actualizada con todas las variables necesarias
    comentarios = await workflow_service.get_historial(conn, id_oportunidad)
    return templates.TemplateResponse(request, "shared/partials/comentarios_list.html", {"comentarios": comentarios,
        "mode": None,  # Mostrar todos los comentarios después de crear uno nuevo
        "has_more": False,
        "total_extra": 0,
        "id_oportunidad": id_oportunidad
    })


@router.get("/partials/bess/{id_oportunidad}", include_in_schema=False)
async def get_bess_partial(
    request: Request, 
    id_oportunidad: UUID,
    conn = Depends(get_db_connection), 
    service: SimulacionService = Depends(get_simulacion_service)
):
    """Partial: Detalles técnicos BESS."""
    bess = await service.get_detalles_bess(conn, id_oportunidad)
    
    
    return templates.TemplateResponse(request, "shared/modals/bess_detalle_modal.html", {"bess": bess
    })

@router.get("/partials/sitios/{id_oportunidad}", include_in_schema=False)
async def get_sitios_partial(
    id_oportunidad: UUID,
    request: Request,
    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion")
):
    """Partial: Lista de sitios de la oportunidad."""
    sitios = await service.get_sitios(conn, id_oportunidad)
    
    # Obtener opciones para dropdown (Excluyendo "Ganada")
    status_ids = await service._get_status_ids(conn)
    # Using DB Service
    estatus_options = await db_service.get_estatus_simulacion_dropdown(conn, exclude_id=status_ids["ganada"])
    
    # Logic to determine effective role for UI (Consistent with Main UI)
    effective_role = "viewer"
    if context.get("role") == "ADMIN":
         effective_role = "admin"
    else:
        effective_role = context.get("module_roles", {}).get("simulacion", "viewer")

    # Validar si la oportunidad está en estado terminal (Bloquear edición)
    op = await db_service.get_oportunidad_by_id(conn, id_oportunidad)
    is_locked = False
    if op:
         # Estados terminales: Entregado(4), Cancelado(3), Perdido(5), Ganada(2) - IDs aproximados standard
         # Usamos el mapa de IDs para ser precisos
         if op['id_estatus_global'] in [status_ids.get('entregado'), status_ids.get('cancelado'), status_ids.get('perdido'), status_ids.get('ganada')]:
             is_locked = True

    return templates.TemplateResponse(request, "simulacion/partials/sitios_list.html", {"sitios": sitios,
        "context": context,
        "current_module_role": effective_role, 
        "estatus_options": [dict(r) for r in estatus_options],
        "id_oportunidad": id_oportunidad,
        "is_locked": is_locked # <--- Variable de bloqueo para UI
    })

# ========================================
# MODALES DE DETALLE
# ========================================
@router.get("/modals/detalle/{id_oportunidad}", include_in_schema=False)
async def get_detalle_modal(
    request: Request,
    id_oportunidad: UUID,
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion", "editor")
):
    """Modal de detalle (solo lectura) usando template compartido."""
    
    # 1. Obtener datos
    op = await db_service.get_oportunidad_by_id(conn, id_oportunidad)
    if not op:
         return JSONResponse(status_code=404, content={"message": "Oportunidad no encontrada"})

    # 2. Logic flags (Simulacion usually readonly for comercial actions)
    # But we follow the template requirements
    can_edit_comercial = False 
    can_close_sale = False
    
    return templates.TemplateResponse(request, "shared/modals/detalle_oportunidad_modal.html", {"op": dict(op),
        "can_edit_comercial": can_edit_comercial,
        "can_close_sale": can_close_sale,
        "context": context
    })

# --- ENDPOINTS DE GESTIÓN (MODALES Y UPDATES) ---

def _compose_historial_ctx(
    op,
    estatus_global: list,
    historial_timeline: list,
    umbral_lag_registro_min: int,
    context: dict,
    permisos_update: dict,
    *,
    historial_message: Optional[str] = None,
    historial_error: Optional[str] = None,
) -> dict:
    current_status = next((s for s in estatus_global if s["id"] == op["id_estatus_global"]), None)
    estatus_reversion = [s for s in estatus_global if not s["es_estatus_final"] and s["orden"] in (1, 2, 3, 4)]
    # current_status is None when the op's status was excluded from the dropdown
    # (e.g. 'ganada', which is filtered out but is still a terminal state).
    is_terminal = bool(current_status and current_status["es_estatus_final"]) or current_status is None
    return {
        "op": dict(op),
        "historial_timeline": historial_timeline,
        "umbral_lag_registro_min": umbral_lag_registro_min,
        "estatus_global": [dict(r) for r in estatus_global],
        "estatus_reversion": estatus_reversion,
        "can_reconstruct_history": permisos_update["can_edit_sensitive"],
        "can_reverse_terminal": context.get("role") == "ADMIN" and is_terminal,
        "historial_message": historial_message,
        "historial_error": historial_error,
    }


async def _build_edit_modal_context(
    conn,
    id_oportunidad: UUID,
    service: SimulacionService,
    db_service: SimulacionDBService,
    context: dict,
    form_message: Optional[str] = None,
    historial_message: Optional[str] = None,
    historial_error: Optional[str] = None
) -> Optional[dict]:
    op = await db_service.get_oportunidad_by_id(conn, id_oportunidad)
    if not op:
        return None

    responsables = await service.get_responsables_dropdown(conn)
    status_ids = await service._get_status_ids(conn)
    estatus_global = await db_service.get_estatus_simulacion_dropdown(conn, exclude_id=status_ids["ganada"])
    motivos_cierre = await db_service.get_motivos_cierre(conn)
    sitios_oportunidad = await db_service.get_sitios_by_oportunidad(conn, id_oportunidad)
    motivos_retrabajo = await db_service.get_motivos_retrabajo(conn)
    es_multisitio = ComercialService.is_originally_multisite(dict(op))

    permisos_update = resolve_update_permissions(context)
    current_user_id = str(context.get("user_db_id") or "")
    if permisos_update["can_edit_assignment_fields"] and current_user_id:
        existe_usuario_actual = any(str(resp["id_usuario"]) == current_user_id for resp in responsables)
        if not existe_usuario_actual:
            responsables.append({
                "id_usuario": context["user_db_id"],
                "nombre": context.get("user_name") or "Usuario actual",
                "departamento": context.get("department") or ""
            })

    simulaciones_adicionales = await db_service.get_simulaciones_adicionales(conn, id_oportunidad)
    fv_terminado = await db_service.get_fv_terminado(conn, id_oportunidad)
    historial_timeline = await service.get_historial_timeline(conn, id_oportunidad)
    umbral_lag_registro_min = await ConfigService.get_global_config(conn, "UMBRAL_LAG_NOTIFICACION", 1440, int)

    return {
        **_compose_historial_ctx(op, estatus_global, historial_timeline, umbral_lag_registro_min, context, permisos_update, historial_message=historial_message, historial_error=historial_error),
        "responsables": responsables,
        "motivos_cierre": [dict(r) for r in motivos_cierre],
        "status_ids": status_ids,
        "can_manage": permisos_update["can_manage"],
        "can_edit_sensitive": permisos_update["can_edit_sensitive"],
        "can_edit_assignment_fields": permisos_update["can_edit_assignment_fields"],
        "can_assign_others": permisos_update["can_assign_others"],
        "current_user_id": current_user_id,
        "context": context,
        "sitios_oportunidad": [dict(r) for r in sitios_oportunidad],
        "motivos_retrabajo": [dict(r) for r in motivos_retrabajo],
        "es_multisitio": es_multisitio,
        "is_bess_related": op["id_tecnologia"] in [2, 3],
        "is_bess_only": op["id_tecnologia"] == 2,
        "fv_terminado_fecha": fv_terminado["fecha_entrega"] if fv_terminado else None,
        "simulaciones_adicionales": simulaciones_adicionales,
        "form_message": form_message,
    }


@router.get("/modals/edit/{id_oportunidad}", include_in_schema=False)
async def get_edit_modal(
    request: Request,
    id_oportunidad: UUID,
    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion"),
):
    modal_context = await _build_edit_modal_context(conn, id_oportunidad, service, db_service, context)
    if not modal_context:
        return JSONResponse(status_code=404, content={"message": "Oportunidad no encontrada"})
    return templates.TemplateResponse(request, "simulacion/modals/update_oportunidades.html", modal_context)


@router.post("/fv-terminado/{id_oportunidad}", include_in_schema=False)
async def marcar_fv_terminado(
    request: Request,
    id_oportunidad: UUID,
    service: SimulacionService = Depends(get_simulacion_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion", "editor"),
):
    """Marca la parte FV de un hibrido (FV+BESS) como terminada. Accion independiente del estatus."""
    fecha, _filas = await service.marcar_fv_terminado(conn, id_oportunidad, context)
    msg = f"FV marcado como terminado el {fecha.strftime('%d/%m/%Y %H:%M')}."
    return templates.TemplateResponse(
        request,
        "simulacion/partials/messages/success_inline.html",
        {"message": msg, "id_oportunidad": id_oportunidad},
    )


@router.put("/update/{id_oportunidad}")
async def update_simulacion(
    request: Request,
    id_oportunidad: UUID,
    # Form Data explícito para HTMX
    id_estatus_global: int = Form(...),
    id_interno_simulacion: Optional[str] = Form(None),
    responsable_simulacion_id: Optional[UUID] = Form(None),
    fecha_entrega_simulacion: Optional[str] = Form(None), # Recibe string ISO
    deadline_negociado: Optional[str] = Form(None),       # Recibe string ISO
    id_motivo_cierre: Optional[int] = Form(None),
    monto_cierre_usd: Optional[Decimal] = Form(None),
    potencia_cierre_fv_kwp: Optional[Decimal] = Form(None),
    capacidad_cierre_bess_kwh: Optional[Decimal] = Form(None),
    fecha_cambio_real: Optional[str] = Form(None),          # backdating del cambio de estatus
    # Campos de retrabajo
    es_retrabajo: Optional[bool] = Form(False),
    id_motivo_retrabajo: Optional[int] = Form(None),
    sitios_retrabajo: Optional[str] = Form(None),          # JSON string de UUIDs
    simulaciones_adicionales_json: Optional[str] = Form(None),  # JSON string de SimulacionAdicionalItem

    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion", "editor") 
):
    """Procesa el update del padre con datos de formulario HTMX."""
    # NOTA: No necesitamos notificación aquí, el service se encarga.
    import json
    
    try:
        # Parsear sitios_retrabajo si viene como JSON
        sitios_retrabajo_ids = None
        if sitios_retrabajo:
            try:
                sitios_retrabajo_ids = [UUID(s) for s in json.loads(sitios_retrabajo)]
            except (json.JSONDecodeError, ValueError):
                pass

        # Parsear simulaciones_adicionales si viene como JSON
        sims_adicionales = []
        if simulaciones_adicionales_json:
            try:
                raw_sims = json.loads(simulaciones_adicionales_json)
                sims_adicionales = [SimulacionAdicionalItem(**item) for item in raw_sims if isinstance(item, dict)]
            except (json.JSONDecodeError, ValueError):
                pass

        # Parsear fecha_cambio_real si viene del form
        fecha_cambio_real_dt = None
        if fecha_cambio_real:
            try:
                fecha_cambio_real_dt = datetime.fromisoformat(fecha_cambio_real)
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail="Formato de fecha inválido. Use el selector de fecha del formulario."
                )

        # Reconstruir modelo Pydantic
        datos = SimulacionUpdate(
            id_estatus_global=id_estatus_global,
            id_interno_simulacion=id_interno_simulacion,
            responsable_simulacion_id=responsable_simulacion_id,
            fecha_entrega_simulacion=fecha_entrega_simulacion,
            deadline_negociado=deadline_negociado,
            fecha_cambio_real=fecha_cambio_real_dt,
            id_motivo_cierre=id_motivo_cierre,
            monto_cierre_usd=monto_cierre_usd,
            potencia_cierre_fv_kwp=potencia_cierre_fv_kwp,
            capacidad_cierre_bess_kwh=capacidad_cierre_bess_kwh,
            es_retrabajo=es_retrabajo,
            id_motivo_retrabajo=id_motivo_retrabajo,
            sitios_retrabajo_ids=sitios_retrabajo_ids,
            simulaciones_adicionales=sims_adicionales
        )

        kpi_sla_interno, kpi_compromiso, has_negotiated_deadline, es_cierre_terminal = \
            await service.update_simulacion_padre(conn, id_oportunidad, datos, context)

        # Avance intermedio (no terminal): confirma en #form-feedback y recarga el modal vía JS
        if not es_cierre_terminal:
            return templates.TemplateResponse(
                request,
                "simulacion/partials/messages/success_inline.html",
                {"message": "Estatus actualizado correctamente.", "id_oportunidad": id_oportunidad},
            )

        # Cierre terminal (Entregado/Cancelado/Perdido): cerrar modal y refrescar lista
        show_confetti = False
        if kpi_sla_interno is not None or kpi_compromiso is not None:
            if has_negotiated_deadline:
                show_confetti = (kpi_compromiso == "Entrega a tiempo")
            else:
                show_confetti = (kpi_sla_interno == "Entrega a tiempo")

        redirect_url = "/simulacion/ui"
        if show_confetti:
            redirect_url += "?confetti=1"

        return templates.TemplateResponse(request, "simulacion/partials/messages/success_redirect.html", {"title": "Actualización Exitosa",
            "message": "La oportunidad se ha actualizado correctamente.",
            "redirect_url": redirect_url
        })
    except HTTPException as e:
        # UX IMPROVEMENT: Mostrar errores de validación dentro del modal como mensajes inline
        # para que el usuario los vea en contexto y pueda corregirlos fácilmente
        if e.status_code == 400:
             return templates.TemplateResponse(request, "simulacion/partials/messages/error_inline.html", {"message": e.detail
            }, status_code=200) # Forzamos 200 para que HTMX renderice el contenido
            
        return templates.TemplateResponse(request, "simulacion/partials/messages/error_inline.html", {"message": e.detail
        }, status_code=e.status_code)

async def _build_historial_context(
    conn,
    id_oportunidad: UUID,
    service: SimulacionService,
    db_service: SimulacionDBService,
    context: dict,
    historial_message: Optional[str] = None,
    historial_error: Optional[str] = None,
) -> Optional[dict]:
    op = await db_service.get_oportunidad_by_id(conn, id_oportunidad)
    if not op:
        return None
    status_ids = await service._get_status_ids(conn)
    estatus_global = await db_service.get_estatus_simulacion_dropdown(conn, exclude_id=status_ids["ganada"])
    historial_timeline = await service.get_historial_timeline(conn, id_oportunidad)
    umbral_lag_registro_min = await ConfigService.get_global_config(conn, "UMBRAL_LAG_NOTIFICACION", 1440, int)
    permisos_update = resolve_update_permissions(context)
    return _compose_historial_ctx(
        op, estatus_global, historial_timeline, umbral_lag_registro_min, context, permisos_update,
        historial_message=historial_message, historial_error=historial_error,
    )


async def _render_historial_timeline_partial(
    request: Request,
    conn,
    id_oportunidad: UUID,
    service: SimulacionService,
    db_service: SimulacionDBService,
    context: dict,
    historial_message: Optional[str] = None,
    historial_error: Optional[str] = None,
):
    ctx = await _build_historial_context(
        conn, id_oportunidad, service, db_service, context,
        historial_message=historial_message,
        historial_error=historial_error,
    )
    if not ctx:
        raise HTTPException(status_code=404, detail="Oportunidad no encontrada")
    return templates.TemplateResponse(request, "simulacion/partials/historial_timeline.html", ctx)


@router.post("/historial/{id_oportunidad}/insertar-transicion", include_in_schema=False)
async def insertar_transicion_historica(
    request: Request,
    id_oportunidad: UUID,
    id_estatus: int = Form(...),
    fecha_cambio_real: str = Form(...),
    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_manager_access("simulacion"),
):
    try:
        try:
            fecha_real = datetime.fromisoformat(fecha_cambio_real)
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail="Formato de fecha inválido. Use el selector de fecha del formulario.",
            ) from exc

        await service.insertar_transicion_historica(
            conn,
            id_oportunidad,
            id_estatus,
            fecha_real,
            context,
        )
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_message="Evento histórico insertado correctamente.",
        )
    except HTTPException as exc:
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_error=str(exc.detail),
        )
    except asyncpg.PostgresError:
        logger.exception("Error de base de datos insertando transición histórica %s", id_oportunidad)
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_error="Error de base de datos al insertar el evento histórico.",
        )


@router.post("/historial/{id_oportunidad}/revertir-cierre", include_in_schema=False)
async def revertir_cierre_admin(
    request: Request,
    id_oportunidad: UUID,
    id_estatus_destino: int = Form(...),
    confirmar_reversion: Optional[str] = Form(None),
    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_role(["ADMIN"]),
):
    try:
        if confirmar_reversion != "on":
            raise HTTPException(
                status_code=400,
                detail="Debe confirmar explícitamente la reversión del cierre.",
            )

        await service.revertir_cierre_admin(
            conn,
            id_oportunidad,
            id_estatus_destino,
            context,
        )
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_message="Cierre revertido correctamente.",
        )
    except HTTPException as exc:
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_error=str(exc.detail),
        )
    except asyncpg.PostgresError:
        logger.exception("Error de base de datos revirtiendo cierre %s", id_oportunidad)
        return await _render_historial_timeline_partial(
            request,
            conn,
            id_oportunidad,
            service,
            db_service,
            context,
            historial_error="Error de base de datos al revertir el cierre.",
        )


@router.put("/sitios/batch-update")
async def batch_update_sitios(
    request: Request,
    datos: SitiosBatchUpdate, # FastAPI Pydantic Injection (Handles JSON automatically)
    service: SimulacionService = Depends(get_simulacion_service),
    db_service: SimulacionDBService = Depends(get_db_service),
    conn = Depends(get_db_connection),
    context = Depends(get_current_user_context),
    _ = require_module_access("simulacion", "editor") 
):
    """
    Actualiza múltiples sitios en batch.
    Refactorizado para usar Pydantic + IDOR Check en Service.
    Payload esperado: JSON (hx-ext="json-enc" en frontend).
    """
    try:
        if not datos.ids_sitios:
             # Retorno vacío seguro si no hubo selección
            return templates.TemplateResponse(request, "simulacion/partials/sitios_list.html", {"sitios": [], 
                "context": context
            })

        # Obtener id_oportunidad del primer sitio (Validación de consistencia)
        id_op = await db_service.get_id_oportunidad_from_sitio(conn, datos.ids_sitios[0])
        
        if not id_op:
            raise HTTPException(status_code=404, detail="Sitio no encontrado o sin oportunidad asociada")
        
        # Execute Service (Ahora con validación IDOR interna)
        await service.update_sitios_batch(conn, id_op, datos, context)
        
        # Response (Refresh Table)
        sitios = await service.get_sitios(conn, id_op)
        status_ids = await service._get_status_ids(conn)
        estatus_options = await db_service.get_estatus_simulacion_dropdown(conn, exclude_id=status_ids["ganada"])
        
        # Logic to determine effective role for UI
        effective_role = "viewer"
        if context.get("role") == "ADMIN":
             effective_role = "admin"
        else:
            effective_role = context.get("module_roles", {}).get("simulacion", "viewer")

        # Validar si la oportunidad está en estado terminal (Bloquear edición visualmente)
        op = await db_service.get_oportunidad_by_id(conn, id_op)
        is_locked = False
        if op and op['id_estatus_global'] in [status_ids.get('entregado'), status_ids.get('cancelado'), status_ids.get('perdido'), status_ids.get('ganada')]:
             is_locked = True
            
        return templates.TemplateResponse(request, "simulacion/partials/sitios_list.html", {"sitios": sitios,
            "context": context,
            "current_module_role": effective_role,
            "estatus_options": [dict(r) for r in estatus_options],
            "id_oportunidad": id_op,
            "is_locked": is_locked
        })

    except HTTPException as e:
        # Return error as OOB swap or inline message?
        # For this partial, usually a toast notification is better but we don't have easy toast trigger from here without HX-Trigger header.
        # Let's return a simple error alert replacing the table for now, or just raise.
        # Better: HX-Trigger for toast.
        from fastapi import Response
        return Response(status_code=e.status_code, headers={"HX-Retarget": "#error-container-if-exists", "HX-Reswap": "none"})
    except Exception as e:
        logger.error(f"Error in batch update: {e}")
        raise HTTPException(status_code=500, detail="Error interno del servidor")# The 'require_module_access' dependency actually returns the context if assigned, but here it is assigned to `_`.
        # Let's see if I can add context to args safely or if it's already there implicitly.
        # Looking at original code: `batch_update_sitios` did NOT have `context` in args.
        # I will inject it.
        # But actually, I can just hardcode "editor" or "admin" if I knew, but that's risky.
        # Best way: Add `context = Depends(get_current_user_context)` to the function signature in a separate edit, 
        # OR assume the user has rights since they passed the check. 
        # However, to render the checkboxes again, we need to know if they are still allowed.
        # Since they just did an update, they ARE allowed.
        # But the template checks `current_module_role in ['editor', 'admin']`.
        # So I can pass "editor" as a fallback if I can't get context, but adding context is better.
        
        # NOTE: I am editing the BODY here. I cannot easily change the signature in `replace_file_content` if it spans many lines above.
        # Let's check the signature lines in the file view... 
        # Signature is lines 616-623. I am editing lines 722-733.
        # I cannot access `context` if it's not in args.
        # Workaround: Use `request.state.user` if available, or just pass a flag.
        # Wait, the previous code had `"context": {"role": "ADMIN", "module_role": "editor"},`.
        # This was a HARDCODED fake context!
        # `context={"role": "ADMIN", "module_role": "editor"}`.
        # My new template logic uses `current_module_role`.
        # So I can just pass `current_module_role="editor"` (or "admin") into the template.
        # Since this endpoint is protected by `require_module_access("simulacion", "editor")`, the user is at least an editor.
        # So passing "editor" is safe for the purpose of showing the checkboxes again.
        
        return templates.TemplateResponse(request, "simulacion/partials/sitios_list.html", {"sitios": sitios,
            "context": {"role": "ADMIN"}, # Dummy context to avoid jinja errors if used elsewhere
            "current_module_role": "editor", # Force enable checkboxes after update
            "estatus_options": [dict(r) for r in estatus_options],
            "id_oportunidad": id_op
        })

    except Exception as e:
        logger.error(f"[BATCH UPDATE ERROR] {str(e)}")
        return templates.TemplateResponse(request, "shared/partials/toasts/toast_error.html", {"title": "Error Batch",
            "message": f"Error procesando solicitud: {str(e)}"
        })

@router.patch("/update-responsable/{id_oportunidad}")
async def update_responsable(
    request: Request,
    id_oportunidad: UUID,
    responsable_simulacion_id: UUID = Form(...),
    context = Depends(get_current_user_context),
    conn = Depends(get_db_connection),
    db_service: SimulacionDBService = Depends(get_db_service),
    _ = require_module_access("simulacion", "editor")
):
    """Actualización rápida de responsable (Inline)."""
    try:
        permisos_update = resolve_update_permissions(context)
        can_assign_others = permisos_update["can_assign_others"]
        is_self_assignment = str(responsable_simulacion_id) == str(context.get("user_db_id"))

        if not (can_assign_others or is_self_assignment):
            raise HTTPException(
                status_code=403,
                detail="No autorizado. Solo puedes autoasignarte oportunidades."
            )

        await db_service.update_responsable(conn, id_oportunidad, responsable_simulacion_id)
        return templates.TemplateResponse(request, "shared/partials/toasts/toast_success.html", {"title": "Asignación Actualizada",
            "message": "El responsable ha sido actualizado correctamente."
        })
    except HTTPException as e:
        return templates.TemplateResponse(request, "shared/partials/toasts/toast_error.html", {"title": "Error Asignación",
            "message": e.detail
        }, status_code=e.status_code)
    except asyncpg.PostgresError:
        logger.exception("Error de base de datos al actualizar responsable de simulación")
        return templates.TemplateResponse(request, "shared/partials/toasts/toast_error.html", {"title": "Error Asignación",
            "message": "No se pudo actualizar el responsable."
        }, status_code=500)
