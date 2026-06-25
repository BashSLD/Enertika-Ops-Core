import zipfile
from io import BytesIO
from pathlib import Path
import sys
import types
from uuid import uuid4

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

from modules.cfe.analysis import construir_analisis_recibos
from modules.shared.services.cfe import CfeXmlInput, construir_datos_recibo_cfe, generar_excel_cfe
from modules.shared.services.cfe.extractor import extraer_datos_xml
from modules.shared.services.cfe.profiles import obtener_perfil_cfe


CFE_REAL_DIR = Path("CFE")
CFE_REAL_GDMTO_XML = CFE_REAL_DIR / "JB-000090976983(3).xml"

CFE_XML = b"""<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    Serie="GC" Folio="12345" Fecha="2025-02-28T10:30:00"
    SubTotal="4490.00" Total="5208.40" Moneda="MXN" MetodoPago="PUE" FormaPago="03">
  <cfdi:Emisor Rfc="CFE370814QI0" Nombre="COMISION FEDERAL DE ELECTRICIDAD"/>
  <cfdi:Receptor Rfc="AAA010101AAA" Nombre="CLIENTE PRUEBA" UsoCFDI="G03"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveUnidad="E48" Cantidad="1" Descripcion="Servicio de energia electrica" Importe="4490.00"/>
  </cfdi:Conceptos>
  <cfdi:Impuestos TotalImpuestosTrasladados="718.40"/>
  <cfdi:Complemento>
    <cfe:clsRegArchFact xmlns:cfe="https://www.cfe.mx/cfd/recibo">
      <TARIFA_REG>GDMTH</TARIFA_REG>
      <TARIFA>GDMTH</TARIFA>
      <RPU>123456789012</RPU>
      <FECDESDE>01 FEB 25</FECDESDE>
      <FECHASTA>28 FEB 25</FECHASTA>
      <FECLIMITE>15 MAR 25</FECLIMITE>
      <FECORTE>16 MAR 25</FECORTE>
      <NUMMED1>MED123</NUMMED1>
      <HILOS>3</HILOS>
      <CONSUMO_R>6000</CONSUMO_R>
      <CONSUMO3F>1000</CONSUMO3F>
      <CONSUMO2F>2000</CONSUMO2F>
      <CONSUMO1F>3000</CONSUMO1F>
      <DEMANDA>22</DEMANDA>
      <DEMANDA3P>10</DEMANDA3P>
      <DEMANDA2P>15</DEMANDA2P>
      <DEMANDA1P>20</DEMANDA1P>
      <KVARH>400</KVARH>
      <FacPot>95.5</FacPot>
      <MOTIVO_REG_1>EGB</MOTIVO_REG_1>
      <IMPTE_TOT_REG_1>100.00</IMPTE_TOT_REG_1>
      <MOTIVO_REG_2>EGI</MOTIVO_REG_2>
      <IMPTE_TOT_REG_2>200.00</IMPTE_TOT_REG_2>
      <MOTIVO_REG_3>EGP</MOTIVO_REG_3>
      <IMPTE_TOT_REG_3>300.00</IMPTE_TOT_REG_3>
      <MOTIVO_REG_4>ETB</MOTIVO_REG_4>
      <IMPTE_TOT_REG_4>400.00</IMPTE_TOT_REG_4>
      <MOTIVO_REG_5>ED1</MOTIVO_REG_5>
      <IMPTE_TOT_REG_5>500.00</IMPTE_TOT_REG_5>
      <MOTIVO_REG_6>EID</MOTIVO_REG_6>
      <IMPTE_TOT_REG_6>600.00</IMPTE_TOT_REG_6>
      <MOTIVO_REG_7>EMB</MOTIVO_REG_7>
      <IMPTE_TOT_REG_7>700.00</IMPTE_TOT_REG_7>
      <MOTIVO_REG_8>ES1</MOTIVO_REG_8>
      <IMPTE_TOT_REG_8>800.00</IMPTE_TOT_REG_8>
      <MOTIVO_REG_9>ECB</MOTIVO_REG_9>
      <IMPTE_TOT_REG_9>900.00</IMPTE_TOT_REG_9>
      <Conceptos>
        <Concepto1>2% Baja Tension((3))</Concepto1>
        <Concepto2>Bonificacion Factor de Potencia((3))</Concepto2>
        <Concepto3>Subtotal</Concepto3>
      </Conceptos>
      <Importes>
        <Importe1>20.00</Importe1>
        <Importe2>-30.00</Importe2>
        <Importe3>4490.00</Importe3>
      </Importes>
    </cfe:clsRegArchFact>
  </cfdi:Complemento>
</cfdi:Comprobante>
"""

NON_CFE_XML = b"""<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    xmlns:tfd="http://www.sat.gob.mx/TimbreFiscalDigital"
    Version="4.0" Fecha="2025-03-01T12:00:00" Total="100.00"
    SubTotal="86.21" Moneda="MXN" TipoDeComprobante="I">
  <cfdi:Emisor Rfc="AAA010101AAA" Nombre="Proveedor No CFE"/>
  <cfdi:Receptor Rfc="BBB020202BBB" Nombre="Cliente"/>
  <cfdi:Conceptos>
    <cfdi:Concepto Descripcion="Servicio no CFE" Cantidad="1" ValorUnitario="86.21" Importe="86.21"/>
  </cfdi:Conceptos>
  <cfdi:Complemento>
    <tfd:TimbreFiscalDigital UUID="AAAAAAAA-1111-2222-3333-444444444444" Version="1.1"/>
  </cfdi:Complemento>
</cfdi:Comprobante>
"""


def _cfe_xml_con_metricas(
    *,
    subtotal: str = "4490.00",
    total: str = "5208.40",
    consumo_base: str = "1000",
    consumo_intermedio: str = "2000",
    consumo_punta: str = "3000",
    demanda: str = "22",
    fp: str = "95.5",
) -> bytes:
    xml = CFE_XML
    reemplazos = [
        (b'SubTotal="4490.00"', f'SubTotal="{subtotal}"'.encode()),
        (b'Total="5208.40"', f'Total="{total}"'.encode()),
        (b"<CONSUMO3F>1000</CONSUMO3F>", f"<CONSUMO3F>{consumo_base}</CONSUMO3F>".encode()),
        (b"<CONSUMO2F>2000</CONSUMO2F>", f"<CONSUMO2F>{consumo_intermedio}</CONSUMO2F>".encode()),
        (b"<CONSUMO1F>3000</CONSUMO1F>", f"<CONSUMO1F>{consumo_punta}</CONSUMO1F>".encode()),
        (b"<DEMANDA>22</DEMANDA>", f"<DEMANDA>{demanda}</DEMANDA>".encode()),
        (b"<FacPot>95.5</FacPot>", f"<FacPot>{fp}</FacPot>".encode()),
        (b"<Importe3>4490.00</Importe3>", f"<Importe3>{subtotal}</Importe3>".encode()),
    ]
    for original, nuevo in reemplazos:
        xml = xml.replace(original, nuevo)
    return xml


CFE_GDMTO_XML = b"""<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/4"
    Serie="GC" Folio="67890" Fecha="2026-03-31T10:30:00"
    SubTotal="4500.00" Total="5220.00" Moneda="MXN" MetodoPago="PUE" FormaPago="03">
  <cfdi:Emisor Rfc="CFE370814QI0" Nombre="COMISION FEDERAL DE ELECTRICIDAD"/>
  <cfdi:Receptor Rfc="AAA010101AAA" Nombre="CLIENTE PRUEBA" UsoCFDI="G03"/>
  <cfdi:Conceptos>
    <cfdi:Concepto ClaveUnidad="E48" Cantidad="1" Descripcion="Servicio de energia electrica" Importe="4500.00"/>
  </cfdi:Conceptos>
  <cfdi:Impuestos TotalImpuestosTrasladados="720.00"/>
  <cfdi:Complemento>
    <cfe:clsRegArchFact xmlns:cfe="https://www.cfe.mx/cfd/recibo">
      <TARIFA_REG>GDMTO</TARIFA_REG>
      <TARIFA>GDMTO</TARIFA>
      <RPU>226160800958</RPU>
      <FECDESDE>01 MAR 26</FECDESDE>
      <FECHASTA>31 MAR 26</FECHASTA>
      <FECLIMITE>15 ABR 26</FECLIMITE>
      <FECORTE>16 ABR 26</FECORTE>
      <NUMMED1>MED456</NUMMED1>
      <HILOS>3</HILOS>
      <CONSUMO_R>4500</CONSUMO_R>
      <DEMANDA>28</DEMANDA>
      <DEMANDA_CAPACIDAD>30</DEMANDA_CAPACIDAD>
      <DEMANDA_DISTRIBUCION>31</DEMANDA_DISTRIBUCION>
      <FacPot>97.1</FacPot>
      <MESR01>MAR</MESR01>
      <CONSUMOR01>4500</CONSUMOR01>
      <DEMANDAR01>28</DEMANDAR01>
      <FACPOTR01>97.1</FACPOTR01>
      <FACARR01>42.5</FACARR01>
      <PMVR01>3.14</PMVR01>
      <MESR02>FEB</MESR02>
      <CONSUMOR02>4100</CONSUMOR02>
      <DEMANDAR02>26</DEMANDAR02>
      <FACPOTR02>96.8</FACPOTR02>
      <FACARR02>41.2</FACARR02>
      <PMVR02>3.01</PMVR02>
      <MOTIVO_REG_1>ES1</MOTIVO_REG_1>
      <IMPTE_TOT_REG_1>100.00</IMPTE_TOT_REG_1>
      <MOTIVO_REG_2>ED1</MOTIVO_REG_2>
      <IMPTE_TOT_REG_2>200.00</IMPTE_TOT_REG_2>
      <MOTIVO_REG_3>ETB</MOTIVO_REG_3>
      <IMPTE_TOT_REG_3>300.00</IMPTE_TOT_REG_3>
      <MOTIVO_REG_4>ECB</MOTIVO_REG_4>
      <IMPTE_TOT_REG_4>40.00</IMPTE_TOT_REG_4>
      <MOTIVO_REG_5>EID</MOTIVO_REG_5>
      <IMPTE_TOT_REG_5>50.00</IMPTE_TOT_REG_5>
      <MOTIVO_REG_6>EMB</MOTIVO_REG_6>
      <IMPTE_TOT_REG_6>60.00</IMPTE_TOT_REG_6>
      <MOTIVO_REG_7>EG1</MOTIVO_REG_7>
      <IMPTE_TOT_REG_7>4200.24</IMPTE_TOT_REG_7>
      <MOTIVO_REG_8>X08</MOTIVO_REG_8>
      <IMPTE_TOT_REG_8>8.00</IMPTE_TOT_REG_8>
      <MOTIVO_REG_9>X09</MOTIVO_REG_9>
      <IMPTE_TOT_REG_9>9.00</IMPTE_TOT_REG_9>
      <MOTIVO_REG_10>X10</MOTIVO_REG_10>
      <IMPTE_TOT_REG_10>10.00</IMPTE_TOT_REG_10>
      <MOTIVO_REG_11>X11</MOTIVO_REG_11>
      <IMPTE_TOT_REG_11>11.00</IMPTE_TOT_REG_11>
      <MOTIVO_REG_12>X12</MOTIVO_REG_12>
      <IMPTE_TOT_REG_12>12.00</IMPTE_TOT_REG_12>
      <MOTIVO_REG_13>X13</MOTIVO_REG_13>
      <IMPTE_TOT_REG_13>13.00</IMPTE_TOT_REG_13>
      <MOTIVO_REG_14>X14</MOTIVO_REG_14>
      <IMPTE_TOT_REG_14>14.00</IMPTE_TOT_REG_14>
      <MOTIVO_REG_15>X15</MOTIVO_REG_15>
      <IMPTE_TOT_REG_15>15.00</IMPTE_TOT_REG_15>
      <MOTIVO_REG_16>X16</MOTIVO_REG_16>
      <IMPTE_TOT_REG_16>16.00</IMPTE_TOT_REG_16>
      <MOTIVO_REG_17>X17</MOTIVO_REG_17>
      <IMPTE_TOT_REG_17>17.00</IMPTE_TOT_REG_17>
      <MOTIVO_REG_18>X18</MOTIVO_REG_18>
      <IMPTE_TOT_REG_18>18.00</IMPTE_TOT_REG_18>
      <MOTIVO_REG_19>X19</MOTIVO_REG_19>
      <IMPTE_TOT_REG_19>19.00</IMPTE_TOT_REG_19>
      <MOTIVO_REG_20>ET1</MOTIVO_REG_20>
      <IMPTE_TOT_REG_20>700.00</IMPTE_TOT_REG_20>
      <Conceptos>
        <Concepto1>Subtotal</Concepto1>
        <Concepto2>2% Baja Tension((3))</Concepto2>
        <Concepto3>Bonificacion Factor de Potencia((3))</Concepto3>
        <Concepto4>Cargo 4</Concepto4>
        <Concepto5>Cargo 5</Concepto5>
        <Concepto6>Cargo 6</Concepto6>
        <Concepto7>Cargo 7</Concepto7>
        <Concepto8>Cargo 8</Concepto8>
        <Concepto9>Cargo 9</Concepto9>
        <Concepto10>Cargo 10</Concepto10>
        <Concepto11>Cargo 11</Concepto11>
        <Concepto12>Cargo 12</Concepto12>
        <Concepto13>Cargo 13</Concepto13>
        <Concepto14>Cargo 14</Concepto14>
        <Concepto15>Cargo 15</Concepto15>
        <Concepto16>Cargo 16</Concepto16>
        <Concepto17>Cargo 17</Concepto17>
        <Concepto18>Cargo 18</Concepto18>
        <Concepto19>Cargo 19</Concepto19>
        <Concepto20>Cargo dinamico</Concepto20>
      </Conceptos>
      <Importes>
        <Importe1>4500.00</Importe1>
        <Importe2>20.00</Importe2>
        <Importe3>-10.00</Importe3>
        <Importe4>4.00</Importe4>
        <Importe5>5.00</Importe5>
        <Importe6>6.00</Importe6>
        <Importe7>7.00</Importe7>
        <Importe8>8.00</Importe8>
        <Importe9>9.00</Importe9>
        <Importe10>10.00</Importe10>
        <Importe11>11.00</Importe11>
        <Importe12>12.00</Importe12>
        <Importe13>13.00</Importe13>
        <Importe14>14.00</Importe14>
        <Importe15>15.00</Importe15>
        <Importe16>16.00</Importe16>
        <Importe17>17.00</Importe17>
        <Importe18>18.00</Importe18>
        <Importe19>19.00</Importe19>
        <Importe20>20.00</Importe20>
      </Importes>
    </cfe:clsRegArchFact>
  </cfdi:Complemento>
</cfdi:Comprobante>
"""


def _valor_en_fila(ws, header: str, row: int = 2):
    headers = [cell.value for cell in ws[1]]
    return ws.cell(row=row, column=headers.index(header) + 1).value


def _pdf_bytes():
    return (
        b"%PDF-1.4\n"
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
        b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 72 72] >>\nendobj\n"
        b"xref\n0 4\n"
        b"0000000000 65535 f \n"
        b"0000000009 00000 n \n"
        b"0000000058 00000 n \n"
        b"0000000115 00000 n \n"
        b"trailer\n<< /Size 4 /Root 1 0 R >>\n"
        b"startxref\n186\n%%EOF\n"
    )


class _FakeCfeZipDB:
    def __init__(self, servicio: dict, descargas: list[dict], global_rows: list[dict] | None = None):
        self.servicio = servicio
        self.descargas = descargas
        self.global_rows = global_rows or []

    async def get_servicio_by_id(self, conn, servicio_id):
        return self.servicio if self.servicio["id"] == servicio_id else None

    async def get_descargas_por_servicio(self, conn, servicio_id):
        return self.descargas if self.servicio["id"] == servicio_id else []

    async def get_ultimas_descargas_completadas_por_modulo(
        self, conn, modulos, creado_por_ids=None, servicio_ids=None
    ):
        return self.global_rows


def _install_fake_redis(monkeypatch):
    redis_module = types.ModuleType("redis")
    redis_asyncio_module = types.ModuleType("redis.asyncio")
    redis_exceptions_module = types.ModuleType("redis.exceptions")
    pypdf_module = types.ModuleType("pypdf")

    class FakeRedis:
        pass

    class FakeRedisError(Exception):
        pass

    redis_asyncio_module.Redis = FakeRedis
    redis_asyncio_module.from_url = lambda *args, **kwargs: FakeRedis()
    redis_exceptions_module.RedisError = FakeRedisError
    redis_module.asyncio = redis_asyncio_module
    redis_module.exceptions = redis_exceptions_module

    class FakePyPdfError(Exception):
        pass

    class FakePdfWriter:
        def __init__(self):
            self.pages = []

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def append(self, _content):
            self.pages.append(object())

        def write(self, buffer):
            buffer.write(b"%PDF-1.4\n% merged\n")

    pypdf_module.PdfWriter = FakePdfWriter
    pypdf_module.errors = types.SimpleNamespace(PyPdfError=FakePyPdfError)

    monkeypatch.setitem(sys.modules, "redis", redis_module)
    monkeypatch.setitem(sys.modules, "redis.asyncio", redis_asyncio_module)
    monkeypatch.setitem(sys.modules, "redis.exceptions", redis_exceptions_module)
    monkeypatch.setitem(sys.modules, "pypdf", pypdf_module)


@pytest.mark.asyncio
async def test_generar_zip_servicio_incluye_xml_pdf_y_excel(monkeypatch):
    _install_fake_redis(monkeypatch)
    from modules.cfe.service import CfeService

    servicio_id = uuid4()
    servicio = {
        "id": servicio_id,
        "numero_servicio": "123456789012",
        "nombre": "SERVICIO PRUEBA",
    }
    descargas = [
        {
            "id": uuid4(),
            "periodo": "2026-05",
            "tipo": "xml",
            "estatus": "completado",
            "nombre_archivo": "recibo.xml",
            "ruta_sharepoint": "https://sharepoint.test/recibo.xml",
        },
        {
            "id": uuid4(),
            "periodo": "2026-05",
            "tipo": "pdf",
            "estatus": "completado",
            "nombre_archivo": "recibo.pdf",
            "ruta_sharepoint": "https://sharepoint.test/recibo.pdf",
        },
    ]
    service = CfeService(_FakeCfeZipDB(servicio, descargas))

    async def fake_descargar_archivos(rows):
        contenidos = {"xml": CFE_XML, "pdf": b"%PDF-1.4\n"}
        return [(row, contenidos[row["tipo"]]) for row in rows], []

    monkeypatch.setattr(service, "_descargar_archivos_sharepoint_con_reintentos", fake_descargar_archivos)

    zip_bytes, filename = await service.generar_zip_servicio(None, servicio_id)

    assert filename == "CFE_123456789012.zip"
    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "CFE_123456789012/XML/2026-05_recibo.xml" in names
        assert "CFE_123456789012/PDF/2026-05_recibo.pdf" in names
        assert "CFE_123456789012/CFE_123456789012.xlsx" in names
        assert zf.read("CFE_123456789012/XML/2026-05_recibo.xml") == CFE_XML
        assert zf.read("CFE_123456789012/PDF/2026-05_recibo.pdf") == b"%PDF-1.4\n"

        workbook = load_workbook(BytesIO(zf.read("CFE_123456789012/CFE_123456789012.xlsx")))
        assert workbook.active.max_row == 2


@pytest.mark.asyncio
async def test_descarga_zip_reintenta_hasta_descargar(monkeypatch):
    _install_fake_redis(monkeypatch)
    import modules.cfe.service as cfe_service_module
    from modules.cfe.service import CfeService

    class FakeAuth:
        async def get_application_token(self):
            return "token"

    service = CfeService(_FakeCfeZipDB({"id": uuid4(), "numero_servicio": "1"}, []))
    intentos = {"total": 0}

    async def fake_http(_client, _token, _row):
        intentos["total"] += 1
        if intentos["total"] < 3:
            return None, "fallo temporal"
        return CFE_XML, None

    monkeypatch.setattr(cfe_service_module, "get_ms_auth", lambda: FakeAuth())
    monkeypatch.setattr(service, "_descargar_archivo_sharepoint_http", fake_http)

    rows = [{
        "id": uuid4(),
        "periodo": "2026-06",
        "tipo": "xml",
        "nombre_archivo": "recibo.xml",
        "ruta_sharepoint": "https://sharepoint.test/recibo.xml",
        "numero_servicio": "123456789012",
        "servicio_nombre": "SERVICIO",
    }]

    archivos, faltantes = await service._descargar_archivos_sharepoint_con_reintentos(rows)

    assert intentos["total"] == 3
    assert len(archivos) == 1
    assert archivos[0][1] == CFE_XML
    assert faltantes == []


@pytest.mark.asyncio
async def test_descarga_zip_reporta_faltante_tras_tres_intentos(monkeypatch):
    _install_fake_redis(monkeypatch)
    import modules.cfe.service as cfe_service_module
    from modules.cfe.service import CfeService

    class FakeAuth:
        async def get_application_token(self):
            return "token"

    service = CfeService(_FakeCfeZipDB({"id": uuid4(), "numero_servicio": "1"}, []))
    intentos = {"total": 0}

    async def fake_http(_client, _token, _row):
        intentos["total"] += 1
        return None, "fallo permanente"

    monkeypatch.setattr(cfe_service_module, "get_ms_auth", lambda: FakeAuth())
    monkeypatch.setattr(service, "_descargar_archivo_sharepoint_http", fake_http)

    rows = [{
        "id": uuid4(),
        "periodo": "2026-06",
        "tipo": "pdf",
        "nombre_archivo": "recibo.pdf",
        "ruta_sharepoint": "https://sharepoint.test/recibo.pdf",
        "numero_servicio": "123456789012",
        "servicio_nombre": "SERVICIO",
    }]

    archivos, faltantes = await service._descargar_archivos_sharepoint_con_reintentos(rows)

    assert intentos["total"] == 3
    assert archivos == []
    assert faltantes[0]["periodo"] == "2026-06"
    assert faltantes[0]["tipo"] == "PDF"
    assert faltantes[0]["motivo"] == "fallo permanente"


@pytest.mark.asyncio
async def test_generar_zip_servicio_bloquea_y_luego_permite_faltantes(monkeypatch):
    _install_fake_redis(monkeypatch)
    from modules.cfe.service import CfeService, CfeZipFaltantesError

    servicio_id = uuid4()
    servicio = {
        "id": servicio_id,
        "numero_servicio": "123456789012",
        "nombre": "SERVICIO PRUEBA",
    }
    descargas = [
        {
            "id": uuid4(),
            "periodo": "2026-06",
            "tipo": "xml",
            "estatus": "completado",
            "nombre_archivo": "recibo.xml",
            "ruta_sharepoint": "https://sharepoint.test/recibo.xml",
        },
        {
            "id": uuid4(),
            "periodo": "2026-06",
            "tipo": "pdf",
            "estatus": "completado",
            "nombre_archivo": "recibo.pdf",
            "ruta_sharepoint": "https://sharepoint.test/recibo.pdf",
        },
    ]
    service = CfeService(_FakeCfeZipDB(servicio, descargas))

    async def fake_descargar_archivos(rows):
        xml_row = next(row for row in rows if row["tipo"] == "xml")
        pdf_row = next(row for row in rows if row["tipo"] == "pdf")
        return [(xml_row, CFE_XML)], [{
            "servicio": "SERVICIO PRUEBA",
            "numero_servicio": "123456789012",
            "periodo": "2026-06",
            "tipo": "PDF",
            "nombre_archivo": pdf_row["nombre_archivo"],
            "motivo": "fallo permanente",
        }]

    monkeypatch.setattr(service, "_descargar_archivos_sharepoint_con_reintentos", fake_descargar_archivos)

    with pytest.raises(CfeZipFaltantesError):
        await service.generar_zip_servicio(None, servicio_id)

    zip_bytes, _filename = await service.generar_zip_servicio(
        None, servicio_id, permitir_incompleto=True
    )

    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert "CFE_123456789012/XML/2026-06_recibo.xml" in names
        assert "CFE_123456789012/PDF/2026-06_recibo.pdf" not in names
        assert "_FALTANTES.txt" in names
        assert "recibo.pdf" in zf.read("_FALTANTES.txt").decode("utf-8")


@pytest.mark.asyncio
async def test_generar_zip_servicio_simulacion_renombra_xml_pdf(monkeypatch):
    _install_fake_redis(monkeypatch)
    from modules.cfe.service import CfeService

    servicio_id = uuid4()
    servicio = {
        "id": servicio_id,
        "numero_servicio": "237110414099",
        "nombre": "SERVICIO PRUEBA",
        "alias": "UNIVERSIDAD TECNOLOGICA DE TEHUACAN",
    }
    descargas = [
        {
            "id": uuid4(),
            "periodo": "2026-06",
            "tipo": "xml",
            "estatus": "completado",
            "nombre_archivo": "origen.xml",
            "ruta_sharepoint": "https://sharepoint.test/origen.xml",
        },
        {
            "id": uuid4(),
            "periodo": "2026-06",
            "tipo": "pdf",
            "estatus": "completado",
            "nombre_archivo": "origen.pdf",
            "ruta_sharepoint": "https://sharepoint.test/origen.pdf",
        },
    ]
    service = CfeService(_FakeCfeZipDB(servicio, descargas))

    async def fake_descargar_archivos(rows):
        contenidos = {"xml": CFE_XML, "pdf": b"%PDF-1.4\n"}
        return [(row, contenidos[row["tipo"]]) for row in rows], []

    monkeypatch.setattr(service, "_descargar_archivos_sharepoint_con_reintentos", fake_descargar_archivos)

    zip_bytes, _filename = await service.generar_zip_servicio(
        None, servicio_id, perfil_slug="simulacion"
    )

    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        base = "JUN 26_237110414099_UNIVERSIDAD TECNOLOGICA DE TEHUACAN"
        assert f"CFE_237110414099/XML/{base}.xml" in names
        assert f"CFE_237110414099/PDF/{base}.pdf" in names


@pytest.mark.asyncio
async def test_generar_zip_global_incluye_excel_global(monkeypatch):
    _install_fake_redis(monkeypatch)
    from modules.cfe.service import CfeService

    servicio_id = uuid4()
    global_rows = [
        {
            "id": uuid4(),
            "servicio_id": servicio_id,
            "numero_servicio": "123456789012",
            "servicio_nombre": "SERVICIO PRUEBA",
            "alias": "ALIAS PRUEBA",
            "periodo": "2026-06",
            "tipo": "xml",
            "estatus": "completado",
            "nombre_archivo": "recibo.xml",
            "ruta_sharepoint": "https://sharepoint.test/recibo.xml",
        },
        {
            "id": uuid4(),
            "servicio_id": servicio_id,
            "numero_servicio": "123456789012",
            "servicio_nombre": "SERVICIO PRUEBA",
            "alias": "ALIAS PRUEBA",
            "periodo": "2026-06",
            "tipo": "pdf",
            "estatus": "completado",
            "nombre_archivo": "recibo.pdf",
            "ruta_sharepoint": "https://sharepoint.test/recibo.pdf",
        },
    ]
    service = CfeService(
        _FakeCfeZipDB({"id": servicio_id, "numero_servicio": "123456789012"}, [], global_rows)
    )
    pdf_content = _pdf_bytes()

    async def fake_descargar_archivos(rows):
        contenidos = {"xml": CFE_XML, "pdf": pdf_content}
        return [(row, contenidos[row["tipo"]]) for row in rows], []

    monkeypatch.setattr(service, "_descargar_archivos_sharepoint_con_reintentos", fake_descargar_archivos)

    zip_bytes, _filename = await service.generar_zip_global(
        None, modulos=["simulacion"], perfil_slug="simulacion"
    )

    with zipfile.ZipFile(BytesIO(zip_bytes)) as zf:
        names = set(zf.namelist())
        assert any(name.startswith("CFE_todos_los_recibos_") and name.endswith(".pdf") for name in names)
        excel_global = [
            name for name in names
            if name.startswith("CFE_todos_los_recibos_") and name.endswith(".xlsx")
        ]
        assert len(excel_global) == 1
        workbook = load_workbook(BytesIO(zf.read(excel_global[0])))
        assert workbook.active.max_row == 2


def test_extrae_datos_cfe_para_excel():
    datos = extraer_datos_xml(CFE_XML, "recibo.xml")

    importes = {linea["concepto"]: linea["importe"] for linea in datos["lineas_excel"]}
    assert datos["servicio"]["rpu"] == "123456789012"
    assert datos["periodo"]["mes_nombre"] == "Feb"
    assert importes["kWh base"] == 1000
    assert importes["Generación B"] == 100
    assert importes["Subtotal"] == 4490


def test_construir_datos_recibo_cfe_publico_expone_normalizacion_excel():
    recibo = extraer_datos_xml(CFE_XML, "recibo.xml")

    datos = construir_datos_recibo_cfe(recibo)

    assert datos["mes"] == "Feb-25"
    assert datos["consumo"] == 6000
    assert datos["kwmax"] == 22
    assert datos["total"] == 4490
    assert "Tarifa: GDMTH" in datos["observaciones"]


def test_construir_analisis_recibos_calcula_ultimo_comparativos_y_alertas():
    recibo_anterior = extraer_datos_xml(
        _cfe_xml_con_metricas(
            subtotal="3000.00",
            total="3480.00",
            consumo_base="1000",
            consumo_intermedio="1000",
            consumo_punta="1000",
            demanda="12",
            fp="96.0",
        ),
        "ene.xml",
    )
    recibo_ultimo = extraer_datos_xml(
        _cfe_xml_con_metricas(
            subtotal="6000.00",
            total="6960.00",
            consumo_base="1000",
            consumo_intermedio="2000",
            consumo_punta="3000",
            demanda="22",
            fp="89.0",
        ),
        "feb.xml",
    )

    analisis = construir_analisis_recibos(
        {"id": "servicio-1", "nombre": "SERVICIO PRUEBA", "numero_servicio": "123"},
        [
            ({"periodo": "2025-01"}, recibo_anterior),
            ({"periodo": "2025-02"}, recibo_ultimo),
        ],
    )

    assert analisis["hay_datos"] is True
    assert analisis["ultimo"]["periodo"] == "2025-02"
    assert analisis["ultimo"]["consumo"] == 6000
    assert analisis["ultimo"]["total_facturado"] == 6960
    assert analisis["ultimo"]["costo_kwh"] == pytest.approx(1.16)
    assert analisis["baseline_periodos"] == 1

    consumo = next(item for item in analisis["comparativos"] if item["key"] == "consumo")
    assert consumo["promedio_12"]["disponible"] is True
    assert consumo["promedio_12"]["valor"] == 3000
    assert consumo["promedio_12"]["delta_pct"] == pytest.approx(100)

    assert analisis["variacion_historico"]["disponible"] is True
    assert analisis["variacion_historico"]["costo_esperado"] == pytest.approx(6960)
    assert any(alerta["titulo"] == "Factor de potencia bajo" for alerta in analisis["alertas"])

    kpi_total = next(item for item in analisis["kpis"] if item["key"] == "total_facturado")
    assert kpi_total["subtexto"]["label"] == "Subtotal"
    assert kpi_total["subtexto"]["valor"] == analisis["ultimo"]["subtotal"]


def test_construir_analisis_recibos_gdmto_no_infiere_perfil_horario_ni_punta():
    recibo_anterior = extraer_datos_xml(CFE_GDMTO_XML, "gdmto-ene.xml")
    recibo_ultimo = extraer_datos_xml(CFE_GDMTO_XML, "gdmto-feb.xml")

    analisis = construir_analisis_recibos(
        {"id": "servicio-1", "nombre": "SERVICIO GDMTO", "numero_servicio": "123"},
        [
            ({"periodo": "2026-02"}, recibo_anterior),
            ({"periodo": "2026-03"}, recibo_ultimo),
        ],
    )

    assert analisis["perfil_analisis"]["key"] == "GDMTO"
    assert analisis["secciones"]["perfil_horario"] is False
    assert analisis["ultimo"]["perfil_horario"] == []
    assert [kpi["key"] for kpi in analisis["kpis"]] == [
        "total_facturado",
        "consumo",
        "costo_kwh",
        "kwmax",
        "kw_cap",
        "kw_dist",
        "fp",
    ]
    assert "consumo_punta" not in analisis["graficas"]["metricas"]
    assert not any(alerta["titulo"] == "Consumo punta elevado" for alerta in analisis["alertas"])
    assert "GDMTO no se analiza con perfil horario base/intermedia/punta." in analisis["calidad_datos"]["limitaciones"]


def test_construir_analisis_recibos_excluye_baseline_con_tarifa_distinta():
    recibo_gdmth = extraer_datos_xml(CFE_XML, "gdmth.xml")
    recibo_gdmto = extraer_datos_xml(CFE_GDMTO_XML, "gdmto.xml")

    analisis = construir_analisis_recibos(
        {"id": "servicio-1", "nombre": "SERVICIO MIXTO", "numero_servicio": "123"},
        [
            ({"periodo": "2026-02"}, recibo_gdmth),
            ({"periodo": "2026-03"}, recibo_gdmto),
        ],
    )

    assert analisis["perfil_analisis"]["key"] == "GDMTO"
    assert analisis["baseline_periodos"] == 0
    assert analisis["periodos_comparables"] == 1
    assert analisis["periodos_excluidos_tarifa"] == 1
    assert any(alerta["titulo"] == "Baseline filtrado por tarifa" for alerta in analisis["alertas"])


def test_construir_analisis_recibos_tarifa_no_soportada_usa_basico():
    xml_pdbt = CFE_XML.replace(b"<TARIFA_REG>GDMTH</TARIFA_REG>", b"<TARIFA_REG>PDBT</TARIFA_REG>")
    xml_pdbt = xml_pdbt.replace(b"<TARIFA>GDMTH</TARIFA>", b"<TARIFA>PDBT</TARIFA>")
    recibo = extraer_datos_xml(xml_pdbt, "pdbt.xml")

    analisis = construir_analisis_recibos(
        {"id": "servicio-1", "nombre": "SERVICIO PDBT", "numero_servicio": "123"},
        [({"periodo": "2026-03"}, recibo)],
    )

    assert analisis["perfil_analisis"]["key"] == "NO_SOPORTADA"
    assert analisis["secciones"]["perfil_horario"] is False
    assert [kpi["key"] for kpi in analisis["kpis"]] == [
        "total_facturado",
        "consumo",
        "costo_kwh",
    ]
    assert [item["key"] for item in analisis["comparativos"]] == [
        "consumo",
        "total_facturado",
        "costo_kwh",
    ]
    assert any("La tarifa PDBT no tiene reglas especificas" in item for item in analisis["calidad_datos"]["limitaciones"])


def test_extrae_gdmto_demandas_historial_y_campos_dinamicos():
    datos = extraer_datos_xml(CFE_GDMTO_XML, "gdmto.xml")

    assert datos["servicio"]["tarifa"] == "GDMTO"
    assert datos["servicio"]["tarifa_reconocida"] is True
    assert datos["medicion"]["demanda_capacidad"] == 30
    assert datos["medicion"]["demanda_distribucion"] == 31
    assert datos["historial"] == [
        {
            "mes": "MAR",
            "consumo_kwh": 4500,
            "demanda_kw": 28,
            "factor_potencia_pct": 97.1,
            "factor_carga_pct": 42.5,
            "precio_medio_mxn": 3.14,
        },
        {
            "mes": "FEB",
            "consumo_kwh": 4100,
            "demanda_kw": 26,
            "factor_potencia_pct": 96.8,
            "factor_carga_pct": 41.2,
            "precio_medio_mxn": 3.01,
        },
    ]
    assert datos["componentes_tarifarios"][-1]["codigo"] == "ET1"
    importes = {linea["concepto"]: linea["importe"] for linea in datos["lineas_excel"]}
    assert importes["Generación I"] == 4200.24
    assert any(
        concepto["concepto"] == "Cargo dinamico"
        for concepto in datos["conceptos_importes"]
    )


def test_generar_excel_gdmto_usa_directos_na_e_historial():
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="gdmto.xml", content=CFE_GDMTO_XML)],
        perfil_slug="simulacion",
        modo_calculo="calculado",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active

    assert ws.title == "226160800958"
    assert _valor_en_fila(ws, "Mes") == "Mar-26"
    assert _valor_en_fila(ws, "Consumo") == 4500
    assert _valor_en_fila(ws, "Consumo Base") == "N/A"
    assert _valor_en_fila(ws, "Potencia Base") == "N/A"
    assert _valor_en_fila(ws, "Reactiva") == "N/A"
    assert _valor_en_fila(ws, "KW CAP") == 30
    assert _valor_en_fila(ws, "kW DIST") == 31
    assert _valor_en_fila(ws, "Coste Energía (Intermedia)") == 4200.24

    hojas_historial = [name for name in wb.sheetnames if "Historial" in name]
    assert hojas_historial == ["226160800958 Historial"]
    historial = wb[hojas_historial[0]]
    assert [cell.value for cell in historial[1]] == [
        "Mes",
        "Consumo kWh",
        "Demanda kW",
        "Factor Potencia %",
        "Factor Carga %",
        "Precio Medio MXN",
    ]
    assert historial.max_row == 3
    assert historial["A2"].value == "MAR"
    assert historial["B2"].value == 4500


def test_generar_excel_gdmto_formulas_conserva_valores_directos():
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="gdmto.xml", content=CFE_GDMTO_XML)],
        perfil_slug="simulacion",
        modo_calculo="formulas",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active

    assert _valor_en_fila(ws, "Consumo") == 4500
    assert _valor_en_fila(ws, "KW CAP") == 30
    assert _valor_en_fila(ws, "kW DIST") == 31


def test_gdmth_no_remapea_eg1_como_generacion_intermedia():
    xml = CFE_XML.replace(
        b"<MOTIVO_REG_2>EGI</MOTIVO_REG_2>",
        b"<MOTIVO_REG_2>EG1</MOTIVO_REG_2>",
    )
    datos = extraer_datos_xml(xml, "gdmth-eg1.xml")

    importes = {linea["concepto"]: linea["importe"] for linea in datos["lineas_excel"]}
    assert importes["Generacion"] == 200
    assert "Generación I" not in importes


@pytest.mark.skipif(
    not CFE_REAL_GDMTO_XML.exists(),
    reason="XML real GDMTO no disponible",
)
def test_xml_real_gdmto_eg1_alimenta_coste_energia_intermedia():
    content = CFE_REAL_GDMTO_XML.read_bytes()
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="JB-000090976983(3).xml", content=content)],
        perfil_slug="simulacion",
        modo_calculo="calculado",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active

    assert _valor_en_fila(ws, "Coste Energía (Intermedia)") == 4200.24


def test_generar_excel_simulacion_calculado():
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="recibo.xml", content=CFE_XML)],
        perfil_slug="simulacion",
        modo_calculo="calculado",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    assert headers[-1] == "Observaciones"
    assert headers == [column.header for column in obtener_perfil_cfe("simulacion").columns]
    assert ws["A2"].value == "Feb-25"
    assert ws["B2"].value == 6000
    assert ws["I2"].value == 28
    assert ws["J2"].value == 16
    assert ws["K2"].value == 16
    assert ws["Z2"].value == 4490
    assert "Tarifa: GDMTH" in ws["AA2"].value


def test_generar_excel_oym_mantiene_observaciones():
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="recibo.xml", content=CFE_XML)],
        perfil_slug="oym",
        modo_calculo="calculado",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    assert headers[-1] == "Observaciones"
    assert "Tarifa: GDMTH" in ws.cell(row=2, column=len(headers)).value


def test_generar_excel_omite_xml_no_cfe_y_procesa_validos():
    buffer = generar_excel_cfe(
        [
            CfeXmlInput(filename="proveedor.xml", content=NON_CFE_XML),
            CfeXmlInput(filename="recibo.xml", content=CFE_XML),
        ],
        perfil_slug="simulacion",
        modo_calculo="calculado",
    )
    wb = load_workbook(buffer, data_only=False)

    assert "Validacion" in wb.sheetnames
    ws = wb.active
    assert ws.max_row == 2
    assert ws["A2"].value == "Feb-25"

    validacion = wb["Validacion"]
    assert validacion["A2"].value == "proveedor.xml"
    assert validacion["B2"].value == "Omitido"
    assert validacion["C2"].value == "No es un XML de recibo CFE"
    assert validacion["A3"].value == "recibo.xml"
    assert validacion["B3"].value == "Procesado"


def test_generar_excel_error_si_no_hay_xml_cfe():
    with pytest.raises(ValueError, match="No se encontraron XML de CFE válidos"):
        generar_excel_cfe(
            [CfeXmlInput(filename="proveedor.xml", content=NON_CFE_XML)],
            perfil_slug="simulacion",
            modo_calculo="calculado",
        )


def test_generar_excel_soporta_formulas_internas():
    buffer = generar_excel_cfe(
        [CfeXmlInput(filename="recibo.xml", content=CFE_XML)],
        perfil_slug="simulacion",
        modo_calculo="formulas",
    )
    wb = load_workbook(buffer, data_only=False)
    ws = wb.active

    assert ws["B2"].value == "=SUM(C2,D2,E2)"
    assert ws["J2"].value.startswith("=ROUNDUP(MIN(")
    assert ws["K2"].value.startswith("=ROUNDUP(MIN(MAX(")


def test_generar_excel_rechaza_modo_invalido():
    with pytest.raises(ValueError, match="Modo de Excel CFE no soportado"):
        generar_excel_cfe(
            [CfeXmlInput(filename="recibo.xml", content=CFE_XML)],
            perfil_slug="simulacion",
            modo_calculo="mixto",
        )


def test_modal_cfe_compila_con_contexto_de_modulo():
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("shared/modals/cfe_upload_modal.html")

    html = template.render(
        module_slug="simulacion",
        module_label="Simulación",
        post_url="/simulacion/cfe/excel",
        accent="indigo",
    )

    assert "Recibos CFE - Simulación" in html
    assert "/simulacion/cfe/excel" in html


def test_lista_cfe_contiene_flujo_zip_con_faltantes():
    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template("cfe/partials/lista_servicios.html")
    servicio_id = uuid4()

    html = template.render(
        servicios=[{
            "id": servicio_id,
            "nombre": "SERVICIO PRUEBA",
            "alias": "ALIAS",
            "numero_servicio": "123456789012",
            "miespacio_estatus": "registrado",
            "miespacio_error": None,
            "ultima_descarga": None,
            "total_descargas": 1,
            "descarga_activa": False,
            "busqueda_activa_id": None,
            "busqueda_activa_estatus": None,
        }],
        estado_sesion={"sesion_estado": "activa"},
        modulo="simulacion",
        modulos_accesibles=["simulacion"],
        user={},
    )

    assert "descargarZipCfe(this, '/cfe/servicios/zip-global?modulo=simulacion')" in html
    assert f"descargarZipCfe(this, '/cfe/servicios/{servicio_id}/zip?modulo=simulacion')" in html
    assert "Descargar ZIP con faltantes" in html


def test_router_zip_servicio_faltantes_devuelve_409(monkeypatch):
    _install_fake_redis(monkeypatch)
    from core.database import get_db_connection
    from core.security import get_current_user_context
    import modules.cfe.router as cfe_router_module
    from modules.cfe.router import router as cfe_router
    from modules.cfe.service import CfeZipFaltantesError

    servicio_id = uuid4()

    class FakeDB:
        async def get_servicio_by_id(self, _conn, _servicio_id):
            return {
                "id": servicio_id,
                "numero_servicio": "123456789012",
                "nombre": "SERVICIO PRUEBA",
                "modulos": ["oym"],
            }

    class FakeService:
        db = FakeDB()

        async def generar_zip_servicio(self, *_args, **_kwargs):
            raise CfeZipFaltantesError([{
                "servicio": "SERVICIO PRUEBA",
                "numero_servicio": "123456789012",
                "periodo": "2026-06",
                "tipo": "PDF",
                "nombre_archivo": "recibo.pdf",
                "motivo": "fallo permanente",
            }])

    async def fake_conn():
        yield None

    app = FastAPI()
    app.include_router(cfe_router)
    app.dependency_overrides[get_db_connection] = fake_conn
    app.dependency_overrides[get_current_user_context] = lambda: {
        "user_db_id": uuid4(),
        "role": "USER",
        "module_roles": {"oym": "viewer"},
    }
    monkeypatch.setattr(cfe_router_module, "get_cfe_service", lambda: FakeService())

    response = TestClient(app).get(f"/cfe/servicios/{servicio_id}/zip?modulo=oym")

    assert response.status_code == 409
    assert response.json()["faltantes"][0]["nombre_archivo"] == "recibo.pdf"
