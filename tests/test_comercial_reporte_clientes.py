from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from unittest.mock import AsyncMock, patch
from uuid import uuid4

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from core.database import get_db_connection
from core.pdf_service.service import get_pdf_service
from core.security import get_current_user_context
from modules.comercial import reportes_service
from modules.comercial.reportes_excel_builder import (
    build_reporte_clientes_general_workbook,
    build_reporte_clientes_por_cliente_workbook,
    generar_nombre_archivo,
)
from modules.comercial.router import router as comercial_router


# ---------------------------------------------------------------------------
# Filtros / fechas
# ---------------------------------------------------------------------------

def test_parse_filtros_sin_fechas_permite_historico_completo():
    filtros = reportes_service.parse_filtros_reporte_clientes()

    assert filtros.fecha_inicio is None
    assert filtros.fecha_fin is None
    assert filtros.filtro_cliente_id is None


def test_parse_filtros_fechas_validas():
    filtros = reportes_service.parse_filtros_reporte_clientes(
        filtro_fecha_inicio="2026-01-01", filtro_fecha_fin="2026-06-30"
    )

    assert filtros.fecha_inicio == date(2026, 1, 1)
    assert filtros.fecha_fin == date(2026, 6, 30)


def test_parse_filtros_formato_invalido_lanza_value_error():
    with pytest.raises(ValueError, match="invalida"):
        reportes_service.parse_filtros_reporte_clientes(filtro_fecha_inicio="01/01/2026")


def test_parse_filtros_rango_invertido_lanza_value_error():
    with pytest.raises(ValueError):
        reportes_service.parse_filtros_reporte_clientes(
            filtro_fecha_inicio="2026-06-30", filtro_fecha_fin="2026-01-01"
        )


def test_parse_filtros_solo_activos_con_cliente_lanza_value_error():
    """solo_activos no aplica en modo por-cliente; combinarlos es un 400, no un no-op silencioso."""
    with pytest.raises(ValueError, match="solo_activos"):
        reportes_service.parse_filtros_reporte_clientes(
            filtro_cliente_id=uuid4(), solo_activos=True
        )


def test_limites_timestamptz_fin_es_exclusivo_dia_siguiente():
    filtros = reportes_service.FiltrosReporteClientes(
        fecha_inicio=date(2026, 1, 1), fecha_fin=date(2026, 1, 31)
    )

    inicio_mx, fin_mx = reportes_service._limites_timestamptz(filtros)

    assert inicio_mx.date() == date(2026, 1, 1)
    assert fin_mx.date() == date(2026, 2, 1)
    assert inicio_mx.tzinfo is not None and fin_mx.tzinfo is not None


def test_limites_timestamptz_sin_fechas_es_none():
    filtros = reportes_service.FiltrosReporteClientes()

    assert reportes_service._limites_timestamptz(filtros) == (None, None)


# ---------------------------------------------------------------------------
# Guardrails (mockeando ConfigService y el db_service)
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_generar_dataset_general_excede_limite_solicitudes():
    filtros = reportes_service.FiltrosReporteClientes()
    conn = AsyncMock()

    with patch.object(reportes_service.ConfigService, "get_global_config", return_value=10) as mock_cfg, \
         patch("modules.comercial.reportes_service.db.contar_oportunidades_filtradas", return_value=11):
        with pytest.raises(ValueError, match="maximo de solicitudes"):
            await reportes_service.generar_dataset_general(
                conn, filtros, formato="excel", solicitante_email="a@b.com"
            )
    assert mock_cfg.await_count >= 1


@pytest.mark.asyncio
async def test_generar_dataset_general_excede_limite_resumen():
    filtros = reportes_service.FiltrosReporteClientes()
    conn = AsyncMock()

    with patch.object(reportes_service.ConfigService, "get_global_config", return_value=10_000), \
         patch("modules.comercial.reportes_service.db.contar_oportunidades_filtradas", return_value=5), \
         patch("modules.comercial.reportes_service.db.contar_filas_resumen_general", return_value=999_999):
        with pytest.raises(ValueError, match="maximo de clientes"):
            await reportes_service.generar_dataset_general(
                conn, filtros, formato="pdf", solicitante_email="a@b.com"
            )


@pytest.mark.asyncio
async def test_generar_dataset_por_cliente_requiere_cliente_id():
    filtros = reportes_service.FiltrosReporteClientes()
    conn = AsyncMock()

    with pytest.raises(ValueError, match="filtro_cliente_id"):
        await reportes_service.generar_dataset_por_cliente(
            conn, filtros, formato="excel", solicitante_email="a@b.com"
        )


@pytest.mark.asyncio
async def test_generar_dataset_por_cliente_excede_limite_detalle():
    filtros = reportes_service.FiltrosReporteClientes(filtro_cliente_id=uuid4())
    conn = AsyncMock()

    with patch.object(reportes_service.ConfigService, "get_global_config", return_value=10_000), \
         patch("modules.comercial.reportes_service.db.contar_oportunidades_filtradas", return_value=1), \
         patch("modules.comercial.reportes_service.db.obtener_detalle_por_cliente", return_value=[{}] * 10_001):
        with pytest.raises(ValueError, match="maximo de filas"):
            await reportes_service.generar_dataset_por_cliente(
                conn, filtros, formato="excel", solicitante_email="a@b.com"
            )


@pytest.mark.asyncio
async def test_generar_dataset_general_ok_arma_dto():
    filtros = reportes_service.FiltrosReporteClientes()
    conn = AsyncMock()
    resumen = [{"grupo_id": "1", "cliente_nombre": "Cliente A", "total_solicitudes": 2, "desglose_estatus": "GANADA: 2"}]
    detalle = [{"cliente_nombre": "Cliente A", "folio": "OP-1", "fecha_solicitud": datetime(2026, 1, 5)}]

    with patch.object(reportes_service.ConfigService, "get_global_config", return_value=10_000), \
         patch("modules.comercial.reportes_service.db.contar_oportunidades_filtradas", return_value=1), \
         patch("modules.comercial.reportes_service.db.contar_filas_resumen_general", return_value=1), \
         patch("modules.comercial.reportes_service.db.obtener_resumen_clientes", return_value=resumen), \
         patch("modules.comercial.reportes_service.db.obtener_detalle_general", return_value=detalle):
        dataset = await reportes_service.generar_dataset_general(
            conn, filtros, formato="excel", solicitante_email="a@b.com"
        )

    assert dataset["modo"] == "general"
    assert dataset["resumen"] == resumen
    assert dataset["detalle"] == detalle


@pytest.mark.asyncio
async def test_generar_dataset_por_cliente_ok_incluye_nombre_cliente():
    cliente_id = uuid4()
    filtros = reportes_service.FiltrosReporteClientes(filtro_cliente_id=cliente_id)
    conn = AsyncMock()
    detalle = [{"folio": "OP-1", "fecha_solicitud": datetime(2026, 1, 5)}]

    with patch.object(reportes_service.ConfigService, "get_global_config", return_value=10_000), \
         patch("modules.comercial.reportes_service.db.contar_oportunidades_filtradas", return_value=1), \
         patch("modules.comercial.reportes_service.db.obtener_detalle_por_cliente", return_value=detalle), \
         patch("modules.comercial.reportes_service.db.obtener_nombre_cliente", return_value="Cliente A"):
        dataset = await reportes_service.generar_dataset_por_cliente(
            conn, filtros, formato="excel", solicitante_email="a@b.com"
        )

    assert dataset["modo"] == "cliente"
    assert dataset["detalle"] == detalle
    assert dataset["cliente_nombre"] == "Cliente A"


@pytest.mark.asyncio
async def test_describir_filtros_formatea_fechas_dd_mm_yyyy():
    filtros = reportes_service.FiltrosReporteClientes(
        fecha_inicio=date(2026, 1, 5), fecha_fin=date(2026, 6, 30)
    )
    conn = AsyncMock()
    etiquetas = {"tipo_nombre": None, "tecnologia_nombre": None, "estatus_nombre": None, "cliente_nombre": None}

    with patch("modules.comercial.reportes_service.db.obtener_etiquetas_filtros", return_value=etiquetas):
        resumen = await reportes_service.describir_filtros(conn, filtros)

    assert "05/01/2026" in resumen
    assert "30/06/2026" in resumen
    assert "2026-01-05" not in resumen


@pytest.mark.asyncio
async def test_describir_filtros_sin_fechas_es_historico_completo():
    filtros = reportes_service.FiltrosReporteClientes()
    conn = AsyncMock()
    etiquetas = {"tipo_nombre": None, "tecnologia_nombre": None, "estatus_nombre": None, "cliente_nombre": None}

    with patch("modules.comercial.reportes_service.db.obtener_etiquetas_filtros", return_value=etiquetas):
        resumen = await reportes_service.describir_filtros(conn, filtros)

    assert resumen == (
        "Fechas: histórico completo | "
        "Vista: todos los clientes (incluye sin actividad en el rango)"
    )


# ---------------------------------------------------------------------------
# Excel builder: seguridad e inyeccion de formulas, fechas nativas, dataset vacio
# ---------------------------------------------------------------------------

def test_excel_neutraliza_texto_peligroso_en_resumen():
    resumen = [{"cliente_nombre": "=cmd|'/c calc'!A1", "total_solicitudes": 1, "desglose_estatus": "+HYPERLINK(1)"}]
    workbook = build_reporte_clientes_general_workbook(resumen, [])
    ws = workbook["Resumen por cliente"]

    assert ws["A2"].value.startswith("'=")
    assert ws["C2"].value.startswith("'+")


def test_excel_fecha_detalle_general_es_valor_nativo_no_string():
    detalle = [{
        "cliente_nombre": "Cliente A",
        "folio": "OP-1",
        "fecha_solicitud": datetime(2026, 3, 15, 10, 30, tzinfo=None),
        "estatus_nombre": "GANADA",
        "fase_proyecto": "N/A",
    }]
    workbook = build_reporte_clientes_general_workbook([], detalle)
    ws = workbook["Detalle de solicitudes"]
    cell = ws["C2"]

    assert isinstance(cell.value, datetime)
    assert cell.number_format == "dd/mm/yyyy hh:mm"


def test_excel_general_dataset_vacio_solo_encabezados():
    workbook = build_reporte_clientes_general_workbook([], [])

    assert workbook.sheetnames == ["Resumen por cliente", "Detalle de solicitudes"]
    assert workbook["Resumen por cliente"].max_row == 1
    assert workbook["Detalle de solicitudes"].max_row == 1


def test_excel_por_cliente_hoja_unica_y_columnas():
    detalle = [{
        "folio": "OP-1",
        "fecha_solicitud": datetime(2026, 3, 15),
        "estatus_nombre": "GANADA",
        "sitio_nombre": "Sucursal Norte",
        "sitio_direccion": "Av. Reforma 123",
        "proyecto_id_estandar": "FV-001-PROY",
        "fase_proyecto": "CONSTRUCCION",
    }]
    workbook = build_reporte_clientes_por_cliente_workbook(detalle)

    assert workbook.sheetnames == ["Detalle por cliente"]
    ws = workbook["Detalle por cliente"]
    assert ws["A2"].value == "OP-1"
    assert ws["D2"].value == "Sucursal Norte"


def test_excel_por_cliente_dataset_vacio():
    workbook = build_reporte_clientes_por_cliente_workbook([])

    assert workbook["Detalle por cliente"].max_row == 1


def test_excel_por_cliente_hoja_incluye_nombre_cliente():
    workbook = build_reporte_clientes_por_cliente_workbook([], cliente_nombre="Acme S.A. de C.V.")

    assert workbook.sheetnames == ["Detalle - Acme S.A. de C.V."]


def test_generar_nombre_archivo_incluye_cliente_sanitizado():
    nombre = generar_nombre_archivo("Acme / Ríos S.A.?")

    assert nombre.startswith("reporte_clientes_")
    assert "Acme" in nombre
    assert nombre.endswith(".xlsx")
    assert not any(c in nombre for c in ("/", "?", " "))


def test_generar_nombre_archivo_sin_cliente_es_generico():
    nombre = generar_nombre_archivo()

    assert nombre.startswith("reporte_clientes_")
    assert "None" not in nombre


# ---------------------------------------------------------------------------
# Autorizacion (HTTP end-to-end, ejercitando require_module_access y
# require_role reales — el gate ADMIN/MANAGER ya no es un helper imperativo,
# es require_role(["ADMIN", "MANAGER"]) stackeado como Depends; su
# comportamiento generico (401/403) ya esta cubierto en test_permissions.py)
# ---------------------------------------------------------------------------

def _build_test_app() -> FastAPI:
    """App aislada con solo el router de Comercial; get_db_connection va mockeado,
    get_current_user_context lo sobreescribe cada test. require_module_access queda
    SIN mockear para verificar el gate de RBAC real, no solo el helper interno."""
    app = FastAPI()
    app.include_router(comercial_router)

    async def fake_conn():
        yield AsyncMock()

    app.dependency_overrides[get_db_connection] = fake_conn
    return app


def test_reporte_excel_manager_sin_acceso_403():
    """MANAGER sin 'comercial' en module_roles debe ser rechazado por require_module_access,
    antes de llegar siquiera al gate require_role(["ADMIN", "MANAGER"])."""
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: {
        "user_db_id": uuid4(),
        "email": "manager@test.com",
        "role": "MANAGER",
        "module_roles": {},
    }

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 403


def test_reporte_excel_user_module_admin_403():
    """USER con module-admin en Comercial pasa require_module_access (tiene rol suficiente)
    pero debe ser bloqueado por el gate require_role(["ADMIN", "MANAGER"]): es el caso
    exacto que motivo no usar require_manager_access para este reporte."""
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: {
        "user_db_id": uuid4(),
        "email": "user@test.com",
        "role": "USER",
        "module_roles": {"comercial": "admin"},
    }

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 403


# ---------------------------------------------------------------------------
# Autorizacion positiva + contrato de respuesta (HTTP end-to-end)
# ---------------------------------------------------------------------------

def _dataset_general_fake():
    return {
        "modo": "general",
        "resumen": [
            {"grupo_id": "1", "cliente_nombre": "Cliente Test", "total_solicitudes": 2, "desglose_estatus": "Ganada: 2"}
        ],
        "detalle": [
            {
                "cliente_nombre": "Cliente Test", "folio": "OP-1",
                "fecha_solicitud": datetime(2026, 1, 5), "estatus_nombre": "Ganada", "fase_proyecto": "N/A",
            }
        ],
        "filtros": None,
    }


def test_reporte_excel_admin_global_sin_module_roles_200(monkeypatch, admin_context):
    """ADMIN global sin entrada en module_roles debe pasar por el bypass de
    require_module_access y llegar a 200: es el caso borde central de la autorizacion."""
    monkeypatch.setattr(
        reportes_service, "generar_dataset_general", AsyncMock(return_value=_dataset_general_fake())
    )
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 200


def test_reporte_excel_manager_con_acceso_200(monkeypatch):
    """MANAGER con viewer+ en Comercial debe pasar ambas verificaciones y llegar a 200."""
    monkeypatch.setattr(
        reportes_service, "generar_dataset_general", AsyncMock(return_value=_dataset_general_fake())
    )
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: {
        "user_db_id": uuid4(),
        "email": "manager@test.com",
        "role": "MANAGER",
        "module_roles": {"comercial": "viewer"},
    }

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 200


def test_reporte_excel_content_disposition_y_workbook_legible(monkeypatch, admin_context):
    """La respuesta 200 trae Content-Disposition con nombre esperado y bytes de un
    workbook xlsx real y legible (no se mockea el builder, solo el dataset)."""
    monkeypatch.setattr(
        reportes_service, "generar_dataset_general", AsyncMock(return_value=_dataset_general_fake())
    )
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert "attachment" in disposition
    assert "reporte_clientes_" in disposition
    assert disposition.strip().endswith('.xlsx"')

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Resumen por cliente", "Detalle de solicitudes"]
    assert workbook["Resumen por cliente"]["A2"].value == "Cliente Test"


def test_reporte_excel_400_por_limite_excedido(monkeypatch, admin_context):
    """El ValueError de reportes_service (limite de guardrail) se traduce a 400 con detail."""
    monkeypatch.setattr(
        reportes_service,
        "generar_dataset_general",
        AsyncMock(side_effect=ValueError("El reporte excede el maximo de solicitudes permitido (20000 > 10000).")),
    )
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context

    response = TestClient(app).get("/comercial/reportes/clientes.xlsx")

    assert response.status_code == 400
    assert "maximo de solicitudes" in response.json()["detail"]


class _FakePdfService:
    async def generate(self, template_name, context):
        return b"%PDF-1.4 fake reporte clientes"

    def generate_filename(self, prefix, suffix=""):
        return f"{prefix}_fake.pdf"


def test_reporte_pdf_200_content_type(monkeypatch, admin_context):
    """El endpoint PDF, con PDFService mockeado, responde 200 y application/pdf."""
    monkeypatch.setattr(
        reportes_service, "generar_dataset_general", AsyncMock(return_value=_dataset_general_fake())
    )
    monkeypatch.setattr(reportes_service, "describir_filtros", AsyncMock(return_value="Fechas: histórico completo"))
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context
    app.dependency_overrides[get_pdf_service] = lambda: _FakePdfService()

    response = TestClient(app).get("/comercial/reportes/clientes.pdf")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


def test_reporte_pdf_manager_sin_acceso_403():
    """Mismo gate de autorizacion que el Excel, verificado tambien en la ruta PDF."""
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: {
        "user_db_id": uuid4(),
        "email": "manager@test.com",
        "role": "MANAGER",
        "module_roles": {},
    }

    response = TestClient(app).get("/comercial/reportes/clientes.pdf")

    assert response.status_code == 403


# ---------------------------------------------------------------------------
# Regresion: el parseo de filtros debe seguir mapeando a 400, no a un 500
# sin manejar (bug real: _preparar_reporte_clientes se llamaba fuera del try).
# ---------------------------------------------------------------------------

def test_reporte_excel_fecha_invalida_400_no_500(admin_context):
    """Una fecha con formato invalido debe traducirse a 400 con detail, no a un 500
    sin manejar — _preparar_reporte_clientes corre real (sin mock) para ejercitar
    reportes_service.parse_filtros_reporte_clientes de punta a punta."""
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context

    response = TestClient(app).get(
        "/comercial/reportes/clientes.xlsx", params={"filtro_fecha_inicio": "not-a-date"}
    )

    assert response.status_code == 400
    assert "invalida" in response.json()["detail"]


def test_reporte_pdf_fecha_invalida_400_no_500(admin_context):
    """Mismo bug de mapeo de ValueError, verificado tambien en la ruta PDF."""
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context

    response = TestClient(app).get(
        "/comercial/reportes/clientes.pdf", params={"filtro_fecha_inicio": "not-a-date"}
    )

    assert response.status_code == 400
    assert "invalida" in response.json()["detail"]


class _FakePdfServiceTemplateNotFound:
    async def generate(self, template_name, context):
        raise ValueError(f"Template PDF no encontrado: {template_name}")

    def generate_filename(self, prefix, suffix=""):
        return f"{prefix}_fake.pdf"


def test_reporte_pdf_error_generando_pdf_es_500_no_400(monkeypatch, admin_context):
    """Un ValueError propio de PDFService.generate (ej. plantilla no encontrada) debe
    traducirse a 500, no confundirse con el 400 de los guardrails de reportes_service —
    son dos ValueError distintos que no deben compartir el mismo except."""
    monkeypatch.setattr(
        reportes_service, "generar_dataset_general", AsyncMock(return_value=_dataset_general_fake())
    )
    monkeypatch.setattr(reportes_service, "describir_filtros", AsyncMock(return_value="Fechas: histórico completo"))
    app = _build_test_app()
    app.dependency_overrides[get_current_user_context] = lambda: admin_context
    app.dependency_overrides[get_pdf_service] = lambda: _FakePdfServiceTemplateNotFound()

    response = TestClient(app).get("/comercial/reportes/clientes.pdf")

    assert response.status_code == 500
    assert "Template PDF no encontrado" not in response.json()["detail"]
